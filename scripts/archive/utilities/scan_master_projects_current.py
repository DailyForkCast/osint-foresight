"""
Scan current Master_Projects structure
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("MASTER_PROJECTS - CURRENT STRUCTURE")
print("="*80)

ws = wb['Master_Projects']

print("\nAll column headers:")
for col in range(1, min(35, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    if header:
        print(f"  {col_letter}: {header}")
