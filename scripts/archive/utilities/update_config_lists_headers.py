"""
Update Config_Lists sheet column headers
"""
import openpyxl

print("="*80)
print("UPDATING CONFIG_LISTS HEADERS")
print("="*80)

# Load v21
print("\nLoading v21...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v21.xlsx')

ws = wb['Config_Lists']

print("\nSetting column headers in row 1...")

ws['A1'] = 'Status'
ws['B1'] = 'Priority'
ws['C1'] = 'Stage'
ws['D1'] = 'Country_Code'
ws['E1'] = 'Country'
ws['F1'] = 'Region'

print("  A1: Status")
print("  B1: Priority")
print("  C1: Stage")
print("  D1: Country_Code")
print("  E1: Country")
print("  F1: Region")

# Save
print("\nSaving v21...")
wb.save('2025-10-26-Tracker-v21.xlsx')

print("\n" + "="*80)
print("CONFIG_LISTS HEADERS UPDATED!")
print("="*80)

print("\nColumn layout:")
print("  A: Status (Started, On Hold, Completed, etc.)")
print("  B: Priority (High, Medium, Low)")
print("  C: Stage (Planning, Implementation, Closing, etc.)")
print("  D: Country_Code (AF, CN, EUR, etc.)")
print("  E: Country (Afghanistan, China, etc.)")
print("  F: Region (AF, EAP, EUR, NEA, SCA, WHA)")

print("\nv21 is ready!")
