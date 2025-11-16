"""
Fix critical issues found in deep dive - v45
"""
import openpyxl

print("="*80)
print("FIXING CRITICAL ISSUES - V45")
print("="*80)

# Load v44
print("\nLoading v44...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v44.xlsx')

print("\n" + "="*80)
print("FIX 1: CONTROL E5 #REF! ERROR")
print("="*80)

ws_control = wb['Control']

print("\nCurrent E5 formula:")
print(f"  {ws_control['E5'].value}")

print("\nFix: Change to =B9/E8 (Total Projects / Countries Active)")
ws_control['E5'] = '=B9/E8'

print(f"  New: =B9/E8")

print("\n" + "="*80)
print("FIX 2: COUNTRY_DASHBOARD B8 #REF! ERROR")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nCurrent B8 formula:")
print(f"  {ws_country['B8'].value}")

print("\nFix: Change to proper Country PM lookup")
# First, let me check if Country_PM_Assignments has a table
ws_pm = wb['Country_PM_Assignments']
if 'T_Country_PM_Assignments' in ws_pm.tables:
    print("  Using table reference: T_Country_PM_Assignments")
    new_formula = '=IFERROR(INDEX(T_Country_PM_Assignments[Project_Manager],MATCH($B$2,T_Country_PM_Assignments[Country_Code],0)),"")'
else:
    print("  No table found, using direct reference to Country_PM_Assignments sheet")
    new_formula = '=IFERROR(INDEX(Country_PM_Assignments!D:D,MATCH($B$2,Country_PM_Assignments!A:A,0)),"")'

ws_country['B8'] = new_formula
print(f"  New: {new_formula}")

print("\n" + "="*80)
print("FIX 3: REGIONAL_SUMMARY REGION NAMES")
print("="*80)

ws_regional = wb['Regional_Summary']

print("\nCurrent region names (incorrect):")
for row in range(2, 8):
    region_code = ws_regional.cell(row, 1).value
    region_name = ws_regional.cell(row, 2).value
    print(f"  {region_code}: {region_name}")

print("\nFixing region names...")

# Correct mapping
region_names = {
    'AF': 'Africa',
    'EAP': 'East Asia Pacific',
    'EUR': 'Europe',
    'NEA': 'Near East Asia',
    'SCA': 'South Central Asia',
    'WHA': 'Western Hemisphere'
}

for row in range(2, 8):
    region_code = ws_regional.cell(row, 1).value
    if region_code in region_names:
        ws_regional.cell(row, 2).value = region_names[region_code]
        print(f"  {region_code} -> {region_names[region_code]}")

print("\n" + "="*80)
print("SAVING V45")
print("="*80)

wb.save('2025-10-26-Tracker-v45.xlsx')

print("\nOK - v45 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v45:")
print("  1. Control E5: Now calculates Projects/Country correctly")
print("  2. Country_Dashboard B8: Now looks up Country PM correctly")
print("  3. Regional_Summary: Fixed all region names")

print("\nAll critical #REF! errors resolved!")
print("\nv45 ready!")
