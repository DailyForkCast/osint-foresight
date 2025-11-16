#!/usr/bin/env python3
"""
Final Comprehensive TED Review Excel
Includes all expanded intelligence categories:
  - China Cooperation
  - Trade Events (NEW)
  - INTPA Radar (NEW)
  - Promotional Events (NEW)
  - Plus all original categories
"""

import sqlite3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from collections import defaultdict


class FinalComprehensiveExcel:
    """Generate final comprehensive Excel with all intelligence categories"""

    def __init__(self):
        self.db_path = Path("F:/OSINT_WAREHOUSE/osint_master.db")

        # Load expanded influence analysis
        self.influence_file = self.find_latest_file("ted_influence_expanded_*.json")
        with open(self.influence_file, 'r', encoding='utf-8') as f:
            self.influence_data = json.load(f)

        # Load revalidation for contractor analysis
        self.revalidation_file = self.find_latest_file("ted_revalidation_detailed_*.json")
        with open(self.revalidation_file, 'r', encoding='utf-8') as f:
            self.revalidation_results = json.load(f)

        # Create lookup
        self.influence_by_id = {}
        for contract in self.influence_data.get('contracts', []):
            self.influence_by_id[contract['contract_id']] = contract

    def find_latest_file(self, pattern):
        """Find most recent file matching pattern"""
        analysis_dir = Path("analysis")
        files = list(analysis_dir.glob(pattern))
        if not files:
            raise FileNotFoundError(f"No files found for: {pattern}")
        return max(files, key=lambda p: p.stat().st_mtime)

    def get_priority_info(self, category):
        """Get priority and color for category"""
        priority_map = {
            'BRI_RELATED': ('🔴 HIGH', 'FFE6E6'),
            'CHINA_COOPERATION': ('🔴 HIGH', 'FFE6E6'),
            'TRADE_MISSION': ('🔴 HIGH', 'FFE6E6'),
            'TRADE_EVENT': ('🟠 MED-HIGH', 'FFF4E6'),
            'PROMOTIONAL_EVENT': ('🟠 MED-HIGH', 'FFF4E6'),
            'CHINESE_FUNDED': ('🟠 MED-HIGH', 'FFF4E6'),
            '17PLUS1_INITIATIVE': ('🟠 MED-HIGH', 'FFF4E6'),
            'INTPA_RADAR': ('🟡 MEDIUM', 'FFFCE6'),
            'HK_REFERENCE': ('🟡 LOW-MED', 'FFFFEB'),
            'GEOGRAPHIC_REFERENCE': ('🟢 LOW', 'F5F5F5'),
            'NO_INFLUENCE_DETECTED': ('⚪ NONE', 'FFFFFF')
        }
        return priority_map.get(category, ('UNKNOWN', 'FFFFFF'))

    def create_excel(self):
        """Create comprehensive Excel"""

        print("\n" + "="*80)
        print("FINAL COMPREHENSIVE EXCEL GENERATION")
        print("="*80)
        print()

        # Categorize by intelligence priority
        by_category = defaultdict(list)

        for contract in self.influence_data.get('contracts', []):
            category = contract['intelligence_category']
            by_category[category].append(contract)

        # Create Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet order by priority
        sheet_order = [
            ('CHINA_COOPERATION', 'EU-China Cooperation'),
            ('BRI_RELATED', 'Belt & Road Initiative'),
            ('TRADE_MISSION', 'Trade Missions'),
            ('TRADE_EVENT', 'Trade Events'),
            ('PROMOTIONAL_EVENT', 'Promotional Events'),
            ('CHINESE_FUNDED', 'Chinese Funding'),
            ('17PLUS1_INITIATIVE', '17+1 Initiative'),
            ('INTPA_RADAR', 'INTPA Radar'),
            ('HK_REFERENCE', 'Hong Kong Refs'),
            ('GEOGRAPHIC_REFERENCE', 'Geographic Only'),
            ('NO_INFLUENCE_DETECTED', 'No Influence')
        ]

        for category, sheet_name in sheet_order:
            contracts = by_category.get(category, [])
            if contracts:
                print(f"  Creating sheet: {sheet_name} ({len(contracts)} contracts)")
                self.create_category_sheet(wb, sheet_name, category, contracts)

        # Create summary sheet
        self.create_summary_sheet(wb, by_category)

        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = Path(f"analysis/ted_final_comprehensive_review_{timestamp}.xlsx")
        wb.save(excel_file)

        print()
        print(f"[SUCCESS] Final comprehensive Excel created: {excel_file}")
        print(f"  File size: {excel_file.stat().st_size / 1024:.1f} KB")
        print()

        return excel_file

    def create_category_sheet(self, wb, sheet_name, category, contracts):
        """Create sheet for specific category"""

        ws = wb.create_sheet(title=sheet_name[:31])

        # Headers
        headers = [
            'ID', 'Notice', 'Date', 'Country', 'Contracting Authority',
            'Contractor', 'Contract Title',
            'INTELLIGENCE CATEGORY', 'PRIORITY',
            'Influence Patterns', 'Reference Patterns', 'INTPA Contract',
            'Value', 'Currency',
            'MANUAL REVIEW', 'NOTES'
        ]

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        priority, fill_color = self.get_priority_info(category)
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for row_num, contract in enumerate(contracts, 2):
            cursor = conn.cursor()
            full = cursor.execute("""
                SELECT ca_name, ca_country, contract_title,
                       contractor_name, value_total, currency, publication_date
                FROM ted_contracts_production WHERE id = ?
            """, (contract['contract_id'],)).fetchone()

            if not full:
                continue

            # Populate columns
            ws.cell(row=row_num, column=1, value=contract['contract_id'])
            ws.cell(row=row_num, column=2, value=contract.get('notice_number', ''))
            ws.cell(row=row_num, column=3, value=full['publication_date'])
            ws.cell(row=row_num, column=4, value=contract.get('ca_country', ''))
            ws.cell(row=row_num, column=5, value=(full['ca_name'] or '')[:60])
            ws.cell(row=row_num, column=6, value=(full['contractor_name'] or 'N/A')[:40])
            ws.cell(row=row_num, column=7, value=(full['contract_title'] or '')[:80])

            ws.cell(row=row_num, column=8, value=category.replace('_', ' '))
            ws.cell(row=row_num, column=9, value=priority)

            influence_patterns = contract.get('influence_patterns', {})
            ws.cell(row=row_num, column=10, value=', '.join(influence_patterns.keys()) if influence_patterns else 'None')

            reference_patterns = contract.get('reference_patterns', {})
            ws.cell(row=row_num, column=11, value=', '.join(reference_patterns.keys()) if reference_patterns else 'None')

            ws.cell(row=row_num, column=12, value='YES' if contract.get('is_intpa') else 'No')

            ws.cell(row=row_num, column=13, value=full['value_total'])
            ws.cell(row=row_num, column=14, value=full['currency'])

            # Apply color coding
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = row_fill
                ws.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)

        conn.close()

        # Column widths
        column_widths = [8, 12, 12, 8, 40, 30, 50, 22, 12, 30, 30, 12, 12, 8, 25, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes and filter
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    def create_summary_sheet(self, wb, by_category):
        """Create summary sheet"""

        ws = wb.create_sheet(title="SUMMARY", index=0)

        # Title
        ws['A1'] = "TED Chinese Influence - Final Comprehensive Review"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')

        # Metadata
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = "Total Contracts:"
        ws['B4'] = len(self.influence_data.get('contracts', []))
        ws['A5'] = "Data Source:"
        ws['B5'] = self.influence_file.name

        # Intelligence summary
        ws['A7'] = "INTELLIGENCE CATEGORIZATION"
        ws['A7'].font = Font(size=14, bold=True)

        headers = ['Category', 'Count', '%', 'Priority', 'Description', 'Sheet']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        categories_info = [
            ('CHINA_COOPERATION', 'EU-China cooperation programs', 'EU-China Cooperation'),
            ('TRADE_EVENT', 'Trade exhibitions, conferences involving China', 'Trade Events'),
            ('INTPA_RADAR', 'European Commission INTPA contracts (radar)', 'INTPA Radar'),
            ('BRI_RELATED', 'Belt and Road Initiative projects', 'Belt & Road Initiative'),
            ('TRADE_MISSION', 'Trade/business missions with China', 'Trade Missions'),
            ('PROMOTIONAL_EVENT', 'Investment/business promotion events', 'Promotional Events'),
            ('CHINESE_FUNDED', 'Chinese funding indicators', 'Chinese Funding'),
            ('17PLUS1_INITIATIVE', 'China-CEEC 17+1 cooperation', '17+1 Initiative'),
            ('HK_REFERENCE', 'Hong Kong office references', 'Hong Kong Refs'),
            ('GEOGRAPHIC_REFERENCE', 'Geographic mentions only', 'Geographic Only'),
            ('NO_INFLUENCE_DETECTED', 'No influence patterns detected', 'No Influence')
        ]

        total = len(self.influence_data.get('contracts', []))
        row = 9
        for category, description, sheet in categories_info:
            count = len(by_category.get(category, []))
            pct = count / total * 100 if total else 0
            priority, _ = self.get_priority_info(category)

            ws.cell(row=row, column=1, value=category.replace('_', ' '))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
            ws.cell(row=row, column=4, value=priority)
            ws.cell(row=row, column=5, value=description)
            ws.cell(row=row, column=6, value=sheet if count > 0 else 'N/A')

            row += 1

        # Key findings
        ws['A' + str(row + 2)] = "KEY INTELLIGENCE FINDINGS"
        ws['A' + str(row + 2)].font = Font(size=14, bold=True)

        findings = [
            "",
            f"Total Intelligence Contracts: {total - len(by_category.get('NO_INFLUENCE_DETECTED', []))} ({(total - len(by_category.get('NO_INFLUENCE_DETECTED', [])))/total*100:.1f}%)",
            "",
            "HIGH PRIORITY (🔴):",
            f"  • {len(by_category.get('CHINA_COOPERATION', []))} EU-China Cooperation programs",
            f"  • {len(by_category.get('BRI_RELATED', []))} Belt and Road Initiative projects",
            f"  • {len(by_category.get('TRADE_MISSION', []))} Trade missions",
            "",
            "MEDIUM-HIGH PRIORITY (🟠):",
            f"  • {len(by_category.get('TRADE_EVENT', []))} Trade events/exhibitions",
            f"  • {len(by_category.get('PROMOTIONAL_EVENT', []))} Promotional events",
            "",
            "MEDIUM PRIORITY (🟡):",
            f"  • {len(by_category.get('INTPA_RADAR', []))} INTPA radar contracts",
            "",
            "Notable INTPA Contracts:",
            "  • Observatory on China's overseas investments in critical raw materials",
            "  • EU-China Dialogue on ETS policies",
            "  • Understanding Chinese Legal Reform (EUCLERA)",
            "",
            "Trade Event Examples:",
            "  • Shenzhen Trade Event October 2025",
            "  • French Pavilions on Hainan Expo (CICPE)",
            "  • European Fresh Pears campaign in China"
        ]

        for i, finding in enumerate(findings, row + 3):
            ws['A' + str(i)] = finding
            ws['A' + str(i)].alignment = Alignment(wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 25

    def generate(self):
        """Generate final comprehensive Excel"""

        print("\n" + "="*80)
        print("GENERATING FINAL COMPREHENSIVE REVIEW EXCEL")
        print("="*80)
        print()
        print("Combining:")
        print("  - China Cooperation programs (verified)")
        print("  - Trade Events (NEW)")
        print("  - INTPA Radar tracking (NEW)")
        print("  - All other influence patterns")
        print()

        excel_file = self.create_excel()

        print("="*80)
        print("GENERATION COMPLETE")
        print("="*80)
        print()
        print("Review Priority:")
        print("  1. EU-China Cooperation (🔴 HIGH)")
        print("  2. Trade Events (🟠 MED-HIGH)")
        print("  3. INTPA Radar (🟡 MEDIUM)")
        print("  4. No Influence (⚪ NONE)")
        print()

        return excel_file


if __name__ == '__main__':
    generator = FinalComprehensiveExcel()
    excel_file = generator.generate()
