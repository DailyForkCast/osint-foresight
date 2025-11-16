"""
Diagnose Country_Dashboard and Portfolio issues in v48
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v48.xlsx')

print("="*80)
print("DIAGNOSIS")
print("="*80)

# ============================================================================
# ISSUE 1: Country_Dashboard not showing projects
# ============================================================================
print("\n" + "="*80)
print("ISSUE 1: COUNTRY_DASHBOARD - WHY NO PROJECTS?")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nCurrent values:")
print(f"  B2 (Country Name): {ws_country['B2'].value}")
print(f"  G2 (Country Code): {ws_country['G2'].value}")
print(f"  D2 (Region): {ws_country['D2'].value}")

print("\nG2 formula:")
print(f"  {ws_country['G2'].value}")

print("\nD2 formula:")
print(f"  {ws_country['D2'].value}")

print("\nA12 formula (first project):")
print(f"  {ws_country['A12'].value}")

print("\nB12 formula (project name):")
print(f"  {ws_country['B12'].value}")

# Check Country_Budgets for Germany data
ws_budgets = wb['Country_Budgets']
print("\n" + "="*80)
print("CHECKING COUNTRY_BUDGETS FOR GERMANY/FRANCE DATA")
print("="*80)

germany_count = 0
france_count = 0

for row in range(2, min(100, ws_budgets.max_row + 1)):
    country_name = ws_budgets.cell(row, 5).value  # Column E - Country_Name
    country_code = ws_budgets.cell(row, 4).value  # Column D - Country_Code

    if country_name == 'Germany' or country_code == 'DE':
        germany_count += 1
        if germany_count <= 3:
            unique_id = ws_budgets.cell(row, 2).value
            print(f"  Row {row}: {country_name} ({country_code}) - Project: {unique_id}")

    if country_name == 'France' or country_code == 'FR':
        france_count += 1
        if france_count <= 3:
            unique_id = ws_budgets.cell(row, 2).value
            print(f"  Row {row}: {country_name} ({country_code}) - Project: {unique_id}")

print(f"\nTotal Germany entries: {germany_count}")
print(f"Total France entries: {france_count}")

# Check Country_Regions for exact names
ws_regions = wb['Country_Regions']
print("\n" + "="*80)
print("CHECKING COUNTRY_REGIONS FOR EXACT NAMES")
print("="*80)

for row in range(2, min(ws_regions.max_row + 1, 100)):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value

    if code in ['DE', 'FR']:
        print(f"  {code}: '{name}'")

# ============================================================================
# ISSUE 2: Portfolio_Dashboard red in columns L and O
# ============================================================================
print("\n" + "="*80)
print("ISSUE 2: PORTFOLIO_DASHBOARD - CONDITIONAL FORMATTING")
print("="*80)

ws_portfolio = wb['Portfolio_Dashboard']

print("\nConditional formatting rules on Portfolio_Dashboard:")
print(f"  Total rules: {len(ws_portfolio.conditional_formatting._cf_rules)}")

for idx, (range_str, rules) in enumerate(ws_portfolio.conditional_formatting._cf_rules.items()):
    print(f"\n  Rule set {idx + 1}:")
    print(f"    Applies to: {range_str}")
    for rule in rules:
        print(f"    Rule type: {type(rule).__name__}")
        if hasattr(rule, 'formula'):
            print(f"    Formula: {rule.formula}")

wb.close()

print("\n" + "="*80)
print("DIAGNOSIS COMPLETE")
print("="*80)
