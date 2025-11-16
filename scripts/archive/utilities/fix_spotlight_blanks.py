"""
Fix Spotlight_PMWorkspace to show blank instead of 0 for all fields
"""
import openpyxl

print("="*80)
print("FIXING SPOTLIGHT BLANK DISPLAY")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("PART 1: FIX EXISTING FORMULAS")
print("="*80)

print("\nUpdating existing formulas to show blank instead of 0...")

# Map of cells that need fixing
formulas_to_fix = {
    # Column B fields
    'B5': ('Project_Name', 'text'),
    'B6': ('Project_Priority', 'text'),
    'B7': ('Project_Start_Date', 'date'),
    'B8': ('POP_Days_Remaining', 'numeric'),
    'B9': ('Implementer_POC', 'text'),
    'B10': ('Countries', 'text'),
    'B11': ('Project_Summary', 'text'),

    # Financial fields
    'B14': ('Total_Proposed', 'numeric'),
    'E14': ('Total_Obligated', 'numeric'),
    'B15': ('Total_Allocation', 'numeric'),
    'E15': ('Total_Spent', 'numeric'),

    # Award/Date fields
    'F2': ('Award_Number', 'text'),
    'F3': ('POP_Start', 'date'),
}

for cell_ref, (field_name, field_type) in formulas_to_fix.items():
    cell = ws_spotlight[cell_ref]

    # Current formula pattern: =IFERROR(INDEX(T_Master_Projects[Field],MATCH(...)),"")
    # New pattern: =IF($B$2="","",IFERROR(IF(INDEX(...)="","",INDEX(...)),""))

    if field_type == 'numeric':
        # For numeric fields, check for 0
        new_formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))=0,"",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'
    else:
        # For text/date fields, check for empty string
        new_formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'

    cell.value = new_formula
    print(f"  {cell_ref}: Updated ({field_type})")

print("\nPart 1 complete - existing formulas updated")

print("\n" + "="*80)
print("PART 2: ADD MISSING F COLUMN FORMULAS")
print("="*80)

print("\nAdding formulas for F column fields...")

# F column fields that need formulas
f_column_fields = {
    'F5': ('Project_Status', 'text'),
    'F6': ('Project_Progress', 'numeric'),  # User specifically mentioned Progress
    'F7': ('POP_End', 'date'),
    'F8': ('Implementer_Organization', 'text'),  # User specifically mentioned Implementer
    'F9': ('Implementer_POC_Email', 'text'),
    'F10': ('Country_Count', 'numeric'),
}

for cell_ref, (field_name, field_type) in f_column_fields.items():
    if field_type == 'numeric':
        # For numeric fields, check for 0
        formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))=0,"",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'
    else:
        # For text/date fields, check for empty string
        formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'

    ws_spotlight[cell_ref] = formula
    print(f"  {cell_ref}: {field_name} ({field_type})")

print("\nPart 2 complete - F column formulas added")

print("\n" + "="*80)
print("PART 3: CHECK FOR NCE FIELDS")
print("="*80)

print("\nSearching for NCE fields in rows 1-20...")

# Search for NCE fields
nce_found = False
for row in range(1, 21):
    for col in range(1, 7):
        cell = ws_spotlight.cell(row, col)
        if cell.value and isinstance(cell.value, str):
            if 'nce' in cell.value.lower():
                col_letter = openpyxl.utils.get_column_letter(col)
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
                print(f"  Found at {col_letter}{row}: {clean_value}")
                nce_found = True

if not nce_found:
    print("  No NCE fields found in rows 1-20")
    print("  NCE fields are in Master_Projects columns N-O:")
    print("    - NCE_Eligible (Yes/No)")
    print("    - NCE_Status (None, NCE 1 Approved, etc.)")
    print("\n  Should these be added to Spotlight?")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Spotlight formulas updated!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed existing formulas (blank instead of 0):")
for cell_ref, (field_name, field_type) in formulas_to_fix.items():
    print(f"  {cell_ref}: {field_name}")

print("\nAdded missing F column formulas:")
for cell_ref, (field_name, field_type) in f_column_fields.items():
    print(f"  {cell_ref}: {field_name}")

print("\nAll fields now show BLANK when:")
print("  - No project selected ($B$2 is empty)")
print("  - Value is 0 (for numeric fields)")
print("  - Value is empty (for text/date fields)")

print("\nv28 updated!")
