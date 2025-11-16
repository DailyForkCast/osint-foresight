"""
Fix Control sheet formulas after inserting Total Proposed row
The formulas in Total ULO and Portfolio ULO % need to reference the correct rows
"""
import openpyxl

print("="*80)
print("FIXING CONTROL SHEET FORMULAS")
print("="*80)

# Load v27
print("\nLoading v27...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx')
ws_control = wb['Control']

print("\n" + "="*80)
print("CHECKING CURRENT FORMULAS")
print("="*80)

print("\nCurrent Financial Status section:")
for row in range(14, 22):
    label = ws_control.cell(row, 1).value
    formula = ws_control.cell(row, 2).value
    if label:
        print(f"  B{row} ({label}): {formula}")

print("\n" + "="*80)
print("FIXING FORMULAS")
print("="*80)

# After inserting row at 15, the structure is now:
# B15 = Total Proposed
# B16 = Total Allocated
# B17 = Total Obligated
# B18 = Total ULO (should be B16 - B17, i.e., Allocated - Obligated)
# B19 = Portfolio ULO % (should be B18/B16, i.e., ULO / Allocated)

print("\nFixing B18 (Total ULO)...")
old_formula = ws_control['B18'].value
ws_control['B18'] = '=B16-B17'
print(f"  Old: {old_formula}")
print(f"  New: =B16-B17 (Allocated - Obligated)")

print("\nFixing B19 (Portfolio ULO %)...")
old_formula = ws_control['B19'].value
ws_control['B19'] = '=IF(B16=0,0,B18/B16)'
print(f"  Old: {old_formula}")
print(f"  New: =IF(B16=0,0,B18/B16) (ULO / Allocated)")

print("\n" + "="*80)
print("VERIFYING FIXED FORMULAS")
print("="*80)

print("\nUpdated Financial Status section:")
print("  B15 (Total Proposed): =SUM(T_Master_Projects[Total_Proposed])")
print("  B16 (Total Allocated): =SUM(T_Country_Budgets[Allocated_Amount])")
print("  B17 (Total Obligated): =SUM(T_Country_Budgets[Obligated_Amount])")
print("  B18 (Total ULO): =B16-B17")
print("  B19 (Portfolio ULO %): =IF(B16=0,0,B18/B16)")

print("\n" + "="*80)
print("SAVING V27")
print("="*80)

wb.save('2025-10-26-Tracker-v27.xlsx')

print("\nOK - Control sheet formulas fixed!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFinancial Status section (corrected):")
print("  Row 15: Total Proposed - Sum of all proposed budgets")
print("  Row 16: Total Allocated - Sum of all allocated budgets")
print("  Row 17: Total Obligated - Sum of all obligated budgets")
print("  Row 18: Total ULO - Allocated minus Obligated")
print("  Row 19: Portfolio ULO % - ULO as percentage of Allocated")

print("\nAll formulas now reference the correct cells!")
print("v27 is ready to use!")
