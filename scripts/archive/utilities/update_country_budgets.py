import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("UPDATING COUNTRY_BUDGETS STRUCTURE")
print("=" * 70)

# Load the workbook
print("\n[1/3] Loading workbook...")
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-CLEAN.xlsx')
ws = wb['Country_Budgets']
print("✓ Workbook loaded")

print("\n[2/3] Restructuring Country_Budgets columns...")

# Current structure (from inspection):
# A: Budget_ID
# B: Unique_ID
# C: Country_Code
# D: Country_Name
# E: Allocated_Amount
# F: Obligated_Amount
# G: ULO (currently = E-F, but should be F-Spent)
# H: ULO_Percent
# I: Spend_Health

# NEW structure should be:
# A: Budget_ID
# B: Unique_ID (Project_ID)
# C: Country_Code
# D: Country_Name
# E: Allocated_Amount (budgeted)
# F: Obligated_Amount (received)
# G: Spent_Amount (NEW - actually disbursed)
# H: ULO (Obligated - Spent) (FIXED FORMULA)
# I: ULO_Percent (ULO / Obligated) (FIXED FORMULA)
# J: Funding_Gap (NEW - Allocated - Obligated)
# K: Spend_Health

# Insert new "Spent" column at position G
ws.insert_cols(7)  # Insert at column G
print("✓ Inserted 'Spent' column at position G")

# Insert new "Funding_Gap" column at position K (after ULO_Percent)
ws.insert_cols(11)  # Insert at column K
print("✓ Inserted 'Funding_Gap' column at position K")

# Update headers
ws.cell(row=1, column=7, value='Spent')
ws.cell(row=1, column=7).font = Font(bold=True)
ws.cell(row=1, column=7).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

ws.cell(row=1, column=8, value='ULO')
ws.cell(row=1, column=9, value='ULO_Percent')
ws.cell(row=1, column=10, value='Funding_Gap')
ws.cell(row=1, column=10).font = Font(bold=True)
ws.cell(row=1, column=10).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

ws.cell(row=1, column=11, value='Spend_Health')

print("✓ Updated column headers")

print("\n[3/3] Adding formulas and sample data...")

# For each existing row, add formulas
max_row = ws.max_row
for row in range(2, max_row + 1):
    allocated = ws.cell(row=row, column=5).value  # Column E
    obligated = ws.cell(row=row, column=6).value  # Column F

    # Add sample Spent data (50-90% of obligated)
    if obligated and isinstance(obligated, (int, float)) and obligated > 0:
        import random
        spent_amount = int(obligated * random.uniform(0.5, 0.9))
        ws.cell(row=row, column=7, value=spent_amount)  # Spent
    else:
        ws.cell(row=row, column=7, value=0)

    # ULO = Obligated - Spent (column H)
    ws.cell(row=row, column=8, value=f'=F{row}-G{row}')

    # ULO_Percent = ULO / Obligated (column I)
    ws.cell(row=row, column=9, value=f'=IF(F{row}=0,0,H{row}/F{row})')
    ws.cell(row=row, column=9).number_format = '0%'

    # Funding_Gap = Allocated - Obligated (column J)
    ws.cell(row=row, column=10, value=f'=E{row}-F{row}')

    # Spend_Health (column K) - keep existing formula or create new
    # Green if ULO% is reasonable (20-40%), Red if too low (<10%) or too high (>60%)
    ws.cell(row=row, column=11, value=f'=IF(I{row}<0.1,"Low Execution",IF(I{row}>0.6,"Slow Execution","On Track"))')

print(f"✓ Updated {max_row - 1} rows with formulas and sample data")

# Adjust column widths
ws.column_dimensions['A'].width = 15  # Budget_ID
ws.column_dimensions['B'].width = 12  # Unique_ID
ws.column_dimensions['C'].width = 12  # Country_Code
ws.column_dimensions['D'].width = 20  # Country_Name
ws.column_dimensions['E'].width = 15  # Allocated
ws.column_dimensions['F'].width = 15  # Obligated
ws.column_dimensions['G'].width = 15  # Spent
ws.column_dimensions['H'].width = 15  # ULO
ws.column_dimensions['I'].width = 12  # ULO_Percent
ws.column_dimensions['J'].width = 15  # Funding_Gap
ws.column_dimensions['K'].width = 15  # Spend_Health

print("✓ Adjusted column widths")

# Save
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-CLEAN.xlsx'
print(f"\n{'='*70}")
print("Saving updated workbook...")
wb.save(output_file)
print(f"✓ Saved: {output_file}")

print(f"\n{'='*70}")
print("COUNTRY_BUDGETS UPDATE COMPLETE!")
print(f"{'='*70}")
print("\nNew Country_Budgets Structure:")
print("  A: Budget_ID")
print("  B: Unique_ID (Project_ID)")
print("  C: Country_Code")
print("  D: Country_Name")
print("  E: Allocated (budgeted amount)")
print("  F: Obligated (amount received)")
print("  G: Spent (amount disbursed) ← NEW")
print("  H: ULO (Obligated - Spent) ← FIXED FORMULA")
print("  I: ULO_Percent (ULO / Obligated) ← FIXED FORMULA")
print("  J: Funding_Gap (Allocated - Obligated) ← NEW")
print("  K: Spend_Health (execution status)")
print("\nDefinitions:")
print("  • Allocated = Amount budgeted for project")
print("  • Obligated = Amount actually received (have in hand)")
print("  • Spent = Amount actually disbursed/used")
print("  • ULO = Obligated - Spent (money on hand but not spent)")
print("  • Funding Gap = Allocated - Obligated (promised but not received)")
print("\nNOTE: Sample 'Spent' data added (50-90% of Obligated)")
print("      Replace with actual spending data")
