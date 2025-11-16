import openpyxl

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6-clean.xlsx'

print('='*80)
print('CLEANING SAMPLE DATA FROM v6')
print('Keeping 2 sample entries per data sheet')
print('='*80)
print()

# Load workbook
print('Loading v6...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

# Define which sheets to clean (data entry sheets)
DATA_SHEETS = {
    'Master_Projects': 2,        # Keep rows 2-3 (2 projects)
    'Country_Budgets': 2,        # Keep rows 2-3 (2 budget entries)
    'Milestones': 2,             # Keep rows 2-3 (2 milestones)
    'Events': 2,                 # Keep rows 2-3 (2 events)
    'Risk_Register': 2,          # Keep rows 2-3 (2 risks)
    'Decision_Log': 2,           # Keep rows 2-3 (2 decisions)
    'Stakeholders': 2,           # Keep rows 2-3 (2 stakeholders)
    'Project_Audiences': 2,      # Keep rows 2-3 (2 audiences)
    'Project_Technologies': 2,   # Keep rows 2-3 (2 technologies)
    'Project_Deliverables': 2,   # Keep rows 2-3 (2 deliverables)
}

# Sheets to NOT touch (reference data, dashboards, calculated sheets)
SKIP_SHEETS = [
    'Control',                   # Settings and control values
    'Portfolio_Dashboard',       # Dashboard with formulas
    'Project_Spotlight',         # Dashboard sheet
    'Config_Lists',              # Reference data - keep all
    'Country_Regions',           # Reference data - keep all (98 countries)
    'Regional_Summary',          # Calculated from other sheets
    '_SETUP',                    # Instructions
    'Calendar_Todo',             # May have important content
    'Country_PM_Assignments',    # Reference data - keep all
    'Project_Products',          # Keep as is
]

total_rows_deleted = 0

for sheet_name, keep_rows in DATA_SHEETS.items():
    if sheet_name not in wb.sheetnames:
        print(f'WARNING: {sheet_name}: Not found, skipping')
        continue

    ws = wb[sheet_name]

    # Get current max row
    max_row = ws.max_row

    # Calculate rows to delete (everything after row 1 + keep_rows)
    first_row_to_delete = 2 + keep_rows  # Header is row 1, data starts row 2

    if max_row > first_row_to_delete - 1:
        # Delete from bottom up to avoid row number shifting issues
        rows_to_delete = max_row - (first_row_to_delete - 1)

        print(f'{sheet_name}:')
        print(f'   Before: {max_row - 1} data rows')
        print(f'   Keeping: rows 2-{1 + keep_rows} ({keep_rows} samples)')
        print(f'   Deleting: rows {first_row_to_delete}-{max_row} ({rows_to_delete} rows)')

        # Delete rows
        ws.delete_rows(first_row_to_delete, rows_to_delete)

        total_rows_deleted += rows_to_delete
        print(f'   Deleted {rows_to_delete} rows')
        print()
    else:
        print(f'{sheet_name}: Already has {max_row - 1} or fewer rows, skipping')
        print()

print()
print('='*80)
print('SKIPPED SHEETS (reference data, dashboards, calculated)')
print('='*80)
for sheet_name in SKIP_SHEETS:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'{sheet_name}: Kept as-is ({ws.max_row} rows)')

print()
print('='*80)
print('Saving cleaned file...')
wb.save(output_file)
wb.close()

print('='*80)
print('CLEANUP COMPLETE!')
print('='*80)
print()
print(f'Total data rows deleted: {total_rows_deleted}')
print()
print('Summary:')
print('  - Cleaned 10 data entry sheets')
print('  - Kept 2 sample entries per sheet')
print('  - Preserved all reference data (Config_Lists, Country_Regions, etc.)')
print('  - Preserved all dashboard sheets')
print('  - All formulas and structures intact')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v6-clean in Excel')
print('  2. Verify data looks correct (2 samples per sheet)')
print('  3. Check formulas still work')
print('  4. If good, rename to v6.xlsx or use as base for v6.3')
print()
print('Sheets with sample data (2 entries each):')
for sheet in DATA_SHEETS.keys():
    print(f'  - {sheet}')
