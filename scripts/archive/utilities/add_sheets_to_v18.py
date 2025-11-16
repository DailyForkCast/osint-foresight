"""
Copy sheets from v12 to v18, excluding Risk_Register, Decision_Log, Project_Products
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from copy import copy

print("="*80)
print("COPYING SHEETS FROM V12 TO V18")
print("="*80)

# Load both workbooks
print("\nLoading workbooks...")
wb_old = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=True)
wb_new = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

# Sheets to skip
SKIP_SHEETS = {
    'Risk_Register',
    'Decision_Log',
    'Project_Products',
    # Already in v18:
    'Master_Projects',
    'Country_Budgets',
    'Country_Regions',
    'Config_Lists',
    'Country_PM_Assignments'
}

# Get sheets to copy
sheets_to_copy = [name for name in wb_old.sheetnames if name not in SKIP_SHEETS]

print(f"\nCopying {len(sheets_to_copy)} sheets:")
for name in sheets_to_copy:
    print(f"  - {name}")

print("\nCopying sheets...")
for sheet_name in sheets_to_copy:
    print(f"\n  {sheet_name}...", end=" ")

    ws_old = wb_old[sheet_name]
    ws_new = wb_new.create_sheet(sheet_name)

    # Copy column widths
    for col_letter in ws_old.column_dimensions:
        if ws_old.column_dimensions[col_letter].width:
            ws_new.column_dimensions[col_letter].width = ws_old.column_dimensions[col_letter].width

    # Copy row heights
    for row_num in ws_old.row_dimensions:
        if ws_old.row_dimensions[row_num].height:
            ws_new.row_dimensions[row_num].height = ws_old.row_dimensions[row_num].height

    # Copy cells
    for row in ws_old.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column)

            # Copy value
            new_cell.value = cell.value

            # Copy formatting
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.border = copy(cell.border)
                    new_cell.number_format = cell.number_format
                except:
                    pass  # Skip if formatting fails

    # Copy merged cells
    for merged_range in ws_old.merged_cells.ranges:
        try:
            ws_new.merge_cells(str(merged_range))
        except:
            pass  # Skip if merge fails

    print("Done!")

print("\nSaving v18...")
wb_new.save('2025-10-26-Tracker-v18.xlsx')

print("\n" + "="*80)
print("COPY COMPLETE!")
print("="*80)

print("\nV18 now has the following sheets:")
wb_check = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')
for i, name in enumerate(wb_check.sheetnames, 1):
    print(f"  {i:2}. {name}")

print(f"\nTotal: {len(wb_check.sheetnames)} sheets")
print("\nSkipped sheets:")
for name in sorted(SKIP_SHEETS):
    if name in wb_old.sheetnames:
        print(f"  - {name}")
