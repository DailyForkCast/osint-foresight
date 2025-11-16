"""
Fix Portfolio_Dashboard column O formatting - v50
Add blank check to prevent formatting on empty cells
"""
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule

print("="*80)
print("FIXING COLUMN O FORMATTING - V50")
print("="*80)

# Load v49
print("\nLoading v49...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v49.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

# Define colors
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_text = Font(color='9C0006')
orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_text = Font(color='9C6500')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_text = Font(color='006100')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

print("\n" + "="*80)
print("CLEARING ALL CONDITIONAL FORMATTING")
print("="*80)

print("\nClearing all existing rules...")
ws_portfolio.conditional_formatting._cf_rules.clear()
print("  All rules cleared")

print("\n" + "="*80)
print("RE-ADDING CONDITIONAL FORMATTING WITH BLANK CHECKS")
print("="*80)

# 1. ULO % Alert (Column O) - Red if > 75% AND not blank
print("\n1. ULO % alert (O11:O20) - Red if >75% AND has data...")
rule = FormulaRule(
    formula=['AND($O11<>"",$O11>0.75)'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('O11:O20', rule)
print("   Added: ULO % > 75% = Red (only if cell has data)")

# 2. Days Remaining Alert (Column P) - Red if < 0 AND not blank
print("\n2. Days Remaining negative alert (P11:P20) - Red if <0 AND has data...")
rule = FormulaRule(
    formula=['AND($P11<>"",$P11<0)'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('P11:P20', rule)
print("   Added: Days < 0 = Red (only if cell has data)")

# 3. Priority Color Coding (Column D)
print("\n3. Priority color coding (D11:D20)...")
# Critical = Red
rule = CellIsRule(
    operator='equal',
    formula=['"Critical"'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# High = Orange
rule = CellIsRule(
    operator='equal',
    formula=['"High"'],
    stopIfTrue=True,
    fill=orange_fill,
    font=orange_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# Medium = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"Medium"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# Low = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Low"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)
print("   Added: Priority colors (Critical=Red, High=Orange, Medium=Yellow, Low=Green)")

# 4. Status Color Coding (Column C)
print("\n4. Status color coding (C11:C20)...")
# Active = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Active"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)

# On Hold = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"On Hold"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)

# Completed = Light Green
rule = CellIsRule(
    operator='equal',
    formula=['"Completed"'],
    stopIfTrue=True,
    fill=light_green_fill
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)
print("   Added: Status colors (Active=Green, On Hold=Yellow, Completed=Light Green)")

print("\n" + "="*80)
print("SAVING V50")
print("="*80)

wb.save('2025-10-26-Tracker-v50.xlsx')

print("\nOK - v50 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed Portfolio_Dashboard conditional formatting:")
print("\n  Column O (ULO %):")
print("    - NOW: Only shows red if cell has data AND value > 75%")
print("    - Uses formula: AND($O11<>\"\", $O11>0.75)")
print("    - Empty cells stay white")

print("\n  Column P (Days Remaining):")
print("    - NOW: Only shows red if cell has data AND value < 0")
print("    - Uses formula: AND($P11<>\"\", $P11<0)")
print("    - Empty cells stay white")

print("\n  Columns C & D (Status & Priority):")
print("    - Unchanged - still working correctly")

print("\nv50 ready - column O should now work correctly!")
