#!/usr/bin/env python3
"""
Comprehensive TED Review Excel Generator
Combines:
  1. Contractor detection heuristics (True/False Chinese contractor)
  2. Intelligence influence patterns (BRI, Cooperation, Trade Missions)

Provides full context for manual review
"""

import sqlite3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from collections import defaultdict


class ComprehensiveReviewGenerator:
    """Generate comprehensive Excel for manual review with all categorizations"""

    def __init__(self):
        self.db_path = Path("F:/OSINT_WAREHOUSE/osint_master.db")

        # Load revalidation results
        self.revalidation_file = self.find_latest_file("ted_revalidation_detailed_*.json")
        with open(self.revalidation_file, 'r', encoding='utf-8') as f:
            self.revalidation_results = json.load(f)

        # Load influence analysis
        self.influence_file = self.find_latest_file("ted_chinese_influence_analysis_*.json")
        with open(self.influence_file, 'r', encoding='utf-8') as f:
            self.influence_data = json.load(f)

        # Load deep analysis if exists
        try:
            self.deep_analysis_file = self.find_latest_file("ted_uncertain_deep_analysis_*.json")
            with open(self.deep_analysis_file, 'r', encoding='utf-8') as f:
                self.deep_analysis = json.load(f)
        except:
            self.deep_analysis = None

        # Create lookup for influence by contract ID
        self.influence_by_id = {}
        for contract in self.influence_data.get('contracts', []):
            self.influence_by_id[contract['contract_id']] = contract

    def find_latest_file(self, pattern):
        """Find most recent file matching pattern"""
        analysis_dir = Path("analysis")
        files = list(analysis_dir.glob(pattern))
        if not files:
            raise FileNotFoundError(f"No files found for: {pattern}")
        return max(files, key=lambda p: p.stat().st_mtime)

    def get_priority_emoji(self, category):
        """Get priority indicator for intelligence category"""
        priority_map = {
            'BRI_RELATED': '🔴 HIGH',
            'CHINA_COOPERATION': '🔴 HIGH',
            'CHINESE_FUNDED': '🟠 MED-HIGH',
            '17PLUS1_INITIATIVE': '🟠 MED-HIGH',
            'HK_REFERENCE': '🟡 LOW-MED',
            'GEOGRAPHIC_REFERENCE': '🟢 LOW',
            'NO_INFLUENCE_DETECTED': '⚪ NONE'
        }
        return priority_map.get(category, '')

    def get_deep_analysis_category(self, contract_id):
        """Get deep analysis category if available"""
        if not self.deep_analysis:
            return None, None

        for category, contracts in self.deep_analysis.get('categories', {}).items():
            for contract in contracts:
                if contract['id'] == contract_id:
                    analysis = contract.get('enhanced_analysis', {})
                    return category, analysis
        return None, None

    def create_comprehensive_excel(self):
        """Create comprehensive Excel with all categorizations"""

        print("\n" + "="*80)
        print("COMPREHENSIVE REVIEW EXCEL GENERATOR")
        print("="*80)
        print()

        # Get uncertain contracts only
        uncertain_contracts = [
            r for r in self.revalidation_results
            if r['category'] == 'UNCERTAIN'
        ]

        print(f"Processing {len(uncertain_contracts)} uncertain contracts...")
        print()

        # Categorize by intelligence priority
        by_priority = defaultdict(list)

        for contract in uncertain_contracts:
            contract_id = contract['id']
            influence = self.influence_by_id.get(contract_id, {})
            intel_category = influence.get('intelligence_category', 'NO_INFLUENCE_DETECTED')

            deep_category, deep_analysis = self.get_deep_analysis_category(contract_id)

            # Combine data
            combined = {
                'contract': contract,
                'influence': influence,
                'deep_category': deep_category,
                'deep_analysis': deep_analysis or {}
            }

            by_priority[intel_category].append(combined)

        # Create Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Priority order for sheets
        priority_order = [
            ('BRI_RELATED', 'BRI - Belt & Road'),
            ('CHINA_COOPERATION', 'EU-China Cooperation'),
            ('CHINESE_FUNDED', 'Chinese Funding'),
            ('17PLUS1_INITIATIVE', '17+1 Initiative'),
            ('HK_REFERENCE', 'Hong Kong References'),
            ('GEOGRAPHIC_REFERENCE', 'Geographic Only'),
            ('NO_INFLUENCE_DETECTED', 'No Influence')
        ]

        for category, sheet_name in priority_order:
            contracts = by_priority.get(category, [])
            if not contracts:
                continue

            self.create_category_sheet(wb, sheet_name, category, contracts)

        # Create summary sheet
        self.create_summary_sheet(wb, by_priority, uncertain_contracts)

        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = Path(f"analysis/ted_comprehensive_review_{timestamp}.xlsx")
        wb.save(excel_file)

        print(f"[SUCCESS] Comprehensive Excel created: {excel_file}")
        print(f"  File size: {excel_file.stat().st_size / 1024:.1f} KB")
        print()

        return excel_file

    def create_category_sheet(self, wb, sheet_name, category, contracts):
        """Create sheet for specific intelligence category"""

        ws = wb.create_sheet(title=sheet_name[:31])

        # Headers
        headers = [
            'ID', 'Notice', 'Date', 'Country', 'Contracting Authority',
            'Contractor Name', 'Contract Title', 'Value', 'Currency',
            'INTELLIGENCE CATEGORY', 'PRIORITY',
            'Influence Patterns', 'Reference Patterns',
            'CONTRACTOR ANALYSIS', 'Heuristic Category', 'Confidence',
            'Detection Reasons',
            'MANUAL CLASSIFICATION', 'NOTES'
        ]

        # Style headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add data
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        for row_num, item in enumerate(contracts, 2):
            contract = item['contract']
            influence = item['influence']
            deep_analysis = item['deep_analysis']

            # Get full contract data
            cursor = conn.cursor()
            full = cursor.execute("""
                SELECT ca_name, ca_country, contract_title, contract_description,
                       contractor_name, value_total, currency, publication_date
                FROM ted_contracts_production WHERE id = ?
            """, (contract['id'],)).fetchone()

            if not full:
                continue

            # Populate columns
            ws.cell(row=row_num, column=1, value=contract['id'])
            ws.cell(row=row_num, column=2, value=contract.get('notice_number', ''))
            ws.cell(row=row_num, column=3, value=full['publication_date'])
            ws.cell(row=row_num, column=4, value=full['ca_country'])
            ws.cell(row=row_num, column=5, value=(full['ca_name'] or '')[:60])
            ws.cell(row=row_num, column=6, value=(full['contractor_name'] or '')[:50])
            ws.cell(row=row_num, column=7, value=(full['contract_title'] or '')[:80])
            ws.cell(row=row_num, column=8, value=full['value_total'])
            ws.cell(row=row_num, column=9, value=full['currency'])

            # Intelligence categorization
            ws.cell(row=row_num, column=10, value=category.replace('_', ' '))
            ws.cell(row=row_num, column=11, value=self.get_priority_emoji(category))

            # Influence patterns
            influence_patterns = influence.get('influence_patterns', {})
            ws.cell(row=row_num, column=12, value=', '.join(influence_patterns.keys()) if influence_patterns else 'None')

            reference_patterns = influence.get('reference_patterns', {})
            ws.cell(row=row_num, column=13, value=', '.join(reference_patterns.keys()) if reference_patterns else 'None')

            # Contractor analysis
            ws.cell(row=row_num, column=14, value=contract.get('category', 'UNCERTAIN'))
            ws.cell(row=row_num, column=15, value=deep_analysis.get('final_category', 'N/A'))
            ws.cell(row=row_num, column=16, value=deep_analysis.get('confidence', ''))
            ws.cell(row=row_num, column=17, value=deep_analysis.get('primary_reason', contract.get('reasoning', '')))

            # Color code rows by priority
            if category in ['BRI_RELATED', 'CHINA_COOPERATION']:
                fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Red tint
            elif category in ['CHINESE_FUNDED', '17PLUS1_INITIATIVE']:
                fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # Orange tint
            elif category == 'HK_REFERENCE':
                fill = PatternFill(start_color="FFFCE6", end_color="FFFCE6", fill_type="solid")  # Yellow tint
            else:
                fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Gray tint

            for col in range(1, 20):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)

        conn.close()

        # Column widths
        column_widths = [8, 12, 12, 8, 35, 30, 50, 12, 8, 22, 12, 25, 25, 18, 22, 12, 40, 25, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes
        ws.freeze_panes = 'A2'

        # Add filter
        ws.auto_filter.ref = ws.dimensions

    def create_summary_sheet(self, wb, by_priority, all_contracts):
        """Create summary overview sheet"""

        ws = wb.create_sheet(title="SUMMARY", index=0)

        # Title
        ws['A1'] = "TED Uncertain Contracts - Comprehensive Review"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        # Metadata
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = "Total Contracts:"
        ws['B4'] = len(all_contracts)
        ws['A5'] = "Revalidation File:"
        ws['B5'] = self.revalidation_file.name
        ws['A6'] = "Influence File:"
        ws['B6'] = self.influence_file.name

        # Intelligence categorization breakdown
        ws['A8'] = "INTELLIGENCE CATEGORIZATION"
        ws['A8'].font = Font(size=14, bold=True)

        headers = ['Category', 'Count', '%', 'Priority', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        categories_info = [
            ('BRI_RELATED', '🔴 HIGH', 'Belt and Road Initiative projects/studies'),
            ('CHINA_COOPERATION', '🔴 HIGH', 'EU-China cooperation programs, trade missions'),
            ('CHINESE_FUNDED', '🟠 MED-HIGH', 'Chinese funding indicators'),
            ('17PLUS1_INITIATIVE', '🟠 MED-HIGH', 'China-CEEC 17+1 cooperation'),
            ('HK_REFERENCE', '🟡 LOW-MED', 'Hong Kong office references'),
            ('GEOGRAPHIC_REFERENCE', '🟢 LOW', 'Geographic mentions only'),
            ('NO_INFLUENCE_DETECTED', '⚪ NONE', 'No influence patterns detected')
        ]

        row = 10
        for category, priority, description in categories_info:
            count = len(by_priority.get(category, []))
            pct = count / len(all_contracts) * 100 if all_contracts else 0

            ws.cell(row=row, column=1, value=category.replace('_', ' '))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
            ws.cell(row=row, column=4, value=priority)
            ws.cell(row=row, column=5, value=description)

            row += 1

        # Instructions
        ws['A' + str(row + 2)] = "MANUAL REVIEW INSTRUCTIONS"
        ws['A' + str(row + 2)].font = Font(size=14, bold=True)

        instructions = [
            "1. Review contracts by priority (HIGH priority sheets first)",
            "2. For each contract, determine:",
            "   - Is the contractor actually Chinese? (Column N)",
            "   - Does this represent Chinese influence/participation? (Columns J-M)",
            "3. Fill in 'MANUAL CLASSIFICATION' column with your assessment",
            "4. Add notes in 'NOTES' column for uncertain cases",
            "5. Color coding:",
            "   - Red tint = HIGH priority (BRI, Cooperation)",
            "   - Orange tint = MEDIUM-HIGH priority (Funding, 17+1)",
            "   - Yellow tint = LOW-MEDIUM priority (HK references)",
            "   - Gray tint = LOW priority (Geographic only)"
        ]

        for i, instruction in enumerate(instructions, row + 3):
            ws['A' + str(i)] = instruction
            ws['A' + str(i)].alignment = Alignment(wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50

    def generate_report(self):
        """Generate comprehensive review Excel"""

        print("\n" + "="*80)
        print("GENERATING COMPREHENSIVE REVIEW EXCEL")
        print("="*80)
        print()
        print("Combining:")
        print("  - Revalidation results (True/False Chinese contractors)")
        print("  - Intelligence influence patterns (BRI, Cooperation, etc.)")
        print("  - Deep heuristic analysis")
        print()

        excel_file = self.create_comprehensive_excel()

        print("="*80)
        print("GENERATION COMPLETE")
        print("="*80)
        print()
        print("Next Steps:")
        print("  1. Open the Excel file for manual review")
        print("  2. Start with HIGH priority sheets (BRI, Cooperation)")
        print("  3. Compare automated categorizations vs your assessment")
        print("  4. Fill in MANUAL CLASSIFICATION and NOTES columns")
        print()

        return excel_file


if __name__ == '__main__':
    generator = ComprehensiveReviewGenerator()
    excel_file = generator.generate_report()
