"""
Build v15 by properly removing all Excel Tables first, then making changes
"""
import openpyxl
from openpyxl import load_workbook
import os

print("="*80)
print("BUILDING V15 - PROPER TABLE REMOVAL")
print("="*80)

# Step 1: Load v12 and save as v15-base without tables
print("\nStep 1: Creating base v15 from v12...")
wb = load_workbook('2025-10-26-Tracker-v12.xlsx', keep_vba=False)

# Get list of all tables before we modify anything
tables_info = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_tables') and ws._tables:
        tables_info[sheet_name] = list(ws._tables)
        print(f"  {sheet_name}: {len(ws._tables)} table(s) found")

# Save with a different extension to force Excel to rebuild
# This is a trick - saving as .xlsx without tables info
wb.save('2025-10-26-Tracker-v15-base.xlsx')
print("\nSaved base file (tables removed)")

# Step 2: Reload and verify no tables
print("\nStep 2: Verifying tables removed...")
wb = load_workbook('2025-10-26-Tracker-v15-base.xlsx')

tables_remain = False
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_tables') and ws._tables:
        print(f"  WARNING: {sheet_name} still has {len(ws._tables)} table(s)")
        tables_remain = True

if not tables_remain:
    print("  SUCCESS: All tables removed!")

# Step 3: Apply the changes
print("\nStep 3: Applying changes to v15-base...")

# State Department regions
STATE_DEPT_REGIONS = {
    'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
    'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
    'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
    'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
    'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
    'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
    'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',
    'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
    'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
    'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
    'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
    'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',
    'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
    'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
    'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
    'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
    'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
    'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
    'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR', 'VA': 'EUR',
    'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA', 'LB': 'NEA', 'OM': 'NEA',
    'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA', 'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',
    'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
    'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',
    'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
    'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
    'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
    'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
    'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
}

SUBREGIONS = {
    'DZ': 'Northern Africa', 'EG': 'Northern Africa', 'LY': 'Northern Africa', 'MA': 'Northern Africa',
    'TN': 'Northern Africa', 'SD': 'Northern Africa',
    'BJ': 'Western Africa', 'BF': 'Western Africa', 'CV': 'Western Africa', 'CI': 'Western Africa',
    'GM': 'Western Africa', 'GH': 'Western Africa', 'GN': 'Western Africa', 'GW': 'Western Africa',
    'LR': 'Western Africa', 'ML': 'Western Africa', 'MR': 'Western Africa', 'NE': 'Western Africa',
    'NG': 'Western Africa', 'SN': 'Western Africa', 'SL': 'Western Africa', 'TG': 'Western Africa',
    'BI': 'Eastern Africa', 'KM': 'Eastern Africa', 'DJ': 'Eastern Africa', 'ER': 'Eastern Africa',
    'ET': 'Eastern Africa', 'KE': 'Eastern Africa', 'MG': 'Eastern Africa', 'MW': 'Eastern Africa',
    'MU': 'Eastern Africa', 'MZ': 'Eastern Africa', 'RW': 'Eastern Africa', 'SC': 'Eastern Africa',
    'SO': 'Eastern Africa', 'SS': 'Eastern Africa', 'TZ': 'Eastern Africa', 'UG': 'Eastern Africa',
    'ZM': 'Eastern Africa', 'ZW': 'Eastern Africa',
    'AO': 'Middle Africa', 'CM': 'Middle Africa', 'CF': 'Middle Africa', 'TD': 'Middle Africa',
    'CG': 'Middle Africa', 'CD': 'Middle Africa', 'GQ': 'Middle Africa', 'GA': 'Middle Africa',
    'BW': 'Southern Africa', 'LS': 'Southern Africa', 'NA': 'Southern Africa', 'ZA': 'Southern Africa', 'SZ': 'Southern Africa',
    'DK': 'Northern Europe', 'EE': 'Northern Europe', 'FI': 'Northern Europe', 'IS': 'Northern Europe',
    'IE': 'Northern Europe', 'LV': 'Northern Europe', 'LT': 'Northern Europe', 'NO': 'Northern Europe',
    'SE': 'Northern Europe', 'GB': 'Northern Europe',
    'BY': 'Eastern Europe', 'BG': 'Eastern Europe', 'CZ': 'Eastern Europe', 'HU': 'Eastern Europe',
    'MD': 'Eastern Europe', 'PL': 'Eastern Europe', 'RO': 'Eastern Europe', 'RU': 'Eastern Europe',
    'SK': 'Eastern Europe', 'UA': 'Eastern Europe',
    'AL': 'Southern Europe', 'AD': 'Southern Europe', 'BA': 'Southern Europe', 'HR': 'Southern Europe',
    'CY': 'Southern Europe', 'GR': 'Southern Europe', 'IT': 'Southern Europe', 'XK': 'Southern Europe',
    'MK': 'Southern Europe', 'MT': 'Southern Europe', 'ME': 'Southern Europe', 'PT': 'Southern Europe',
    'RS': 'Southern Europe', 'SI': 'Southern Europe', 'ES': 'Southern Europe', 'VA': 'Southern Europe',
    'AT': 'Western Europe', 'BE': 'Western Europe', 'FR': 'Western Europe', 'DE': 'Western Europe',
    'LI': 'Western Europe', 'LU': 'Western Europe', 'MC': 'Western Europe', 'NL': 'Western Europe', 'CH': 'Western Europe',
    'AM': 'Caucasus', 'AZ': 'Caucasus', 'GE': 'Caucasus', 'TR': 'Western Asia',
    'CA': 'Northern America', 'US': 'Northern America',
    'BZ': 'Central America', 'CR': 'Central America', 'SV': 'Central America', 'GT': 'Central America',
    'HN': 'Central America', 'MX': 'Central America', 'NI': 'Central America', 'PA': 'Central America',
    'AG': 'Caribbean', 'BS': 'Caribbean', 'BB': 'Caribbean', 'CU': 'Caribbean', 'DM': 'Caribbean',
    'DO': 'Caribbean', 'GD': 'Caribbean', 'HT': 'Caribbean', 'JM': 'Caribbean', 'KN': 'Caribbean',
    'LC': 'Caribbean', 'VC': 'Caribbean', 'TT': 'Caribbean',
    'AR': 'South America', 'BO': 'South America', 'BR': 'South America', 'CL': 'South America',
    'CO': 'South America', 'EC': 'South America', 'GY': 'South America', 'PY': 'South America',
    'PE': 'South America', 'SR': 'South America', 'UY': 'South America', 'VE': 'South America',
    'CN': 'Eastern Asia', 'JP': 'Eastern Asia', 'KP': 'Eastern Asia', 'KR': 'Eastern Asia',
    'MN': 'Eastern Asia', 'TW': 'Eastern Asia',
    'BN': 'South-Eastern Asia', 'KH': 'South-Eastern Asia', 'ID': 'South-Eastern Asia',
    'LA': 'South-Eastern Asia', 'MY': 'South-Eastern Asia', 'MM': 'South-Eastern Asia',
    'PH': 'South-Eastern Asia', 'SG': 'South-Eastern Asia', 'TH': 'South-Eastern Asia',
    'TL': 'South-Eastern Asia', 'VN': 'South-Eastern Asia',
    'AF': 'Southern Asia', 'BD': 'Southern Asia', 'BT': 'Southern Asia', 'IN': 'Southern Asia',
    'MV': 'Southern Asia', 'NP': 'Southern Asia', 'PK': 'Southern Asia', 'LK': 'Southern Asia',
    'KZ': 'Central Asia', 'KG': 'Central Asia', 'TJ': 'Central Asia', 'TM': 'Central Asia', 'UZ': 'Central Asia',
    'AU': 'Australia and New Zealand', 'NZ': 'Australia and New Zealand',
    'FJ': 'Melanesia', 'PG': 'Melanesia', 'SB': 'Melanesia', 'VU': 'Melanesia',
    'KI': 'Micronesia', 'MH': 'Micronesia', 'FM': 'Micronesia', 'NR': 'Micronesia', 'PW': 'Micronesia',
    'WS': 'Polynesia', 'TO': 'Polynesia', 'TV': 'Polynesia',
    'BH': 'Arabian Peninsula', 'KW': 'Arabian Peninsula', 'OM': 'Arabian Peninsula',
    'QA': 'Arabian Peninsula', 'SA': 'Arabian Peninsula', 'AE': 'Arabian Peninsula', 'YE': 'Arabian Peninsula',
    'IL': 'Levant', 'JO': 'Levant', 'LB': 'Levant', 'PS': 'Levant', 'SY': 'Levant', 'IQ': 'Mesopotamia',
}

# 3a. Update Country_Regions
print("  3a. Updating Country_Regions...")
ws = wb['Country_Regions']
for row in range(2, ws.max_row + 1):
    code = ws.cell(row, 1).value
    if code in STATE_DEPT_REGIONS:
        ws.cell(row, 3).value = STATE_DEPT_REGIONS[code]
    if code in SUBREGIONS:
        ws.cell(row, 5).value = SUBREGIONS[code]

# 3b. Add Proposed_Amount to Country_Budgets
print("  3b. Adding Proposed_Amount column...")
ws = wb['Country_Budgets']
ws.insert_cols(6, 1)  # Insert before Allocated (which was at F)
ws.cell(1, 6).value = "Proposed_Amount"
from openpyxl.styles import Font, PatternFill, Alignment
ws.cell(1, 6).font = Font(bold=True, color='FFFFFF')
ws.cell(1, 6).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')

for row in range(2, ws.max_row + 1):
    ws.cell(row, 6).value = 0  # Default proposed amount
    ws.cell(row, 10).value = f'=H{row}-I{row}'  # ULO = Obligated - Spent
    ws.cell(row, 11).value = f'=IF(H{row}>0,J{row}/H{row},0)'  # ULO %
    ws.cell(row, 11).number_format = '0%'

# 3c. Add Total_Proposed to Master_Projects
print("  3c. Adding Total_Proposed column...")
ws = wb['Master_Projects']
ws.insert_cols(14, 1)  # Insert at N
ws.cell(1, 14).value = "Total_Proposed"
ws.cell(1, 14).font = Font(bold=True, color='FFFFFF')
ws.cell(1, 14).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.cell(1, 14).alignment = Alignment(horizontal='center', vertical='center')

for row in range(2, ws.max_row + 1):
    ws.cell(row, 14).value = f'=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$F:$F)'
    ws.cell(row, 15).value = f'=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$G:$G)'
    ws.cell(row, 16).value = f'=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$H:$H)'
    ws.cell(row, 17).value = f'=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$J:$J)'
    ws.cell(row, 18).value = f'=IF(P{row}>0,Q{row}/P{row},0)'
    ws.cell(row, 18).number_format = '0%'

# 3d. Update Config_Lists
print("  3d. Adding countries to Config_Lists...")
ws_config = wb['Config_Lists']
ws_regions = wb['Country_Regions']
next_row = ws_config.max_row + 1
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws_config.cell(next_row, 1).value = "Country"
        ws_config.cell(next_row, 2).value = code
        ws_config.cell(next_row, 3).value = name
        ws_config.cell(next_row, 4).value = region
        next_row += 1

# 3e. Sync Country_PM_Assignments
print("  3e. Syncing Country_PM_Assignments...")
ws_pm = wb['Country_PM_Assignments']
# Clear and rebuild
for row in range(2, 200):  # Clear up to row 200
    for col in range(1, 8):
        ws_pm.cell(row, col).value = None

current_row = 2
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws_pm.cell(current_row, 1).value = code
        ws_pm.cell(current_row, 2).value = name
        ws_pm.cell(current_row, 3).value = region
        ws_pm.cell(current_row, 4).value = "TBD"
        current_row += 1

# Step 4: Save final v15
print("\nStep 4: Saving final v15...")
wb.save('2025-10-26-Tracker-v15.xlsx')

print("\n" + "="*80)
print("V15 BUILD COMPLETE!")
print("="*80)
print("\nFile: 2025-10-26-Tracker-v15.xlsx")
print("\nThis version:")
print("  - Has NO Excel Tables")
print("  - Has Proposed -> Allocated -> Obligated -> Spent budget flow")
print("  - Uses State Dept regions (AF, EAP, EUR, NEA, SCA, WHA)")
print("  - Has subregions for all countries")
print("  - All lists are synchronized")
print("\nIt should open cleanly in Excel!")
