"""
Check Calendar_Todo sheet to understand its purpose and what formulas it needs
"""
import openpyxl

print("="*80)
print("CHECKING CALENDAR_TODO SHEET")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx', data_only=False)

ws_calendar = wb['Calendar_Todo']

print("\n" + "="*80)
print("CALENDAR_TODO STRUCTURE")
print("="*80)

# Check headers
print("\nHeaders (Row 1):")
headers = []
for col in range(1, 15):
    header = ws_calendar.cell(1, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        headers.append((col_letter, header))
        print(f"  {col_letter}: {header}")

# Check if there's any data
print("\nChecking for existing data (rows 2-10)...")
has_data = False
for row in range(2, 11):
    row_data = []
    for col in range(1, len(headers) + 1):
        value = ws_calendar.cell(row, col).value
        if value:
            row_data.append(value)
    if row_data:
        has_data = True
        print(f"  Row {row}: {', '.join([str(v)[:30] for v in row_data[:3]])}...")

if not has_data:
    print("  No data found in rows 2-10")

# Check for formulas
print("\nChecking for formulas (row 2)...")
has_formulas = False
for col in range(1, len(headers) + 1):
    cell = ws_calendar.cell(2, col)
    if cell.data_type == 'f':
        has_formulas = True
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = str(cell.value)[:60]
        print(f"  {col_letter}2: {formula}...")

if not has_formulas:
    print("  No formulas found in row 2")

print("\n" + "="*80)
print("ANALYSIS")
print("="*80)

print("\nCalendar_Todo sheet purpose:")
if not headers:
    print("  EMPTY - No headers found")
    print("  Recommendation: Delete sheet or define its purpose")
elif has_formulas:
    print("  Has formulas - likely pulls data from other sheets")
    print("  Recommendation: Review formulas for optimization")
elif has_data:
    print("  Has data but no formulas - manual entry sheet")
    print("  Recommendation: Add formulas if it should auto-populate")
else:
    print("  Headers exist but no data/formulas")
    print("  Recommendation: Determine if this should:")
    print("    A) Pull data from other sheets (add formulas)")
    print("    B) Be manually populated (leave as is)")
    print("    C) Be removed (if not needed)")

# Typical uses for Calendar/Todo sheets:
print("\n" + "-"*80)
print("Common Calendar/Todo sheet patterns:")
print("-"*80)
print("\n1. Task Management:")
print("   - Task ID, Task Name, Due Date, Assigned To, Status")
print("   - Usually manual entry")
print("\n2. Milestone Tracker:")
print("   - Pulls from Projects sheet (Project Name, Milestone, Date)")
print("   - Auto-populates with formulas")
print("\n3. Deadline Calendar:")
print("   - Pulls upcoming deadlines from multiple sources")
print("   - Formulas pull from Master_Projects, Deliverables, etc.")

if headers:
    print("\n" + "-"*80)
    print("Based on your headers:")
    print("-"*80)
    for col_letter, header in headers:
        if 'date' in header.lower():
            print(f"  {header}: Could pull from Master_Projects or Milestones")
        elif 'project' in header.lower():
            print(f"  {header}: Could pull from Master_Projects[Project_Name]")
        elif 'status' in header.lower():
            print(f"  {header}: Could pull from Master_Projects[Project_Status]")
        elif 'task' in header.lower():
            print(f"  {header}: Likely manual entry")

print("\n" + "="*80)
print("RECOMMENDATION")
print("="*80)

print("\nWithout knowing the intended use, here are options:")
print("\nOption A - Manual Task Tracker:")
print("  Keep as-is for manual entry")
print("  No formulas needed")
print("\nOption B - Automated Milestone Calendar:")
print("  Add formulas to pull upcoming dates from:")
print("    - Master_Projects (POP dates, Project End Date)")
print("    - Milestones sheet")
print("    - Project_Deliverables")
print("\nOption C - Remove Sheet:")
print("  If not actively used, consider removing")

print("\nPlease let me know which approach you prefer!")
