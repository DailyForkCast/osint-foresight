"""
Fix T_Stakeholders table (table4) - v40
The issue started when we deleted Stakeholder_ID column
"""
import openpyxl

print("="*80)
print("FIXING T_STAKEHOLDERS TABLE - V40")
print("="*80)

# Load v39
print("\nLoading v39...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v39.xlsx')

ws_stakeholders = wb['Stakeholders']

print("\n" + "="*80)
print("CHECKING CURRENT TABLE STRUCTURE")
print("="*80)

print("\nHeaders (row 1):")
for col in range(1, 10):
    header = ws_stakeholders.cell(1, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}1: {clean_header}")

print("\nChecking if T_Stakeholders table exists...")
if 'T_Stakeholders' in ws_stakeholders.tables:
    table = ws_stakeholders.tables['T_Stakeholders']
    print(f"  Table exists")
    print(f"  Table name: {table.name}")
    print(f"  Display name: {table.displayName}")

    # Get table reference
    table_ref = table.ref
    print(f"  Current reference: {table_ref}")
else:
    print("  Table NOT found!")

print("\n" + "="*80)
print("FIXING TABLE")
print("="*80)

print("\nApproach: Delete and recreate the table with correct structure...")

# Delete the existing table
if 'T_Stakeholders' in ws_stakeholders.tables:
    print("  Deleting old T_Stakeholders table...")
    del ws_stakeholders.tables['T_Stakeholders']
    print("  Deleted")

# Count how many rows have data
data_rows = 1  # Start at 1 for header
row = 2
while ws_stakeholders.cell(row, 1).value:
    data_rows += 1
    row += 1
    if row > 200:  # Safety limit
        break

print(f"  Found {data_rows} rows (including header)")

# Count columns with headers
data_cols = 0
col = 1
while ws_stakeholders.cell(1, col).value:
    data_cols += 1
    col += 1
    if col > 20:  # Safety limit
        break

print(f"  Found {data_cols} columns")

# Create new table
from openpyxl.worksheet.table import Table, TableStyleInfo

# Determine table range
if data_rows > 1:
    end_col_letter = openpyxl.utils.get_column_letter(data_cols)
    table_ref = f"A1:{end_col_letter}{data_rows}"
else:
    # If no data, just include header row
    end_col_letter = openpyxl.utils.get_column_letter(data_cols)
    table_ref = f"A1:{end_col_letter}1"

print(f"\n  Creating new table with range: {table_ref}")

tab = Table(displayName="T_Stakeholders", ref=table_ref)

# Add a default style
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

# Add the table
ws_stakeholders.add_table(tab)

print("  New T_Stakeholders table created")

print("\n" + "="*80)
print("SAVING V40")
print("="*80)

wb.save('2025-10-26-Tracker-v40.xlsx')

print("\nOK - v40 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v40:")
print("  - T_Stakeholders table (table4) rebuilt")
print("  - Table now correctly references current column structure")
print("  - No more Stakeholder_ID column issues")

print(f"\nTable structure:")
print(f"  Range: {table_ref}")
print(f"  Rows: {data_rows} (including header)")
print(f"  Columns: {data_cols}")

print("\nv40 should open without ANY Excel errors!")
