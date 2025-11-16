"""
Fix Portfolio_Dashboard text columns (ID, Name, Status, Priority, Progress)
to show blank instead of 0 when no data
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO_DASHBOARD TEXT COLUMNS")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("FIXING TEXT COLUMNS (A-E)")
print("="*80)

print("\nColumns to fix:")
print("  A: ID (Project_Unique_ID)")
print("  B: Project Name")
print("  C: Status")
print("  D: Priority")
print("  E: Progress")

print("\nUpdating formulas for rows 11-20...")

for row in range(11, 21):
    # Column A: Project ID
    # If INDEX returns empty/0, show blank
    ws_port.cell(row, 1).value = f'=IFERROR(IF(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)="","",INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)),"")'

    # Column B: Project Name
    ws_port.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Name],ROW()-10)="","",INDEX(T_Master_Projects[Project_Name],ROW()-10)),""))'

    # Column C: Status
    ws_port.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Status],ROW()-10)="","",INDEX(T_Master_Projects[Project_Status],ROW()-10)),""))'

    # Column D: Priority
    ws_port.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Priority],ROW()-10)="","",INDEX(T_Master_Projects[Project_Priority],ROW()-10)),""))'

    # Column E: Progress
    ws_port.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Progress],ROW()-10)="","",INDEX(T_Master_Projects[Project_Progress],ROW()-10)),""))'

print("  Updated formulas for columns A-E")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Portfolio_Dashboard text columns fixed!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed columns A-E (ID, Name, Status, Priority, Progress):")
print("  - Now show blank instead of 0 when no data")
print("  - Uses double check: IFERROR + IF(value=\"\",\"\",value)")
print("  - Prevents zeros from displaying in text columns")

print("\nColumns F-M (Countries, budgets) already handled separately")
print("Columns N-Q: Ready for your manual restructuring")

print("\nv28 updated!")
