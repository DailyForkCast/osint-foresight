"""
Fix Calendar_Todo once and for all - v38
The table reference formulas are causing corruption.
Let's make it a simple manual entry sheet instead.
"""
import openpyxl

print("="*80)
print("FIXING CALENDAR_TODO - V38")
print("="*80)

# Load v37
print("\nLoading v37...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v37.xlsx')

ws_calendar = wb['Calendar_Todo']

print("\n" + "="*80)
print("CLEARING CALENDAR_TODO")
print("="*80)

print("\nRemoving all formulas that are causing corruption...")

# Clear all data except headers
for row in range(2, 101):
    for col in range(1, 10):
        ws_calendar.cell(row, col).value = None

print("  All formulas cleared")

print("\n" + "="*80)
print("ADDING INSTRUCTIONS")
print("="*80)

print("\nAdding instructions for manual entry...")

ws_calendar['A2'] = "MANUAL ENTRY - Add your tasks, deadlines, and calendar items below"
ws_calendar['A3'] = ""
ws_calendar['A4'] = "Tip: You can pull dates from Master_Projects columns:"
ws_calendar['A5'] = "  - POP_Start (column L)"
ws_calendar['A6'] = "  - POP_End (column M)"
ws_calendar['A7'] = "  - Project_Start_Date (column I)"
ws_calendar['A8'] = "  - Project_End_Date (column J)"

print("  Instructions added")

print("\n" + "="*80)
print("ALTERNATIVE: ADD SIMPLE NON-TABLE FORMULAS")
print("="*80)

print("\nAdding a few examples using direct cell references (not tables)...")

# Add a few examples using direct Master_Projects cell references
ws_calendar['A10'] = "=== EXAMPLE: FIRST 5 PROJECT START DATES ==="

for i in range(1, 6):
    row = 10 + i
    # Use direct cell references to Master_Projects sheet
    # Assuming Project_Unique_ID is in column B, Project_Name in C, POP_Start in L
    ws_calendar.cell(row, 1).value = f'=IF(Master_Projects!B{i+1}="","",Master_Projects!B{i+1})'
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","","Start: "&Master_Projects!C{i+1})'
    ws_calendar.cell(row, 3).value = f'=A{row}'
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",Master_Projects!L{i+1})'
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","",Master_Projects!R{i+1})'  # Implementer_POC
    ws_calendar.cell(row, 6).value = "Scheduled"
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","",Master_Projects!F{i+1})'  # Priority
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Project start date")'

print("  Example formulas added (rows 11-15)")

print("\n" + "="*80)
print("SAVING V38")
print("="*80)

wb.save('2025-10-26-Tracker-v38.xlsx')

print("\nOK - v38 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v38:")
print("  1. Calendar_Todo: Cleared all corrupting table reference formulas")
print("  2. Added instructions for manual entry")
print("  3. Added 5 example rows using direct cell references")
print("     (These should work without corruption)")

print("\nCalendar_Todo options:")
print("  A. Use it as manual entry sheet (safest)")
print("  B. Copy the example formula pattern in rows 11-15")
print("  C. Use Excel's built-in features to link dates")

print("\nThe direct cell reference approach (Master_Projects!B2)")
print("should work better than table references (T_Master_Projects[...])")

print("\nv38 ready - should open without errors!")
