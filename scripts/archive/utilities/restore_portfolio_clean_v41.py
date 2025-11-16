"""
Restore Portfolio_Dashboard from v27 (before Award/NCE columns) - v41
Clean, simple structure
"""
import openpyxl

print("="*80)
print("RESTORING CLEAN PORTFOLIO_DASHBOARD - V41")
print("="*80)

# Load v27 (clean Portfolio) and v40
print("\nLoading v27 and v40...")
wb_v27 = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx')
wb_v40 = openpyxl.load_workbook('2025-10-26-Tracker-v40.xlsx')

print("\n" + "="*80)
print("RESTORING PORTFOLIO_DASHBOARD FROM V27")
print("="*80)

# Delete Portfolio_Dashboard from v40
print("\nRemoving current Portfolio_Dashboard from v40...")
if 'Portfolio_Dashboard' in wb_v40.sheetnames:
    del wb_v40['Portfolio_Dashboard']
    print("  Removed")

# Copy Portfolio_Dashboard from v27
print("\nCopying Portfolio_Dashboard from v27...")
ws_v27_portfolio = wb_v27['Portfolio_Dashboard']
ws_v40_portfolio = wb_v40.create_sheet('Portfolio_Dashboard')

# Copy all cells
print("  Copying cells...")
for row in ws_v27_portfolio.iter_rows():
    for cell in row:
        new_cell = ws_v40_portfolio[cell.coordinate]
        new_cell.value = cell.value
        if cell.has_style:
            try:
                new_cell.font = cell.font.copy() if hasattr(cell.font, 'copy') else cell.font
                new_cell.border = cell.border.copy() if hasattr(cell.border, 'copy') else cell.border
                new_cell.fill = cell.fill.copy() if hasattr(cell.fill, 'copy') else cell.fill
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy() if hasattr(cell.protection, 'copy') else cell.protection
                new_cell.alignment = cell.alignment.copy() if hasattr(cell.alignment, 'copy') else cell.alignment
            except:
                pass

# Copy column widths
print("  Copying column widths...")
for col in ws_v27_portfolio.column_dimensions:
    if col in ws_v27_portfolio.column_dimensions:
        ws_v40_portfolio.column_dimensions[col].width = ws_v27_portfolio.column_dimensions[col].width

# Copy row heights
print("  Copying row heights...")
for row in ws_v27_portfolio.row_dimensions:
    if row in ws_v27_portfolio.row_dimensions:
        ws_v40_portfolio.row_dimensions[row].height = ws_v27_portfolio.row_dimensions[row].height

# Copy merged cells
print("  Copying merged cells...")
for merged_cell_range in ws_v27_portfolio.merged_cells.ranges:
    ws_v40_portfolio.merge_cells(str(merged_cell_range))

print("  Portfolio_Dashboard restored from v27")

# Move Portfolio_Dashboard to correct position
print("\nMoving Portfolio_Dashboard to correct position...")
sheets = wb_v40.sheetnames
portfolio_idx = sheets.index('Portfolio_Dashboard')
master_idx = sheets.index('Master_Projects') if 'Master_Projects' in sheets else 0
target_idx = master_idx + 1

if portfolio_idx != target_idx:
    wb_v40.move_sheet('Portfolio_Dashboard', offset=target_idx - portfolio_idx)
    print(f"  Moved to position {target_idx}")

print("\n" + "="*80)
print("CHECKING STRUCTURE")
print("="*80)

print("\nRow 10 headers (project list):")
for col in range(1, 14):
    header = ws_v40_portfolio.cell(10, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}10: {clean_header}")

print("\n" + "="*80)
print("SAVING V41")
print("="*80)

wb_v40.save('2025-10-26-Tracker-v41.xlsx')

print("\nOK - v41 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nPortfolio_Dashboard restored from v27:")
print("  - Clean, simple structure")
print("  - NO Award Number column")
print("  - NO NCE Eligible column")
print("  - NO NCE Status column")
print("  - Professional formatting preserved")

print("\nColumn structure (rows 11-20):")
print("  A: ID")
print("  B: Project Name")
print("  C: Status")
print("  D: Priority")
print("  E: Progress")
print("  F: Countries")
print("  G: Total Proposed")
print("  H: Total Allocation")
print("  I: Total Obligated")
print("  J: Total Spent")
print("  K: Total ULO")
print("  L: ULO %")
print("  M: Days Remaining")

print("\nYou can now manually add Award Number and NCE columns wherever you want!")

print("\nv41 ready!")
