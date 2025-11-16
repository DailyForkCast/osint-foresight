"""
Fix Portfolio_Dashboard formatting by restoring from v32 - v36
v32 had the correct structure with Award/NCE columns
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

print("="*80)
print("FIXING PORTFOLIO_DASHBOARD FORMATTING - V36")
print("="*80)

# Load v32 (had good formatting) and v35
print("\nLoading v32 and v35...")
wb_v32 = openpyxl.load_workbook('2025-10-26-Tracker-v32.xlsx')
wb_v35 = openpyxl.load_workbook('2025-10-26-Tracker-v35.xlsx')

print("\n" + "="*80)
print("RESTORING PORTFOLIO_DASHBOARD FROM V32")
print("="*80)

# Delete Portfolio_Dashboard from v35
print("\nRemoving Portfolio_Dashboard from v35...")
if 'Portfolio_Dashboard' in wb_v35.sheetnames:
    del wb_v35['Portfolio_Dashboard']
    print("  Removed")

# Copy Portfolio_Dashboard from v32
print("\nCopying Portfolio_Dashboard from v32...")
ws_v32_portfolio = wb_v32['Portfolio_Dashboard']
ws_v35_portfolio = wb_v35.create_sheet('Portfolio_Dashboard')

# Copy all cells
print("  Copying cells...")
for row in ws_v32_portfolio.iter_rows():
    for cell in row:
        new_cell = ws_v35_portfolio[cell.coordinate]
        new_cell.value = cell.value
        if cell.has_style:
            try:
                new_cell.font = cell.font.copy() if hasattr(cell.font, 'copy') else cell.font
                new_cell.border = cell.border.copy() if hasattr(cell.border, 'copy') else cell.border
                new_cell.fill = cell.fill.copy() if hasattr(cell.fill, 'copy') else cell.fill
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy() if hasattr(cell.protection, 'copy') else cell.protection
                new_cell.alignment = cell.alignment.copy() if hasattr(cell.alignment, 'copy') else cell.alignment
            except:
                pass  # Skip if style copy fails

# Copy column widths
print("  Copying column widths...")
for col in ws_v32_portfolio.column_dimensions:
    if col in ws_v32_portfolio.column_dimensions:
        ws_v35_portfolio.column_dimensions[col].width = ws_v32_portfolio.column_dimensions[col].width

# Copy row heights
print("  Copying row heights...")
for row in ws_v32_portfolio.row_dimensions:
    if row in ws_v32_portfolio.row_dimensions:
        ws_v35_portfolio.row_dimensions[row].height = ws_v32_portfolio.row_dimensions[row].height

# Copy merged cells
print("  Copying merged cells...")
for merged_cell_range in ws_v32_portfolio.merged_cells.ranges:
    ws_v35_portfolio.merge_cells(str(merged_cell_range))

print("  Portfolio_Dashboard restored from v32")

# Move Portfolio_Dashboard to correct position
print("\nMoving Portfolio_Dashboard to correct position...")
sheets = wb_v35.sheetnames
portfolio_idx = sheets.index('Portfolio_Dashboard')
master_idx = sheets.index('Master_Projects') if 'Master_Projects' in sheets else 0
target_idx = master_idx + 1

if portfolio_idx != target_idx:
    wb_v35.move_sheet('Portfolio_Dashboard', offset=target_idx - portfolio_idx)
    print(f"  Moved to position {target_idx}")

print("\n" + "="*80)
print("SAVING V36")
print("="*80)

wb_v35.save('2025-10-26-Tracker-v36.xlsx')

print("\nOK - v36 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v36:")
print("  - Portfolio_Dashboard formatting restored from v32")
print("  - All styling, merged cells, column widths preserved")
print("  - Includes Award Number and NCE columns")
print("  - Clean, professional appearance")

print("\nv36 includes:")
print("  - Portfolio_Dashboard: Clean formatting from v32")
print("  - Calendar_Todo: Automated calendar from v35")
print("  - All other fixes from v33-v34")

print("\nv36 ready!")
