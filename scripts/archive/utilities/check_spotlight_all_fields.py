"""
Check ALL fields in Spotlight including Award, NCE, Progress, Implementer
"""
import openpyxl

print("="*80)
print("COMPREHENSIVE SPOTLIGHT CHECK")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("ROWS 1-20 DETAILED")
print("="*80)

for row in range(1, 21):
    print(f"\nRow {row}:")
    for col in range(1, 7):  # A-F
        cell = ws_spotlight.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)[:60]
                print(f"  {col_letter}: FORMULA - {formula}...")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')[:50]
                print(f"  {col_letter}: {clean_value}")

print("\n" + "="*80)
print("SEARCH FOR SPECIFIC FIELDS")
print("="*80)

# Look for Award Number, NCE fields specifically in rows 1-10
print("\nRows 1-10, Columns E-F (Award/NCE area):")
for row in range(1, 11):
    for col in [5, 6]:  # E, F
        cell = ws_spotlight.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)[:80]
                print(f"  {col_letter}{row}: FORMULA - {formula}...")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
                print(f"  {col_letter}{row}: {clean_value}")

print("\n" + "="*80)
print("FIELDS THAT NEED FIXING")
print("="*80)

print("\nBased on formulas found, these fields might show 0 instead of blank:")
print("  - B6: Priority (text field)")
print("  - B7: Start Date (date field)")
print("  - B8: Days Remaining (numeric)")
print("  - B9: POC/Implementer (text field)")
print("  - B14: Total Proposed (numeric)")
print("  - E14: Total Obligated (numeric)")
print("  - B15: Total Allocated (numeric)")
print("  - E15: Total Spent (numeric)")
print("  - F2: Award Number (text)")
print("  - F3: POP Start Date (date)")
print("  - Any NCE fields in E/F columns")

print("\nAll these should show BLANK when:")
print("  1. No project selected ($B$2 is empty)")
print("  2. The field value is 0 or empty")

print("\nReady to fix!")
