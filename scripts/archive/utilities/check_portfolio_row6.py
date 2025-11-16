"""
Check Portfolio_Dashboard row 6, columns D-F to see what's wrong with the percentage
"""
import openpyxl

print("="*80)
print("CHECKING PORTFOLIO_DASHBOARD ROW 6")
print("="*80)

# Load v30
print("\nLoading v30...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v30.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("ROWS 1-10, COLUMNS A-F")
print("="*80)

for row in range(1, 11):
    print(f"\nRow {row}:")
    for col in range(1, 7):  # A-F
        cell = ws_portfolio.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)[:80]
                print(f"  {col_letter}{row}: FORMULA - {formula}...")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')[:60]
                print(f"  {col_letter}{row}: {clean_value}")

                # Also check number format
                if cell.number_format and cell.number_format != 'General':
                    print(f"        (Format: {cell.number_format})")

print("\n" + "="*80)
print("SPECIFIC CHECK: D6, E6, F6")
print("="*80)

for col_letter in ['D', 'E', 'F']:
    cell = ws_portfolio[f'{col_letter}6']
    print(f"\n{col_letter}6:")
    print(f"  Value: {cell.value}")
    print(f"  Data type: {cell.data_type}")
    print(f"  Number format: {cell.number_format}")
    if cell.data_type == 'f':
        print(f"  Formula: {cell.value}")

print("\nReady to fix!")
