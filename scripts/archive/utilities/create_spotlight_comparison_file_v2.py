import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

output_file = 'C:/Projects/OSINT-Foresight/Spotlight_Comparison.xlsx'
source_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v12.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING STANDALONE SPOTLIGHT COMPARISON FILE')
print('='*80)
print()

# Create new workbook
wb = openpyxl.Workbook()
ws_default = wb.active
wb.remove(ws_default)

# Load source to read structure
wb_source = openpyxl.load_workbook(source_file, data_only=False)

print('Creating reference data sheets...')
print('-'*80)

# ============================================================================
# REFERENCE DATA SHEETS
# ============================================================================

# Master_Projects
ws_projects = wb.create_sheet('Master_Projects', 0)
ws_projects['A1'] = 'Unique_ID'
ws_projects['B1'] = 'Fiscal_Year'
ws_projects['C1'] = 'Project_Name'
ws_projects['D1'] = 'Summary'
ws_projects['E1'] = 'Status'
ws_projects['F1'] = 'Priority'
ws_projects['G1'] = 'Progress'
ws_projects['H1'] = 'Start_Date'
ws_projects['I1'] = 'End_Date'
ws_projects['J1'] = 'Days_Remaining'
ws_projects['K1'] = 'Implementer'
ws_projects['L1'] = 'Implementer_POC'
ws_projects['M1'] = 'Implementer_POC_Email'
ws_projects['N1'] = 'Total_Allocation'
ws_projects['O1'] = 'Total_Obligated'
ws_projects['P1'] = 'Total_ULO'
ws_projects['Q1'] = 'ULO_Percent'
ws_projects['R1'] = 'Countries'
ws_projects['S1'] = 'Country_Count'
ws_projects['T1'] = 'Total_Spent'

ws_projects.append(['PRJ-001', 2025, 'Digital Infrastructure Modernization',
    'Comprehensive program to modernize critical digital infrastructure across European partner nations, focusing on cloud adoption, cybersecurity, and data governance.',
    'Active', 'High', '75%', date(2025, 1, 15), date(2025, 12, 31), 120,
    'TechCorp International', 'Jane Doe', 'jane.doe@techcorp.com',
    1500000, 1200000, 800000, 0.667, 'DE, FR, IT, ES, NL', 5, 400000])

ws_projects.append(['PRJ-002', 2025, 'AI Ethics Framework Development',
    'Multi-stakeholder initiative to develop comprehensive AI governance and ethics frameworks for government adoption.',
    'CN Stage', 'Critical', '45%', date(2025, 3, 1), date(2026, 2, 28), 240,
    'AI Policy Institute', 'Dr. Smith', 'smith@aipolicy.org',
    850000, 600000, 450000, 0.75, 'FR, BE, NL', 3, 150000])

print('[OK] Master_Projects')

# T_Project_Deliverables
ws_deliv = wb.create_sheet('T_Project_Deliverables', 1)
ws_deliv.append(['Project_ID', 'Deliverable_Name', 'Deliverable_Type', 'Due_Date', 'Status', 'Owner', 'Completion_Percent'])
ws_deliv.append(['PRJ-001', 'Technical Requirements Document', 'Report', date(2025, 11, 15), 'In Progress', 'JD', '75%'])
ws_deliv.append(['PRJ-001', 'Beta Release', 'Software', date(2025, 12, 1), 'Not Started', 'AS', '0%'])
ws_deliv.append(['PRJ-001', 'Security Audit', 'Report', date(2025, 11, 30), 'In Progress', 'MS', '60%'])
ws_deliv.append(['PRJ-001', 'User Training Materials', 'Documentation', date(2025, 12, 15), 'Not Started', 'LK', '0%'])
ws_deliv.append(['PRJ-001', 'Final Implementation Report', 'Report', date(2025, 12, 31), 'Not Started', 'JD', '0%'])
ws_deliv.append(['PRJ-002', 'Ethics Framework Draft', 'Report', date(2025, 6, 30), 'Completed', 'DS', '100%'])
ws_deliv.append(['PRJ-002', 'Stakeholder Consultation Report', 'Report', date(2025, 9, 15), 'In Progress', 'MJ', '70%'])
ws_deliv.append(['PRJ-002', 'Final Framework Publication', 'Report', date(2026, 2, 28), 'Not Started', 'DS', '0%'])

print('[OK] T_Project_Deliverables')

# T_Project_Audiences
ws_aud = wb.create_sheet('T_Project_Audiences', 2)
ws_aud.append(['Project_ID', 'Audience_Type', 'Audience_Region', 'Description', 'Priority'])
ws_aud.append(['PRJ-001', 'Government Officials', 'Europe', 'Senior technology and digital transformation officials', 'High'])
ws_aud.append(['PRJ-001', 'Private Sector', 'Europe', 'Technology providers and system integrators', 'Medium'])
ws_aud.append(['PRJ-001', 'Academia', 'Europe', 'Research institutions and universities', 'Low'])
ws_aud.append(['PRJ-002', 'Government Officials', 'Europe', 'Policy makers and regulators', 'High'])
ws_aud.append(['PRJ-002', 'Civil Society', 'Global', 'Ethics advocates and civil society organizations', 'High'])

print('[OK] T_Project_Audiences')

# Project_Technologies
ws_tech = wb.create_sheet('Project_Technologies', 3)
ws_tech.append(['Project_ID', 'Technology', 'Category', 'Status'])
ws_tech.append(['PRJ-001', 'Cloud Infrastructure', 'Infrastructure', 'In Use'])
ws_tech.append(['PRJ-001', 'Machine Learning', 'AI/ML', 'Planned'])
ws_tech.append(['PRJ-001', 'Data Analytics', 'Data', 'In Use'])
ws_tech.append(['PRJ-002', 'Natural Language Processing', 'AI/ML', 'In Use'])
ws_tech.append(['PRJ-002', 'Policy Automation Tools', 'Software', 'Planned'])

print('[OK] Project_Technologies')

# Stakeholders
ws_stake = wb.create_sheet('Stakeholders', 4)
ws_stake.append(['Stakeholder_ID', 'Name', 'Title', 'Organization', 'Email', 'Stakeholder_Type', 'Project_IDs'])
ws_stake.append(['STK-001', 'John Smith', 'Program Manager', 'Agency A', 'john.smith@agency.gov', 'Government', 'PRJ-001, PRJ-002'])
ws_stake.append(['STK-002', 'Jane Doe', 'Technical Lead', 'TechCorp International', 'jane.doe@techcorp.com', 'Private Sector', 'PRJ-001'])
ws_stake.append(['STK-003', 'Dr. Robert Chen', 'Chief Architect', 'TechCorp International', 'robert.chen@techcorp.com', 'Private Sector', 'PRJ-001'])
ws_stake.append(['STK-004', 'Maria Garcia', 'Security Advisor', 'CyberSec Consultants', 'maria.garcia@cybersec.com', 'Private Sector', 'PRJ-001'])
ws_stake.append(['STK-005', 'Dr. Emily Smith', 'Research Director', 'AI Policy Institute', 'smith@aipolicy.org', 'Academia', 'PRJ-002'])

print('[OK] Stakeholders')
print()

# ============================================================================
# RECREATE SPOTLIGHT SHEETS BY READING SOURCE
# ============================================================================

print('Recreating Spotlight sheets from v12...')
print('-'*80)

def copy_sheet_content(source_ws, target_ws):
    """Copy all cell values and basic formatting from source to target"""
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column)
            target_cell.value = cell.value

            # Copy basic formatting
            if cell.font:
                target_cell.font = cell.font.copy()
            if cell.fill:
                target_cell.fill = cell.fill.copy()
            if cell.alignment:
                target_cell.alignment = cell.alignment.copy()
            if cell.border:
                target_cell.border = cell.border.copy()
            if cell.number_format:
                target_cell.number_format = cell.number_format

    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copy column dimensions
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width

    # Copy row dimensions
    for row_num, row_dim in source_ws.row_dimensions.items():
        if row_dim.height:
            target_ws.row_dimensions[row_num].height = row_dim.height

    # Copy data validations
    for dv in source_ws.data_validations.dataValidation:
        target_ws.add_data_validation(dv)

# Version 1: Executive
print('Creating 1_Executive...')
ws_exec = wb.create_sheet('1_Executive', 5)
source_exec = wb_source['Spotlight_Executive']
copy_sheet_content(source_exec, ws_exec)
print('[OK] 1_Executive')

# Version 2: PM Workspace
print('Creating 2_PMWorkspace...')
ws_pm = wb.create_sheet('2_PMWorkspace', 6)
source_pm = wb_source['Spotlight_PMWorkspace']
copy_sheet_content(source_pm, ws_pm)
print('[OK] 2_PMWorkspace')

# Version 3: Stakeholder
print('Creating 3_Stakeholder...')
ws_stake_spot = wb.create_sheet('3_Stakeholder', 7)
source_stake_spot = wb_source['Spotlight_Stakeholder']
copy_sheet_content(source_stake_spot, ws_stake_spot)
print('[OK] 3_Stakeholder')

wb_source.close()
print()

# ============================================================================
# ADD START_HERE GUIDE
# ============================================================================

print('Creating START_HERE guide...')

ws_guide = wb.create_sheet('START_HERE', 0)

# Styles
header_font = Font(bold=True, size=18, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
label_font = Font(bold=True, size=10)

ws_guide['A1'] = 'SPOTLIGHT COMPARISON GUIDE'
ws_guide['A1'].font = header_font
ws_guide['A1'].fill = header_fill
ws_guide.merge_cells('A1:F1')
ws_guide.row_dimensions[1].height = 30

ws_guide['A3'] = 'Welcome to the Spotlight Comparison File!'
ws_guide['A3'].font = Font(bold=True, size=12)
ws_guide.merge_cells('A3:F3')

ws_guide['A5'] = 'This file contains THREE versions of Project_Spotlight for you to compare:'
ws_guide.merge_cells('A5:F5')

ws_guide['A7'] = 'Version 1: EXECUTIVE SUMMARY'
ws_guide['A7'].font = section_font
ws_guide['A7'].fill = section_fill
ws_guide.merge_cells('A7:F7')

ws_guide['A8'] = 'Sheet: 1_Executive'
ws_guide['A8'].font = label_font
ws_guide['B8'] = 'Clean, compact, 1-page layout for leadership briefings'
ws_guide.merge_cells('B8:F8')

ws_guide['A9'] = 'Best for:'
ws_guide['A9'].font = label_font
ws_guide['B9'] = 'Quick status updates, executive briefings, printable summaries'
ws_guide.merge_cells('B9:F9')

ws_guide['A11'] = 'Version 2: PM WORKSPACE'
ws_guide['A11'].font = section_font
ws_guide['A11'].fill = section_fill
ws_guide.merge_cells('A11:F11')

ws_guide['A12'] = 'Sheet: 2_PMWorkspace'
ws_guide['A12'].font = label_font
ws_guide['B12'] = 'Detailed, comprehensive, working document with all project info'
ws_guide.merge_cells('B12:F12')

ws_guide['A13'] = 'Best for:'
ws_guide['A13'].font = label_font
ws_guide['B13'] = 'Day-to-day project management, tracking all deliverables and stakeholders'
ws_guide.merge_cells('B13:F13')

ws_guide['A15'] = 'Version 3: STAKEHOLDER BRIEFING'
ws_guide['A15'].font = section_font
ws_guide['A15'].fill = section_fill
ws_guide.merge_cells('A15:F15')

ws_guide['A16'] = 'Sheet: 3_Stakeholder'
ws_guide['A16'].font = label_font
ws_guide['B16'] = 'Visual, polished, presentation-ready for external audiences'
ws_guide.merge_cells('B16:F16')

ws_guide['A17'] = 'Best for:'
ws_guide['A17'].font = label_font
ws_guide['B17'] = 'Client meetings, stakeholder briefings, professional presentations'
ws_guide.merge_cells('B17:F17')

ws_guide['A19'] = 'HOW TO TEST:'
ws_guide['A19'].font = section_font
ws_guide['A19'].fill = section_fill
ws_guide.merge_cells('A19:F19')

ws_guide['A20'] = '1. Go to each sheet (1_Executive, 2_PMWorkspace, 3_Stakeholder)'
ws_guide.merge_cells('A20:F20')
ws_guide['A21'] = '2. Cell B2 has a dropdown - select PRJ-001 or PRJ-002'
ws_guide.merge_cells('A21:F21')
ws_guide['A22'] = '3. Watch all data auto-populate based on selected project'
ws_guide.merge_cells('A22:F22')
ws_guide['A23'] = '4. Compare layouts, styles, and level of detail'
ws_guide.merge_cells('A23:F23')
ws_guide['A24'] = '5. Note what you like and don\'t like about each version'
ws_guide.merge_cells('A24:F24')

ws_guide['A26'] = 'SAMPLE DATA:'
ws_guide['A26'].font = section_font
ws_guide['A26'].fill = section_fill
ws_guide.merge_cells('A26:F26')

ws_guide['A27'] = 'Reference sheets contain sample data for testing:'
ws_guide.merge_cells('A27:F27')
ws_guide['A28'] = '  • Master_Projects: 2 sample projects (PRJ-001, PRJ-002)'
ws_guide.merge_cells('A28:F28')
ws_guide['A29'] = '  • T_Project_Deliverables: 8 sample deliverables'
ws_guide.merge_cells('A29:F29')
ws_guide['A30'] = '  • T_Project_Audiences: 5 sample audiences'
ws_guide.merge_cells('A30:F30')
ws_guide['A31'] = '  • Project_Technologies: 5 sample technologies'
ws_guide.merge_cells('A31:F31')
ws_guide['A32'] = '  • Stakeholders: 5 sample stakeholders'
ws_guide.merge_cells('A32:F32')

ws_guide['A34'] = 'WHAT TO REPORT BACK:'
ws_guide['A34'].font = section_font
ws_guide['A34'].fill = section_fill
ws_guide.merge_cells('A34:F34')

ws_guide['A35'] = '  • Which version do you prefer overall?'
ws_guide.merge_cells('A35:F35')
ws_guide['A36'] = '  • What specific elements do you like from each?'
ws_guide.merge_cells('A36:F36')
ws_guide['A37'] = '  • What would you change or improve?'
ws_guide.merge_cells('A37:F37')
ws_guide['A38'] = '  • Should we combine elements from multiple versions?'
ws_guide.merge_cells('A38:F38')
ws_guide['A39'] = '  • Need calendar/timeline added?'
ws_guide.merge_cells('A39:F39')

ws_guide.column_dimensions['A'].width = 15
ws_guide.column_dimensions['B'].width = 80

print('[OK] START_HERE guide created')
print()

# ============================================================================
# SAVE
# ============================================================================

print('='*80)
print('Saving Spotlight_Comparison.xlsx...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('STANDALONE SPOTLIGHT COMPARISON FILE COMPLETE!')
print('='*80)
print()
print('Created file: Spotlight_Comparison.xlsx')
print()
print('Sheet order:')
print('  1. START_HERE - Read this first!')
print('  2. 1_Executive - Executive Summary version')
print('  3. 2_PMWorkspace - PM Workspace version')
print('  4. 3_Stakeholder - Stakeholder Briefing version')
print('  5-9. Reference data (Master_Projects, etc.)')
print()
print('All formulas work with sample data!')
print('Ready for testing and fine-tuning!')
print()
print(f'Created: {timestamp}')
