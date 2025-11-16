import openpyxl
from openpyxl import load_workbook
import sys

# Set UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = load_workbook('c:/Users/mrear/AppData/Local/Temp/2025-10-05-Tracker (1).xlsx')

print("=" * 70)
print("WORKBOOK INSPECTION: 2025-10-05-Tracker (1).xlsx")
print("=" * 70)
print(f"\nTotal Sheets: {len(wb.sheetnames)}")
print(f"Sheet Names: {', '.join(wb.sheetnames)}\n")

# Inspect each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("\n" + "=" * 70)
    print(f"SHEET: {sheet_name}")
    print("=" * 70)

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")

    # Get headers (first row)
    headers = []
    for col in range(1, min(max_col + 1, 26)):  # Limit to first 25 columns
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            try:
                headers.append(f"{col}:{cell_value}")
            except:
                headers.append(f"{col}:[special chars]")

    if headers:
        print(f"\nHeaders (Row 1):")
        for header in headers:
            print(f"  - {header}")

    # Show first few rows of data
    print(f"\nFirst 5 data rows:")
    for row in range(2, min(7, max_row + 1)):
        row_data = []
        for col in range(1, min(max_col + 1, 6)):  # First 5 columns
            val = ws.cell(row=row, column=col).value
            if val is not None:
                try:
                    row_data.append(str(val)[:30])  # Truncate long values
                except:
                    row_data.append("[special]")
            else:
                row_data.append("")
        if any(row_data):  # Only print if row has data
            print(f"  Row {row}: {' | '.join(row_data)}")

    # Check for merged cells
    if ws.merged_cells:
        print(f"\nMerged Cells: {len(ws.merged_cells.ranges)} ranges")

    # Check for formulas in first 10 rows
    formulas_found = []
    for row in range(1, min(11, max_row + 1)):
        for col in range(1, min(max_col + 1, 26)):
            cell = ws.cell(row=row, column=col)
            if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas_found.append(f"{cell.coordinate}: {cell.value[:50]}")

    if formulas_found:
        print(f"\nFormulas Found (first 10 rows):")
        for formula in formulas_found[:5]:  # Show first 5
            print(f"  - {formula}")
        if len(formulas_found) > 5:
            print(f"  ... and {len(formulas_found) - 5} more")

print("\n" + "=" * 70)
print("END OF INSPECTION")
print("=" * 70)
