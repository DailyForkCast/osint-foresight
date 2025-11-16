"""
Check if Country_Budgets has any data
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("COUNTRY_BUDGETS DATA CHECK")
print("="*80)

ws = wb['Country_Budgets']

print("\nChecking first 20 rows for data:")
print(f"{'Row':<5} {'Project ID (B)':<15} {'Country Code (D)':<15} {'Country Name (E)':<25}")
print("-"*65)

data_count = 0
for row in range(2, min(21, ws.max_row + 1)):
    project_id = ws.cell(row, 2).value or ""
    country_code = ws.cell(row, 4).value or ""
    country_name = ws.cell(row, 5).value or ""

    if project_id or country_code or country_name:
        print(f"{row:<5} {str(project_id):<15} {str(country_code):<15} {str(country_name):<25}")
        data_count += 1

if data_count == 0:
    print("(No data found)")

print("\n" + "="*80)
print("DIAGNOSIS")
print("="*80)

if data_count == 0:
    print("\n❌ Country_Budgets is EMPTY - no countries assigned to projects yet!")
    print("\nTo populate the Countries column in Master_Projects:")
    print("  1. Go to Country_Budgets sheet")
    print("  2. Add rows for each project-country combination")
    print("  3. Column B: Enter Project ID (e.g., FY2025)")
    print("  4. Column E: Select country from dropdown")
    print("  5. Column D will auto-fill with country code")
    print("  6. Column A will auto-create Budget_ID")
    print("\nExample:")
    print("  Row 2: B2=FY2025, E2=Albania → D2 auto-fills 'AL', A2='FY2025-AL'")
    print("  Row 3: B2=FY2025, E2=Austria → D3 auto-fills 'AT', A3='FY2025-AT'")
    print("\nOnce you add countries, Column S will automatically show them!")
else:
    print(f"\n✅ Found {data_count} entries in Country_Budgets")
    print("\nIf Countries column is still blank, the formula may need checking.")
