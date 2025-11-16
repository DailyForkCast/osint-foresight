"""
Fix Portfolio_Dashboard D6 formula - showing wrong percentage
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO_DASHBOARD D6 FORMULA")
print("="*80)

# Load v30
print("\nLoading v30...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v30.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']
ws_control = wb['Control']

print("\n" + "="*80)
print("CHECKING CONTROL!B18")
print("="*80)

print("\nControl sheet, rows 15-20 (Financial Status):")
for row in range(15, 21):
    a_cell = ws_control.cell(row, 1)
    b_cell = ws_control.cell(row, 2)

    label = ""
    if a_cell.value:
        label = str(a_cell.value).encode('ascii', 'ignore').decode('ascii')[:30]

    print(f"  Row {row}: {label}")
    if b_cell.value:
        if b_cell.data_type == 'f':
            formula = str(b_cell.value)[:60]
            print(f"    B{row}: FORMULA - {formula}...")
        else:
            print(f"    B{row}: {b_cell.value}")

        print(f"    Format: {b_cell.number_format}")

print("\n" + "="*80)
print("CURRENT D6 FORMULA")
print("="*80)

d6_formula = ws_portfolio['D6'].value
print(f"\nCurrent D6: {d6_formula}")

print("\n" + "="*80)
print("FIXING D6")
print("="*80)

print("\nAnalysis:")
print("  The formula uses TEXT(Control!B18,\"0%\") which multiplies the value by 100")
print("  If B18 is already formatted as a percentage, this creates a huge number")
print("  Solution: Check if B18 already shows as percentage, then don't multiply")

# Check what B18 actually is
b18_format = ws_control['B18'].number_format
print(f"\n  Control!B18 format: {b18_format}")

# Fix: If B18 is already a percentage format, just concatenate
# Otherwise, B18 should be the ULO amount, not percentage
# Let me check what should be shown

print("\nFix: ULO % should come from Portfolio ULO % (row 19)")
print("  Changing D6 to reference Control!B19 (Portfolio ULO %)")

# New formula - just reference B19 which is already formatted as percentage
new_formula = '=TEXT(Control!B19,"0%")&" Unobligated"'
ws_portfolio['D6'] = new_formula

print(f"\n  New D6: {new_formula}")

print("\n" + "="*80)
print("SAVING V30")
print("="*80)

wb.save('2025-10-26-Tracker-v30.xlsx')

print("\nOK - D6 formula fixed!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed:")
print("  D6: Now references Control!B19 (Portfolio ULO %)")
print("  Old: =TEXT(Control!B18,\"0%\")&\" Unobligated\"")
print("  New: =TEXT(Control!B19,\"0%\")&\" Unobligated\"")

print("\nControl sheet Financial Status:")
print("  B18: Total ULO (dollar amount)")
print("  B19: Portfolio ULO % (percentage)")

print("\nv30 updated!")
