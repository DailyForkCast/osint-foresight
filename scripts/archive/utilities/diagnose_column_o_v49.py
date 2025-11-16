"""
Diagnose Portfolio_Dashboard column O issue in v49
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v49.xlsx')
ws = wb['Portfolio_Dashboard']

print("="*80)
print("DIAGNOSING PORTFOLIO_DASHBOARD COLUMN O")
print("="*80)

# Check what's actually in column O
print("\nColumn O header and values:")
print(f"  O10 (header): {ws['O10'].value}")

for row in range(11, 21):
    cell = ws.cell(row, 15)  # Column O
    value = cell.value
    data_type = cell.data_type

    if value:
        if data_type == 'f':
            print(f"  O{row}: [formula] {value[:80]}")
        else:
            print(f"  O{row}: {value} (type: {data_type})")
    else:
        print(f"  O{row}: (empty)")

# Check conditional formatting on column O
print("\n" + "="*80)
print("CONDITIONAL FORMATTING ON COLUMN O")
print("="*80)

print(f"\nTotal conditional formatting rules: {len(ws.conditional_formatting._cf_rules)}")

for idx, (range_str, rules) in enumerate(ws.conditional_formatting._cf_rules.items()):
    if 'O' in str(range_str):
        print(f"\nRule on range: {range_str}")
        for rule in rules:
            print(f"  Rule type: {type(rule).__name__}")
            if hasattr(rule, 'formula'):
                print(f"  Formula: {rule.formula}")
            if hasattr(rule, 'operator'):
                print(f"  Operator: {rule.operator}")

# Check what column O actually is by looking at row 10
print("\n" + "="*80)
print("CHECKING ALL COLUMNS IN ROW 10")
print("="*80)

for col in range(1, 21):
    header = ws.cell(10, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}10: {header}")

wb.close()

print("\n" + "="*80)
print("DIAGNOSIS COMPLETE")
print("="*80)
