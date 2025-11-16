"""
Create v30:
1. Restore Portfolio_Dashboard from v27
2. Fix Spotlight - restore labels in F, update formulas in G with blank logic
3. Keep the good changes: Total_Proposed in Country_Dashboard/Regional_Summary, NCE in Spotlight row 2
"""
import openpyxl

print("="*80)
print("CREATING V30 - COMPREHENSIVE FIX")
print("="*80)

# Load v27 for Portfolio and v29 for everything else
print("\nLoading v27 and v29...")
wb_v27 = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx')
wb_v29 = openpyxl.load_workbook('2025-10-26-Tracker-v29.xlsx')

print("\n" + "="*80)
print("PART 1: RESTORE PORTFOLIO_DASHBOARD FROM V27")
print("="*80)

# Delete Portfolio_Dashboard from v29
print("\nRemoving Portfolio_Dashboard from v29...")
if 'Portfolio_Dashboard' in wb_v29.sheetnames:
    del wb_v29['Portfolio_Dashboard']
    print("  Removed")

# Copy Portfolio_Dashboard from v27
print("\nCopying Portfolio_Dashboard from v27...")
ws_v27_portfolio = wb_v27['Portfolio_Dashboard']
ws_v29_portfolio = wb_v29.create_sheet('Portfolio_Dashboard')

# Copy all cells
for row in ws_v27_portfolio.iter_rows():
    for cell in row:
        new_cell = ws_v29_portfolio[cell.coordinate]
        new_cell.value = cell.value
        if cell.has_style:
            new_cell.font = cell.font.copy()
            new_cell.border = cell.border.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection.copy()
            new_cell.alignment = cell.alignment.copy()

# Copy column widths
for col in ws_v27_portfolio.column_dimensions:
    if col in ws_v27_portfolio.column_dimensions:
        ws_v29_portfolio.column_dimensions[col].width = ws_v27_portfolio.column_dimensions[col].width

# Copy row heights
for row in ws_v27_portfolio.row_dimensions:
    if row in ws_v27_portfolio.row_dimensions:
        ws_v29_portfolio.row_dimensions[row].height = ws_v27_portfolio.row_dimensions[row].height

# Copy merged cells
for merged_cell_range in ws_v27_portfolio.merged_cells.ranges:
    ws_v29_portfolio.merge_cells(str(merged_cell_range))

print("  Portfolio_Dashboard restored from v27")

print("\n" + "="*80)
print("PART 2: FIX SPOTLIGHT COLUMN F (RESTORE LABELS)")
print("="*80)

ws_spotlight = wb_v29['Spotlight_PMWorkspace']

print("\nRestoring labels in column F...")

# Restore the labels that were there originally
spotlight_labels = {
    5: 'Status:',
    6: 'Progress:',
    7: 'End Date:',
    8: 'Implementer:',
    9: 'POC Email:',
    10: 'Country Count:',
}

for row, label in spotlight_labels.items():
    ws_spotlight.cell(row, 6).value = label  # Column F
    print(f"  F{row}: {label}")

print("\n" + "="*80)
print("PART 3: FIX SPOTLIGHT COLUMN G (UPDATE FORMULAS)")
print("="*80)

print("\nUpdating column G formulas with blank-instead-of-zero logic...")

# Update formulas in column G with proper blank handling
g_formulas = {
    5: ('Project_Status', 'text'),
    6: ('Project_Progress', 'numeric'),
    7: ('Project_End_Date', 'date'),
    8: ('Implementer', 'text'),
    9: ('Implementer_POC_Email', 'text'),
    10: ('Country_Count', 'numeric'),
}

for row, (field_name, field_type) in g_formulas.items():
    if field_type == 'numeric':
        # For numeric fields, check for 0
        formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))=0,"",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'
    else:
        # For text/date fields, check for empty string
        formula = f'=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[{field_name}],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'

    ws_spotlight.cell(row, 7).value = formula  # Column G
    print(f"  G{row}: {field_name} ({field_type})")

print("\n" + "="*80)
print("SAVING V30")
print("="*80)

# Move Portfolio_Dashboard to correct position (after Master_Projects)
sheets = wb_v29.sheetnames
portfolio_idx = sheets.index('Portfolio_Dashboard')
master_idx = sheets.index('Master_Projects') if 'Master_Projects' in sheets else 0
target_idx = master_idx + 1

if portfolio_idx != target_idx:
    wb_v29.move_sheet('Portfolio_Dashboard', offset=target_idx - portfolio_idx)
    print(f"  Moved Portfolio_Dashboard to position {target_idx}")

wb_v29.save('2025-10-26-Tracker-v30.xlsx')

print("\nOK - v30 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nv30 contains:")
print("  1. Portfolio_Dashboard restored from v27 (no Award/NCE columns)")
print("  2. Spotlight column F: Labels restored (Status, Progress, etc.)")
print("  3. Spotlight column G: Formulas updated with blank-instead-of-zero logic")
print("  4. Spotlight row 2: NCE_Eligible in H2 with label in G2 (kept from v29)")
print("  5. Country_Dashboard: Total_Proposed added (kept from v28)")
print("  6. Regional_Summary: Total_Proposed added (kept from v28)")

print("\nv30 ready!")
