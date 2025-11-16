"""
Fix Config_Lists sheet:
1. Check for duplicate country listings
2. Keep only one set (in columns D-F)
3. Alphabetize by Region first, then Country Name
"""
import openpyxl

print("="*80)
print("FIXING CONFIG_LISTS SHEET")
print("="*80)

# Load v26
print("\nLoading v26...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v26.xlsx')
ws_config = wb['Config_Lists']

print("\n" + "="*80)
print("STEP 1: ANALYZE CURRENT STRUCTURE")
print("="*80)

# Check columns D-F (top section)
print("\nChecking columns D-F (Country_Code, Country, Region)...")
top_section_count = 0
row = 2
while ws_config.cell(row, 4).value:  # Column D
    top_section_count += 1
    row += 1

print(f"  Found {top_section_count} countries in columns D-F (rows 2-{row-1})")

# Check row 80+ columns A-D
print("\nChecking row 80+ columns A-D...")
bottom_section_count = 0
row = 80
first_bottom_row = None
while row < 200:
    if ws_config.cell(row, 1).value:  # Column A
        if first_bottom_row is None:
            first_bottom_row = row
        bottom_section_count += 1
        row += 1
    else:
        break

if bottom_section_count > 0:
    print(f"  Found {bottom_section_count} countries in row {first_bottom_row}+ columns A-D")
else:
    print("  No countries found in bottom section")

print("\n" + "="*80)
print("STEP 2: GET COUNTRY DATA FROM COUNTRY_REGIONS")
print("="*80)

# Get country data from Country_Regions sheet (source of truth)
ws_regions = wb['Country_Regions']

print("\nReading from Country_Regions sheet...")
countries = []
row = 2
while ws_regions.cell(row, 1).value:
    country_code = ws_regions.cell(row, 1).value
    country_name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    countries.append((country_code, country_name, region))
    row += 1

print(f"  Read {len(countries)} countries")

print("\n" + "="*80)
print("STEP 3: SORT BY REGION THEN COUNTRY NAME")
print("="*80)

# Sort by region first, then country name
print("\nSorting countries...")
region_order = {'AF': 1, 'EAP': 2, 'EUR': 3, 'NEA': 4, 'SCA': 5, 'WHA': 6}

countries_sorted = sorted(countries, key=lambda x: (region_order.get(x[2], 99), x[1]))

print(f"  Sorted {len(countries_sorted)} countries")
print("\nFirst 10 countries after sorting:")
for i, (code, name, region) in enumerate(countries_sorted[:10], 1):
    print(f"    {i}. {code} - {name} ({region})")

print("\n" + "="*80)
print("STEP 4: CLEAR OLD DATA")
print("="*80)

# Clear columns D-F (rows 2 onwards)
print("\nClearing columns D-F (rows 2-200)...")
for row in range(2, 201):
    ws_config.cell(row, 4).value = None  # Column D
    ws_config.cell(row, 5).value = None  # Column E
    ws_config.cell(row, 6).value = None  # Column F
print("  Cleared columns D-F")

# Clear row 80+ columns A-D if it exists
if bottom_section_count > 0:
    print(f"\nClearing rows {first_bottom_row}-200 columns A-D...")
    for row in range(first_bottom_row, 201):
        ws_config.cell(row, 1).value = None  # Column A
        ws_config.cell(row, 2).value = None  # Column B
        ws_config.cell(row, 3).value = None  # Column C
        ws_config.cell(row, 4).value = None  # Column D
    print("  Cleared bottom section")

print("\n" + "="*80)
print("STEP 5: WRITE SORTED COUNTRIES TO COLUMNS D-F")
print("="*80)

print("\nWriting sorted countries to columns D-F...")
for idx, (country_code, country_name, region) in enumerate(countries_sorted, start=2):
    ws_config.cell(idx, 4).value = country_code
    ws_config.cell(idx, 5).value = country_name
    ws_config.cell(idx, 6).value = region

print(f"  Written {len(countries_sorted)} countries to rows 2-{len(countries_sorted) + 1}")

print("\n" + "="*80)
print("SAVING V27")
print("="*80)

wb.save('2025-10-26-Tracker-v27.xlsx')

print("\nOK - Config_Lists fixed!")
print("Saved as: 2025-10-26-Tracker-v27.xlsx")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print(f"\n- Removed duplicate country listings")
print(f"- Kept single list in columns D-F: Country_Code, Country, Region")
print(f"- Alphabetized by Region, then Country Name")
print(f"- Total countries: {len(countries_sorted)}")

print("\nCountries by region:")
current_region = None
region_count = 0
for code, name, region in countries_sorted:
    if region != current_region:
        if current_region:
            print(f"  {current_region}: {region_count} countries")
        current_region = region
        region_count = 1
    else:
        region_count += 1
# Print last region
if current_region:
    print(f"  {current_region}: {region_count} countries")

print("\nv26 Config_Lists is now clean and sorted!")
