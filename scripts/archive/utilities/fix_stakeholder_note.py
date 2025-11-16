"""
Fix the stakeholder note - add it BELOW the table instead of above
This won't break the table structure
"""
import openpyxl

print("="*80)
print("FIXING STAKEHOLDER NOTE PLACEMENT")
print("="*80)

# Load v20 (before we broke it)
print("\nLoading v20 (working version)...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v20.xlsx')

# First, copy all changes from v21 back to v20
print("\nApplying v21 changes to v20...")

# Step 1: Portfolio Dashboard proposed column changes
ws_port = wb['Portfolio_Dashboard']
print("\n1. Updating Portfolio_Dashboard formulas...")
for row in range(11, 21):
    ws_port.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")'
    ws_port.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Name],ROW()-10),""))'
    ws_port.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Status],ROW()-10),""))'
    ws_port.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],ROW()-10),""))'
    ws_port.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Progress],ROW()-10),""))'
    ws_port.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Countries],ROW()-10),""))'
    ws_port.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Proposed],ROW()-10),""))'
    ws_port.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Allocation],ROW()-10),""))'
    ws_port.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Obligated],ROW()-10),""))'
    ws_port.cell(row, 10).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Spent],ROW()-10),""))'
    ws_port.cell(row, 11).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_ULO],ROW()-10),""))'
    ws_port.cell(row, 12).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[ULO_Percent],ROW()-10),""))'
    ws_port.cell(row, 13).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Days_Remaining],ROW()-10),""))'
print("  OK - Portfolio Dashboard updated")

# Step 2: Spotlight Award/NCE fields
ws_spot = wb['Spotlight_PMWorkspace']
print("\n2. Adding Spotlight Award/NCE fields...")
ws_spot['E2'] = 'Award Number:'
ws_spot['F2'] = '=IFERROR(INDEX(T_Master_Projects[Award_Number],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['H2'] = 'NCE Status:'
ws_spot['I2'] = '=IFERROR(INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['E3'] = 'POP Start:'
ws_spot['F3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Start],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['G3'] = 'POP End:'
ws_spot['H3'] = '=IFERROR(INDEX(T_Master_Projects[POP_End],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['I3'] = 'POP Days Remaining:'
ws_spot['J3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Days_Remaining],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
print("  OK - Award/NCE fields added")

# Step 3: Spotlight stakeholder formulas
print("\n3. Adding Spotlight stakeholder formulas...")
for row in range(33, 43):
    idx = row - 32
    ws_spot.cell(row, 1).value = f'=IFERROR(INDEX(T_Project_Stakeholders[Stakeholder_Name],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),"")'
    ws_spot.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Organization],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Role],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Email],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Engagement_Level],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
print("  OK - Stakeholder formulas added")

# Step 4: Config_Lists headers
ws_config = wb['Config_Lists']
print("\n4. Updating Config_Lists headers...")
ws_config['A1'] = 'Status'
ws_config['B1'] = 'Priority'
ws_config['C1'] = 'Stage'
ws_config['D1'] = 'Country_Code'
ws_config['E1'] = 'Country'
ws_config['F1'] = 'Region'
print("  OK - Config_Lists headers updated")

# Step 5: Add stakeholder note - but in a SAFE location
# Add it in column J (to the right of the table, won't interfere)
ws_stake = wb['Stakeholders']
print("\n5. Adding stakeholder note (safe location - column J)...")

note_text = """STAKEHOLDER CATEGORIZATION GUIDE:

Government Officials: Gov agency contacts, embassy staff, ministry officials

Implementing Partners: Contractors, grantees, vendors executing project work

Internal Team: DOS staff, project managers, technical advisors

Beneficiaries: End users, communities, target populations

Oversight: Congress, GAO, OIG, monitoring entities

Stakeholder_ID Format: ProjectID-STK-001, ProjectID-STK-002, etc."""

ws_stake['J1'] = note_text
ws_stake['J1'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
ws_stake['J1'].fill = openpyxl.styles.PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
ws_stake['J1'].font = openpyxl.styles.Font(size=9, italic=True)

# Make column J wider to accommodate the note
ws_stake.column_dimensions['J'].width = 60

print("  OK - Note added in column J (won't break table)")

# Save as v22
print("\n" + "="*80)
print("SAVING AS V22...")
print("="*80)

wb.save('2025-10-26-Tracker-v22.xlsx')

print("\nOK - V22 CREATED!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nv22 includes all v21 features:")
print("  OK - Portfolio Dashboard with Total_Proposed and blank row handling")
print("  OK - Spotlight with Award/NCE fields and stakeholder formulas")
print("  OK - Config_Lists headers updated")
print("  OK - Stakeholder categorization note (in column J, safe location)")

print("\nThe stakeholder note is now in column J (to the right)")
print("This preserves the T_Project_Stakeholders table (columns A-H)")

print("\nv22 should open without errors!")
