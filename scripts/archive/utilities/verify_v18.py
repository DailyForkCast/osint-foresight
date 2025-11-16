"""
Verify v18 has zero formulas
"""
import openpyxl

print("="*80)
print("VERIFYING V18 - FORMULA CHECK")
print("="*80)

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

formula_count = 0
formula_locations = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nChecking {sheet_name}...")

    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula type
                formula_count += 1
                formula_locations.append(f"{sheet_name}!{cell.coordinate}: {cell.value}")

    print(f"  Formulas found: {len([f for f in formula_locations if sheet_name in f])}")

print("\n" + "="*80)
if formula_count == 0:
    print("SUCCESS! ZERO FORMULAS FOUND")
    print("="*80)
    print("\nFile should open in Excel without any errors!")
    print("\nNext steps:")
    print("1. Open 2025-10-26-Tracker-v18.xlsx in Excel")
    print("2. Go to FORMULA_GUIDE sheet")
    print("3. Copy formulas from column B and paste into appropriate cells")
    print("   (Should take about 5 minutes)")
else:
    print(f"WARNING! {formula_count} FORMULAS FOUND:")
    print("="*80)
    for loc in formula_locations:
        print(f"  {loc}")
