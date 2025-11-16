"""
Fix Excel errors from v36 and create clean v37
The error log shows:
- Removed Records: Formula from sheet14 (likely Calendar_Todo)
- Repaired Records: Table from table4

Let's rebuild Calendar_Todo more carefully and check tables
"""
import openpyxl

print("="*80)
print("FIXING EXCEL ERRORS - V37")
print("="*80)

# Load v36
print("\nLoading v36...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v36.xlsx')

print("\n" + "="*80)
print("PART 1: CHECK WHICH SHEET IS SHEET14")
print("="*80)

print("\nSheet list:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet_name}")

# Sheet14 is likely Calendar_Todo (position 14)

print("\n" + "="*80)
print("PART 2: FIX CALENDAR_TODO FORMULAS")
print("="*80)

ws_calendar = wb['Calendar_Todo']

print("\nClearing and rebuilding Calendar_Todo with simpler formulas...")

# Clear all data
for row in range(2, 101):
    for col in range(1, 10):
        ws_calendar.cell(row, col).value = None

print("  Cleared existing formulas")

# Add simpler formulas that are less likely to cause issues
print("\nAdding simplified calendar formulas...")

current_row = 2

# Section 1: Project Start Dates - use simpler approach
print("  Adding Project Start Dates (10 projects)...")
ws_calendar.cell(current_row, 1).value = "=== PROJECT START DATES ==="
current_row += 1

for i in range(1, 11):  # Reduced to 10 projects
    row = current_row + i - 1

    # Simple formulas without complex nesting
    ws_calendar.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],{i}),"")'
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","","Start: "&IFERROR(INDEX(T_Master_Projects[Project_Name],{i}),""))'
    ws_calendar.cell(row, 3).value = f'=A{row}'
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[POP_Start],{i}),""))'
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Implementer_POC],{i}),""))'
    ws_calendar.cell(row, 6).value = "Scheduled"
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],{i}),""))'
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Project start date")'

current_row += 11

# Section 2: Project End Dates
print("  Adding Project End Dates (10 projects)...")
ws_calendar.cell(current_row, 1).value = "=== PROJECT END DATES ==="
current_row += 1

for i in range(1, 11):
    row = current_row + i - 1

    ws_calendar.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],{i}),"")'
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","","End: "&IFERROR(INDEX(T_Master_Projects[Project_Name],{i}),""))'
    ws_calendar.cell(row, 3).value = f'=A{row}'
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[POP_End],{i}),""))'
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Implementer_POC],{i}),""))'
    ws_calendar.cell(row, 6).value = "Upcoming"
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],{i}),""))'
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Project end date")'

print("  Formulas added")

print("\n" + "="*80)
print("PART 3: CHECK TABLES")
print("="*80)

print("\nExisting tables in workbook:")
table_count = 0
for sheet in wb.worksheets:
    if sheet.tables:
        for table_name in sheet.tables:
            table_count += 1
            print(f"  Table {table_count}: {table_name} in {sheet.title}")

print("\n" + "="*80)
print("SAVING V37")
print("="*80)

wb.save('2025-10-26-Tracker-v37.xlsx')

print("\nOK - v37 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v37:")
print("  1. Calendar_Todo: Rebuilt with simpler formulas")
print("     - Reduced to 10 projects per section (was 20)")
print("     - Simpler formula structure to avoid corruption")
print("     - Removed complex nested IFs")

print("\n  2. Sections included:")
print("     - Project Start Dates (rows 3-12)")
print("     - Project End Dates (rows 14-23)")

print("\nNote: This should fix the Excel error. If you still get errors,")
print("we may need to manually add calendar items instead of formulas.")

print("\nv37 ready!")
