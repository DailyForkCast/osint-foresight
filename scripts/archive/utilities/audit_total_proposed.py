"""
Audit all sheets to find where budget columns exist and check for Total_Proposed
"""
import openpyxl

print("="*80)
print("AUDITING ALL SHEETS FOR TOTAL_PROPOSED")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

print(f"\nFound {len(wb.sheetnames)} sheets")

# Budget-related keywords to look for
budget_keywords = ['allocated', 'obligated', 'spent', 'ulo', 'proposed', 'budget', 'amount']

print("\n" + "="*80)
print("SHEET-BY-SHEET ANALYSIS")
print("="*80)

results = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    print(f"\n{sheet_name}")
    print("-" * len(sheet_name))

    # Check first 20 rows for headers
    budget_cols = {}
    has_proposed = False

    for row_num in range(1, min(21, ws.max_row + 1)):
        for col_num in range(1, min(30, ws.max_column + 1)):
            cell_value = ws.cell(row_num, col_num).value
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower().replace('_', ' ').replace('-', ' ')

                # Check for budget keywords
                for keyword in budget_keywords:
                    if keyword in cell_lower:
                        col_letter = openpyxl.utils.get_column_letter(col_num)
                        if col_letter not in budget_cols:
                            budget_cols[col_letter] = []
                        budget_cols[col_letter].append((row_num, cell_value, keyword))

                        if 'proposed' in cell_lower:
                            has_proposed = True

    if budget_cols:
        print(f"  Budget-related columns found:")
        for col_letter, matches in sorted(budget_cols.items()):
            for row_num, cell_value, keyword in matches:
                # Clean cell_value to avoid unicode issues
                clean_value = str(cell_value).encode('ascii', 'ignore').decode('ascii')
                print(f"    {col_letter}{row_num}: {clean_value} ('{keyword}')")

        if has_proposed:
            print(f"  -> HAS Total_Proposed")
        else:
            print(f"  -> MISSING Total_Proposed")

        results[sheet_name] = {
            'has_budget_cols': True,
            'has_proposed': has_proposed,
            'columns': budget_cols
        }
    else:
        print(f"  No budget columns found")
        results[sheet_name] = {
            'has_budget_cols': False,
            'has_proposed': False,
            'columns': {}
        }

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

needs_proposed = []
has_proposed = []
no_budget = []

for sheet_name, info in results.items():
    if info['has_budget_cols']:
        if info['has_proposed']:
            has_proposed.append(sheet_name)
        else:
            needs_proposed.append(sheet_name)
    else:
        no_budget.append(sheet_name)

print(f"\nSheets with budget columns AND Total_Proposed ({len(has_proposed)}):")
for sheet in has_proposed:
    print(f"  - {sheet}")

print(f"\nSheets with budget columns but MISSING Total_Proposed ({len(needs_proposed)}):")
for sheet in needs_proposed:
    print(f"  - {sheet}")

print(f"\nSheets with no budget columns ({len(no_budget)}):")
for sheet in no_budget:
    print(f"  - {sheet}")

print("\n" + "="*80)
print("NEXT STEPS")
print("="*80)

if needs_proposed:
    print(f"\nNeed to add Total_Proposed to {len(needs_proposed)} sheets:")
    for sheet in needs_proposed:
        print(f"  1. {sheet}")
        print(f"     Columns found: {list(results[sheet]['columns'].keys())}")
else:
    print("\nAll sheets with budget columns already have Total_Proposed!")

print("\nAudit complete!")
