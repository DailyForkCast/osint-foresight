"""
Add Total_Proposed column to Portfolio_Dashboard column G
Shift other budget columns to the right
"""
import openpyxl

print("="*80)
print("UPDATING PORTFOLIO_DASHBOARD - ADDING PROPOSED COLUMN")
print("="*80)

# Load v20
print("\nLoading v20...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v20.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\nUpdating project list formulas (rows 11-20)...")
print("New column order:")
print("  G: Total_Proposed")
print("  H: Total_Allocation")
print("  I: Total_Obligated")
print("  J: Total_Spent")
print("  K: Total_ULO")
print("  L: ULO_Percent")
print("  M: Days_Remaining")

for row in range(11, 21):
    # A-F stay the same (Project ID, Name, Status, Priority, Progress, Countries)

    # G: Total_Proposed (NEW)
    ws_port.cell(row, 7).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Proposed],ROW()-10),"")'

    # H: Total_Allocation (shifted from G)
    ws_port.cell(row, 8).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Allocation],ROW()-10),"")'

    # I: Total_Obligated (shifted from H)
    ws_port.cell(row, 9).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Obligated],ROW()-10),"")'

    # J: Total_Spent (NEW - good to show)
    ws_port.cell(row, 10).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Spent],ROW()-10),"")'

    # K: Total_ULO (shifted from I)
    ws_port.cell(row, 11).value = f'=IFERROR(INDEX(T_Master_Projects[Total_ULO],ROW()-10),"")'

    # L: ULO_Percent (shifted from J)
    ws_port.cell(row, 12).value = f'=IFERROR(INDEX(T_Master_Projects[ULO_Percent],ROW()-10),"")'

    # M: Days_Remaining (shifted from K)
    ws_port.cell(row, 13).value = f'=IFERROR(INDEX(T_Master_Projects[Days_Remaining],ROW()-10),"")'

print("\nOK - Portfolio Dashboard updated with Proposed column")

# Save as v21 (v20 is open in Excel)
print("\nSaving as v21...")
wb.save('2025-10-26-Tracker-v21.xlsx')

print("\n" + "="*80)
print("PORTFOLIO_DASHBOARD UPDATED!")
print("="*80)

print("\nSaved as: 2025-10-26-Tracker-v21.xlsx")
print("\nColumn G now shows Total_Proposed for each project")
print("All other budget columns shifted right")
print("\nDon't forget to update the column headers in row 10:")
print("  G10: Total Proposed")
print("  H10: Total Allocation")
print("  I10: Total Obligated")
print("  J10: Total Spent")
print("  K10: Total ULO")
print("  L10: ULO %")
print("  M10: Days Remaining")
print("\nYou can close v20 and use v21, or copy v21 over v20 after closing it")
