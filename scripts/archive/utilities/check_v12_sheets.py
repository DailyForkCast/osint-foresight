"""
Check what sheets are in v12
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=True)

print("="*80)
print("V12 SHEETS")
print("="*80)

for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"{i:2}. {name:<35} ({ws.max_row} rows x {ws.max_column} cols)")

print("\n" + "="*80)
print("V18 SHEETS (Current)")
print("="*80)

wb_v18 = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')
for i, name in enumerate(wb_v18.sheetnames, 1):
    ws = wb_v18[name]
    print(f"{i:2}. {name:<35} ({ws.max_row} rows x {ws.max_column} cols)")

print("\n" + "="*80)
print("MISSING SHEETS (in v12 but not v18)")
print("="*80)

v12_sheets = set(wb.sheetnames)
v18_sheets = set(wb_v18.sheetnames)
missing = v12_sheets - v18_sheets

for name in sorted(missing):
    print(f"  - {name}")
