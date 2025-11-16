"""
Update Portfolio_Dashboard formulas to show truly blank rows when no project data exists
Uses IF to check if Project_Unique_ID exists before displaying other columns
"""
import openpyxl

print("="*80)
print("UPDATING PORTFOLIO_DASHBOARD - BLANK ROW HANDLING")
print("="*80)

# Load v21
print("\nLoading v21...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v21.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\nUpdating project list formulas (rows 11-20)...")
print("Strategy: Column A shows Project ID, other columns are blank if A is blank")

for row in range(11, 21):
    # A: Project ID (master check - if this is blank, row has no project)
    ws_port.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")'

    # B: Project Name (blank if A is blank)
    ws_port.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Name],ROW()-10),""))'

    # C: Status
    ws_port.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Status],ROW()-10),""))'

    # D: Priority
    ws_port.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],ROW()-10),""))'

    # E: Progress
    ws_port.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Progress],ROW()-10),""))'

    # F: Countries
    ws_port.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Countries],ROW()-10),""))'

    # G: Total_Proposed
    ws_port.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Proposed],ROW()-10),""))'

    # H: Total_Allocation
    ws_port.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Allocation],ROW()-10),""))'

    # I: Total_Obligated
    ws_port.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Obligated],ROW()-10),""))'

    # J: Total_Spent
    ws_port.cell(row, 10).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Spent],ROW()-10),""))'

    # K: Total_ULO
    ws_port.cell(row, 11).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_ULO],ROW()-10),""))'

    # L: ULO_Percent
    ws_port.cell(row, 12).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[ULO_Percent],ROW()-10),""))'

    # M: Days_Remaining
    ws_port.cell(row, 13).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Days_Remaining],ROW()-10),""))'

print("  OK - All columns now blank when Project ID is blank")

# Save
print("\nSaving v21...")
wb.save('2025-10-26-Tracker-v21.xlsx')

print("\n" + "="*80)
print("PORTFOLIO_DASHBOARD UPDATED!")
print("="*80)

print("\nFormula pattern:")
print("  Column A: =IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),\"\")")
print("  Columns B-M: =IF(A{row}=\"\",\"\",IFERROR(INDEX(...))")
print("\nResult:")
print("  - If there's no project in that position, entire row is blank")
print("  - If there's a project but a field is missing, just that cell is blank")
print("\nv21 is ready to use!")
