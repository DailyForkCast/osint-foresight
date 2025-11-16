"""
Detailed recommendations for each sheet
"""
import openpyxl

print("="*80)
print("DETAILED SHEET-BY-SHEET RECOMMENDATIONS")
print("="*80)

wb = openpyxl.load_workbook('2025-10-26-Tracker-v44.xlsx')

recommendations = {}

# ============================================================================
# CONTROL
# ============================================================================
print("\n" + "="*80)
print("1. CONTROL")
print("="*80)

ws = wb['Control']
print("\nISSUES FOUND:")
print("  - E5 has #REF! error: =B9/COUNTA(#REF!)")

print("\nRECOMMENDATIONS:")
print("  AUTOMATED:")
print("    - Fix E5 formula (Projects/Country)")
print("      Current: =B9/COUNTA(#REF!)")
print("      Should be: =B9/E8 (Total Projects / Countries Active)")
print("  MANUAL:")
print("    - Add data validation for ULO Warning threshold (B5)")
print("    - Add conditional formatting for ULO warning indicators")
print("    - Consider adding refresh date automation")

# ============================================================================
# MASTER_PROJECTS
# ============================================================================
print("\n" + "="*80)
print("2. MASTER_PROJECTS")
print("="*80)

ws = wb['Master_Projects']

# Count how many projects have data
project_count = 0
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 2).value:  # Project_Unique_ID column
        project_count += 1
    else:
        break

print(f"\nCurrent projects: {project_count}")
print(f"Table size: {ws.max_row} rows")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Fill in missing project data (only 2 projects currently)")
print("    - Add data validation for:")
print("      * Fiscal_Year (dropdown)")
print("      * Project_Status (dropdown from Config_Lists)")
print("      * Project_Priority (dropdown from Config_Lists)")
print("      * Countries (multi-select or validation)")
print("    - Add conditional formatting:")
print("      * Highlight overdue projects (POP_End_Date < TODAY)")
print("      * Color-code by priority")
print("      * Progress bars for Project_Progress")
print("    - Consider protecting column formulas")

# ============================================================================
# PORTFOLIO_DASHBOARD
# ============================================================================
print("\n" + "="*80)
print("3. PORTFOLIO_DASHBOARD")
print("="*80)

ws = wb['Portfolio_Dashboard']

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add conditional formatting:")
print("      * Red text for ULO > 75%")
print("      * Green/Yellow/Red for Progress levels")
print("      * Days Remaining < 90 highlighted")
print("    - Add sparklines or charts for visual summary")
print("    - Consider adding filters/slicers")
print("    - Add Award Number column (user will do this)")
print("    - Add NCE columns (user will do this)")

# ============================================================================
# COUNTRY_DASHBOARD
# ============================================================================
print("\n" + "="*80)
print("4. COUNTRY_DASHBOARD")
print("="*80)

ws = wb['Country_Dashboard']

print("\nISSUES FOUND:")
e5_val = ws['E5'].value
print(f"  - B8 has #REF! error: {ws['B8'].value}")

print("\nRECOMMENDATIONS:")
print("  AUTOMATED:")
print("    - Fix B8 formula (Country PM lookup)")
print("      Current: =IFERROR(INDEX(#REF!,MATCH($B$2,#REF!,0)),\"\")")
print("      Should be: =IFERROR(INDEX(Country_PM_Assignments[Project_Manager],")
print("                   MATCH($B$2,Country_PM_Assignments[Country_Code],0)),\"\")")
print("  MANUAL:")
print("    - Add data validation for B2 (Country dropdown)")
print("    - Add conditional formatting:")
print("      * ULO % thresholds")
print("      * Project status colors")
print("    - Consider adding country flag or region indicator")

# ============================================================================
# SPOTLIGHT_PMWORKSPACE
# ============================================================================
print("\n" + "="*80)
print("5. SPOTLIGHT_PMWORKSPACE")
print("="*80)

ws = wb['Spotlight_PMWorkspace']

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add data validation for B2 (Project ID dropdown)")
print("    - Add conditional formatting:")
print("      * Status colors")
print("      * Progress indicator")
print("      * Days Remaining alerts")
print("    - Fill in NCE data (H2, I2)")
print("    - Complete deliverables section with formulas")
print("    - Add milestones section")
print("    - Add events section")

# ============================================================================
# COUNTRY_BUDGETS
# ============================================================================
print("\n" + "="*80)
print("6. COUNTRY_BUDGETS")
print("="*80)

ws = wb['Country_Budgets']

# Count rows with data
data_rows = 0
for row in range(2, min(100, ws.max_row + 1)):
    if ws.cell(row, 2).value:  # Unique_ID
        data_rows += 1

print(f"\nData rows: {data_rows}")
print(f"Table size: 1001 rows")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Fill in budget data for all projects")
print("    - Add data validation:")
print("      * Unique_ID (from Master_Projects)")
print("      * Country_Code (from Country_Regions)")
print("    - Add conditional formatting:")
print("      * ULO % thresholds")
print("      * My_Country = True highlighting")
print("    - Consider adding totals row")

# ============================================================================
# COUNTRY_REGIONS
# ============================================================================
print("\n" + "="*80)
print("7. COUNTRY_REGIONS")
print("="*80)

ws = wb['Country_Regions']

print(f"\nCountries: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Verify all countries are correct")
print("    - Add missing countries if needed")
print("    - Review EU_Member status")
print("    - Review Subregion assignments")

# ============================================================================
# CONFIG_LISTS
# ============================================================================
print("\n" + "="*80)
print("8. CONFIG_LISTS")
print("="*80)

ws = wb['Config_Lists']

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Review Status list (A2-A10)")
print("    - Review Priority list (B2-B5)")
print("    - Review Stage list (C2-C5)")
print("    - Add more configuration lists as needed:")
print("      * Document types")
print("      * Event types")
print("      * Deliverable types")
print("      * Technology categories")

# ============================================================================
# COUNTRY_PM_ASSIGNMENTS
# ============================================================================
print("\n" + "="*80)
print("9. COUNTRY_PM_ASSIGNMENTS")
print("="*80)

ws = wb['Country_PM_Assignments']

tbd_count = 0
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 4).value == 'TBD':
        tbd_count += 1

print(f"\nCountries with TBD PM: {tbd_count} of {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Assign Project Managers to all countries")
print("    - Fill in PM_Email addresses")
print("    - Add PM_Phone if needed")
print("    - Consider adding backup PM column")

# ============================================================================
# MILESTONES
# ============================================================================
print("\n" + "="*80)
print("10. MILESTONES")
print("="*80)

ws = wb['Milestones']

print(f"\nMilestones: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add milestones for all projects")
print("    - Add data validation:")
print("      * Unique_ID (from Master_Projects)")
print("      * Status (dropdown)")
print("    - Add conditional formatting:")
print("      * Overdue milestones (Due_Date < TODAY and Status != Complete)")
print("      * Upcoming milestones (Due_Date within 30 days)")
print("    - Consider converting to table (T_Milestones)")

# ============================================================================
# EVENTS
# ============================================================================
print("\n" + "="*80)
print("11. EVENTS")
print("="*80)

ws = wb['Events']

print(f"\nEvents: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add events for projects")
print("    - Add data validation:")
print("      * Unique_ID (from Master_Projects)")
print("      * Event_Type (dropdown)")
print("    - Add Date column for event date")
print("    - Consider converting to table (T_Events)")
print("    - Add conditional formatting for upcoming events")

# ============================================================================
# STAKEHOLDERS
# ============================================================================
print("\n" + "="*80)
print("12. STAKEHOLDERS")
print("="*80)

ws = wb['Stakeholders']

print(f"\nStakeholders: {ws.max_row - 1}")
print(f"Columns: {ws.max_column}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add stakeholder data")
print("    - Add data validation:")
print("      * Location_Country (from Country_Regions)")
print("    - Consider linking to projects")
print("    - Add Role/Type column if needed")

# ============================================================================
# REGIONAL_SUMMARY
# ============================================================================
print("\n" + "="*80)
print("13. REGIONAL_SUMMARY")
print("="*80)

ws = wb['Regional_Summary']

print("\nRECOMMENDATIONS:")
print("  AUTOMATED:")
print("    - Verify all formulas are working correctly")
print("  MANUAL:")
print("    - Fix Region_Name column (seems incorrect)")
print("      AF should be Africa, not Europe")
print("      EUR should be Europe, not Western Hemisphere")
print("      etc.")
print("    - Add conditional formatting:")
print("      * Highlight regions with high ULO %")
print("      * Color-code by region")
print("    - Add charts/visualizations")

# ============================================================================
# CALENDAR_TODO
# ============================================================================
print("\n" + "="*80)
print("14. CALENDAR_TODO")
print("="*80)

ws = wb['Calendar_Todo']

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add tasks as needed")
print("    - Add data validation:")
print("      * Unique_ID (from Master_Projects)")
print("      * Status (dropdown)")
print("      * Assigned_To (dropdown of team members)")
print("    - Add conditional formatting:")
print("      * Overdue tasks")
print("      * Tasks due soon")
print("    - Consider converting to table")

# ============================================================================
# PROJECT_DELIVERABLES
# ============================================================================
print("\n" + "="*80)
print("15. PROJECT_DELIVERABLES")
print("="*80)

ws = wb['Project_Deliverables']

print(f"\nDeliverables: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add more deliverables")
print("    - Add data validation:")
print("      * Project_ID (from Master_Projects)")
print("      * Deliverable_Type (dropdown)")
print("      * Status (dropdown)")
print("    - Fix Completion_Percent formatting (remove %)")
print("    - Add conditional formatting:")
print("      * Overdue deliverables")
print("      * Color-code by status")

# ============================================================================
# PROJECT_AUDIENCES
# ============================================================================
print("\n" + "="*80)
print("16. PROJECT_AUDIENCES")
print("="*80)

ws = wb['Project_Audiences']

print(f"\nAudiences: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add audience data for all projects")
print("    - Add data validation:")
print("      * Project_ID (from Master_Projects)")
print("      * Audience_Type (dropdown)")
print("      * Priority (dropdown)")
print("    - Consider expanding table if more audience types needed")

# ============================================================================
# PROJECT_TECHNOLOGIES
# ============================================================================
print("\n" + "="*80)
print("17. PROJECT_TECHNOLOGIES")
print("="*80)

ws = wb['Project_Technologies']

print(f"\nTechnologies: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add technology data for all projects")
print("    - Add data validation:")
print("      * Project_ID (from Master_Projects)")
print("      * Category (dropdown)")
print("      * Status (dropdown)")
print("    - Consider expanding table")

# ============================================================================
# PROJECT_DOCUMENTS
# ============================================================================
print("\n" + "="*80)
print("18. PROJECT_DOCUMENTS")
print("="*80)

ws = wb['Project_Documents']

print(f"\nDocuments: {ws.max_row - 1}")

print("\nRECOMMENDATIONS:")
print("  MANUAL:")
print("    - Add document links")
print("    - Add data validation:")
print("      * Project_Unique_ID (from Master_Projects)")
print("      * Document_Type (dropdown)")
print("    - Consider adding upload date")
print("    - Consider adding document owner")

wb.close()

print("\n" + "="*80)
print("SUMMARY OF TOP PRIORITIES")
print("="*80)

print("\n**CRITICAL AUTOMATED FIXES:**")
print("  1. Fix Control E5 #REF! error")
print("  2. Fix Country_Dashboard B8 #REF! error")
print("  3. Fix Regional_Summary region names (AF=Africa, EUR=Europe, etc.)")

print("\n**HIGH PRIORITY MANUAL TASKS:**")
print("  1. Add data validation to Master_Projects:")
print("     - Status, Priority, Countries dropdowns")
print("  2. Add data validation to Country_Dashboard:")
print("     - Country dropdown in B2")
print("  3. Add data validation to Spotlight_PMWorkspace:")
print("     - Project ID dropdown in B2")
print("  4. Add conditional formatting to Portfolio_Dashboard:")
print("     - ULO % thresholds, Progress, Days Remaining")
print("  5. Fill in Country PM assignments (currently all TBD)")

print("\n**MEDIUM PRIORITY MANUAL TASKS:**")
print("  6. Add more project data to Master_Projects")
print("  7. Add budget data to Country_Budgets")
print("  8. Add milestones for all projects")
print("  9. Add conditional formatting across all dashboards")
print("  10. Add stakeholder data")

print("\n**LOW PRIORITY ENHANCEMENTS:**")
print("  11. Add charts and visualizations")
print("  12. Add events and deliverables")
print("  13. Add audience and technology mappings")
print("  14. Add document links")
print("  15. Expand Config_Lists with more options")

print("\nDone!")
