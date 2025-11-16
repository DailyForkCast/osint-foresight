"""
Check v18 structure to verify all changes are in place
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

print("="*80)
print("V18 STRUCTURE VERIFICATION")
print("="*80)

# Check Master_Projects headers
print("\n1. MASTER_PROJECTS HEADERS:")
ws = wb['Master_Projects']
headers = [ws.cell(1, col).value for col in range(1, 21)]
for i, h in enumerate(headers, 1):
    print(f"   Col {chr(64+i)}: {h}")

# Check Country_Budgets headers
print("\n2. COUNTRY_BUDGETS HEADERS:")
ws = wb['Country_Budgets']
headers = [ws.cell(1, col).value for col in range(1, 13)]
for i, h in enumerate(headers, 1):
    print(f"   Col {chr(64+i)}: {h}")

# Verify budget structure
print("\n3. BUDGET COLUMN ORDER:")
print(f"   F: {ws.cell(1, 6).value} (NEW!)")
print(f"   G: {ws.cell(1, 7).value}")
print(f"   H: {ws.cell(1, 8).value}")
print(f"   I: {ws.cell(1, 9).value}")
print(f"   J: {ws.cell(1, 10).value} (awaiting formula: =H-I)")
print(f"   K: {ws.cell(1, 11).value} (awaiting formula: =IF(H>0,J/H,0))")

# Check Country_Regions sample
print("\n4. COUNTRY_REGIONS SAMPLE (First 5 countries):")
ws = wb['Country_Regions']
print(f"   {'Code':<6} {'Name':<30} {'Region':<6} {'EU':<4} {'Subregion':<25}")
print(f"   {'-'*6} {'-'*30} {'-'*6} {'-'*4} {'-'*25}")
for row in range(2, min(7, ws.max_row + 1)):
    code = ws.cell(row, 1).value or ""
    name = ws.cell(row, 2).value or ""
    region = ws.cell(row, 3).value or ""
    eu = ws.cell(row, 4).value or ""
    sub = ws.cell(row, 5).value or ""
    print(f"   {code:<6} {name:<30} {region:<6} {str(eu):<4} {sub:<25}")

# Check region codes
print("\n5. REGION CODES USED:")
regions = set()
for row in range(2, ws.max_row + 1):
    region = ws.cell(row, 3).value
    if region:
        regions.add(region)
print(f"   {', '.join(sorted(regions))}")
if regions <= {'AF', 'EAP', 'EUR', 'NEA', 'SCA', 'WHA'}:
    print("   ✅ All regions use State Dept codes!")
else:
    print("   ❌ Non-standard regions found!")

# Count sheets
print("\n6. SHEETS IN WORKBOOK:")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"   {i}. {name}")

print("\n" + "="*80)
print("VERIFICATION COMPLETE")
print("="*80)
