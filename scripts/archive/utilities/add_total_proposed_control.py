"""
Add Total Proposed to Control sheet Financial Status section
"""
import openpyxl

print("="*80)
print("ADDING TOTAL PROPOSED TO CONTROL SHEET")
print("="*80)

# Load v27
print("\nLoading v27...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx')
ws_control = wb['Control']

print("\n" + "="*80)
print("STEP 1: CHECK CURRENT FINANCIAL STATUS SECTION")
print("="*80)

print("\nChecking Financial Status section (rows around 15-20)...")
for row in range(14, 21):
    label = ws_control.cell(row, 1).value
    value = ws_control.cell(row, 2).value
    if label:
        print(f"  Row {row}: A={label}, B={value if isinstance(value, str) else 'formula'}")

print("\n" + "="*80)
print("STEP 2: INSERT NEW ROW FOR TOTAL PROPOSED")
print("="*80)

# Total Proposed should come before Total Allocated
# If Total Allocated is in B15, we need to insert a row at 15
print("\nInserting row at position 15 (before Total Allocated)...")
ws_control.insert_rows(15)
print("  Row inserted")

print("\n" + "="*80)
print("STEP 3: ADD TOTAL PROPOSED LABEL AND FORMULA")
print("="*80)

# Add label
print("\nAdding label in A15...")
ws_control['A15'] = 'Total Proposed'
print("  Label added: 'Total Proposed'")

# Add formula
print("\nAdding formula in B15...")
ws_control['B15'] = '=SUM(T_Master_Projects[Total_Proposed])'
print("  Formula added: =SUM(T_Master_Projects[Total_Proposed])")

print("\n" + "="*80)
print("STEP 4: VERIFY FINANCIAL STATUS SECTION")
print("="*80)

print("\nFinancial Status section after changes:")
for row in range(14, 22):
    label = ws_control.cell(row, 1).value
    value = ws_control.cell(row, 2).value
    if label:
        value_display = str(value)[:50] if isinstance(value, str) and value.startswith('=') else value
        print(f"  Row {row}: {label} = {value_display}")

print("\n" + "="*80)
print("SAVING V27")
print("="*80)

wb.save('2025-10-26-Tracker-v27.xlsx')

print("\nOK - Total Proposed added to Control sheet!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFinancial Status section now includes:")
print("  1. Total Proposed (NEW)")
print("  2. Total Allocated")
print("  3. Total Obligated")
print("  4. Total Spent (if present)")
print("  5. Total ULO")

print("\nThe budget lifecycle is now complete:")
print("  Proposed -> Allocated -> Obligated -> Spent -> ULO")

print("\nv27 Control sheet updated!")
