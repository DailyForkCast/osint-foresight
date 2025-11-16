"""
Update Portfolio_Dashboard formulas to show blank instead of 0
for numeric columns (budgets, percentages, days remaining)
"""
import openpyxl

print("="*80)
print("UPDATING PORTFOLIO_DASHBOARD - NO ZEROS")
print("="*80)

# Load v23
print("\nLoading v23...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v23.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\nUpdating formulas for rows 11-20...")
print("Numeric columns will show blank instead of 0")

for row in range(11, 21):
    # A: Project ID (master check)
    ws_port.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")'

    # B-F: Text columns (Project Name, Status, Priority, Progress, Countries)
    # These stay as-is (blank if no data)
    ws_port.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Name],ROW()-10),""))'
    ws_port.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Status],ROW()-10),""))'
    ws_port.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],ROW()-10),""))'
    ws_port.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Progress],ROW()-10),""))'
    ws_port.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Countries],ROW()-10),""))'

    # G-M: Numeric columns (budgets, percentages, days)
    # Updated to show blank if value is 0 or blank

    # G: Total_Proposed
    ws_port.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Proposed],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Proposed],ROW()-10)),""))'

    # H: Total_Allocation
    ws_port.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Allocation],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Allocation],ROW()-10)),""))'

    # I: Total_Obligated
    ws_port.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Obligated],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Obligated],ROW()-10)),""))'

    # J: Total_Spent
    ws_port.cell(row, 10).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Spent],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Spent],ROW()-10)),""))'

    # K: Total_ULO
    ws_port.cell(row, 11).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_ULO],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_ULO],ROW()-10)),""))'

    # L: ULO_Percent
    ws_port.cell(row, 12).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[ULO_Percent],ROW()-10)=0,"",INDEX(T_Master_Projects[ULO_Percent],ROW()-10)),""))'

    # M: Days_Remaining
    ws_port.cell(row, 13).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Days_Remaining],ROW()-10)=0,"",INDEX(T_Master_Projects[Days_Remaining],ROW()-10)),""))'

print("  OK - All numeric columns updated (blank instead of 0)")

# Save as v24 (v23 is open in Excel)
print("\nSaving as v24...")
wb.save('2025-10-26-Tracker-v24.xlsx')

print("\n" + "="*80)
print("PORTFOLIO_DASHBOARD UPDATED!")
print("="*80)

print("\nSaved as: 2025-10-26-Tracker-v24.xlsx")

print("\nFormula logic:")
print("  Text columns (B-F): Blank if no project or no data")
print("  Numeric columns (G-M): Blank if no project, no data, OR value is 0")

print("\nResult:")
print("  - Empty budget fields show blank (not $0)")
print("  - Zero percentages show blank (not 0%)")
print("  - Missing days remaining shows blank")
print("  - Much cleaner appearance!")

print("\nv23 is ready to use!")
