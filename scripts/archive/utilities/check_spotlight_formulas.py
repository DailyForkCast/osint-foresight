"""
Check what formulas Spotlight_PMWorkspace needs
"""
import openpyxl

wb_v12 = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=False)

print("="*80)
print("SPOTLIGHT_PMWORKSPACE - V12 FORMULAS")
print("="*80)

ws = wb_v12['Spotlight_PMWorkspace']

print("\nAll formulas in Spotlight_PMWorkspace:")
print(f"{'Cell':<8} {'Formula':<70}")
print("-"*80)

formulas = []
for row in range(1, min(51, ws.max_row + 1)):
    for col in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row, col)
        if cell.data_type == 'f':
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            formula = str(cell.value)
            formulas.append((cell_ref, formula))

if formulas:
    for cell_ref, formula in formulas:
        # Truncate long formulas
        if len(formula) > 65:
            formula_display = formula[:62] + "..."
        else:
            formula_display = formula
        print(f"{cell_ref:<8} {formula_display:<70}")
else:
    print("NO FORMULAS FOUND")

print("\n" + "="*80)
print("SPOTLIGHT LAYOUT (First 25 rows, cols A-D)")
print("="*80)

ws_v12 = wb_v12['Spotlight_PMWorkspace']
print(f"\n{'Row':<5} {'A':<30} {'B':<30} {'C':<30}")
print("-"*95)

for row in range(1, 26):
    val_a = ws_v12.cell(row, 1).value or ""
    val_b = ws_v12.cell(row, 2).value or ""
    val_c = ws_v12.cell(row, 3).value or ""

    # Clean up values
    val_a = str(val_a)[:28] if val_a else ""
    val_b = str(val_b)[:28] if val_b else ""
    val_c = str(val_c)[:28] if val_c else ""

    print(f"{row:<5} {val_a:<30} {val_b:<30} {val_c:<30}")

print("\n" + "="*80)
print("CONTROL SHEET FORMULAS")
print("="*80)

ws_control = wb_v12['Control']
print("\nKey formulas in Control sheet:")
print(f"{'Cell':<8} {'Formula':<70}")
print("-"*80)

for row in range(1, 26):
    for col in range(1, 6):
        cell = ws_control.cell(row, col)
        if cell.data_type == 'f':
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            formula = str(cell.value)[:65]
            print(f"{cell_ref:<8} {formula:<70}")
