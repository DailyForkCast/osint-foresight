"""
Check Spotlight columns F and G rows 4-9 for duplicates
"""
import openpyxl

print("="*80)
print("CHECKING SPOTLIGHT DUPLICATES")
print("="*80)

# Load v29
print("\nLoading v29...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v29.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("ROWS 4-10, COLUMNS E-G")
print("="*80)

for row in range(4, 11):
    print(f"\nRow {row}:")
    for col in range(5, 8):  # E, F, G
        cell = ws_spotlight.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)[:80]
                print(f"  {col_letter}{row}: FORMULA - {formula}...")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
                print(f"  {col_letter}{row}: {clean_value}")
        else:
            print(f"  {col_letter}{row}: [empty]")

print("\n" + "="*80)
print("ANALYSIS")
print("="*80)

print("\nChecking F4 vs G5:")
f4 = ws_spotlight['F4'].value
g5 = ws_spotlight['G5'].value
print(f"  F4: {f4}")
print(f"  G5: {g5}")

print("\nChecking F9 vs G9:")
f9 = ws_spotlight['F9'].value
g9 = ws_spotlight['G9'].value
print(f"  F9: {f9}")
print(f"  G9: {g9}")

print("\nReady to fix!")
