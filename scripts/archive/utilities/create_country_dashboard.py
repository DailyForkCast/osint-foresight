import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.1.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING COUNTRY DASHBOARD')
print('='*80)
print()

wb = openpyxl.load_workbook(input_file, data_only=False)

# Create or get Country_Dashboard sheet
if 'Country_Dashboard' in wb.sheetnames:
    print('Country_Dashboard sheet exists - removing old version...')
    del wb['Country_Dashboard']

print('Creating new Country_Dashboard sheet...')
ws = wb.create_sheet('Country_Dashboard', 3)  # Insert after Portfolio_Dashboard
print('[OK] Sheet created')
print()

# ============================================================================
# DESIGN: COUNTRY DASHBOARD
# ============================================================================

print('Building Country Dashboard structure...')
print('-'*80)

# Styles
header_font = Font(bold=True, size=14, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
subheader_font = Font(bold=True, size=11, color='FFFFFF')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
metric_font = Font(bold=True, size=16)
label_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# HEADER SECTION
# ============================================================================

print('Step 1: Header section...')

# Row 1: Title
ws['A1'] = 'COUNTRY DASHBOARD'
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws.merge_cells('A1:H1')

# Row 2: Country selector
ws['A2'] = 'Selected Country:'
ws['A2'].font = label_font
ws['B2'] = 'DE'  # Default to Germany
ws.merge_cells('B2:C2')

# Add dropdown for country selection
ws_regions = wb['Country_Regions']
last_country_row = ws_regions.max_row
country_dv = DataValidation(
    type='list',
    formula1=f'Country_Regions!$A$2:$A${last_country_row}',
    allow_blank=False,
    showDropDown=True
)
country_dv.promptTitle = 'Select Country'
country_dv.prompt = 'Choose a country to view its dashboard'
country_dv.add('B2')
ws.add_data_validation(country_dv)

# Country Name display
ws['D2'] = '=IFERROR(INDEX(Country_Regions!$B:$B,MATCH(B2,Country_Regions!$A:$A,0)),"")'
ws['D2'].font = Font(bold=True, size=12)
ws.merge_cells('D2:F2')

print('[OK] Header with country selector')
print()

# ============================================================================
# SUMMARY METRICS SECTION (Row 4-7)
# ============================================================================

print('Step 2: Summary metrics...')

# Row 4: Metric labels
metrics_row = 4
ws[f'A{metrics_row}'] = 'TOTAL ALLOCATED'
ws[f'C{metrics_row}'] = 'TOTAL OBLIGATED'
ws[f'E{metrics_row}'] = 'TOTAL SPENT'
ws[f'G{metrics_row}'] = 'ULO'

for col in ['A', 'C', 'E', 'G']:
    cell = ws[f'{col}{metrics_row}']
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')

# Row 5: Metric values
values_row = 5

# Total Allocated for selected country
ws[f'A{values_row}'] = '=SUMIF(Country_Budgets!$D:$D,$B$2,Country_Budgets!$F:$F)'
ws[f'A{values_row}'].number_format = '$#,##0'
ws[f'A{values_row}'].font = metric_font
ws[f'A{values_row}'].alignment = Alignment(horizontal='center')

# Total Obligated for selected country
ws[f'C{values_row}'] = '=SUMIF(Country_Budgets!$D:$D,$B$2,Country_Budgets!$G:$G)'
ws[f'C{values_row}'].number_format = '$#,##0'
ws[f'C{values_row}'].font = metric_font
ws[f'C{values_row}'].alignment = Alignment(horizontal='center')

# Total Spent for selected country
ws[f'E{values_row}'] = '=SUMIF(Country_Budgets!$D:$D,$B$2,Country_Budgets!$H:$H)'
ws[f'E{values_row}'].number_format = '$#,##0'
ws[f'E{values_row}'].font = metric_font
ws[f'E{values_row}'].alignment = Alignment(horizontal='center')

# ULO (Obligated - Spent)
ws[f'G{values_row}'] = f'=C{values_row}-E{values_row}'
ws[f'G{values_row}'].number_format = '$#,##0'
ws[f'G{values_row}'].font = metric_font
ws[f'G{values_row}'].alignment = Alignment(horizontal='center')

# Row 6: Secondary metrics
ws[f'A6'] = 'Number of Projects:'
ws[f'B6'] = '=COUNTIF(Country_Budgets!$D:$D,$B$2)'
ws[f'B6'].font = label_font

ws[f'E6'] = 'ULO %:'
ws[f'F6'] = f'=IF(C{values_row}=0,0,G{values_row}/C{values_row})'
ws[f'F6'].number_format = '0.0%'
ws[f'F6'].font = label_font

print('[OK] Summary metrics added')
print()

# ============================================================================
# PROJECT MANAGER SECTION (Row 8-9)
# ============================================================================

print('Step 3: Project Manager info...')

ws['A8'] = 'Country PM:'
ws['A8'].font = label_font
ws['B8'] = '=IFERROR(INDEX(Country_PM_Assignments!$D:$D,MATCH($B$2,Country_PM_Assignments!$A:$A,0)),"TBD")'
ws.merge_cells('B8:C8')

ws['E8'] = 'Email:'
ws['E8'].font = label_font
ws['F8'] = '=IFERROR(INDEX(Country_PM_Assignments!$E:$E,MATCH($B$2,Country_PM_Assignments!$A:$A,0)),"")'
ws.merge_cells('F8:H8')

print('[OK] PM info added')
print()

# ============================================================================
# PROJECTS LIST SECTION (Starting Row 11)
# ============================================================================

print('Step 4: Projects list...')

list_start_row = 11

# Headers (row 10)
ws['A10'] = 'PROJECTS IN THIS COUNTRY'
ws['A10'].font = subheader_font
ws['A10'].fill = subheader_fill
ws.merge_cells('A10:H10')

# Column headers (row 11)
headers = [
    ('A', 'Project ID'),
    ('B', 'Project Name'),
    ('C', 'Status'),
    ('D', 'Priority'),
    ('E', 'Allocated'),
    ('F', 'Obligated'),
    ('G', 'Spent'),
    ('H', 'ULO')
]

for col, header_text in headers:
    cell = ws[f'{col}{list_start_row}']
    cell.value = header_text
    cell.font = label_font
    cell.border = thin_border

# Data rows (12-31) - 20 rows of data
data_start_row = 12

# We need to show projects that have budgets for the selected country
# This is tricky - we need to find unique project IDs that have budgets for the country
# Then display them

# Approach: Use helper formulas to extract unique project IDs for selected country
# Then use INDEX/MATCH to get project details

# For each row, we'll get the Nth unique project that has a budget for selected country
for row_offset in range(20):  # 20 rows of data
    row_num = data_start_row + row_offset
    row_index = row_offset + 1  # 1st, 2nd, 3rd... project

    # Project ID - Get Nth unique project ID that has selected country
    # This is complex - simplified approach: get Nth budget record's project ID
    # More sophisticated: use FILTER in Excel 365, but we're avoiding that due to corruption

    # Simplified approach: Show budget records directly (not unique projects)
    # Column A: Project ID from Nth budget record for this country
    ws[f'A{row_num}'] = f'=IFERROR(INDEX(Country_Budgets!$B:$B,SMALL(IF(Country_Budgets!$D:$D=$B$2,ROW(Country_Budgets!$D:$D)),{row_index})),"")'

    # Column B: Project Name (lookup from Master_Projects)
    ws[f'B{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Master_Projects!$C:$C,MATCH(A{row_num},Master_Projects!$A:$A,0)),""))'

    # Column C: Status
    ws[f'C{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Master_Projects!$D:$D,MATCH(A{row_num},Master_Projects!$A:$A,0)),""))'

    # Column D: Priority
    ws[f'D{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Master_Projects!$E:$E,MATCH(A{row_num},Master_Projects!$A:$A,0)),""))'

    # Column E: Allocated (from Country_Budgets for this specific budget record)
    ws[f'E{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Country_Budgets!$F:$F,SMALL(IF(Country_Budgets!$D:$D=$B$2,ROW(Country_Budgets!$F:$F)),{row_index})),0))'
    ws[f'E{row_num}'].number_format = '$#,##0'

    # Column F: Obligated
    ws[f'F{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Country_Budgets!$G:$G,SMALL(IF(Country_Budgets!$D:$D=$B$2,ROW(Country_Budgets!$G:$G)),{row_index})),0))'
    ws[f'F{row_num}'].number_format = '$#,##0'

    # Column G: Spent
    ws[f'G{row_num}'] = f'=IF(A{row_num}="","",IFERROR(INDEX(Country_Budgets!$H:$H,SMALL(IF(Country_Budgets!$D:$D=$B$2,ROW(Country_Budgets!$H:$H)),{row_index})),0))'
    ws[f'G{row_num}'].number_format = '$#,##0'

    # Column H: ULO
    ws[f'H{row_num}'] = f'=IF(A{row_num}="","",F{row_num}-G{row_num})'
    ws[f'H{row_num}'].number_format = '$#,##0'

print('[OK] Projects list with array formulas (requires Ctrl+Shift+Enter)')
print()

# ============================================================================
# COLUMN WIDTHS
# ============================================================================

print('Step 5: Formatting...')

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15

print('[OK] Column widths set')
print()

# ============================================================================
# SAVE
# ============================================================================

print('='*80)
print('Saving v10.1...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('COUNTRY DASHBOARD CREATED!')
print('='*80)
print()
print('Features:')
print('  1. [OK] Country selector dropdown (cell B2)')
print('  2. [OK] Summary metrics (Allocated, Obligated, Spent, ULO)')
print('  3. [OK] Project count and ULO %')
print('  4. [OK] Country PM info from Country_PM_Assignments')
print('  5. [OK] List of all budget records for selected country')
print('  6. [OK] Auto-updates when you change country')
print()
print('How to use:')
print('  1. Open v10.1 in Excel')
print('  2. Go to Country_Dashboard sheet')
print('  3. Click cell B2, select a country from dropdown')
print('  4. All metrics and project list update automatically')
print()
print('IMPORTANT - Array Formulas:')
print('  - Project list uses array formulas (SMALL/IF pattern)')
print('  - In Excel 2016/2019: Select column A12:A31, press F2, Ctrl+Shift+Enter')
print('  - Repeat for columns E12:E31, F12:F31, G12:G31')
print('  - In Excel 365: Should work automatically')
print()
print(f'Output: {output_file}')
print()
print(f'Created: {timestamp}')
