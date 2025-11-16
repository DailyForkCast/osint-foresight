import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load the v1 file
input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v1.xlsx'

print('='*80)
print('PRIORITY 5: Creating Country_PM_Assignments Sheet')
print('='*80)
print()

# Load workbook
print('Loading workbook...')
wb = openpyxl.load_workbook(input_file, data_only=False)

# Check if sheet already exists
if 'Country_PM_Assignments' in wb.sheetnames:
    print('WARNING: Country_PM_Assignments sheet already exists!')
    print('Removing old sheet...')
    wb.remove(wb['Country_PM_Assignments'])

# Create new sheet
print('Creating Country_PM_Assignments sheet...')
ws = wb.create_sheet('Country_PM_Assignments')

# Define headers
headers = [
    'Country_Code',
    'Country_Name',
    'Region',
    'Project_Manager',
    'PM_Email',
    'PM_Phone',
    'Notes'
]

# Add headers with formatting
print('Adding headers...')
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)

    # Format header: Bold, white text, dark blue background
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Get country data from Country_Regions sheet
print('Reading country data from Country_Regions...')
if 'Country_Regions' not in wb.sheetnames:
    print('ERROR: Country_Regions sheet not found!')
    wb.close()
    exit(1)

country_ws = wb['Country_Regions']

# Extract all countries
countries = []
for row in range(2, country_ws.max_row + 1):
    country_code = country_ws.cell(row, 1).value
    country_name = country_ws.cell(row, 2).value
    region = country_ws.cell(row, 3).value

    if country_code and country_name:
        countries.append({
            'code': country_code,
            'name': country_name,
            'region': region
        })

print(f'Found {len(countries)} countries')
print()

# Add all countries to the new sheet
print('Adding countries to Country_PM_Assignments...')
for idx, country in enumerate(countries, start=2):
    ws.cell(row=idx, column=1, value=country['code'])
    ws.cell(row=idx, column=2, value=country['name'])
    ws.cell(row=idx, column=3, value=country['region'])
    ws.cell(row=idx, column=4, value='TBD')  # Project_Manager - to be filled in
    ws.cell(row=idx, column=5, value='')     # PM_Email - empty
    ws.cell(row=idx, column=6, value='')     # PM_Phone - empty
    ws.cell(row=idx, column=7, value='')     # Notes - empty

print(f'Added {len(countries)} countries')
print()

# Set column widths for better readability
print('Setting column widths...')
column_widths = {
    'A': 15,  # Country_Code
    'B': 30,  # Country_Name
    'C': 10,  # Region
    'D': 25,  # Project_Manager
    'E': 30,  # PM_Email
    'F': 20,  # PM_Phone
    'G': 40   # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze top row
print('Freezing header row...')
ws.freeze_panes = 'A2'

# Summary by region
print()
print('Regional breakdown:')
regions_count = {}
for country in countries:
    region = country['region']
    regions_count[region] = regions_count.get(region, 0) + 1

for region, count in sorted(regions_count.items()):
    print(f'  {region}: {count} countries')

print()
print('Sheet structure:')
print('  Column A: Country_Code (e.g., DE, FR, US)')
print('  Column B: Country_Name (e.g., Germany, France, United States)')
print('  Column C: Region (e.g., EUR, WHA, EAP)')
print('  Column D: Project_Manager (currently "TBD" - ready for your PM names)')
print('  Column E: PM_Email (empty - ready for email addresses)')
print('  Column F: PM_Phone (empty - ready for phone numbers)')
print('  Column G: Notes (empty - for any additional notes)')
print()

# Save
print('Saving workbook...')
wb.save(input_file)
wb.close()

print('='*80)
print('COMPLETE: Country_PM_Assignments sheet created!')
print('='*80)
print()
print(f'File updated: {input_file}')
print()
print('NEXT STEPS:')
print('1. Open the Excel file')
print('2. Go to the Country_PM_Assignments sheet')
print('3. Replace "TBD" in column D with actual PM names for each country')
print('4. Optionally add PM emails and phone numbers')
print()
print('TIP: You can use Find & Replace to quickly assign PMs by region:')
print('  - Filter by EUR in column C, then fill in EUR PM names')
print('  - Filter by WHA in column C, then fill in WHA PM names')
print('  - Etc.')
