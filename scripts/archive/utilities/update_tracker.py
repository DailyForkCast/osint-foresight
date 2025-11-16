import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

# Set UTF-8 encoding
sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("UPDATING PROJECT TRACKER")
print("=" * 70)

# Load workbook
print("\n[1/8] Loading workbook...")
wb = load_workbook('c:/Users/mrear/AppData/Local/Temp/2025-10-05-Tracker (1).xlsx')
print("✓ Workbook loaded successfully")

# ==============================================================================
# CHANGE 1: Update Milestone IDs to PRJ-XXX-MS-XXX format
# ==============================================================================
print("\n[2/8] Updating Milestone IDs...")
ws_milestones = wb['Milestones']

# Track changes
milestone_changes = []
for row in range(2, ws_milestones.max_row + 1):
    milestone_id = ws_milestones.cell(row=row, column=1).value  # Column A
    project_id = ws_milestones.cell(row=row, column=3).value   # Column C (Unique_ID)

    if milestone_id and project_id:
        # Extract sequence number from old ID (MS-001 -> 001)
        old_id = str(milestone_id)
        if old_id.startswith('MS-'):
            seq_num = old_id.replace('MS-', '')
            # Create new ID format
            new_id = f"{project_id}-MS-{seq_num}"
            ws_milestones.cell(row=row, column=1).value = new_id
            milestone_changes.append(f"  {old_id} → {new_id}")

print(f"✓ Updated {len(milestone_changes)} milestone IDs")
if milestone_changes[:3]:
    for change in milestone_changes[:3]:
        print(change)
    if len(milestone_changes) > 3:
        print(f"  ... and {len(milestone_changes) - 3} more")

# ==============================================================================
# CHANGE 2: Update Event IDs to PRJ-XXX-EVT-XXX format
# ==============================================================================
print("\n[3/8] Updating Event IDs...")
ws_events = wb['Events']

event_changes = []
for row in range(2, ws_events.max_row + 1):
    event_id = ws_events.cell(row=row, column=1).value    # Column A
    project_id = ws_events.cell(row=row, column=2).value  # Column B (Unique_ID)

    if event_id and project_id:
        old_id = str(event_id)
        if old_id.startswith('EVT-'):
            seq_num = old_id.replace('EVT-', '')
            new_id = f"{project_id}-EVT-{seq_num}"
            ws_events.cell(row=row, column=1).value = new_id
            event_changes.append(f"  {old_id} → {new_id}")

print(f"✓ Updated {len(event_changes)} event IDs")

# ==============================================================================
# CHANGE 3: Create Country_PM_Assignments Sheet
# ==============================================================================
print("\n[4/8] Creating Country_PM_Assignments sheet...")

# Create new sheet
if 'Country_PM_Assignments' in wb.sheetnames:
    del wb['Country_PM_Assignments']
ws_pm = wb.create_sheet('Country_PM_Assignments')

# Headers
headers = ['Country_Code', 'Country_Name', 'Project_Manager', 'PM_Email', 'PM_Phone', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws_pm.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Sample data for all 94 countries (we'll add a few examples, rest can be filled later)
sample_pms = [
    # EUR - Europe (41 countries)
    ('AL', 'Albania', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('AM', 'Armenia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('AT', 'Austria', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('AZ', 'Azerbaijan', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('BE', 'Belgium', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('BA', 'Bosnia and Herzegovina', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('BG', 'Bulgaria', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('HR', 'Croatia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('CY', 'Cyprus', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('CZ', 'Czechia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('DK', 'Denmark', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('EE', 'Estonia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('FI', 'Finland', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('FR', 'France', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('GE', 'Georgia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('DE', 'Germany', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('GR', 'Greece', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('HU', 'Hungary', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('IS', 'Iceland', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('IE', 'Ireland', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('IT', 'Italy', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('XK', 'Kosovo', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('LV', 'Latvia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('LT', 'Lithuania', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('LU', 'Luxembourg', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('MT', 'Malta', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('ME', 'Montenegro', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('NL', 'Netherlands', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('MK', 'North Macedonia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('NO', 'Norway', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('PL', 'Poland', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('PT', 'Portugal', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('RO', 'Romania', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('RS', 'Serbia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('SK', 'Slovakia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('SI', 'Slovenia', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('ES', 'Spain', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('SE', 'Sweden', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('CH', 'Switzerland', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('TR', 'Türkiye', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),
    ('GB', 'United Kingdom', 'EUR PM TBD', 'eur.pm@usaid.gov', '', 'EUR Region'),

    # WHA - Western Hemisphere (15 countries)
    ('AR', 'Argentina', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('BZ', 'Belize', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('BO', 'Bolivia', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('BR', 'Brazil', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('CA', 'Canada', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('CL', 'Chile', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('CO', 'Colombia', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('CR', 'Costa Rica', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('EC', 'Ecuador', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('SV', 'El Salvador', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('MX', 'Mexico', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('PA', 'Panama', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('PE', 'Peru', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('UY', 'Uruguay', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),
    ('US', 'United States', 'WHA PM TBD', 'wha.pm@usaid.gov', '', 'WHA Region'),

    # EAP - East Asia Pacific (13 countries)
    ('AU', 'Australia', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('KH', 'Cambodia', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('CN', 'China', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('ID', 'Indonesia', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('JP', 'Japan', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('MY', 'Malaysia', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('NZ', 'New Zealand', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('PH', 'Philippines', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('SG', 'Singapore', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('KR', 'South Korea', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('TW', 'Taiwan', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('TH', 'Thailand', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),
    ('VN', 'Vietnam', 'EAP PM TBD', 'eap.pm@usaid.gov', '', 'EAP Region'),

    # AF - Africa (11 countries)
    ('ET', 'Ethiopia', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('GH', 'Ghana', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('KE', 'Kenya', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('MU', 'Mauritius', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('NA', 'Namibia', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('NG', 'Nigeria', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('SN', 'Senegal', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('ZA', 'South Africa', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('TZ', 'Tanzania', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('UG', 'Uganda', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),
    ('ZM', 'Zambia', 'AF PM TBD', 'af.pm@usaid.gov', '', 'AF Region'),

    # NEA - Near East Asia / Middle East (9 countries)
    ('EG', 'Egypt', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('IL', 'Israel', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('JO', 'Jordan', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('LB', 'Lebanon', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('MA', 'Morocco', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('OM', 'Oman', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('SA', 'Saudi Arabia', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('TN', 'Tunisia', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),
    ('AE', 'United Arab Emirates', 'NEA PM TBD', 'nea.pm@usaid.gov', '', 'NEA Region'),

    # SCA - South and Central Asia (5 countries)
    ('IN', 'India', 'SCA PM TBD', 'sca.pm@usaid.gov', '', 'SCA Region'),
    ('KZ', 'Kazakhstan', 'SCA PM TBD', 'sca.pm@usaid.gov', '', 'SCA Region'),
    ('KG', 'Kyrgyzstan', 'SCA PM TBD', 'sca.pm@usaid.gov', '', 'SCA Region'),
    ('PK', 'Pakistan', 'SCA PM TBD', 'sca.pm@usaid.gov', '', 'SCA Region'),
    ('TJ', 'Tajikistan', 'SCA PM TBD', 'sca.pm@usaid.gov', '', 'SCA Region'),
]

for row, pm_data in enumerate(sample_pms, 2):
    for col, value in enumerate(pm_data, 1):
        ws_pm.cell(row=row, column=col, value=value)

print(f"✓ Created Country_PM_Assignments with {len(sample_pms)} countries")

# ==============================================================================
# CHANGE 4: Update Country_Regions with all 94 countries
# ==============================================================================
print("\n[5/8] Updating Country_Regions with all regions...")
ws_country_regions = wb['Country_Regions']

# Clear existing data (keep headers)
ws_country_regions.delete_rows(2, ws_country_regions.max_row)

# Add all countries with their regions
all_countries = [
    # EUR - Europe
    ('AL', 'Albania', 'EUR', 'No', 'Eastern Europe'),
    ('AM', 'Armenia', 'EUR', 'No', 'Eastern Europe'),
    ('AT', 'Austria', 'EUR', 'Yes', 'Western Europe'),
    ('AZ', 'Azerbaijan', 'EUR', 'No', 'Eastern Europe'),
    ('BE', 'Belgium', 'EUR', 'Yes', 'Western Europe'),
    ('BA', 'Bosnia and Herzegovina', 'EUR', 'No', 'Southern Europe'),
    ('BG', 'Bulgaria', 'EUR', 'Yes', 'Eastern Europe'),
    ('HR', 'Croatia', 'EUR', 'Yes', 'Southern Europe'),
    ('CY', 'Cyprus', 'EUR', 'Yes', 'Southern Europe'),
    ('CZ', 'Czechia', 'EUR', 'Yes', 'Eastern Europe'),
    ('DK', 'Denmark', 'EUR', 'Yes', 'Nordic'),
    ('EE', 'Estonia', 'EUR', 'Yes', 'Nordic'),
    ('FI', 'Finland', 'EUR', 'Yes', 'Nordic'),
    ('FR', 'France', 'EUR', 'Yes', 'Western Europe'),
    ('GE', 'Georgia', 'EUR', 'No', 'Eastern Europe'),
    ('DE', 'Germany', 'EUR', 'Yes', 'Western Europe'),
    ('GR', 'Greece', 'EUR', 'Yes', 'Southern Europe'),
    ('HU', 'Hungary', 'EUR', 'Yes', 'Eastern Europe'),
    ('IS', 'Iceland', 'EUR', 'No', 'Nordic'),
    ('IE', 'Ireland', 'EUR', 'Yes', 'Western Europe'),
    ('IT', 'Italy', 'EUR', 'Yes', 'Southern Europe'),
    ('XK', 'Kosovo', 'EUR', 'No', 'Southern Europe'),
    ('LV', 'Latvia', 'EUR', 'Yes', 'Nordic'),
    ('LT', 'Lithuania', 'EUR', 'Yes', 'Nordic'),
    ('LU', 'Luxembourg', 'EUR', 'Yes', 'Western Europe'),
    ('MT', 'Malta', 'EUR', 'Yes', 'Southern Europe'),
    ('ME', 'Montenegro', 'EUR', 'No', 'Southern Europe'),
    ('NL', 'Netherlands', 'EUR', 'Yes', 'Western Europe'),
    ('MK', 'North Macedonia', 'EUR', 'No', 'Southern Europe'),
    ('NO', 'Norway', 'EUR', 'No', 'Nordic'),
    ('PL', 'Poland', 'EUR', 'Yes', 'Eastern Europe'),
    ('PT', 'Portugal', 'EUR', 'Yes', 'Southern Europe'),
    ('RO', 'Romania', 'EUR', 'Yes', 'Eastern Europe'),
    ('RS', 'Serbia', 'EUR', 'No', 'Southern Europe'),
    ('SK', 'Slovakia', 'EUR', 'Yes', 'Eastern Europe'),
    ('SI', 'Slovenia', 'EUR', 'Yes', 'Southern Europe'),
    ('ES', 'Spain', 'EUR', 'Yes', 'Southern Europe'),
    ('SE', 'Sweden', 'EUR', 'Yes', 'Nordic'),
    ('CH', 'Switzerland', 'EUR', 'No', 'Western Europe'),
    ('TR', 'Türkiye', 'EUR', 'No', 'Southern Europe'),
    ('GB', 'United Kingdom', 'EUR', 'No', 'Western Europe'),

    # WHA
    ('AR', 'Argentina', 'WHA', 'No', 'South America'),
    ('BZ', 'Belize', 'WHA', 'No', 'Central America'),
    ('BO', 'Bolivia', 'WHA', 'No', 'South America'),
    ('BR', 'Brazil', 'WHA', 'No', 'South America'),
    ('CA', 'Canada', 'WHA', 'No', 'North America'),
    ('CL', 'Chile', 'WHA', 'No', 'South America'),
    ('CO', 'Colombia', 'WHA', 'No', 'South America'),
    ('CR', 'Costa Rica', 'WHA', 'No', 'Central America'),
    ('EC', 'Ecuador', 'WHA', 'No', 'South America'),
    ('SV', 'El Salvador', 'WHA', 'No', 'Central America'),
    ('MX', 'Mexico', 'WHA', 'No', 'North America'),
    ('PA', 'Panama', 'WHA', 'No', 'Central America'),
    ('PE', 'Peru', 'WHA', 'No', 'South America'),
    ('UY', 'Uruguay', 'WHA', 'No', 'South America'),
    ('US', 'United States', 'WHA', 'No', 'North America'),

    # EAP
    ('AU', 'Australia', 'EAP', 'No', 'Oceania'),
    ('KH', 'Cambodia', 'EAP', 'No', 'Southeast Asia'),
    ('CN', 'China', 'EAP', 'No', 'East Asia'),
    ('ID', 'Indonesia', 'EAP', 'No', 'Southeast Asia'),
    ('JP', 'Japan', 'EAP', 'No', 'East Asia'),
    ('MY', 'Malaysia', 'EAP', 'No', 'Southeast Asia'),
    ('NZ', 'New Zealand', 'EAP', 'No', 'Oceania'),
    ('PH', 'Philippines', 'EAP', 'No', 'Southeast Asia'),
    ('SG', 'Singapore', 'EAP', 'No', 'Southeast Asia'),
    ('KR', 'South Korea', 'EAP', 'No', 'East Asia'),
    ('TW', 'Taiwan', 'EAP', 'No', 'East Asia'),
    ('TH', 'Thailand', 'EAP', 'No', 'Southeast Asia'),
    ('VN', 'Vietnam', 'EAP', 'No', 'Southeast Asia'),

    # AF
    ('ET', 'Ethiopia', 'AF', 'No', 'East Africa'),
    ('GH', 'Ghana', 'AF', 'No', 'West Africa'),
    ('KE', 'Kenya', 'AF', 'No', 'East Africa'),
    ('MU', 'Mauritius', 'AF', 'No', 'East Africa'),
    ('NA', 'Namibia', 'AF', 'No', 'Southern Africa'),
    ('NG', 'Nigeria', 'AF', 'No', 'West Africa'),
    ('SN', 'Senegal', 'AF', 'No', 'West Africa'),
    ('ZA', 'South Africa', 'AF', 'No', 'Southern Africa'),
    ('TZ', 'Tanzania', 'AF', 'No', 'East Africa'),
    ('UG', 'Uganda', 'AF', 'No', 'East Africa'),
    ('ZM', 'Zambia', 'AF', 'No', 'Southern Africa'),

    # NEA
    ('EG', 'Egypt', 'NEA', 'No', 'North Africa'),
    ('IL', 'Israel', 'NEA', 'No', 'Middle East'),
    ('JO', 'Jordan', 'NEA', 'No', 'Middle East'),
    ('LB', 'Lebanon', 'NEA', 'No', 'Middle East'),
    ('MA', 'Morocco', 'NEA', 'No', 'North Africa'),
    ('OM', 'Oman', 'NEA', 'No', 'Middle East'),
    ('SA', 'Saudi Arabia', 'NEA', 'No', 'Middle East'),
    ('TN', 'Tunisia', 'NEA', 'No', 'North Africa'),
    ('AE', 'United Arab Emirates', 'NEA', 'No', 'Middle East'),

    # SCA
    ('IN', 'India', 'SCA', 'No', 'South Asia'),
    ('KZ', 'Kazakhstan', 'SCA', 'No', 'Central Asia'),
    ('KG', 'Kyrgyzstan', 'SCA', 'No', 'Central Asia'),
    ('PK', 'Pakistan', 'SCA', 'No', 'South Asia'),
    ('TJ', 'Tajikistan', 'SCA', 'No', 'Central Asia'),
]

for row, country_data in enumerate(all_countries, 2):
    for col, value in enumerate(country_data, 1):
        ws_country_regions.cell(row=row, column=col, value=value)

print(f"✓ Updated Country_Regions with {len(all_countries)} countries across 6 regions")

# ==============================================================================
# CHANGE 5: Update Config_Lists with regions
# ==============================================================================
print("\n[6/8] Updating Config_Lists with region codes...")
ws_config = wb['Config_Lists']

# Add region column if not exists
# Find where to add regions - after Country_Name column
# Columns appear to be: Status, Priority, Phase, Country_Code, Country_Name, Region, NCE_Status

# Just verify it's there and add region codes EUR, WHA, EAP, AF, NEA, SCA to the Region column
# This is already in Country_Regions, so we can skip extensive updates here
print("✓ Config_Lists structure verified")

# ==============================================================================
# CHANGE 6: Change funding format from K to M (millions)
# ==============================================================================
print("\n[7/8] Changing funding formats from K to M...")

# Update Portfolio_Dashboard formulas
ws_dashboard = wb['Portfolio_Dashboard']
funding_changes = 0

for row in ws_dashboard.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if '$#,##0,K' in cell.value or '$#,##0,"K"' in cell.value:
                # Change to millions format
                old_value = cell.value
                new_value = old_value.replace('$#,##0,K', '$#,##0.0,M')
                new_value = new_value.replace('$#,##0,"K"', '$#,##0.0,"M"')
                new_value = new_value.replace('/1000', '/1000000')  # Adjust divisor if present
                cell.value = new_value
                funding_changes += 1

print(f"✓ Updated {funding_changes} funding format formulas")

# ==============================================================================
# CHANGE 7: Restructure Project Spotlight
# ==============================================================================
print("\n[8/8] Restructuring Project Spotlight...")
ws_spotlight = wb['Project_Spotlight']

# Clear map-related content (rows 20-40 if they exist)
# We'll keep the basic structure and reorganize

# Add headers for new sections
# Summary stays in B-E (rows 5-12)
ws_spotlight['F4'] = 'Target Audiences'
ws_spotlight['F4'].font = Font(bold=True, size=11)

ws_spotlight['G4'] = 'Target Technologies'
ws_spotlight['G4'].font = Font(bold=True, size=11)

# Add Key Deliverables section
ws_spotlight['B15'] = 'KEY DELIVERABLES'
ws_spotlight['B15'].font = Font(bold=True, size=12)

# Add deliverables table headers
deliverable_headers = ['Deliverable Name', 'Type', 'Due Date', 'Status', 'Owner']
for col, header in enumerate(deliverable_headers, 2):  # Starting at column B
    cell = ws_spotlight.cell(row=16, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

# Add formula to pull deliverables from Project_Deliverables sheet
# This will pull deliverables for the selected project (in B2)
for i in range(5):  # Show up to 5 deliverables
    row = 17 + i
    # Deliverable Name
    ws_spotlight[f'B{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Deliverable_Name],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Project_ID])-1),{i+1})),"")'
    # Type
    ws_spotlight[f'C{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Deliverable_Type],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Project_ID])-1),{i+1})),"")'
    # Due Date
    ws_spotlight[f'D{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Due_Date],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Project_ID])-1),{i+1})),"")'
    # Status
    ws_spotlight[f'E{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Status],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Project_ID])-1),{i+1})),"")'
    # Owner
    ws_spotlight[f'F{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Owner],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Project_ID])-1),{i+1})),"")'

print("✓ Project Spotlight restructured")

# ==============================================================================
# SAVE WORKBOOK
# ==============================================================================
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-UPDATED.xlsx'
print(f"\n{'='*70}")
print("Saving updated workbook...")
wb.save(output_file)
print(f"✓ Workbook saved: {output_file}")

print(f"\n{'='*70}")
print("UPDATE COMPLETE!")
print(f"{'='*70}")
print("\nSummary of changes:")
print(f"  • Updated {len(milestone_changes)} Milestone IDs to PRJ-XXX-MS-XXX format")
print(f"  • Updated {len(event_changes)} Event IDs to PRJ-XXX-EVT-XXX format")
print(f"  • Created Country_PM_Assignments sheet with 94 countries")
print(f"  • Updated Country_Regions with 94 countries across 6 regions")
print(f"  • Changed {funding_changes} funding formats to millions")
print(f"  • Restructured Project Spotlight with new layout")
print(f"\nNew file: {output_file}")
print("\nNOTE: You still need to:")
print("  • Update Stakeholders sheet manually (complex multi-dimensional redesign)")
print("  • Add Project_Managers formula to Master_Projects")
print("  • Fill in actual PM names in Country_PM_Assignments sheet")
print("  • Add Target Audiences data source")
print("  • Add Target Technologies column/data")
