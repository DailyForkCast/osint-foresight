"""
Add POP/NCE fields to Spotlight_PMWorkspace rows 2-3
"""
import openpyxl

print("="*80)
print("ADDING POP/NCE FIELDS TO SPOTLIGHT")
print("="*80)

# Load v19
print("\nLoading v19...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v19.xlsx')

ws = wb['Spotlight_PMWorkspace']

print("\nAdding Row 2 formulas (Award Number, NCE Status)...")

# Row 2 - Labels and Formulas
ws['E2'] = 'Award Number:'
ws['F2'] = '=IFERROR(INDEX(T_Master_Projects[Award_Number],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['H2'] = 'NCE Status:'
ws['I2'] = '=IFERROR(INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

print("  OK - Award Number (E2:F2)")
print("  OK - NCE Status (H2:I2)")

print("\nAdding Row 3 formulas (POP_Start, POP_End, POP_Days_Remaining)...")

# Row 3 - Labels and Formulas
ws['E3'] = 'POP Start:'
ws['F3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Start],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['G3'] = 'POP End:'
ws['H3'] = '=IFERROR(INDEX(T_Master_Projects[POP_End],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['I3'] = 'POP Days Remaining:'
ws['J3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Days_Remaining],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

print("  OK - POP Start (E3:F3)")
print("  OK - POP End (G3:H3)")
print("  OK - POP Days Remaining (I3:J3)")

# Save
print("\nSaving v19...")
wb.save('2025-10-26-Tracker-v19.xlsx')

print("\n" + "="*80)
print("POP/NCE FIELDS ADDED TO SPOTLIGHT!")
print("="*80)

print("\nRow 2 layout:")
print("  A2-B2: Project ID selector")
print("  E2: Award Number: | F2: [formula]")
print("  H2: NCE Status: | I2: [formula]")

print("\nRow 3 layout:")
print("  E3: POP Start: | F3: [formula]")
print("  G3: POP End: | H3: [formula]")
print("  I3: POP Days Remaining: | J3: [formula]")

print("\nV19 is ready to use!")
print("\nNext steps:")
print("  1. Open v19 in Excel")
print("  2. Add data validation for NCE_Eligible and NCE_Status in Master_Projects")
print("  3. Test the formulas by selecting different projects")
