"""
Comprehensive check of v12 features to find what's missing in v18
"""
import openpyxl

print("="*80)
print("COMPREHENSIVE V12 FEATURE CHECK")
print("="*80)

wb_v12 = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=False)

# Check all sheets for formulas
print("\n1. ALL FORMULAS IN V12 (BY SHEET)")
print("="*80)

for sheet_name in wb_v12.sheetnames:
    ws = wb_v12[sheet_name]
    formulas_found = []

    # Check first 50 rows for formulas
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.data_type == 'f':
                col_letter = openpyxl.utils.get_column_letter(col)
                header = ws.cell(1, col).value if row > 1 else "HEADER"
                formula = str(cell.value)[:60]
                formulas_found.append(f"  {col_letter}{row}: {formula}")

    if formulas_found:
        print(f"\n{sheet_name}:")
        # Show first 10 formulas per sheet
        for f in formulas_found[:10]:
            print(f)
        if len(formulas_found) > 10:
            print(f"  ... and {len(formulas_found) - 10} more formulas")

# Check for data validation
print("\n\n2. DATA VALIDATION IN V12")
print("="*80)

for sheet_name in wb_v12.sheetnames:
    ws = wb_v12[sheet_name]
    if ws.data_validations.dataValidation:
        print(f"\n{sheet_name}:")
        for dv in ws.data_validations.dataValidation:
            print(f"  Range: {dv.sqref}")
            print(f"  Type: {dv.type}")
            if hasattr(dv, 'formula1') and dv.formula1:
                formula = str(dv.formula1)[:70]
                print(f"  Source: {formula}")

# Check Config_Lists for list types
print("\n\n3. CONFIG_LISTS CONTENT")
print("="*80)

ws = wb_v12['Config_Lists']
list_types = {}

for row in range(2, min(100, ws.max_row + 1)):
    list_type = ws.cell(row, 1).value
    if list_type:
        if list_type not in list_types:
            list_types[list_type] = []
        # Get all values in this row
        values = []
        for col in range(2, 8):
            val = ws.cell(row, col).value
            if val:
                values.append(str(val))
        if values:
            list_types[list_type].extend(values)

print("\nList types found:")
for lt, values in sorted(list_types.items()):
    print(f"\n  {lt}:")
    # Show first 5 values
    for v in values[:5]:
        print(f"    - {v}")
    if len(values) > 5:
        print(f"    ... and {len(values) - 5} more values")

# Check for named ranges
print("\n\n4. NAMED RANGES IN V12")
print("="*80)

if wb_v12.defined_names.definedName:
    print("\nNamed ranges found:")
    for name in wb_v12.defined_names.definedName:
        print(f"  - {name.name}")
        if hasattr(name, 'value') and name.value:
            val_str = str(name.value)[:60]
            print(f"    = {val_str}")
else:
    print("\nNo named ranges found")

# Check Master_Projects ALL columns
print("\n\n5. MASTER_PROJECTS - ALL COLUMNS WITH FORMULAS")
print("="*80)

ws = wb_v12['Master_Projects']
print("\nRow 2 formulas:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.data_type == 'f':
        formula = str(cell.value)[:70]
        print(f"\n  {col_letter} ({header}):")
        print(f"    {formula}")

# Check Country_Budgets ALL columns
print("\n\n6. COUNTRY_BUDGETS - ALL COLUMNS WITH FORMULAS")
print("="*80)

ws = wb_v12['Country_Budgets']
print("\nRow 2 formulas:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.data_type == 'f':
        formula = str(cell.value)[:70]
        print(f"\n  {col_letter} ({header}):")
        print(f"    {formula}")

# Check Spotlight_PMWorkspace for formulas
print("\n\n7. SPOTLIGHT_PMWORKSPACE FORMULAS")
print("="*80)

ws = wb_v12['Spotlight_PMWorkspace']
print("\nFormulas found:")
formula_count = 0
for row in range(1, min(51, ws.max_row + 1)):
    for col in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row, col)
        if cell.data_type == 'f':
            col_letter = openpyxl.utils.get_column_letter(col)
            formula = str(cell.value)[:60]
            print(f"  {col_letter}{row}: {formula}")
            formula_count += 1
            if formula_count >= 20:
                break
    if formula_count >= 20:
        break

if formula_count == 0:
    print("  No formulas found in first 50 rows")

print("\n" + "="*80)
print("SCAN COMPLETE")
print("="*80)
