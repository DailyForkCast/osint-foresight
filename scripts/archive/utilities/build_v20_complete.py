"""
Build v20 - Complete tracker with all formulas in all dashboards
Starting from v18, adding all missing formulas carefully
"""
import openpyxl
from datetime import datetime

print("="*80)
print("BUILDING V20 - COMPLETE TRACKER")
print("="*80)

# Load v18 as base (it works!)
print("\nLoading v18 as base...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

print("\n" + "="*80)
print("STEP 1: UPDATE CONTROL SHEET FORMULAS")
print("="*80)

ws_control = wb['Control']

# Update Control sheet formulas to match new column structure
control_formulas = {
    'B3': '=TODAY()',
    'B4': '=NOW()',
    'B6': '=TEXT(B3,"MMM DD, YYYY")',
    'B9': '=COUNTA(T_Master_Projects[Project_Unique_ID])',
    'B10': '=COUNTIF(T_Master_Projects[Project_Status],"Started*")',
    'B11': '=COUNTIF(T_Master_Projects[Project_Status],"On Hold")',
    'B12': '=COUNTIF(T_Master_Projects[Project_Status],"Completed")',
    'B15': '=SUM(T_Country_Budgets[Allocated_Amount])',
    'B16': '=SUM(T_Country_Budgets[Obligated_Amount])',
    'B17': '=B15-B16',
    'B18': '=IF(B15=0,0,B17/B15)',
    'B19': '=SUMIF(T_Master_Projects[ULO_Percent],">0.5",T_Master_Projects[Total_ULO])',
    'B22': '=COUNTIF(T_Master_Projects[ULO_Percent],">0.75")',
    'B23': '=COUNTIFS(T_Master_Projects[ULO_Percent],">0.5",T_Master_Projects[ULO_Percent],"<=0.75")',
    'B24': '=COUNTIF(T_Master_Projects[ULO_Percent],"<=0.5")',
    'B25': '=COUNTIFS(T_Master_Projects[Days_Remaining],"<90",T_Master_Projects[Days_Remaining],">0")',
    'E4': '=AVERAGE(T_Master_Projects[ULO_Percent])',
    'E5': '=B9/COUNTA(T_Country_Regions[Country_Code])',
    'E6': '=B15/B9',
    'E7': '=B16/B15',
    'E8': '=COUNTA(T_Country_Budgets[Unique_ID])',
    'E9': '=COUNTA(Milestones!A:A)-1',
    'E10': '=COUNTIF(Milestones!I:I,"<0")',
    'E11': '=COUNTIFS(Milestones!I:I,">=0",Milestones!I:I,"<=30")',
    'E12': '=IF(B9=0,0,B12/B9)',
}

print("Adding Control sheet formulas...")
for cell_ref, formula in control_formulas.items():
    ws_control[cell_ref] = formula
    print(f"  {cell_ref}")

print("\n" + "="*80)
print("STEP 2: BUILD PORTFOLIO_DASHBOARD")
print("="*80)

ws_port = wb['Portfolio_Dashboard']

# Header formulas
ws_port['L1'] = '=Control!B6'
ws_port['A4'] = '=Control!B10'
ws_port['D4'] = '=TEXT(Control!B15,"$#,##0,K")'
ws_port['G4'] = '=TEXT(Control!B19,"$#,##0,K")'
ws_port['J4'] = '=Control!B25'

# Summary text
ws_port['A6'] = '="of "&Control!B9&" total"'
ws_port['D6'] = '=TEXT(Control!B18,"0%")&" Unobligated"'
ws_port['G6'] = '=Control!B22&" Critical Projects"'

print("Adding Portfolio Dashboard summary formulas...")

# Project list (rows 11-20, auto-increment with ROW())
print("Adding project list formulas (rows 11-20)...")
for row in range(11, 21):
    # A: Project ID
    ws_port.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")'
    # B: Project Name
    ws_port.cell(row, 2).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Name],ROW()-10),"")'
    # C: Status
    ws_port.cell(row, 3).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Status],ROW()-10),"")'
    # D: Priority
    ws_port.cell(row, 4).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Priority],ROW()-10),"")'
    # E: Progress
    ws_port.cell(row, 5).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Progress],ROW()-10),"")'
    # F: Countries
    ws_port.cell(row, 6).value = f'=IFERROR(INDEX(T_Master_Projects[Countries],ROW()-10),"")'
    # G: Total Allocation
    ws_port.cell(row, 7).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Allocation],ROW()-10),"")'
    # H: Total Obligated
    ws_port.cell(row, 8).value = f'=IFERROR(INDEX(T_Master_Projects[Total_Obligated],ROW()-10),"")'
    # I: Total ULO
    ws_port.cell(row, 9).value = f'=IFERROR(INDEX(T_Master_Projects[Total_ULO],ROW()-10),"")'
    # J: ULO %
    ws_port.cell(row, 10).value = f'=IFERROR(INDEX(T_Master_Projects[ULO_Percent],ROW()-10),"")'
    # K: Days Remaining
    ws_port.cell(row, 11).value = f'=IFERROR(INDEX(T_Master_Projects[Days_Remaining],ROW()-10),"")'

print("  OK - Portfolio Dashboard complete")

print("\n" + "="*80)
print("STEP 3: BUILD COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

# Country name lookup
ws_country['D2'] = '=IFERROR(INDEX(T_Country_Regions[Country_Name],MATCH(B2,T_Country_Regions[Country_Code],0)),"")'

# Summary metrics
ws_country['A5'] = '=SUMIF(T_Country_Budgets[Country_Code],$B$2,T_Country_Budgets[Allocated_Amount])'
ws_country['C5'] = '=SUMIF(T_Country_Budgets[Country_Code],$B$2,T_Country_Budgets[Obligated_Amount])'
ws_country['E5'] = '=SUMIF(T_Country_Budgets[Country_Code],$B$2,T_Country_Budgets[Spent_Amount])'
ws_country['G5'] = '=C5-E5'
ws_country['B6'] = '=COUNTIF(T_Country_Budgets[Country_Code],$B$2)'
ws_country['F6'] = '=IF(C5=0,0,G5/C5)'

# PM lookup
ws_country['B8'] = '=IFERROR(INDEX(T_Country_PM_Assignments[Project_Manager],MATCH($B$2,T_Country_PM_Assignments[Country_Code],0)),"")'
ws_country['F8'] = '=IFERROR(INDEX(T_Country_PM_Assignments[PM_Email],MATCH($B$2,T_Country_PM_Assignments[Country_Code],0)),"")'

print("Adding Country Dashboard formulas (rows 12-31)...")
# Project list for selected country (rows 12-31)
for row in range(12, 32):
    idx = row - 11  # 1, 2, 3...

    # A: Project ID
    ws_country.cell(row, 1).value = f'=IFERROR(INDEX(T_Country_Budgets[Unique_ID],SMALL(IF(T_Country_Budgets[Country_Code]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{idx})),"")'

    # B: Project Name (lookup from Master_Projects)
    ws_country.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Name],MATCH(A{row},T_Master_Projects[Project_Unique_ID],0)),""))'

    # C: Summary
    ws_country.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Summary],MATCH(A{row},T_Master_Projects[Project_Unique_ID],0)),""))'

    # D: Status
    ws_country.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Status],MATCH(A{row},T_Master_Projects[Project_Unique_ID],0)),""))'

    # E-H: Budget amounts (these will need array formulas, simpler to reference the budget row directly)
    ws_country.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Country_Budgets[Allocated_Amount],SMALL(IF(T_Country_Budgets[Country_Code]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{idx})),""))'
    ws_country.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Country_Budgets[Obligated_Amount],SMALL(IF(T_Country_Budgets[Country_Code]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{idx})),""))'
    ws_country.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Country_Budgets[Spent_Amount],SMALL(IF(T_Country_Budgets[Country_Code]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{idx})),""))'
    ws_country.cell(row, 8).value = f'=IF(A{row}="","",F{row}-G{row})'

print("  OK - Country Dashboard complete")

print("\n" + "="*80)
print("STEP 4: BUILD REGIONAL_SUMMARY")
print("="*80)

ws_regional = wb['Regional_Summary']

print("Adding Regional Summary formulas (rows 2-7)...")
# Regions in column A (should already have: AF, EAP, EUR, NEA, SCA, WHA)
for row in range(2, 8):
    # C: Project Count
    ws_regional.cell(row, 3).value = f'=SUMPRODUCT((ISNUMBER(FIND(A{row},T_Master_Projects[Countries])))*1)'

    # D: Country Count
    ws_regional.cell(row, 4).value = f'=COUNTIF(T_Country_Regions[Region],A{row})'

    # E: Active Budgets
    ws_regional.cell(row, 5).value = f'=SUMPRODUCT((T_Country_Budgets[Country_Code]<>"")*(T_Country_Regions[Region]=A{row}))'

    # F: Total Allocated (region)
    ws_regional.cell(row, 6).value = f'=SUMPRODUCT((T_Country_Regions[Region]=A{row})*(ISNUMBER(MATCH(T_Country_Regions[Country_Code],T_Country_Budgets[Country_Code],0)))*(OFFSET(T_Country_Budgets[[#Headers],[Allocated_Amount]],MATCH(T_Country_Regions[Country_Code],T_Country_Budgets[Country_Code],0),0,1)))'

    # Simplified version for F, G, H
    ws_regional.cell(row, 6).value = f'=SUMIF(T_Country_Regions[Region],A{row},T_Country_Budgets[Allocated_Amount])'
    ws_regional.cell(row, 7).value = f'=SUMIF(T_Country_Regions[Region],A{row},T_Country_Budgets[Obligated_Amount])'
    ws_regional.cell(row, 8).value = f'=SUMIF(T_Country_Regions[Region],A{row},T_Country_Budgets[Spent_Amount])'

    # I: ULO
    ws_regional.cell(row, 9).value = f'=F{row}-G{row}'

    # J: ULO %
    ws_regional.cell(row, 10).value = f'=IF(F{row}=0,0,I{row}/F{row})'

    # K: Execution Rate
    ws_regional.cell(row, 11).value = f'=IF(G{row}=0,0,H{row}/G{row})'

    # L: Status
    ws_regional.cell(row, 12).value = f'=IF(J{row}>0.5,"High ULO",IF(K{row}<0.25,"Low Execution","On Track"))'

print("  OK - Regional Summary complete")

# Save
print("\n" + "="*80)
print("SAVING V20...")
print("="*80)

wb.save('2025-10-26-Tracker-v20.xlsx')

print("\nOK - V20 CREATED!")
print("\nFile: 2025-10-26-Tracker-v20.xlsx")
print("\nDashboards with formulas:")
print("  OK - Control: All metrics")
print("  OK - Portfolio_Dashboard: Project list with metrics")
print("  OK - Country_Dashboard: Country-specific view")
print("  OK - Regional_Summary: Region rollups")
print("\nNote: Some Country_Dashboard and Regional_Summary formulas use array syntax")
print("You may need to enter them with Ctrl+Shift+Enter in Excel")
print("\nNext: Test by opening v20 and checking if dashboards calculate correctly")
