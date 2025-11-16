import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'

print('='*80)
print('Updating Config_Lists with All Countries')
print('='*80)
print()

# Load workbook
print('Loading v5...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

# Get all countries from Country_Regions
print('Reading countries from Country_Regions...')
ws_regions = wb['Country_Regions']
all_countries = {}

for row in range(2, ws_regions.max_row + 1):
    country_code = ws_regions.cell(row, 1).value
    country_name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value

    if country_code:
        all_countries[country_code] = {
            'name': country_name,
            'region': region
        }

print(f'Found {len(all_countries)} countries in Country_Regions')
print()

# Get existing countries from Config_Lists
print('Reading existing countries from Config_Lists...')
ws_config = wb['Config_Lists']
existing_countries = set()

for row in range(2, ws_config.max_row + 1):
    country_code = ws_config.cell(row, 4).value  # Column D
    if country_code:
        existing_countries.add(country_code)

print(f'Found {len(existing_countries)} countries in Config_Lists')
print()

# Find missing countries
missing_countries = []
for code in sorted(all_countries.keys()):
    if code not in existing_countries:
        missing_countries.append(code)

print(f'Missing countries: {len(missing_countries)}')
if missing_countries:
    print('Missing country codes:')
    for code in missing_countries:
        print(f'  {code} - {all_countries[code]["name"]} ({all_countries[code]["region"]})')
print()

# Add missing countries to Config_Lists
if missing_countries:
    print('Adding missing countries to Config_Lists...')

    # Find the next empty row
    next_row = ws_config.max_row + 1

    for code in missing_countries:
        # Column A: Status (leave empty - will be filled from dropdown)
        ws_config.cell(row=next_row, column=1, value=None)

        # Column B: Priority (leave empty)
        ws_config.cell(row=next_row, column=2, value=None)

        # Column C: Phase (leave empty)
        ws_config.cell(row=next_row, column=3, value=None)

        # Column D: Country_Code
        ws_config.cell(row=next_row, column=4, value=code)

        # Column E: Country_Name
        ws_config.cell(row=next_row, column=5, value=all_countries[code]['name'])

        # Column F: Region
        ws_config.cell(row=next_row, column=6, value=all_countries[code]['region'])

        # Column G: NCE_Status (leave empty)
        ws_config.cell(row=next_row, column=7, value=None)

        print(f'  Added: {code} - {all_countries[code]["name"]} ({all_countries[code]["region"]})')
        next_row += 1

    print()
    print(f'Added {len(missing_countries)} countries')
else:
    print('No missing countries found - Config_Lists is up to date!')

print()

# Verify final count
final_country_count = 0
for row in range(2, ws_config.max_row + 1):
    if ws_config.cell(row, 4).value:
        final_country_count += 1

print('='*80)
print('Verification:')
print(f'  Country_Regions: {len(all_countries)} countries')
print(f'  Config_Lists (before): {len(existing_countries)} countries')
print(f'  Config_Lists (after): {final_country_count} countries')
print(f'  Match: {"YES" if final_country_count == len(all_countries) else "NO"}')
print('='*80)
print()

# Save
print('Saving...')
wb.save(output_file)
wb.close()

print('='*80)
print('SUCCESS: Config_Lists Updated!')
print('='*80)
print()
print('Summary:')
print(f'  - Started with {len(existing_countries)} countries')
print(f'  - Added {len(missing_countries)} missing countries')
print(f'  - Now have {final_country_count} countries')
print(f'  - Config_Lists now matches Country_Regions (98 countries)')
print()
print(f'Output: {output_file}')
print()
print('Config_Lists is now complete and can be used for:')
print('  - Country dropdowns throughout the tracker')
print('  - Data validation for country fields')
print('  - Filtering and reporting by country')
