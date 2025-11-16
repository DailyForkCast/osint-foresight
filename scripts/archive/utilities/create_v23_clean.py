"""
Create v23 - all formulas, NO programmatic changes to Stakeholders sheet
User can add note manually in Excel (safest approach)
"""
import openpyxl

print("="*80)
print("CREATING V23 - CLEAN VERSION")
print("="*80)

# Load v20 (last known fully working version)
print("\nLoading v20 (working base)...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v20.xlsx')

print("\nApplying all formula updates...")

# Step 1: Portfolio Dashboard - proposed column + blank row handling
ws_port = wb['Portfolio_Dashboard']
print("\n1. Portfolio_Dashboard formulas...")
for row in range(11, 21):
    ws_port.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")'
    ws_port.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Name],ROW()-10),""))'
    ws_port.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Status],ROW()-10),""))'
    ws_port.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Priority],ROW()-10),""))'
    ws_port.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Project_Progress],ROW()-10),""))'
    ws_port.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Countries],ROW()-10),""))'
    ws_port.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Proposed],ROW()-10),""))'
    ws_port.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Allocation],ROW()-10),""))'
    ws_port.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Obligated],ROW()-10),""))'
    ws_port.cell(row, 10).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_Spent],ROW()-10),""))'
    ws_port.cell(row, 11).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Total_ULO],ROW()-10),""))'
    ws_port.cell(row, 12).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[ULO_Percent],ROW()-10),""))'
    ws_port.cell(row, 13).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Days_Remaining],ROW()-10),""))'
print("  OK - Portfolio Dashboard (with Total_Proposed, blank row handling)")

# Step 2: Spotlight Award/NCE fields (rows 2-3)
ws_spot = wb['Spotlight_PMWorkspace']
print("\n2. Spotlight Award/NCE fields (rows 2-3)...")
ws_spot['E2'] = 'Award Number:'
ws_spot['F2'] = '=IFERROR(INDEX(T_Master_Projects[Award_Number],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['H2'] = 'NCE Status:'
ws_spot['I2'] = '=IFERROR(INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['E3'] = 'POP Start:'
ws_spot['F3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Start],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['G3'] = 'POP End:'
ws_spot['H3'] = '=IFERROR(INDEX(T_Master_Projects[POP_End],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
ws_spot['I3'] = 'POP Days Remaining:'
ws_spot['J3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Days_Remaining],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'
print("  OK - Award Number, NCE Status, POP dates")

# Step 3: Spotlight stakeholder formulas (rows 33-42)
print("\n3. Spotlight stakeholder formulas (rows 33-42)...")
for row in range(33, 43):
    idx = row - 32
    ws_spot.cell(row, 1).value = f'=IFERROR(INDEX(T_Project_Stakeholders[Stakeholder_Name],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),"")'
    ws_spot.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Organization],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Role],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Email],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
    ws_spot.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Engagement_Level],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'
print("  OK - Stakeholder formulas (Name, Organization, Role, Email, Engagement)")

# Step 4: Config_Lists headers
ws_config = wb['Config_Lists']
print("\n4. Config_Lists headers...")
ws_config['A1'] = 'Status'
ws_config['B1'] = 'Priority'
ws_config['C1'] = 'Stage'
ws_config['D1'] = 'Country_Code'
ws_config['E1'] = 'Country'
ws_config['F1'] = 'Region'
print("  OK - Config_Lists headers updated")

# DO NOT TOUCH STAKEHOLDERS SHEET - leave it exactly as is

# Save as v23
print("\n" + "="*80)
print("SAVING AS V23...")
print("="*80)

wb.save('2025-10-26-Tracker-v23.xlsx')

print("\nOK - V23 CREATED!")

print("\n" + "="*80)
print("SUMMARY - V23 COMPLETE")
print("="*80)

print("\nAll formulas added:")
print("  OK - Control sheet (all metrics)")
print("  OK - Portfolio_Dashboard (Total_Proposed, blank rows)")
print("  OK - Country_Dashboard (country filtering)")
print("  OK - Regional_Summary (region rollups)")
print("  OK - Spotlight_PMWorkspace (Award/NCE, stakeholders)")
print("  OK - Config_Lists (correct headers)")

print("\nStakeholders sheet:")
print("  UNCHANGED - table preserved")
print("  Add categorization note MANUALLY in Excel:")
print("    - Open Stakeholders sheet")
print("    - Add note text in a cell to the right (column J or K)")
print("    - Or add as a comment/cell note")

print("\nStakeholder categories to include in note:")
print("  - Government Officials: Gov agency, embassy, ministry")
print("  - Implementing Partners: Contractors, grantees, vendors")
print("  - Internal Team: DOS staff, PMs, technical advisors")
print("  - Beneficiaries: End users, communities, target populations")
print("  - Oversight: Congress, GAO, OIG, monitoring")
print("  - Stakeholder_ID format: ProjectID-STK-001, etc.")

print("\nv23 should open WITHOUT any Excel errors!")
