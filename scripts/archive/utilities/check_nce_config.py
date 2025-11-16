"""
Check Config_Lists for NCE_Status values
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=True)

print("="*80)
print("CONFIG_LISTS - NCE_STATUS VALUES")
print("="*80)

ws = wb['Config_Lists']

print("\nLooking for NCE_Status list:")
for row in range(2, ws.max_row + 1):
    list_type = ws.cell(row, 1).value
    if list_type and 'NCE' in str(list_type).upper():
        print(f"\nList Type: {list_type}")
        values = []
        for col in range(2, 8):
            val = ws.cell(row, col).value
            if val:
                values.append(val)
        print(f"Values: {', '.join(values)}")
