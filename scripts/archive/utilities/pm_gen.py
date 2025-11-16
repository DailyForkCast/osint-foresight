from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime

def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for s in ["_SETUP","Portfolio_Dashboard","Project_Spotlight","Master_Projects","Country_Budgets","Milestones","Events","Calendar_Todo","Decision_Log","Risk_Register","Stakeholders","Config_Lists","Country_Regions","Regional_Summary"]:
        wb.create_sheet(s)
    return wb

if __name__ == "__main__":
    wb = create_workbook()
    wb.save("PM_Dashboard_Template.xlsx")
    print("Dashboard template created successfully!")
