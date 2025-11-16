"""
Create v26 with Priority 1 fixes:
1. Fix Regional_Summary region codes
2. Add missing Master_Projects formulas (columns J, Z)
3. Create T_Country_Regions table
4. Fix Country_Dashboard D2 formula
5. Fix Portfolio_Dashboard column headers
"""
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

print("="*80)
print("CREATING V26 - PRIORITY 1 FIXES")
print("="*80)

# Load v25
print("\nLoading v25...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v25.xlsx')

# ============================================================================
# FIX 1: REGIONAL_SUMMARY REGION CODES
# ============================================================================
print("\n" + "="*80)
print("FIX 1: REGIONAL_SUMMARY REGION CODES")
print("="*80)

ws_regional = wb['Regional_Summary']

correct_regions = {
    2: 'AF',
    3: 'EAP',
    4: 'EUR',
    5: 'NEA',
    6: 'SCA',
    7: 'WHA'
}

print("\nFixing region codes in column A...")
for row, region_code in correct_regions.items():
    old_value = ws_regional.cell(row, 1).value
    ws_regional.cell(row, 1).value = region_code
    print(f"  Row {row}: '{old_value}' -> '{region_code}'")

print("\n  OK - Regional_Summary region codes corrected")

# ============================================================================
# FIX 2: MASTER_PROJECTS MISSING FORMULAS
# ============================================================================
print("\n" + "="*80)
print("FIX 2: MASTER_PROJECTS MISSING FORMULAS")
print("="*80)

ws_mp = wb['Master_Projects']

print("\nAdding formulas to Master_Projects...")

# Column J: Days_Remaining
print("\n  Column J (Days_Remaining):")
print("    Formula: =IF(OR(I2=\"\",E2=\"\"),\"\",IF(E2=\"Completed\",\"Complete\",INT(I2-TODAY())))")

# Add to row 2 first
ws_mp['J2'] = '=IF(OR(I2="",E2=""),"",IF(E2="Completed","Complete",INT(I2-TODAY())))'
print("    Added to row 2")

# Copy down to row 201 (end of table)
for row in range(3, 202):
    ws_mp.cell(row, 10).value = f'=IF(OR(I{row}="",E{row}=""),"",IF(E{row}="Completed","Complete",INT(I{row}-TODAY())))'

print("    Copied to rows 3-201")

# Column Z: Total_Spent
print("\n  Column Z (Total_Spent):")
print("    Formula: =SUMIF(Country_Budgets!$B:$B,B2,Country_Budgets!$I:$I)")

# Add to row 2 first
ws_mp['Z2'] = '=IF(B2="","",SUMIF(Country_Budgets!$B$2:INDEX(Country_Budgets!$B:$B,COUNTA(Country_Budgets!$B:$B)),B2,Country_Budgets!$I$2:INDEX(Country_Budgets!$I:$I,COUNTA(Country_Budgets!$I:$I))))'
print("    Added to row 2")

# Copy down to row 201
for row in range(3, 202):
    ws_mp.cell(row, 26).value = f'=IF(B{row}="","",SUMIF(Country_Budgets!$B$2:INDEX(Country_Budgets!$B:$B,COUNTA(Country_Budgets!$B:$B)),B{row},Country_Budgets!$I$2:INDEX(Country_Budgets!$I:$I,COUNTA(Country_Budgets!$I:$I))))'

print("    Copied to rows 3-201")

print("\n  OK - Master_Projects formulas added")

# ============================================================================
# FIX 3: CREATE T_COUNTRY_REGIONS TABLE
# ============================================================================
print("\n" + "="*80)
print("FIX 3: CREATE T_COUNTRY_REGIONS TABLE")
print("="*80)

ws_regions = wb['Country_Regions']

print("\nChecking Country_Regions sheet structure...")

# Check if headers exist
if ws_regions['A1'].value is None:
    print("  Creating headers...")
    ws_regions['A1'] = 'Country_Code'
    ws_regions['B1'] = 'Country_Name'
    ws_regions['C1'] = 'Region'
    print("    Headers created: Country_Code, Country_Name, Region")
else:
    print(f"  Headers already exist: {ws_regions['A1'].value}, {ws_regions['B1'].value}, {ws_regions['C1'].value}")

# Create the table
print("\nCreating T_Country_Regions table...")

# Define table range (A1:C100 to allow for data entry)
table_range = 'A1:C100'

# Create table
tab = Table(displayName="T_Country_Regions", ref=table_range)

# Add a style
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

# Add table to sheet
ws_regions.add_table(tab)

print(f"  Table created: T_Country_Regions")
print(f"  Range: {table_range}")
print(f"  Columns: Country_Code, Country_Name, Region")

print("\n  OK - T_Country_Regions table created")

# ============================================================================
# FIX 4: COUNTRY_DASHBOARD D2 FORMULA
# ============================================================================
print("\n" + "="*80)
print("FIX 4: COUNTRY_DASHBOARD D2 FORMULA")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nFixing country name lookup in D2...")

old_formula = ws_country['D2'].value
print(f"  Old formula: {old_formula}")

# New formula using T_Country_Regions table
new_formula = '=IFERROR(INDEX(T_Country_Regions[Country_Name],MATCH(B2,T_Country_Regions[Country_Code],0)),"")'
ws_country['D2'] = new_formula

print(f"  New formula: {new_formula}")
print("\n  OK - Country_Dashboard D2 formula fixed")

# ============================================================================
# FIX 5: PORTFOLIO_DASHBOARD COLUMN HEADERS
# ============================================================================
print("\n" + "="*80)
print("FIX 5: PORTFOLIO_DASHBOARD COLUMN HEADERS")
print("="*80)

ws_port = wb['Portfolio_Dashboard']

print("\nFixing column headers in row 10...")

# Current incorrect headers
header_fixes = {
    'G': ('Proposed', 'Total Proposed'),
    'H': ('Allocation', 'Total Allocation'),
    'I': ('Obligated', 'Total Obligated'),
    'J': ('ULO', 'Total Spent'),
    'K': ('ULO %', 'Total ULO'),
    'L': ('Days Left', 'ULO %'),
    'M': ('Health', 'Days Remaining')
}

for col_letter, (old_header, new_header) in header_fixes.items():
    cell = ws_port[f'{col_letter}10']
    old = cell.value
    cell.value = new_header
    print(f"  {col_letter}10: '{old}' -> '{new_header}'")

print("\n  OK - Portfolio_Dashboard headers corrected")

# ============================================================================
# SAVE V26
# ============================================================================
print("\n" + "="*80)
print("SAVING V26")
print("="*80)

print("\nSaving as 2025-10-26-Tracker-v26.xlsx...")
wb.save('2025-10-26-Tracker-v26.xlsx')

print("\n" + "="*80)
print("V26 CREATED SUCCESSFULLY!")
print("="*80)

print("\n" + "="*80)
print("SUMMARY OF FIXES")
print("="*80)

print("\nOK FIX 1: Regional_Summary region codes corrected (AF, EAP, EUR, NEA, SCA, WHA)")
print("OK FIX 2: Master_Projects formulas added:")
print("    - Column J (Days_Remaining): Date calculation from Project_End_Date")
print("    - Column Z (Total_Spent): SUMIF from Country_Budgets")
print("OK FIX 3: T_Country_Regions table created")
print("    - Headers: Country_Code, Country_Name, Region")
print("    - Range: A1:C100")
print("    - Ready for data entry")
print("OK FIX 4: Country_Dashboard D2 formula fixed")
print("    - Now uses T_Country_Regions table for country name lookup")
print("OK FIX 5: Portfolio_Dashboard column headers corrected")
print("    - G: Total Proposed")
print("    - H: Total Allocation")
print("    - I: Total Obligated")
print("    - J: Total Spent")
print("    - K: Total ULO")
print("    - L: ULO %")
print("    - M: Days Remaining")

print("\n" + "="*80)
print("REMAINING MANUAL WORK")
print("="*80)

print("\n1. Populate T_Country_Regions table (Country_Regions sheet)")
print("   Add rows with: Country_Code, Country_Name, Region")
print("   Example:")
print("     AF | Afghanistan | SCA")
print("     CN | China | EAP")
print("     DE | Germany | EUR")
print("     etc.")

print("\n2. Add data validation dropdowns:")
print("   - Master_Projects E (Status) → Config_Lists!A:A")
print("   - Master_Projects F (Priority) → Config_Lists!B:B")
print("   - Master_Projects N (NCE_Eligible) → Yes, No")
print("   - Master_Projects O (NCE_Status) → custom list")
print("   - Country_Budgets D (Country_Code) → Config_Lists!D:D")
print("   - Spotlight B2 (Project_Unique_ID) → T_Master_Projects[Project_Unique_ID]")

print("\n3. Populate Config_Lists with values:")
print("   - Column A: Status values")
print("   - Column B: Priority values")
print("   - Column C: Stage values")
print("   - Column D: Country codes")
print("   - Column E: Country names")
print("   - Column F: Region codes")

print("\n4. Add stakeholder categorization note (Stakeholders sheet, column J)")

print("\nv26 is ready to use!")
