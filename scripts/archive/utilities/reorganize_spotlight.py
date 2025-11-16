import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("REORGANIZING PROJECT SPOTLIGHT - REMOVING MAP SPACE")
print("=" * 70)

# Load the workbook
print("\n[1/2] Loading workbook...")
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-FIXED.xlsx')
ws = wb['Project_Spotlight']
print("✓ Workbook loaded")

print("\n[2/2] Reorganizing layout...")

# Clear everything from row 14 onwards (where map would have been)
max_row = ws.max_row
if max_row > 14:
    ws.delete_rows(14, max_row - 13)
    print(f"✓ Cleared {max_row - 13} rows of old content")

# NEW LAYOUT STRUCTURE:
# Rows 1-2: Header and Project Selector (keep as-is)
# Rows 3-13: Summary in B-E, Target Audiences in F-G (keep as-is)
# Row 14: Section divider
# Rows 15-30: TARGET TECHNOLOGIES section
# Rows 32-50: KEY DELIVERABLES section

# === TARGET TECHNOLOGIES SECTION ===
ws['B14'] = 'TARGET TECHNOLOGIES'
ws['B14'].font = Font(bold=True, size=12, color="FFFFFF")
ws['B14'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells('B14:G14')
ws['B14'].alignment = Alignment(horizontal='left', vertical='center')

# Add Technology table headers
tech_headers = ['Technology Area', 'Priority', 'Status', 'Notes']
for col, header in enumerate(tech_headers, 2):  # Starting at column B
    cell = ws.cell(row=15, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Add sample technology rows (user will fill these in)
tech_examples = [
    ('Cloud Computing', 'High', 'Active', ''),
    ('Artificial Intelligence', 'High', 'Active', ''),
    ('Cybersecurity', 'Critical', 'Active', ''),
    ('IoT/Edge Computing', 'Medium', 'Planned', ''),
    ('Data Analytics', 'High', 'Active', ''),
]

for i, (tech, priority, status, notes) in enumerate(tech_examples, 16):
    ws.cell(row=i, column=2, value=tech)
    ws.cell(row=i, column=3, value=priority)
    ws.cell(row=i, column=4, value=status)
    ws.cell(row=i, column=5, value=notes)

print("✓ Added Target Technologies section (rows 14-20)")

# === KEY DELIVERABLES SECTION ===
ws['B22'] = 'KEY DELIVERABLES'
ws['B22'].font = Font(bold=True, size=12, color="FFFFFF")
ws['B22'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells('B22:G22')
ws['B22'].alignment = Alignment(horizontal='left', vertical='center')

# Add Deliverables table headers
deliv_headers = ['Deliverable Name', 'Type', 'Due Date', 'Status', 'Owner', 'Progress']
for col, header in enumerate(deliv_headers, 2):  # Starting at column B
    cell = ws.cell(row=23, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Add instruction row
ws['B24'] = '[Add deliverables here - can link to Project_Deliverables sheet]'
ws['B24'].font = Font(italic=True, color="666666")
ws.merge_cells('B24:G24')

# Add sample deliverable rows for structure
deliv_examples = [
    ('Technical Requirements Document', 'Document', '01/15/2025', 'Complete', 'John Smith', '100%'),
    ('Beta Release', 'Software', '02/28/2025', 'In Progress', 'Sarah Chen', '75%'),
    ('User Training Materials', 'Training', '03/15/2025', 'Not Started', 'Maria Garcia', '0%'),
    ('', '', '', '', '', ''),
    ('', '', '', '', '', ''),
]

for i, (name, dtype, due, status, owner, progress) in enumerate(deliv_examples, 25):
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=dtype)
    ws.cell(row=i, column=4, value=due)
    ws.cell(row=i, column=5, value=status)
    ws.cell(row=i, column=6, value=owner)
    ws.cell(row=i, column=7, value=progress)

print("✓ Added Key Deliverables section (rows 22-29)")

# === ADJUST COLUMN WIDTHS ===
ws.column_dimensions['B'].width = 30  # Deliverable/Tech name
ws.column_dimensions['C'].width = 15  # Type/Priority
ws.column_dimensions['D'].width = 12  # Due Date/Status
ws.column_dimensions['E'].width = 15  # Status/Notes
ws.column_dimensions['F'].width = 15  # Owner
ws.column_dimensions['G'].width = 10  # Progress

print("✓ Adjusted column widths")

# Save
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-FINAL.xlsx'
print(f"\n{'='*70}")
print("Saving reorganized workbook...")
wb.save(output_file)
print(f"✓ Saved: {output_file}")

print(f"\n{'='*70}")
print("REORGANIZATION COMPLETE!")
print(f"{'='*70}")
print("\nProject Spotlight New Layout:")
print("  Rows 1-2:   Header & Project Selector")
print("  Rows 3-13:  Project Summary & Details")
print("  Rows 14-20: TARGET TECHNOLOGIES section")
print("  Rows 22-29: KEY DELIVERABLES section")
print("\n✓ Map space removed - clean, efficient layout!")
print(f"\nFile: {output_file}")
print("\nNOTE: Sample data included - replace with actual data")
