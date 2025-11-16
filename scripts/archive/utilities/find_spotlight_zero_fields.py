"""
Find all fields in Spotlight that might show 0 instead of blank
"""
import openpyxl

print("="*80)
print("FINDING SPOTLIGHT FIELDS THAT MIGHT SHOW 0")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("ALL FORMULAS IN SPOTLIGHT (Rows 1-60, Columns A-F)")
print("="*80)

formula_cells = []

for row in range(1, 61):
    for col in range(1, 7):  # A-F
        cell = ws_spotlight.cell(row, col)

        if cell.data_type == 'f':
            # It's a formula
            label_cell = None

            # Check if there's a label in the same row
            if col == 2:  # Column B - check column A for label
                label_cell = ws_spotlight.cell(row, 1)
            elif col == 4:  # Column D - check column C for label
                label_cell = ws_spotlight.cell(row, 3)

            label = ""
            if label_cell and label_cell.value:
                label = str(label_cell.value).encode('ascii', 'ignore').decode('ascii')[:40]

            formula = str(cell.value)

            # Check if it's INDEX formula (likely pulling from Master_Projects)
            if 'INDEX' in formula:
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_cells.append({
                    'cell': f"{col_letter}{row}",
                    'label': label,
                    'formula': formula[:80]
                })

print(f"\nFound {len(formula_cells)} INDEX formulas:")
for item in formula_cells:
    print(f"\n  {item['cell']}: {item['label']}")
    print(f"    Formula: {item['formula']}...")

print("\n" + "="*80)
print("FIELDS TO FIX")
print("="*80)

print("\nBased on user request, need to fix these types of fields:")
print("  - Progress (percentage field)")
print("  - Implementer (text field)")
print("  - NCE fields (numeric)")
print("  - Any other numeric fields showing 0")

print("\nLet's find specific fields by searching labels:")

# Search for specific terms in row 2-30
search_for = [
    'progress', 'percent', 'completion',
    'implement', 'contractor', 'vendor',
    'nce', 'award', 'contract',
    'allocated', 'obligated', 'spent', 'proposed', 'ulo'
]

found_fields = {}

for row in range(1, 31):
    for col in [1, 3]:  # Label columns A and C
        label_cell = ws_spotlight.cell(row, col)
        if label_cell.value:
            label = str(label_cell.value).lower()

            for term in search_for:
                if term in label:
                    value_col = col + 1
                    value_cell = ws_spotlight.cell(row, value_col)

                    if value_cell.data_type == 'f':
                        col_letter = openpyxl.utils.get_column_letter(value_col)
                        if term not in found_fields:
                            found_fields[term] = []

                        found_fields[term].append({
                            'row': row,
                            'label_col': chr(64+col),
                            'value_col': chr(64+value_col),
                            'label': str(label_cell.value).encode('ascii', 'ignore').decode('ascii'),
                            'formula': str(value_cell.value)[:80]
                        })

print("\nFields found by term:")
for term, fields in found_fields.items():
    print(f"\n  {term.upper()}:")
    for field in fields:
        print(f"    Row {field['row']}: {field['label']}")
        print(f"      Cell: {field['value_col']}{field['row']}")
        print(f"      Formula: {field['formula']}...")

print("\n" + "="*80)
print("NEXT STEPS")
print("="*80)

print("\nWill update formulas to wrap with blank-if-zero logic:")
print("  Old: =INDEX(T_Master_Projects[Field],MATCH(...))")
print("  New: =IF($B$2=\"\",\"\",IFERROR(IF(INDEX(...)=0,\"\",INDEX(...)),\"\"))")

print("\nReady to fix!")
