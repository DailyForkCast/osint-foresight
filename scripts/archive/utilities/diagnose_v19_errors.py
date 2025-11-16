"""
Diagnose v19 formula errors
"""
import openpyxl

print("="*80)
print("DIAGNOSING V19 ERRORS")
print("="*80)

# Load v19 without data_only to see formulas
print("\nLoading v19 to check formulas...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v19.xlsx', data_only=False)

print("\nSheet order (sheet3 is the problem):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  Sheet {i}: {name}")

# Check Master_Projects (likely sheet3)
sheet3_name = wb.sheetnames[2] if len(wb.sheetnames) > 2 else None
if sheet3_name:
    print(f"\nSheet 3 is: {sheet3_name}")
    ws = wb[sheet3_name]

    print("\nChecking for problematic formulas in row 2...")
    for col in range(1, min(35, ws.max_column + 1)):
        cell = ws.cell(2, col)
        header = ws.cell(1, col).value
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.data_type == 'f':
            formula = str(cell.value)
            # Check if formula uses table syntax
            if '[@' in formula or 'T_' in formula:
                print(f"\n  {col_letter} ({header}):")
                print(f"    Formula: {formula[:80]}...")

                # Check if it's a table cell (should work) or regular cell (might not)
                if ws.tables:
                    print(f"    INFO: Sheet has {len(ws.tables)} table(s)")

# Check if tables exist
print("\n" + "="*80)
print("CHECKING FOR EXCEL TABLES")
print("="*80)

table_names = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.tables:
        for table in ws.tables.values():
            table_names.append((sheet_name, table.name))
            print(f"  {sheet_name}: {table.name}")

if not table_names:
    print("\n  WARNING: NO TABLES FOUND!")
    print("  This is likely the problem - formulas reference tables that don't exist")

# Check v18 to compare
print("\n" + "="*80)
print("COMPARING WITH V18")
print("="*80)

print("\nTables in v18:")
wb18 = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=False)
for sheet_name in wb18.sheetnames:
    ws = wb18[sheet_name]
    if ws.tables:
        for table in ws.tables.values():
            print(f"  {sheet_name}: {table.name}")

print("\n" + "="*80)
print("DIAGNOSIS")
print("="*80)

print("\nMost likely cause:")
print("  The script added formulas with table syntax (like T_Master_Projects)")
print("  BUT when saving the file, the tables were lost or not preserved correctly")
print("\nSolution:")
print("  Option 1: Rebuild tables in v19 manually")
print("  Option 2: Open v18, verify tables exist, then manually add the few missing formulas")
print("  Option 3: Create v20 using openpyxl more carefully to preserve tables")
