"""
Scan current V18 structure
"""
import openpyxl

try:
    wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

    print("="*80)
    print("V18 CURRENT STRUCTURE")
    print("="*80)

    print("\nALL SHEETS:")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {i:2}. {name:<35} ({ws.max_row} rows x {ws.max_column} cols)")

    # Focus on Spotlight_PMWorkspace
    if 'Spotlight_PMWorkspace' in wb.sheetnames:
        print("\n" + "="*80)
        print("SPOTLIGHT_PMWORKSPACE LAYOUT")
        print("="*80)

        ws = wb['Spotlight_PMWorkspace']

        # Show rows 1-25 to see the structure
        print("\nFirst 25 rows (showing columns A-C for context):")
        print(f"{'Row':<5} {'Col A':<30} {'Col B':<30} {'Col C':<30}")
        print("-"*100)

        for row in range(1, min(26, ws.max_row + 1)):
            val_a = ws.cell(row, 1).value or ""
            val_b = ws.cell(row, 2).value or ""
            val_c = ws.cell(row, 3).value or ""

            # Convert to string and truncate
            val_a = str(val_a)[:28]
            val_b = str(val_b)[:28]
            val_c = str(val_c)[:28]

            print(f"{row:<5} {val_a:<30} {val_b:<30} {val_c:<30}")

        # Highlight rows 13-15
        print("\n" + "="*80)
        print("FINANCIAL DETAILS (Rows 13-15):")
        print("="*80)
        for row in range(13, 16):
            print(f"\nRow {row}:")
            for col in range(1, min(11, ws.max_column + 1)):
                val = ws.cell(row, col).value
                if val:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"  {col_letter}: {val}")

    # Check Master_Projects structure
    print("\n" + "="*80)
    print("MASTER_PROJECTS HEADERS")
    print("="*80)
    ws = wb['Master_Projects']
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {header}")

    # Check Country_Budgets structure
    print("\n" + "="*80)
    print("COUNTRY_BUDGETS HEADERS")
    print("="*80)
    ws = wb['Country_Budgets']
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {header}")

except PermissionError:
    print("="*80)
    print("WARNING: FILE IS LOCKED")
    print("="*80)
    print("\nThe file is open in Excel and locked for reading.")
    print("Please close Excel and I'll try again.")

except Exception as e:
    print(f"Error: {e}")
