"""
Build Tracker v15 from scratch with correct structure
This avoids all the Excel Table issues by building everything correctly from the start
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

print("="*80)
print("BUILDING TRACKER V15 FROM SCRATCH")
print("="*80)

# Load v12 to get the data
print("\nLoading v12 to extract data...")
wb_v12 = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx')

# Create new workbook
print("Creating new workbook...")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Helper function for headers
def create_header_row(ws, row, headers, bg_color='366092'):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

# State Department regions mapping
STATE_DEPT_REGIONS = {
    # Africa (AF)
    'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
    'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
    'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
    'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
    'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
    'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
    'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',
    # East Asia & Pacific (EAP)
    'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
    'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
    'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
    'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
    'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',
    # Europe (EUR)
    'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
    'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
    'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
    'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
    'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
    'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
    'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR',
    'VA': 'EUR',
    # Near East (NEA)
    'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA',
    'LB': 'NEA', 'OM': 'NEA', 'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA',
    'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',
    # South & Central Asia (SCA)
    'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
    'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',
    # Western Hemisphere (WHA)
    'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
    'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
    'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
    'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
    'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
}

SUBREGIONS = {
    'DZ': 'Northern Africa', 'EG': 'Northern Africa', 'LY': 'Northern Africa', 'MA': 'Northern Africa',
    'TN': 'Northern Africa', 'SD': 'Northern Africa',
    'BJ': 'Western Africa', 'BF': 'Western Africa', 'CV': 'Western Africa', 'CI': 'Western Africa',
    'GM': 'Western Africa', 'GH': 'Western Africa', 'GN': 'Western Africa', 'GW': 'Western Africa',
    'LR': 'Western Africa', 'ML': 'Western Africa', 'MR': 'Western Africa', 'NE': 'Western Africa',
    'NG': 'Western Africa', 'SN': 'Western Africa', 'SL': 'Western Africa', 'TG': 'Western Africa',
    'BI': 'Eastern Africa', 'KM': 'Eastern Africa', 'DJ': 'Eastern Africa', 'ER': 'Eastern Africa',
    'ET': 'Eastern Africa', 'KE': 'Eastern Africa', 'MG': 'Eastern Africa', 'MW': 'Eastern Africa',
    'MU': 'Eastern Africa', 'MZ': 'Eastern Africa', 'RW': 'Eastern Africa', 'SC': 'Eastern Africa',
    'SO': 'Eastern Africa', 'SS': 'Eastern Africa', 'TZ': 'Eastern Africa', 'UG': 'Eastern Africa',
    'ZM': 'Eastern Africa', 'ZW': 'Eastern Africa',
    'AO': 'Middle Africa', 'CM': 'Middle Africa', 'CF': 'Middle Africa', 'TD': 'Middle Africa',
    'CG': 'Middle Africa', 'CD': 'Middle Africa', 'GQ': 'Middle Africa', 'GA': 'Middle Africa',
    'BW': 'Southern Africa', 'LS': 'Southern Africa', 'NA': 'Southern Africa', 'ZA': 'Southern Africa',
    'SZ': 'Southern Africa',
    'DK': 'Northern Europe', 'EE': 'Northern Europe', 'FI': 'Northern Europe', 'IS': 'Northern Europe',
    'IE': 'Northern Europe', 'LV': 'Northern Europe', 'LT': 'Northern Europe', 'NO': 'Northern Europe',
    'SE': 'Northern Europe', 'GB': 'Northern Europe',
    'BY': 'Eastern Europe', 'BG': 'Eastern Europe', 'CZ': 'Eastern Europe', 'HU': 'Eastern Europe',
    'MD': 'Eastern Europe', 'PL': 'Eastern Europe', 'RO': 'Eastern Europe', 'RU': 'Eastern Europe',
    'SK': 'Eastern Europe', 'UA': 'Eastern Europe',
    'AL': 'Southern Europe', 'AD': 'Southern Europe', 'BA': 'Southern Europe', 'HR': 'Southern Europe',
    'CY': 'Southern Europe', 'GR': 'Southern Europe', 'IT': 'Southern Europe', 'XK': 'Southern Europe',
    'MK': 'Southern Europe', 'MT': 'Southern Europe', 'ME': 'Southern Europe', 'PT': 'Southern Europe',
    'RS': 'Southern Europe', 'SI': 'Southern Europe', 'ES': 'Southern Europe', 'VA': 'Southern Europe',
    'AT': 'Western Europe', 'BE': 'Western Europe', 'FR': 'Western Europe', 'DE': 'Western Europe',
    'LI': 'Western Europe', 'LU': 'Western Europe', 'MC': 'Western Europe', 'NL': 'Western Europe',
    'CH': 'Western Europe',
    'AM': 'Caucasus', 'AZ': 'Caucasus', 'GE': 'Caucasus',
    'TR': 'Western Asia',
    'CA': 'Northern America', 'US': 'Northern America',
    'BZ': 'Central America', 'CR': 'Central America', 'SV': 'Central America', 'GT': 'Central America',
    'HN': 'Central America', 'MX': 'Central America', 'NI': 'Central America', 'PA': 'Central America',
    'AG': 'Caribbean', 'BS': 'Caribbean', 'BB': 'Caribbean', 'CU': 'Caribbean', 'DM': 'Caribbean',
    'DO': 'Caribbean', 'GD': 'Caribbean', 'HT': 'Caribbean', 'JM': 'Caribbean', 'KN': 'Caribbean',
    'LC': 'Caribbean', 'VC': 'Caribbean', 'TT': 'Caribbean',
    'AR': 'South America', 'BO': 'South America', 'BR': 'South America', 'CL': 'South America',
    'CO': 'South America', 'EC': 'South America', 'GY': 'South America', 'PY': 'South America',
    'PE': 'South America', 'SR': 'South America', 'UY': 'South America', 'VE': 'South America',
    'CN': 'Eastern Asia', 'JP': 'Eastern Asia', 'KP': 'Eastern Asia', 'KR': 'Eastern Asia',
    'MN': 'Eastern Asia', 'TW': 'Eastern Asia',
    'BN': 'South-Eastern Asia', 'KH': 'South-Eastern Asia', 'ID': 'South-Eastern Asia',
    'LA': 'South-Eastern Asia', 'MY': 'South-Eastern Asia', 'MM': 'South-Eastern Asia',
    'PH': 'South-Eastern Asia', 'SG': 'South-Eastern Asia', 'TH': 'South-Eastern Asia',
    'TL': 'South-Eastern Asia', 'VN': 'South-Eastern Asia',
    'AF': 'Southern Asia', 'BD': 'Southern Asia', 'BT': 'Southern Asia', 'IN': 'Southern Asia',
    'MV': 'Southern Asia', 'NP': 'Southern Asia', 'PK': 'Southern Asia', 'LK': 'Southern Asia',
    'KZ': 'Central Asia', 'KG': 'Central Asia', 'TJ': 'Central Asia', 'TM': 'Central Asia',
    'UZ': 'Central Asia',
    'AU': 'Australia and New Zealand', 'NZ': 'Australia and New Zealand',
    'FJ': 'Melanesia', 'PG': 'Melanesia', 'SB': 'Melanesia', 'VU': 'Melanesia',
    'KI': 'Micronesia', 'MH': 'Micronesia', 'FM': 'Micronesia', 'NR': 'Micronesia', 'PW': 'Micronesia',
    'WS': 'Polynesia', 'TO': 'Polynesia', 'TV': 'Polynesia',
    'BH': 'Arabian Peninsula', 'KW': 'Arabian Peninsula', 'OM': 'Arabian Peninsula',
    'QA': 'Arabian Peninsula', 'SA': 'Arabian Peninsula', 'AE': 'Arabian Peninsula', 'YE': 'Arabian Peninsula',
    'IL': 'Levant', 'JO': 'Levant', 'LB': 'Levant', 'PS': 'Levant', 'SY': 'Levant',
    'IQ': 'Mesopotamia',
}

# ========== 1. Control Sheet ==========
print("\n1. Creating Control sheet...")
ws_control = wb.create_sheet("Control")
ws_control['A1'] = "TRACKER CONTROL PANEL"
ws_control['A1'].font = Font(size=16, bold=True)
ws_control['A3'] = "Version: 15.0"
ws_control['A4'] = "Last Updated: Auto"
ws_control['A5'] = "Status: Production"

# ========== 2. Master_Projects ==========
print("2. Creating Master_Projects sheet...")
ws_old = wb_v12['Master_Projects']
ws = wb.create_sheet("Master_Projects")

# New header structure with Proposed BEFORE Allocated
headers = [
    'Project_ID', 'Project_Unique_ID', 'Project_Name', 'Project_Status', 'Project_Priority',
    'Project_Progress', 'Project_Start_Date', 'Project_End_Date', 'Days_Remaining',
    'Implementer', 'Implementer_POC', 'Implementer_POC_Phone', 'Implementer_POC_Email',
    'Total_Proposed', 'Total_Allocation', 'Total_Obligated', 'Total_ULO', 'ULO_Percent',
    'Countries', 'Country_Count'
]
create_header_row(ws, 1, headers)

# Copy data from v12, mapping to new structure
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    # Copy basic fields (A-L)
    for col in range(1, 13):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value

    # Add budget formulas in new positions (N-R)
    ws.cell(new_row, 14).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$F:$F)'  # Total_Proposed
    ws.cell(new_row, 15).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$G:$G)'  # Total_Allocation
    ws.cell(new_row, 16).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$H:$H)'  # Total_Obligated
    ws.cell(new_row, 17).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$J:$J)'  # Total_ULO
    ws.cell(new_row, 18).value = f'=IF(P{new_row}>0,Q{new_row}/P{new_row},0)'  # ULO %
    ws.cell(new_row, 18).number_format = '0%'

    # Copy Countries and Country_Count from old positions
    ws.cell(new_row, 19).value = ws_old.cell(old_row, 18).value  # Countries
    ws.cell(new_row, 20).value = ws_old.cell(old_row, 19).value  # Country_Count

print(f"  Copied {ws_old.max_row - 1} projects")

# ========== 3. Country_Budgets ==========
print("3. Creating Country_Budgets sheet...")
ws_old = wb_v12['Country_Budgets']
ws = wb.create_sheet("Country_Budgets")

# New header with Proposed column
headers = [
    'Budget_ID', 'Unique_ID', 'My_Country', 'Country_Code', 'Country_Name',
    'Proposed_Amount', 'Allocated_Amount', 'Obligated_Amount', 'Spent_Amount',
    'ULO', 'ULO_Percent', 'Spend_Health'
]
create_header_row(ws, 1, headers)

# Copy data
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    # Copy A-E (IDs and country info)
    for col in range(1, 6):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value

    # F: Proposed_Amount (default to 0)
    ws.cell(new_row, 6).value = 0

    # G-I: Copy old F-H (Allocated, Obligated, Spent)
    ws.cell(new_row, 7).value = ws_old.cell(old_row, 6).value  # Allocated
    ws.cell(new_row, 8).value = ws_old.cell(old_row, 7).value  # Obligated
    ws.cell(new_row, 9).value = ws_old.cell(old_row, 8).value  # Spent

    # J-K: ULO formulas
    ws.cell(new_row, 10).value = f'=H{new_row}-I{new_row}'  # ULO = Obligated - Spent
    ws.cell(new_row, 11).value = f'=IF(H{new_row}>0,J{new_row}/H{new_row},0)'  # ULO %
    ws.cell(new_row, 11).number_format = '0%'

    # L: Spend Health (copy from old K)
    ws.cell(new_row, 12).value = ws_old.cell(old_row, 11).value

print(f"  Copied {ws_old.max_row - 1} budget entries")

# ========== 4. Country_Regions ==========
print("4. Creating Country_Regions with State Dept regions...")
ws_old = wb_v12['Country_Regions']
ws = wb.create_sheet("Country_Regions")

headers = ['Country_Code', 'Country_Name', 'Region', 'EU_Member', 'Subregion']
create_header_row(ws, 1, headers)

# Copy and update regions
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    code = ws_old.cell(old_row, 1).value
    name = ws_old.cell(old_row, 2).value
    eu_member = ws_old.cell(old_row, 4).value

    ws.cell(new_row, 1).value = code  # Country_Code
    ws.cell(new_row, 2).value = name  # Country_Name
    ws.cell(new_row, 3).value = STATE_DEPT_REGIONS.get(code, 'EUR')  # Region (State Dept)
    ws.cell(new_row, 4).value = eu_member  # EU_Member
    ws.cell(new_row, 5).value = SUBREGIONS.get(code, '')  # Subregion

print(f"  Created {ws_old.max_row - 1} country records with State Dept regions")

# ========== Copy remaining sheets without tables ==========
sheets_to_copy = [
    'Country_Dashboard', 'Portfolio_Dashboard', 'Milestones',
    'Spotlight_Executive', 'Spotlight_PMWorkspace', 'Spotlight_Stakeholder',
    'Events', 'Risk_Register', 'Decision_Log', 'Stakeholders',
    'Config_Lists', 'Regional_Summary', '_SETUP', 'Calendar_Todo',
    'Project_Deliverables', 'Project_Audiences', 'Project_Products',
    'Country_PM_Assignments', 'Project_Technologies'
]

print("\n5. Copying remaining sheets...")
for sheet_name in sheets_to_copy:
    if sheet_name in wb_v12.sheetnames:
        print(f"  Copying {sheet_name}...")
        ws_old = wb_v12[sheet_name]
        ws_new = wb.create_sheet(sheet_name)

        # Copy all cells
        for row in ws_old.iter_rows():
            for cell in row:
                new_cell = ws_new[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()

        # Copy column widths
        for col_letter in ws_old.column_dimensions:
            if col_letter in ws_old.column_dimensions:
                ws_new.column_dimensions[col_letter].width = ws_old.column_dimensions[col_letter].width

        # Copy merged cells
        for merged_range in ws_old.merged_cells.ranges:
            ws_new.merge_cells(str(merged_range))

# ========== Update Config_Lists with countries ==========
print("\n6. Adding countries to Config_Lists...")
ws_config = wb['Config_Lists']
ws_regions = wb['Country_Regions']

# Find next empty row
next_row = ws_config.max_row + 1

# Add all countries
country_count = 0
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value

    if code and name:
        ws_config.cell(next_row, 1).value = "Country"
        ws_config.cell(next_row, 2).value = code
        ws_config.cell(next_row, 3).value = name
        ws_config.cell(next_row, 4).value = region
        next_row += 1
        country_count += 1

print(f"  Added {country_count} countries to Config_Lists")

# ========== Update Country_PM_Assignments ==========
print("\n7. Updating Country_PM_Assignments...")
ws_pm = wb['Country_PM_Assignments']

# Clear existing data except header
for row in range(2, ws_pm.max_row + 1):
    for col in range(1, 8):
        ws_pm.cell(row, col).value = None

# Add all countries from Country_Regions
current_row = 2
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value

    if code:
        ws_pm.cell(current_row, 1).value = code
        ws_pm.cell(current_row, 2).value = name
        ws_pm.cell(current_row, 3).value = region
        ws_pm.cell(current_row, 4).value = "TBD"
        current_row += 1

print(f"  Updated {current_row - 2} countries in Country_PM_Assignments")

# ========== Update Spotlight sheets ==========
print("\n8. Updating Spotlight_PMWorkspace financial section...")
ws_spotlight = wb['Spotlight_PMWorkspace']

# Update financial section (around row 20-22)
ws_spotlight.cell(21, 1).value = "Total Proposed:"
ws_spotlight.cell(21, 2).value = "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$F:$F)"
ws_spotlight.cell(21, 4).value = "Total Allocated:"
ws_spotlight.cell(21, 5).value = "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$G:$G)"
ws_spotlight.cell(21, 7).value = "Total Obligated:"
ws_spotlight.cell(21, 8).value = "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$H:$H)"

ws_spotlight.cell(22, 1).value = "Total Spent:"
ws_spotlight.cell(22, 2).value = "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$I:$I)"
ws_spotlight.cell(22, 4).value = "ULO:"
ws_spotlight.cell(22, 5).value = "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$J:$J)"
ws_spotlight.cell(22, 7).value = "ULO %:"
ws_spotlight.cell(22, 8).value = "=IF(H21>0,E22/H21,0)"
ws_spotlight.cell(22, 8).number_format = '0%'

print("  Updated PMWorkspace")

# ========== Save ==========
print("\nSaving v15...")
wb.save('2025-10-26-Tracker-v15.xlsx')

print("\n" + "="*80)
print("V15 BUILD COMPLETE!")
print("="*80)

print("\nNew structure:")
print("  Country_Budgets: Proposed | Allocated | Obligated | Spent | ULO | ULO%")
print("  Master_Projects: Total_Proposed | Total_Allocation | Total_Obligated | Total_ULO | ULO%")
print("  Regions: State Dept codes (AF, EAP, EUR, NEA, SCA, WHA)")
print("  Subregions: Added for all countries")
print("  Config_Lists: Contains all countries")
print("  Country_PM_Assignments: Synced with Country_Regions")
print("\nFile: 2025-10-26-Tracker-v15.xlsx")
print("\nThis version has NO Excel Tables - should open cleanly!")
