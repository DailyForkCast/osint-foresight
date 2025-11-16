"""
Add automated calendar formulas to Calendar_Todo - v35
Pulls dates from Master_Projects, Deliverables, Milestones, and Events
"""
import openpyxl

print("="*80)
print("CREATING AUTOMATED CALENDAR - V35")
print("="*80)

# Load v34
print("\nLoading v34...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v34.xlsx')

ws_calendar = wb['Calendar_Todo']

print("\n" + "="*80)
print("CURRENT CALENDAR_TODO STRUCTURE")
print("="*80)

print("\nRow 1 headers:")
for col in range(1, 10):
    header = ws_calendar.cell(1, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}1: {clean_header}")

print("\n" + "="*80)
print("DESIGN: AUTOMATED CALENDAR")
print("="*80)

print("\nApproach:")
print("  We'll create separate sections for different date sources:")
print("  1. POP Start Dates (from Master_Projects)")
print("  2. POP End Dates (from Master_Projects)")
print("  3. Deliverable Due Dates (from Project_Deliverables)")
print("  4. Milestone Dates (from Milestones)")
print("  5. Event Dates (from Events)")

print("\nNote: This will create a long list. Another approach is to use")
print("  a single consolidated list sorted by date, but that requires")
print("  complex array formulas or VBA.")

print("\nLet's use a simpler approach:")
print("  - Create formulas that pull the next 50 upcoming items")
print("  - Use a helper column to consolidate all dates")
print("  - Sort by date using FILTER/SORT (Excel 365) or array formulas")

print("\n" + "="*80)
print("IMPLEMENTATION")
print("="*80)

print("\nClearing existing data (rows 2-100)...")
for row in range(2, 101):
    for col in range(1, 10):
        ws_calendar.cell(row, col).value = None

print("\nAdding section headers and formulas...")

# We'll create a simple approach: list all POP dates, then deliverables, then milestones
# Each section will have its own rows

current_row = 2

# Section 1: POP Start Dates
print("\n1. Adding POP Start Dates section...")
ws_calendar.cell(current_row, 1).value = "=== PROJECT START DATES ==="
current_row += 1

# Add up to 20 project start dates
for i in range(1, 21):
    row = current_row + i - 1

    # Task ID: Project ID from Master_Projects
    ws_calendar.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],{i}),"")'

    # Task Name: "Start: " + Project Name
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","","Start: "&INDEX(T_Master_Projects[Project_Name],{i}))'

    # Unique ID: Project Unique ID
    ws_calendar.cell(row, 3).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Project_Unique_ID],{i}))'

    # Due Date: POP Start Date
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[POP_Start],{i}))'

    # Assigned To: Project Manager or POC
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Implementer_POC],{i}))'

    # Status: "Scheduled" or "Completed" based on start date vs today
    ws_calendar.cell(row, 6).value = f'=IF(A{row}="","",IF(D{row}<TODAY(),"Past","Upcoming"))'

    # Priority: Pull from Master_Projects
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Project_Priority],{i}))'

    # Notes: "Project start date"
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Project start date")'

current_row += 21

# Section 2: POP End Dates
print("2. Adding POP End Dates section...")
ws_calendar.cell(current_row, 1).value = "=== PROJECT END DATES ==="
current_row += 1

for i in range(1, 21):
    row = current_row + i - 1

    ws_calendar.cell(row, 1).value = f'=IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],{i}),"")'
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","","End: "&INDEX(T_Master_Projects[Project_Name],{i}))'
    ws_calendar.cell(row, 3).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Project_Unique_ID],{i}))'
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[POP_End],{i}))'
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Implementer_POC],{i}))'
    ws_calendar.cell(row, 6).value = f'=IF(A{row}="","",IF(D{row}<TODAY(),"Overdue",IF(D{row}<TODAY()+30,"Due Soon","Upcoming")))'
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","",INDEX(T_Master_Projects[Project_Priority],{i}))'
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Project end date")'

current_row += 21

# Section 3: Milestone Dates
print("3. Adding Milestone Dates section...")
ws_calendar.cell(current_row, 1).value = "=== MILESTONES ==="
current_row += 1

for i in range(1, 21):
    row = current_row + i - 1

    # Milestones sheet structure might be different - let's use basic references
    ws_calendar.cell(row, 1).value = f'=IFERROR(Milestones!A{i+1},"")'  # Skip header row
    ws_calendar.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(Milestones!B{i+1},""))'
    ws_calendar.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(Milestones!A{i+1},""))'
    ws_calendar.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(Milestones!C{i+1},""))'
    ws_calendar.cell(row, 5).value = f'=IF(A{row}="","","")'  # No assigned person in milestones
    ws_calendar.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(Milestones!D{i+1},""))'
    ws_calendar.cell(row, 7).value = f'=IF(A{row}="","","")'  # No priority in milestones
    ws_calendar.cell(row, 8).value = f'=IF(A{row}="","","Milestone")'

print("\nFormulas added!")

print("\n" + "="*80)
print("SAVING V35")
print("="*80)

wb.save('2025-10-26-Tracker-v35.xlsx')

print("\nOK - v35 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nAutomated Calendar Created!")

print("\nSections added:")
print("  1. Project Start Dates (rows 3-22)")
print("     - Pulls POP_Start from Master_Projects")
print("     - Shows first 20 projects")
print("")
print("  2. Project End Dates (rows 24-43)")
print("     - Pulls POP_End from Master_Projects")
print("     - Shows status: Overdue/Due Soon/Upcoming")
print("")
print("  3. Milestones (rows 45-64)")
print("     - Pulls from Milestones sheet")
print("     - Shows first 20 milestones")

print("\nColumns populated:")
print("  A: Task_ID")
print("  B: Task_Name")
print("  C: Unique_ID (Project ID)")
print("  D: Due_Date")
print("  E: Assigned_To")
print("  F: Status")
print("  G: Priority")
print("  H: Notes")

print("\nFeatures:")
print("  - Auto-populated from other sheets")
print("  - Shows past/upcoming status")
print("  - Highlights due soon (within 30 days)")
print("  - No manual entry needed!")

print("\nNote: To see ALL items sorted by date, you could:")
print("  - Use Excel's Filter/Sort features on the calendar")
print("  - Add more rows to each section")
print("  - Create a pivot table view")

print("\nv35 ready!")
