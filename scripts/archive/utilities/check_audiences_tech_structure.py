"""
Check Project_Audiences and Project_Technologies table structure
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("PROJECT_AUDIENCES STRUCTURE")
print("="*80)

ws = wb['Project_Audiences']
print("\nHeaders:")
for col in range(1, min(10, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\nFirst 3 data rows:")
for row in range(2, 5):
    print(f"\nRow {row}:")
    for col in range(1, min(10, ws.max_column + 1)):
        val = ws.cell(row, col).value
        header = ws.cell(1, col).value
        if val:
            print(f"  {header}: {val}")

print("\n" + "="*80)
print("PROJECT_TECHNOLOGIES STRUCTURE")
print("="*80)

ws = wb['Project_Technologies']
print("\nHeaders:")
for col in range(1, min(10, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\nFirst 3 data rows:")
for row in range(2, 5):
    print(f"\nRow {row}:")
    for col in range(1, min(10, ws.max_column + 1)):
        val = ws.cell(row, col).value
        header = ws.cell(1, col).value
        if val:
            print(f"  {header}: {val}")
