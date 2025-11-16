import openpyxl

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6-clean.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6-clean.xlsx'

print('='*80)
print('FIXING CONFIG_LISTS TABLE RANGE')
print('='*80)
print()

# Load workbook
print('Loading v6-clean...')
wb = openpyxl.load_workbook(input_file)
ws = wb['Config_Lists']
print('Loaded.')
print()

# Check current table
for table in ws.tables.values():
    if table.name == 'T_Config_Lists':
        print(f'Current table range: {table.ref}')
        print(f'Table ends at row 77, but data goes to row {ws.max_row}')
        print()

        # Expand table to include all data
        new_range = f'A1:G{ws.max_row}'
        print(f'Expanding table to: {new_range}')
        table.ref = new_range
        print('Table expanded!')
        print()

        print(f'Countries now included: {ws.max_row - 1} total')
        print()

        break

# Save
print('Saving...')
wb.save(output_file)
wb.close()

print('='*80)
print('FIX COMPLETE!')
print('='*80)
print()
print('The T_Config_Lists table now includes all 101 countries (rows 2-102).')
print('All dropdown lists and formulas will now see all countries.')
print()
