"""
Check current Master_Projects column headers
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("MASTER_PROJECTS - CURRENT COLUMN HEADERS")
print("="*80)

ws = wb['Master_Projects']

print("\nAll headers:")
for col in range(1, min(25, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\n" + "="*80)
print("WHAT NEEDS TO CHANGE")
print("="*80)

print("\nUser says:")
print("  D should be: Project_Summary")
print("  E should be: Project_Status")
print("  F should be: Project_Priority")
print("  Need to add: Project_Progress")
print("\nCurrent:")
print(f"  D: {ws.cell(1, 4).value}")
print(f"  E: {ws.cell(1, 5).value}")
print(f"  F: {ws.cell(1, 6).value}")
