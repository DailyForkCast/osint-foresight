"""
Check Project_Deliverables and Stakeholders table structure
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("PROJECT_DELIVERABLES STRUCTURE")
print("="*80)

ws = wb['Project_Deliverables']
print("\nHeaders:")
for col in range(1, min(10, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\nFirst 3 data rows:")
for row in range(2, 5):
    print(f"\nRow {row}:")
    for col in range(1, min(10, ws.max_column + 1)):
        val = ws.cell(row, col).value
        header = ws.cell(1, col).value
        if val:
            print(f"  {header}: {val}")

print("\n" + "="*80)
print("STAKEHOLDERS STRUCTURE")
print("="*80)

ws = wb['Stakeholders']
print("\nHeaders:")
for col in range(1, min(25, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    if header:
        print(f"  {col_letter}: {header}")

print("\nFirst 3 data rows:")
for row in range(2, 5):
    print(f"\nRow {row}:")
    for col in range(1, min(10, ws.max_column + 1)):
        val = ws.cell(row, col).value
        header = ws.cell(1, col).value
        if val:
            print(f"  {header}: {val}")

# Check Spotlight layout for these sections
print("\n" + "="*80)
print("SPOTLIGHT DELIVERABLES SECTION")
print("="*80)

ws = wb['Spotlight_PMWorkspace']
print("\nRow 18 (Deliverable headers):")
for col in range(1, 10):
    val = ws.cell(18, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {val}")

print("\n" + "="*80)
print("SPOTLIGHT STAKEHOLDERS SECTION")
print("="*80)

# Find stakeholders section
for row in range(35, 50):
    val = ws.cell(row, 1).value
    if val and "STAKEHOLDER" in str(val).upper():
        print(f"\nStakeholders section starts at row {row}")
        print(f"Row {row+1} headers:")
        for col in range(1, 10):
            val = ws.cell(row+1, col).value
            if val:
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"  {col_letter}: {val}")
        break
