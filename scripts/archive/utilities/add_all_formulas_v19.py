"""
Create v19 with all formulas added to current structure
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

print("="*80)
print("CREATING V19 WITH ALL FORMULAS")
print("="*80)

# Load current v18
print("\nLoading v18...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

# Get Master_Projects sheet
ws_mp = wb['T_Master_Projects'] if 'T_Master_Projects' in wb.sheetnames else wb['Master_Projects']

print("Adding formulas to Master_Projects...")

# Get the data rows (skip header)
max_row = ws_mp.max_row
print(f"  Processing {max_row - 1} project rows...")

# Add formulas to each project row
for row in range(2, max_row + 1):
    # J: Days_Remaining (based on Project_End_Date in column I)
    if not ws_mp.cell(row, 10).value or str(ws_mp.cell(row, 10).value).startswith('='):
        ws_mp.cell(row, 10).value = f'=IF(OR(I{row}="",E{row}=""),"",IF(E{row}="Completed","Complete",INT(I{row}-TODAY())))'

    # R: POP_Days_Remaining (based on POP dates with NCE logic)
    if not ws_mp.cell(row, 18).value or str(ws_mp.cell(row, 18).value).startswith('='):
        ws_mp.cell(row, 18).value = f'=IF(OR(M{row}="",E{row}=""),"",IF(E{row}="Completed","Complete",IF(Q{row}<>"",INT(Q{row}-TODAY()),IF(P{row}<>"",INT(P{row}-TODAY()),INT(M{row}-TODAY())))))'

print("  OK - Days_Remaining and POP_Days_Remaining formulas added")

# Verify other key formulas exist
print("\nVerifying key formulas...")

# Check if Total_Spent exists (column Z)
if ws_mp.cell(2, 26).value is None or not str(ws_mp.cell(2, 26).value).startswith('='):
    print("  Adding Total_Spent formula (column Z)...")
    for row in range(2, max_row + 1):
        # Using table syntax if available
        ws_mp.cell(row, 26).value = f'=SUMIF(T_Country_Budgets[Unique_ID],[@[Project_Unique_ID]],T_Country_Budgets[Spent_Amount])'

# Check if Total_ULO exists (column AA)
if ws_mp.cell(2, 27).value is None or not str(ws_mp.cell(2, 27).value).startswith('='):
    print("  Adding Total_ULO formula (column AA)...")
    for row in range(2, max_row + 1):
        ws_mp.cell(row, 27).value = f'=SUMIF(T_Country_Budgets[Unique_ID],[@[Project_Unique_ID]],T_Country_Budgets[ULO])'

# Check if ULO_Percent exists (column AB)
if ws_mp.cell(2, 28).value is None or not str(ws_mp.cell(2, 28).value).startswith('='):
    print("  Adding ULO_Percent formula (column AB)...")
    for row in range(2, max_row + 1):
        ws_mp.cell(row, 28).value = f'=IF(OR([@[Project_Unique_ID]]="",[@[Total_Obligated]]=""),"",IF([@[Total_Obligated]]>0,[@[Total_ULO]]/[@[Total_Obligated]],0))'
        ws_mp.cell(row, 28).number_format = '0%'

# Check if Countries exists (column AC)
if ws_mp.cell(2, 29).value is None or not str(ws_mp.cell(2, 29).value).startswith('='):
    print("  Adding Countries formula (column AC)...")
    for row in range(2, max_row + 1):
        ws_mp.cell(row, 29).value = f'=IF([@[Project_Unique_ID]]="","",TEXTJOIN(", ",TRUE,IF(T_Country_Budgets[Unique_ID]=[@[Project_Unique_ID]],T_Country_Budgets[Country_Name],"")))'

# Check if Country_Count exists (column AD)
if ws_mp.cell(2, 30).value is None or not str(ws_mp.cell(2, 30).value).startswith('='):
    print("  Adding Country_Count formula (column AD)...")
    for row in range(2, max_row + 1):
        ws_mp.cell(row, 30).value = f'=IF([@[Project_Unique_ID]]="","",COUNTIF(T_Country_Budgets[Unique_ID],[@[Project_Unique_ID]]))'

print("\n" + "="*80)
print("Checking Spotlight_PMWorkspace formulas...")
print("="*80)

ws_spot = wb['Spotlight_PMWorkspace']

# Add/verify Spotlight formulas
spotlight_formulas = {
    # Project Details
    'B5': '=IFERROR(INDEX(T_Master_Projects[Project_Name],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B6': '=IFERROR(INDEX(T_Master_Projects[Project_Priority],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B7': '=IFERROR(INDEX(T_Master_Projects[Project_Start_Date],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B8': '=IFERROR(INDEX(T_Master_Projects[Days_Remaining],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B9': '=IFERROR(INDEX(T_Master_Projects[Implementer_POC],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B10': '=IFERROR(INDEX(T_Master_Projects[Countries],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B11': '=IFERROR(INDEX(T_Master_Projects[Project_Summary],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',

    # Right side
    'G5': '=IFERROR(INDEX(T_Master_Projects[Project_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'G6': '=IFERROR(INDEX(T_Master_Projects[Project_Progress],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'G7': '=IFERROR(INDEX(T_Master_Projects[Project_End_Date],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'G8': '=IFERROR(INDEX(T_Master_Projects[Implementer],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'G9': '=IFERROR(INDEX(T_Master_Projects[Implementer_POC_Email],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'G10': '=IFERROR(INDEX(T_Master_Projects[Country_Count],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',

    # Audiences & Technologies (row 12)
    'B12': '=IFERROR(TEXTJOIN(", ",TRUE,IF(T_Project_Audiences[Project_ID]=$B$2,T_Project_Audiences[Audience_Type],"")),"?")',
    'G12': '=IFERROR(TEXTJOIN(", ",TRUE,IF(T_Project_Technologies[Project_ID]=$B$2,T_Project_Technologies[Technology],"")),"?")',

    # Financial Details
    'B14': '=IFERROR(INDEX(T_Master_Projects[Total_Proposed],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'E14': '=IFERROR(INDEX(T_Master_Projects[Total_Obligated],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'H14': '=IFERROR(INDEX(T_Master_Projects[Total_ULO],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'B15': '=IFERROR(INDEX(T_Master_Projects[Total_Allocation],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'E15': '=IFERROR(INDEX(T_Master_Projects[Total_Spent],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
    'H15': '=IFERROR(INDEX(T_Master_Projects[ULO_Percent],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")',
}

print("\nAdding Spotlight formulas...")
for cell_ref, formula in spotlight_formulas.items():
    cell = ws_spot[cell_ref]
    if not cell.value or not str(cell.value).startswith('='):
        cell.value = formula
        print(f"  OK - {cell_ref}")

# Add Deliverables formulas (rows 19-29, auto-increment with ROW())
print("\nAdding Deliverables formulas (rows 19-29)...")
for row in range(19, 30):
    # A: Deliverable Name
    ws_spot.cell(row, 1).value = f'=IFERROR(INDEX(T_Project_Deliverables[Deliverable_Name],AGGREGATE(15,6,ROW(T_Project_Deliverables[Project_ID])-ROW(INDEX(T_Project_Deliverables[Project_ID],1,1))+1/(T_Project_Deliverables[Project_ID]=$B$2),ROW()-18)),"")'

    # E: Due Date
    ws_spot.cell(row, 5).value = f'=IFERROR(INDEX(T_Project_Deliverables[Due_Date],AGGREGATE(15,6,ROW(T_Project_Deliverables[Project_ID])-ROW(INDEX(T_Project_Deliverables[Project_ID],1,1))+1/(T_Project_Deliverables[Project_ID]=$B$2),ROW()-18)),"")'

    # G: Status
    ws_spot.cell(row, 7).value = f'=IFERROR(INDEX(T_Project_Deliverables[Status],AGGREGATE(15,6,ROW(T_Project_Deliverables[Project_ID])-ROW(INDEX(T_Project_Deliverables[Project_ID],1,1))+1/(T_Project_Deliverables[Project_ID]=$B$2),ROW()-18)),"")'

    # H: Owner
    ws_spot.cell(row, 8).value = f'=IFERROR(INDEX(T_Project_Deliverables[Owner],AGGREGATE(15,6,ROW(T_Project_Deliverables[Project_ID])-ROW(INDEX(T_Project_Deliverables[Project_ID],1,1))+1/(T_Project_Deliverables[Project_ID]=$B$2),ROW()-18)),"")'

    # I: Complete %
    ws_spot.cell(row, 9).value = f'=IFERROR(INDEX(T_Project_Deliverables[Completion_Percent],AGGREGATE(15,6,ROW(T_Project_Deliverables[Project_ID])-ROW(INDEX(T_Project_Deliverables[Project_ID],1,1))+1/(T_Project_Deliverables[Project_ID]=$B$2),ROW()-18))/100,"")'

print("  OK - Deliverables formulas added")

# Add Documents formulas (rows 32-42, auto-increment with ROW())
print("\nAdding Documents formulas (rows 32-42)...")
for row in range(32, 43):
    # A: Document Name
    ws_spot.cell(row, 1).value = f'=IFERROR(INDEX(T_Project_Documents[Document_Name],AGGREGATE(15,6,ROW(T_Project_Documents[Project_Unique_ID])-ROW(INDEX(T_Project_Documents[Project_Unique_ID],1,1))+1/(T_Project_Documents[Project_Unique_ID]=$B$2),ROW()-31)),"")'

    # C: Type
    ws_spot.cell(row, 3).value = f'=IFERROR(INDEX(T_Project_Documents[Document_Type],AGGREGATE(15,6,ROW(T_Project_Documents[Project_Unique_ID])-ROW(INDEX(T_Project_Documents[Project_Unique_ID],1,1))+1/(T_Project_Documents[Project_Unique_ID]=$B$2),ROW()-31)),"")'

    # E: Link
    ws_spot.cell(row, 5).value = f'=IFERROR(INDEX(T_Project_Documents[Document_Link],AGGREGATE(15,6,ROW(T_Project_Documents[Project_Unique_ID])-ROW(INDEX(T_Project_Documents[Project_Unique_ID],1,1))+1/(T_Project_Documents[Project_Unique_ID]=$B$2),ROW()-31)),"")'

    # G: Date Added
    ws_spot.cell(row, 7).value = f'=IFERROR(INDEX(T_Project_Documents[Date_Added],AGGREGATE(15,6,ROW(T_Project_Documents[Project_Unique_ID])-ROW(INDEX(T_Project_Documents[Project_Unique_ID],1,1))+1/(T_Project_Documents[Project_Unique_ID]=$B$2),ROW()-31)),"")'

print("  OK - Documents formulas added")

# Add Stakeholders formulas (rows 45+, auto-increment with ROW())
print("\nAdding Stakeholders formulas (rows 45-54)...")
for row in range(45, 55):
    # A: Name
    ws_spot.cell(row, 1).value = f'=IFERROR(INDEX(T_Stakeholders[Name],AGGREGATE(15,6,ROW(T_Stakeholders[Project_IDs])-ROW(INDEX(T_Stakeholders[Project_IDs],1,1))+1/(ISNUMBER(SEARCH($B$2,T_Stakeholders[Project_IDs]))),ROW()-44)),"")'

    # C: Title
    ws_spot.cell(row, 3).value = f'=IFERROR(INDEX(T_Stakeholders[Title],AGGREGATE(15,6,ROW(T_Stakeholders[Project_IDs])-ROW(INDEX(T_Stakeholders[Project_IDs],1,1))+1/(ISNUMBER(SEARCH($B$2,T_Stakeholders[Project_IDs]))),ROW()-44)),"")'

    # E: Organization
    ws_spot.cell(row, 5).value = f'=IFERROR(INDEX(T_Stakeholders[Organization],AGGREGATE(15,6,ROW(T_Stakeholders[Project_IDs])-ROW(INDEX(T_Stakeholders[Project_IDs],1,1))+1/(ISNUMBER(SEARCH($B$2,T_Stakeholders[Project_IDs]))),ROW()-44)),"")'

    # G: Email
    ws_spot.cell(row, 7).value = f'=IFERROR(INDEX(T_Stakeholders[Email],AGGREGATE(15,6,ROW(T_Stakeholders[Project_IDs])-ROW(INDEX(T_Stakeholders[Project_IDs],1,1))+1/(ISNUMBER(SEARCH($B$2,T_Stakeholders[Project_IDs]))),ROW()-44)),"")'

print("  OK - Stakeholders formulas added")

# Save as v19
print("\n" + "="*80)
print("Saving as v19...")
print("="*80)

wb.save('2025-10-26-Tracker-v19.xlsx')

print("\nOK - V19 CREATED SUCCESSFULLY!")
print("\nFile: 2025-10-26-Tracker-v19.xlsx")
print("\nFormulas added:")
print("  OK - Master_Projects: Days_Remaining, POP_Days_Remaining")
print("  OK - Master_Projects: All budget rollup formulas")
print("  OK - Master_Projects: Countries, Country_Count")
print("  OK - Spotlight: All project details")
print("  OK - Spotlight: Financial details")
print("  OK - Spotlight: Audiences & Technologies")
print("  OK - Spotlight: Deliverables (11 rows)")
print("  OK - Spotlight: Documents (11 rows)")
print("  OK - Spotlight: Stakeholders (10 rows)")
print("\nNext step: Add data validation dropdowns for NCE_Eligible and NCE_Status")
