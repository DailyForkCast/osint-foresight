"""
Check v18 Country_Regions sheet to see what countries are already there
"""
import openpyxl

print("="*80)
print("CHECKING V18 COUNTRY_REGIONS DATA")
print("="*80)

# Load v18
print("\nLoading v18...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

ws_regions = wb['Country_Regions']

print("\nReading Country_Regions sheet data...")

# Get headers
headers = []
for col in range(1, 10):
    cell = ws_regions.cell(1, col)
    if cell.value:
        headers.append(cell.value)
    else:
        break

print(f"\nHeaders found: {', '.join(headers)}")

# Read all data rows
countries_data = []
row = 2
while True:
    # Check if row is empty
    first_cell = ws_regions.cell(row, 1).value
    if first_cell is None:
        break

    row_data = []
    for col in range(1, len(headers) + 1):
        row_data.append(ws_regions.cell(row, col).value)

    countries_data.append(row_data)
    row += 1

print(f"\nTotal countries found: {len(countries_data)}")

if countries_data:
    print("\nCountry data:")
    print("-" * 80)

    # Print header row
    header_line = " | ".join([f"{h:20s}" for h in headers])
    print(header_line)
    print("-" * 80)

    # Print data rows
    for row_data in countries_data:
        formatted_row = []
        for val in row_data:
            if val is None:
                formatted_row.append(" " * 20)
            else:
                formatted_row.append(f"{str(val):20s}")
        print(" | ".join(formatted_row))

# Save to file for easy reference
print("\n" + "="*80)
print("SAVING TO FILE")
print("="*80)

with open('v18_country_regions_data.txt', 'w') as f:
    f.write("COUNTRY_REGIONS DATA FROM V18\n")
    f.write("="*80 + "\n\n")

    f.write("Headers: " + ", ".join(headers) + "\n\n")

    if countries_data:
        # Header row
        f.write(" | ".join([f"{h:20s}" for h in headers]) + "\n")
        f.write("-" * 80 + "\n")

        # Data rows
        for row_data in countries_data:
            formatted_row = []
            for val in row_data:
                if val is None:
                    formatted_row.append(" " * 20)
                else:
                    formatted_row.append(f"{str(val):20s}")
            f.write(" | ".join(formatted_row) + "\n")
    else:
        f.write("No data found\n")

print("\nSaved to: v18_country_regions_data.txt")

print("\n" + "="*80)
print("DONE")
print("="*80)
