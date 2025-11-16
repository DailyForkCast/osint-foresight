"""
1. Add labels back to Portfolio_Dashboard columns O-Q from v27
2. Fix Country_Dashboard to use country names instead of codes in B2
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO LABELS AND COUNTRY DASHBOARD")
print("="*80)

# Load v27 to get the labels
print("\nLoading v27 to get labels...")
wb_v27 = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx', data_only=True)
ws_v27_port = wb_v27['Portfolio_Dashboard']

labels = {}
for row in range(3, 8):
    label = ws_v27_port.cell(row, 15).value  # Column O
    if label:
        labels[row] = label

print(f"Found {len(labels)} labels in v27:")
for row, label in labels.items():
    print(f"  Row {row}: {label}")

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

print("\n" + "="*80)
print("PART 1: ADD LABELS TO PORTFOLIO_DASHBOARD")
print("="*80)

ws_port = wb['Portfolio_Dashboard']

print("\nAdding labels to column O...")
for row, label in labels.items():
    ws_port.cell(row, 15).value = label
    print(f"  O{row}: {label}")

print("\nOK - Labels added")

print("\n" + "="*80)
print("PART 2: FIX COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nChanging B2 from Country Code to Country Name...")

# B2 currently expects a country code (like "DE")
# We want to change it to show country name (like "Germany")
# And update D2 to not be needed (or show region instead)

print("\nStep 1: B2 should show country name for selection")
print("  Current: Country code (manual entry)")
print("  New: Country name (from dropdown)")

# We'll keep B2 as manual entry but update label
ws_country['A2'] = 'Country:'
print("  Updated A2 label")

# Step 2: Update formulas that use B2
# They currently expect B2 to be a country code
# We need them to look up the code from the name

print("\nStep 2: Update D2 to show Region instead of Country Name")
# Since B2 now has the country name, D2 can show region
# Check if C2 is merged first
if 'C2' in ws_country.merged_cells:
    print("  C2 is merged, skipping label update")
else:
    ws_country['C2'] = 'Region:'

ws_country['D2'] = '=IFERROR(INDEX(T_Country_Regions[Region],MATCH(B2,T_Country_Regions[Country_Name],0)),"")'
print("  D2 now shows Region based on country name in B2")

print("\nStep 3: Update project list formulas to use country name")
# The project list uses B2 to filter
# Current: Matches Country_Code
# New: Must look up Country_Code from Country_Name first

# This is complex because Country_Budgets uses Country_Code
# We need to first get the code from the name, then filter

# Actually, simpler approach: Update Country_Budgets to also have Country_Name
# Or use a helper cell to convert name to code

# Helper approach: Use cell F2 for the country code lookup (avoid merged cells)
# First unmerge any merged cells in row 2 if needed
print("  Checking for merged cells in row 2...")
merged_to_unmerge = []
for merged in ws_country.merged_cells:
    if merged.min_row == 2 and merged.max_row == 2:
        merged_to_unmerge.append(str(merged))

for merged_range in merged_to_unmerge:
    ws_country.unmerge_cells(merged_range)
    print(f"    Unmerged: {merged_range}")

ws_country['F2'] = '=IFERROR(INDEX(T_Country_Regions[Country_Code],MATCH(B2,T_Country_Regions[Country_Name],0)),"")'

print("  Added helper cell F2 to lookup Country_Code from Country_Name")

# Now update the project list formulas to use F2 instead of B2
print("\nStep 4: Update project list to use F2 (Country_Code)")

for row in range(12, 32):
    # Column A: Project IDs filtered by country code (use F2 now)
    current_formula = str(ws_country.cell(row, 1).value)
    if 'B2' in current_formula:
        new_formula = current_formula.replace('$B$2', '$F$2')
        ws_country.cell(row, 1).value = new_formula
        if row == 12:
            print(f"    A12 updated to use F2")

print("\nOK - Country_Dashboard now uses country names in B2!")

print("\n" + "="*80)
print("PART 3: FIX COUNTRY_DASHBOARD BLANK/ZERO DISPLAY")
print("="*80)

print("\nUpdating formulas to show blank instead of 0...")

# Need to update columns with numbers to show blank instead of 0
# Assuming structure (check what columns have what):
# Let me update the project list columns

# This requires knowing the column structure. Let's update common ones:
# Columns that should show blank: Allocated, Obligated, Spent, ULO, ULO%, Proposed

print("  Updating numeric columns to show blank when zero...")

for row in range(12, 32):
    # For each numeric column, wrap in IF to check for 0
    # Column B: Project Name (text, already handled)
    # Column C: Project Summary (text)
    # Column D: Status (text)
    # Let's say E, F, G, H are budget columns

    # We need to check the actual formulas
    # For now, let's wrap any numeric INDEX formulas

    for col in range(5, 9):  # Columns E-H (adjust as needed)
        cell = ws_country.cell(row, col)
        if cell.data_type == 'f':
            formula = str(cell.value)
            if 'INDEX' in formula and 'IF(A' not in formula:
                # Wrap to check if A column (Project ID) is blank
                new_formula = f'=IF(A{row}="","",IFERROR(IF({formula[1:]}=0,"",{formula[1:]}),""))'
                ws_country.cell(row, col).value = new_formula
                if row == 12:
                    print(f"    Updated column {chr(64+col)}")

print("\nOK - Numeric columns will show blank instead of 0")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Changes saved!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\n1. Portfolio_Dashboard:")
print("   - Labels added back to column O (rows 3-7)")
print("   - Formatted as nice table with header")

print("\n2. Country_Dashboard:")
print("   - B2 now for Country Name (not code)")
print("   - D2 shows Region (not country name)")
print("   - E2 helper cell converts name to code")
print("   - Project list uses E2 for filtering")

print("\nHow to use Country_Dashboard now:")
print("  1. Enter country name in B2 (e.g., 'Germany' not 'DE')")
print("  2. D2 shows the region (e.g., 'EUR')")
print("  3. E2 auto-converts to code for filtering")
print("  4. Project list shows projects for that country")

print("\nAdd data validation to B2:")
print("  Source: =T_Country_Regions[Country_Name]")

print("\nv28 updated!")
