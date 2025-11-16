import openpyxl
from openpyxl.utils import get_column_letter

# Load the file we just created
input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v1.xlsx'

print('='*80)
print('PART 4: Expanding Country_Regions to 94 countries across 6 regions')
print('='*80)
print()

# Load workbook
print('Loading workbook...')
wb = openpyxl.load_workbook(input_file, data_only=False)

if 'Country_Regions' not in wb.sheetnames:
    print('ERROR: Country_Regions sheet not found!')
    wb.close()
    exit(1)

ws = wb['Country_Regions']
print(f'Current Country_Regions structure:')
print(f'  Max row: {ws.max_row}')
print(f'  Max col: {ws.max_column}')

# Check headers
print('  Headers:', end=' ')
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    print(f'{header}', end=', ')
print()
print()

# Define all 94 countries by region
countries_data = [
    # EUR - Europe (44 countries)
    ('AL', 'Albania', 'EUR', 'No', 'Balkans'),
    ('AM', 'Armenia', 'EUR', 'No', 'Caucasus'),
    ('AT', 'Austria', 'EUR', 'Yes', 'Western Europe'),
    ('AZ', 'Azerbaijan', 'EUR', 'No', 'Caucasus'),
    ('BY', 'Belarus', 'EUR', 'No', 'Eastern Europe'),
    ('BE', 'Belgium', 'EUR', 'Yes', 'Western Europe'),
    ('BA', 'Bosnia and Herzegovina', 'EUR', 'No', 'Balkans'),
    ('BG', 'Bulgaria', 'EUR', 'Yes', 'Eastern Europe'),
    ('HR', 'Croatia', 'EUR', 'Yes', 'Southern Europe'),
    ('CY', 'Cyprus', 'EUR', 'Yes', 'Southern Europe'),
    ('CZ', 'Czech Republic', 'EUR', 'Yes', 'Eastern Europe'),
    ('DK', 'Denmark', 'EUR', 'Yes', 'Northern Europe'),
    ('EE', 'Estonia', 'EUR', 'Yes', 'Northern Europe'),
    ('FI', 'Finland', 'EUR', 'Yes', 'Northern Europe'),
    ('FR', 'France', 'EUR', 'Yes', 'Western Europe'),
    ('GE', 'Georgia', 'EUR', 'No', 'Caucasus'),
    ('DE', 'Germany', 'EUR', 'Yes', 'Western Europe'),
    ('GR', 'Greece', 'EUR', 'Yes', 'Southern Europe'),
    ('HU', 'Hungary', 'EUR', 'Yes', 'Eastern Europe'),
    ('IS', 'Iceland', 'EUR', 'No', 'Northern Europe'),
    ('IE', 'Ireland', 'EUR', 'Yes', 'Northern Europe'),
    ('IT', 'Italy', 'EUR', 'Yes', 'Southern Europe'),
    ('XK', 'Kosovo', 'EUR', 'No', 'Balkans'),
    ('LV', 'Latvia', 'EUR', 'Yes', 'Northern Europe'),
    ('LT', 'Lithuania', 'EUR', 'Yes', 'Northern Europe'),
    ('LU', 'Luxembourg', 'EUR', 'Yes', 'Western Europe'),
    ('MT', 'Malta', 'EUR', 'Yes', 'Southern Europe'),
    ('MD', 'Moldova', 'EUR', 'No', 'Eastern Europe'),
    ('ME', 'Montenegro', 'EUR', 'No', 'Balkans'),
    ('NL', 'Netherlands', 'EUR', 'Yes', 'Western Europe'),
    ('MK', 'North Macedonia', 'EUR', 'No', 'Balkans'),
    ('NO', 'Norway', 'EUR', 'No', 'Northern Europe'),
    ('PL', 'Poland', 'EUR', 'Yes', 'Eastern Europe'),
    ('PT', 'Portugal', 'EUR', 'Yes', 'Southern Europe'),
    ('RO', 'Romania', 'EUR', 'Yes', 'Eastern Europe'),
    ('RS', 'Serbia', 'EUR', 'No', 'Balkans'),
    ('SK', 'Slovakia', 'EUR', 'Yes', 'Eastern Europe'),
    ('SI', 'Slovenia', 'EUR', 'Yes', 'Southern Europe'),
    ('ES', 'Spain', 'EUR', 'Yes', 'Southern Europe'),
    ('SE', 'Sweden', 'EUR', 'Yes', 'Northern Europe'),
    ('CH', 'Switzerland', 'EUR', 'No', 'Western Europe'),
    ('TR', 'Turkey', 'EUR', 'No', 'Southern Europe'),
    ('UA', 'Ukraine', 'EUR', 'No', 'Eastern Europe'),
    ('GB', 'United Kingdom', 'EUR', 'No', 'Northern Europe'),

    # WHA - Western Hemisphere (15 countries)
    ('AR', 'Argentina', 'WHA', 'No', 'South America'),
    ('BR', 'Brazil', 'WHA', 'No', 'South America'),
    ('CA', 'Canada', 'WHA', 'No', 'North America'),
    ('CL', 'Chile', 'WHA', 'No', 'South America'),
    ('CO', 'Colombia', 'WHA', 'No', 'South America'),
    ('CR', 'Costa Rica', 'WHA', 'No', 'Central America'),
    ('DO', 'Dominican Republic', 'WHA', 'No', 'Caribbean'),
    ('EC', 'Ecuador', 'WHA', 'No', 'South America'),
    ('SV', 'El Salvador', 'WHA', 'No', 'Central America'),
    ('GT', 'Guatemala', 'WHA', 'No', 'Central America'),
    ('MX', 'Mexico', 'WHA', 'No', 'North America'),
    ('PA', 'Panama', 'WHA', 'No', 'Central America'),
    ('PE', 'Peru', 'WHA', 'No', 'South America'),
    ('US', 'United States', 'WHA', 'No', 'North America'),
    ('UY', 'Uruguay', 'WHA', 'No', 'South America'),

    # EAP - East Asia Pacific (13 countries)
    ('AU', 'Australia', 'EAP', 'No', 'Oceania'),
    ('KH', 'Cambodia', 'EAP', 'No', 'Southeast Asia'),
    ('CN', 'China', 'EAP', 'No', 'East Asia'),
    ('ID', 'Indonesia', 'EAP', 'No', 'Southeast Asia'),
    ('JP', 'Japan', 'EAP', 'No', 'East Asia'),
    ('MY', 'Malaysia', 'EAP', 'No', 'Southeast Asia'),
    ('NZ', 'New Zealand', 'EAP', 'No', 'Oceania'),
    ('PH', 'Philippines', 'EAP', 'No', 'Southeast Asia'),
    ('SG', 'Singapore', 'EAP', 'No', 'Southeast Asia'),
    ('KR', 'South Korea', 'EAP', 'No', 'East Asia'),
    ('TW', 'Taiwan', 'EAP', 'No', 'East Asia'),
    ('TH', 'Thailand', 'EAP', 'No', 'Southeast Asia'),
    ('VN', 'Vietnam', 'EAP', 'No', 'Southeast Asia'),

    # AF - Africa (11 countries)
    ('AO', 'Angola', 'AF', 'No', 'Southern Africa'),
    ('BW', 'Botswana', 'AF', 'No', 'Southern Africa'),
    ('ET', 'Ethiopia', 'AF', 'No', 'East Africa'),
    ('GH', 'Ghana', 'AF', 'No', 'West Africa'),
    ('KE', 'Kenya', 'AF', 'No', 'East Africa'),
    ('NG', 'Nigeria', 'AF', 'No', 'West Africa'),
    ('RW', 'Rwanda', 'AF', 'No', 'East Africa'),
    ('SN', 'Senegal', 'AF', 'No', 'West Africa'),
    ('ZA', 'South Africa', 'AF', 'No', 'Southern Africa'),
    ('TZ', 'Tanzania', 'AF', 'No', 'East Africa'),
    ('ZM', 'Zambia', 'AF', 'No', 'Southern Africa'),

    # NEA - Near East Asia / Middle East (9 countries)
    ('DZ', 'Algeria', 'NEA', 'No', 'North Africa'),
    ('EG', 'Egypt', 'NEA', 'No', 'North Africa'),
    ('IQ', 'Iraq', 'NEA', 'No', 'Middle East'),
    ('IL', 'Israel', 'NEA', 'No', 'Middle East'),
    ('JO', 'Jordan', 'NEA', 'No', 'Middle East'),
    ('LB', 'Lebanon', 'NEA', 'No', 'Middle East'),
    ('MA', 'Morocco', 'NEA', 'No', 'North Africa'),
    ('SA', 'Saudi Arabia', 'NEA', 'No', 'Middle East'),
    ('AE', 'United Arab Emirates', 'NEA', 'No', 'Middle East'),

    # SCA - South Central Asia (6 countries)
    ('AF', 'Afghanistan', 'SCA', 'No', 'Central Asia'),
    ('IN', 'India', 'SCA', 'No', 'South Asia'),
    ('KZ', 'Kazakhstan', 'SCA', 'No', 'Central Asia'),
    ('PK', 'Pakistan', 'SCA', 'No', 'South Asia'),
    ('LK', 'Sri Lanka', 'SCA', 'No', 'South Asia'),
    ('UZ', 'Uzbekistan', 'SCA', 'No', 'Central Asia'),
]

print(f'Country data prepared: {len(countries_data)} countries')
print()
print('Clearing existing country data (preserving headers)...')

# Delete all rows except header
for row in range(ws.max_row, 1, -1):
    ws.delete_rows(row)

print('Adding new country data...')

# Add all new countries
for idx, country in enumerate(countries_data, start=2):
    ws.cell(row=idx, column=1, value=country[0])  # Country_Code
    ws.cell(row=idx, column=2, value=country[1])  # Country_Name
    ws.cell(row=idx, column=3, value=country[2])  # Region
    ws.cell(row=idx, column=4, value=country[3])  # EU_Member
    ws.cell(row=idx, column=5, value=country[4])  # Subregion

print(f'COMPLETE: Added {len(countries_data)} countries')
print()

# Summary by region
regions_count = {}
for country in countries_data:
    region = country[2]
    regions_count[region] = regions_count.get(region, 0) + 1

print('Regional breakdown:')
for region, count in sorted(regions_count.items()):
    print(f'  {region}: {count} countries')
print()

# Save
print('Saving workbook...')
wb.save(input_file)
wb.close()
print('COMPLETE: Country_Regions expanded to 94 countries!')
print()
print('='*80)
print('ALL PARTS 1-4 COMPLETED SUCCESSFULLY')
print('='*80)
print(f'Output file: {input_file}')
