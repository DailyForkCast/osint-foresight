import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'

print('='*80)
print('Adding Project_Technologies Sheet to v5')
print('='*80)
print()

# Load workbook
print('Loading v5...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print(f'Loaded. Current sheets: {", ".join(wb.sheetnames)}')
print()

# Check if sheet already exists
if 'Project_Technologies' in wb.sheetnames:
    print('WARNING: Project_Technologies already exists. Removing and recreating...')
    wb.remove(wb['Project_Technologies'])

# Create new sheet
print('Creating Project_Technologies sheet...')
ws = wb.create_sheet('Project_Technologies')

# Define headers
headers = ['Project_ID', 'Technology', 'Category', 'Status', 'Notes']

# Add headers
print('Adding headers...')
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Sample technology data for each project
# Based on project descriptions from Master_Projects
technologies_data = [
    # PRJ-001: Digital Transformation Initiative
    ('PRJ-001', 'Cloud Computing', 'Infrastructure', 'Active', 'AWS/Azure migration'),
    ('PRJ-001', 'AI/ML', 'Software', 'Active', 'Process automation'),
    ('PRJ-001', 'Data Analytics', 'Software', 'Active', 'Business intelligence'),
    ('PRJ-001', 'API Integration', 'Software', 'Planned', 'Legacy system integration'),

    # PRJ-002: Cybersecurity Enhancement Program
    ('PRJ-002', 'Zero Trust Architecture', 'Security', 'Active', 'Network segmentation'),
    ('PRJ-002', 'SIEM/SOC', 'Security', 'Active', 'Security monitoring'),
    ('PRJ-002', 'Threat Intelligence', 'Security', 'Active', 'Real-time analysis'),
    ('PRJ-002', 'Encryption', 'Security', 'Deployed', 'End-to-end encryption'),

    # PRJ-003: Green Energy Transition
    ('PRJ-003', 'Smart Grid', 'Infrastructure', 'Active', 'Grid modernization'),
    ('PRJ-003', 'IoT Sensors', 'Hardware', 'Active', 'Energy monitoring'),
    ('PRJ-003', 'Renewable Integration', 'Infrastructure', 'Planned', 'Solar/wind systems'),
    ('PRJ-003', 'Energy Analytics', 'Software', 'Active', 'Consumption optimization'),

    # PRJ-004: Healthcare Data Integration
    ('PRJ-004', 'HL7 FHIR', 'Standards', 'Active', 'Healthcare interoperability'),
    ('PRJ-004', 'Blockchain', 'Software', 'Pilot', 'Medical records'),
    ('PRJ-004', 'Data Privacy', 'Security', 'Active', 'HIPAA compliance'),
    ('PRJ-004', 'Clinical Analytics', 'Software', 'Planned', 'Patient outcomes'),

    # PRJ-005: Smart Cities Infrastructure
    ('PRJ-005', 'IoT Platform', 'Infrastructure', 'On Hold', 'City-wide sensors'),
    ('PRJ-005', '5G Network', 'Infrastructure', 'Planned', 'High-speed connectivity'),
    ('PRJ-005', 'Traffic Management', 'Software', 'On Hold', 'Real-time optimization'),

    # PRJ-006: Education Platform Modernization
    ('PRJ-006', 'LMS Platform', 'Software', 'Deployed', 'Learning management'),
    ('PRJ-006', 'Virtual Reality', 'Hardware', 'Pilot', 'Immersive learning'),
    ('PRJ-006', 'Collaboration Tools', 'Software', 'Deployed', 'Video conferencing'),
    ('PRJ-006', 'Analytics Dashboard', 'Software', 'Active', 'Student performance'),

    # PRJ-007: Border Security Enhancement
    ('PRJ-007', 'Biometric Systems', 'Hardware', 'Deployed', 'Facial recognition'),
    ('PRJ-007', 'Surveillance AI', 'Software', 'Deployed', 'Threat detection'),
    ('PRJ-007', 'Command & Control', 'Software', 'Deployed', 'Central monitoring'),

    # PRJ-008: Agricultural Innovation Hub
    ('PRJ-008', 'Precision Agriculture', 'Software', 'Not Started', 'Crop optimization'),
    ('PRJ-008', 'Drone Technology', 'Hardware', 'Not Started', 'Field monitoring'),
    ('PRJ-008', 'Soil Sensors', 'Hardware', 'Not Started', 'Ground monitoring'),

    # PRJ-009: Tourism Promotion Initiative
    ('PRJ-009', 'Mobile Apps', 'Software', 'Active', 'Tourist information'),
    ('PRJ-009', 'AR Experiences', 'Software', 'Active', 'Augmented reality tours'),
    ('PRJ-009', 'Digital Marketing', 'Software', 'Active', 'Social media campaigns'),

    # PRJ-010: Transport Logistics Optimization
    ('PRJ-010', 'AI Route Optimization', 'Software', 'Deployed', 'Logistics planning'),
    ('PRJ-010', 'Fleet Management', 'Software', 'Deployed', 'Vehicle tracking'),
    ('PRJ-010', 'Predictive Maintenance', 'Software', 'Deployed', 'Asset management'),
]

# Add data
print('Adding technology data...')
for idx, (project_id, technology, category, status, notes) in enumerate(technologies_data, start=2):
    ws.cell(row=idx, column=1, value=project_id)
    ws.cell(row=idx, column=2, value=technology)
    ws.cell(row=idx, column=3, value=category)
    ws.cell(row=idx, column=4, value=status)
    ws.cell(row=idx, column=5, value=notes)

print(f'  Added {len(technologies_data)} technology records')
print()

# Set column widths
column_widths = {
    'A': 12,  # Project_ID
    'B': 25,  # Technology
    'C': 18,  # Category
    'D': 15,  # Status
    'E': 35   # Notes
}

print('Setting column widths...')
for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze top row
ws.freeze_panes = 'A2'

# Create Excel Table
print('Creating Excel Table...')
from openpyxl.worksheet.table import Table, TableStyleInfo

table_ref = f'A1:E{len(technologies_data) + 1}'
table = Table(displayName='Project_Technologies', ref=table_ref)

style = TableStyleInfo(
    name='TableStyleMedium2',
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

print(f'  Table created: {table_ref}')
print()

# Now update Project_Spotlight to use the new sheet
print('Updating Project_Spotlight formulas...')
if 'Project_Spotlight' in wb.sheetnames:
    ws_spotlight = wb['Project_Spotlight']

    # Update the placeholder in J5-K5 with real headers
    ws_spotlight['J5'].value = 'Technology'
    ws_spotlight['K5'].value = 'Category'

    # Add formulas for technologies (rows 6-14)
    updates = 0
    for row_offset in range(1, 10):
        row_num = 5 + row_offset

        # Check if cell is empty or has placeholder text
        current_val = ws_spotlight[f'J{row_num}'].value
        if not current_val or '[Create' in str(current_val):
            # Technology name (Column J)
            ws_spotlight[f'J{row_num}'].value = f'=IFERROR(INDEX(Project_Technologies!B:B,SMALL(IF(Project_Technologies!$A:$A=$B$2,ROW(Project_Technologies!$A:$A)),{row_offset})),"")'

            # Category (Column K)
            ws_spotlight[f'K{row_num}'].value = f'=IF(J{row_num}="","",INDEX(Project_Technologies!C:C,MATCH(J{row_num},Project_Technologies!B:B,0)))'

            updates += 1

    print(f'  Updated {updates} formula rows in Project_Spotlight')
else:
    print('  WARNING: Project_Spotlight not found')

print()

# Save
print('='*80)
print('Saving...')
wb.save(output_file)
wb.close()

print('='*80)
print('SUCCESS: Project_Technologies Sheet Added!')
print('='*80)
print()
print('Summary:')
print(f'  - Created Project_Technologies sheet')
print(f'  - Added {len(technologies_data)} technology records across 10 projects')
print(f'  - Technologies by project:')
print(f'    PRJ-001 (Digital Transform): 4 technologies')
print(f'    PRJ-002 (Cybersecurity): 4 technologies')
print(f'    PRJ-003 (Green Energy): 4 technologies')
print(f'    PRJ-004 (Healthcare): 4 technologies')
print(f'    PRJ-005 (Smart Cities): 3 technologies')
print(f'    PRJ-006 (Education): 4 technologies')
print(f'    PRJ-007 (Border Security): 3 technologies')
print(f'    PRJ-008 (Agriculture): 3 technologies')
print(f'    PRJ-009 (Tourism): 3 technologies')
print(f'    PRJ-010 (Transport): 3 technologies')
print(f'  - Updated Project_Spotlight to use the new sheet')
print()
print(f'Output: {output_file}')
print()
print('Next: Open the file and test Project_Spotlight!')
print('  1. Go to Project_Spotlight sheet')
print('  2. Select a project in cell B2')
print('  3. Watch the technologies section (J5:K14) populate automatically')
