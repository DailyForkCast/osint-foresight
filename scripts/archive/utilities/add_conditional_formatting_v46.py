"""
Add conditional formatting to key sheets - v46
Visual enhancements - color coding and alerts
"""
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule

print("="*80)
print("ADDING CONDITIONAL FORMATTING - V46")
print("="*80)

# Load v45 (skip v45.1 since data validation didn't work)
print("\nLoading v45...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v45.xlsx')

# Define colors
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_text = Font(color='9C0006')
orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_text = Font(color='9C6500')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_text = Font(color='006100')
light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

print("\n" + "="*80)
print("1. PORTFOLIO_DASHBOARD")
print("="*80)

ws_portfolio = wb['Portfolio_Dashboard']

# ULO % Alert (Column L) - Red text if > 75%
print("\nAdding ULO % alert (L11:L20) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('L11:L20', rule)
print("  Added: ULO % > 75% = Red")

# Days Remaining Alert (Column M) - Red if < 90
print("\nAdding Days Remaining alert (M11:M20) - Red if <90...")
rule = CellIsRule(
    operator='lessThan',
    formula=['90'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('M11:M20', rule)
print("  Added: Days < 90 = Red")

# Priority Color Coding (Column D)
print("\nAdding Priority color coding (D11:D20)...")
# Critical = Red
rule = CellIsRule(
    operator='equal',
    formula=['"Critical"'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# High = Orange
rule = CellIsRule(
    operator='equal',
    formula=['"High"'],
    stopIfTrue=True,
    fill=orange_fill,
    font=orange_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# Medium = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"Medium"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)

# Low = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Low"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_portfolio.conditional_formatting.add('D11:D20', rule)
print("  Added: Priority colors (Critical=Red, High=Orange, Medium=Yellow, Low=Green)")

# Status Color Coding (Column C)
print("\nAdding Status color coding (C11:C20)...")
# Active = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Active"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)

# On Hold = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"On Hold"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)

# Completed = Light Green
rule = CellIsRule(
    operator='equal',
    formula=['"Completed"'],
    stopIfTrue=True,
    fill=light_green_fill
)
ws_portfolio.conditional_formatting.add('C11:C20', rule)
print("  Added: Status colors (Active=Green, On Hold=Yellow, Completed=Light Green)")

print("\n" + "="*80)
print("2. COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

# ULO % Alert (E5 and G6) - Red if > 75%
print("\nAdding ULO % alerts (E5, G6) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_country.conditional_formatting.add('E5', rule)
ws_country.conditional_formatting.add('G6', rule)
print("  Added: ULO % > 75% = Red")

# Project Status Colors (C12:C31)
print("\nAdding Status color coding (C12:C31)...")
# Active = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Active"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_country.conditional_formatting.add('C12:C31', rule)

# On Hold = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"On Hold"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_country.conditional_formatting.add('C12:C31', rule)

# Completed = Light Green
rule = CellIsRule(
    operator='equal',
    formula=['"Completed"'],
    stopIfTrue=True,
    fill=light_green_fill
)
ws_country.conditional_formatting.add('C12:C31', rule)
print("  Added: Status colors")

# Priority Color Coding (D12:D31)
print("\nAdding Priority color coding (D12:D31)...")
# Critical = Red
rule = CellIsRule(
    operator='equal',
    formula=['"Critical"'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_country.conditional_formatting.add('D12:D31', rule)

# High = Orange
rule = CellIsRule(
    operator='equal',
    formula=['"High"'],
    stopIfTrue=True,
    fill=orange_fill,
    font=orange_text
)
ws_country.conditional_formatting.add('D12:D31', rule)

# Medium = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"Medium"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_country.conditional_formatting.add('D12:D31', rule)

# Low = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Low"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_country.conditional_formatting.add('D12:D31', rule)
print("  Added: Priority colors")

print("\n" + "="*80)
print("3. SPOTLIGHT_PMWORKSPACE")
print("="*80)

ws_spotlight = wb['Spotlight_PMWorkspace']

# ULO % Alert (H15) - Red if > 75%
print("\nAdding ULO % alert (H15) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_spotlight.conditional_formatting.add('H15', rule)
print("  Added: ULO % > 75% = Red")

# Days Remaining Alert (J3) - Red if < 90
print("\nAdding Days Remaining alert (J3) - Red if <90...")
rule = CellIsRule(
    operator='lessThan',
    formula=['90'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_spotlight.conditional_formatting.add('J3', rule)
print("  Added: Days < 90 = Red")

# Status Color (G5)
print("\nAdding Status color coding (G5)...")
# Active = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Active"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_spotlight.conditional_formatting.add('G5', rule)

# On Hold = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"On Hold"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_spotlight.conditional_formatting.add('G5', rule)
print("  Added: Status colors")

# Priority Color (B6)
print("\nAdding Priority color coding (B6)...")
# Critical = Red
rule = CellIsRule(
    operator='equal',
    formula=['"Critical"'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_spotlight.conditional_formatting.add('B6', rule)

# High = Orange
rule = CellIsRule(
    operator='equal',
    formula=['"High"'],
    stopIfTrue=True,
    fill=orange_fill,
    font=orange_text
)
ws_spotlight.conditional_formatting.add('B6', rule)
print("  Added: Priority colors")

print("\n" + "="*80)
print("4. CONTROL")
print("="*80)

ws_control = wb['Control']

# Portfolio ULO % Alert (B19) - Red if > 75%
print("\nAdding Portfolio ULO % alert (B19) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_control.conditional_formatting.add('B19', rule)
print("  Added: Portfolio ULO % > 75% = Red")

# Avg ULO % Alert (E4) - Red if > 75%
print("\nAdding Avg ULO % alert (E4) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_control.conditional_formatting.add('E4', rule)
print("  Added: Avg ULO % > 75% = Red")

print("\n" + "="*80)
print("5. MILESTONES")
print("="*80)

ws_milestones = wb['Milestones']

# Overdue milestones - Red if Due_Date < TODAY and Status != Complete
print("\nAdding overdue milestone alert (D2:D100)...")
# This uses a formula rule: Due date is past AND status is not Complete
rule = FormulaRule(
    formula=['AND($D2<TODAY(),$E2<>"Complete")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_milestones.conditional_formatting.add('A2:J100', rule)
print("  Added: Overdue milestones = Red (Due Date < Today AND Status != Complete)")

# Upcoming milestones - Yellow if within 30 days
print("\nAdding upcoming milestone alert (D2:D100)...")
rule = FormulaRule(
    formula=['AND($D2>=TODAY(),$D2<=TODAY()+30,$E2<>"Complete")'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_milestones.conditional_formatting.add('A2:J100', rule)
print("  Added: Upcoming milestones = Yellow (Due within 30 days)")

print("\n" + "="*80)
print("6. MASTER_PROJECTS")
print("="*80)

ws_master = wb['Master_Projects']

# Overdue projects - Red if POP_End_Date < TODAY
print("\nAdding overdue project alert...")
# POP_End_Date is column W (column 23)
rule = FormulaRule(
    formula=['$W2<TODAY()'],
    stopIfTrue=True,
    fill=light_red_fill
)
ws_master.conditional_formatting.add('A2:AD201', rule)
print("  Added: Overdue projects = Light Red (POP_End_Date < Today)")

# Priority colors in Priority column (F)
print("\nAdding Priority color coding (F2:F201)...")
# Critical = Red
rule = CellIsRule(
    operator='equal',
    formula=['"Critical"'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_master.conditional_formatting.add('F2:F201', rule)

# High = Orange
rule = CellIsRule(
    operator='equal',
    formula=['"High"'],
    stopIfTrue=True,
    fill=orange_fill,
    font=orange_text
)
ws_master.conditional_formatting.add('F2:F201', rule)

# Medium = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"Medium"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_master.conditional_formatting.add('F2:F201', rule)

# Low = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Low"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_master.conditional_formatting.add('F2:F201', rule)
print("  Added: Priority colors")

print("\n" + "="*80)
print("7. COUNTRY_BUDGETS")
print("="*80)

ws_budgets = wb['Country_Budgets']

# ULO % Alert (M2:M1001) - Red if > 75%
print("\nAdding ULO % alert (M2:M1001) - Red if >75%...")
rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_budgets.conditional_formatting.add('M2:M1001', rule)
print("  Added: ULO % > 75% = Red")

# My_Country = TRUE highlight (C2:C1001) - Green
print("\nAdding My_Country highlight (C2:C1001)...")
rule = CellIsRule(
    operator='equal',
    formula=['TRUE'],
    stopIfTrue=True,
    fill=light_green_fill
)
ws_budgets.conditional_formatting.add('C2:C1001', rule)
print("  Added: My_Country = TRUE = Light Green")

print("\n" + "="*80)
print("8. PROJECT_DELIVERABLES")
print("="*80)

ws_deliverables = wb['Project_Deliverables']

# Overdue deliverables - Red if Due_Date < TODAY and Status != Completed
print("\nAdding overdue deliverable alert...")
rule = FormulaRule(
    formula=['AND($D2<TODAY(),$E2<>"Completed")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_deliverables.conditional_formatting.add('A2:H100', rule)
print("  Added: Overdue deliverables = Red")

# Status colors (E2:E100)
print("\nAdding Status color coding (E2:E100)...")
# Completed = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Completed"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_deliverables.conditional_formatting.add('E2:E100', rule)

# In Progress = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"In Progress"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_deliverables.conditional_formatting.add('E2:E100', rule)
print("  Added: Status colors")

print("\n" + "="*80)
print("9. CALENDAR_TODO")
print("="*80)

ws_calendar = wb['Calendar_Todo']

# Overdue tasks - Red if Due_Date < TODAY and Status != Completed
print("\nAdding overdue task alert...")
rule = FormulaRule(
    formula=['AND($D2<TODAY(),$F2<>"Completed")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_calendar.conditional_formatting.add('A2:H100', rule)
print("  Added: Overdue tasks = Red")

# Due soon (within 7 days) - Yellow
print("\nAdding due soon alert...")
rule = FormulaRule(
    formula=['AND($D2>=TODAY(),$D2<=TODAY()+7,$F2<>"Completed")'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_calendar.conditional_formatting.add('A2:H100', rule)
print("  Added: Due within 7 days = Yellow")

print("\n" + "="*80)
print("SAVING V46")
print("="*80)

wb.save('2025-10-26-Tracker-v46.xlsx')

print("\nOK - v46 created!")

print("\n" + "="*80)
print("SUMMARY - CONDITIONAL FORMATTING ADDED")
print("="*80)

print("\n>> Portfolio_Dashboard:")
print("  - ULO % > 75% = Red")
print("  - Days < 90 = Red")
print("  - Priority: Critical=Red, High=Orange, Medium=Yellow, Low=Green")
print("  - Status: Active=Green, On Hold=Yellow, Completed=Light Green")

print("\n>> Country_Dashboard:")
print("  - ULO % > 75% = Red (both E5 and G6)")
print("  - Status colors")
print("  - Priority colors")

print("\n>> Spotlight_PMWorkspace:")
print("  - ULO % > 75% = Red")
print("  - Days < 90 = Red")
print("  - Status colors")
print("  - Priority colors")

print("\n>> Control:")
print("  - Portfolio ULO % > 75% = Red")
print("  - Avg ULO % > 75% = Red")

print("\n>> Milestones:")
print("  - Overdue (Due Date < Today AND Status != Complete) = Red")
print("  - Due within 30 days = Yellow")

print("\n>> Master_Projects:")
print("  - Overdue projects (POP_End_Date < Today) = Light Red")
print("  - Priority colors")

print("\n>> Country_Budgets:")
print("  - ULO % > 75% = Red")
print("  - My_Country = TRUE = Light Green")

print("\n>> Project_Deliverables:")
print("  - Overdue deliverables = Red")
print("  - Status: Completed=Green, In Progress=Yellow")

print("\n>> Calendar_Todo:")
print("  - Overdue tasks = Red")
print("  - Due within 7 days = Yellow")

print("\nAll conditional formatting added successfully!")
print("Colors will appear automatically based on your data!")
print("v46 ready!")
