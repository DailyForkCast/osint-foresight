"""
Fix Country_Dashboard ULO% formula and Project ID zero - v44
Based on user moving things around
"""
import openpyxl

print("="*80)
print("FIXING COUNTRY_DASHBOARD - V44")
print("="*80)

# Load v43
print("\nLoading v43...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')

ws_country = wb['Country_Dashboard']

print("\n" + "="*80)
print("CURRENT ISSUES IDENTIFIED")
print("="*80)

print("\nRow 4 labels:")
print("  A4: TOTAL PROPOSED")
print("  B4: TOTAL OBLIGATED")
print("  C4: TOTAL SPENT")
print("  D4: ULO")
print("  E4: ULO %")

print("\nRow 5 formulas:")
print("  A5: Sum of Proposed - OK")
print("  B5: Sum of Obligated - OK")
print("  C5: Sum of Spent - OK")
print("  D5: =B5-E5 - BROKEN (E5 is empty!)")
print("  E5: (empty) - NEEDS FORMULA")

print("\nAlso found:")
print("  F6: ULO %: (duplicate label)")
print("  G6: =IF(B5=0,0,G5/B5) - BROKEN (G5 is empty!)")

print("\n" + "="*80)
print("FIXES TO APPLY")
print("="*80)

print("\n1. Fix D5: Should be =B5-C5 (Obligated - Spent)")
print("2. Fix E5: Should be =IF(B5=0,0,D5/B5) (ULO % = ULO / Obligated)")
print("3. Fix G6: Should reference E5 or be removed")
print("4. Fix Project ID formulas to show blank instead of 0")

print("\n" + "="*80)
print("APPLYING FIXES")
print("="*80)

# Fix 1: D5 - ULO should be Obligated - Spent
print("\n1. Fixing D5 (ULO)...")
old_d5 = ws_country['D5'].value
ws_country['D5'] = '=B5-C5'
print(f"   Old: {old_d5}")
print(f"   New: =B5-C5")

# Fix 2: E5 - ULO % should be ULO / Obligated
print("\n2. Fixing E5 (ULO %)...")
ws_country['E5'] = '=IF(B5=0,0,D5/B5)'
print(f"   New: =IF(B5=0,0,D5/B5)")

# Fix 3: G6 - should reference E5
print("\n3. Fixing G6...")
old_g6 = ws_country['G6'].value
ws_country['G6'] = '=E5'
print(f"   Old: {old_g6}")
print(f"   New: =E5")

# Fix 4: Project ID - add blank check
print("\n4. Fixing Project ID formulas (A12-A30)...")
print("   Adding check for empty values...")

# Count how many project rows we have
project_rows = 0
for row in range(12, 50):
    if ws_country.cell(row, 1).value:
        project_rows += 1
    else:
        break

print(f"   Found {project_rows} project rows")

for i in range(1, project_rows + 1):
    row = 11 + i
    # The original formula uses SMALL with index i to get the i-th matching project
    # We need to keep that logic but add a blank check
    new_formula = f'=IFERROR(IF(INDEX(T_Country_Budgets[Unique_ID],SMALL(IF(T_Country_Budgets[[#This Row],[Country_Code]]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{i}))="","",INDEX(T_Country_Budgets[Unique_ID],SMALL(IF(T_Country_Budgets[[#This Row],[Country_Code]]=$B$2,ROW(T_Country_Budgets[Country_Code])-1),{i}))),"")'

    ws_country.cell(row, 1).value = new_formula

print(f"   Updated rows 12-{11+project_rows}")

print("\n" + "="*80)
print("SAVING V44")
print("="*80)

wb.save('2025-10-26-Tracker-v44.xlsx')

print("\nOK - v44 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v44:")
print("  1. D5: Now calculates ULO correctly (Obligated - Spent)")
print("  2. E5: Now calculates ULO % correctly (ULO / Obligated)")
print("  3. G6: Now references E5 for consistent ULO %")
print("  4. Project ID (A12-A30): Now shows blank instead of 0")

print("\nAll formulas should work correctly now!")
print("\nv44 ready!")
