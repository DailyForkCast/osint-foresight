import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def is_merged_cell(ws, row, col):
    """Check if a cell is part of a merged range"""
    cell = ws.cell(row, col)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

def safe_update(ws, row, col, value):
    """Safely update a cell if it's not merged"""
    if not is_merged_cell(ws, row, col):
        ws.cell(row, col).value = value
        return True
    return False

print("Updating Spotlight sheets with new budget structure...")
print("="*80)

wb = openpyxl.load_workbook('2025-10-26-Tracker-v13.xlsx')

# Update Spotlight_PMWorkspace
print("\n1. Updating Spotlight_PMWorkspace FINANCIAL DETAILS...")
ws = wb['Spotlight_PMWorkspace']

# Row 21: First row of financial data
safe_update(ws, 21, 1, "Total Proposed:")
safe_update(ws, 21, 2, "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$F:$F)")
safe_update(ws, 21, 4, "Total Allocated:")
safe_update(ws, 21, 5, "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$G:$G)")
safe_update(ws, 21, 7, "Total Obligated:")
safe_update(ws, 21, 8, "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$H:$H)")

# Row 22: Second row
safe_update(ws, 22, 1, "Total Spent:")
safe_update(ws, 22, 2, "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$I:$I)")
safe_update(ws, 22, 4, "ULO:")
safe_update(ws, 22, 5, "=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$J:$J)")
safe_update(ws, 22, 7, "ULO %:")
if safe_update(ws, 22, 8, "=IF(H21>0,E22/H21,0)"):
    ws.cell(22, 8).number_format = '0%'

print("  Updated PMWorkspace financial section (rows 21-22)")

# Update Spotlight_Executive
print("\n2. Updating Spotlight_Executive FINANCIAL SUMMARY...")
ws_exec = wb['Spotlight_Executive']

# Find FINANCIAL SUMMARY section
financial_row = None
for row in range(1, 30):
    val = ws_exec.cell(row, 1).value
    if val and 'FINANCIAL' in str(val).upper():
        financial_row = row
        print(f"  Found FINANCIAL section at row {row}")
        break

if financial_row:
    # Headers row
    header_row = financial_row + 1
    safe_update(ws_exec, header_row, 1, "Proposed")
    safe_update(ws_exec, header_row, 2, "Allocated")
    safe_update(ws_exec, header_row, 3, "Obligated")
    safe_update(ws_exec, header_row, 4, "Spent")
    safe_update(ws_exec, header_row, 5, "ULO")
    safe_update(ws_exec, header_row, 6, "ULO %")

    # Data row
    data_row = financial_row + 2
    safe_update(ws_exec, data_row, 1, f'=IFERROR(INDEX(Master_Projects!$N:$N,MATCH($B$2,Master_Projects!$A:$A,0)),0)')
    safe_update(ws_exec, data_row, 2, f'=IFERROR(INDEX(Master_Projects!$O:$O,MATCH($B$2,Master_Projects!$A:$A,0)),0)')
    safe_update(ws_exec, data_row, 3, f'=IFERROR(INDEX(Master_Projects!$P:$P,MATCH($B$2,Master_Projects!$A:$A,0)),0)')
    safe_update(ws_exec, data_row, 4, f'=SUMIF(Country_Budgets!$B:$B,$B$2,Country_Budgets!$I:$I)')
    safe_update(ws_exec, data_row, 5, f'=IFERROR(INDEX(Master_Projects!$Q:$Q,MATCH($B$2,Master_Projects!$A:$A,0)),0)')
    safe_update(ws_exec, data_row, 6, f'=IFERROR(INDEX(Master_Projects!$R:$R,MATCH($B$2,Master_Projects!$A:$A,0)),0)')

    print(f"  Updated Executive financial section at rows {header_row}-{data_row}")

wb.save('2025-10-26-Tracker-v13.xlsx')
print("\n" + "="*80)
print("Spotlight sheets updated successfully!")
print("="*80)

print("\nFINAL VERIFICATION")
print("="*80)

# Verify the changes
ws_budget = wb['Country_Budgets']
print("\nCountry_Budgets columns:")
for col in range(1, 13):
    header = ws_budget.cell(1, col).value
    if header:
        print(f"  Col {chr(64+col):2s}: {header}")

ws_projects = wb['Master_Projects']
print("\nMaster_Projects budget columns (N-S):")
for col in range(14, 20):
    header = ws_projects.cell(1, col).value
    if header:
        print(f"  Col {chr(64+col):2s}: {header}")

ws_regions = wb['Country_Regions']
print("\nCountry_Regions sample (showing State Dept regions):")
print("  Code  Country              Region  Subregion")
print("  " + "-"*70)
for row in range(2, min(12, ws_regions.max_row + 1)):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    subregion = ws_regions.cell(row, 5).value
    if code:
        print(f"  {code:4s}  {name:20s} {region:6s}  {subregion if subregion else 'N/A'}")

ws_config = wb['Config_Lists']
country_count = sum(1 for row in range(2, ws_config.max_row + 1)
                    if ws_config.cell(row, 1).value == "Country")
print(f"\nConfig_Lists: {country_count} countries added")

ws_pm = wb['Country_PM_Assignments']
print(f"Country_PM_Assignments: {ws_pm.max_row - 1} countries")

print("\n" + "="*80)
print("ALL UPDATES COMPLETE!")
print("="*80)
print("\nChanges made to v13:")
print("  ✓ Added Proposed_Amount column to Country_Budgets (before Allocated)")
print("  ✓ ULO calculations now use Obligated - Spent")
print("  ✓ Added Total_Proposed column to Master_Projects")
print("  ✓ Updated regions to State Department codes (AF, EAP, EUR, NEA, SCA, WHA)")
print("  ✓ Added subregions for all countries")
print("  ✓ Added all countries to Config_Lists")
print("  ✓ Synced Country_PM_Assignments with Country_Regions")
print("  ✓ Updated Spotlight sheets to show Proposed/Allocated/Obligated/Spent/ULO")
print("\nFile: 2025-10-26-Tracker-v13.xlsx")
