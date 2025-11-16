import openpyxl

file_path = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.2.xlsx'

print('='*80)
print('INVESTIGATING AUTOMATION OPPORTUNITIES')
print('='*80)
print()

wb = openpyxl.load_workbook(file_path, data_only=False)

print('Available sheets:')
for idx, sheet in enumerate(wb.sheetnames, 1):
    print(f'{idx:2d}. {sheet}')
print()

# ============================================================================
# CHECK CONFIG_LISTS
# ============================================================================

if 'Config_Lists' in wb.sheetnames:
    ws = wb['Config_Lists']
    print('='*80)
    print('CONFIG_LISTS SHEET')
    print('='*80)
    print()
    print('Structure (first 30 rows, first 10 columns):')
    print()

    for row in range(1, min(31, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row, col)
            value = str(cell.value) if cell.value else ''
            row_data.append(value[:20])
        if any(row_data):  # Only print non-empty rows
            print(f'Row {row:2d}: {" | ".join(row_data)}')
    print()

# ============================================================================
# CHECK MASTER_PROJECTS for dropdowns
# ============================================================================

print('='*80)
print('MASTER_PROJECTS - CHECKING FOR DROPDOWNS')
print('='*80)
print()

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    # Check headers
    print('Column headers:')
    for col in range(1, min(20, ws.max_column + 1)):
        header = ws.cell(1, col).value
        if header:
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f'  {col_letter}: {header}')
    print()

    # Check for existing data validations
    print('Data validations in Master_Projects:')
    if hasattr(ws, 'data_validations') and ws.data_validations.dataValidation:
        for idx, dv in enumerate(ws.data_validations.dataValidation):
            print(f'  Validation #{idx + 1}:')
            print(f'    Type: {dv.type}')
            print(f'    Formula: {dv.formula1}')
            print(f'    Cells: {dv.sqref}')
    else:
        print('  None found')
    print()

# ============================================================================
# CHECK OTHER SHEETS for missing dropdowns
# ============================================================================

sheets_to_check = [
    'Milestones',
    'Events',
    'Risk_Register',
    'Decision_Log',
    'Stakeholders',
    'Project_Deliverables',
    'Project_Audiences',
    'Project_Products',
    'Project_Technologies'
]

print('='*80)
print('CHECKING OTHER SHEETS FOR AUTOMATION OPPORTUNITIES')
print('='*80)
print()

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'{sheet_name}:')
        print(f'  Rows: {ws.max_row}, Columns: {ws.max_column}')

        # Show headers
        headers = []
        for col in range(1, min(15, ws.max_column + 1)):
            header = ws.cell(1, col).value
            if header:
                headers.append(str(header))
        print(f'  Headers: {", ".join(headers)}')

        # Check for data validations
        has_validations = hasattr(ws, 'data_validations') and ws.data_validations.dataValidation
        print(f'  Has dropdowns: {"Yes" if has_validations else "No"}')
        print()

wb.close()

print('='*80)
print('AUTOMATION OPPORTUNITIES IDENTIFIED:')
print('='*80)
print()
print('Potential improvements:')
print('  1. Add dropdowns to Master_Projects (Status, Priority, etc.)')
print('  2. Review Config_Lists structure and usefulness')
print('  3. Add dropdowns to Milestones, Risk_Register, etc.')
print('  4. Add Project_ID dropdowns where needed')
print('  5. Add Country_Code dropdowns where needed')
print('  6. Standardize date formats')
print()
