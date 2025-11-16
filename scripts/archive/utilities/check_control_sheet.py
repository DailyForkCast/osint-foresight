"""
Check Control sheet to see what's in B3
"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("CONTROL SHEET INSPECTION")
print("="*80)

ws = wb['Control']

print("\nFirst 10 rows of Control sheet:")
print(f"{'Row':<5} {'Col A':<30} {'Col B':<30} {'Col C':<30}")
print("-"*95)

for row in range(1, 11):
    val_a = ws.cell(row, 1).value or ""
    val_b = ws.cell(row, 2).value or ""
    val_c = ws.cell(row, 3).value or ""

    print(f"{row:<5} {str(val_a)[:28]:<30} {str(val_b)[:28]:<30} {str(val_c)[:28]:<30}")

print("\n" + "="*80)
print("CELL B3 DETAILS")
print("="*80)

cell = ws['B3']
print(f"\nValue: {cell.value}")
print(f"Type: {type(cell.value)}")
print(f"Number format: {cell.number_format}")

if cell.value:
    print(f"\nIs this a date? {isinstance(cell.value, datetime)}")
    if isinstance(cell.value, (int, float)):
        print(f"Numeric value: {cell.value}")
        # Try to interpret as Excel date
        try:
            from datetime import datetime, timedelta
            excel_epoch = datetime(1899, 12, 30)
            date_value = excel_epoch + timedelta(days=cell.value)
            print(f"As date: {date_value.strftime('%Y-%m-%d')}")
        except:
            pass

print("\n" + "="*80)
print("SOLUTION")
print("="*80)
print("\nOption 1: Update Control!B3 to have TODAY() formula")
print("Option 2: Change Days_Remaining formula to use TODAY() directly")
print("\nRecommendation: Use TODAY() in the formula for simplicity")
