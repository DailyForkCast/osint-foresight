"""
Add note about stakeholder categorization to Project_Stakeholders sheet
"""
import openpyxl

print("="*80)
print("ADDING STAKEHOLDER CATEGORIZATION NOTE")
print("="*80)

# Load v21
print("\nLoading v21...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v21.xlsx')

ws = wb['Stakeholders']

print("\nAdding note in cells A1 and A2...")

# Add note at the top
note_text = """NOTE: Stakeholder Categorization - Use the following guidelines for categorizing stakeholders:
- Government Officials: Gov agency contacts, embassy staff, ministry officials
- Implementing Partners: Contractors, grantees, vendors executing project work
- Internal Team: DOS staff, project managers, technical advisors
- Beneficiaries: End users, communities, target populations
- Oversight: Congress, GAO, OIG, monitoring entities
Stakeholder_ID auto-generates as: ProjectID-STK-001, ProjectID-STK-002, etc."""

# Check if we need to insert rows or if there's space
# Typically the headers would be in row 1, so we'll add a merged cell note above

# Insert 2 rows at the top
ws.insert_rows(1, 2)

# Add the note in A1 (merge A1:H2 for visibility)
ws.merge_cells('A1:H2')
ws['A1'] = note_text
ws['A1'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

# Style the note
ws['A1'].fill = openpyxl.styles.PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
ws['A1'].font = openpyxl.styles.Font(size=9, italic=True)

print("  OK - Note added in merged cells A1:H2")
print("  OK - Headers shifted down to row 3")

# Save
print("\nSaving v21...")
wb.save('2025-10-26-Tracker-v21.xlsx')

print("\n" + "="*80)
print("STAKEHOLDER NOTE ADDED!")
print("="*80)

print("\nNote content:")
print("  - Government Officials (Gov agency, embassy, ministry)")
print("  - Implementing Partners (Contractors, grantees, vendors)")
print("  - Internal Team (DOS staff, PMs, technical advisors)")
print("  - Beneficiaries (End users, communities, target populations)")
print("  - Oversight (Congress, GAO, OIG, monitoring)")
print("  - Stakeholder_ID format: ProjectID-STK-001, ProjectID-STK-002, etc.")

print("\nv21 is ready!")
