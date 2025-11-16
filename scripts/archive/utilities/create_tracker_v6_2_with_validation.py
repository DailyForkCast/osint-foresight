import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6.2.xlsx'

print('='*80)
print('CREATING v6.2 WITH DATA VALIDATION')
print('='*80)
print()

# Load workbook
print('Loading v6...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

total_validations = 0

# ============================================================================
# PRIORITY 1: Project_Spotlight B2 (CRITICAL)
# ============================================================================
print('='*80)
print('PRIORITY 1: Project_Spotlight B2 - Project Selection')
print('='*80)
print()

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']

    # Create validation for project selection
    dv = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=False
    )
    dv.error = 'Please select a valid project ID'
    dv.errorTitle = 'Invalid Project'
    dv.prompt = 'Select a project from the list'
    dv.promptTitle = 'Project Selection'

    ws.add_data_validation(dv)
    dv.add('B2')

    print('  B2: Added project dropdown (Master_Projects[Unique_ID])')
    total_validations += 1
else:
    print('  WARNING: Project_Spotlight not found')

print()

# ============================================================================
# PRIORITY 2: Master_Projects Status, Priority, NCE_Status
# ============================================================================
print('='*80)
print('PRIORITY 2: Master_Projects - Status, Priority, NCE_Status')
print('='*80)
print()

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    # E column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    dv_status.prompt = 'Select project status'
    dv_status.promptTitle = 'Status'

    ws.add_data_validation(dv_status)
    dv_status.add('E2:E1000')

    print('  E2:E1000 (Status): Added dropdown (List_Status)')
    total_validations += 1

    # F column - Priority
    dv_priority = DataValidation(
        type="list",
        formula1="=List_Priority",
        allow_blank=True
    )
    dv_priority.prompt = 'Select priority level'
    dv_priority.promptTitle = 'Priority'

    ws.add_data_validation(dv_priority)
    dv_priority.add('F2:F1000')

    print('  F2:F1000 (Priority): Added dropdown (List_Priority)')
    total_validations += 1

    # X column - NCE_Status
    dv_nce = DataValidation(
        type="list",
        formula1="=L_NCE_Status",
        allow_blank=True
    )
    dv_nce.prompt = 'Select NCE status'
    dv_nce.promptTitle = 'NCE Status'

    ws.add_data_validation(dv_nce)
    dv_nce.add('X2:X1000')

    print('  X2:X1000 (NCE_Status): Added dropdown (L_NCE_Status)')
    total_validations += 1
else:
    print('  WARNING: Master_Projects not found')

print()

# ============================================================================
# PRIORITY 3: Country_Budgets - Project ID and Country Code
# ============================================================================
print('='*80)
print('PRIORITY 3: Country_Budgets - Project ID and Country Code')
print('='*80)
print()

if 'Country_Budgets' in wb.sheetnames:
    ws = wb['Country_Budgets']

    # B column - Unique_ID (Project ID)
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    dv_project.prompt = 'Select project ID'
    dv_project.promptTitle = 'Project'

    ws.add_data_validation(dv_project)
    dv_project.add('B2:B5000')

    print('  B2:B5000 (Unique_ID): Added dropdown (Master_Projects[Unique_ID])')
    total_validations += 1

    # C column - Country_Code
    dv_country = DataValidation(
        type="list",
        formula1="=List_CountryCodes",
        allow_blank=True
    )
    dv_country.prompt = 'Select country code'
    dv_country.promptTitle = 'Country'

    ws.add_data_validation(dv_country)
    dv_country.add('C2:C5000')

    print('  C2:C5000 (Country_Code): Added dropdown (List_CountryCodes)')
    total_validations += 1
else:
    print('  WARNING: Country_Budgets not found')

print()

# ============================================================================
# PRIORITY 4: Milestones - Project ID, Status, Phase, Priority
# ============================================================================
print('='*80)
print('PRIORITY 4: Milestones - Project ID, Status, Phase, Priority')
print('='*80)
print()

if 'Milestones' in wb.sheetnames:
    ws = wb['Milestones']

    # C column - Unique_ID (Project ID)
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('C2:C5000')
    print('  C2:C5000 (Unique_ID): Added dropdown')
    total_validations += 1

    # E column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('E2:E5000')
    print('  E2:E5000 (Status): Added dropdown')
    total_validations += 1

    # F column - Phase
    dv_phase = DataValidation(
        type="list",
        formula1="=Config_Lists!$C$2:$C$5",
        allow_blank=True
    )
    ws.add_data_validation(dv_phase)
    dv_phase.add('F2:F5000')
    print('  F2:F5000 (Phase): Added dropdown')
    total_validations += 1

    # G column - Priority
    dv_priority = DataValidation(
        type="list",
        formula1="=List_Priority",
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add('G2:G5000')
    print('  G2:G5000 (Priority): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Milestones not found')

print()

# ============================================================================
# PRIORITY 5: Events - Project ID and Status
# ============================================================================
print('='*80)
print('PRIORITY 5: Events - Project ID and Status')
print('='*80)
print()

if 'Events' in wb.sheetnames:
    ws = wb['Events']

    # C column - Unique_ID (Project ID)
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('C2:C2000')
    print('  C2:C2000 (Unique_ID): Added dropdown')
    total_validations += 1

    # F column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('F2:F2000')
    print('  F2:F2000 (Status): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Events not found')

print()

# ============================================================================
# PRIORITY 6: Stakeholders - Location Country
# ============================================================================
print('='*80)
print('PRIORITY 6: Stakeholders - Location Country')
print('='*80)
print()

if 'Stakeholders' in wb.sheetnames:
    ws = wb['Stakeholders']

    # F column - Location_Country
    dv_country = DataValidation(
        type="list",
        formula1="=List_CountryCodes",
        allow_blank=True
    )
    ws.add_data_validation(dv_country)
    dv_country.add('F2:F1000')
    print('  F2:F1000 (Location_Country): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Stakeholders not found')

print()

# ============================================================================
# PRIORITY 7: Risk_Register - Project ID, Status, Priority
# ============================================================================
print('='*80)
print('PRIORITY 7: Risk_Register - Project ID, Status, Priority')
print('='*80)
print()

if 'Risk_Register' in wb.sheetnames:
    ws = wb['Risk_Register']

    # B column - Unique_ID (Project ID)
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('B2:B2000')
    print('  B2:B2000 (Unique_ID): Added dropdown')
    total_validations += 1

    # E column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('E2:E2000')
    print('  E2:E2000 (Status): Added dropdown')
    total_validations += 1

    # F column - Priority
    dv_priority = DataValidation(
        type="list",
        formula1="=List_Priority",
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add('F2:F2000')
    print('  F2:F2000 (Priority): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Risk_Register not found')

print()

# ============================================================================
# PRIORITY 8: Decision_Log - Project ID and Status
# ============================================================================
print('='*80)
print('PRIORITY 8: Decision_Log - Project ID and Status')
print('='*80)
print()

if 'Decision_Log' in wb.sheetnames:
    ws = wb['Decision_Log']

    # B column - Unique_ID (Project ID)
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('B2:B2000')
    print('  B2:B2000 (Unique_ID): Added dropdown')
    total_validations += 1

    # E column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('E2:E2000')
    print('  E2:E2000 (Status): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Decision_Log not found')

print()

# ============================================================================
# PRIORITY 9: Project_Audiences - Project ID
# ============================================================================
print('='*80)
print('PRIORITY 9: Project_Audiences - Project ID')
print('='*80)
print()

if 'Project_Audiences' in wb.sheetnames:
    ws = wb['Project_Audiences']

    # A column - Project_ID
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('A2:A1000')
    print('  A2:A1000 (Project_ID): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Project_Audiences not found')

print()

# ============================================================================
# PRIORITY 10: Project_Technologies - Project ID
# ============================================================================
print('='*80)
print('PRIORITY 10: Project_Technologies - Project ID')
print('='*80)
print()

if 'Project_Technologies' in wb.sheetnames:
    ws = wb['Project_Technologies']

    # A column - Project_ID
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('A2:A1000')
    print('  A2:A1000 (Project_ID): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Project_Technologies not found')

print()

# ============================================================================
# PRIORITY 11: Project_Deliverables - Project ID and Status
# ============================================================================
print('='*80)
print('PRIORITY 11: Project_Deliverables - Project ID and Status')
print('='*80)
print()

if 'Project_Deliverables' in wb.sheetnames:
    ws = wb['Project_Deliverables']

    # A column - Project_ID
    dv_project = DataValidation(
        type="list",
        formula1="=Master_Projects[Unique_ID]",
        allow_blank=True
    )
    ws.add_data_validation(dv_project)
    dv_project.add('A2:A2000')
    print('  A2:A2000 (Project_ID): Added dropdown')
    total_validations += 1

    # D column - Status
    dv_status = DataValidation(
        type="list",
        formula1="=List_Status",
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('D2:D2000')
    print('  D2:D2000 (Status): Added dropdown')
    total_validations += 1
else:
    print('  WARNING: Project_Deliverables not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v6.2...')
wb.save(output_file)
wb.close()

print('='*80)
print('DATA VALIDATION COMPLETE!')
print('='*80)
print()
print(f'Total validation rules added: {total_validations}')
print()
print('Validations added:')
print('  1. Project_Spotlight B2 - Project selection (CRITICAL)')
print('  2. Master_Projects - Status, Priority, NCE_Status')
print('  3. Country_Budgets - Project ID, Country Code')
print('  4. Milestones - Project ID, Status, Phase, Priority')
print('  5. Events - Project ID, Status')
print('  6. Stakeholders - Location Country')
print('  7. Risk_Register - Project ID, Status, Priority')
print('  8. Decision_Log - Project ID, Status')
print('  9. Project_Audiences - Project ID')
print(' 10. Project_Technologies - Project ID')
print(' 11. Project_Deliverables - Project ID, Status')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v6.2 in Excel')
print('  2. Test Project_Spotlight B2 - should have dropdown')
print('  3. Click any validated cell - should see dropdown arrow')
print('  4. Try entering data - dropdowns should work!')
print()
print('If validations work correctly:')
print('  - Use v6.2 as your base')
print('  - If you want FILTER functions, apply them to v6.2 manually')
