"""
Check Master_Projects column B values
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("MASTER_PROJECTS - ALL ID COLUMNS")
print("="*80)

ws = wb['Master_Projects']

print(f"\n{'Row':<5} {'A: Fiscal_Year':<20} {'B: Project_Unique_ID':<25}")
print("-"*50)

for row in range(2, 5):
    fiscal_year = ws.cell(row, 1).value or ""
    unique_id = ws.cell(row, 2).value or ""

    if fiscal_year or unique_id:
        print(f"{row:<5} {str(fiscal_year):<20} {str(unique_id):<25}")

print("\n" + "="*80)
print("COUNTRY_BUDGETS - UNIQUE_ID COLUMN")
print("="*80)

ws_cb = wb['Country_Budgets']

print(f"\n{'Row':<5} {'B: Unique_ID':<20} {'E: Country_Name':<25}")
print("-"*50)

for row in range(2, 6):
    unique_id = ws_cb.cell(row, 2).value or ""
    country_name = ws_cb.cell(row, 5).value or ""

    if unique_id or country_name:
        print(f"{row:<5} {str(unique_id):<20} {str(country_name):<25}")

print("\n" + "="*80)
print("STATUS")
print("="*80)

mp_unique = ws.cell(2, 2).value
cb_unique = ws_cb.cell(2, 2).value
cb_name = ws_cb.cell(2, 5).value

print(f"\nMaster_Projects B2 (Project_Unique_ID): '{mp_unique}'")
print(f"Country_Budgets B2 (Unique_ID): '{cb_unique}'")
print(f"Country_Budgets E2 (Country_Name): '{cb_name}'")

if mp_unique == cb_unique:
    print("\nOK - Project IDs MATCH!")
else:
    print(f"\nISSUE - IDs don't match: '{mp_unique}' != '{cb_unique}'")

if cb_name:
    print("OK - Country names are populated!")
else:
    print("ISSUE - Country names are still empty in column E")
