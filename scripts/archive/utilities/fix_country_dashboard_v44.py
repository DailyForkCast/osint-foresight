"""
Fix Country_Dashboard ULO% formula and Project ID zero - v44
"""
import openpyxl

print("="*80)
print("FIXING COUNTRY_DASHBOARD - V44")
print("="*80)

# Load v43
print("\nLoading v43...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')

ws_country = wb['Country_Dashboard']

print("\n" + "="*80)
print("CHECKING CURRENT STRUCTURE")
print("="*80)

print("\nFirst 20 rows, first 10 columns:")
for row in range(1, 21):
    row_data = []
    for col in range(1, 11):
        cell = ws_country.cell(row, col)
        value = cell.value
        if value:
            # Clean for display
            clean_value = str(value)[:30].encode('ascii', 'ignore').decode('ascii')
            row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}: {clean_value}")
    if row_data:
        print(f"  Row {row}: {', '.join(row_data[:3])}")

print("\nLooking for ULO% and Project ID...")

# Find ULO% cell
ulo_percent_cell = None
for row in range(1, 30):
    for col in range(1, 15):
        cell_value = ws_country.cell(row, col).value
        if cell_value and 'ULO%' in str(cell_value):
            ulo_percent_cell = (row, col)
            print(f"  Found 'ULO%' at {openpyxl.utils.get_column_letter(col)}{row}")
            # Check formula in next row or same row
            for check_row in range(row, row + 3):
                for check_col in range(col, col + 2):
                    check_cell = ws_country.cell(check_row, check_col)
                    if check_cell.data_type == 'f':
                        print(f"    Formula at {openpyxl.utils.get_column_letter(check_col)}{check_row}: {str(check_cell.value)[:80]}")

# Find Project ID
project_id_cell = None
for row in range(1, 30):
    for col in range(1, 15):
        cell_value = ws_country.cell(row, col).value
        if cell_value and 'Project' in str(cell_value) and 'ID' in str(cell_value):
            project_id_cell = (row, col)
            print(f"  Found 'Project ID' at {openpyxl.utils.get_column_letter(col)}{row}")
            # Check formula in next few rows
            for check_row in range(row, row + 5):
                for check_col in range(col, col + 2):
                    check_cell = ws_country.cell(check_row, check_col)
                    if check_cell.data_type == 'f':
                        print(f"    Formula at {openpyxl.utils.get_column_letter(check_col)}{check_row}: {str(check_cell.value)[:80]}")

print("\n" + "="*80)
print("IDENTIFYING FIXES NEEDED")
print("="*80)

# We need to see the actual formulas to fix them
# Let me check more systematically

print("\nAll formulas in rows 1-20:")
for row in range(1, 21):
    for col in range(1, 15):
        cell = ws_country.cell(row, col)
        if cell.data_type == 'f':
            col_letter = openpyxl.utils.get_column_letter(col)
            formula = str(cell.value)[:80]
            print(f"  {col_letter}{row}: {formula}...")

print("\n" + "="*80)
print("MANUAL INSPECTION COMPLETE")
print("="*80)

print("\nPlease review the output above.")
print("I'll create a proper fix script once we identify the exact cells to update.")

wb.close()
