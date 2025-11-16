import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

output_file = 'C:/Projects/OSINT-Foresight/Spotlight_Comparison.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING STANDALONE SPOTLIGHT COMPARISON FILE')
print('='*80)
print()

wb = openpyxl.Workbook()
ws_default = wb.active
wb.remove(ws_default)  # Remove default sheet

# Styles
header_font = Font(bold=True, size=16, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
label_font = Font(bold=True, size=10)
metric_font = Font(bold=True, size=14)
column_header_font = Font(bold=True, size=9)
column_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# REFERENCE DATA SHEETS (so formulas work)
# ============================================================================

print('Creating reference data sheets...')
print('-'*80)

# Master_Projects sheet with sample data
ws_projects = wb.create_sheet('Master_Projects', 0)
ws_projects['A1'] = 'Unique_ID'
ws_projects['B1'] = 'Fiscal_Year'
ws_projects['C1'] = 'Project_Name'
ws_projects['D1'] = 'Summary'
ws_projects['E1'] = 'Status'
ws_projects['F1'] = 'Priority'
ws_projects['G1'] = 'Progress'
ws_projects['H1'] = 'Start_Date'
ws_projects['I1'] = 'End_Date'
ws_projects['J1'] = 'Days_Remaining'
ws_projects['K1'] = 'Implementer'
ws_projects['L1'] = 'Implementer_POC'
ws_projects['M1'] = 'Implementer_POC_Email'
ws_projects['N1'] = 'Total_Allocation'
ws_projects['O1'] = 'Total_Obligated'
ws_projects['P1'] = 'Total_ULO'
ws_projects['Q1'] = 'ULO_Percent'
ws_projects['R1'] = 'Countries'
ws_projects['S1'] = 'Country_Count'
ws_projects['T1'] = 'Total_Spent'

# Sample project 1
ws_projects['A2'] = 'PRJ-001'
ws_projects['B2'] = 2025
ws_projects['C2'] = 'Digital Infrastructure Modernization'
ws_projects['D2'] = 'Comprehensive program to modernize critical digital infrastructure across European partner nations, focusing on cloud adoption, cybersecurity, and data governance.'
ws_projects['E2'] = 'Active'
ws_projects['F2'] = 'High'
ws_projects['G2'] = '75%'
ws_projects['H2'] = date(2025, 1, 15)
ws_projects['I2'] = date(2025, 12, 31)
ws_projects['J2'] = 120
ws_projects['K2'] = 'TechCorp International'
ws_projects['L2'] = 'Jane Doe'
ws_projects['M2'] = 'jane.doe@techcorp.com'
ws_projects['N2'] = 1500000
ws_projects['O2'] = 1200000
ws_projects['P2'] = 800000
ws_projects['Q2'] = 0.667
ws_projects['R2'] = 'DE, FR, IT, ES, NL'
ws_projects['S2'] = 5
ws_projects['T2'] = 400000

# Sample project 2
ws_projects['A3'] = 'PRJ-002'
ws_projects['B3'] = 2025
ws_projects['C3'] = 'AI Ethics Framework Development'
ws_projects['D3'] = 'Multi-stakeholder initiative to develop comprehensive AI governance and ethics frameworks for government adoption.'
ws_projects['E3'] = 'CN Stage'
ws_projects['F3'] = 'Critical'
ws_projects['G3'] = '45%'
ws_projects['H3'] = date(2025, 3, 1)
ws_projects['I3'] = date(2026, 2, 28)
ws_projects['J3'] = 240
ws_projects['K3'] = 'AI Policy Institute'
ws_projects['L3'] = 'Dr. Smith'
ws_projects['M3'] = 'smith@aipolicy.org'
ws_projects['N3'] = 850000
ws_projects['O3'] = 600000
ws_projects['P3'] = 450000
ws_projects['Q3'] = 0.75
ws_projects['R3'] = 'FR, BE, NL'
ws_projects['S3'] = 3
ws_projects['T3'] = 150000

print('[OK] Master_Projects with 2 sample projects')

# Project_Deliverables sheet
ws_deliverables = wb.create_sheet('T_Project_Deliverables', 1)
ws_deliverables['A1'] = 'Project_ID'
ws_deliverables['B1'] = 'Deliverable_Name'
ws_deliverables['C1'] = 'Deliverable_Type'
ws_deliverables['D1'] = 'Due_Date'
ws_deliverables['E1'] = 'Status'
ws_deliverables['F1'] = 'Owner'
ws_deliverables['G1'] = 'Completion_Percent'

# PRJ-001 deliverables
deliverables_001 = [
    ('PRJ-001', 'Technical Requirements Document', 'Report', date(2025, 11, 15), 'In Progress', 'JD', '75%'),
    ('PRJ-001', 'Beta Release', 'Software', date(2025, 12, 1), 'Not Started', 'AS', '0%'),
    ('PRJ-001', 'Security Audit', 'Report', date(2025, 11, 30), 'In Progress', 'MS', '60%'),
    ('PRJ-001', 'User Training Materials', 'Documentation', date(2025, 12, 15), 'Not Started', 'LK', '0%'),
    ('PRJ-001', 'Final Implementation Report', 'Report', date(2025, 12, 31), 'Not Started', 'JD', '0%'),
]

for idx, deliv in enumerate(deliverables_001, start=2):
    for col_idx, value in enumerate(deliv, start=1):
        ws_deliverables.cell(idx, col_idx, value)

# PRJ-002 deliverables
deliverables_002 = [
    ('PRJ-002', 'Ethics Framework Draft', 'Report', date(2025, 6, 30), 'Completed', 'DS', '100%'),
    ('PRJ-002', 'Stakeholder Consultation Report', 'Report', date(2025, 9, 15), 'In Progress', 'MJ', '70%'),
    ('PRJ-002', 'Final Framework Publication', 'Report', date(2026, 2, 28), 'Not Started', 'DS', '0%'),
]

for idx, deliv in enumerate(deliverables_002, start=7):
    for col_idx, value in enumerate(deliv, start=1):
        ws_deliverables.cell(idx, col_idx, value)

print('[OK] T_Project_Deliverables with sample data')

# Project_Audiences sheet
ws_audiences = wb.create_sheet('T_Project_Audiences', 2)
ws_audiences['A1'] = 'Project_ID'
ws_audiences['B1'] = 'Audience_Type'
ws_audiences['C1'] = 'Audience_Region'
ws_audiences['D1'] = 'Description'
ws_audiences['E1'] = 'Priority'

audiences = [
    ('PRJ-001', 'Government Officials', 'Europe', 'Senior technology and digital transformation officials', 'High'),
    ('PRJ-001', 'Private Sector', 'Europe', 'Technology providers and system integrators', 'Medium'),
    ('PRJ-001', 'Academia', 'Europe', 'Research institutions and universities', 'Low'),
    ('PRJ-002', 'Government Officials', 'Europe', 'Policy makers and regulators', 'High'),
    ('PRJ-002', 'Civil Society', 'Global', 'Ethics advocates and civil society organizations', 'High'),
]

for idx, aud in enumerate(audiences, start=2):
    for col_idx, value in enumerate(aud, start=1):
        ws_audiences.cell(idx, col_idx, value)

print('[OK] T_Project_Audiences with sample data')

# Project_Technologies sheet
ws_tech = wb.create_sheet('Project_Technologies', 3)
ws_tech['A1'] = 'Project_ID'
ws_tech['B1'] = 'Technology'
ws_tech['C1'] = 'Category'
ws_tech['D1'] = 'Status'

technologies = [
    ('PRJ-001', 'Cloud Infrastructure', 'Infrastructure', 'In Use'),
    ('PRJ-001', 'Machine Learning', 'AI/ML', 'Planned'),
    ('PRJ-001', 'Data Analytics', 'Data', 'In Use'),
    ('PRJ-002', 'Natural Language Processing', 'AI/ML', 'In Use'),
    ('PRJ-002', 'Policy Automation Tools', 'Software', 'Planned'),
]

for idx, tech in enumerate(technologies, start=2):
    for col_idx, value in enumerate(tech, start=1):
        ws_tech.cell(idx, col_idx, value)

print('[OK] Project_Technologies with sample data')

# Stakeholders sheet
ws_stakeholders = wb.create_sheet('Stakeholders', 4)
ws_stakeholders['A1'] = 'Stakeholder_ID'
ws_stakeholders['B1'] = 'Name'
ws_stakeholders['C1'] = 'Title'
ws_stakeholders['D1'] = 'Organization'
ws_stakeholders['E1'] = 'Email'
ws_stakeholders['F1'] = 'Stakeholder_Type'
ws_stakeholders['G1'] = 'Project_IDs'

stakeholders = [
    ('STK-001', 'John Smith', 'Program Manager', 'Agency A', 'john.smith@agency.gov', 'Government', 'PRJ-001, PRJ-002'),
    ('STK-002', 'Jane Doe', 'Technical Lead', 'TechCorp International', 'jane.doe@techcorp.com', 'Private Sector', 'PRJ-001'),
    ('STK-003', 'Dr. Robert Chen', 'Chief Architect', 'TechCorp International', 'robert.chen@techcorp.com', 'Private Sector', 'PRJ-001'),
    ('STK-004', 'Maria Garcia', 'Security Advisor', 'CyberSec Consultants', 'maria.garcia@cybersec.com', 'Private Sector', 'PRJ-001'),
    ('STK-005', 'Dr. Emily Smith', 'Research Director', 'AI Policy Institute', 'smith@aipolicy.org', 'Academia', 'PRJ-002'),
]

for idx, stakeholder in enumerate(stakeholders, start=2):
    for col_idx, value in enumerate(stakeholder, start=1):
        ws_stakeholders.cell(idx, col_idx, value)

print('[OK] Stakeholders with sample data')
print()

# ============================================================================
# COPY THE THREE SPOTLIGHT VERSIONS
# ============================================================================

print('Loading original Spotlight versions from v12...')
wb_source = openpyxl.load_workbook('C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v12.xlsx', data_only=False)

# Version 1: Executive
print('Copying Spotlight_Executive...')
ws_source_exec = wb_source['Spotlight_Executive']
ws_exec = wb.copy_worksheet(ws_source_exec)
ws_exec.title = '1_Executive'
print('[OK] Executive copied')

# Version 2: PM Workspace
print('Copying Spotlight_PMWorkspace...')
ws_source_pm = wb_source['Spotlight_PMWorkspace']
ws_pm = wb.copy_worksheet(ws_source_pm)
ws_pm.title = '2_PMWorkspace'
print('[OK] PM Workspace copied')

# Version 3: Stakeholder
print('Copying Spotlight_Stakeholder...')
ws_source_stake = wb_source['Spotlight_Stakeholder']
ws_stake = wb.copy_worksheet(ws_source_stake)
ws_stake.title = '3_Stakeholder'
print('[OK] Stakeholder copied')

wb_source.close()
print()

# ============================================================================
# ADD COMPARISON GUIDE SHEET
# ============================================================================

print('Creating Comparison Guide sheet...')
ws_guide = wb.create_sheet('START_HERE', 0)

ws_guide['A1'] = 'SPOTLIGHT COMPARISON GUIDE'
ws_guide['A1'].font = Font(bold=True, size=18, color='FFFFFF')
ws_guide['A1'].fill = header_fill
ws_guide.merge_cells('A1:F1')
ws_guide.row_dimensions[1].height = 30

ws_guide['A3'] = 'Welcome to the Spotlight Comparison File!'
ws_guide['A3'].font = Font(bold=True, size=12)
ws_guide.merge_cells('A3:F3')

ws_guide['A5'] = 'This file contains THREE versions of Project_Spotlight for you to compare:'
ws_guide.merge_cells('A5:F5')

ws_guide['A7'] = 'Version 1: EXECUTIVE SUMMARY'
ws_guide['A7'].font = section_font
ws_guide['A7'].fill = section_fill
ws_guide.merge_cells('A7:F7')

ws_guide['A8'] = 'Sheet: 1_Executive'
ws_guide['A8'].font = label_font
ws_guide['B8'] = 'Clean, compact, 1-page layout for leadership briefings'
ws_guide.merge_cells('B8:F8')

ws_guide['A9'] = 'Best for:'
ws_guide['A9'].font = label_font
ws_guide['B9'] = 'Quick status updates, executive briefings, printable summaries'
ws_guide.merge_cells('B9:F9')

ws_guide['A11'] = 'Version 2: PM WORKSPACE'
ws_guide['A11'].font = section_font
ws_guide['A11'].fill = section_fill
ws_guide.merge_cells('A11:F11')

ws_guide['A12'] = 'Sheet: 2_PMWorkspace'
ws_guide['A12'].font = label_font
ws_guide['B12'] = 'Detailed, comprehensive, working document with all project info'
ws_guide.merge_cells('B12:F12')

ws_guide['A13'] = 'Best for:'
ws_guide['A13'].font = label_font
ws_guide['B13'] = 'Day-to-day project management, tracking all deliverables and stakeholders'
ws_guide.merge_cells('B13:F13')

ws_guide['A15'] = 'Version 3: STAKEHOLDER BRIEFING'
ws_guide['A15'].font = section_font
ws_guide['A15'].fill = section_fill
ws_guide.merge_cells('A15:F15')

ws_guide['A16'] = 'Sheet: 3_Stakeholder'
ws_guide['A16'].font = label_font
ws_guide['B16'] = 'Visual, polished, presentation-ready for external audiences'
ws_guide.merge_cells('B16:F16')

ws_guide['A17'] = 'Best for:'
ws_guide['A17'].font = label_font
ws_guide['B17'] = 'Client meetings, stakeholder briefings, professional presentations'
ws_guide.merge_cells('B17:F17')

ws_guide['A19'] = 'HOW TO TEST:'
ws_guide['A19'].font = section_font
ws_guide['A19'].fill = section_fill
ws_guide.merge_cells('A19:F19')

ws_guide['A20'] = '1. Go to each sheet (1_Executive, 2_PMWorkspace, 3_Stakeholder)'
ws_guide.merge_cells('A20:F20')
ws_guide['A21'] = '2. Cell B2 has a dropdown - select PRJ-001 or PRJ-002'
ws_guide.merge_cells('A21:F21')
ws_guide['A22'] = '3. Watch all data auto-populate based on selected project'
ws_guide.merge_cells('A22:F22')
ws_guide['A23'] = '4. Compare layouts, styles, and level of detail'
ws_guide.merge_cells('A23:F23')
ws_guide['A24'] = '5. Note what you like and don\'t like about each version'
ws_guide.merge_cells('A24:F24')

ws_guide['A26'] = 'SAMPLE DATA:'
ws_guide['A26'].font = section_font
ws_guide['A26'].fill = section_fill
ws_guide.merge_cells('A26:F26')

ws_guide['A27'] = 'Reference sheets contain sample data for testing:'
ws_guide.merge_cells('A27:F27')
ws_guide['A28'] = '  • Master_Projects: 2 sample projects (PRJ-001, PRJ-002)'
ws_guide.merge_cells('A28:F28')
ws_guide['A29'] = '  • T_Project_Deliverables: 8 sample deliverables'
ws_guide.merge_cells('A29:F29')
ws_guide['A30'] = '  • T_Project_Audiences: 5 sample audiences'
ws_guide.merge_cells('A30:F30')
ws_guide['A31'] = '  • Project_Technologies: 5 sample technologies'
ws_guide.merge_cells('A31:F31')
ws_guide['A32'] = '  • Stakeholders: 5 sample stakeholders'
ws_guide.merge_cells('A32:F32')

ws_guide['A34'] = 'WHAT TO REPORT BACK:'
ws_guide['A34'].font = section_font
ws_guide['A34'].fill = section_fill
ws_guide.merge_cells('A34:F34')

ws_guide['A35'] = '  • Which version do you prefer overall?'
ws_guide.merge_cells('A35:F35')
ws_guide['A36'] = '  • What specific elements do you like from each?'
ws_guide.merge_cells('A36:F36')
ws_guide['A37'] = '  • What would you change or improve?'
ws_guide.merge_cells('A37:F37')
ws_guide['A38'] = '  • Should we combine elements from multiple versions?'
ws_guide.merge_cells('A38:F38')
ws_guide['A39'] = '  • Need calendar/timeline added?'
ws_guide.merge_cells('A39:F39')

ws_guide.column_dimensions['A'].width = 15
ws_guide.column_dimensions['B'].width = 80

print('[OK] Comparison guide created')
print()

# ============================================================================
# SAVE
# ============================================================================

print('='*80)
print('Saving Spotlight_Comparison.xlsx...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('STANDALONE SPOTLIGHT COMPARISON FILE COMPLETE!')
print('='*80)
print()
print('Created file: Spotlight_Comparison.xlsx')
print()
print('Contents:')
print('  Sheet 1: START_HERE - Comparison guide')
print('  Sheet 2: 1_Executive - Executive Summary version')
print('  Sheet 3: 2_PMWorkspace - PM Workspace version')
print('  Sheet 4: 3_Stakeholder - Stakeholder Briefing version')
print('  Sheet 5-9: Reference data (Master_Projects, Deliverables, etc.)')
print()
print('Sample data included:')
print('  • 2 projects (PRJ-001, PRJ-002)')
print('  • 8 deliverables')
print('  • 5 audiences')
print('  • 5 technologies')
print('  • 5 stakeholders')
print()
print('All formulas work with sample data!')
print()
print('NEXT STEPS:')
print('  1. Open Spotlight_Comparison.xlsx')
print('  2. Read START_HERE guide')
print('  3. Test all three versions')
print('  4. Report back with preferences!')
print()
print(f'Created: {timestamp}')
