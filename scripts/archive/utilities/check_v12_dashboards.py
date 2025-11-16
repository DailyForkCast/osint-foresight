"""
Check what formulas exist in v12 dashboard sheets
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=False)

print("="*80)
print("V12 DASHBOARD FORMULAS")
print("="*80)

dashboard_sheets = [
    'Portfolio_Dashboard',
    'Country_Dashboard',
    'Regional_Summary',
    'Control'
]

for sheet_name in dashboard_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"\n{sheet_name}: NOT FOUND")
        continue

    ws = wb[sheet_name]

    print(f"\n{'='*80}")
    print(f"{sheet_name.upper()}")
    print(f"{'='*80}")

    formulas_found = []

    # Check first 50 rows
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.data_type == 'f':
                col_letter = openpyxl.utils.get_column_letter(col)
                formula = str(cell.value)
                # Truncate long formulas
                if len(formula) > 80:
                    formula = formula[:77] + "..."
                formulas_found.append(f"  {col_letter}{row}: {formula}")

    if formulas_found:
        print(f"\nFound {len(formulas_found)} formulas:")
        # Show first 30
        for f in formulas_found[:30]:
            print(f)
        if len(formulas_found) > 30:
            print(f"  ... and {len(formulas_found) - 30} more")
    else:
        print("\n  No formulas found")

print("\n" + "="*80)
print("CONTROL SHEET - KEY METRICS")
print("="*80)

ws = wb['Control']
print("\nFormulas in Control sheet (dashboard metrics):")
for row in range(1, min(30, ws.max_row + 1)):
    for col in range(1, min(6, ws.max_column + 1)):
        cell = ws.cell(row, col)
        if cell.data_type == 'f':
            col_letter = openpyxl.utils.get_column_letter(col)
            label_cell = ws.cell(row, col-1).value if col > 1 else ""
            formula = str(cell.value)[:70]
            print(f"  {col_letter}{row} ({label_cell}): {formula}")
