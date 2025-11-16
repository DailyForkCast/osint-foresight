import openpyxl

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'

print('='*80)
print('Fixing Millions Format (adding second comma)')
print('='*80)
print()

print('IMPORTANT: Close the Excel file before running this!')
print()
input('Press Enter when file is closed...')
print()

# Load workbook
print('Loading v5...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

# Fix Portfolio_Dashboard formulas
print('Fixing Portfolio_Dashboard formulas...')
if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']
    updates = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Look for the incorrect format
                if '"$#,##0.0,M"' in str(cell.value):
                    old_formula = cell.value
                    # Replace with correct format (two commas)
                    new_formula = old_formula.replace('"$#,##0.0,M"', '"$#,##0.0,,M"')
                    cell.value = new_formula
                    updates += 1
                    print(f'  Updated {cell.coordinate}:')
                    print(f'    Old: {old_formula}')
                    print(f'    New: {new_formula}')

    print()
    print(f'Fixed {updates} formulas')
else:
    print('WARNING: Portfolio_Dashboard not found')

print()

# Also update Config_Lists while we're at it
print('='*80)
print('Also updating Config_Lists with all countries...')
print('='*80)
print()

# Get all countries from Country_Regions
ws_regions = wb['Country_Regions']
all_countries = {}

for row in range(2, ws_regions.max_row + 1):
    country_code = ws_regions.cell(row, 1).value
    country_name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value

    if country_code:
        all_countries[country_code] = {
            'name': country_name,
            'region': region
        }

print(f'Found {len(all_countries)} countries in Country_Regions')

# Get existing countries from Config_Lists
ws_config = wb['Config_Lists']
existing_countries = set()

for row in range(2, ws_config.max_row + 1):
    country_code = ws_config.cell(row, 4).value
    if country_code:
        existing_countries.add(country_code)

print(f'Found {len(existing_countries)} countries in Config_Lists')

# Find missing countries
missing_countries = []
for code in sorted(all_countries.keys()):
    if code not in existing_countries:
        missing_countries.append(code)

print(f'Missing: {len(missing_countries)} countries')
print()

# Add missing countries
if missing_countries:
    print('Adding missing countries...')
    next_row = ws_config.max_row + 1

    for code in missing_countries:
        ws_config.cell(row=next_row, column=4, value=code)
        ws_config.cell(row=next_row, column=5, value=all_countries[code]['name'])
        ws_config.cell(row=next_row, column=6, value=all_countries[code]['region'])
        print(f'  Added: {code} - {all_countries[code]["name"]}')
        next_row += 1

    print(f'Added {len(missing_countries)} countries')
print()

# Save
print('='*80)
print('Saving...')
wb.save(output_file)
wb.close()

print('='*80)
print('SUCCESS!')
print('='*80)
print()
print('Changes made:')
print('  1. Fixed millions format: "$#,##0.0,M" -> "$#,##0.0,,M"')
print(f'  2. Added {len(missing_countries)} missing countries to Config_Lists')
print()
print(f'Output: {output_file}')
print()
print('The #VALUE! errors should now be fixed!')
print('The formulas now correctly divide by 1,000,000 (millions).')
