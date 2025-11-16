"""
Delete _SETUP and FORMULA_GUIDE sheets - v34
"""
import openpyxl

print("="*80)
print("DELETING _SETUP AND FORMULA_GUIDE SHEETS - V34")
print("="*80)

# Load v33
print("\nLoading v33...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v33.xlsx')

print("\nCurrent sheets:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet_name}")

print("\n" + "="*80)
print("DELETING SHEETS")
print("="*80)

# Delete _SETUP
if '_SETUP' in wb.sheetnames:
    del wb['_SETUP']
    print("\n_SETUP sheet deleted")
else:
    print("\n_SETUP sheet not found")

# Delete FORMULA_GUIDE
if 'FORMULA_GUIDE' in wb.sheetnames:
    del wb['FORMULA_GUIDE']
    print("FORMULA_GUIDE sheet deleted")
else:
    print("FORMULA_GUIDE sheet not found")

print("\n" + "="*80)
print("REMAINING SHEETS")
print("="*80)

print(f"\nTotal sheets: {len(wb.sheetnames)}")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet_name}")

print("\n" + "="*80)
print("SAVING V34")
print("="*80)

wb.save('2025-10-26-Tracker-v34.xlsx')

print("\nOK - v34 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nDeleted sheets:")
print("  - _SETUP (blank)")
print("  - FORMULA_GUIDE (outdated)")

print(f"\nRemaining sheets: {len(wb.sheetnames)}")

print("\nv34 ready!")
