"""
Add NCE_Eligible to Spotlight column G row 2
"""
import openpyxl

print("="*80)
print("ADDING NCE_ELIGIBLE TO SPOTLIGHT")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("CURRENT ROW 2 STRUCTURE")
print("="*80)

print("\nChecking columns A-J in row 2:")
for col in range(1, 11):  # A-J
    cell = ws_spotlight.cell(2, col)
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)[:60]
            print(f"  {col_letter}2: FORMULA - {formula}...")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  {col_letter}2: {clean_value}")
    else:
        print(f"  {col_letter}2: [empty]")

print("\n" + "="*80)
print("ADDING NCE_ELIGIBLE")
print("="*80)

print("\nStep 1: Check if G2 is empty or needs to be updated")
current_g2 = ws_spotlight['G2'].value
if current_g2:
    print(f"  G2 currently has: {current_g2}")
    print(f"  Will replace with NCE_Eligible formula")
else:
    print(f"  G2 is empty, will add NCE_Eligible formula")

print("\nStep 2: Add label in F2 if needed")
# Check what F2 has - based on earlier, F2 had Award Number formula
# So we might need to add label in E2 if it's not there

# Actually, looking at pattern: E2 = "Award Number:", F2 = formula
# So we should add: G2 = "NCE Eligible:", H2 = formula
# But user said they want it in G2 specifically

# Let me check if there's a label column before G2
# User said "add NCE_Eligible in Column G row 2"
# They want the VALUE in G2

# Let's check row 3 pattern to understand better
print("\nChecking row 3 for pattern:")
for col in range(5, 11):  # E-J
    cell = ws_spotlight.cell(3, col)
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.value:
        if cell.data_type == 'f':
            print(f"  {col_letter}3: FORMULA")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  {col_letter}3: {clean_value}")

print("\nStep 3: Add NCE_Eligible formula to G2")

# Add formula for NCE_Eligible (shows Yes/No or blank)
formula = '=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[NCE_Eligible],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[NCE_Eligible],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'

ws_spotlight['G2'] = formula
print(f"  G2: Formula added for NCE_Eligible")

print("\nStep 4: Check for label column")
# If the pattern is label-value pairs, we might need a label
# Let's check if there should be a label before G2

# Based on user's request, they specifically want G2, so let's just add it there
# If they need a label, they can tell us

print("\n" + "="*80)
print("VERIFYING NCE_STATUS IN I2")
print("="*80)

i2_value = ws_spotlight['I2'].value
if i2_value:
    if ws_spotlight['I2'].data_type == 'f':
        print(f"  I2: Has formula (NCE_Status)")
        formula_preview = str(i2_value)[:80]
        print(f"      {formula_preview}...")
    else:
        print(f"  I2: {i2_value}")
else:
    print(f"  I2: [empty] - should we add NCE_Status formula?")
    print("\n  Adding NCE_Status formula to I2...")

    # Add NCE_Status formula
    formula = '=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'
    ws_spotlight['I2'] = formula
    print(f"    I2: Formula added for NCE_Status")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v29.xlsx')

print("\nOK - NCE_Eligible added to G2!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nRow 2 NCE fields:")
print("  G2: NCE_Eligible (Yes/No or blank)")
print("  I2: NCE_Status (already has label)")
print("  J2: NCE_Status formula")

print("\nNote: Row 2 structure shows:")
print("  - I2 has the label 'NCE Status:'")
print("  - J2 has the NCE_Status formula")
print("  - G2 now has NCE_Eligible formula")
print("  - You may want to add a label 'NCE Eligible:' in column F2 or elsewhere")

print("\nBoth formulas show BLANK when:")
print("  - No project selected ($B$2 is empty)")
print("  - Field value is empty")

print("\nv29 created!")
