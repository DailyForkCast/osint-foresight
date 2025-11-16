"""
Check what happened to Portfolio_Dashboard between v28 and v29
"""
import openpyxl

print("="*80)
print("CHECKING PORTFOLIO_DASHBOARD")
print("="*80)

# Load both versions
print("\nLoading v28 and v29...")
wb_v28 = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')
wb_v29 = openpyxl.load_workbook('2025-10-26-Tracker-v29.xlsx')

ws_v28 = wb_v28['Portfolio_Dashboard']
ws_v29 = wb_v29['Portfolio_Dashboard']

print("\n" + "="*80)
print("V28 PORTFOLIO_DASHBOARD HEADERS (ROW 10)")
print("="*80)

print("\nv28 headers:")
for col in range(1, 17):  # A-P
    cell = ws_v28.cell(10, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}10: {clean_value}")

print("\n" + "="*80)
print("V29 PORTFOLIO_DASHBOARD HEADERS (ROW 10)")
print("="*80)

print("\nv29 headers:")
for col in range(1, 17):  # A-P
    cell = ws_v29.cell(10, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}10: {clean_value}")

print("\n" + "="*80)
print("COMPARISON")
print("="*80)

print("\nChecking if they match...")
columns_match = True
for col in range(1, 17):
    v28_value = ws_v28.cell(10, col).value
    v29_value = ws_v29.cell(10, col).value

    if v28_value != v29_value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}10: DIFFERENT")
        print(f"    v28: {v28_value}")
        print(f"    v29: {v29_value}")
        columns_match = False

if columns_match:
    print("  All columns match!")

print("\nAnalysis complete!")
