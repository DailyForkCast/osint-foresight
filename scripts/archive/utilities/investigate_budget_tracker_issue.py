import openpyxl

file_path = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.1.xlsx'

print('='*80)
print('INVESTIGATING BUDGET TRACKER - COUNTRIES NOT LOADING ISSUE')
print('='*80)
print()

wb = openpyxl.load_workbook(file_path, data_only=False)

print('Available sheets:')
for sheet in wb.sheetnames:
    print(f'  - {sheet}')
print()

# Look for budget-related sheets
budget_sheets = [s for s in wb.sheetnames if 'budget' in s.lower() or 'country' in s.lower()]

if budget_sheets:
    print(f'Budget-related sheets found: {budget_sheets}')
    print()

    for sheet_name in budget_sheets:
        ws = wb[sheet_name]
        print('='*80)
        print(f'SHEET: {sheet_name}')
        print('='*80)

        # Show first 10 rows and up to 15 columns
        print('First 10 rows:')
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value is None:
                    value = ''
                row_data.append(str(value)[:20])  # Truncate long values
            print(f'Row {row_idx}: {" | ".join(row_data)}')
        print()

        # Check for data validation (dropdowns)
        if hasattr(ws, 'data_validations') and ws.data_validations.dataValidation:
            print('Data Validations found:')
            for dv in ws.data_validations.dataValidation:
                print(f'  - Type: {dv.type}')
                print(f'  - Formula1: {dv.formula1}')
                print(f'  - Cells: {dv.sqref}')
                print()

wb.close()

print('='*80)
print('Investigation complete. Please describe the issue you\'re seeing.')
print('='*80)
