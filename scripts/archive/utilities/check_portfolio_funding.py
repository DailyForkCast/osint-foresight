"""
Check Portfolio_Dashboard funding formulas showing $0K
"""
import openpyxl

print("="*80)
print("CHECKING PORTFOLIO_DASHBOARD FUNDING FORMULAS")
print("="*80)

# Load v31
print("\nLoading v31...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v31.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("ROWS 1-8, COLUMNS A-F")
print("="*80)

for row in range(1, 9):
    print(f"\nRow {row}:")
    for col in range(1, 7):  # A-F
        cell = ws_portfolio.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)
                print(f"  {col_letter}{row}: {formula}")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
                print(f"  {col_letter}{row}: {clean_value}")

print("\n" + "="*80)
print("CHECKING CONTROL SHEET")
print("="*80)

ws_control = wb['Control']

print("\nControl sheet financial rows:")
for row in range(15, 21):
    label = ws_control.cell(row, 1).value
    formula = ws_control.cell(row, 2).value

    if label:
        clean_label = str(label).encode('ascii', 'ignore').decode('ascii')
        print(f"\nRow {row}: {clean_label}")
        if formula:
            if ws_control.cell(row, 2).data_type == 'f':
                print(f"  B{row}: {formula}")
            else:
                print(f"  B{row}: {formula}")

print("\n" + "="*80)
print("ANALYSIS")
print("="*80)

print("\nTotal Funding (D4):")
d4 = ws_portfolio['D4'].value
print(f"  Formula: {d4}")
print("  Should reference: Control!B15 (Total Proposed) or Control!B16 (Total Allocated)?")

print("\nAt Risk Funding:")
print("  Need to find which cell shows 'At Risk' - checking row 6...")
d6 = ws_portfolio['D6'].value
print(f"  D6: {d6}")

print("\nReady to fix!")
