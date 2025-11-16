"""
Comprehensive audit of v25
Check all sheets, formulas, tables, and identify what's missing or needs work
"""
import openpyxl
import json

print("="*80)
print("COMPREHENSIVE AUDIT OF V25")
print("="*80)

# Load v25
print("\nLoading v25...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v25.xlsx', data_only=False)

audit_results = {
    "sheets_found": [],
    "tables_found": {},
    "formulas_by_sheet": {},
    "issues": [],
    "recommendations": []
}

print("\n" + "="*80)
print("1. SHEETS INVENTORY")
print("="*80)

for sheet_name in wb.sheetnames:
    print(f"\n  - {sheet_name}")
    audit_results["sheets_found"].append(sheet_name)

print("\n" + "="*80)
print("2. TABLES INVENTORY")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.tables:
        for table in ws.tables.values():
            print(f"\n  Sheet: {sheet_name}")
            print(f"    Table: {table.name}")
            print(f"    Range: {table.ref}")
            audit_results["tables_found"][table.name] = {
                "sheet": sheet_name,
                "range": table.ref
            }

print("\n" + "="*80)
print("3. MASTER_PROJECTS SHEET AUDIT")
print("="*80)

ws_mp = wb['Master_Projects']
mp_headers = []
for col in range(1, 31):  # Check first 30 columns
    cell = ws_mp.cell(1, col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        mp_headers.append((col_letter, cell.value))

print("\nColumn headers found:")
for col_letter, header in mp_headers:
    print(f"  {col_letter}: {header}")

# Check for formulas in row 2
print("\nChecking formulas in Master_Projects row 2...")
mp_formulas = []
for col in range(1, 31):
    cell = ws_mp.cell(2, col)
    if cell.data_type == 'f':
        col_letter = openpyxl.utils.get_column_letter(col)
        header = ws_mp.cell(1, col).value
        mp_formulas.append((col_letter, header, str(cell.value)[:60]))

if mp_formulas:
    print("\nFormulas found:")
    for col_letter, header, formula in mp_formulas:
        print(f"  {col_letter} ({header}): {formula}...")
else:
    print("  WARNING: No formulas found in row 2!")
    audit_results["issues"].append("Master_Projects has no formulas in data rows")

print("\n" + "="*80)
print("4. COUNTRY_BUDGETS SHEET AUDIT")
print("="*80)

ws_cb = wb['Country_Budgets']
cb_headers = []
for col in range(1, 15):
    cell = ws_cb.cell(1, col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        cb_headers.append((col_letter, cell.value))

print("\nColumn headers found:")
for col_letter, header in cb_headers:
    print(f"  {col_letter}: {header}")

# Check for formulas in row 2
print("\nChecking formulas in Country_Budgets row 2...")
cb_formulas = []
for col in range(1, 15):
    cell = ws_cb.cell(2, col)
    if cell.data_type == 'f':
        col_letter = openpyxl.utils.get_column_letter(col)
        header = ws_cb.cell(1, col).value
        cb_formulas.append((col_letter, header))

if cb_formulas:
    print("\nFormulas found in columns:")
    for col_letter, header in cb_formulas:
        print(f"  {col_letter}: {header}")
else:
    print("  INFO: No formulas in row 2 (likely all manual entry)")

print("\n" + "="*80)
print("5. PORTFOLIO_DASHBOARD AUDIT")
print("="*80)

ws_port = wb['Portfolio_Dashboard']

# Check headers
print("\nRow 10 headers (should be project list headers):")
port_headers = []
for col in range(1, 15):
    cell = ws_port.cell(10, col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        port_headers.append((col_letter, cell.value))
        print(f"  {col_letter}: {cell.value}")

# Check formulas in row 11
print("\nChecking formulas in row 11 (first project row)...")
port_formulas = []
for col in range(1, 15):
    cell = ws_port.cell(11, col)
    if cell.data_type == 'f':
        col_letter = openpyxl.utils.get_column_letter(col)
        port_formulas.append((col_letter, str(cell.value)[:60]))

if port_formulas:
    print("\nFormulas found:")
    for col_letter, formula in port_formulas:
        print(f"  {col_letter}: {formula}...")
else:
    print("  WARNING: No formulas in row 11!")
    audit_results["issues"].append("Portfolio_Dashboard missing project list formulas")

print("\n" + "="*80)
print("6. SPOTLIGHT_PMWORKSPACE AUDIT")
print("="*80)

ws_spot = wb['Spotlight_PMWorkspace']

# Check Project ID cell
print("\nProject ID selector:")
cell_b2 = ws_spot['B2']
if cell_b2.data_type == 'f':
    print(f"  B2 formula: {cell_b2.value}")
else:
    print(f"  B2 value: {cell_b2.value} (manual entry)")

# Check Award Number / NCE fields (row 2-3)
print("\nAward Number / NCE fields:")
award_fields = [
    ('E2', 'Label'), ('F2', 'Value'),
    ('H2', 'Label'), ('I2', 'Value'),
    ('E3', 'Label'), ('F3', 'Value'),
    ('G3', 'Label'), ('H3', 'Value'),
    ('I3', 'Label'), ('J3', 'Value'),
]
for cell_ref, field_type in award_fields:
    cell = ws_spot[cell_ref]
    if cell.value:
        if cell.data_type == 'f':
            print(f"  {cell_ref} ({field_type}): FORMULA - {str(cell.value)[:50]}...")
        else:
            print(f"  {cell_ref} ({field_type}): {cell.value}")
    else:
        print(f"  {cell_ref} ({field_type}): EMPTY")
        if field_type == 'Value':
            audit_results["issues"].append(f"Spotlight {cell_ref} missing formula")

# Check Deliverables section (row 19-29)
print("\nDeliverables section (rows 19-29):")
deliv_formula = ws_spot.cell(19, 1)
if deliv_formula.data_type == 'f':
    print(f"  A19 formula: {str(deliv_formula.value)[:60]}...")
else:
    print(f"  WARNING: A19 has no formula (deliverables won't populate)")
    audit_results["issues"].append("Spotlight Deliverables section missing formulas")

# Check Stakeholders section (row 33-42)
print("\nStakeholders section (rows 33-42):")
stake_cell = ws_spot.cell(33, 1)
if stake_cell.value:
    print(f"  A33: {stake_cell.value} (likely manual entry)")
else:
    print(f"  A33: EMPTY (ready for manual entry)")
print("  INFO: Stakeholders section cleared - manual entry required")

print("\n" + "="*80)
print("7. CONTROL SHEET AUDIT")
print("="*80)

ws_control = wb['Control']

# Check key metrics
print("\nChecking key metric formulas...")
control_cells = [
    ('B3', 'Current Date'),
    ('B9', 'Total Projects'),
    ('B10', 'Started Projects'),
    ('B15', 'Total Allocated'),
    ('B16', 'Total Obligated'),
]

for cell_ref, metric_name in control_cells:
    cell = ws_control[cell_ref]
    if cell.data_type == 'f':
        print(f"  OK - {cell_ref} ({metric_name}): FORMULA")
    else:
        print(f"  WARNING - {cell_ref} ({metric_name}): NO FORMULA (value: {cell.value})")
        audit_results["issues"].append(f"Control {cell_ref} ({metric_name}) missing formula")

print("\n" + "="*80)
print("8. COUNTRY_DASHBOARD AUDIT")
print("="*80)

ws_country = wb['Country_Dashboard']

# Check country selector
print("\nCountry selector:")
cell_b2 = ws_country['B2']
print(f"  B2: {cell_b2.value} (manual entry - user selects country code)")

# Check country name lookup
cell_d2 = ws_country['D2']
if cell_d2.data_type == 'f':
    print(f"  D2 (Country Name): FORMULA - {str(cell_d2.value)[:50]}...")
else:
    print(f"  WARNING: D2 has no formula for country name lookup")
    audit_results["issues"].append("Country_Dashboard D2 missing country name lookup")

# Check project list (row 12)
print("\nProject list (starting row 12):")
proj_cell = ws_country.cell(12, 1)
if proj_cell.data_type == 'f':
    print(f"  A12 formula: {str(proj_cell.value)[:60]}...")
else:
    print(f"  WARNING: A12 has no formula")
    audit_results["issues"].append("Country_Dashboard missing project list formulas")

print("\n" + "="*80)
print("9. REGIONAL_SUMMARY AUDIT")
print("="*80)

ws_regional = wb['Regional_Summary']

# Check region rows
print("\nRegion formulas (rows 2-7)...")
regions = ['AF', 'EAP', 'EUR', 'NEA', 'SCA', 'WHA']
for idx, region in enumerate(regions, start=2):
    cell_a = ws_regional.cell(idx, 1)
    cell_c = ws_regional.cell(idx, 3)  # Project count column

    print(f"\n  Row {idx} ({region}):")
    if cell_a.value == region:
        print(f"    Region code: OK")
    else:
        print(f"    Region code: {cell_a.value} (should be {region})")

    if cell_c.data_type == 'f':
        print(f"    Project count formula: OK")
    else:
        print(f"    WARNING: No formula in C{idx}")
        audit_results["issues"].append(f"Regional_Summary row {idx} missing formulas")

print("\n" + "="*80)
print("10. DATA VALIDATION / DROPDOWNS")
print("="*80)

print("\nData validation checks:")
print("  - Config_Lists sheet exists: YES" if 'Config_Lists' in wb.sheetnames else "  - Config_Lists sheet: MISSING")

ws_config = wb['Config_Lists']
print("\n  Config_Lists headers:")
for col in range(1, 7):
    header = ws_config.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"    {col_letter}: {header}")

print("\n  INFO: Dropdowns need to be added manually in Excel")
print("  Recommended dropdowns:")
print("    - Master_Projects: Status, Priority, Stage, NCE_Eligible, NCE_Status")
print("    - Country_Budgets: Country_Code")
print("    - Spotlight: Project_Unique_ID selector in B2")

print("\n" + "="*80)
print("11. MISSING TABLES CHECK")
print("="*80)

expected_tables = {
    'T_Master_Projects': 'Master_Projects',
    'T_Country_Budgets': 'Country_Budgets',
    'T_Country_Regions': 'Country_Regions',
    'T_Country_PM_Assignments': 'Country_PM_Assignments',
    'T_Stakeholders': 'Stakeholders',
    'T_Project_Deliverables': 'Project_Deliverables',
    'T_Project_Audiences': 'Project_Audiences',
    'T_Project_Technologies': 'Project_Technologies',
    'T_Project_Documents': 'Project_Documents',
}

print("\nExpected tables:")
for table_name, expected_sheet in expected_tables.items():
    if table_name in audit_results["tables_found"]:
        actual_sheet = audit_results["tables_found"][table_name]["sheet"]
        print(f"  OK - {table_name} (in {actual_sheet})")
    else:
        print(f"  MISSING - {table_name} (should be in {expected_sheet})")
        audit_results["issues"].append(f"Table {table_name} missing from {expected_sheet}")
        audit_results["recommendations"].append(f"Create table {table_name} in {expected_sheet} sheet")

# Save audit results
print("\n" + "="*80)
print("SAVING AUDIT RESULTS")
print("="*80)

with open('audit_v25_results.json', 'w') as f:
    json.dump(audit_results, f, indent=2)

print("\nAudit results saved to: audit_v25_results.json")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print(f"\nSheets found: {len(audit_results['sheets_found'])}")
print(f"Tables found: {len(audit_results['tables_found'])}")
print(f"Issues identified: {len(audit_results['issues'])}")
print(f"Recommendations: {len(audit_results['recommendations'])}")

if audit_results["issues"]:
    print("\n" + "="*80)
    print("ISSUES FOUND")
    print("="*80)
    for i, issue in enumerate(audit_results["issues"], 1):
        print(f"\n{i}. {issue}")

if audit_results["recommendations"]:
    print("\n" + "="*80)
    print("RECOMMENDATIONS")
    print("="*80)
    for i, rec in enumerate(audit_results["recommendations"], 1):
        print(f"\n{i}. {rec}")

print("\n" + "="*80)
print("AUDIT COMPLETE")
print("="*80)
