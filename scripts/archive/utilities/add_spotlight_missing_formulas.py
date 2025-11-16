"""
Add missing formulas to Spotlight_PMWorkspace:
1. Award Number and NCE info (rows 2-3)
2. Stakeholder formulas (rows 33-42)
"""
import openpyxl

print("="*80)
print("ADDING MISSING SPOTLIGHT FORMULAS")
print("="*80)

# Load v21
print("\nLoading v21...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v21.xlsx')

ws = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("STEP 1: ADD AWARD NUMBER AND NCE INFO (ROWS 2-3)")
print("="*80)

# Row 2 - Award Number and NCE Status
print("\nAdding Row 2 formulas...")
ws['E2'] = 'Award Number:'
ws['F2'] = '=IFERROR(INDEX(T_Master_Projects[Award_Number],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['H2'] = 'NCE Status:'
ws['I2'] = '=IFERROR(INDEX(T_Master_Projects[NCE_Status],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

print("  OK - Award Number (E2:F2)")
print("  OK - NCE Status (H2:I2)")

# Row 3 - POP dates and POP Days Remaining
print("\nAdding Row 3 formulas...")
ws['E3'] = 'POP Start:'
ws['F3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Start],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['G3'] = 'POP End:'
ws['H3'] = '=IFERROR(INDEX(T_Master_Projects[POP_End],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

ws['I3'] = 'POP Days Remaining:'
ws['J3'] = '=IFERROR(INDEX(T_Master_Projects[POP_Days_Remaining],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0)),"")'

print("  OK - POP Start (E3:F3)")
print("  OK - POP End (G3:H3)")
print("  OK - POP Days Remaining (I3:J3)")

print("\n" + "="*80)
print("STEP 2: ADD STAKEHOLDER FORMULAS (ROWS 33-42)")
print("="*80)

# Stakeholder section header should be in row 31
# Formulas start in row 33

print("\nAdding stakeholder formulas (rows 33-42, auto-incrementing)...")

for row in range(33, 43):
    idx = row - 32  # 1, 2, 3... 10

    # A: Stakeholder Name
    ws.cell(row, 1).value = f'=IFERROR(INDEX(T_Project_Stakeholders[Stakeholder_Name],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),"")'

    # E: Organization
    ws.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Organization],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'

    # G: Role
    ws.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Role],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'

    # H: Email
    ws.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Email],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'

    # I: Engagement Level
    ws.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Project_Stakeholders[Engagement_Level],AGGREGATE(15,6,ROW(T_Project_Stakeholders[Project_ID])-ROW(INDEX(T_Project_Stakeholders[Project_ID],1,1))+1/(T_Project_Stakeholders[Project_ID]=$B$2),{idx})),\"\"))'

print("  OK - Stakeholder formulas added (rows 33-42)")
print("      Columns: A (Name), E (Organization), G (Role), H (Email), I (Engagement Level)")

# Save
print("\n" + "="*80)
print("SAVING V21...")
print("="*80)

wb.save('2025-10-26-Tracker-v21.xlsx')

print("\nOK - SPOTLIGHT FORMULAS ADDED!")

print("\n" + "="*80)
print("SUMMARY OF CHANGES")
print("="*80)

print("\nRow 2 layout:")
print("  E2: Award Number: | F2: [formula]")
print("  H2: NCE Status: | I2: [formula]")

print("\nRow 3 layout:")
print("  E3: POP Start: | F3: [formula]")
print("  G3: POP End: | H3: [formula]")
print("  I3: POP Days Remaining: | J3: [formula]")

print("\nStakeholder section (rows 33-42):")
print("  A: Stakeholder Name")
print("  E: Organization")
print("  G: Role")
print("  H: Email")
print("  I: Engagement Level")
print("  (Auto-increments, shows up to 10 stakeholders per project)")

print("\nv21 is now complete with all Spotlight formulas!")
