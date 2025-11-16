"""
Check Portfolio_Dashboard column O in v27 to see what data/formatting is there
"""
import openpyxl

print("="*80)
print("CHECKING PORTFOLIO_DASHBOARD COLUMN O")
print("="*80)

# Load v27
print("\nLoading v27...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx', data_only=False)

ws_port = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("COLUMN O ANALYSIS")
print("="*80)

# Check what's in column O
col_o = 15  # Column O is 15

print("\nColumn O contents (first 30 rows):")
for row in range(1, 31):
    cell = ws_port.cell(row, col_o)
    value = cell.value

    if value:
        # Check if it's a formula
        if cell.data_type == 'f':
            print(f"  Row {row}: FORMULA - {str(value)[:60]}...")
        else:
            print(f"  Row {row}: {value}")

# Check for any existing tables in Portfolio_Dashboard
print("\n" + "="*80)
print("EXISTING TABLES")
print("="*80)

if ws_port.tables:
    print("\nTables found in Portfolio_Dashboard:")
    for table_name, table in ws_port.tables.items():
        print(f"  {table_name}: {table.ref}")
else:
    print("\nNo tables found in Portfolio_Dashboard")

# Get dimensions
print("\n" + "="*80)
print("SHEET DIMENSIONS")
print("="*80)

print(f"\nMax row with data: {ws_port.max_row}")
print(f"Max column with data: {ws_port.max_column}")

print("\n" + "="*80)
print("RECOMMENDATION")
print("="*80)

print("\nPlease describe what data is in column O and what kind of")
print("formatting you'd like. For example:")
print("  - Should it be a separate table with its own headers?")
print("  - What columns should the table include?")
print("  - What styling do you want (colors, borders, etc.)?")

print("\nI'll create formatting based on what's there!")
