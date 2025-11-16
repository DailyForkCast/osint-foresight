"""
Add Total_Proposed to Country_Dashboard and Regional_Summary
"""
import openpyxl
from openpyxl.utils import get_column_letter

print("="*80)
print("ADDING TOTAL_PROPOSED TO DASHBOARDS")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

print("\n" + "="*80)
print("PART 1: COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nAdding TOTAL PROPOSED to budget summary...")

# Insert column at position B (between Country and Total Allocated)
# This will shift Total Allocated from A to C, etc.
# Current: A4=Total Allocated, C4=Total Obligated, E4=Total Spent, G4=ULO
# After: A4=Total Proposed, C4=Total Allocated, E4=Total Obligated, G4=Total Spent, I4=ULO

print("\nStep 1: Insert column at position B (will shift everything right)")
ws_country.insert_cols(2, 2)  # Insert 2 columns at position B
print("  Inserted 2 columns at position B")
print("  Old structure: A=Allocated, C=Obligated, E=Spent, G=ULO")
print("  New structure: C=Allocated, E=Obligated, G=Spent, I=ULO")
print("  Will add: A=Proposed")

print("\nStep 2: Update Row 4 headers")
# Move existing headers right
# Old A4 (Total Allocated) is now C4
# Old C4 (Total Obligated) is now E4
# Old E4 (Total Spent) is now G4
# Old G4 (ULO) is now I4

# Add new header
ws_country['A4'] = 'TOTAL PROPOSED'
print("  A4: TOTAL PROPOSED")

print("\nStep 3: Add formula for Total Proposed (A5)")
ws_country['A5'] = '=SUMIF(T_Country_Budgets[Country_Code],$F$2,T_Country_Budgets[Proposed_Amount])'
print("  A5: Sum of proposed amounts for selected country")

print("\nStep 4: Check project list headers (row 11)")
print("  Current headers:")
for col in range(1, 13):
    header = ws_country.cell(11, col).value
    col_letter = get_column_letter(col)
    if header:
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')[:20]
        print(f"    {col_letter}11: {clean_header}")

print("\nStep 5: Add Proposed column to project list")
# The project list headers are in row 11
# Current: A=Project ID, B/C=Project Name (merged?), D=Status, E=Priority, F=Allocated, G=Obligated, H=Spent, I=ULO
# Wait, after inserting 2 columns, everything shifted right
# Need to check actual structure

# Let's just add Proposed column headers and formulas
# Find where the project budget columns start
# Based on the earlier output: E11=Allocated, F11=Obligated, G11=Spent, H11=ULO (AFTER the shift)

print("\n  Checking where to add Proposed column in project list...")
# After shift: Project ID, blank, blank, Project Name, Status, Priority, then budget columns
# Need to find Allocated column and insert before it

# Let's look for Allocated in row 11
allocated_col = None
for col in range(1, 15):
    cell_value = ws_country.cell(11, col).value
    if cell_value and 'allocated' in str(cell_value).lower():
        allocated_col = col
        print(f"  Found 'Allocated' at column {get_column_letter(col)}")
        break

if allocated_col:
    # Insert column before Allocated
    print(f"\n  Inserting Proposed column before Allocated (column {get_column_letter(allocated_col)})")
    ws_country.insert_cols(allocated_col, 1)

    # Add header
    ws_country.cell(11, allocated_col).value = 'Proposed'
    print(f"  Added header: {get_column_letter(allocated_col)}11 = Proposed")

    # Add formulas for rows 12-31
    print(f"\n  Adding formulas to rows 12-31...")
    for row in range(12, 32):
        # Formula to get proposed amount from Country_Budgets
        # Match on Project ID (column A)
        formula = f'=IF(A{row}="","",IFERROR(INDEX(T_Country_Budgets[Proposed_Amount],MATCH(A{row},T_Country_Budgets[Unique_ID],0)),""))'
        ws_country.cell(row, allocated_col).value = formula

    print(f"  Formulas added to column {get_column_letter(allocated_col)}")

print("\nCountry_Dashboard: Total_Proposed added!")

print("\n" + "="*80)
print("PART 2: REGIONAL_SUMMARY")
print("="*80)

ws_regional = wb['Regional_Summary']

print("\nAdding Total_Proposed column...")

print("\nStep 1: Check current structure")
print("  Current headers (row 1):")
for col in range(1, 13):
    header = ws_regional.cell(1, col).value
    col_letter = get_column_letter(col)
    if header:
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')[:20]
        print(f"    {col_letter}1: {clean_header}")

print("\nStep 2: Insert column before Total_Allocated (column F)")
ws_regional.insert_cols(6, 1)  # Insert 1 column at position F
print("  Inserted 1 column at position F")
print("  Old F (Total_Allocated) is now G")

print("\nStep 3: Add header")
ws_regional['F1'] = 'Total_Proposed'
print("  F1: Total_Proposed")

print("\nStep 4: Add formulas for each region (rows 2-7)")
for row in range(2, 8):
    region_code = ws_regional.cell(row, 1).value
    if region_code:
        # Sum proposed amounts for this region
        formula = f'=SUMIF(T_Country_Regions[Region],A{row},T_Country_Budgets[Proposed_Amount])'
        ws_regional.cell(row, 6).value = formula
        print(f"  Row {row} ({region_code}): Formula added")

print("\nRegional_Summary: Total_Proposed added!")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Total_Proposed added to both dashboards!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nCountry_Dashboard:")
print("  - Added TOTAL PROPOSED to budget summary (row 4-5)")
print("  - Added Proposed column to project list (before Allocated)")
print("  - Formulas pull from T_Country_Budgets[Proposed_Amount]")

print("\nRegional_Summary:")
print("  - Added Total_Proposed column (column F, before Total_Allocated)")
print("  - Formulas sum proposed amounts by region")

print("\nv28 updated!")
