"""
Fix Portfolio_Dashboard formulas to show blank instead of 0 - v43
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO BLANKS - V43")
print("="*80)

# Load v42
print("\nLoading v42...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v42.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("CHECKING CURRENT FORMULAS")
print("="*80)

print("\nRow 11 formulas (first data row):")
for col in range(1, 14):
    cell = ws_portfolio.cell(11, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value and cell.data_type == 'f':
        formula = str(cell.value)[:80]
        print(f"  {col_letter}11: {formula}...")

print("\n" + "="*80)
print("FIXING FORMULAS TO SHOW BLANK INSTEAD OF 0")
print("="*80)

print("\nUpdating rows 11-20...")

for row in range(11, 21):
    # Column A: ID - double-check for blank
    ws_portfolio.cell(row, 1).value = f'=IFERROR(IF(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)="","",INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)),"")'

    # Column B: Project Name - only if A is not blank
    ws_portfolio.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Name],ROW()-10)="","",INDEX(T_Master_Projects[Project_Name],ROW()-10)),""))'

    # Column C: Status - only if A is not blank
    ws_portfolio.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Status],ROW()-10)="","",INDEX(T_Master_Projects[Project_Status],ROW()-10)),""))'

    # Column D: Priority - only if A is not blank
    ws_portfolio.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Priority],ROW()-10)="","",INDEX(T_Master_Projects[Project_Priority],ROW()-10)),""))'

    # Column E: Progress - only if A is not blank, check for 0
    ws_portfolio.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Progress],ROW()-10)=0,"",INDEX(T_Master_Projects[Project_Progress],ROW()-10)),""))'

    # Column F: Countries - only if A is not blank
    ws_portfolio.cell(row, 6).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Countries],ROW()-10)="","",INDEX(T_Master_Projects[Countries],ROW()-10)),""))'

    # Column G: Total Proposed - check for 0
    ws_portfolio.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Proposed],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Proposed],ROW()-10)),""))'

    # Column H: Total Allocation - check for 0
    ws_portfolio.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Allocation],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Allocation],ROW()-10)),""))'

    # Column I: Total Obligated - check for 0
    ws_portfolio.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Obligated],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Obligated],ROW()-10)),""))'

    # Column J: Total Spent - check for 0
    ws_portfolio.cell(row, 10).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_Spent],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_Spent],ROW()-10)),""))'

    # Column K: Total ULO - check for 0
    ws_portfolio.cell(row, 11).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Total_ULO],ROW()-10)=0,"",INDEX(T_Master_Projects[Total_ULO],ROW()-10)),""))'

    # Column L: ULO % - check for 0
    ws_portfolio.cell(row, 12).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[ULO_Percent],ROW()-10)=0,"",INDEX(T_Master_Projects[ULO_Percent],ROW()-10)),""))'

    # Column M: Days Remaining - check for 0
    ws_portfolio.cell(row, 13).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[POP_Days_Remaining],ROW()-10)=0,"",INDEX(T_Master_Projects[POP_Days_Remaining],ROW()-10)),""))'

print("  Rows 11-20 updated")

print("\n" + "="*80)
print("SAVING V43")
print("="*80)

wb.save('2025-10-26-Tracker-v43.xlsx')

print("\nOK - v43 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed all formulas in Portfolio_Dashboard (rows 11-20):")
print("  - Column A (ID): Shows blank when empty")
print("  - Column B (Project Name): Shows blank when A is blank or empty")
print("  - Column C (Status): Shows blank when A is blank or empty")
print("  - Column D (Priority): Shows blank when A is blank or empty")
print("  - Column E (Progress): Shows blank when A is blank or 0")
print("  - Column F (Countries): Shows blank when A is blank or empty")
print("  - Column G (Total Proposed): Shows blank when A is blank or 0")
print("  - Column H (Total Allocation): Shows blank when A is blank or 0")
print("  - Column I (Total Obligated): Shows blank when A is blank or 0")
print("  - Column J (Total Spent): Shows blank when A is blank or 0")
print("  - Column K (Total ULO): Shows blank when A is blank or 0")
print("  - Column L (ULO %): Shows blank when A is blank or 0")
print("  - Column M (Days Remaining): Shows blank when A is blank or 0")

print("\nAll columns now show clean blanks instead of zeros!")

print("\nv43 ready!")
