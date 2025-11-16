"""
Fix conditional formatting on empty rows - v48
The issue: Formatting triggers on empty cells
The fix: Add blank checks to all conditional formatting rules
"""
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule, CellIsRule

print("="*80)
print("FIXING EMPTY ROW FORMATTING - V48")
print("="*80)

# Load v47
print("\nLoading v47...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v47.xlsx')

# Define colors
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_text = Font(color='9C0006')
orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_text = Font(color='9C6500')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_text = Font(color='006100')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

print("\n" + "="*80)
print("STRATEGY")
print("="*80)
print("\nThe problem: Empty cells are triggering conditional formatting")
print("The solution: Clear all existing formatting and re-add with blank checks")
print("\nClearing all conditional formatting from affected sheets...")

# Clear conditional formatting from problem sheets
wb['Country_Budgets'].conditional_formatting._cf_rules.clear()
wb['Milestones'].conditional_formatting._cf_rules.clear()
wb['Calendar_Todo'].conditional_formatting._cf_rules.clear()
wb['Project_Deliverables'].conditional_formatting._cf_rules.clear()

print("  Cleared: Country_Budgets, Milestones, Calendar_Todo, Project_Deliverables")

print("\n" + "="*80)
print("1. COUNTRY_BUDGETS - FIX ULO % AND MY_COUNTRY")
print("="*80)

ws_budgets = wb['Country_Budgets']

# ULO % Alert (M2:M1001) - Red if > 75% AND not blank
print("\nAdding ULO % alert with blank check (M2:M1001)...")
rule = FormulaRule(
    formula=['AND($M2<>"",$M2>0.75)'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_budgets.conditional_formatting.add('M2:M1001', rule)
print("  Added: ULO % > 75% = Red (only if cell has data)")

# My_Country = TRUE highlight (C2:C1001) - Green
print("\nAdding My_Country highlight (C2:C1001)...")
rule = CellIsRule(
    operator='equal',
    formula=['TRUE'],
    stopIfTrue=True,
    fill=light_green_fill
)
ws_budgets.conditional_formatting.add('C2:C1001', rule)
print("  Added: My_Country = TRUE = Light Green")

print("\n" + "="*80)
print("2. MILESTONES - FIX OVERDUE AND UPCOMING")
print("="*80)

ws_milestones = wb['Milestones']

# Overdue milestones - Red if Due_Date < TODAY and Status != Complete AND not blank
print("\nAdding overdue milestone alert with blank check...")
rule = FormulaRule(
    formula=['AND($D2<>"",$D2<TODAY(),$E2<>"Complete")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_milestones.conditional_formatting.add('A2:J100', rule)
print("  Added: Overdue milestones = Red (only if Due_Date has data)")

# Upcoming milestones - Yellow if within 30 days AND not blank
print("\nAdding upcoming milestone alert with blank check...")
rule = FormulaRule(
    formula=['AND($D2<>"",$D2>=TODAY(),$D2<=TODAY()+30,$E2<>"Complete")'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_milestones.conditional_formatting.add('A2:J100', rule)
print("  Added: Upcoming milestones = Yellow (only if Due_Date has data)")

print("\n" + "="*80)
print("3. CALENDAR_TODO - FIX OVERDUE AND DUE SOON")
print("="*80)

ws_calendar = wb['Calendar_Todo']

# Overdue tasks - Red if Due_Date < TODAY and Status != Completed AND not blank
print("\nAdding overdue task alert with blank check...")
rule = FormulaRule(
    formula=['AND($D2<>"",$D2<TODAY(),$F2<>"Completed")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_calendar.conditional_formatting.add('A2:H100', rule)
print("  Added: Overdue tasks = Red (only if Due_Date has data)")

# Due soon (within 7 days) - Yellow AND not blank
print("\nAdding due soon alert with blank check...")
rule = FormulaRule(
    formula=['AND($D2<>"",$D2>=TODAY(),$D2<=TODAY()+7,$F2<>"Completed")'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_calendar.conditional_formatting.add('A2:H100', rule)
print("  Added: Due within 7 days = Yellow (only if Due_Date has data)")

print("\n" + "="*80)
print("4. PROJECT_DELIVERABLES - FIX OVERDUE AND STATUS")
print("="*80)

ws_deliverables = wb['Project_Deliverables']

# Overdue deliverables - Red if Due_Date < TODAY and Status != Completed AND not blank
print("\nAdding overdue deliverable alert with blank check...")
rule = FormulaRule(
    formula=['AND($D2<>"",$D2<TODAY(),$E2<>"Completed")'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_deliverables.conditional_formatting.add('A2:H100', rule)
print("  Added: Overdue deliverables = Red (only if Due_Date has data)")

# Status colors (E2:E100)
print("\nAdding Status color coding (E2:E100)...")
# Completed = Green
rule = CellIsRule(
    operator='equal',
    formula=['"Completed"'],
    stopIfTrue=True,
    fill=green_fill,
    font=green_text
)
ws_deliverables.conditional_formatting.add('E2:E100', rule)

# In Progress = Yellow
rule = CellIsRule(
    operator='equal',
    formula=['"In Progress"'],
    stopIfTrue=True,
    fill=yellow_fill
)
ws_deliverables.conditional_formatting.add('E2:E100', rule)
print("  Added: Status colors (Completed=Green, In Progress=Yellow)")

print("\n" + "="*80)
print("SAVING V48")
print("="*80)

wb.save('2025-10-26-Tracker-v48.xlsx')

print("\nOK - v48 created!")

print("\n" + "="*80)
print("SUMMARY - EMPTY ROW FORMATTING FIXED")
print("="*80)

print("\nFixed conditional formatting to ignore empty rows:")

print("\n>> Country_Budgets:")
print("   - ULO % > 75% = Red (only if cell has data)")
print("   - My_Country = TRUE = Light Green")

print("\n>> Milestones:")
print("   - Overdue (Due Date < Today AND Status != Complete) = Red")
print("   - Due within 30 days = Yellow")
print("   - ONLY if Due_Date column has data")

print("\n>> Calendar_Todo:")
print("   - Overdue tasks = Red")
print("   - Due within 7 days = Yellow")
print("   - ONLY if Due_Date column has data")

print("\n>> Project_Deliverables:")
print("   - Overdue deliverables = Red")
print("   - Status: Completed=Green, In Progress=Yellow")
print("   - ONLY if Due_Date column has data")

print("\nAll other sheets (Portfolio_Dashboard, Country_Dashboard, etc.)")
print("already have proper formatting - no changes needed")

print("\nv48 ready - no more red on empty rows!")
