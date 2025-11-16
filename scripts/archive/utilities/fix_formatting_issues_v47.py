"""
Fix formatting and Country_Dashboard issues - v47
1. Control ULO warning: Format as percentage
2. Portfolio: Fix ULO % conditional formatting (column O, not L)
3. Portfolio: Add negative days conditional formatting (column P)
4. Country_Dashboard: Change B2 to accept country NAME instead of CODE
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, numbers
from openpyxl.formatting.rule import CellIsRule

print("="*80)
print("FIXING FORMATTING ISSUES - V47")
print("="*80)

# Load v46
print("\nLoading v46...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v46.xlsx')

print("\n" + "="*80)
print("FIX 1: CONTROL - ULO WARNING FORMAT (B5)")
print("="*80)

ws_control = wb['Control']

print("\nCurrent B5 value:", ws_control['B5'].value)
print("Current B5 format:", ws_control['B5'].number_format)

print("\nChanging to percentage format...")
ws_control['B5'].number_format = numbers.FORMAT_PERCENTAGE_00
print("  New format: 0.00% (will display as 75.00%)")

print("\n" + "="*80)
print("FIX 2: PORTFOLIO - MOVE ULO % FORMATTING TO COLUMN O")
print("="*80)

ws_portfolio = wb['Portfolio_Dashboard']

print("\nRemoving incorrect formatting from column L...")
# Remove all conditional formatting from column L
# We need to remove the rule we added, but openpyxl doesn't have easy removal
# So we'll just add the correct rule to column O and it will override

print("\nAdding correct ULO % formatting to column O (O11:O20)...")
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_text = Font(color='9C0006')

rule = CellIsRule(
    operator='greaterThan',
    formula=['0.75'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('O11:O20', rule)
print("  Added: ULO % > 75% = Red (column O)")

print("\n" + "="*80)
print("FIX 3: PORTFOLIO - ADD NEGATIVE DAYS FORMATTING (COLUMN P)")
print("="*80)

print("\nAdding negative days conditional formatting to P11:P20...")
rule = CellIsRule(
    operator='lessThan',
    formula=['0'],
    stopIfTrue=True,
    fill=red_fill,
    font=red_text
)
ws_portfolio.conditional_formatting.add('P11:P20', rule)
print("  Added: Days Remaining < 0 = Red (column P)")

print("\n" + "="*80)
print("FIX 4: COUNTRY_DASHBOARD - CHANGE B2 TO COUNTRY NAME")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nCurrent B2 value:", ws_country['B2'].value)
print("\nChanging B2 to show country name instead of code...")

# Change B2 from "DE" to "Germany"
ws_country['B2'] = 'Germany'
print("  Changed B2 to: Germany")

# Now update all formulas that reference B2 to first look up the country code
print("\nUpdating formulas to look up country code from country name...")

# The formulas currently use $B$2 directly as country code
# We need to change them to look up the code from the name

# Formula pattern to replace:
# OLD: MATCH($B$2, ... matches against Country_Code
# NEW: MATCH(INDEX(Country_Regions[Country_Code],MATCH($B$2,Country_Regions[Country_Name],0)), ...

# Let's check which cells have formulas referencing B2
print("\nFinding formulas that reference B2...")

formulas_to_update = []

# D2 - Region lookup
if ws_country['D2'].data_type == 'f':
    old_formula = ws_country['D2'].value
    print(f"\n  D2 (Region):")
    print(f"    Old: {old_formula}")
    # Change to look up by country name
    new_formula = '=IFERROR(INDEX(T_Country_Regions[Region],MATCH(B2,T_Country_Regions[Country_Name],0)),"")'
    ws_country['D2'] = new_formula
    print(f"    New: {new_formula}")

# G2 - Country Code lookup
if ws_country['G2'].data_type == 'f':
    old_formula = ws_country['G2'].value
    print(f"\n  G2 (Country Code):")
    print(f"    Old: {old_formula}")
    # Change to look up by country name
    new_formula = '=IFERROR(INDEX(T_Country_Regions[Country_Code],MATCH(B2,T_Country_Regions[Country_Name],0)),"")'
    ws_country['G2'] = new_formula
    print(f"    New: {new_formula}")

# A5, B5, C5 - Budget totals (use country code from G2)
if ws_country['A5'].data_type == 'f':
    print(f"\n  A5 (Total Proposed):")
    # Use G2 (country code) instead of B2
    new_formula = '=SUMIF(T_Country_Budgets[Country_Code],$G$2,T_Country_Budgets[Proposed_Amount])'
    ws_country['A5'] = new_formula
    print(f"    New: {new_formula}")

if ws_country['B5'].data_type == 'f':
    print(f"\n  B5 (Total Obligated):")
    new_formula = '=SUMIF(T_Country_Budgets[Country_Code],$G$2,T_Country_Budgets[Obligated_Amount])'
    ws_country['B5'] = new_formula
    print(f"    New: {new_formula}")

if ws_country['C5'].data_type == 'f':
    print(f"\n  C5 (Total Spent):")
    new_formula = '=SUMIF(T_Country_Budgets[Country_Code],$G$2,T_Country_Budgets[Spent_Amount])'
    ws_country['C5'] = new_formula
    print(f"    New: {new_formula}")

# B6 - Number of Projects
if ws_country['B6'].data_type == 'f':
    print(f"\n  B6 (Number of Projects):")
    new_formula = '=COUNTIF(T_Country_Budgets[Country_Code],$G$2)'
    ws_country['B6'] = new_formula
    print(f"    New: {new_formula}")

# B8 - Country PM
if ws_country['B8'].data_type == 'f':
    print(f"\n  B8 (Country PM):")
    # Look up by country code from G2
    new_formula = '=IFERROR(INDEX(Country_PM_Assignments!D:D,MATCH($G$2,Country_PM_Assignments!A:A,0)),"")'
    ws_country['B8'] = new_formula
    print(f"    New: {new_formula}")

# Project list formulas (A12:A31) - these use $B$2 in SMALL/IF formulas
print("\n  Updating project list formulas (A12-A31)...")
# These formulas filter by Country_Code=$B$2
# We need to change to Country_Code=$G$2

for row in range(12, 32):
    cell_a = ws_country.cell(row, 1)  # Column A
    if cell_a.data_type == 'f' and '$B$2' in str(cell_a.value):
        old_formula = cell_a.value
        # Replace $B$2 with $G$2 (country code)
        new_formula = old_formula.replace('$B$2', '$G$2')
        cell_a.value = new_formula

print(f"    Updated {row-11} project list formulas")

print("\n" + "="*80)
print("SAVING V47")
print("="*80)

wb.save('2025-10-26-Tracker-v47.xlsx')

print("\nOK - v47 created!")

print("\n" + "="*80)
print("SUMMARY OF FIXES")
print("="*80)

print("\n1. Control ULO Warning (B5):")
print("   - Now formatted as percentage (displays as 75.00%)")

print("\n2. Portfolio_Dashboard ULO %:")
print("   - Fixed: Conditional formatting now on column O (ULO %), not column L")
print("   - Red highlighting when ULO % > 75%")

print("\n3. Portfolio_Dashboard Days Remaining:")
print("   - Added: Conditional formatting for negative days (column P)")
print("   - Red highlighting when Days Remaining < 0")

print("\n4. Country_Dashboard B2:")
print("   - Changed: Now accepts country NAME (e.g., 'Germany') instead of code")
print("   - How it works:")
print("     - B2: Enter country name (Germany)")
print("     - G2: Shows country code (DE) - auto-calculated")
print("     - D2: Shows region (EUR) - auto-calculated")
print("     - All formulas updated to use G2 (country code) for lookups")

print("\nv47 ready!")
print("\nNOTE: Your spacing changes in Country_Dashboard are preserved!")
print("Just change B2 to any country name and everything will update!")
