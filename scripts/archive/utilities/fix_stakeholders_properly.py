import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v2.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v3-FIXED.xlsx'

print('='*80)
print('PROPER FIX: Decision_Log, Risk_Register, and Stakeholders')
print('='*80)
print()

# Load workbook
print('Loading v2 (last known good)...')
wb = openpyxl.load_workbook(input_file, data_only=False)

# ============================================================================
# PART 1: Update Decision_Log IDs (NO TABLE ISSUES)
# ============================================================================
print('='*80)
print('PART 1: Updating Decision_Log IDs')
print('='*80)

if 'Decision_Log' in wb.sheetnames:
    ws = wb['Decision_Log']
    updated_count = 0

    for row in range(2, ws.max_row + 1):
        decision_id = ws.cell(row, 1).value
        project_id = ws.cell(row, 3).value

        if decision_id and project_id:
            if not str(decision_id).startswith('PRJ-'):
                if str(decision_id).startswith('DEC-'):
                    dec_num = str(decision_id).replace('DEC-', '')
                    new_id = f'{project_id}-DEC-{dec_num}'
                    ws.cell(row, 1).value = new_id
                    updated_count += 1
                    print(f'  Row {row}: {decision_id} -> {new_id}')

    print(f'COMPLETE: Updated {updated_count} decision IDs')
else:
    print('WARNING: Decision_Log sheet not found')

print()

# ============================================================================
# PART 2: Update Risk_Register IDs (NO TABLE ISSUES)
# ============================================================================
print('='*80)
print('PART 2: Updating Risk_Register IDs')
print('='*80)

if 'Risk_Register' in wb.sheetnames:
    ws = wb['Risk_Register']
    updated_count = 0

    for row in range(2, ws.max_row + 1):
        risk_id = ws.cell(row, 1).value
        project_id = ws.cell(row, 3).value

        if risk_id and project_id:
            if not str(risk_id).startswith('PRJ-'):
                if str(risk_id).startswith('RISK-'):
                    risk_num = str(risk_id).replace('RISK-', '')
                    new_id = f'{project_id}-RISK-{risk_num}'
                    ws.cell(row, 1).value = new_id
                    updated_count += 1
                    print(f'  Row {row}: {risk_id} -> {new_id}')

    print(f'COMPLETE: Updated {updated_count} risk IDs')
else:
    print('WARNING: Risk_Register sheet not found')

print()

# ============================================================================
# PART 3: Rebuild Stakeholders - PROPER TABLE HANDLING
# ============================================================================
print('='*80)
print('PART 3: Rebuilding Stakeholders (PROPER TABLE HANDLING)')
print('='*80)

if 'Stakeholders' in wb.sheetnames:
    ws = wb['Stakeholders']

    # STEP 1: Save existing data
    print('Step 1: Saving existing stakeholder data...')
    old_data = []
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            stk_id = ws.cell(row, 1).value
            name = ws.cell(row, 2).value
            if stk_id or name:
                old_data.append({'id': stk_id, 'name': name})
    print(f'  Saved {len(old_data)} existing stakeholders')

    # STEP 2: Remove the old Excel Table (THIS IS THE KEY!)
    print('Step 2: Removing old Excel Table...')
    if hasattr(ws, 'tables') and ws.tables:
        table_names = list(ws.tables.keys())
        for table_name in table_names:
            print(f'  Removing table: {table_name}')
            del ws.tables[table_name]
    print('  Table removed successfully')

    # STEP 3: Clear the sheet but keep row 1
    print('Step 3: Clearing old data...')
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # STEP 4: Add new 22-column headers
    print('Step 4: Adding 22 new column headers...')
    headers = [
        'Stakeholder_ID',      # A
        'Name',                # B
        'Title',               # C
        'Organization',        # D
        'Location_City',       # E
        'Location_Country',    # F
        'Time_Zone_Offset',    # G
        'Local_Time',          # H
        'Email',               # I
        'Phone',               # J
        'Stakeholder_Type',    # K
        'Project_IDs',         # L
        'Countries',           # M
        'Products',            # N
        'Region',              # O
        'Theme',               # P
        'Influence_Level',     # Q
        'Interest_Level',      # R
        'Contact_Frequency',   # S
        'Last_Contact',        # T
        'Next_Contact',        # U
        'Notes'                # V
    ]

    # Clear old headers first
    for col in range(1, 100):  # Clear way more than needed
        ws.cell(1, col).value = None

    # Add new headers with formatting
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f'  Added {len(headers)} headers')

    # STEP 5: Add formula for Local_Time
    ws['H2'] = '=IF(G2<>"",NOW()+(G2/24),"")'

    # STEP 6: Restore old data
    if old_data:
        print('Step 5: Restoring existing stakeholder...')
        ws['A2'] = old_data[0]['id']
        ws['B2'] = old_data[0]['name']

    # STEP 7: Set column widths
    print('Step 6: Setting column widths...')
    column_widths = {
        'A': 15, 'B': 25, 'C': 25, 'D': 30, 'E': 20, 'F': 20,
        'G': 12, 'H': 20, 'I': 30, 'J': 18, 'K': 20, 'L': 20,
        'M': 20, 'N': 20, 'O': 10, 'P': 25, 'Q': 15, 'R': 15,
        'S': 18, 'T': 15, 'U': 15, 'V': 40
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # STEP 8: Create NEW Excel Table with correct range
    print('Step 7: Creating new Excel Table...')
    table_ref = f'A1:V2'  # Start with just header + 1 row
    new_table = Table(displayName='Stakeholders', ref=table_ref)

    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    new_table.tableStyleInfo = style

    ws.add_table(new_table)
    print(f'  New table created: {table_ref}')

    # Freeze top row
    ws.freeze_panes = 'A2'

    print('COMPLETE: Stakeholders rebuilt with proper table handling')
else:
    print('WARNING: Stakeholders sheet not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v3-FIXED...')
wb.save(output_file)
wb.close()

print('='*80)
print('SUCCESS: All updates complete with proper table handling!')
print('='*80)
print()
print('Changes:')
print('  1. Decision_Log: DEC-001 -> PRJ-XXX-DEC-XXX')
print('  2. Risk_Register: RISK-001 -> PRJ-XXX-RISK-XXX')
print('  3. Stakeholders: Properly rebuilt (12 -> 22 columns)')
print('     - Old table removed FIRST')
print('     - Structure rebuilt')
print('     - NEW table created with correct range')
print()
print(f'Output: {output_file}')
print()
print('This should open without any table corruption errors!')
