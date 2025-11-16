"""
Check Control sheet formulas and Country Dashboard structure
"""
import openpyxl

print("="*80)
print("CHECKING CONTROL AND COUNTRY_DASHBOARD")
print("="*80)

# Load v32
print("\nLoading v32...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v32.xlsx')

print("\n" + "="*80)
print("PART 1: CONTROL SHEET - AVG BUDGET AND BURNRATE")
print("="*80)

ws_control = wb['Control']

print("\nChecking rows 1-10:")
for row in range(1, 11):
    col_a = ws_control.cell(row, 1).value
    col_b = ws_control.cell(row, 2).value
    col_c = ws_control.cell(row, 3).value
    col_d = ws_control.cell(row, 4).value
    col_e = ws_control.cell(row, 5).value

    if col_a or col_b or col_c or col_d or col_e:
        print(f"\nRow {row}:")
        if col_a:
            print(f"  A: {str(col_a).encode('ascii', 'ignore').decode('ascii')[:40]}")
        if col_b:
            if ws_control.cell(row, 2).data_type == 'f':
                print(f"  B: FORMULA - {str(col_b)[:60]}...")
            else:
                print(f"  B: {col_b}")
        if col_c:
            print(f"  C: {str(col_c).encode('ascii', 'ignore').decode('ascii')[:40]}")
        if col_d:
            print(f"  D: {str(col_d).encode('ascii', 'ignore').decode('ascii')[:40]}")
        if col_e:
            if ws_control.cell(row, 5).data_type == 'f':
                print(f"  E: FORMULA - {str(col_e)[:60]}...")
            else:
                print(f"  E: {col_e}")

print("\n" + "="*80)
print("PART 2: COUNTRY_DASHBOARD - CHECK COLUMNS")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nRow 11 headers (project list):")
for col in range(1, 20):
    header = ws_country.cell(11, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}11: {clean_header}")

print("\nRow 12 (first data row) - checking for formulas:")
for col in range(1, 20):
    cell = ws_country.cell(12, col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        if cell.data_type == 'f':
            print(f"  {col_letter}12: FORMULA")
        else:
            print(f"  {col_letter}12: {cell.value}")

print("\n" + "="*80)
print("PART 3: STAKEHOLDERS SHEET - CHECK STRUCTURE")
print("="*80)

ws_stakeholders = wb['Stakeholders']

print("\nRow 1 headers:")
for col in range(1, 10):
    header = ws_stakeholders.cell(1, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}1: {clean_header}")

print("\nAnalysis complete!")
