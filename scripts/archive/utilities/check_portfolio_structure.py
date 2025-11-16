"""
Check Portfolio_Dashboard structure to understand column P
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v46.xlsx')
ws = wb['Portfolio_Dashboard']

print("="*80)
print("PORTFOLIO_DASHBOARD STRUCTURE")
print("="*80)

print("\nRow 10 headers (columns A-Q):")
for col in range(1, 18):
    cell = ws.cell(10, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        print(f"  {col_letter}10: {cell.value}")

print("\nRow 11 values (columns A-Q):")
for col in range(1, 18):
    cell = ws.cell(11, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        if cell.data_type == 'f':
            print(f"  {col_letter}11: [formula]")
        else:
            print(f"  {col_letter}11: {cell.value}")

print("\nColumn P specifically:")
print(f"  P10: {ws['P10'].value}")
print(f"  P11: {ws['P11'].value}")
print(f"  P11 type: {ws['P11'].data_type}")

wb.close()
