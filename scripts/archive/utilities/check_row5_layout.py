"""
Check row 4 and 5 layout to understand ULO% formula
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')
ws_country = wb['Country_Dashboard']

print("="*80)
print("ROW 4 - LABELS")
print("="*80)

for col in range(1, 12):
    cell = ws_country.cell(4, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}4: {clean_value}")

print("\n" + "="*80)
print("ROW 5 - VALUES/FORMULAS")
print("="*80)

for col in range(1, 12):
    cell = ws_country.cell(5, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)[:60]
            print(f"  {col_letter}5: {formula}...")
        else:
            print(f"  {col_letter}5: {cell.value}")

print("\n" + "="*80)
print("UNDERSTANDING ULO% FORMULA")
print("="*80)

print("\nCurrent G6 formula: =IF(B5=0,0,G5/B5)")
print("\nThis divides G5 by B5")

g5 = ws_country['G5']
b5 = ws_country['B5']

print(f"\nB5 formula: {b5.value if b5.data_type == 'f' else b5.value}")
print(f"G5 formula: {g5.value if g5.data_type == 'f' else g5.value}")

wb.close()
