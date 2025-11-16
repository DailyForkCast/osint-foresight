import openpyxl

file_path = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.1.xlsx'

print('='*80)
print('COUNTRY BUDGETS - DETAILED INVESTIGATION')
print('='*80)
print()

wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['Country_Budgets']

print('COLUMN E (Country_Name) - Formula Analysis:')
print('-'*80)
print(f'E2 formula: {ws["E2"].value}')
print(f'E3 formula: {ws["E3"].value}')
print()

print('COLUMN A (Budget_ID) - Formula Analysis:')
print('-'*80)
print(f'A2 formula: {ws["A2"].value}')
print(f'A3 formula: {ws["A3"].value}')
print()

print('DATA VALIDATIONS:')
print('-'*80)
if hasattr(ws, 'data_validations'):
    if ws.data_validations.dataValidation:
        for idx, dv in enumerate(ws.data_validations.dataValidation):
            print(f'\nValidation #{idx + 1}:')
            print(f'  Type: {dv.type}')
            print(f'  Formula1: {dv.formula1}')
            print(f'  Formula2: {dv.formula2}')
            print(f'  Cells: {dv.sqref}')
            print(f'  Allow Blank: {dv.allowBlank}')
            print(f'  Show Dropdown: {dv.showDropDown}')
    else:
        print('  No data validations found')
else:
    print('  Sheet does not have data_validations attribute')

print()
print('ROW 4 (should be empty or template row):')
print('-'*80)
for col in range(1, 12):
    cell = ws.cell(4, col)
    if cell.value:
        print(f'  Column {openpyxl.utils.get_column_letter(col)}: {cell.value}')

wb.close()

print()
print('='*80)
print('Now checking Country_Regions table...')
print('='*80)

wb = openpyxl.load_workbook(file_path, data_only=False)
ws_regions = wb['Country_Regions']

print(f'\nCountry_Regions has {ws_regions.max_row} rows')
print('Sample countries (first 10):')
for row in range(2, min(12, ws_regions.max_row + 1)):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    print(f'  {code}: {name}')

wb.close()
