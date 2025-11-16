"""
Diagnose why Countries column isn't populating
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("DIAGNOSING COUNTRIES COLUMN ISSUE")
print("="*80)

# Check Master_Projects
ws_mp = wb['Master_Projects']
print("\nMASTER_PROJECTS - Project IDs (Column A):")
for row in range(2, 5):
    project_id = ws_mp.cell(row, 1).value
    if project_id:
        print(f"  Row {row}: '{project_id}'")

# Check Country_Budgets
ws_cb = wb['Country_Budgets']
print("\nCOUNTRY_BUDGETS - Data:")
print(f"{'Row':<5} {'Col B (Unique_ID)':<20} {'Col D (Code)':<12} {'Col E (Name)':<25}")
print("-"*65)

for row in range(2, 10):
    unique_id = ws_cb.cell(row, 2).value or ""
    country_code = ws_cb.cell(row, 4).value or ""
    country_name = ws_cb.cell(row, 5).value or ""

    if unique_id or country_code or country_name:
        print(f"{row:<5} {str(unique_id):<20} {str(country_code):<12} {str(country_name):<25}")

print("\n" + "="*80)
print("PROBLEM DIAGNOSIS")
print("="*80)

# Get actual values
mp_id = ws_mp.cell(2, 1).value
cb_id = ws_cb.cell(2, 2).value

print(f"\nMaster_Projects A2: '{mp_id}'")
print(f"Country_Budgets B2: '{cb_id}'")

if mp_id != cb_id:
    print("\nISSUE #1: PROJECT IDs DON'T MATCH!")
    print(f"  Master_Projects uses: '{mp_id}'")
    print(f"  Country_Budgets uses: '{cb_id}'")
    print(f"  Formula can't find matches because IDs are different")

cb_name = ws_cb.cell(2, 5).value
if not cb_name:
    print("\nISSUE #2: COUNTRY NAMES ARE EMPTY!")
    print("  Country_Budgets column E (Country_Name) is empty")
    print("  Formula looks for names in column E but finds nothing")

print("\n" + "="*80)
print("SOLUTION")
print("="*80)

if mp_id != cb_id:
    print("\n1. FIX PROJECT ID MISMATCH:")
    print(f"   Option A: Change Country_Budgets column B to use '{mp_id}'")
    print(f"   Option B: Change Master_Projects column A to use '{cb_id}'")

if not cb_name:
    print("\n2. ADD COUNTRY NAMES:")
    print("   Go to Country_Budgets and use the dropdown in column E to select:")
    print("   - Row 2: Select 'Germany' (for code DE)")
    print("   - Row 3: Select 'France' (for code FR)")
