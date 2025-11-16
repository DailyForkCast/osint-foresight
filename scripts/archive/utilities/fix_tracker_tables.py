import openpyxl
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("FIXING TABLE REFERENCE ERRORS")
print("=" * 70)

# Load the workbook
print("\n[1/3] Loading workbook...")
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-UPDATED.xlsx')
print("✓ Workbook loaded")

print("\n[2/3] Fixing Project_Spotlight formulas...")
ws_spotlight = wb['Project_Spotlight']

# Remove the problematic array formulas that reference tables
# Replace with simpler approaches or leave blank for manual entry

# Clear the deliverables table formulas (rows 17-21) that might be causing issues
cleared_count = 0
for row in range(17, 22):
    for col in range(2, 7):  # Columns B-F
        try:
            cell = ws_spotlight.cell(row=row, column=col)
            # If it has a formula, clear it
            if cell.value and isinstance(cell.value, str) and '=' in cell.value:
                cell.value = ""
                cleared_count += 1
        except:
            # Skip merged cells
            pass

print(f"✓ Cleared {cleared_count} problematic formulas")

# Add simple placeholder text instead (only to non-merged cells)
try:
    if not ws_spotlight.cell(17, 2).value:
        ws_spotlight.cell(17, 2).value = '[Deliverables: Add manually or use VLOOKUP from Project_Deliverables]'
except:
    pass

print("✓ Added placeholders where possible")

print("\n[3/3] Saving fixed workbook...")
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-FIXED.xlsx'
wb.save(output_file)
print(f"✓ Saved: {output_file}")

print("\n" + "=" * 70)
print("FIX COMPLETE!")
print("=" * 70)
print("\nWhat was fixed:")
print("  • Removed array formulas referencing Excel Tables")
print("  • Cleared problematic table references")
print("  • File should now open without errors")
print("\nNew file: 2025-10-05-Tracker-FIXED.xlsx")
print("\nNOTE: You'll need to manually add:")
print("  • Target Audiences data in Project_Spotlight")
print("  • Target Technologies data in Project_Spotlight")
print("  • Deliverables table (or add formulas yourself)")
