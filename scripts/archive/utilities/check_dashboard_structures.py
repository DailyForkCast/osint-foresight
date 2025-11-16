"""
Check Country_Dashboard and Regional_Summary structures to see where to add Total_Proposed
"""
import openpyxl

print("="*80)
print("CHECKING DASHBOARD STRUCTURES")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

print("\n" + "="*80)
print("COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

# Check first 15 rows and columns for structure
print("\nHeaders and structure (first 12 rows, columns A-M):")
for row in range(1, 13):
    row_data = []
    for col in range(1, 14):  # A-M
        cell = ws_country.cell(row, col)
        value = cell.value
        col_letter = openpyxl.utils.get_column_letter(col)

        if value:
            clean_value = str(value).encode('ascii', 'ignore').decode('ascii')[:30]
            row_data.append(f"{col_letter}{row}:{clean_value}")

    if row_data:
        print(f"  Row {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
print("REGIONAL_SUMMARY")
print("="*80)

ws_regional = wb['Regional_Summary']

# Check headers
print("\nHeaders (row 1, columns A-L):")
for col in range(1, 13):  # A-L
    cell = ws_regional.cell(1, col)
    value = cell.value
    col_letter = openpyxl.utils.get_column_letter(col)

    if value:
        clean_value = str(value).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}1: {clean_value}")

print("\nData rows (rows 2-7, columns A-L):")
for row in range(2, 8):
    region_name = ws_regional.cell(row, 1).value
    if region_name:
        print(f"\n  Row {row} ({region_name}):")
        for col in range(1, 13):
            cell = ws_regional.cell(row, col)
            value = cell.value
            col_letter = openpyxl.utils.get_column_letter(col)

            if value and col > 1:  # Skip region name column
                if cell.data_type == 'f':
                    print(f"    {col_letter}: FORMULA")
                else:
                    clean_value = str(value).encode('ascii', 'ignore').decode('ascii')[:40]
                    print(f"    {col_letter}: {clean_value}")

print("\n" + "="*80)
print("RECOMMENDATIONS")
print("="*80)

print("\nCountry_Dashboard:")
print("  - Need to find where Allocated, Obligated, Spent columns are")
print("  - Add Total_Proposed column before Allocated")

print("\nRegional_Summary:")
print("  - Currently has: Projects, Countries, Allocated, Obligated, Spent, ULO, ULO%, Execution Rate")
print("  - Need to add Total_Proposed column (likely between Countries and Allocated)")

print("\nReady to add columns!")
