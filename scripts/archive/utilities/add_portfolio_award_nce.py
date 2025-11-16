"""
Add Award Number and NCE tracking to Portfolio_Dashboard
Insert between Countries (F) and Total Proposed (G)
"""
import openpyxl

print("="*80)
print("ADDING AWARD NUMBER AND NCE TO PORTFOLIO_DASHBOARD")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("CURRENT STRUCTURE")
print("="*80)

print("\nCurrent columns (row 10 headers):")
for col in range(1, 15):
    header = ws_port.cell(10, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    if header:
        print(f"  {col_letter}: {header}")

print("\n" + "="*80)
print("INSERTING NEW COLUMNS")
print("="*80)

print("\nInserting 3 columns after F (Countries)...")
# Insert 3 columns at position G
ws_port.insert_cols(7, 3)
print("  Inserted columns G, H, I")
print("  Old G (Total Proposed) is now J")
print("  Old H-M shifted to K-P")

print("\n" + "="*80)
print("ADDING HEADERS")
print("="*80)

# Add headers in row 10
print("\nAdding new column headers...")
ws_port['G10'] = 'Award Number'
ws_port['H10'] = 'NCE Eligible'
ws_port['I10'] = 'NCE Status'

print("  G10: Award Number")
print("  H10: NCE Eligible")
print("  I10: NCE Status")

print("\n" + "="*80)
print("ADDING FORMULAS (ROWS 11-20)")
print("="*80)

print("\nAdding formulas for Award Number, NCE Eligible, NCE Status...")

for row in range(11, 21):
    # G: Award Number
    ws_port.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[Award_Number],ROW()-10),""))'

    # H: NCE Eligible
    ws_port.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[NCE_Eligible],ROW()-10),""))'

    # I: NCE Status (shows current NCE status/count)
    ws_port.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(INDEX(T_Master_Projects[NCE_Status],ROW()-10),""))'

print("  Formulas added for rows 11-20")

print("\n" + "="*80)
print("NEW COLUMN STRUCTURE")
print("="*80)

print("\nUpdated columns:")
print("  A: ID")
print("  B: Project Name")
print("  C: Status")
print("  D: Priority")
print("  E: Progress")
print("  F: Countries")
print("  G: Award Number (NEW)")
print("  H: NCE Eligible (NEW)")
print("  I: NCE Status (NEW)")
print("  J: Total Proposed (was G)")
print("  K: Total Allocation (was H)")
print("  L: Total Obligated (was I)")
print("  M: Total Spent (was J)")
print("  N: Total ULO (was K)")
print("  O: ULO % (was L)")
print("  P: Days Remaining (was M)")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Award Number and NCE columns added!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nAdded 3 new columns to Portfolio_Dashboard:")
print("  1. Award Number - Shows contract/award number")
print("  2. NCE Eligible - Yes/No")
print("  3. NCE Status - Shows which NCE (None, NCE 1 Approved, etc.)")

print("\nThese columns pull from Master_Projects:")
print("  - Award_Number")
print("  - NCE_Eligible")
print("  - NCE_Status")

print("\nAll budget columns shifted right by 3 columns")
print("Headers in row 10 updated")

print("\nv28 updated!")
