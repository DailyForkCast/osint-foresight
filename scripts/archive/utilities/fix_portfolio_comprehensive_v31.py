"""
Fix Portfolio_Dashboard v31:
1. Fix D6 percentage formula
2. Fix columns A-E (ID, Project Name, Status, Priority, Progress) to show blank instead of 0
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO_DASHBOARD - CREATING V31")
print("="*80)

# Load v30
print("\nLoading v30...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v30.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("PART 1: FIX D6 PERCENTAGE FORMULA")
print("="*80)

print("\nCurrent D6: =TEXT(Control!B18,\"0%\")&\" Unobligated\"")
print("  Problem: B18 is Total ULO (dollar amount), not percentage")
print("  Fix: Use B19 (Portfolio ULO %)")

new_formula = '=TEXT(Control!B19,"0%")&" Unobligated"'
ws_portfolio['D6'] = new_formula

print(f"  New D6: {new_formula}")

print("\n" + "="*80)
print("PART 2: FIX BLANK/ZERO DISPLAY IN COLUMNS A-E")
print("="*80)

print("\nUpdating rows 11-20 to show blank instead of 0...")

# Columns to fix:
# A: ID
# B: Project Name
# C: Status
# D: Priority
# E: Progress

for row in range(11, 21):
    # Column A: ID
    # Current: =IFERROR(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10),"")
    # New: Double-check for blank
    ws_portfolio.cell(row, 1).value = f'=IFERROR(IF(INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)="","",INDEX(T_Master_Projects[Project_Unique_ID],ROW()-10)),"")'

    # Column B: Project Name - only if A is not blank
    ws_portfolio.cell(row, 2).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Name],ROW()-10)="","",INDEX(T_Master_Projects[Project_Name],ROW()-10)),""))'

    # Column C: Status - only if A is not blank
    ws_portfolio.cell(row, 3).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Status],ROW()-10)="","",INDEX(T_Master_Projects[Project_Status],ROW()-10)),""))'

    # Column D: Priority - only if A is not blank
    ws_portfolio.cell(row, 4).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Priority],ROW()-10)="","",INDEX(T_Master_Projects[Project_Priority],ROW()-10)),""))'

    # Column E: Progress - only if A is not blank, and check for 0
    ws_portfolio.cell(row, 5).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Project_Progress],ROW()-10)=0,"",INDEX(T_Master_Projects[Project_Progress],ROW()-10)),""))'

print("  Rows 11-20 updated")

print("\n" + "="*80)
print("SAVING V31")
print("="*80)

wb.save('2025-10-26-Tracker-v31.xlsx')

print("\nOK - v31 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v31:")
print("\n1. D6 Percentage Formula:")
print("   - Now references Control!B19 (Portfolio ULO %)")
print("   - Shows correct percentage (e.g., 25% not 25500000%)")

print("\n2. Blank/Zero Display (Rows 11-20):")
print("   - Column A (ID): Shows blank when empty")
print("   - Column B (Project Name): Shows blank when A is blank or value is empty")
print("   - Column C (Status): Shows blank when A is blank or value is empty")
print("   - Column D (Priority): Shows blank when A is blank or value is empty")
print("   - Column E (Progress): Shows blank when A is blank or value is 0")

print("\nAll columns show clean blank cells instead of 0s")

print("\nv31 ready!")
