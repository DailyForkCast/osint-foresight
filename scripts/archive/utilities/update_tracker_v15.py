import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# State Department region mapping
STATE_DEPT_REGIONS = {
    # Africa (AF)
    'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
    'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
    'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
    'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
    'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
    'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
    'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',

    # East Asia & Pacific (EAP)
    'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
    'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
    'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
    'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
    'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',

    # Europe (EUR)
    'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
    'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
    'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
    'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
    'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
    'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
    'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR',
    'VA': 'EUR',

    # Near East (NEA) - Note: Some countries appear in both AF and NEA lists
    'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA',
    'LB': 'NEA', 'OM': 'NEA', 'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA',
    'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',

    # South & Central Asia (SCA)
    'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
    'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',

    # Western Hemisphere (WHA)
    'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
    'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
    'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
    'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
    'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
}

# Subregions
SUBREGIONS = {
    # Africa subregions
    'DZ': 'Northern Africa', 'EG': 'Northern Africa', 'LY': 'Northern Africa', 'MA': 'Northern Africa',
    'TN': 'Northern Africa', 'SD': 'Northern Africa',

    'BJ': 'Western Africa', 'BF': 'Western Africa', 'CV': 'Western Africa', 'CI': 'Western Africa',
    'GM': 'Western Africa', 'GH': 'Western Africa', 'GN': 'Western Africa', 'GW': 'Western Africa',
    'LR': 'Western Africa', 'ML': 'Western Africa', 'MR': 'Western Africa', 'NE': 'Western Africa',
    'NG': 'Western Africa', 'SN': 'Western Africa', 'SL': 'Western Africa', 'TG': 'Western Africa',

    'BI': 'Eastern Africa', 'KM': 'Eastern Africa', 'DJ': 'Eastern Africa', 'ER': 'Eastern Africa',
    'ET': 'Eastern Africa', 'KE': 'Eastern Africa', 'MG': 'Eastern Africa', 'MW': 'Eastern Africa',
    'MU': 'Eastern Africa', 'MZ': 'Eastern Africa', 'RW': 'Eastern Africa', 'SC': 'Eastern Africa',
    'SO': 'Eastern Africa', 'SS': 'Eastern Africa', 'TZ': 'Eastern Africa', 'UG': 'Eastern Africa',
    'ZM': 'Eastern Africa', 'ZW': 'Eastern Africa',

    'AO': 'Middle Africa', 'CM': 'Middle Africa', 'CF': 'Middle Africa', 'TD': 'Middle Africa',
    'CG': 'Middle Africa', 'CD': 'Middle Africa', 'GQ': 'Middle Africa', 'GA': 'Middle Africa',

    'BW': 'Southern Africa', 'LS': 'Southern Africa', 'NA': 'Southern Africa', 'ZA': 'Southern Africa',
    'SZ': 'Southern Africa',

    # Europe subregions
    'DK': 'Northern Europe', 'EE': 'Northern Europe', 'FI': 'Northern Europe', 'IS': 'Northern Europe',
    'IE': 'Northern Europe', 'LV': 'Northern Europe', 'LT': 'Northern Europe', 'NO': 'Northern Europe',
    'SE': 'Northern Europe', 'GB': 'Northern Europe',

    'BY': 'Eastern Europe', 'BG': 'Eastern Europe', 'CZ': 'Eastern Europe', 'HU': 'Eastern Europe',
    'MD': 'Eastern Europe', 'PL': 'Eastern Europe', 'RO': 'Eastern Europe', 'RU': 'Eastern Europe',
    'SK': 'Eastern Europe', 'UA': 'Eastern Europe',

    'AL': 'Southern Europe', 'AD': 'Southern Europe', 'BA': 'Southern Europe', 'HR': 'Southern Europe',
    'CY': 'Southern Europe', 'GR': 'Southern Europe', 'IT': 'Southern Europe', 'XK': 'Southern Europe',
    'MK': 'Southern Europe', 'MT': 'Southern Europe', 'ME': 'Southern Europe', 'PT': 'Southern Europe',
    'RS': 'Southern Europe', 'SI': 'Southern Europe', 'ES': 'Southern Europe', 'VA': 'Southern Europe',

    'AT': 'Western Europe', 'BE': 'Western Europe', 'FR': 'Western Europe', 'DE': 'Western Europe',
    'LI': 'Western Europe', 'LU': 'Western Europe', 'MC': 'Western Europe', 'NL': 'Western Europe',
    'CH': 'Western Europe',

    # Caucasus
    'AM': 'Caucasus', 'AZ': 'Caucasus', 'GE': 'Caucasus',

    # Turkey
    'TR': 'Western Asia',

    # Americas subregions
    'CA': 'Northern America', 'US': 'Northern America',

    'BZ': 'Central America', 'CR': 'Central America', 'SV': 'Central America', 'GT': 'Central America',
    'HN': 'Central America', 'MX': 'Central America', 'NI': 'Central America', 'PA': 'Central America',

    'AG': 'Caribbean', 'BS': 'Caribbean', 'BB': 'Caribbean', 'CU': 'Caribbean', 'DM': 'Caribbean',
    'DO': 'Caribbean', 'GD': 'Caribbean', 'HT': 'Caribbean', 'JM': 'Caribbean', 'KN': 'Caribbean',
    'LC': 'Caribbean', 'VC': 'Caribbean', 'TT': 'Caribbean',

    'AR': 'South America', 'BO': 'South America', 'BR': 'South America', 'CL': 'South America',
    'CO': 'South America', 'EC': 'South America', 'GY': 'South America', 'PY': 'South America',
    'PE': 'South America', 'SR': 'South America', 'UY': 'South America', 'VE': 'South America',

    # Asia subregions
    'CN': 'Eastern Asia', 'JP': 'Eastern Asia', 'KP': 'Eastern Asia', 'KR': 'Eastern Asia',
    'MN': 'Eastern Asia', 'TW': 'Eastern Asia',

    'BN': 'South-Eastern Asia', 'KH': 'South-Eastern Asia', 'ID': 'South-Eastern Asia',
    'LA': 'South-Eastern Asia', 'MY': 'South-Eastern Asia', 'MM': 'South-Eastern Asia',
    'PH': 'South-Eastern Asia', 'SG': 'South-Eastern Asia', 'TH': 'South-Eastern Asia',
    'TL': 'South-Eastern Asia', 'VN': 'South-Eastern Asia',

    'AF': 'Southern Asia', 'BD': 'Southern Asia', 'BT': 'Southern Asia', 'IN': 'Southern Asia',
    'MV': 'Southern Asia', 'NP': 'Southern Asia', 'PK': 'Southern Asia', 'LK': 'Southern Asia',

    'KZ': 'Central Asia', 'KG': 'Central Asia', 'TJ': 'Central Asia', 'TM': 'Central Asia',
    'UZ': 'Central Asia',

    'AU': 'Australia and New Zealand', 'NZ': 'Australia and New Zealand',

    'FJ': 'Melanesia', 'PG': 'Melanesia', 'SB': 'Melanesia', 'VU': 'Melanesia',

    'KI': 'Micronesia', 'MH': 'Micronesia', 'FM': 'Micronesia', 'NR': 'Micronesia', 'PW': 'Micronesia',

    'WS': 'Polynesia', 'TO': 'Polynesia', 'TV': 'Polynesia',

    # Middle East
    'BH': 'Arabian Peninsula', 'KW': 'Arabian Peninsula', 'OM': 'Arabian Peninsula',
    'QA': 'Arabian Peninsula', 'SA': 'Arabian Peninsula', 'AE': 'Arabian Peninsula', 'YE': 'Arabian Peninsula',

    'IL': 'Levant', 'JO': 'Levant', 'LB': 'Levant', 'PS': 'Levant', 'SY': 'Levant',

    'IQ': 'Mesopotamia',
}

print("="*80)
print("UPDATING TRACKER V13")
print("="*80)

print("\nLoading v15...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v15.xlsx')

# ====== STEP 1: Update Country_Regions ======
print("\n" + "="*80)
print("STEP 1: Updating Country_Regions with State Dept regions and subregions")
print("="*80)
ws = wb['Country_Regions']

updated_count = 0
missing_region = []
missing_subregion = []

for row in range(2, ws.max_row + 1):
    code = ws.cell(row, 1).value
    name = ws.cell(row, 2).value

    if code:
        code = code.strip().upper()
        # Update region
        if code in STATE_DEPT_REGIONS:
            ws.cell(row, 3).value = STATE_DEPT_REGIONS[code]
            updated_count += 1
        else:
            missing_region.append(f"{code} - {name}")

        # Update subregion
        if code in SUBREGIONS:
            ws.cell(row, 5).value = SUBREGIONS[code]
        else:
            missing_subregion.append(f"{code} - {name}")

print(f"Updated {updated_count} country regions")
if missing_region:
    print(f"WARNING: {len(missing_region)} countries missing region mapping:")
    for item in missing_region[:5]:
        print(f"  {item}")
if missing_subregion:
    print(f"NOTE: {len(missing_subregion)} countries missing subregion")

wb.save('2025-10-26-Tracker-v15.xlsx')
print("Saved Country_Regions updates")

# ====== STEP 2: Add Proposed_Amount to Country_Budgets ======
print("\n" + "="*80)
print("STEP 2: Adding Proposed_Amount column to Country_Budgets")
print("="*80)

ws = wb['Country_Budgets']

# Insert column F (before current Allocated_Amount which is at F)
print("Inserting column at F for Proposed_Amount...")
ws.insert_cols(6)  # Insert before column F

# Set header
ws.cell(1, 6).value = "Proposed_Amount"
ws.cell(1, 6).font = Font(bold=True, color='FFFFFF')
ws.cell(1, 6).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')

# Add default value of 0 for all rows
for row in range(2, ws.max_row + 1):
    ws.cell(row, 6).value = 0

print("Added Proposed_Amount column")
print("Column order is now: A-E (IDs/Country), F=Proposed, G=Allocated, H=Obligated, I=Spent, J=ULO, K=ULO%, L=Health")

# Update ULO formula to use Obligated (col H) - Spent (col I)
print("\nUpdating ULO formulas to use Obligated - Spent...")
for row in range(2, ws.max_row + 1):
    # ULO = Obligated - Spent (col H - col I, was I - J before insertion)
    ws.cell(row, 10).value = f"=H{row}-I{row}"
    # ULO % = ULO / Obligated (col J / col H)
    ws.cell(row, 11).value = f"=IF(H{row}>0,J{row}/H{row},0)"

print("Updated ULO formulas")

wb.save('2025-10-26-Tracker-v15.xlsx')
print("Saved Country_Budgets updates")

# ====== STEP 3: Add Total_Proposed to Master_Projects ======
print("\n" + "="*80)
print("STEP 3: Adding Total_Proposed column to Master_Projects")
print("="*80)

ws = wb['Master_Projects']

# Insert column at N (before current Total_Allocation)
print("Inserting column at N for Total_Proposed...")
ws.insert_cols(14)  # Column N is 14

# Set header
ws.cell(1, 14).value = "Total_Proposed"
ws.cell(1, 14).font = Font(bold=True, color='FFFFFF')
ws.cell(1, 14).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.cell(1, 14).alignment = Alignment(horizontal='center', vertical='center')

# Add SUMIF formula to sum proposed amounts by project
for row in range(2, ws.max_row + 1):
    project_id_col = 'A'  # Assuming Project_ID is in column A
    ws.cell(row, 14).value = f'=SUMIF(Country_Budgets!$B:$B,{project_id_col}{row},Country_Budgets!$F:$F)'

print("Added Total_Proposed column")
print("Budget columns now: N=Total_Proposed, O=Total_Allocation, P=Total_Obligated, Q=Total_ULO, R=ULO%")

# Update Total_ULO formula
print("\nUpdating Total_ULO formulas to use Total_Obligated - Total_Spent...")
for row in range(2, ws.max_row + 1):
    # Total_ULO should be Total_Obligated - Total_Spent
    # But we don't have Total_Spent in Master_Projects, so ULO should be summed from Country_Budgets
    ws.cell(row, 17).value = f'=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$J:$J)'
    # ULO % = Total_ULO / Total_Obligated
    ws.cell(row, 18).value = f'=IF(P{row}>0,Q{row}/P{row},0)'

print("Updated ULO formulas in Master_Projects")

wb.save('2025-10-26-Tracker-v15.xlsx')
print("Saved Master_Projects updates")

# ====== STEP 4: Update Config_Lists with countries ======
print("\n" + "="*80)
print("STEP 4: Adding countries to Config_Lists")
print("="*80)

ws_config = wb['Config_Lists']
ws_regions = wb['Country_Regions']

# Get all countries from Country_Regions
countries = []
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code and name:
        countries.append((code, name, region))

print(f"Found {len(countries)} countries in Country_Regions")

# Find next empty row in Config_Lists
next_row = ws_config.max_row + 1

# Add all countries
for code, name, region in countries:
    ws_config.cell(next_row, 1).value = "Country"
    ws_config.cell(next_row, 2).value = code
    ws_config.cell(next_row, 3).value = name
    ws_config.cell(next_row, 4).value = region
    next_row += 1

print(f"Added {len(countries)} countries to Config_Lists")

wb.save('2025-10-26-Tracker-v15.xlsx')
print("Saved Config_Lists updates")

# ====== STEP 5: Sync Country_PM_Assignments with Country_Regions ======
print("\n" + "="*80)
print("STEP 5: Syncing Country_PM_Assignments with Country_Regions")
print("="*80)

ws_pm = wb['Country_PM_Assignments']

# Get existing PM assignments
existing_assignments = {}
for row in range(2, ws_pm.max_row + 1):
    code = ws_pm.cell(row, 1).value
    if code:
        existing_assignments[code] = {
            'name': ws_pm.cell(row, 2).value,
            'region': ws_pm.cell(row, 3).value,
            'pm': ws_pm.cell(row, 4).value,
            'email': ws_pm.cell(row, 5).value,
            'phone': ws_pm.cell(row, 6).value,
            'notes': ws_pm.cell(row, 7).value,
        }

# Clear the sheet except header
for row in range(2, ws_pm.max_row + 1):
    for col in range(1, 8):
        ws_pm.cell(row, col).value = None

# Rebuild with countries from Country_Regions
current_row = 2
for code, name, region in sorted(countries, key=lambda x: (x[2], x[1])):  # Sort by region, then name
    ws_pm.cell(current_row, 1).value = code
    ws_pm.cell(current_row, 2).value = name
    ws_pm.cell(current_row, 3).value = region

    # If we had existing PM data, preserve it
    if code in existing_assignments:
        ws_pm.cell(current_row, 4).value = existing_assignments[code]['pm']
        ws_pm.cell(current_row, 5).value = existing_assignments[code]['email']
        ws_pm.cell(current_row, 6).value = existing_assignments[code]['phone']
        ws_pm.cell(current_row, 7).value = existing_assignments[code]['notes']
    else:
        ws_pm.cell(current_row, 4).value = "TBD"

    current_row += 1

print(f"Synced {current_row - 2} countries in Country_PM_Assignments")
print(f"Preserved PM assignments for {sum(1 for c in countries if c[0] in existing_assignments)} countries")

wb.save('2025-10-26-Tracker-v15.xlsx')
print("Saved Country_PM_Assignments updates")

print("\n" + "="*80)
print("ALL UPDATES COMPLETE!")
print("="*80)
print("\nSummary of changes:")
print(f"  1. Updated {updated_count} countries to State Dept regions in Country_Regions")
print("  2. Added Proposed_Amount column to Country_Budgets")
print("  3. Updated ULO calculations to use Obligated - Spent")
print("  4. Added Total_Proposed column to Master_Projects")
print(f"  5. Added {len(countries)} countries to Config_Lists")
print(f"  6. Synced Country_PM_Assignments with {len(countries)} countries")
print("\nFile: 2025-10-26-Tracker-v15.xlsx")
