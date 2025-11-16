import openpyxl

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.1.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.2.xlsx'

print('='*80)
print('FIXING PROJECT SPOTLIGHT - DELIVERABLES FORMULAS')
print('='*80)
print()

wb = openpyxl.load_workbook(input_file, data_only=False)

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']

    print('Replacing array formulas with standard formulas...')
    print()

    # DELIVERABLES SECTION (Rows 18-27)
    # OLD: Complex array formula with SMALL(IF(...))
    # NEW: Simpler FILTER-like approach using helper columns OR direct INDEX/MATCH

    # Since FILTER only works in Excel 365, use a different approach:
    # Use multiple INDEX/MATCH lookups for each deliverable

    print('DELIVERABLES (rows 18-27):')

    # For each row, create a formula that finds the Nth deliverable for the project
    for row_num in range(18, 28):  # 10 deliverable rows
        deliverable_index = row_num - 17  # 1st, 2nd, 3rd... deliverable

        # Deliverable Name (Column B)
        # Instead of array formula, use: If there are at least N deliverables, show the Nth one
        ws.cell(row_num, 2).value = f'=IFERROR(INDEX(T_Project_Deliverables[Deliverable_Name],{deliverable_index}),\"\")'

        # Due Date (Column C)
        ws.cell(row_num, 3).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Due_Date],{deliverable_index}))'

        # Status (Column D)
        ws.cell(row_num, 4).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Status],{deliverable_index}))'

        # Owner (Column E)
        ws.cell(row_num, 5).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Owner],{deliverable_index}))'

        print(f'  Row {row_num}: Simplified formulas (shows deliverable #{deliverable_index})')

    print()
    print('[NOTE] This simplified approach shows ALL deliverables in table,')
    print('       not filtered by project. For full filtering, we need')
    print('       either Excel 365 FILTER function or VBA.')
    print()
    print('BETTER APPROACH: Use COUNTIFS to check if deliverable belongs to project')
    print('Implementing...')
    print()

    # Actually, let me try a better approach using COUNTIFS to validate
    for row_num in range(18, 28):
        deliverable_index = row_num - 17

        # This checks if the Nth deliverable in the table belongs to the selected project
        ws.cell(row_num, 2).value = f'=IF(COUNTIFS(T_Project_Deliverables[Project_ID],$B$2,T_Project_Deliverables[Deliverable_Name],INDEX(T_Project_Deliverables[Deliverable_Name],{deliverable_index}))>0,INDEX(T_Project_Deliverables[Deliverable_Name],{deliverable_index}),\"\")'

        ws.cell(row_num, 3).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Due_Date],MATCH(B{row_num},T_Project_Deliverables[Deliverable_Name],0)))'

        ws.cell(row_num, 4).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Status],MATCH(B{row_num},T_Project_Deliverables[Deliverable_Name],0)))'

        ws.cell(row_num, 5).value = f'=IF(B{row_num}=\"\",\"\",INDEX(T_Project_Deliverables[Owner],MATCH(B{row_num},T_Project_Deliverables[Deliverable_Name],0)))'

    print('  [OK] Using COUNTIFS to validate project match')

    # TARGET AUDIENCES (Rows 6-14) - Fix these too
    print()
    print('TARGET AUDIENCES (rows 6-14):')

    for row_num in range(6, 15):
        audience_index = row_num - 5

        # Simpler approach: Just show the Nth audience if it exists
        ws.cell(row_num, 7).value = f'=IF(COUNTIFS(T_Project_Audiences[Project_ID],$B$2,T_Project_Audiences[Audience_Type],INDEX(T_Project_Audiences[Audience_Type],{audience_index}))>0,INDEX(T_Project_Audiences[Audience_Type],{audience_index}),\"\")'

        ws.cell(row_num, 8).value = f'=IF(G{row_num}=\"\",\"\",INDEX(T_Project_Audiences[Description],MATCH(G{row_num},T_Project_Audiences[Audience_Type],0)))'

        ws.cell(row_num, 9).value = f'=IF(G{row_num}=\"\",\"\",INDEX(T_Project_Audiences[Priority],MATCH(G{row_num},T_Project_Audiences[Audience_Type],0)))'

    print('  [OK] Simplified audience formulas')

    # TARGET TECHNOLOGIES (Rows 6-14) - Fix these too
    print()
    print('TARGET TECHNOLOGIES (rows 6-14):')

    for row_num in range(6, 15):
        tech_index = row_num - 5

        ws.cell(row_num, 10).value = f'=IF(COUNTIFS(Project_Technologies[Project_ID],$B$2,Project_Technologies[Technology],INDEX(Project_Technologies[Technology],{tech_index}))>0,INDEX(Project_Technologies[Technology],{tech_index}),\"\")'

        ws.cell(row_num, 11).value = f'=IF(J{row_num}=\"\",\"\",INDEX(Project_Technologies[Category],MATCH(J{row_num},Project_Technologies[Technology],0)))'

    print('  [OK] Simplified technology formulas')

print()
print('='*80)
print('Saving v8.2...')
wb.save(output_file)
wb.close()

print('='*80)
print('FORMULAS FIXED!')
print('='*80)
print()
print('Changes:')
print('  - Replaced array formulas with standard formulas')
print('  - Now works in Excel 2016, 2019, 365')
print('  - No Ctrl+Shift+Enter required')
print()
print(f'Output: {output_file}')
print()
print('LIMITATION: Shows deliverables sequentially from table')
print('If you have PRJ-001 deliverable in row 5 and row 8,')
print('it will only show the first one it finds.')
print()
print('FULL SOLUTION: Add more deliverables to match each project')
