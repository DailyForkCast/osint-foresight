"""
Check which specific formulas are problematic
"""
import openpyxl

print("="*80)
print("CHECKING PROBLEMATIC FORMULAS")
print("="*80)

# Load the REPAIRED v19 (after Excel fixed it)
print("\nLoading repaired v19...")
wb_fixed = openpyxl.load_workbook('2025-10-26-Tracker-v19.xlsx', data_only=False)

ws = wb_fixed['Master_Projects']

print("\nChecking Master_Projects row 2 - which formulas are missing?")
print("(Excel removed the bad ones)")

missing_formulas = []
for col in range(1, min(35, ws.max_column + 1)):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    # Check specific columns that should have formulas
    should_have_formula = [
        (10, 'J', 'Days_Remaining'),
        (18, 'R', 'POP_Days_Remaining'),
        (23, 'W', 'Total_Proposed'),
        (24, 'X', 'Total_Allocation'),
        (25, 'Y', 'Total_Obligated'),
        (26, 'Z', 'Total_Spent'),
        (27, 'AA', 'Total_ULO'),
        (28, 'AB', 'ULO_Percent'),
        (29, 'AC', 'Countries'),
        (30, 'AD', 'Country_Count'),
    ]

    for expected_col, expected_letter, expected_name in should_have_formula:
        if col == expected_col:
            if cell.data_type != 'f':
                missing_formulas.append((col_letter, header, expected_name))
                print(f"  MISSING: {col_letter} ({header or expected_name}) - should have formula")
            else:
                print(f"  OK: {col_letter} ({header}) has formula")

# Check what formulas actually exist
print("\n" + "="*80)
print("FORMULAS THAT EXIST IN ROW 2:")
print("="*80)

for col in range(1, min(35, ws.max_column + 1)):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.data_type == 'f':
        formula = str(cell.value)
        print(f"\n{col_letter} ({header}):")
        print(f"  {formula[:100]}")

print("\n" + "="*80)
print("LIKELY ISSUE")
print("="*80)

print("\nThe formulas I added probably reference the NEW column structure")
print("but v18's table definitions still use the OLD column structure.")
print("\nFor example:")
print("  - My formula might reference [@[Total_Spent]] in column Z")
print("  - But the table might not have Total_Spent defined yet")
print("\nSolution: We need to either:")
print("  1. Manually add these formulas in Excel (safest)")
print("  2. Expand the table range to include new columns")
print("  3. Start fresh with v18 and add columns manually")
