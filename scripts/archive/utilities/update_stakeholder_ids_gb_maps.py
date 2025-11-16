import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v3-FIXED.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v4.xlsx'

print('='*80)
print('UPDATES: Stakeholder IDs, GB->UK, Remove Maps')
print('='*80)
print()

wb = openpyxl.load_workbook(input_file, data_only=False)

# ============================================================================
# PART 1: Change GB to UK
# ============================================================================
print('='*80)
print('PART 1: Changing GB to UK')
print('='*80)
print()

# Country_Regions
if 'Country_Regions' in wb.sheetnames:
    ws = wb['Country_Regions']
    changes = 0

    for row in range(1, ws.max_row + 1):
        country_code = ws.cell(row, 1).value
        country_name = ws.cell(row, 2).value

        if country_code == 'GB':
            ws.cell(row, 1).value = 'UK'
            print(f'  Country_Regions Row {row}: GB -> UK ({country_name})')
            changes += 1

    print(f'COMPLETE: {changes} changes in Country_Regions')
else:
    print('WARNING: Country_Regions not found')

print()

# Country_PM_Assignments
if 'Country_PM_Assignments' in wb.sheetnames:
    ws = wb['Country_PM_Assignments']
    changes = 0

    for row in range(1, ws.max_row + 1):
        country_code = ws.cell(row, 1).value
        country_name = ws.cell(row, 2).value

        if country_code == 'GB':
            ws.cell(row, 1).value = 'UK'
            print(f'  Country_PM_Assignments Row {row}: GB -> UK ({country_name})')
            changes += 1

    print(f'COMPLETE: {changes} changes in Country_PM_Assignments')
else:
    print('WARNING: Country_PM_Assignments not found')

print()

# Country_Budgets (check if any GB exists)
if 'Country_Budgets' in wb.sheetnames:
    ws = wb['Country_Budgets']
    changes = 0

    for row in range(1, ws.max_row + 1):
        country_code = ws.cell(row, 3).value  # Column C

        if country_code == 'GB':
            ws.cell(row, 3).value = 'UK'
            print(f'  Country_Budgets Row {row}: GB -> UK')
            changes += 1

    if changes > 0:
        print(f'COMPLETE: {changes} changes in Country_Budgets')
    else:
        print('  No GB found in Country_Budgets')

print()

# ============================================================================
# PART 2: Remove Map References from Project_Spotlight
# ============================================================================
print('='*80)
print('PART 2: Removing Map References from Project_Spotlight')
print('='*80)
print()

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']

    # Clear G3 (EUROPE MAP)
    if ws['G3'].value:
        print(f'  G3: Removing "{ws["G3"].value}"')
        ws['G3'].value = None
        ws['G3'].font = Font()
        ws['G3'].fill = PatternFill()

    # Clear G5 ([Map visualization area])
    if ws['G5'].value:
        print(f'  G5: Removing "{ws["G5"].value}"')
        ws['G5'].value = None
        ws['G5'].font = Font()
        ws['G5'].fill = PatternFill()

    print('COMPLETE: Map references removed')
else:
    print('WARNING: Project_Spotlight not found')

print()

# ============================================================================
# PART 3: Add Stakeholder ID Documentation
# ============================================================================
print('='*80)
print('PART 3: Adding Stakeholder ID Structure Documentation')
print('='*80)
print()

if 'Stakeholders' in wb.sheetnames:
    ws = wb['Stakeholders']

    # Add comment to Stakeholder_ID header (A1) with ID structure guide
    id_guide = """STAKEHOLDER ID STRUCTURE:

Priority Order (use first applicable):
1. Country-Specific: [CC]-STK-XXX
   Example: DE-STK-001, UK-STK-001, FR-STK-002

2. Multi-Country: MC-STK-XXX
   Example: MC-STK-001 (works across multiple countries)

3. Regional: [REGION]-STK-XXX
   Example: EUR-STK-001, WHA-STK-001, EAP-STK-001

4. Thematic/Technology: [THEME]-STK-XXX
   Example: CYBER-STK-001, AI-STK-001, CLIMATE-STK-001

5. Project-Specific: PRJ-XXX-STK-XXX
   Example: PRJ-001-STK-001, PRJ-003-STK-002

Use the MOST SPECIFIC category that applies.
If stakeholder fits multiple, use highest priority."""

    # Add comment to cell A1
    comment = Comment(id_guide, 'System')
    ws['A1'].comment = comment

    print('COMPLETE: Added ID structure guide as comment to cell A1')
    print()
    print('ID Structure Priority:')
    print('  1. Country-Specific (DE-STK-001)')
    print('  2. Multi-Country (MC-STK-001)')
    print('  3. Regional (EUR-STK-001)')
    print('  4. Thematic (CYBER-STK-001)')
    print('  5. Project-Specific (PRJ-001-STK-001)')

    # Update the sample stakeholder ID
    if ws['A2'].value == 'STK-001':
        ws['A2'].value = 'EXAMPLE-STK-001'
        print()
        print('Updated sample ID: STK-001 -> EXAMPLE-STK-001')
else:
    print('WARNING: Stakeholders not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v4...')
wb.save(output_file)
wb.close()

print('='*80)
print('ALL UPDATES COMPLETE!')
print('='*80)
print()
print('Summary:')
print('  1. Changed GB to UK in all relevant sheets')
print('  2. Removed map references from Project_Spotlight (G3, G5)')
print('  3. Added Stakeholder ID structure guide (see comment on cell A1)')
print()
print(f'Output: {output_file}')
print()
print('STAKEHOLDER ID EXAMPLES:')
print('  Country:      DE-STK-001 (Germany contact #1)')
print('  Country:      UK-STK-001 (United Kingdom contact #1)')
print('  Multi-Country: MC-STK-001 (works across DE, FR, IT)')
print('  Regional:     EUR-STK-001 (European Regional Director)')
print('  Thematic:     CYBER-STK-001 (Cybersecurity SME)')
print('  Project:      PRJ-001-STK-001 (Project 001 tech lead)')
print()
print('Next: Open v4 and hover over Stakeholder_ID header (A1) to see full guide')
