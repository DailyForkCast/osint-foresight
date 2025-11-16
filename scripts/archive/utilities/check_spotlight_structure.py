"""
Check Spotlight_PMWorkspace to see which fields show 0 instead of blank
"""
import openpyxl

print("="*80)
print("CHECKING SPOTLIGHT_PMWORKSPACE STRUCTURE")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("SPOTLIGHT STRUCTURE (First 50 rows)")
print("="*80)

# Check rows 1-50, columns A-D
print("\nLabel/Value pairs (A=Label, B=Value, C=Label, D=Value):")
for row in range(1, 51):
    # Check columns A-B
    label_a = ws_spotlight.cell(row, 1).value
    value_b = ws_spotlight.cell(row, 2).value
    label_c = ws_spotlight.cell(row, 3).value
    value_d = ws_spotlight.cell(row, 4).value

    if label_a or value_b or label_c or value_d:
        parts = []

        if label_a:
            clean_label = str(label_a).encode('ascii', 'ignore').decode('ascii')[:30]
            parts.append(f"A{row}:{clean_label}")

        if value_b:
            if ws_spotlight.cell(row, 2).data_type == 'f':
                parts.append(f"B{row}:FORMULA")
            else:
                clean_value = str(value_b).encode('ascii', 'ignore').decode('ascii')[:30]
                parts.append(f"B{row}:{clean_value}")

        if label_c:
            clean_label = str(label_c).encode('ascii', 'ignore').decode('ascii')[:30]
            parts.append(f"C{row}:{clean_label}")

        if value_d:
            if ws_spotlight.cell(row, 4).data_type == 'f':
                parts.append(f"D{row}:FORMULA")
            else:
                clean_value = str(value_d).encode('ascii', 'ignore').decode('ascii')[:30]
                parts.append(f"D{row}:{clean_value}")

        if parts:
            print(f"  Row {row:2d}: {' | '.join(parts)}")

print("\n" + "="*80)
print("LOOK FOR THESE FIELDS")
print("="*80)

# Search for specific fields mentioned by user
search_terms = ['progress', 'implementer', 'nce', 'allocated', 'obligated', 'spent']

print("\nSearching for fields that might show 0:")
for term in search_terms:
    print(f"\n  {term.upper()}:")
    found = False
    for row in range(1, 51):
        for col in [1, 3]:  # Label columns
            cell_value = ws_spotlight.cell(row, col).value
            if cell_value and term.lower() in str(cell_value).lower():
                clean_value = str(cell_value).encode('ascii', 'ignore').decode('ascii')
                value_col = col + 1
                value_cell = ws_spotlight.cell(row, value_col)

                if value_cell.data_type == 'f':
                    formula = str(value_cell.value)[:60]
                    print(f"    Row {row}, Col {chr(64+col)}: {clean_value}")
                    print(f"      -> Value in Col {chr(64+value_col)}: FORMULA")
                    print(f"         {formula}...")
                else:
                    print(f"    Row {row}, Col {chr(64+col)}: {clean_value}")
                    print(f"      -> Value in Col {chr(64+value_col)}: {value_cell.value}")
                found = True

    if not found:
        print(f"    Not found")

print("\n" + "="*80)
print("NUMERIC FIELDS TO CHECK")
print("="*80)

print("\nFields that likely need blank-instead-of-0 treatment:")
print("  - Progress (percentage)")
print("  - Implementer (text field)")
print("  - NCE fields")
print("  - Any budget amounts (Allocated, Obligated, Spent, Proposed)")

print("\nReady to fix!")
