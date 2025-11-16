"""
Check column E to see what's there
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')
ws_country = wb['Country_Dashboard']

print("="*80)
print("COLUMN E")
print("="*80)

for row in range(1, 15):
    cell = ws_country.cell(row, 5)  # Column E
    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)
            print(f"  E{row}: {formula}")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  E{row}: {clean_value}")

print("\n" + "="*80)
print("COLUMN F")
print("="*80)

for row in range(1, 15):
    cell = ws_country.cell(row, 6)  # Column F
    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)
            print(f"  F{row}: {formula}")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  F{row}: {clean_value}")

print("\n" + "="*80)
print("COLUMN G")
print("="*80)

for row in range(1, 15):
    cell = ws_country.cell(row, 7)  # Column G
    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)
            print(f"  G{row}: {formula}")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  G{row}: {clean_value}")

wb.close()
