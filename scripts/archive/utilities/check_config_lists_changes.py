"""
Check Config_Lists sheet to see user's table changes
"""
import openpyxl

print("="*80)
print("CHECKING CONFIG_LISTS SHEET")
print("="*80)

# Try to load the most recent file (user may have saved changes)
try:
    wb = openpyxl.load_workbook('2025-10-26-Tracker-v51.xlsx')
    print("\nLoaded: v51")
except:
    print("\nERROR: Could not load v51")
    exit()

ws = wb['Config_Lists']

print("\n" + "="*80)
print("CHECKING FOR TABLES")
print("="*80)

tables = list(ws.tables.values())
if tables:
    print(f"\nFound {len(tables)} table(s):")
    for table in tables:
        print(f"\n  Table: {table.displayName}")
        print(f"  Range: {table.ref}")
else:
    print("\nNo tables found yet")

print("\n" + "="*80)
print("CURRENT STRUCTURE")
print("="*80)

print("\nAll data in Config_Lists:")
print("\nFirst 10 rows, all columns:")

for row in range(1, 11):
    row_data = []
    for col in range(1, 10):
        cell = ws.cell(row, col)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col)
            val_str = str(cell.value)[:20]
            row_data.append(f"{col_letter}:{val_str}")

    if row_data:
        print(f"  Row {row}: {', '.join(row_data)}")

print("\nContinuing through row 80...")
for row in range(11, 81):
    row_data = []
    for col in range(1, 10):
        cell = ws.cell(row, col)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col)
            val_str = str(cell.value)[:15]
            row_data.append(f"{col_letter}:{val_str}")

    if row_data and row % 10 == 0:
        print(f"  Row {row}: {', '.join(row_data[:3])}")

wb.close()

print("\n" + "="*80)
print("DONE")
print("="*80)
