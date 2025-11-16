import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("CREATING DYNAMIC PROJECT SPOTLIGHT - OPTION 1")
print("=" * 70)

# Load the workbook
print("\n[1/4] Loading workbook...")
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-CLEAN.xlsx')
print("✓ Workbook loaded")

# ==============================================================================
# STEP 1: Create Project_Technologies Sheet
# ==============================================================================
print("\n[2/4] Creating Project_Technologies sheet...")

if 'Project_Technologies' in wb.sheetnames:
    del wb['Project_Technologies']

ws_tech = wb.create_sheet('Project_Technologies')

# Headers
tech_headers = ['Project_ID', 'Technology_Area', 'Priority', 'Status', 'Notes']
for col, header in enumerate(tech_headers, 1):
    cell = ws_tech.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Sample data for existing projects
sample_technologies = [
    # PRJ-001: Digital Transformation
    ('PRJ-001', 'Cloud Computing', 'High', 'Active', 'AWS/Azure migration'),
    ('PRJ-001', 'API Integration', 'High', 'Active', 'RESTful APIs'),
    ('PRJ-001', 'Data Analytics', 'Medium', 'Planned', 'BI dashboards'),

    # PRJ-002: Cybersecurity Enhancement
    ('PRJ-002', 'Cybersecurity', 'Critical', 'Active', 'Zero Trust implementation'),
    ('PRJ-002', 'Encryption', 'High', 'Active', 'End-to-end encryption'),
    ('PRJ-002', 'Network Security', 'High', 'Active', 'Firewall upgrades'),

    # PRJ-003: Green Energy
    ('PRJ-003', 'IoT', 'High', 'Active', 'Smart grid sensors'),
    ('PRJ-003', 'Data Analytics', 'Medium', 'Active', 'Energy optimization'),

    # PRJ-004: Healthcare Data
    ('PRJ-004', 'Database Management', 'High', 'Active', 'SQL/NoSQL integration'),
    ('PRJ-004', 'Data Security', 'Critical', 'Active', 'HIPAA compliance'),
    ('PRJ-004', 'Machine Learning', 'Medium', 'Planned', 'Predictive analytics'),

    # PRJ-005: Smart Cities
    ('PRJ-005', 'IoT', 'Critical', 'Active', 'City-wide sensor network'),
    ('PRJ-005', 'Edge Computing', 'High', 'Active', 'Real-time processing'),
    ('PRJ-005', 'Data Analytics', 'High', 'Active', 'Traffic optimization'),
]

for row, tech_data in enumerate(sample_technologies, 2):
    for col, value in enumerate(tech_data, 1):
        ws_tech.cell(row=row, column=col, value=value)

# Set column widths
ws_tech.column_dimensions['A'].width = 12
ws_tech.column_dimensions['B'].width = 25
ws_tech.column_dimensions['C'].width = 12
ws_tech.column_dimensions['D'].width = 12
ws_tech.column_dimensions['E'].width = 30

print(f"✓ Created Project_Technologies with {len(sample_technologies)} sample records")

# ==============================================================================
# STEP 2: Update Project Spotlight with Dynamic Sections
# ==============================================================================
print("\n[3/4] Setting up dynamic Project Spotlight...")

ws_spotlight = wb['Project_Spotlight']

# First, unmerge any merged cells in the area we'll be working with (rows 14-50)
merged_ranges = list(ws_spotlight.merged_cells.ranges)
for merged_range in merged_ranges:
    if merged_range.min_row >= 14:
        try:
            ws_spotlight.unmerge_cells(str(merged_range))
            print(f"  Unmerged: {merged_range}")
        except KeyError:
            # Cell already unmerged or doesn't exist, skip
            print(f"  Skipped (already unmerged): {merged_range}")

print("✓ Unmerged cells in working area")

# === TARGET TECHNOLOGIES SECTION ===
row = 14
ws_spotlight.merge_cells(f'B{row}:G{row}')
ws_spotlight[f'B{row}'] = 'TARGET TECHNOLOGIES'
ws_spotlight[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws_spotlight[f'B{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_spotlight[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

# Technology table headers
row = 15
tech_display_headers = ['Technology Area', 'Priority', 'Status', 'Notes']
for col, header in enumerate(tech_display_headers, 2):
    cell = ws_spotlight.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

# Add placeholder text for technologies (formulas cause issues in older Excel)
ws_spotlight['B16'] = '[Technologies will appear here when project selected]'
ws_spotlight['B17'] = 'Add data to Project_Technologies sheet'
ws_spotlight.merge_cells('B16:E16')
ws_spotlight.merge_cells('B17:E17')
ws_spotlight['B16'].font = Font(italic=True, color="666666")
ws_spotlight['B17'].font = Font(italic=True, color="666666", size=9)

print("✓ Added TARGET TECHNOLOGIES section (rows 14-17)")

# === TARGET AUDIENCES SECTION ===
row = 20
ws_spotlight.merge_cells(f'B{row}:G{row}')
ws_spotlight[f'B{row}'] = 'TARGET AUDIENCES'
ws_spotlight[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws_spotlight[f'B{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_spotlight[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

# Audience table headers
row = 21
aud_display_headers = ['Audience Type', 'Region', 'Description', 'Priority']
for col, header in enumerate(aud_display_headers, 2):
    cell = ws_spotlight.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

# Add placeholder
ws_spotlight['B22'] = '[Audiences will appear here when project selected]'
ws_spotlight['B23'] = 'Data from Project_Audiences sheet'
ws_spotlight.merge_cells('B22:E22')
ws_spotlight.merge_cells('B23:E23')
ws_spotlight['B22'].font = Font(italic=True, color="666666")
ws_spotlight['B23'].font = Font(italic=True, color="666666", size=9)

print("✓ Added TARGET AUDIENCES section (rows 20-23)")

# === KEY DELIVERABLES SECTION ===
row = 26
ws_spotlight.merge_cells(f'B{row}:G{row}')
ws_spotlight[f'B{row}'] = 'KEY DELIVERABLES'
ws_spotlight[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws_spotlight[f'B{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_spotlight[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

# Deliverables table headers
row = 27
deliv_display_headers = ['Deliverable Name', 'Type', 'Due Date', 'Status', 'Owner', 'Progress']
for col, header in enumerate(deliv_display_headers, 2):
    cell = ws_spotlight.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

# Add placeholder
ws_spotlight['B28'] = '[Deliverables will appear here when project selected]'
ws_spotlight['B29'] = 'Data from Project_Deliverables sheet'
ws_spotlight.merge_cells('B28:F28')
ws_spotlight.merge_cells('B29:F29')
ws_spotlight['B28'].font = Font(italic=True, color="666666")
ws_spotlight['B29'].font = Font(italic=True, color="666666", size=9)

print("✓ Added KEY DELIVERABLES section (rows 26-29)")

# Add instructions
ws_spotlight['B31'] = 'INSTRUCTIONS FOR DYNAMIC DATA:'
ws_spotlight['B31'].font = Font(bold=True, size=10)
ws_spotlight['B32'] = '1. Add formulas manually to pull data from source sheets based on $B$2'
ws_spotlight['B33'] = '2. Use FILTER() function (Excel 365) or INDEX/MATCH arrays'
ws_spotlight['B34'] = '3. Example: =FILTER(Project_Technologies!B:B, Project_Technologies!A:A=$B$2)'
ws_spotlight['B35'] = '4. Or manually copy/paste data for each project'

for row in range(32, 36):
    ws_spotlight[f'B{row}'].font = Font(size=9, color="666666")

print("✓ Added instructions")

# Adjust column widths
ws_spotlight.column_dimensions['B'].width = 30
ws_spotlight.column_dimensions['C'].width = 15
ws_spotlight.column_dimensions['D'].width = 15
ws_spotlight.column_dimensions['E'].width = 20
ws_spotlight.column_dimensions['F'].width = 15
ws_spotlight.column_dimensions['G'].width = 12

print("✓ Adjusted column widths")

# ==============================================================================
# STEP 3: Save
# ==============================================================================
print("\n[4/4] Saving workbook...")
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-FINAL-v2.xlsx'
wb.save(output_file)
print(f"✓ Saved: {output_file}")

print(f"\n{'='*70}")
print("DYNAMIC PROJECT SPOTLIGHT STRUCTURE COMPLETE!")
print(f"{'='*70}")
print("\nWhat was created:")
print("  ✓ Project_Technologies sheet (14 sample records)")
print("  ✓ TARGET TECHNOLOGIES section in Project Spotlight")
print("  ✓ TARGET AUDIENCES section")
print("  ✓ KEY DELIVERABLES section")
print("\nNext Steps:")
print("  1. Add formulas manually to pull dynamic data OR")
print("  2. Manually populate data for each project")
print("\nData Sources:")
print("  • Technologies → Project_Technologies sheet")
print("  • Audiences → Project_Audiences sheet")
print("  • Deliverables → Project_Deliverables sheet")
print(f"\nFile: {output_file}")
print("\nNOTE: Placeholders added instead of formulas to avoid Excel compatibility issues")
print("      You can add FILTER() or INDEX/MATCH formulas as needed")
