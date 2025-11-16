"""
Create Project_Documents sheet with table structure
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

print("="*80)
print("CREATING PROJECT_DOCUMENTS SHEET")
print("="*80)

# Load workbook
print("\nLoading v18...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')

# Create new sheet
print("Creating Project_Documents sheet...")
ws = wb.create_sheet("Project_Documents")

# Add headers
headers = [
    'Document_ID',
    'Project_Unique_ID',
    'Document_Name',
    'Document_Type',
    'Document_Link',
    'Date_Added',
    'Notes'
]

print("Adding headers...")
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

# Set column widths
ws.column_dimensions['A'].width = 20  # Document_ID
ws.column_dimensions['B'].width = 18  # Project_Unique_ID
ws.column_dimensions['C'].width = 35  # Document_Name
ws.column_dimensions['D'].width = 15  # Document_Type
ws.column_dimensions['E'].width = 50  # Document_Link
ws.column_dimensions['F'].width = 12  # Date_Added
ws.column_dimensions['G'].width = 30  # Notes

# Create table
print("Creating table...")
tab = Table(displayName="T_Project_Documents", ref="A1:G100")

# Add table style
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style
ws.add_table(tab)

# Add data validation for Project_Unique_ID (column B)
print("Adding data validation...")
dv = DataValidation(
    type="list",
    formula1="ProjectList",
    allow_blank=True
)
dv.error = 'Please select a project from the dropdown'
dv.errorTitle = 'Invalid Project ID'
ws.add_data_validation(dv)
dv.add('B2:B100')

# Add sample formula in A2 for Document_ID auto-generation
print("Adding Document_ID formula...")
ws['A2'] = '=IF(B2="","",B2&"-DOC-"&TEXT(ROW()-1,"000"))'

print("\nSaving workbook...")
wb.save('2025-10-26-Tracker-v18.xlsx')

print("\n" + "="*80)
print("PROJECT_DOCUMENTS SHEET CREATED!")
print("="*80)
print("\nSheet created with:")
print("  - 7 columns: Document_ID, Project_Unique_ID, Document_Name,")
print("               Document_Type, Document_Link, Date_Added, Notes")
print("  - Excel Table: T_Project_Documents")
print("  - Dropdown validation on Project_Unique_ID column")
print("  - Auto-generating Document_ID formula")
print("\nDocument_ID format: PRJ-001-DOC-001, PRJ-001-DOC-002, etc.")
