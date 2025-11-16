"""
Move NCE_Eligible from G2 to H2 and add label in G2
"""
import openpyxl

print("="*80)
print("MOVING NCE_ELIGIBLE TO H2")
print("="*80)

# Load v29
print("\nLoading v29...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v29.xlsx')

ws_spotlight = wb['Spotlight_PMWorkspace']

print("\n" + "="*80)
print("CURRENT ROW 2")
print("="*80)

print("\nBefore changes:")
for col in range(5, 11):  # E-J
    cell = ws_spotlight.cell(2, col)
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)[:50]
            print(f"  {col_letter}2: FORMULA - {formula}...")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  {col_letter}2: {clean_value}")
    else:
        print(f"  {col_letter}2: [empty]")

print("\n" + "="*80)
print("MAKING CHANGES")
print("="*80)

print("\nStep 1: Get formula from G2")
g2_formula = ws_spotlight['G2'].value
if g2_formula:
    print(f"  Found formula in G2")
else:
    print(f"  G2 is empty - nothing to move")

print("\nStep 2: Add label to G2")
ws_spotlight['G2'] = 'NCE Eligible:'
print(f"  G2: 'NCE Eligible:'")

print("\nStep 3: Move formula to H2")
if g2_formula:
    ws_spotlight['H2'] = g2_formula
    print(f"  H2: Formula moved from G2")
else:
    # Create formula if needed
    formula = '=IF($B$2="","",IFERROR(IF(INDEX(T_Master_Projects[NCE_Eligible],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))="","",INDEX(T_Master_Projects[NCE_Eligible],MATCH($B$2,T_Master_Projects[Project_Unique_ID],0))),""))'
    ws_spotlight['H2'] = formula
    print(f"  H2: Formula added")

print("\n" + "="*80)
print("AFTER CHANGES")
print("="*80)

print("\nRow 2 structure now:")
for col in range(5, 11):  # E-J
    cell = ws_spotlight.cell(2, col)
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.value:
        if cell.data_type == 'f':
            formula = str(cell.value)[:50]
            print(f"  {col_letter}2: FORMULA - {formula}...")
        else:
            clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')
            print(f"  {col_letter}2: {clean_value}")
    else:
        print(f"  {col_letter}2: [empty]")

print("\n" + "="*80)
print("SAVING V29")
print("="*80)

wb.save('2025-10-26-Tracker-v29.xlsx')

print("\nOK - NCE_Eligible moved to H2!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nRow 2 structure (matches row 3 pattern):")
print("  E2: 'Award Number:' (label)")
print("  F2: Award Number formula")
print("  G2: 'NCE Eligible:' (label)")
print("  H2: NCE_Eligible formula")
print("  I2: 'NCE Status:' (label)")
print("  J2: NCE_Status formula")

print("\nAll formulas show BLANK when:")
print("  - No project selected ($B$2 is empty)")
print("  - Field value is empty")

print("\nv29 updated!")
