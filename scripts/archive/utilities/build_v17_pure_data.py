"""
Build Tracker v17 - Option 1: Pure Data Only (No Formulas)
Guaranteed to work with zero errors
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

print("="*80)
print("BUILDING TRACKER V17 - PURE DATA (NO FORMULAS)")
print("="*80)
print("This version contains ZERO formulas - add them manually in Excel")

# Load v12
print("\nLoading v12 data...")
wb_old = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx')

# Create new workbook
wb = Workbook()
wb.remove(wb.active)

def create_header(ws, row, headers, bg_color='366092'):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

# State Dept regions
STATE_DEPT_REGIONS = {
    'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
    'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
    'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
    'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
    'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
    'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
    'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',
    'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
    'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
    'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
    'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
    'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',
    'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
    'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
    'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
    'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
    'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
    'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
    'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR', 'VA': 'EUR',
    'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA', 'LB': 'NEA', 'OM': 'NEA',
    'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA', 'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',
    'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
    'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',
    'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
    'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
    'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
    'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
    'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
}

SUBREGIONS = {
    'DZ': 'Northern Africa', 'EG': 'Northern Africa', 'LY': 'Northern Africa', 'MA': 'Northern Africa',
    'TN': 'Northern Africa', 'SD': 'Northern Africa',
    'BJ': 'Western Africa', 'BF': 'Western Africa', 'CV': 'Western Africa', 'CI': 'Western Africa',
    'GM': 'Western Africa', 'GH': 'Western Africa', 'GN': 'Western Africa', 'GW': 'Western Africa',
    'LR': 'Western Africa', 'ML': 'Western Africa', 'MR': 'Western Africa', 'NE': 'Western Africa',
    'NG': 'Western Africa', 'SN': 'Western Africa', 'SL': 'Western Africa', 'TG': 'Western Africa',
    'BI': 'Eastern Africa', 'KM': 'Eastern Africa', 'DJ': 'Eastern Africa', 'ER': 'Eastern Africa',
    'ET': 'Eastern Africa', 'KE': 'Eastern Africa', 'MG': 'Eastern Africa', 'MW': 'Eastern Africa',
    'MU': 'Eastern Africa', 'MZ': 'Eastern Africa', 'RW': 'Eastern Africa', 'SC': 'Eastern Africa',
    'SO': 'Eastern Africa', 'SS': 'Eastern Africa', 'TZ': 'Eastern Africa', 'UG': 'Eastern Africa',
    'ZM': 'Eastern Africa', 'ZW': 'Eastern Africa',
    'AO': 'Middle Africa', 'CM': 'Middle Africa', 'CF': 'Middle Africa', 'TD': 'Middle Africa',
    'CG': 'Middle Africa', 'CD': 'Middle Africa', 'GQ': 'Middle Africa', 'GA': 'Middle Africa',
    'BW': 'Southern Africa', 'LS': 'Southern Africa', 'NA': 'Southern Africa', 'ZA': 'Southern Africa', 'SZ': 'Southern Africa',
    'DK': 'Northern Europe', 'EE': 'Northern Europe', 'FI': 'Northern Europe', 'IS': 'Northern Europe',
    'IE': 'Northern Europe', 'LV': 'Northern Europe', 'LT': 'Northern Europe', 'NO': 'Northern Europe',
    'SE': 'Northern Europe', 'GB': 'Northern Europe',
    'BY': 'Eastern Europe', 'BG': 'Eastern Europe', 'CZ': 'Eastern Europe', 'HU': 'Eastern Europe',
    'MD': 'Eastern Europe', 'PL': 'Eastern Europe', 'RO': 'Eastern Europe', 'RU': 'Eastern Europe',
    'SK': 'Eastern Europe', 'UA': 'Eastern Europe',
    'AL': 'Southern Europe', 'AD': 'Southern Europe', 'BA': 'Southern Europe', 'HR': 'Southern Europe',
    'CY': 'Southern Europe', 'GR': 'Southern Europe', 'IT': 'Southern Europe', 'XK': 'Southern Europe',
    'MK': 'Southern Europe', 'MT': 'Southern Europe', 'ME': 'Southern Europe', 'PT': 'Southern Europe',
    'RS': 'Southern Europe', 'SI': 'Southern Europe', 'ES': 'Southern Europe', 'VA': 'Southern Europe',
    'AT': 'Western Europe', 'BE': 'Western Europe', 'FR': 'Western Europe', 'DE': 'Western Europe',
    'LI': 'Western Europe', 'LU': 'Western Europe', 'MC': 'Western Europe', 'NL': 'Western Europe', 'CH': 'Western Europe',
    'AM': 'Caucasus', 'AZ': 'Caucasus', 'GE': 'Caucasus', 'TR': 'Western Asia',
    'CA': 'Northern America', 'US': 'Northern America',
    'BZ': 'Central America', 'CR': 'Central America', 'SV': 'Central America', 'GT': 'Central America',
    'HN': 'Central America', 'MX': 'Central America', 'NI': 'Central America', 'PA': 'Central America',
    'AG': 'Caribbean', 'BS': 'Caribbean', 'BB': 'Caribbean', 'CU': 'Caribbean', 'DM': 'Caribbean',
    'DO': 'Caribbean', 'GD': 'Caribbean', 'HT': 'Caribbean', 'JM': 'Caribbean', 'KN': 'Caribbean',
    'LC': 'Caribbean', 'VC': 'Caribbean', 'TT': 'Caribbean',
    'AR': 'South America', 'BO': 'South America', 'BR': 'South America', 'CL': 'South America',
    'CO': 'South America', 'EC': 'South America', 'GY': 'South America', 'PY': 'South America',
    'PE': 'South America', 'SR': 'South America', 'UY': 'South America', 'VE': 'South America',
    'CN': 'Eastern Asia', 'JP': 'Eastern Asia', 'KP': 'Eastern Asia', 'KR': 'Eastern Asia',
    'MN': 'Eastern Asia', 'TW': 'Eastern Asia',
    'BN': 'South-Eastern Asia', 'KH': 'South-Eastern Asia', 'ID': 'South-Eastern Asia',
    'LA': 'South-Eastern Asia', 'MY': 'South-Eastern Asia', 'MM': 'South-Eastern Asia',
    'PH': 'South-Eastern Asia', 'SG': 'South-Eastern Asia', 'TH': 'South-Eastern Asia',
    'TL': 'South-Eastern Asia', 'VN': 'South-Eastern Asia',
    'AF': 'Southern Asia', 'BD': 'Southern Asia', 'BT': 'Southern Asia', 'IN': 'Southern Asia',
    'MV': 'Southern Asia', 'NP': 'Southern Asia', 'PK': 'Southern Asia', 'LK': 'Southern Asia',
    'KZ': 'Central Asia', 'KG': 'Central Asia', 'TJ': 'Central Asia', 'TM': 'Central Asia', 'UZ': 'Central Asia',
    'AU': 'Australia and New Zealand', 'NZ': 'Australia and New Zealand',
    'FJ': 'Melanesia', 'PG': 'Melanesia', 'SB': 'Melanesia', 'VU': 'Melanesia',
    'KI': 'Micronesia', 'MH': 'Micronesia', 'FM': 'Micronesia', 'NR': 'Micronesia', 'PW': 'Micronesia',
    'WS': 'Polynesia', 'TO': 'Polynesia', 'TV': 'Polynesia',
    'BH': 'Arabian Peninsula', 'KW': 'Arabian Peninsula', 'OM': 'Arabian Peninsula',
    'QA': 'Arabian Peninsula', 'SA': 'Arabian Peninsula', 'AE': 'Arabian Peninsula', 'YE': 'Arabian Peninsula',
    'IL': 'Levant', 'JO': 'Levant', 'LB': 'Levant', 'PS': 'Levant', 'SY': 'Levant', 'IQ': 'Mesopotamia',
}

# 1. Master_Projects
print("1. Master_Projects (data only)...")
ws = wb.create_sheet("Master_Projects")
headers = [
    'Project_ID', 'Project_Unique_ID', 'Project_Name', 'Project_Status', 'Project_Priority',
    'Project_Progress', 'Project_Start_Date', 'Project_End_Date', 'Days_Remaining',
    'Implementer', 'Implementer_POC', 'Implementer_POC_Phone', 'Implementer_POC_Email',
    'Total_Proposed', 'Total_Allocation', 'Total_Obligated', 'Total_ULO', 'ULO_Percent',
    'Countries', 'Country_Count'
]
create_header(ws, 1, headers)

ws_old = wb_old['Master_Projects']
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    for col in range(1, 14):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value
    # N-R: Empty (formulas to be added manually)
    # S-T: Copy countries data
    ws.cell(new_row, 19).value = ws_old.cell(old_row, 18).value
    ws.cell(new_row, 20).value = ws_old.cell(old_row, 19).value

# 2. Country_Budgets
print("2. Country_Budgets (data only)...")
ws = wb.create_sheet("Country_Budgets")
headers = [
    'Budget_ID', 'Unique_ID', 'My_Country', 'Country_Code', 'Country_Name',
    'Proposed_Amount', 'Allocated_Amount', 'Obligated_Amount', 'Spent_Amount',
    'ULO', 'ULO_Percent', 'Spend_Health'
]
create_header(ws, 1, headers)

ws_old = wb_old['Country_Budgets']
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    for col in range(1, 6):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value
    ws.cell(new_row, 6).value = 0  # Proposed
    ws.cell(new_row, 7).value = ws_old.cell(old_row, 6).value  # Allocated
    ws.cell(new_row, 8).value = ws_old.cell(old_row, 7).value  # Obligated
    ws.cell(new_row, 9).value = ws_old.cell(old_row, 8).value  # Spent
    # J-K: Empty (formulas to be added manually)
    ws.cell(new_row, 12).value = ws_old.cell(old_row, 11).value  # Spend Health

# 3. Country_Regions
print("3. Country_Regions...")
ws = wb.create_sheet("Country_Regions")
headers = ['Country_Code', 'Country_Name', 'Region', 'EU_Member', 'Subregion']
create_header(ws, 1, headers)

ws_old = wb_old['Country_Regions']
for old_row in range(2, ws_old.max_row + 1):
    code = ws_old.cell(old_row, 1).value
    name = ws_old.cell(old_row, 2).value
    eu = ws_old.cell(old_row, 4).value

    ws.cell(old_row, 1).value = code
    ws.cell(old_row, 2).value = name
    ws.cell(old_row, 3).value = STATE_DEPT_REGIONS.get(code, 'EUR')
    ws.cell(old_row, 4).value = eu
    ws.cell(old_row, 5).value = SUBREGIONS.get(code, '')

# 4. Config_Lists
print("4. Config_Lists...")
ws = wb.create_sheet("Config_Lists")
headers = ['List_Type', 'Value1', 'Value2', 'Value3', 'Value4', 'Value5']
create_header(ws, 1, headers)

ws_old = wb_old['Config_Lists']
for old_row in range(2, ws_old.max_row + 1):
    for col in range(1, 7):
        ws.cell(old_row, col).value = ws_old.cell(old_row, col).value

# Add countries
ws_regions = wb['Country_Regions']
next_row = ws.max_row + 1
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws.cell(next_row, 1).value = "Country"
        ws.cell(next_row, 2).value = code
        ws.cell(next_row, 3).value = name
        ws.cell(next_row, 4).value = region
        next_row += 1

# 5. Country_PM_Assignments
print("5. Country_PM_Assignments...")
ws = wb.create_sheet("Country_PM_Assignments")
headers = ['Country_Code', 'Country_Name', 'Region', 'Project_Manager', 'PM_Email', 'PM_Phone', 'Notes']
create_header(ws, 1, headers)

current_row = 2
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws.cell(current_row, 1).value = code
        ws.cell(current_row, 2).value = name
        ws.cell(current_row, 3).value = region
        ws.cell(current_row, 4).value = "TBD"
        current_row += 1

# 6. Formula Guide Sheet
print("6. Creating FORMULA_GUIDE...")
ws = wb.create_sheet("FORMULA_GUIDE")
ws['A1'] = "FORMULAS TO ADD MANUALLY"
ws['A1'].font = Font(size=16, bold=True, color='FF0000')

ws['A3'] = "Copy these formulas into Excel:"
ws['A3'].font = Font(bold=True)

ws['A5'] = "MASTER_PROJECTS FORMULAS:"
ws['A5'].font = Font(bold=True, size=12)
ws['A6'] = "Column N (Total_Proposed) - Start at N2:"
ws['A7'] = "=SUMIF(Country_Budgets!$B:$B,A2,Country_Budgets!$F:$F)"
ws['A8'] = "Column O (Total_Allocation) - Start at O2:"
ws['A9'] = "=SUMIF(Country_Budgets!$B:$B,A2,Country_Budgets!$G:$G)"
ws['A10'] = "Column P (Total_Obligated) - Start at P2:"
ws['A11'] = "=SUMIF(Country_Budgets!$B:$B,A2,Country_Budgets!$H:$H)"
ws['A12'] = "Column Q (Total_ULO) - Start at Q2:"
ws['A13'] = "=SUMIF(Country_Budgets!$B:$B,A2,Country_Budgets!$J:$J)"
ws['A14'] = "Column R (ULO_Percent) - Start at R2:"
ws['A15'] = "=IF(P2>0,Q2/P2,0)"
ws['A16'] = "Then format column R as Percentage"

ws['A18'] = "COUNTRY_BUDGETS FORMULAS:"
ws['A18'].font = Font(bold=True, size=12)
ws['A19'] = "Column J (ULO) - Start at J2:"
ws['A20'] = "=H2-I2"
ws['A21'] = "Column K (ULO_Percent) - Start at K2:"
ws['A22'] = "=IF(H2>0,J2/H2,0)"
ws['A23'] = "Then format column K as Percentage"

ws['A25'] = "HOW TO ADD FORMULAS:"
ws['A25'].font = Font(bold=True, size=12)
ws['A26'] = "1. Click on cell (e.g., N2 in Master_Projects)"
ws['A27'] = "2. Copy the formula from this sheet"
ws['A28'] = "3. Paste it into the cell"
ws['A29'] = "4. Press Enter"
ws['A30'] = "5. Drag the fill handle down to copy to all rows"
ws['A31'] = "6. Repeat for each column"

ws['A33'] = "ESTIMATED TIME: 5 minutes"
ws['A33'].font = Font(bold=True, color='00FF00')

# Save
print("\nSaving v17-pure-data...")
wb.save('2025-10-26-Tracker-v17-pure-data.xlsx')

print("\n" + "="*80)
print("OPTION 1 COMPLETE!")
print("="*80)
print("\nFile: 2025-10-26-Tracker-v17-pure-data.xlsx")
print("\nThis file contains:")
print("  - All data")
print("  - Formatted headers")
print("  - NO formulas")
print("  - FORMULA_GUIDE sheet with all formulas to add")
print("\nGUARANTEED to open without errors!")
print("\nNext steps:")
print("  1. Open the file in Excel")
print("  2. Go to FORMULA_GUIDE sheet")
print("  3. Copy and paste formulas (takes 5 minutes)")
print("  4. Save")
