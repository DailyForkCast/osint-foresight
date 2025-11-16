"""
Fix Completion_Percent formatting in Project_Deliverables
Change from number to percentage format so 100 = 100% (not 10000%)
"""
import openpyxl
from openpyxl.styles import numbers

print("="*80)
print("FIXING COMPLETION_PERCENT FORMATTING")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_deliv = wb['Project_Deliverables']

print("\n" + "="*80)
print("STEP 1: FIND COMPLETION_PERCENT COLUMN")
print("="*80)

# Find the column
completion_col = None
for col in range(1, 15):
    header = ws_deliv.cell(1, col).value
    if header and 'completion' in header.lower() and 'percent' in header.lower():
        completion_col = col
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"\nFound: Column {col_letter} ({header})")
        break

if not completion_col:
    print("\nWARNING: Completion_Percent column not found!")
    print("Looking for any column with 'percent' or 'completion'...")
    for col in range(1, 15):
        header = ws_deliv.cell(1, col).value
        if header:
            print(f"  Column {col}: {header}")
else:
    print("\n" + "="*80)
    print("STEP 2: APPLY PERCENTAGE FORMAT")
    print("="*80)

    col_letter = openpyxl.utils.get_column_letter(completion_col)

    print(f"\nFormatting column {col_letter} with custom percent format...")
    print("  This allows entering: 100 to display as 100%")
    print("  Instead of: 100 displaying as 10000%")
    print("  Using custom format: 0\"%\" (shows % symbol without multiplying by 100)")

    # Apply custom percentage format to entire column (rows 2-100)
    for row in range(2, 101):
        cell = ws_deliv.cell(row, completion_col)
        # Set to custom format: number with % symbol (doesn't multiply by 100)
        cell.number_format = '0"%"'

    print(f"  Applied format '0\"%\"' to {col_letter}2:{col_letter}100")

    print("\n" + "="*80)
    print("SAVING V28")
    print("="*80)

    wb.save('2025-10-26-Tracker-v28.xlsx')

    print("\nOK - Completion_Percent formatting fixed!")

    print("\n" + "="*80)
    print("HOW TO USE")
    print("="*80)

    print("\nNow you can enter completion percentages as whole numbers:")
    print("  Enter: 100  →  Displays: 100%")
    print("  Enter: 75   →  Displays: 75%")
    print("  Enter: 50   →  Displays: 50%")
    print("  Enter: 0    →  Displays: 0%")

    print("\nNo more entering 1 for 100%!")
    print("Just enter the number you want to see.")

    print("\nv28 updated!")
