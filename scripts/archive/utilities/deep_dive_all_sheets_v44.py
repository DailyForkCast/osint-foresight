"""
Deep dive analysis of all sheets in v44
Identify what needs to be done - manual or automated
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

print("="*80)
print("DEEP DIVE - ALL SHEETS IN V44")
print("="*80)

wb = openpyxl.load_workbook('2025-10-26-Tracker-v44.xlsx')

analysis = {}

print(f"\nTotal sheets: {len(wb.sheetnames)}")
print(f"Sheets: {', '.join(wb.sheetnames)}")

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print("="*80)

    ws = wb[sheet_name]
    sheet_analysis = {
        'name': sheet_name,
        'issues': [],
        'recommendations': [],
        'manual_tasks': [],
        'automated_tasks': []
    }

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column

    print(f"\nDimensions: {max_row} rows x {max_col} columns")

    # Check for tables
    tables = list(ws.tables.values())
    if tables:
        print(f"\nTables found: {len(tables)}")
        for table in tables:
            print(f"  - {table.displayName}: {table.ref}")
    else:
        print("\nNo tables found")

    # Check for formulas
    formula_count = 0
    error_formulas = []
    for row in ws.iter_rows(max_row=min(max_row, 100)):
        for cell in row:
            if cell.data_type == 'f':
                formula_count += 1
                # Check for common errors
                if '#REF!' in str(cell.value) or 'REF!' in str(cell.value):
                    error_formulas.append(f"{cell.coordinate}: {str(cell.value)[:60]}")

    print(f"\nFormulas: {formula_count}")
    if error_formulas:
        print(f"ERROR FORMULAS FOUND: {len(error_formulas)}")
        for err in error_formulas[:5]:
            print(f"  - {err}")
        sheet_analysis['issues'].append(f"{len(error_formulas)} #REF! errors found")

    # Check for data validation
    data_validations = 0
    if hasattr(ws, 'data_validations'):
        data_validations = len(ws.data_validations.dataValidation)
    print(f"Data validations: {data_validations}")

    # Check for conditional formatting
    conditional_formats = 0
    if hasattr(ws, 'conditional_formatting'):
        conditional_formats = len(ws.conditional_formatting._cf_rules)
    print(f"Conditional formatting rules: {conditional_formats}")

    # Get first 20 rows to understand structure
    print(f"\nFirst 20 rows preview:")
    for row_num in range(1, min(21, max_row + 1)):
        row_data = []
        for col_num in range(1, min(max_col + 1, 15)):
            cell = ws.cell(row_num, col_num)
            if cell.value:
                if cell.data_type == 'f':
                    row_data.append(f"{get_column_letter(col_num)}{row_num}:FORMULA")
                else:
                    val_str = str(cell.value)[:20].encode('ascii', 'ignore').decode('ascii')
                    row_data.append(f"{get_column_letter(col_num)}{row_num}:{val_str}")
        if row_data:
            print(f"  Row {row_num}: {', '.join(row_data[:5])}")

    analysis[sheet_name] = sheet_analysis

wb.close()

print("\n" + "="*80)
print("SAVING ANALYSIS")
print("="*80)

with open('sheet_analysis_v44.json', 'w') as f:
    json.dump(analysis, f, indent=2)

print("\nAnalysis saved to sheet_analysis_v44.json")
print("\nNow running detailed analysis per sheet...")
