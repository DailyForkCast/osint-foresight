"""
Copy country data from v18 to v26
Includes all 5 columns: Country_Code, Country_Name, Region, EU_Member, Subregion
"""
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

print("="*80)
print("COPYING COUNTRY DATA FROM V18 TO V26")
print("="*80)

# Load both files
print("\nLoading v18...")
wb_v18 = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx')
ws_v18 = wb_v18['Country_Regions']

print("Loading v26...")
wb_v26 = openpyxl.load_workbook('2025-10-26-Tracker-v26.xlsx')
ws_v26 = wb_v26['Country_Regions']

# First, remove the existing T_Country_Regions table
print("\nRemoving old T_Country_Regions table...")
if 'T_Country_Regions' in ws_v26.tables:
    del ws_v26.tables['T_Country_Regions']
    print("  Old table removed")

# Get headers from v18
print("\nReading headers from v18...")
headers = []
for col in range(1, 10):
    cell = ws_v18.cell(1, col)
    if cell.value:
        headers.append(cell.value)
    else:
        break

print(f"  Headers: {', '.join(headers)}")

# Copy headers to v26
print("\nCopying headers to v26...")
for col_idx, header in enumerate(headers, start=1):
    ws_v26.cell(1, col_idx).value = header
print("  Headers copied")

# Count rows in v18
row_count = 0
row = 2
while ws_v18.cell(row, 1).value is not None:
    row_count += 1
    row += 1

print(f"\nFound {row_count} countries in v18")

# Copy all data rows
print("\nCopying data rows...")
for row in range(2, row_count + 2):
    for col in range(1, len(headers) + 1):
        value = ws_v18.cell(row, col).value
        ws_v26.cell(row, col).value = value

print(f"  Copied {row_count} rows")

# Create new table with correct range
print("\nCreating new T_Country_Regions table...")

# Calculate table range (include buffer rows for adding more countries)
last_data_row = row_count + 1
table_end_row = max(last_data_row + 20, 100)  # At least 20 empty rows or row 100
table_range = f'A1:{chr(64 + len(headers))}{table_end_row}'

print(f"  Table range: {table_range}")

# Create table
tab = Table(displayName="T_Country_Regions", ref=table_range)

# Add style
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

# Add table to sheet
ws_v26.add_table(tab)

print(f"  Table T_Country_Regions created")
print(f"  Columns: {', '.join(headers)}")

# Save v26
print("\n" + "="*80)
print("SAVING V26")
print("="*80)

wb_v26.save('2025-10-26-Tracker-v26.xlsx')

print("\nOK - Country data copied to v26!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print(f"\nCopied {row_count} countries from v18 to v26")
print(f"\nColumns included:")
for idx, header in enumerate(headers, 1):
    print(f"  {idx}. {header}")

print(f"\nTable T_Country_Regions:")
print(f"  Range: {table_range}")
print(f"  Data rows: {row_count}")
print(f"  Total capacity: {table_end_row - 1} rows")

print("\nv26 Country_Regions is now fully populated!")
