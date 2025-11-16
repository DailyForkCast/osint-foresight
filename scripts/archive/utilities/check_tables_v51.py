"""
Check which sheets have Excel Tables that can be referenced in data validation
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v51.xlsx')

print("="*80)
print("CHECKING FOR EXCEL TABLES")
print("="*80)

all_tables = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    tables = list(ws.tables.values())

    if tables:
        all_tables[sheet_name] = []
        for table in tables:
            all_tables[sheet_name].append({
                'name': table.displayName,
                'ref': table.ref
            })

print("\nSheets with Excel Tables:")
for sheet, tables in all_tables.items():
    print(f"\n{sheet}:")
    for table in tables:
        print(f"  - {table['name']}: {table['ref']}")

print("\n" + "="*80)
print("SHEETS WITHOUT TABLES (Need dynamic ranges)")
print("="*80)

sheets_without_tables = []
for sheet_name in wb.sheetnames:
    if sheet_name not in all_tables:
        sheets_without_tables.append(sheet_name)

for sheet in sheets_without_tables:
    print(f"  - {sheet}")

wb.close()

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print(f"\nSheets with tables: {len(all_tables)}")
print(f"Sheets without tables: {len(sheets_without_tables)}")
