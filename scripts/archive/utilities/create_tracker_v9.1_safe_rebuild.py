import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.1.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v9.1.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING TRACKER V9.1 - SAFE REBUILD WITH EXCEL 365 FILTER()')
print('='*80)
print()
print('Strategy: Minimal changes, careful formula insertion')
print()

# Load workbook
print('Loading v8.1...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('[OK] Loaded successfully')
print()

# ============================================================================
# PROJECT SPOTLIGHT - CAREFUL FILTER IMPLEMENTATION
# ============================================================================
if 'Project_Spotlight' not in wb.sheetnames:
    print('[ERROR] Project_Spotlight sheet not found!')
    exit(1)

ws = wb['Project_Spotlight']

print('='*80)
print('PROJECT SPOTLIGHT - Adding FILTER() Formulas')
print('='*80)
print()

# First, let's clear the target cells completely
print('Step 1: Clearing target cells...')

# Clear Target Audiences area (G6-I14)
for row in range(6, 15):
    for col in range(7, 10):  # G, H, I
        cell = ws.cell(row, col)
        cell.value = None

# Clear Target Technologies area (J6-K14)
for row in range(6, 15):
    for col in range(10, 12):  # J, K
        cell = ws.cell(row, col)
        cell.value = None

# Clear Key Deliverables area (B18-F27)
for row in range(18, 28):
    for col in range(2, 7):  # B, C, D, E, F
        cell = ws.cell(row, col)
        cell.value = None

print('[OK] Cleared all target cells')
print()

# Now add FILTER formulas one section at a time
print('Step 2: Adding Target Audiences FILTER formulas...')

# Target Audiences - G6, H6, I6
ws['G6'].value = '=IFERROR(FILTER(T_Project_Audiences[Audience_Type],T_Project_Audiences[Project_ID]=$B$2),"")'
ws['H6'].value = '=IFERROR(FILTER(T_Project_Audiences[Description],T_Project_Audiences[Project_ID]=$B$2),"")'
ws['I6'].value = '=IFERROR(FILTER(T_Project_Audiences[Priority],T_Project_Audiences[Project_ID]=$B$2),"")'

print('[OK] Added FILTER formulas to G6, H6, I6')
print('     Will spill to rows 7-14 automatically')
print()

print('Step 3: Adding Target Technologies FILTER formulas...')

# Target Technologies - J6, K6
ws['J6'].value = '=IFERROR(FILTER(Project_Technologies[Technology],Project_Technologies[Project_ID]=$B$2),"")'
ws['K6'].value = '=IFERROR(FILTER(Project_Technologies[Category],Project_Technologies[Project_ID]=$B$2),"")'

print('[OK] Added FILTER formulas to J6, K6')
print('     Will spill to rows 7-14 automatically')
print()

print('Step 4: Adding Key Deliverables FILTER formulas...')

# Key Deliverables - B18, C18, D18, E18, F18
ws['B18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Deliverable_Name],T_Project_Deliverables[Project_ID]=$B$2),"")'
ws['C18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Due_Date],T_Project_Deliverables[Project_ID]=$B$2),"")'
ws['D18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Status],T_Project_Deliverables[Project_ID]=$B$2),"")'
ws['E18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Owner],T_Project_Deliverables[Project_ID]=$B$2),"")'
ws['F18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Completion_Percent],T_Project_Deliverables[Project_ID]=$B$2),"")'

print('[OK] Added FILTER formulas to B18, C18, D18, E18, F18')
print('     Will spill to rows 19-27 automatically')
print()

print('Step 5: Adding Stakeholders section...')

stakeholder_start_row = 30

# Add section header
ws.cell(stakeholder_start_row - 1, 2).value = 'KEY STAKEHOLDERS'
ws.cell(stakeholder_start_row - 1, 2).font = Font(bold=True, size=12, color='FFFFFF')
ws.cell(stakeholder_start_row - 1, 2).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

# Add column headers
ws.cell(stakeholder_start_row, 2).value = 'Name'
ws.cell(stakeholder_start_row, 3).value = 'Title'
ws.cell(stakeholder_start_row, 4).value = 'Organization'
ws.cell(stakeholder_start_row, 5).value = 'Email'
ws.cell(stakeholder_start_row, 6).value = 'Type'

for col in range(2, 7):
    ws.cell(stakeholder_start_row, col).font = Font(bold=True)

# Clear stakeholder data area first
for row in range(stakeholder_start_row + 1, stakeholder_start_row + 11):
    for col in range(2, 7):
        ws.cell(row, col).value = None

# Add stakeholder FILTER formulas
stakeholder_data_row = stakeholder_start_row + 1

ws.cell(stakeholder_data_row, 2).value = '=IFERROR(FILTER(Stakeholders[Name],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
ws.cell(stakeholder_data_row, 3).value = '=IFERROR(FILTER(Stakeholders[Title],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
ws.cell(stakeholder_data_row, 4).value = '=IFERROR(FILTER(Stakeholders[Organization],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
ws.cell(stakeholder_data_row, 5).value = '=IFERROR(FILTER(Stakeholders[Email],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
ws.cell(stakeholder_data_row, 6).value = '=IFERROR(FILTER(Stakeholders[Stakeholder_Type],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'

print(f'[OK] Added stakeholder section at row {stakeholder_start_row}')
print('     FILTER formulas use SEARCH for comma-separated Project_IDs')
print()

print('Step 6: Visual enhancements...')

# Style section headers with professional blue
section_headers = [
    (4, 7),   # TARGET AUDIENCES (approximate row)
    (4, 10),  # TARGET TECHNOLOGIES (approximate row)
    (16, 2),  # KEY DELIVERABLES (approximate row)
    (stakeholder_start_row - 1, 2)  # KEY STAKEHOLDERS
]

for row, col in section_headers:
    cell = ws.cell(row, col)
    if cell.value and 'TARGET' in str(cell.value).upper() or 'KEY' in str(cell.value).upper() or 'STAKEHOLDER' in str(cell.value).upper():
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='left')

print('[OK] Applied professional styling to section headers')
print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving v9.1...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('V9.1 SAFE REBUILD COMPLETE!')
print('='*80)
print()
print('Summary of changes:')
print('  1. [OK] Target Audiences - FILTER() in G6/H6/I6')
print('  2. [OK] Target Technologies - FILTER() in J6/K6')
print('  3. [OK] Key Deliverables - FILTER() in B18/C18/D18/E18/F18')
print('  4. [OK] Key Stakeholders - FILTER() in B31/C31/D31/E31/F31')
print('  5. [OK] Professional visual styling applied')
print()
print('Excel 365 Features:')
print('  - FILTER() function for dynamic arrays')
print('  - Auto-spilling into rows below')
print('  - SEARCH() for comma-separated Project_IDs')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v9.1 in Excel 365')
print('  2. Go to Project_Spotlight sheet')
print('  3. Verify B2 contains "PRJ-001"')
print('  4. Check if formulas exist in: G6, J6, B18, B31')
print('  5. Verify deliverables auto-populate')
print()
print(f'Created: {timestamp}')
print()
