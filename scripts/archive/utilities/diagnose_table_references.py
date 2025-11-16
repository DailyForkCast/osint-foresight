"""
Diagnose table reference issues in v24
Check what tables exist and their names
"""
import openpyxl

print("="*80)
print("DIAGNOSING TABLE REFERENCES IN V24")
print("="*80)

# Load v24
print("\nLoading v24...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v24.xlsx', data_only=False)

print("\n" + "="*80)
print("EXCEL TABLES FOUND")
print("="*80)

tables_found = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.tables:
        for table in ws.tables.values():
            tables_found.append((sheet_name, table.name, table.ref))
            print(f"\nSheet: {sheet_name}")
            print(f"  Table Name: {table.name}")
            print(f"  Range: {table.ref}")

if not tables_found:
    print("\n  WARNING: NO TABLES FOUND!")

print("\n" + "="*80)
print("EXPECTED TABLES")
print("="*80)

expected_tables = [
    ('Master_Projects', 'T_Master_Projects'),
    ('Country_Budgets', 'T_Country_Budgets'),
    ('Country_Regions', 'T_Country_Regions'),
    ('Country_PM_Assignments', 'T_Country_PM_Assignments'),
    ('Stakeholders', 'T_Project_Stakeholders'),
    ('Project_Deliverables', 'T_Project_Deliverables'),
    ('Project_Audiences', 'T_Project_Audiences'),
    ('Project_Technologies', 'T_Project_Technologies'),
    ('Project_Documents', 'T_Project_Documents'),
]

print("\nExpected tables:")
for sheet, table_name in expected_tables:
    found = any(t[1] == table_name for t in tables_found)
    status = "OK" if found else "MISSING"
    print(f"  {status} - {table_name} (should be in {sheet})")

print("\n" + "="*80)
print("CHECKING STAKEHOLDER FORMULAS")
print("="*80)

ws_spot = wb['Spotlight_PMWorkspace']

print("\nChecking row 33 formulas...")
for col in [1, 5, 7, 8, 9]:  # A, E, G, H, I
    cell = ws_spot.cell(33, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.data_type == 'f':
        formula = str(cell.value)
        has_ref_error = '#REF!' in formula
        print(f"\n  {col_letter}33:")
        if has_ref_error:
            print(f"    ERROR - Contains #REF!")
            print(f"    Formula: {formula[:100]}...")
        else:
            print(f"    OK - {formula[:80]}...")
    else:
        print(f"\n  {col_letter}33: Not a formula (value: {cell.value})")

print("\n" + "="*80)
print("DIAGNOSIS")
print("="*80)

print("\nLikely issue:")
print("  The T_Project_Stakeholders table doesn't exist in v24")
print("  This could be because:")
print("    1. The table was never created")
print("    2. The table was lost during file modifications")
print("    3. The table has a different name")

print("\nSolution:")
print("  Check v20 to see if the table exists there")
print("  If yes, recreate from v20")
print("  If no, need to create the table manually or use a different base file")
