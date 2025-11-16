"""
Check E5 specifically
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')
ws_country = wb['Country_Dashboard']

print("="*80)
print("ROW 5 - ALL COLUMNS")
print("="*80)

for col in range(1, 12):
    cell = ws_country.cell(5, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        if cell.data_type == 'f':
            print(f"{col_letter}5: {cell.value}")
        else:
            print(f"{col_letter}5: {cell.value} (value)")
    else:
        print(f"{col_letter}5: (empty)")

wb.close()
