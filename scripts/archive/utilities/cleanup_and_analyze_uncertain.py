#!/usr/bin/env python3
"""
Step 1: Remove false positives from database
Step 2: Deep analyze uncertain contracts with enhanced heuristics
Step 3: Export to Excel for manual review comparison
"""

import sqlite3
import json
import re
import sys
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))
from core.enhanced_validation_v3_complete import CompleteEuropeanValidator


class TEDCleanupAndAnalyzer:
    """Remove false positives and deeply analyze uncertain contracts"""

    def __init__(self):
        self.db_path = Path("F:/OSINT_WAREHOUSE/osint_master.db")
        self.validator = CompleteEuropeanValidator()

        # Load revalidation results
        self.results_file = self.find_latest_revalidation_results()
        with open(self.results_file, 'r', encoding='utf-8') as f:
            self.all_results = json.load(f)

    def find_latest_revalidation_results(self):
        """Find the most recent revalidation detailed results"""
        analysis_dir = Path("analysis")
        files = list(analysis_dir.glob("ted_revalidation_detailed_*.json"))
        if not files:
            raise FileNotFoundError("No revalidation results found!")
        latest = max(files, key=lambda p: p.stat().st_mtime)
        print(f"Using revalidation results: {latest}")
        return latest

    def step1_remove_false_positives(self):
        """Remove 2 false positives from database"""

        print("\n" + "="*80)
        print("STEP 1: REMOVE FALSE POSITIVES FROM DATABASE")
        print("="*80)

        # Get false positive IDs
        false_positives = [r for r in self.all_results if r['category'] == 'FALSE_POSITIVE']

        if not false_positives:
            print("No false positives to remove!")
            return

        print(f"\nFound {len(false_positives)} false positives to remove:")
        for fp in false_positives:
            print(f"  - ID {fp['id']}: {fp.get('contractor_name', 'N/A')}")

        # Confirm
        print(f"\nWill update {len(false_positives)} records in ted_contracts_production")
        print("Setting is_chinese_related = 0, chinese_confidence = 0.0")

        # Execute update
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        for fp in false_positives:
            cursor.execute("""
                UPDATE ted_contracts_production
                SET is_chinese_related = 0,
                    chinese_confidence = 0.0,
                    detection_rationale = ?
                WHERE id = ?
            """, (f"False positive - removed in re-validation {timestamp}", fp['id']))

        conn.commit()
        rows_updated = cursor.rowcount
        conn.close()

        print(f"\n[SUCCESS] Updated {rows_updated} records")
        print(f"[SUCCESS] False positives removed from database")

        return false_positives

    def step2_deep_analyze_uncertain(self):
        """Deep analysis of uncertain contracts with enhanced heuristics"""

        print("\n" + "="*80)
        print("STEP 2: DEEP ANALYSIS OF UNCERTAIN CONTRACTS")
        print("="*80)

        uncertain = [r for r in self.all_results if r['category'] == 'UNCERTAIN']
        print(f"\nAnalyzing {len(uncertain)} uncertain contracts...")

        # Enhanced categorization
        categories = {
            'LIKELY_FALSE_POSITIVE': [],
            'LIKELY_TRUE_CHINESE': [],
            'NEEDS_MANUAL_REVIEW': []
        }

        for contract in uncertain:
            analysis = self.enhanced_heuristic_analysis(contract)
            categories[analysis['final_category']].append({
                **contract,
                'enhanced_analysis': analysis
            })

        # Print summary
        print("\n" + "-"*80)
        print("ENHANCED ANALYSIS RESULTS")
        print("-"*80)
        print(f"Likely False Positives:  {len(categories['LIKELY_FALSE_POSITIVE']):3d} ({len(categories['LIKELY_FALSE_POSITIVE'])/len(uncertain)*100:.1f}%)")
        print(f"Likely True Chinese:     {len(categories['LIKELY_TRUE_CHINESE']):3d} ({len(categories['LIKELY_TRUE_CHINESE'])/len(uncertain)*100:.1f}%)")
        print(f"Needs Manual Review:     {len(categories['NEEDS_MANUAL_REVIEW']):3d} ({len(categories['NEEDS_MANUAL_REVIEW'])/len(uncertain)*100:.1f}%)")
        print("-"*80)

        # Show reasoning breakdown
        print("\nReasoning Breakdown:")
        reasoning_counts = {}
        for category_name, contracts in categories.items():
            for c in contracts:
                reason = c['enhanced_analysis']['primary_reason']
                key = f"{category_name}: {reason}"
                reasoning_counts[key] = reasoning_counts.get(key, 0) + 1

        for reason, count in sorted(reasoning_counts.items(), key=lambda x: x[1], reverse=True)[:15]:
            print(f"  {reason}: {count}")

        # Save detailed analysis
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = Path(f"analysis/ted_uncertain_deep_analysis_{timestamp}.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(categories, f, indent=2, ensure_ascii=False)

        print(f"\n[SUCCESS] Detailed analysis saved to: {output_file}")

        return categories

    def enhanced_heuristic_analysis(self, contract):
        """Apply enhanced heuristics to categorize uncertain contract"""

        detection = contract.get('detection', {})
        patterns = detection.get('matched_patterns', [])
        contractor_name = contract.get('contractor_name', '')
        ca_country = contract.get('ca_country', '')
        contractor_country = detection.get('contractor_country', '')

        # Get original DB data for deeper analysis
        contract_id = contract['id']
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        full_contract = cursor.execute("""
            SELECT contract_title, contract_description, ca_name, contractor_address,
                   contractor_city, cpv_code, value_total, currency
            FROM ted_contracts_production
            WHERE id = ?
        """, (contract_id,)).fetchone()
        conn.close()

        if full_contract:
            contract_title = full_contract['contract_title'] or ''
            contract_desc = full_contract['contract_description'] or ''
            ca_name = full_contract['ca_name'] or ''
            contractor_address = full_contract['contractor_address'] or ''
            contractor_city = full_contract['contractor_city'] or ''
            cpv_code = full_contract['cpv_code'] or ''
        else:
            contract_title = contract_desc = ca_name = ''
            contractor_address = contractor_city = cpv_code = ''

        combined_text = ' '.join([
            contract_title, contract_desc, ca_name,
            contractor_name or '', contractor_address, contractor_city
        ]).lower()

        # Initialize scoring
        chinese_score = 0
        false_positive_score = 0
        reasons = []

        # HEURISTIC 1: Only "china" keyword - check context
        if len(patterns) == 1 and 'china' in patterns:
            # Check for "Made in China" / "Manufactured in China"
            if re.search(r'\b(made|manufactured|produced|sourced)\s+in\s+china\b', combined_text):
                false_positive_score += 50
                reasons.append("Made in China reference")

            # Check for project names mentioning China
            elif re.search(r'\b(china|chinese)\s+(project|initiative|program|cooperation|partnership)\b', combined_text):
                false_positive_score += 40
                reasons.append("China-related project (not Chinese contractor)")

            # Check for historical/cultural references
            elif re.search(r'\b(ancient|historic|traditional|cultural)\s+china\b', combined_text):
                false_positive_score += 60
                reasons.append("Historical/cultural China reference")

            # Check if contractor name is clearly European
            elif contractor_name and any(suffix in contractor_name.lower() for suffix in ['gmbh', 'sarl', 'spa', 'srl', 'ltd', 's.l.', 'bv', 'oy']):
                false_positive_score += 35
                reasons.append("European legal suffix in contractor name")

            # If European CA and no other indicators
            elif ca_country in ['DEU', 'FRA', 'ITA', 'ESP', 'POL']:
                false_positive_score += 25
                reasons.append("EU contracting authority, generic 'china' reference")

            else:
                reasons.append("Uncertain - single 'china' keyword")

        # HEURISTIC 2: Hong Kong / Shenzhen / Beijing mentions
        elif any(city in patterns for city in ['hong kong', 'shenzhen', 'beijing', 'shanghai']):
            # If it's a city name AND there's a contractor name, more likely real
            if contractor_name and contractor_name != 'N/A':
                chinese_score += 40
                reasons.append(f"Chinese city + contractor name present")

            # If just city mention in project description
            elif any(city in contract_title.lower() or city in contract_desc.lower() for city in ['hong kong', 'shenzhen', 'beijing', 'shanghai']):
                false_positive_score += 30
                reasons.append("City name in project title/description")

            else:
                reasons.append("Chinese city mentioned - context unclear")

        # HEURISTIC 3: "Chinese" keyword
        elif 'chinese' in patterns:
            # Check for "Chinese language", "Chinese studies", etc.
            if re.search(r'\bchinese\s+(language|studies|culture|art|history|medicine)\b', combined_text):
                false_positive_score += 50
                reasons.append("Chinese language/culture reference")

            # Check for "Chinese market", "Chinese economy"
            elif re.search(r'\bchinese\s+(market|economy|government|policy)\b', combined_text):
                false_positive_score += 40
                reasons.append("Chinese market/economy analysis")

            # Otherwise could be referring to Chinese company
            else:
                chinese_score += 20
                reasons.append("'Chinese' keyword - could be company reference")

        # HEURISTIC 4: Multiple patterns (higher confidence)
        if len(patterns) >= 2:
            chinese_score += 30
            reasons.append(f"Multiple patterns: {', '.join(patterns[:3])}")

        # HEURISTIC 5: Contractor location
        if contractor_country:
            if contractor_country in ['CN', 'CHN', 'HK', 'HKG', 'MO', 'MAC', 'TWN', 'TW']:
                chinese_score += 60
                reasons.append(f"Contractor in {contractor_country}")
            elif contractor_country in ['DEU', 'FRA', 'ITA', 'ESP', 'POL', 'NLD', 'BEL']:
                false_positive_score += 30
                reasons.append(f"European contractor ({contractor_country})")

        # HEURISTIC 6: CPV codes for goods/products (might indicate "Made in China")
        if cpv_code:
            # CPV 30000000 = Computer equipment, 31000000 = Electrical machinery
            # These often have "Made in China" in descriptions
            if cpv_code.startswith(('30', '31', '32', '35')):
                false_positive_score += 15
                reasons.append("Equipment/goods CPV - possible 'Made in China'")

        # HEURISTIC 7: Contract value (very large contracts more likely to be real)
        if full_contract and full_contract['value_total']:
            value = full_contract['value_total']
            if value > 10000000:  # >10M EUR
                chinese_score += 20
                reasons.append("High value contract (>10M)")

        # FINAL CATEGORIZATION
        if false_positive_score > chinese_score + 20:
            final_category = 'LIKELY_FALSE_POSITIVE'
            confidence = min(false_positive_score / 100, 0.9)
        elif chinese_score > false_positive_score + 20:
            final_category = 'LIKELY_TRUE_CHINESE'
            confidence = min(chinese_score / 100, 0.85)
        else:
            final_category = 'NEEDS_MANUAL_REVIEW'
            confidence = 0.5

        return {
            'final_category': final_category,
            'confidence': round(confidence, 2),
            'chinese_score': chinese_score,
            'false_positive_score': false_positive_score,
            'primary_reason': reasons[0] if reasons else 'No specific indicators',
            'all_reasons': reasons,
            'patterns_matched': patterns
        }

    def step3_export_to_excel(self, deep_analysis_categories):
        """Export uncertain contracts to Excel for manual review"""

        print("\n" + "="*80)
        print("STEP 3: EXPORT TO EXCEL FOR MANUAL REVIEW")
        print("="*80)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = Path(f"analysis/ted_uncertain_contracts_manual_review_{timestamp}.xlsx")

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets for each category
        for category_name, contracts in deep_analysis_categories.items():
            if not contracts:
                continue

            ws = wb.create_sheet(title=category_name[:31])  # Excel limit

            # Headers
            headers = [
                'ID', 'Notice Number', 'Date', 'CA Country', 'CA Name',
                'Contractor Name', 'Contractor Country', 'Contractor City',
                'Contract Title', 'Value', 'Currency',
                'Patterns Matched', 'Auto Category', 'Confidence',
                'Primary Reason', 'All Reasons',
                'MANUAL REVIEW', 'MANUAL NOTES'
            ]

            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add data
            for row_num, contract in enumerate(contracts, 2):
                analysis = contract.get('enhanced_analysis', {})

                # Get full contract data
                conn = sqlite3.connect(self.db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                full = cursor.execute("""
                    SELECT ca_name, contract_title, value_total, currency
                    FROM ted_contracts_production WHERE id = ?
                """, (contract['id'],)).fetchone()
                conn.close()

                ws.cell(row=row_num, column=1, value=contract['id'])
                ws.cell(row=row_num, column=2, value=contract.get('notice_number', ''))
                ws.cell(row=row_num, column=3, value=contract.get('publication_date', ''))
                ws.cell(row=row_num, column=4, value=contract.get('ca_country', ''))
                ws.cell(row=row_num, column=5, value=full['ca_name'] if full else '')
                ws.cell(row=row_num, column=6, value=contract.get('contractor_name', ''))
                ws.cell(row=row_num, column=7, value=contract.get('detection', {}).get('contractor_country', ''))
                ws.cell(row=row_num, column=8, value=contract.get('detection', {}).get('combined_text_sample', '')[:50])
                ws.cell(row=row_num, column=9, value=full['contract_title'][:100] if full and full['contract_title'] else '')
                ws.cell(row=row_num, column=10, value=full['value_total'] if full else '')
                ws.cell(row=row_num, column=11, value=full['currency'] if full else '')
                ws.cell(row=row_num, column=12, value=', '.join(analysis.get('patterns_matched', [])))
                ws.cell(row=row_num, column=13, value=analysis.get('final_category', ''))
                ws.cell(row=row_num, column=14, value=analysis.get('confidence', ''))
                ws.cell(row=row_num, column=15, value=analysis.get('primary_reason', ''))
                ws.cell(row=row_num, column=16, value=' | '.join(analysis.get('all_reasons', [])))

                # Color code by category
                if category_name == 'LIKELY_FALSE_POSITIVE':
                    fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif category_name == 'LIKELY_TRUE_CHINESE':
                    fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                else:
                    fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

                for col in range(1, 17):
                    ws.cell(row=row_num, column=col).fill = fill

            # Adjust column widths
            column_widths = [8, 15, 12, 10, 30, 35, 12, 25, 40, 12, 8, 25, 25, 10, 40, 50, 20, 40]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Freeze header row
            ws.freeze_panes = 'A2'

        # Add summary sheet
        summary_ws = wb.create_sheet(title="SUMMARY", index=0)
        summary_ws['A1'] = "TED Uncertain Contracts - Manual Review"
        summary_ws['A1'].font = Font(size=16, bold=True)

        summary_ws['A3'] = "Generated:"
        summary_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        summary_ws['A4'] = "Total Uncertain:"
        summary_ws['B4'] = sum(len(contracts) for contracts in deep_analysis_categories.values())

        summary_ws['A6'] = "Category Breakdown:"
        summary_ws['A6'].font = Font(bold=True)

        row = 7
        for category, contracts in deep_analysis_categories.items():
            summary_ws[f'A{row}'] = category
            summary_ws[f'B{row}'] = len(contracts)
            summary_ws[f'C{row}'] = f"{len(contracts)/sum(len(c) for c in deep_analysis_categories.values())*100:.1f}%"
            row += 1

        summary_ws['A10'] = "Instructions:"
        summary_ws['A10'].font = Font(bold=True)
        summary_ws['A11'] = "1. Review each contract in the category sheets"
        summary_ws['A12'] = "2. In MANUAL REVIEW column, enter: TRUE_CHINESE, FALSE_POSITIVE, or UNCERTAIN"
        summary_ws['A13'] = "3. Add notes in MANUAL NOTES column"
        summary_ws['A14'] = "4. Compare your manual review with the 'Auto Category' column"
        summary_ws['A15'] = "5. This helps validate the automated heuristics"

        # Save workbook
        wb.save(excel_file)

        print(f"\n[SUCCESS] Excel file created: {excel_file}")
        print(f"\n  Contains {sum(len(c) for c in deep_analysis_categories.values())} contracts across {len(deep_analysis_categories)} sheets")
        print(f"  - Likely False Positives: {len(deep_analysis_categories.get('LIKELY_FALSE_POSITIVE', []))}")
        print(f"  - Likely True Chinese: {len(deep_analysis_categories.get('LIKELY_TRUE_CHINESE', []))}")
        print(f"  - Needs Manual Review: {len(deep_analysis_categories.get('NEEDS_MANUAL_REVIEW', []))}")

        return excel_file

    def run_all_steps(self):
        """Execute all three steps"""

        print("\n" + "="*80)
        print("TED CLEANUP AND UNCERTAIN ANALYSIS")
        print("="*80)
        print()
        print("This script will:")
        print("  1. Remove 2 false positives from database")
        print("  2. Deep analyze 142 uncertain contracts")
        print("  3. Export to Excel for manual review comparison")
        print()

        # Step 1
        false_positives = self.step1_remove_false_positives()

        # Step 2
        deep_analysis = self.step2_deep_analyze_uncertain()

        # Step 3
        excel_file = self.step3_export_to_excel(deep_analysis)

        print("\n" + "="*80)
        print("ALL STEPS COMPLETE!")
        print("="*80)
        print(f"\n[SUCCESS] Removed {len(false_positives) if false_positives else 0} false positives from database")
        print(f"[SUCCESS] Analyzed {sum(len(c) for c in deep_analysis.values())} uncertain contracts")
        print(f"[SUCCESS] Exported to Excel: {excel_file}")
        print()
        print("NEXT: Open the Excel file and manually review contracts")
        print("      Compare your manual classification with the automated 'Auto Category'")
        print("      This will help validate the heuristics!")
        print()

        return deep_analysis, excel_file


if __name__ == '__main__':
    analyzer = TEDCleanupAndAnalyzer()
    deep_analysis, excel_file = analyzer.run_all_steps()
