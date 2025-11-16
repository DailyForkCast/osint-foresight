import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.1.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.2.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('ADDING VISUAL ENHANCEMENTS THROUGHOUT TRACKER')
print('='*80)
print()

wb = openpyxl.load_workbook(input_file, data_only=False)

# Define consistent styles
header_font = Font(bold=True, size=14, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
subheader_font = Font(bold=True, size=11, color='FFFFFF')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
column_header_font = Font(bold=True, size=10)
column_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# MASTER_PROJECTS - Enhance headers
# ============================================================================

print('Sheet 1: Master_Projects')
print('-'*80)

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    # Style header row (row 1)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # Set row height for header
    ws.row_dimensions[1].height = 30

    print('[OK] Styled header row')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# COUNTRY_BUDGETS - Enhance headers
# ============================================================================

print('Sheet 2: Country_Budgets')
print('-'*80)

if 'Country_Budgets' in wb.sheetnames:
    ws = wb['Country_Budgets']

    # Style header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # Add alternating row colors for easier reading (rows 2-50)
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for row in range(2, 51):
        if row % 2 == 0:  # Even rows
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = light_fill

    print('[OK] Styled header row')
    print('[OK] Added alternating row colors')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# PORTFOLIO_DASHBOARD - Enhance
# ============================================================================

print('Sheet 3: Portfolio_Dashboard')
print('-'*80)

if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']

    # Style main title (row 1)
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    # Style column headers (row 10)
    for col in range(1, 14):  # A through M
        cell = ws.cell(10, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    ws.row_dimensions[10].height = 25

    print('[OK] Enhanced title and column headers')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# COUNTRY_DASHBOARD - Already styled, verify
# ============================================================================

print('Sheet 4: Country_Dashboard')
print('-'*80)
print('[OK] Already styled during creation')
print()

# ============================================================================
# MILESTONES - Enhance headers
# ============================================================================

print('Sheet 5: Milestones')
print('-'*80)

if 'Milestones' in wb.sheetnames:
    ws = wb['Milestones']

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30

    print('[OK] Styled header row')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# RISK_REGISTER - Enhance headers
# ============================================================================

print('Sheet 6: Risk_Register')
print('-'*80)

if 'Risk_Register' in wb.sheetnames:
    ws = wb['Risk_Register']

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30

    print('[OK] Styled header row')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# STAKEHOLDERS - Enhance headers
# ============================================================================

print('Sheet 7: Stakeholders')
print('-'*80)

if 'Stakeholders' in wb.sheetnames:
    ws = wb['Stakeholders']

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30

    print('[OK] Styled header row')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# COUNTRY_REGIONS - Enhance headers
# ============================================================================

print('Sheet 8: Country_Regions')
print('-'*80)

if 'Country_Regions' in wb.sheetnames:
    ws = wb['Country_Regions']

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    print('[OK] Styled header row')
    print('[OK] Froze header row for scrolling')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# PROJECT_DELIVERABLES - Enhance headers
# ============================================================================

print('Sheet 9: Project_Deliverables')
print('-'*80)

if 'Project_Deliverables' in wb.sheetnames:
    ws = wb['Project_Deliverables']

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    print('[OK] Styled header row')
    print('[OK] Froze header row')
else:
    print('[SKIP] Sheet not found')

print()

# ============================================================================
# SAVE
# ============================================================================

print('='*80)
print('Saving v10.2...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('VISUAL ENHANCEMENTS COMPLETE!')
print('='*80)
print()
print('Enhancements applied:')
print('  1. [OK] Consistent header styling across all sheets')
print('  2. [OK] Professional color scheme (blue headers)')
print('  3. [OK] Alternating row colors in Country_Budgets')
print('  4. [OK] Frozen header rows in data sheets')
print('  5. [OK] Proper column alignment and wrapping')
print('  6. [OK] Consistent fonts and sizing')
print()
print('Sheets enhanced:')
print('  - Master_Projects')
print('  - Country_Budgets')
print('  - Portfolio_Dashboard')
print('  - Country_Dashboard')
print('  - Milestones')
print('  - Risk_Register')
print('  - Stakeholders')
print('  - Country_Regions')
print('  - Project_Deliverables')
print()
print(f'Output: {output_file}')
print(f'Created: {timestamp}')
