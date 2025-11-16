import openpyxl

file_path = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v9.1.xlsx'

print('='*80)
print('VERIFYING V9.1 FORMULAS')
print('='*80)
print()

# Load with data_only=False to see formulas
wb = openpyxl.load_workbook(file_path, data_only=False)

if 'Project_Spotlight' not in wb.sheetnames:
    print('[ERROR] Project_Spotlight not found!')
    exit(1)

ws = wb['Project_Spotlight']

print('Checking critical formula cells:')
print()

# Check Target Audiences
print('1. TARGET AUDIENCES:')
print(f'   G6 (Audience Type): {ws["G6"].value}')
print(f'   H6 (Description):   {ws["H6"].value}')
print(f'   I6 (Priority):      {ws["I6"].value}')
print()

# Check Target Technologies
print('2. TARGET TECHNOLOGIES:')
print(f'   J6 (Technology): {ws["J6"].value}')
print(f'   K6 (Category):   {ws["K6"].value}')
print()

# Check Key Deliverables
print('3. KEY DELIVERABLES:')
print(f'   B18 (Name):       {ws["B18"].value}')
print(f'   C18 (Due Date):   {ws["C18"].value}')
print(f'   D18 (Status):     {ws["D18"].value}')
print(f'   E18 (Owner):      {ws["E18"].value}')
print(f'   F18 (Completion): {ws["F18"].value}')
print()

# Check Stakeholders
print('4. KEY STAKEHOLDERS:')
print(f'   B31 (Name):         {ws["B31"].value}')
print(f'   C31 (Title):        {ws["C31"].value}')
print(f'   D31 (Organization): {ws["D31"].value}')
print(f'   E31 (Email):        {ws["E31"].value}')
print(f'   F31 (Type):         {ws["F31"].value}')
print()

# Check if any are None
issues = []
critical_cells = {
    'G6': ws["G6"].value,
    'J6': ws["J6"].value,
    'B18': ws["B18"].value,
    'B31': ws["B31"].value
}

for cell, value in critical_cells.items():
    if value is None:
        issues.append(f'{cell} is None')
    elif not str(value).startswith('='):
        issues.append(f'{cell} does not contain a formula')
    elif 'FILTER' not in str(value):
        issues.append(f'{cell} does not contain FILTER function')

if issues:
    print('='*80)
    print('[WARNING] ISSUES DETECTED:')
    for issue in issues:
        print(f'  - {issue}')
else:
    print('='*80)
    print('[OK] ALL FORMULAS VERIFIED!')
    print()
    print('All critical cells contain FILTER formulas.')

wb.close()
print('='*80)
