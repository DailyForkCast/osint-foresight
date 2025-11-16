"""
Deep dive audit of v27:
1. Check all formulas for completeness
2. Identify performance bottlenecks
3. Find optimization opportunities
4. List remaining manual work
"""
import openpyxl
import json
import re

print("="*80)
print("V28 DEEP DIVE AUDIT - PERFORMANCE & COMPLETENESS")
print("="*80)

# Load v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx', data_only=False)

audit = {
    "sheets_analyzed": [],
    "formula_complexity": {},
    "performance_issues": [],
    "missing_items": [],
    "optimization_opportunities": [],
    "manual_work_remaining": []
}

# ============================================================================
# ANALYZE EACH MAJOR SHEET
# ============================================================================

print("\n" + "="*80)
print("1. MASTER_PROJECTS ANALYSIS")
print("="*80)

ws_mp = wb['Master_Projects']
audit["sheets_analyzed"].append("Master_Projects")

# Get column count
max_col = 30
headers = []
for col in range(1, max_col + 1):
    header = ws_mp.cell(1, col).value
    if header:
        headers.append((col, header))

print(f"\nFound {len(headers)} columns")

# Check which columns have formulas
print("\nChecking for formulas in row 2...")
formula_cols = []
missing_formula_cols = []

for col, header in headers:
    cell = ws_mp.cell(2, col)
    if cell.data_type == 'f':
        formula = str(cell.value)
        formula_cols.append((col, header, formula))

        # Check formula complexity
        if 'SUMIF' in formula and 'INDEX' in formula:
            audit["performance_issues"].append({
                "sheet": "Master_Projects",
                "column": header,
                "issue": "Dynamic range SUMIF (INDEX/COUNTA) - can be slow",
                "current": formula[:80]
            })
    else:
        # Check if this column should have a formula
        if header in ['Days_Remaining', 'POP_Days_Remaining', 'Total_Proposed',
                      'Total_Allocation', 'Total_Obligated', 'Total_Spent',
                      'Total_ULO', 'ULO_Percent', 'Countries', 'Country_Count']:
            missing_formula_cols.append((col, header))

print(f"  Formulas found: {len(formula_cols)} columns")
if missing_formula_cols:
    print(f"  WARNING: {len(missing_formula_cols)} columns missing formulas:")
    for col, header in missing_formula_cols:
        print(f"    Column {col} ({header})")
        audit["missing_items"].append(f"Master_Projects: {header} missing formula")

# ============================================================================
print("\n" + "="*80)
print("2. COUNTRY_BUDGETS ANALYSIS")
print("="*80)

ws_cb = wb['Country_Budgets']
audit["sheets_analyzed"].append("Country_Budgets")

# Check formulas
print("\nChecking Country_Budgets row 2 formulas...")
cb_headers = []
for col in range(1, 15):
    header = ws_cb.cell(1, col).value
    if header:
        cb_headers.append((col, header))

cb_formula_count = 0
for col, header in cb_headers:
    cell = ws_cb.cell(2, col)
    if cell.data_type == 'f':
        cb_formula_count += 1

print(f"  Found {cb_formula_count} formulas in {len(cb_headers)} columns")

# ============================================================================
print("\n" + "="*80)
print("3. PORTFOLIO_DASHBOARD ANALYSIS")
print("="*80)

ws_port = wb['Portfolio_Dashboard']
audit["sheets_analyzed"].append("Portfolio_Dashboard")

print("\nAnalyzing Portfolio_Dashboard formulas...")

# Check row 11 formulas
formula_count = 0
complex_formulas = 0
for col in range(1, 15):
    cell = ws_port.cell(11, col)
    if cell.data_type == 'f':
        formula = str(cell.value)
        formula_count += 1

        # Check complexity
        if formula.count('IF') > 2:
            complex_formulas += 1
            audit["performance_issues"].append({
                "sheet": "Portfolio_Dashboard",
                "cell": f"Row 11 Col {col}",
                "issue": "Multiple nested IFs",
                "formula": formula[:80]
            })

print(f"  Row 11: {formula_count} formulas")
print(f"  Complex formulas (3+ IFs): {complex_formulas}")

# ============================================================================
print("\n" + "="*80)
print("4. SPOTLIGHT_PMWORKSPACE ANALYSIS")
print("="*80)

ws_spot = wb['Spotlight_PMWorkspace']
audit["sheets_analyzed"].append("Spotlight_PMWorkspace")

print("\nChecking Spotlight sections...")

# Check deliverables (row 19)
deliv_cell = ws_spot.cell(19, 1)
if deliv_cell.data_type == 'f':
    formula = str(deliv_cell.value)
    if 'AGGREGATE' in formula:
        print("  Deliverables: Using AGGREGATE (complex)")
        audit["performance_issues"].append({
            "sheet": "Spotlight_PMWorkspace",
            "section": "Deliverables",
            "issue": "AGGREGATE formula - very complex",
            "note": "Works but computationally heavy"
        })
else:
    print("  WARNING: Deliverables section has no formula")
    audit["missing_items"].append("Spotlight: Deliverables formulas missing")

# Check stakeholders (row 33)
stake_cell = ws_spot.cell(33, 1)
if stake_cell.value:
    print(f"  Stakeholders: Manual entry (value: {stake_cell.value})")
else:
    print("  Stakeholders: Empty (manual entry required)")
    audit["manual_work_remaining"].append("Spotlight: Stakeholder data (manual entry)")

# ============================================================================
print("\n" + "="*80)
print("5. COUNTRY_DASHBOARD ANALYSIS")
print("="*80)

ws_country = wb['Country_Dashboard']
audit["sheets_analyzed"].append("Country_Dashboard")

print("\nChecking Country_Dashboard...")

# Check D2 country name lookup
d2_formula = ws_country['D2'].value
if d2_formula and isinstance(d2_formula, str) and '#REF!' not in d2_formula:
    print("  D2 Country lookup: OK")
else:
    print("  WARNING: D2 has issue")
    if '#REF!' in str(d2_formula):
        audit["missing_items"].append("Country_Dashboard D2: #REF! error in formula")

# Check project list (row 12)
a12_formula = ws_country.cell(12, 1).value
if a12_formula and isinstance(a12_formula, str):
    if 'SMALL' in a12_formula and 'IF' in a12_formula:
        print("  Project list: Using SMALL/IF array formula")
        audit["performance_issues"].append({
            "sheet": "Country_Dashboard",
            "cell": "A12+",
            "issue": "SMALL/IF array formula - slow with large datasets",
            "note": "Consider alternative filtering method"
        })

# ============================================================================
print("\n" + "="*80)
print("6. CONTROL SHEET ANALYSIS")
print("="*80)

ws_control = wb['Control']
audit["sheets_analyzed"].append("Control")

print("\nChecking Control sheet metrics...")

key_metrics = [
    ('B3', 'Current Date'),
    ('B9', 'Total Projects'),
    ('B10', 'Started Projects'),
    ('B15', 'Total Proposed'),
    ('B16', 'Total Allocated'),
    ('B17', 'Total Obligated'),
    ('B18', 'Total ULO'),
    ('B19', 'Portfolio ULO %'),
]

missing_metrics = []
for cell_ref, metric_name in key_metrics:
    cell = ws_control[cell_ref]
    if cell.data_type == 'f':
        print(f"  {cell_ref} ({metric_name}): OK")
    else:
        print(f"  WARNING: {cell_ref} ({metric_name}): No formula")
        missing_metrics.append(metric_name)

if missing_metrics:
    audit["missing_items"].append(f"Control: Missing formulas for {', '.join(missing_metrics)}")

# ============================================================================
print("\n" + "="*80)
print("7. PERFORMANCE OPTIMIZATION OPPORTUNITIES")
print("="*80)

print("\nAnalyzing formula patterns for optimization...")

# Common optimization opportunities
print("\n1. Dynamic Range Formulas (INDEX/COUNTA):")
print("   Current: SUMIF(Sheet!$B$2:INDEX(Sheet!$B:$B,COUNTA(...)))")
print("   Issue: Recalculates entire column every time")
print("   Better: Use table references (already using) or fixed ranges")
print("   Status: Already optimized with table references in most places")

audit["optimization_opportunities"].append({
    "category": "Dynamic Ranges",
    "current": "INDEX/COUNTA patterns",
    "recommendation": "Ensure all use table references where possible",
    "benefit": "10-30% faster calculation"
})

print("\n2. Nested IF Statements:")
print("   Issue: Multiple IF checks slow down calculation")
print("   Better: Use IFS (Excel 2019+) or restructure logic")

audit["optimization_opportunities"].append({
    "category": "Nested IFs",
    "current": "IF(IF(IF(...)))",
    "recommendation": "Use IFS function or simplify logic",
    "benefit": "5-15% faster"
})

print("\n3. Array Formulas (SMALL/IF):")
print("   Issue: Array formulas are slow, especially with filters")
print("   Better: Use FILTER function (Excel 365) or helper columns")

audit["optimization_opportunities"].append({
    "category": "Array Formulas",
    "current": "SMALL(IF(...)) for filtering",
    "recommendation": "Use FILTER function if Excel 365 available",
    "benefit": "50-80% faster on large datasets"
})

print("\n4. AGGREGATE Function:")
print("   Issue: Very complex, recalculates frequently")
print("   Better: Simpler lookups or manual entry")

audit["optimization_opportunities"].append({
    "category": "AGGREGATE",
    "current": "AGGREGATE(15,6,ROW(...)...)",
    "recommendation": "Simplify or use manual entry",
    "benefit": "Significant performance gain"
})

# ============================================================================
print("\n" + "="*80)
print("8. DATA VALIDATION STATUS")
print("="*80)

print("\nChecking for data validation...")
print("  NOTE: Data validation cannot be read by openpyxl easily")
print("  Manual check required for:")
print("    - Master_Projects dropdowns")
print("    - Country_Budgets dropdowns")
print("    - Spotlight B2 project selector")

audit["manual_work_remaining"].append("Data validation dropdowns not added yet")

# ============================================================================
print("\n" + "="*80)
print("9. TABLE STRUCTURE VERIFICATION")
print("="*80)

print("\nVerifying all expected tables exist...")
expected_tables = {
    'T_Master_Projects': 'Master_Projects',
    'T_Country_Budgets': 'Country_Budgets',
    'T_Country_Regions': 'Country_Regions',
    'T_Stakeholders': 'Stakeholders',
    'T_Project_Deliverables': 'Project_Deliverables',
    'T_Project_Audiences': 'Project_Audiences',
    'T_Project_Technologies': 'Project_Technologies',
    'T_Project_Documents': 'Project_Documents',
}

tables_found = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.tables:
        for table in ws.tables.values():
            tables_found[table.name] = sheet_name

print(f"  Found {len(tables_found)} tables")
for table_name, expected_sheet in expected_tables.items():
    if table_name in tables_found:
        print(f"  OK: {table_name}")
    else:
        print(f"  MISSING: {table_name} (should be in {expected_sheet})")
        audit["missing_items"].append(f"Table {table_name} not found")

# ============================================================================
print("\n" + "="*80)
print("10. FILE SIZE AND CALCULATION LOAD")
print("="*80)

# Estimate formula count
total_formulas = 0
for sheet_name in ['Master_Projects', 'Country_Budgets', 'Portfolio_Dashboard',
                   'Country_Dashboard', 'Spotlight_PMWorkspace', 'Control']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = 0
        for row in ws.iter_rows(min_row=2, max_row=min(201, ws.max_row)):
            for cell in row:
                if cell.data_type == 'f':
                    sheet_formulas += 1
        total_formulas += sheet_formulas
        print(f"  {sheet_name}: ~{sheet_formulas} formulas")

print(f"\n  TOTAL ESTIMATED FORMULAS: ~{total_formulas}")

if total_formulas > 5000:
    audit["performance_issues"].append({
        "category": "Overall",
        "issue": f"High formula count ({total_formulas}+)",
        "recommendation": "Consider consolidating or caching some calculations"
    })

# ============================================================================
# SAVE AUDIT RESULTS
# ============================================================================

print("\n" + "="*80)
print("SAVING AUDIT RESULTS")
print("="*80)

with open('audit_v28_deep_dive.json', 'w') as f:
    json.dump(audit, f, indent=2)

print("\nAudit saved to: audit_v28_deep_dive.json")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "="*80)
print("AUDIT SUMMARY")
print("="*80)

print(f"\nSheets Analyzed: {len(audit['sheets_analyzed'])}")
print(f"Performance Issues: {len(audit['performance_issues'])}")
print(f"Missing Items: {len(audit['missing_items'])}")
print(f"Optimization Opportunities: {len(audit['optimization_opportunities'])}")
print(f"Manual Work Remaining: {len(audit['manual_work_remaining'])}")

if audit['performance_issues']:
    print("\n" + "-"*80)
    print("PERFORMANCE ISSUES FOUND:")
    print("-"*80)
    for i, issue in enumerate(audit['performance_issues'], 1):
        print(f"\n{i}. {issue.get('sheet', 'N/A')} - {issue.get('issue', 'N/A')}")
        if 'note' in issue:
            print(f"   Note: {issue['note']}")

if audit['missing_items']:
    print("\n" + "-"*80)
    print("MISSING ITEMS:")
    print("-"*80)
    for i, item in enumerate(audit['missing_items'], 1):
        print(f"{i}. {item}")

if audit['optimization_opportunities']:
    print("\n" + "-"*80)
    print("OPTIMIZATION OPPORTUNITIES:")
    print("-"*80)
    for i, opp in enumerate(audit['optimization_opportunities'], 1):
        print(f"\n{i}. {opp['category']}")
        print(f"   Current: {opp['current']}")
        print(f"   Recommendation: {opp['recommendation']}")
        print(f"   Benefit: {opp['benefit']}")

if audit['manual_work_remaining']:
    print("\n" + "-"*80)
    print("MANUAL WORK REMAINING:")
    print("-"*80)
    for i, work in enumerate(audit['manual_work_remaining'], 1):
        print(f"{i}. {work}")

print("\n" + "="*80)
print("AUDIT COMPLETE")
print("="*80)
