import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v7.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.1.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING TRACKER V8.1 - CLEAN BUILD')
print('='*80)
print()
print('Building from v7 with proper table handling...')
print()

# Load workbook
wb = openpyxl.load_workbook(input_file, data_only=False)

# ============================================================================
# STEP 1: Rebuild Country_Budgets with new columns
# ============================================================================
print('STEP 1: Rebuilding Country_Budgets...')
print()

if 'Country_Budgets' in wb.sheetnames:
    ws = wb['Country_Budgets']

    # Read existing data
    existing_data = []
    headers_old = []
    for col in range(1, 15):
        h = ws.cell(1, col).value
        if h:
            headers_old.append(h)

    for row in range(2, 100):
        if not ws.cell(row, 1).value:
            break
        row_data = {}
        for col, header in enumerate(headers_old, 1):
            row_data[header] = ws.cell(row, col).value
        existing_data.append(row_data)

    print(f'  Read {len(existing_data)} existing budget rows')

    # Delete the old table if it exists
    if hasattr(ws, 'tables'):
        tables_to_remove = list(ws.tables.keys())
        for table_name in tables_to_remove:
            del ws.tables[table_name]

    # Clear the sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()

    # Write new headers
    new_headers = [
        'Budget_ID',
        'Unique_ID',
        'My_Country',
        'Country_Code',
        'Country_Name',
        'Allocated_Amount',
        'Obligated_Amount',
        'Spent_Amount',
        'ULO',
        'ULO_Percent',
        'Spend_Health'
    ]

    for col, header in enumerate(new_headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True)

    print(f'  Created {len(new_headers)} columns')

    # Write data back with new columns
    for row_idx, old_row in enumerate(existing_data, 2):
        # Budget_ID (formula)
        ws.cell(row_idx, 1).value = f'=IF(AND(B{row_idx}<>"",D{row_idx}<>""),B{row_idx}&"-"&D{row_idx},"")'

        # Unique_ID
        ws.cell(row_idx, 2).value = old_row.get('Unique_ID')

        # My_Country (NEW - default TRUE for first row, FALSE for second, TRUE for rest)
        if row_idx == 2:
            ws.cell(row_idx, 3).value = True
        elif row_idx == 3:
            ws.cell(row_idx, 3).value = False
        else:
            ws.cell(row_idx, 3).value = True

        # Country_Code
        ws.cell(row_idx, 4).value = old_row.get('Country_Code')

        # Country_Name (formula)
        ws.cell(row_idx, 5).value = f'=IFERROR(INDEX(Country_Regions!B:B,MATCH(D{row_idx},Country_Regions!A:A,0)),"")'

        # Allocated_Amount
        ws.cell(row_idx, 6).value = old_row.get('Allocated_Amount')

        # Obligated_Amount
        ws.cell(row_idx, 7).value = old_row.get('Obligated_Amount')

        # Spent_Amount (NEW - default to 0)
        ws.cell(row_idx, 8).value = 0

        # ULO (formula: Obligated - Spent)
        ws.cell(row_idx, 9).value = f'=G{row_idx}-H{row_idx}'

        # ULO_Percent (formula: ULO / Obligated)
        ws.cell(row_idx, 10).value = f'=IF(G{row_idx}=0,0,I{row_idx}/G{row_idx})'

        # Spend_Health (simplified for now)
        ws.cell(row_idx, 11).value = f'=IF(J{row_idx}>0.8,"Good",IF(J{row_idx}>0.5,"Caution","Alert"))'

    print(f'  Wrote {len(existing_data)} budget rows with new structure')

    # Create new table
    last_row = len(existing_data) + 1
    tab = Table(displayName='Country_Budgets', ref=f'A1:K{last_row}')
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    print('  [OK] Created new Country_Budgets table')

print()

# ============================================================================
# STEP 2: Update Master_Projects
# ============================================================================
print('STEP 2: Updating Master_Projects...')
print()

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    # Find or add new columns
    headers = {}
    for col in range(1, 50):
        h = ws.cell(1, col).value
        if h:
            headers[h] = col

    next_col = max(headers.values()) + 1 if headers else 1

    # Add new columns if they don't exist
    new_cols = []

    if 'Project_Manager' not in headers:
        ws.cell(1, next_col).value = 'Project_Manager'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Project_Manager'] = next_col
        new_cols.append('Project_Manager')
        next_col += 1

    if 'Proposed_Amount' not in headers:
        ws.cell(1, next_col).value = 'Proposed_Amount'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Proposed_Amount'] = next_col
        new_cols.append('Proposed_Amount')
        next_col += 1

    if 'Total_Spent' not in headers:
        ws.cell(1, next_col).value = 'Total_Spent'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Total_Spent'] = next_col
        new_cols.append('Total_Spent')
        next_col += 1

    if 'Include_In_Calcs' not in headers:
        ws.cell(1, next_col).value = 'Include_In_Calcs'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Include_In_Calcs'] = next_col
        new_cols.append('Include_In_Calcs')
        next_col += 1

    if 'My_Countries_Count' not in headers:
        ws.cell(1, next_col).value = 'My_Countries_Count'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['My_Countries_Count'] = next_col
        new_cols.append('My_Countries_Count')
        next_col += 1

    print(f'  Added columns: {new_cols}')

    # Update formulas for existing projects
    for row in range(2, 10):
        unique_id = ws.cell(row, headers.get('Unique_ID', 2)).value
        if not unique_id:
            break

        # Demo Project Managers
        pm_col = headers.get('Project_Manager')
        if pm_col and not ws.cell(row, pm_col).value:
            ws.cell(row, pm_col).value = 'Smith' if row == 2 else 'Jones'

        # Total_Obligated - sum MY countries only
        ws.cell(row, headers.get('Total_Obligated', 15)).value = \
            f'=SUMIFS(Country_Budgets[Obligated_Amount],Country_Budgets[Unique_ID],B{row},Country_Budgets[My_Country],TRUE)'

        # Total_Spent - sum MY countries only
        ws.cell(row, headers['Total_Spent']).value = \
            f'=SUMIFS(Country_Budgets[Spent_Amount],Country_Budgets[Unique_ID],B{row},Country_Budgets[My_Country],TRUE)'

        # Total_ULO = Obligated - Spent
        obligated_col = openpyxl.utils.get_column_letter(headers.get('Total_Obligated', 15))
        spent_col = openpyxl.utils.get_column_letter(headers['Total_Spent'])
        ws.cell(row, headers.get('Total_ULO', 16)).value = f'={obligated_col}{row}-{spent_col}{row}'

        # ULO_Percent = ULO / Obligated
        ulo_col = openpyxl.utils.get_column_letter(headers.get('Total_ULO', 16))
        ws.cell(row, headers.get('ULO_Percent', 17)).value = \
            f'=IF({obligated_col}{row}=0,0,{ulo_col}{row}/{obligated_col}{row})'

        # Include_In_Calcs
        status_col = openpyxl.utils.get_column_letter(headers.get('Status', 5))
        ws.cell(row, headers['Include_In_Calcs']).value = \
            f'=NOT(OR({status_col}{row}="Proposed",{status_col}{row}="Archived"))'

        # My_Countries_Count
        ws.cell(row, headers['My_Countries_Count']).value = \
            f'=COUNTIFS(Country_Budgets[Unique_ID],B{row},Country_Budgets[My_Country],TRUE)'

    print('  [OK] Updated Master_Projects formulas')

print()

# ============================================================================
# STEP 3: Update Portfolio_Dashboard
# ============================================================================
print('STEP 3: Updating Portfolio_Dashboard...')
print()

if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']

    # Find next available column
    last_col = 1
    for col in range(1, 50):
        if ws.cell(1, col).value:
            last_col = col

    next_col = last_col + 1

    # Add new columns
    new_headers = ['Project_Manager', 'My_Countries_Count', 'Total_Countries_Count', 'FAR_Notes']

    for header in new_headers:
        ws.cell(1, next_col).value = header
        ws.cell(1, next_col).font = Font(bold=True)
        next_col += 1

    print(f'  Added {len(new_headers)} columns')
    print('  [OK] Updated Portfolio_Dashboard')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving v8.1...')
wb.save(output_file)
wb.close()

print('='*80)
print('CLEAN BUILD COMPLETE!')
print('='*80)
print()
print('v8.1 created without table corruption')
print()
print(f'Output: {output_file}')
print()
print(f'Created: {timestamp}')
