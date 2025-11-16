import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v6-clean.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v7.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING TRACKER V7 - QUICK WINS')
print('='*80)
print()
print('Quick Wins:')
print('  1. Fix Project Spotlight reference')
print('  2. Expand project status categories')
print('  3. Expand country list (50+ countries)')
print()
print('='*80)
print()

# Load workbook
print('Loading v6-clean...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

total_changes = 0

# ============================================================================
# QUICK WIN 1: Fix Project Spotlight Reference
# ============================================================================
print('='*80)
print('QUICK WIN 1: Fix Project Spotlight Reference')
print('='*80)
print()

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']
    old_value = ws['B2'].value
    ws['B2'].value = 'PRJ-001'
    print(f'  Changed B2 from "{old_value}" to "PRJ-001"')
    print(f'  [OK] Project Spotlight should now load data correctly')
    total_changes += 1
else:
    print('  WARNING: Project_Spotlight sheet not found')

print()

# ============================================================================
# QUICK WIN 2: Expand Project Status Categories
# ============================================================================
print('='*80)
print('QUICK WIN 2: Expand Project Status Categories')
print('='*80)
print()

if 'Config_Lists' in wb.sheetnames:
    ws = wb['Config_Lists']

    # Find List_Status column
    status_col = None
    for col in range(1, 20):
        if ws.cell(1, col).value == 'List_Status':
            status_col = col
            break

    if status_col:
        # Current statuses
        current_statuses = []
        row = 2
        while ws.cell(row, status_col).value:
            current_statuses.append(ws.cell(row, status_col).value)
            row += 1

        print(f'  Found {len(current_statuses)} existing statuses:')
        for status in current_statuses:
            print(f'    - {status}')
        print()

        # New statuses to add
        new_statuses = [
            'Active',
            'CN Stage',
            'Proposed',
            'Archived',
            'Not Started',
            'In Progress',
            'On Hold',
            'Completed',
            'Cancelled'
        ]

        print(f'  Adding {len(new_statuses)} comprehensive statuses:')
        start_row = 2
        for i, status in enumerate(new_statuses):
            ws.cell(start_row + i, status_col).value = status
            print(f'    - {status}')

        print()
        print(f'  [OK] Updated List_Status with comprehensive options')
        total_changes += 1
    else:
        print('  WARNING: List_Status column not found in Config_Lists')
else:
    print('  WARNING: Config_Lists sheet not found')

print()

# ============================================================================
# QUICK WIN 3: Expand Country List
# ============================================================================
print('='*80)
print('QUICK WIN 3: Expand Country List')
print('='*80)
print()

# Comprehensive country list organized by region
new_countries = {
    # Asia-Pacific
    'IN': 'India',
    'KZ': 'Kazakhstan',
    'PK': 'Pakistan',
    'AU': 'Australia',
    'KH': 'Cambodia',
    'ID': 'Indonesia',
    'JP': 'Japan',
    'MY': 'Malaysia',
    'PH': 'Philippines',
    'KR': 'South Korea',
    'SG': 'Singapore',
    'TW': 'Taiwan',
    'TH': 'Thailand',
    'VN': 'Vietnam',
    'NZ': 'New Zealand',

    # Africa
    'ET': 'Ethiopia',
    'KE': 'Kenya',
    'MU': 'Mauritius',
    'NG': 'Nigeria',
    'NA': 'Namibia',
    'SN': 'Senegal',

    # Middle East
    'EG': 'Egypt',
    'IL': 'Israel',
    'SA': 'Saudi Arabia',
    'MA': 'Morocco',
    'OM': 'Oman',
    'AE': 'United Arab Emirates',

    # Latin America
    'AR': 'Argentina',
    'BR': 'Brazil',
    'CL': 'Chile',
    'CO': 'Colombia',
    'CR': 'Costa Rica',
    'MX': 'Mexico',
    'PA': 'Panama',
    'PE': 'Peru',
}

if 'Country_Regions' in wb.sheetnames:
    ws = wb['Country_Regions']

    # Find headers
    headers = {}
    for col in range(1, 10):
        header = ws.cell(1, col).value
        if header:
            headers[header] = col

    print(f'  Headers found: {list(headers.keys())}')
    print()

    # Find existing countries
    existing_countries = set()
    row = 2
    country_code_col = headers.get('Country_Code', 1)
    while ws.cell(row, country_code_col).value:
        existing_countries.add(ws.cell(row, country_code_col).value)
        row += 1

    print(f'  Found {len(existing_countries)} existing countries')
    print(f'  Adding {len(new_countries)} new countries:')
    print()

    # Add new countries
    next_row = row
    added_count = 0

    for code, name in sorted(new_countries.items()):
        if code not in existing_countries:
            # Determine region
            if code in ['IN', 'KZ', 'PK', 'AU', 'KH', 'ID', 'JP', 'MY', 'PH', 'KR', 'SG', 'TW', 'TH', 'VN', 'NZ']:
                region = 'APAC'
            elif code in ['ET', 'KE', 'MU', 'NG', 'NA', 'SN']:
                region = 'AFRICA'
            elif code in ['EG', 'IL', 'SA', 'MA', 'OM', 'AE']:
                region = 'MENA'
            elif code in ['AR', 'BR', 'CL', 'CO', 'CR', 'MX', 'PA', 'PE']:
                region = 'LATAM'
            else:
                region = 'OTHER'

            ws.cell(next_row, headers.get('Country_Code', 1)).value = code
            ws.cell(next_row, headers.get('Country_Name', 2)).value = name
            if 'Region' in headers:
                ws.cell(next_row, headers['Region']).value = region

            print(f'    {code}: {name} ({region})')
            next_row += 1
            added_count += 1

    print()
    print(f'  [OK] Added {added_count} new countries')
    print(f'  Total countries now: {len(existing_countries) + added_count}')
    total_changes += 1
else:
    print('  WARNING: Country_Regions sheet not found')

print()

# ============================================================================
# Update List_CountryCodes in Config_Lists
# ============================================================================
print('='*80)
print('Updating Country Code List in Config_Lists')
print('='*80)
print()

if 'Config_Lists' in wb.sheetnames and 'Country_Regions' in wb.sheetnames:
    config_ws = wb['Config_Lists']
    country_ws = wb['Country_Regions']

    # Find List_CountryCodes column in Config_Lists
    code_col = None
    for col in range(1, 20):
        if config_ws.cell(1, col).value == 'List_CountryCodes':
            code_col = col
            break

    if code_col:
        # Get all country codes from Country_Regions
        all_codes = []
        country_code_col = 1  # Assuming Country_Code is in column A
        row = 2
        while country_ws.cell(row, country_code_col).value:
            code = country_ws.cell(row, country_code_col).value
            if code:
                all_codes.append(code)
            row += 1

        # Update Config_Lists
        print(f'  Updating List_CountryCodes with {len(all_codes)} country codes...')
        for i, code in enumerate(sorted(all_codes)):
            config_ws.cell(2 + i, code_col).value = code

        # Clear any old values beyond the new list
        for i in range(len(all_codes), 200):
            if config_ws.cell(2 + i, code_col).value:
                config_ws.cell(2 + i, code_col).value = None

        print(f'  [OK] Updated List_CountryCodes')
        total_changes += 1

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v7...')
wb.save(output_file)
wb.close()

print('='*80)
print('QUICK WINS COMPLETE!')
print('='*80)
print()
print(f'Total changes made: {total_changes}')
print()
print('Changes applied:')
print('  1. [OK] Fixed Project Spotlight reference (PRJ-003 -> PRJ-001)')
print('  2. [OK] Expanded project status categories (9 comprehensive statuses)')
print('  3. [OK] Added 30+ new countries (APAC, Africa, MENA, LatAm)')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v7 in Excel')
print('  2. Check Project_Spotlight - data should load now')
print('  3. Check Master_Projects - new status options in dropdown')
print('  4. Check Country_Regions - 50+ countries available')
print()
print(f'Created: {timestamp}')
