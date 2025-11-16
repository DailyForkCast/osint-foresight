"""
Update Project Management Tracker Workbook
Implements new financial terminology and adds new tracking sheets
"""

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def update_country_budgets(ws):
    """Phase 1: Update Country_Budgets sheet structure"""
    print("  Updating Country_Budgets sheet...")

    # Find the last row with data before modifications
    last_row = ws.max_row

    # Find and update the table range to include new columns
    for table in ws.tables.values():
        # Update table reference to expand to column K
        old_ref = table.ref
        # Get the starting cell (should be A1)
        start_cell = old_ref.split(':')[0]
        # Set new range to include all columns through K
        table.ref = f"{start_cell}:K{last_row}"
        print(f"    Updated table '{table.name}' range from {old_ref} to {table.ref}")

    # Insert new column after F (index 7 becomes the new G)
    ws.insert_cols(7)

    # Update headers
    ws['E1'] = 'Allocated_Amount'
    ws['F1'] = 'Obligated_Amount'
    ws['G1'] = 'Spent_Amount'
    ws['H1'] = 'Available'
    ws['I1'] = 'Spending_Rate'
    ws['J1'] = 'Funding_Gap'
    ws['K1'] = 'Spend_Health'

    # Add formulas starting from row 2
    for row in range(2, last_row + 1):
        # H: Available = Obligated - Spent
        ws[f'H{row}'] = f'=F{row}-G{row}'

        # I: Spending_Rate = Spent / Obligated
        ws[f'I{row}'] = f'=IF(F{row}=0,0,G{row}/F{row})'

        # J: Funding_Gap = Allocated - Obligated
        ws[f'J{row}'] = f'=E{row}-F{row}'

        # K: Spend_Health
        ws[f'K{row}'] = f'=IF(J{row}>0,"⚠️ Awaiting",IF(I{row}<0.25,"🔴 Low",IF(I{row}<0.5,"🟡 Moderate",IF(I{row}<0.75,"🟢 Good","✅ On Track"))))'

    print(f"    Added formulas to {last_row - 1} rows")


def update_master_projects(ws):
    """Phase 2: Update Master_Projects sheet structure"""
    print("  Updating Master_Projects sheet...")

    # Find the last row with data before modifications
    last_row = ws.max_row

    # Find and update the table range to include new columns
    for table in ws.tables.values():
        # Update table reference to expand to column U (or whatever the last column will be)
        old_ref = table.ref
        start_cell = old_ref.split(':')[0]
        # We're adding 2 columns, so expand the range accordingly
        # Assuming the original table ended around column S, we'll expand to U
        table.ref = f"{start_cell}:U{last_row}"
        print(f"    Updated table '{table.name}' range from {old_ref} to {table.ref}")

    # Insert new column after O (index 16 becomes the new P)
    ws.insert_cols(16)
    ws.insert_cols(19)  # For Funding_Gap

    # Update headers
    ws['N1'] = 'Total_Allocated'
    ws['O1'] = 'Total_Obligated'
    ws['P1'] = 'Total_Spent'
    ws['Q1'] = 'Total_Available'
    ws['R1'] = 'Spending_Rate'
    ws['S1'] = 'Funding_Gap'
    ws['T1'] = 'Country_Count'
    ws['U1'] = 'Fund_Status'

    # Add formulas starting from row 2
    for row in range(2, last_row + 1):
        # N: Total_Allocated
        ws[f'N{row}'] = f'=SUMIF(Country_Budgets!B:B,B{row},Country_Budgets!E:E)'

        # O: Total_Obligated
        ws[f'O{row}'] = f'=SUMIF(Country_Budgets!B:B,B{row},Country_Budgets!F:F)'

        # P: Total_Spent
        ws[f'P{row}'] = f'=SUMIF(Country_Budgets!B:B,B{row},Country_Budgets!G:G)'

        # Q: Total_Available
        ws[f'Q{row}'] = f'=O{row}-P{row}'

        # R: Spending_Rate
        ws[f'R{row}'] = f'=IF(O{row}=0,0,P{row}/O{row})'

        # S: Funding_Gap
        ws[f'S{row}'] = f'=N{row}-O{row}'

        # T: Country_Count
        ws[f'T{row}'] = f'=COUNTIF(Country_Budgets!B:B,B{row})'

        # U: Fund_Status
        ws[f'U{row}'] = f'=IF(S{row}>0,"⚠️ Gap",IF(R{row}<0.25,"🔴 Low Spend",IF(R{row}<0.5,"🟡 Slow",IF(R{row}<0.75,"🟢 Active","✅ Complete"))))'

    print(f"    Added formulas to {last_row - 1} rows")


def create_project_audiences(wb):
    """Phase 3: Create Project_Audiences sheet"""
    print("  Creating Project_Audiences sheet...")

    # Remove existing sheet if it exists
    if "Project_Audiences" in wb.sheetnames:
        print("    Removing existing Project_Audiences sheet...")
        wb.remove(wb["Project_Audiences"])

    # Create new sheet
    ws = wb.create_sheet("Project_Audiences")

    # Add headers
    headers = ['Project_ID', 'Audience_Type', 'Audience_Region', 'Description', 'Priority']
    for idx, header in enumerate(headers, 1):
        ws.cell(1, idx, header)

    # Add sample data
    sample_data = [
        ['PRJ-001', 'Government', 'North America', 'Federal agencies and departments', 'Primary'],
        ['PRJ-001', 'Military', 'Global', 'Defense contractors and military bases', 'Secondary'],
        ['PRJ-002', 'Commercial', 'Europe', 'Enterprise businesses', 'Primary']
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    # Add data validation
    # Audience_Type validation (column B)
    dv_audience = DataValidation(type="list",
                                  formula1='"Government,Military,Commercial,Academic,Public,Internal,Partners,Contractors"',
                                  allow_blank=True)
    ws.add_data_validation(dv_audience)
    dv_audience.add(f'B2:B1000')

    # Audience_Region validation (column C)
    dv_region = DataValidation(type="list",
                               formula1='"Global,North America,Europe,Asia Pacific,Middle East,Africa,Latin America"',
                               allow_blank=True)
    ws.add_data_validation(dv_region)
    dv_region.add(f'C2:C1000')

    # Priority validation (column E)
    dv_priority = DataValidation(type="list",
                                 formula1='"Primary,Secondary,Tertiary"',
                                 allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add(f'E2:E1000')

    # Create table
    tab = Table(displayName="T_Project_Audiences", ref=f"A1:E{len(sample_data) + 1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    print(f"    Added {len(sample_data)} sample rows")


def create_project_products(wb):
    """Phase 4: Create Project_Products sheet"""
    print("  Creating Project_Products sheet...")

    # Remove existing sheet if it exists
    if "Project_Products" in wb.sheetnames:
        print("    Removing existing Project_Products sheet...")
        wb.remove(wb["Project_Products"])

    # Create new sheet
    ws = wb.create_sheet("Project_Products")

    # Add headers
    headers = ['Project_ID', 'Product_Name', 'Product_Category', 'Product_Status', 'Notes']
    for idx, header in enumerate(headers, 1):
        ws.cell(1, idx, header)

    # Add sample data
    sample_data = [
        ['PRJ-001', 'Dashboard System', 'Software', 'Deployed', 'Web-based monitoring solution'],
        ['PRJ-001', 'User Training', 'Training', 'Planned', '2-day workshop for end users'],
        ['PRJ-002', 'API Gateway', 'Software', 'In Development', 'RESTful API service']
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    # Add data validation
    # Product_Category validation (column C)
    dv_category = DataValidation(type="list",
                                 formula1='"Software,Hardware,Service,Training,Documentation,Research,Prototype"',
                                 allow_blank=True)
    ws.add_data_validation(dv_category)
    dv_category.add(f'C2:C1000')

    # Product_Status validation (column D)
    dv_status = DataValidation(type="list",
                               formula1='"Planned,In Development,Testing,Deployed,Retired"',
                               allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'D2:D1000')

    # Create table
    tab = Table(displayName="T_Project_Products", ref=f"A1:E{len(sample_data) + 1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    print(f"    Added {len(sample_data)} sample rows")


def create_project_deliverables(wb):
    """Phase 5: Create Project_Deliverables sheet"""
    print("  Creating Project_Deliverables sheet...")

    # Remove existing sheet if it exists
    if "Project_Deliverables" in wb.sheetnames:
        print("    Removing existing Project_Deliverables sheet...")
        wb.remove(wb["Project_Deliverables"])

    # Create new sheet
    ws = wb.create_sheet("Project_Deliverables")

    # Add headers
    headers = ['Project_ID', 'Deliverable_Name', 'Deliverable_Type', 'Due_Date', 'Status', 'Owner', 'Completion_Percent']
    for idx, header in enumerate(headers, 1):
        ws.cell(1, idx, header)

    # Add sample data
    sample_data = [
        ['PRJ-001', 'Technical Requirements Doc', 'Document', datetime(2025, 1, 15), 'Completed', 'Smith, John', 100],
        ['PRJ-001', 'Beta Release', 'Software Release', datetime(2025, 2, 28), 'In Progress', 'Davis, Maria', 75],
        ['PRJ-002', 'Kickoff Presentation', 'Presentation', datetime(2025, 1, 10), 'Completed', 'Johnson, Bob', 100]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    # Add data validation
    # Deliverable_Type validation (column C)
    dv_type = DataValidation(type="list",
                            formula1='"Document,Software Release,Presentation,Report,Training,Milestone,Hardware"',
                            allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f'C2:C1000')

    # Status validation (column E)
    dv_status = DataValidation(type="list",
                               formula1='"Not Started,In Progress,Under Review,Completed,Delayed,Cancelled"',
                               allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'E2:E1000')

    # Completion_Percent validation (column G)
    dv_percent = DataValidation(type="whole", operator="between", formula1="0", formula2="100", allow_blank=True)
    ws.add_data_validation(dv_percent)
    dv_percent.add(f'G2:G1000')

    # Create table
    tab = Table(displayName="T_Project_Deliverables", ref=f"A1:G{len(sample_data) + 1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    print(f"    Added {len(sample_data)} sample rows")


def update_project_spotlight(ws):
    """Phase 6: Update Project_Spotlight sheet"""
    print("  Updating Project_Spotlight sheet...")

    # Unmerge any existing merged cells that might interfere
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Update Budget Summary (Row 6)
    ws['A6'] = 'Budget Summary:'
    ws['B6'] = '=CONCATENATE("Allocated: $",TEXT(SUMIF(Country_Budgets[Unique_ID],B2,Country_Budgets[Allocated_Amount]),"#,##0"))'
    ws['D6'] = '=CONCATENATE("Obligated: $",TEXT(SUMIF(Country_Budgets[Unique_ID],B2,Country_Budgets[Obligated_Amount]),"#,##0"))'
    ws['F6'] = '=CONCATENATE("Spent: $",TEXT(SUMIF(Country_Budgets[Unique_ID],B2,Country_Budgets[Spent_Amount]),"#,##0"))'
    ws['H6'] = '=CONCATENATE("Available: $",TEXT(SUMIF(Country_Budgets[Unique_ID],B2,Country_Budgets[Available]),"#,##0"))'

    # Country Performance Table Headers (Row 7)
    ws['E7'] = 'Country'
    ws['F7'] = 'Allocated'
    ws['G7'] = 'Obligated'
    ws['H7'] = 'Spent'
    ws['I7'] = 'Available'
    ws['J7'] = 'Spend %'
    ws['K7'] = 'Gap'
    ws['L7'] = 'Status'

    # Country Performance Formulas (Rows 8-15)
    for row in range(8, 16):
        row_offset = row - 7
        ws[f'E{row}'] = f'=IFERROR(INDEX(Country_Budgets[Country_Name],SMALL(IF(Country_Budgets[Unique_ID]=$B$2,ROW(Country_Budgets[Country_Name])-ROW(Country_Budgets[#Headers])),{row_offset})),"")'
        ws[f'F{row}'] = f'=IF(E{row}="","",SUMIFS(Country_Budgets[Allocated_Amount],Country_Budgets[Unique_ID],$B$2,Country_Budgets[Country_Name],E{row}))'
        ws[f'G{row}'] = f'=IF(E{row}="","",SUMIFS(Country_Budgets[Obligated_Amount],Country_Budgets[Unique_ID],$B$2,Country_Budgets[Country_Name],E{row}))'
        ws[f'H{row}'] = f'=IF(E{row}="","",SUMIFS(Country_Budgets[Spent_Amount],Country_Budgets[Unique_ID],$B$2,Country_Budgets[Country_Name],E{row}))'
        ws[f'I{row}'] = f'=IF(E{row}="","",G{row}-H{row})'
        ws[f'J{row}'] = f'=IF(G{row}="","",IF(G{row}=0,0,H{row}/G{row}))'
        ws[f'K{row}'] = f'=IF(E{row}="","",F{row}-G{row})'
        ws[f'L{row}'] = f'=IF(E{row}="","",IF(K{row}>0,"⚠️ Gap",IF(J{row}=0,"🔴 None",IF(J{row}<0.5,"🟡 Low","🟢 Good"))))'

    # Deliverables Section Headers (Row 17-18)
    ws.merge_cells('A17:D17')
    ws['A17'] = 'KEY DELIVERABLES'
    ws.merge_cells('E17:H17')
    ws['E17'] = 'TARGET AUDIENCES'
    ws.merge_cells('I17:L17')
    ws['I17'] = 'TARGET PRODUCTS'

    ws['A18'] = 'Deliverable'
    ws['B18'] = 'Due Date'
    ws['C18'] = 'Status'
    ws['D18'] = 'Complete %'
    ws['E18'] = 'Audience'
    ws['F18'] = 'Region'
    ws['G18'] = 'Priority'
    ws['I18'] = 'Product'
    ws['J18'] = 'Category'
    ws['K18'] = 'Status'

    # Deliverables, Audiences, and Products Formulas (Rows 19-24)
    for row in range(19, 25):
        row_offset = row - 18

        # Deliverables
        ws[f'A{row}'] = f'=IFERROR(INDEX(Project_Deliverables[Deliverable_Name],SMALL(IF(Project_Deliverables[Project_ID]=$B$2,ROW(Project_Deliverables[Deliverable_Name])-ROW(Project_Deliverables[#Headers])),{row_offset})),"")'
        ws[f'B{row}'] = f'=IF(A{row}="","",TEXT(INDEX(Project_Deliverables[Due_Date],MATCH(1,(Project_Deliverables[Project_ID]=$B$2)*(Project_Deliverables[Deliverable_Name]=A{row}),0)),"mm/dd/yy"))'
        ws[f'C{row}'] = f'=IF(A{row}="","",INDEX(Project_Deliverables[Status],MATCH(1,(Project_Deliverables[Project_ID]=$B$2)*(Project_Deliverables[Deliverable_Name]=A{row}),0)))'
        ws[f'D{row}'] = f'=IF(A{row}="","",INDEX(Project_Deliverables[Completion_Percent],MATCH(1,(Project_Deliverables[Project_ID]=$B$2)*(Project_Deliverables[Deliverable_Name]=A{row}),0))&"%")'

        # Audiences
        ws[f'E{row}'] = f'=IFERROR(INDEX(Project_Audiences[Audience_Type],SMALL(IF(Project_Audiences[Project_ID]=$B$2,ROW(Project_Audiences[Audience_Type])-ROW(Project_Audiences[#Headers])),{row_offset})),"")'
        ws[f'F{row}'] = f'=IF(E{row}="","",INDEX(Project_Audiences[Audience_Region],MATCH(1,(Project_Audiences[Project_ID]=$B$2)*(Project_Audiences[Audience_Type]=E{row}),0)))'
        ws[f'G{row}'] = f'=IF(E{row}="","",INDEX(Project_Audiences[Priority],MATCH(1,(Project_Audiences[Project_ID]=$B$2)*(Project_Audiences[Audience_Type]=E{row}),0)))'

        # Products
        ws[f'I{row}'] = f'=IFERROR(INDEX(Project_Products[Product_Name],SMALL(IF(Project_Products[Project_ID]=$B$2,ROW(Project_Products[Product_Name])-ROW(Project_Products[#Headers])),{row_offset})),"")'
        ws[f'J{row}'] = f'=IF(I{row}="","",INDEX(Project_Products[Product_Category],MATCH(1,(Project_Products[Project_ID]=$B$2)*(Project_Products[Product_Name]=I{row}),0)))'
        ws[f'K{row}'] = f'=IF(I{row}="","",INDEX(Project_Products[Product_Status],MATCH(1,(Project_Products[Project_ID]=$B$2)*(Project_Products[Product_Name]=I{row}),0)))'

    print("    Updated Project_Spotlight with new sections")


def apply_conditional_formatting(wb):
    """Phase 8: Apply conditional formatting"""
    print("  Applying conditional formatting...")

    # Country_Budgets sheet
    if 'Country_Budgets' in wb.sheetnames:
        ws = wb['Country_Budgets']
        last_row = ws.max_row

        # Spending_Rate color scale (column I)
        ws.conditional_formatting.add(f'I2:I{last_row}',
            ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                          mid_type='num', mid_value=0.5, mid_color='FFEB84',
                          end_type='num', end_value=1, end_color='63BE7B'))

        # Funding_Gap highlight red if >0 (column J)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws.conditional_formatting.add(f'J2:J{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill))

    # Master_Projects sheet
    if 'Master_Projects' in wb.sheetnames:
        ws = wb['Master_Projects']
        last_row = ws.max_row

        # Spending_Rate data bars (column R)
        ws.conditional_formatting.add(f'R2:R{last_row}',
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                       color="638EC6", showValue=True, minLength=None, maxLength=None))

        # Funding_Gap highlight red if >0 (column S)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws.conditional_formatting.add(f'S2:S{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill))

    # Project_Spotlight sheet
    if 'Project_Spotlight' in wb.sheetnames:
        ws = wb['Project_Spotlight']

        # Spend % data bars (column J, rows 8-15)
        ws.conditional_formatting.add('J8:J15',
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                       color="638EC6", showValue=True, minLength=None, maxLength=None))

        # Gap highlight yellow if >0 (column K, rows 8-15)
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        ws.conditional_formatting.add('K8:K15',
            CellIsRule(operator='greaterThan', formula=['0'], fill=yellow_fill))

    print("    Applied color scales, data bars, and cell highlighting")


def main():
    print("=" * 60)
    print("Project Management Tracker Update")
    print("=" * 60)

    input_file = "C:/Projects/OSINT - Foresight/2025-10-05-Tracker_v2.xlsx"
    output_file = "C:/Projects/OSINT - Foresight/2025-10-07 Tracker FIXED.xlsx"

    print(f"\nLoading workbook: {input_file}")
    wb = load_workbook(input_file)

    try:
        # Phase 1: Update Country_Budgets
        if 'Country_Budgets' in wb.sheetnames:
            update_country_budgets(wb['Country_Budgets'])
        else:
            print("  WARNING: Country_Budgets sheet not found")

        # Phase 2: Update Master_Projects
        if 'Master_Projects' in wb.sheetnames:
            update_master_projects(wb['Master_Projects'])
        else:
            print("  WARNING: Master_Projects sheet not found")

        # Phase 3: Create Project_Audiences
        create_project_audiences(wb)

        # Phase 4: Create Project_Products
        create_project_products(wb)

        # Phase 5: Create Project_Deliverables
        create_project_deliverables(wb)

        # Phase 6: Update Project_Spotlight
        if 'Project_Spotlight' in wb.sheetnames:
            update_project_spotlight(wb['Project_Spotlight'])
        else:
            print("  WARNING: Project_Spotlight sheet not found")

        # Phase 8: Apply conditional formatting
        apply_conditional_formatting(wb)

        # Save with new name
        print(f"\nSaving workbook: {output_file}")
        wb.save(output_file)

        print("\n" + "=" * 60)
        print("Update completed successfully!")
        print("=" * 60)
        print("\nChanges made:")
        print("  - Updated Country_Budgets with new financial terminology")
        print("  - Updated Master_Projects with spending and gap tracking")
        print("  - Created Project_Audiences sheet")
        print("  - Created Project_Products sheet")
        print("  - Created Project_Deliverables sheet")
        print("  - Updated Project_Spotlight with enhanced reporting")
        print("  - Applied conditional formatting to all sheets")
        print(f"\nNew file saved as: {output_file}")

    except Exception as e:
        print(f"\nError during update: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
