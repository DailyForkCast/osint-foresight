"""
Compare v12 and v18 to ensure all features are present
"""
import openpyxl

print("="*80)
print("COMPARING V12 vs V18 - FEATURE CHECK")
print("="*80)

wb_v12 = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=False)
wb_v18 = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=False)

print("\n1. SHEET COMPARISON")
print("="*80)

v12_sheets = set(wb_v12.sheetnames)
v18_sheets = set(wb_v18.sheetnames)

print("\nSheets in V12 but NOT in V18:")
missing = v12_sheets - v18_sheets
if missing:
    for sheet in sorted(missing):
        print(f"  X {sheet}")
else:
    print("  OK - None - all sheets accounted for")

print("\nSheets in V18 but NOT in V12 (new additions):")
new = v18_sheets - v12_sheets
if new:
    for sheet in sorted(new):
        print(f"  + {sheet}")
else:
    print("  None")

# Check Master_Projects formulas in v12
print("\n2. MASTER_PROJECTS FORMULAS (V12)")
print("="*80)

ws = wb_v12['Master_Projects']
print(f"\nRow 2 formulas in v12:")
for col in range(1, 21):
    cell = ws.cell(2, col)
    if cell.data_type == 'f':
        col_letter = openpyxl.utils.get_column_letter(col)
        header = ws.cell(1, col).value
        print(f"  {col_letter} ({header}): {cell.value}")

# Check Country_Budgets formulas in v12
print("\n3. COUNTRY_BUDGETS FORMULAS (V12)")
print("="*80)

ws = wb_v12['Country_Budgets']
print(f"\nRow 2 formulas in v12:")
for col in range(1, 15):
    cell = ws.cell(2, col)
    if cell.data_type == 'f':
        col_letter = openpyxl.utils.get_column_letter(col)
        header = ws.cell(1, col).value
        print(f"  {col_letter} ({header}): {cell.value}")

# Check for data validation in v12
print("\n4. DATA VALIDATION (V12)")
print("="*80)

for sheet_name in wb_v12.sheetnames:
    ws = wb_v12[sheet_name]
    if ws.data_validations.dataValidation:
        print(f"\n{sheet_name}:")
        for dv in ws.data_validations.dataValidation:
            print(f"  Range: {dv.sqref}")
            print(f"  Type: {dv.type}")
            if dv.formula1:
                print(f"  Source: {dv.formula1}")

# Check Config_Lists content
print("\n5. CONFIG_LISTS CONTENT (V12)")
print("="*80)

ws = wb_v12['Config_Lists']
print("\nList types in v12:")
list_types = set()
for row in range(2, ws.max_row + 1):
    list_type = ws.cell(row, 1).value
    if list_type:
        list_types.add(list_type)

for lt in sorted(list_types):
    print(f"  - {lt}")

# Check for named ranges
print("\n6. NAMED RANGES (V12)")
print("="*80)

if wb_v12.defined_names:
    print("\nNamed ranges in v12:")
    for name in wb_v12.defined_names.definedName:
        print(f"  - {name.name}: {name.value}")
else:
    print("\n  None defined")

# Check for conditional formatting
print("\n7. CONDITIONAL FORMATTING")
print("="*80)

sheets_with_cf = []
for sheet_name in wb_v12.sheetnames:
    ws = wb_v12[sheet_name]
    if ws.conditional_formatting._cf_rules:
        sheets_with_cf.append(sheet_name)

if sheets_with_cf:
    print("\nSheets with conditional formatting in v12:")
    for sheet in sheets_with_cf:
        print(f"  - {sheet}")
else:
    print("\n  None found")

# Check Master_Projects column structure
print("\n8. MASTER_PROJECTS STRUCTURE COMPARISON")
print("="*80)

ws_v12 = wb_v12['Master_Projects']
ws_v18 = wb_v18['Master_Projects']

print("\nV12 Headers:")
for col in range(1, min(21, ws_v12.max_column + 1)):
    header = ws_v12.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\nV18 Headers:")
for col in range(1, min(21, ws_v18.max_column + 1)):
    header = ws_v18.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

# Check Country_Budgets column structure
print("\n9. COUNTRY_BUDGETS STRUCTURE COMPARISON")
print("="*80)

ws_v12 = wb_v12['Country_Budgets']
ws_v18 = wb_v18['Country_Budgets']

print("\nV12 Headers:")
for col in range(1, min(15, ws_v12.max_column + 1)):
    header = ws_v12.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\nV18 Headers:")
for col in range(1, min(15, ws_v18.max_column + 1)):
    header = ws_v18.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}: {header}")

print("\n" + "="*80)
print("COMPARISON COMPLETE")
print("="*80)
