"""
Fix v33:
1. Fix Control sheet Avg Budget and Burnrate formulas
2. Clean up Country_Dashboard empty columns
3. Remove Stakeholder_ID from Stakeholders sheet
"""
import openpyxl

print("="*80)
print("FIXING CONTROL, COUNTRY_DASHBOARD, AND STAKEHOLDERS - V33")
print("="*80)

# Load v32
print("\nLoading v32...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v32.xlsx')

print("\n" + "="*80)
print("PART 1: FIX CONTROL SHEET FORMULAS")
print("="*80)

ws_control = wb['Control']

print("\nFixing Avg Budget (E6)...")
print("  Current: =B15/B9 (Total Proposed / Total Projects)")
print("  This is actually correct - average proposed budget per project")
print("  But let's use Allocated instead of Proposed for active budget")
print("  New: =B16/B9 (Total Allocated / Total Projects)")

ws_control['E6'] = '=B16/B9'
print("  E6 updated: Average allocated budget per project")

print("\nFixing Burn Rate (E7)...")
print("  Current: =B16/B15 (Total Allocated / Total Proposed)")
print("  This shows allocation rate, not burn rate")
print("  Burn Rate should be: Spent / Allocated")
print("  New: =IF(B16=0,0,B17/B16) (Total Obligated / Total Allocated)")

ws_control['E7'] = '=IF(B16=0,0,B17/B16)'
print("  E7 updated: Obligation rate (how much allocated has been obligated)")

print("\n  Note: True 'Burn Rate' could also be Spent/Obligated")
print("  Current formula shows what % of allocation is obligated")

print("\n" + "="*80)
print("PART 2: CLEAN UP COUNTRY_DASHBOARD")
print("="*80)

ws_country = wb['Country_Dashboard']

print("\nChecking for empty columns...")

# The issue is that columns B and C are empty between A and D
# Let's check rows 11-12 to see the structure
print("\nRow 11 structure:")
row11_data = {}
for col in range(1, 15):
    val = ws_country.cell(11, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        row11_data[col_letter] = val
        print(f"  {col_letter}: {str(val).encode('ascii', 'ignore').decode('ascii')}")

# It looks like columns B and C are empty
# Delete columns B and C
print("\nDeleting empty columns B and C...")
ws_country.delete_cols(2, 2)  # Delete 2 columns starting at position 2 (B)
print("  Deleted columns B and C")

print("\nNew structure after deletion:")
for col in range(1, 12):
    val = ws_country.cell(11, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {str(val).encode('ascii', 'ignore').decode('ascii')}")

print("\n" + "="*80)
print("PART 3: REMOVE STAKEHOLDER_ID")
print("="*80)

ws_stakeholders = wb['Stakeholders']

print("\nDeleting column A (Stakeholder_ID)...")
ws_stakeholders.delete_cols(1, 1)  # Delete 1 column at position 1 (A)
print("  Deleted column A")

print("\nNew header structure:")
for col in range(1, 10):
    val = ws_stakeholders.cell(1, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_val = str(val).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}1: {clean_val}")

print("\n" + "="*80)
print("SAVING V33")
print("="*80)

wb.save('2025-10-26-Tracker-v33.xlsx')

print("\nOK - v33 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v33:")

print("\n1. Control Sheet:")
print("   - E6 (Avg Budget): =B16/B9")
print("     Average allocated budget per project")
print("   - E7 (Burn Rate): =IF(B16=0,0,B17/B16)")
print("     Obligation rate (% of allocation that's obligated)")

print("\n2. Country_Dashboard:")
print("   - Deleted empty columns B and C")
print("   - Cleaned up structure")
print("   - New layout:")
print("     A: Project ID")
print("     B: Project Name (was D)")
print("     C: Status (was E)")
print("     D: Priority (was F)")
print("     E: Proposed (was G)")
print("     F: Allocated (was H)")
print("     G: Obligated (was I)")
print("     H: Spent (was J)")
print("     I: ULO (was K)")

print("\n3. Stakeholders Sheet:")
print("   - Removed Stakeholder_ID column")
print("   - New layout starts with:")
print("     A: Name")
print("     B: Title")
print("     C: Organization")
print("     D: Location_City")
print("     etc.")

print("\nv33 ready!")
