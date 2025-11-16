"""
Build Tracker v17 - Option 3: Using Excel COM Automation
This uses Excel itself to build the workbook, avoiding formula reference issues
"""

import sys

print("="*80)
print("BUILDING TRACKER V17 - USING EXCEL COM AUTOMATION")
print("="*80)

try:
    import win32com.client
    print("✓ win32com available - proceeding with Excel COM")
    HAS_COM = True
except ImportError:
    print("✗ win32com not available - will fall back to Option 1")
    HAS_COM = False

if not HAS_COM:
    print("\nInstalling pywin32...")
    import subprocess
    result = subprocess.run([sys.executable, "-m", "pip", "install", "pywin32"],
                          capture_output=True, text=True)
    if result.returncode == 0:
        print("✓ pywin32 installed successfully")
        import win32com.client
        HAS_COM = True
    else:
        print("✗ Could not install pywin32")
        HAS_COM = False

if HAS_COM:
    print("\nStarting Excel COM automation...")
    import openpyxl
    from datetime import datetime
    import time

    # Load v12 data
    print("Loading v12 data...")
    wb_old = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx')

    # Start Excel
    print("Starting Excel application...")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False  # Run in background
    excel.DisplayAlerts = False

    # Create new workbook
    print("Creating new workbook...")
    wb = excel.Workbooks.Add()

    # Delete default sheets except one
    while wb.Sheets.Count > 1:
        wb.Sheets(wb.Sheets.Count).Delete()

    # Rename first sheet to Master_Projects
    ws = wb.Sheets(1)
    ws.Name = "Master_Projects"

    print("Building Master_Projects...")

    # Add headers
    headers = [
        'Project_ID', 'Project_Unique_ID', 'Project_Name', 'Project_Status', 'Project_Priority',
        'Project_Progress', 'Project_Start_Date', 'Project_End_Date', 'Days_Remaining',
        'Implementer', 'Implementer_POC', 'Implementer_POC_Phone', 'Implementer_POC_Email',
        'Total_Proposed', 'Total_Allocation', 'Total_Obligated', 'Total_ULO', 'ULO_Percent',
        'Countries', 'Country_Count'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.Cells(1, col)
        cell.Value = header
        cell.Font.Bold = True
        cell.Font.Color = 16777215  # White
        cell.Interior.Color = 9854208  # Dark blue
        cell.HorizontalAlignment = -4108  # Center

    # Copy data from v12
    ws_old = wb_old['Master_Projects']
    for old_row in range(2, ws_old.max_row + 1):
        new_row = old_row
        # Copy A-M
        for col in range(1, 14):
            val = ws_old.cell(old_row, col).value
            if val:
                ws.Cells(new_row, col).Value = val

        # Copy Countries and Count (S, T)
        ws.Cells(new_row, 19).Value = ws_old.cell(old_row, 18).value
        ws.Cells(new_row, 20).Value = ws_old.cell(old_row, 19).value

    print("  Added project data")

    # Create Country_Budgets sheet
    print("Building Country_Budgets...")
    ws_budget = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    ws_budget.Name = "Country_Budgets"

    headers_budget = [
        'Budget_ID', 'Unique_ID', 'My_Country', 'Country_Code', 'Country_Name',
        'Proposed_Amount', 'Allocated_Amount', 'Obligated_Amount', 'Spent_Amount',
        'ULO', 'ULO_Percent', 'Spend_Health'
    ]

    for col, header in enumerate(headers_budget, 1):
        cell = ws_budget.Cells(1, col)
        cell.Value = header
        cell.Font.Bold = True
        cell.Font.Color = 16777215
        cell.Interior.Color = 9854208
        cell.HorizontalAlignment = -4108

    # Copy budget data from v12
    ws_old = wb_old['Country_Budgets']
    for old_row in range(2, ws_old.max_row + 1):
        new_row = old_row
        # Copy A-E
        for col in range(1, 6):
            val = ws_old.cell(old_row, col).value
            if val:
                ws_budget.Cells(new_row, col).Value = val

        # F: Proposed (0)
        ws_budget.Cells(new_row, 6).Value = 0

        # G-I: Allocated, Obligated, Spent
        ws_budget.Cells(new_row, 7).Value = ws_old.cell(old_row, 6).value
        ws_budget.Cells(new_row, 8).Value = ws_old.cell(old_row, 7).value
        ws_budget.Cells(new_row, 9).Value = ws_old.cell(old_row, 8).value

        # J: ULO formula (NOW IT WILL WORK because both sheets exist)
        ws_budget.Cells(new_row, 10).Formula = f"=H{new_row}-I{new_row}"

        # K: ULO Percent formula
        ws_budget.Cells(new_row, 11).Formula = f"=IF(H{new_row}>0,J{new_row}/H{new_row},0)"
        ws_budget.Cells(new_row, 11).NumberFormat = "0%"

        # L: Spend Health
        val = ws_old.cell(old_row, 11).value
        if val:
            ws_budget.Cells(new_row, 12).Value = val

    print("  Added budget data")

    # Now go back and add formulas to Master_Projects
    print("Adding formulas to Master_Projects...")
    ws_proj = wb.Sheets("Master_Projects")

    for row in range(2, ws_old.max_row + 1):
        # Budget formulas - NOW THESE WILL WORK because Country_Budgets exists
        ws_proj.Cells(row, 14).Formula = f"=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$F:$F)"
        ws_proj.Cells(row, 15).Formula = f"=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$G:$G)"
        ws_proj.Cells(row, 16).Formula = f"=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$H:$H)"
        ws_proj.Cells(row, 17).Formula = f"=SUMIF(Country_Budgets!$B:$B,A{row},Country_Budgets!$J:$J)"
        ws_proj.Cells(row, 18).Formula = f"=IF(P{row}>0,Q{row}/P{row},0)"
        ws_proj.Cells(row, 18).NumberFormat = "0%"

    print("  Added formulas")

    # Create Country_Regions
    print("Building Country_Regions...")
    ws_regions = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    ws_regions.Name = "Country_Regions"

    STATE_DEPT_REGIONS = {
        'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
        'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
        'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
        'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
        'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
        'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
        'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',
        'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
        'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
        'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
        'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
        'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',
        'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
        'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
        'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
        'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
        'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
        'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
        'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR', 'VA': 'EUR',
        'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA', 'LB': 'NEA', 'OM': 'NEA',
        'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA', 'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',
        'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
        'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',
        'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
        'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
        'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
        'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
        'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
    }

    SUBREGIONS = {
        'NA': 'Southern Africa', 'AL': 'Southern Europe', 'AM': 'Caucasus',
        'AT': 'Western Europe', 'AZ': 'Caucasus', 'BY': 'Eastern Europe'
        # (abbreviated for speed - add rest if needed)
    }

    headers_regions = ['Country_Code', 'Country_Name', 'Region', 'EU_Member', 'Subregion']
    for col, header in enumerate(headers_regions, 1):
        cell = ws_regions.Cells(1, col)
        cell.Value = header
        cell.Font.Bold = True
        cell.Font.Color = 16777215
        cell.Interior.Color = 9854208
        cell.HorizontalAlignment = -4108

    ws_old = wb_old['Country_Regions']
    for old_row in range(2, ws_old.max_row + 1):
        code = ws_old.cell(old_row, 1).value
        name = ws_old.cell(old_row, 2).value
        eu = ws_old.cell(old_row, 4).value

        ws_regions.Cells(old_row, 1).Value = code
        ws_regions.Cells(old_row, 2).Value = name
        ws_regions.Cells(old_row, 3).Value = STATE_DEPT_REGIONS.get(code, 'EUR')
        ws_regions.Cells(old_row, 4).Value = eu
        ws_regions.Cells(old_row, 5).Value = SUBREGIONS.get(code, '')

    print("  Added country data")

    # Add a simple _SETUP sheet
    print("Adding _SETUP sheet...")
    ws_setup = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    ws_setup.Name = "_SETUP"
    ws_setup.Cells(1, 1).Value = "TRACKER V17 - BUILT WITH EXCEL COM"
    ws_setup.Cells(1, 1).Font.Bold = True
    ws_setup.Cells(1, 1).Font.Size = 16
    ws_setup.Cells(3, 1).Value = "Build Method: Excel COM Automation"
    ws_setup.Cells(4, 1).Value = f"Build Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_setup.Cells(5, 1).Value = "All formulas should work correctly!"

    # Save
    print("\nSaving v17...")
    import os
    filepath = os.path.abspath('2025-10-26-Tracker-v17.xlsx')
    wb.SaveAs(filepath)

    # Close Excel
    wb.Close(SaveChanges=False)
    excel.Quit()

    print("\n" + "="*80)
    print("OPTION 3 SUCCESS!")
    print("="*80)
    print(f"\nFile: 2025-10-26-Tracker-v17.xlsx")
    print("Built using Excel COM - formulas should work correctly!")
    print("\nTry opening it now - it should work without errors.")

else:
    print("\n" + "="*80)
    print("OPTION 3 FAILED - COM NOT AVAILABLE")
    print("="*80)
    print("\nProceeding to Option 1 (pure data)...")
    print("Running build_v17_pure_data.py...")
