"""
Format the Portfolio_Dashboard info panel (columns O-Q) as a nice table
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

print("="*80)
print("FORMATTING PORTFOLIO_DASHBOARD INFO PANEL")
print("="*80)

# Load v27 or v28
print("\nLoading v28...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v28.xlsx')

ws_port = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("CURRENT STRUCTURE (Columns O-Q, Rows 3-7)")
print("="*80)

# Check current content
print("\nCurrent content:")
for row in range(3, 8):
    o_val = ws_port.cell(row, 15).value  # Column O
    p_val = ws_port.cell(row, 16).value  # Column P
    q_val = ws_port.cell(row, 17).value  # Column Q
    print(f"  Row {row}: O={o_val} | P={p_val} | Q={q_val}")

print("\n" + "="*80)
print("APPLYING FORMATTING")
print("="*80)

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
label_font = Font(bold=True, size=10)
value_font = Font(size=10)

thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add header row if needed
print("\nAdding header row...")
ws_port['O2'] = 'Portfolio Summary'
ws_port['O2'].font = header_font
ws_port['O2'].fill = header_fill
ws_port['O2'].alignment = Alignment(horizontal='center', vertical='center')
ws_port['O2'].border = thick_border

# Merge O2:Q2 for header
ws_port.merge_cells('O2:Q2')

print("  Header added: 'Portfolio Summary' (merged O2:Q2)")

# Format label column (O3:O7)
print("\nFormatting labels (column O)...")
for row in range(3, 8):
    cell = ws_port.cell(row, 15)  # Column O
    cell.fill = label_fill
    cell.font = label_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border

print("  Labels formatted with light blue background")

# Format value columns (P3:Q7)
print("\nFormatting value columns (P-Q)...")
for row in range(3, 8):
    for col in [16, 17]:  # Columns P, Q
        cell = ws_port.cell(row, col)
        cell.font = value_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

print("  Value columns formatted")

# Adjust column widths
print("\nAdjusting column widths...")
ws_port.column_dimensions['O'].width = 22
ws_port.column_dimensions['P'].width = 18
ws_port.column_dimensions['Q'].width = 18

print("  Column O: 22")
print("  Column P: 18")
print("  Column Q: 18")

# Add outer border to table
print("\nAdding outer borders...")
for row in [3, 7]:  # Top and bottom rows
    for col in range(15, 18):  # Columns O, P, Q
        cell = ws_port.cell(row, col)
        if row == 3:  # Top row
            cell.border = Border(
                left=Side(style='medium' if col == 15 else 'thin'),
                right=Side(style='medium' if col == 17 else 'thin'),
                top=Side(style='medium'),
                bottom=Side(style='thin')
            )
        else:  # Bottom row
            cell.border = Border(
                left=Side(style='medium' if col == 15 else 'thin'),
                right=Side(style='medium' if col == 17 else 'thin'),
                top=Side(style='thin'),
                bottom=Side(style='medium')
            )

# Left and right edges
for row in range(4, 7):
    ws_port.cell(row, 15).border = Border(
        left=Side(style='medium'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws_port.cell(row, 17).border = Border(
        left=Side(style='thin'),
        right=Side(style='medium'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

print("  Outer borders applied")

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Info panel formatted!")

print("\n" + "="*80)
print("FORMATTING APPLIED")
print("="*80)

print("\nTable structure:")
print("  Row 2: Header - 'Portfolio Summary' (blue background, white text)")
print("  Rows 3-7: Info rows")
print("    Column O: Labels (light blue background)")
print("    Columns P-Q: Values (white background)")
print("  Borders: Medium outer border, thin inner borders")

print("\nVisual layout:")
print("  ┌─────────────────────────────────────────┐")
print("  │      Portfolio Summary (Header)         │")
print("  ├─────────────────────┬─────────┬─────────┤")
print("  │ Project_Manager     │ [value] │ [value] │")
print("  │ My_Countries_Count  │ [value] │ [value] │")
print("  │ Total_Countries_Count│ [value]│ [value] │")
print("  │ FAR_Notes           │ [value] │ [value] │")
print("  │ NCE_Notes           │ [value] │ [value] │")
print("  └─────────────────────┴─────────┴─────────┘")

print("\nNow add formulas/values to columns P and Q!")
print("v28 updated!")
