"""
Fix Regional_Summary #REF! errors
Check what's causing the errors and rebuild formulas
"""
import openpyxl

print("="*80)
print("FIXING REGIONAL_SUMMARY #REF! ERRORS")
print("="*80)

# Load v27
print("\nLoading v27...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v27.xlsx', data_only=False)
ws_regional = wb['Regional_Summary']

print("\n" + "="*80)
print("CHECKING CURRENT FORMULAS")
print("="*80)

print("\nRow 2-7 current state:")
for row in range(2, 8):
    region = ws_regional.cell(row, 1).value
    print(f"\nRow {row} ({region}):")
    for col in range(2, 11):
        cell = ws_regional.cell(row, col)
        if cell.value:
            value_str = str(cell.value)[:60]
            has_ref = '#REF!' in value_str
            print(f"  Col {col}: {value_str}{'  <-- REF ERROR' if has_ref else ''}")

print("\n" + "="*80)
print("REBUILDING REGIONAL_SUMMARY FORMULAS")
print("="*80)

# Region codes
regions = ['AF', 'EAP', 'EUR', 'NEA', 'SCA', 'WHA']

print("\nRebuilding formulas for each region...")

for idx, region in enumerate(regions, start=2):
    print(f"\n  Row {idx} ({region}):")

    # Column A: Region code (already correct)
    ws_regional.cell(idx, 1).value = region

    # Column B: Region name (manual entry - skip)

    # Column C: Project count
    formula = f'=SUMPRODUCT((ISNUMBER(FIND(A{idx},T_Master_Projects[Countries])))*1)'
    ws_regional.cell(idx, 3).value = formula
    print(f"    C{idx}: Project count")

    # Column D: Country count
    formula = f'=COUNTIF(T_Country_Regions[Region],A{idx})'
    ws_regional.cell(idx, 4).value = formula
    print(f"    D{idx}: Country count")

    # Column E: Active budgets
    formula = f'=SUMIF(T_Country_Regions[Region],A{idx},T_Country_Budgets[Allocated_Amount])'
    ws_regional.cell(idx, 5).value = formula
    print(f"    E{idx}: Active budgets")

    # Column F: Total Allocated
    formula = f'=SUMIF(T_Country_Regions[Region],A{idx},T_Country_Budgets[Allocated_Amount])'
    ws_regional.cell(idx, 6).value = formula
    print(f"    F{idx}: Total Allocated")

    # Column G: Total Obligated
    formula = f'=SUMIF(T_Country_Regions[Region],A{idx},T_Country_Budgets[Obligated_Amount])'
    ws_regional.cell(idx, 7).value = formula
    print(f"    G{idx}: Total Obligated")

    # Column H: Total Spent
    formula = f'=SUMIF(T_Country_Regions[Region],A{idx},T_Country_Budgets[Spent_Amount])'
    ws_regional.cell(idx, 8).value = formula
    print(f"    H{idx}: Total Spent")

    # Column I: Total ULO
    formula = f'=F{idx}-G{idx}'
    ws_regional.cell(idx, 9).value = formula
    print(f"    I{idx}: Total ULO (Allocated - Obligated)")

    # Column J: ULO Percent
    formula = f'=IF(F{idx}=0,0,I{idx}/F{idx})'
    ws_regional.cell(idx, 10).value = formula
    print(f"    J{idx}: ULO Percent")

    # Column K: Execution Rate
    formula = f'=IF(G{idx}=0,0,H{idx}/G{idx})'
    ws_regional.cell(idx, 11).value = formula
    print(f"    K{idx}: Execution Rate (Spent/Obligated)")

    # Column L: Status indicator (manual - skip)

print("\n" + "="*80)
print("SAVING V28")
print("="*80)

wb.save('2025-10-26-Tracker-v28.xlsx')

print("\nOK - Regional_Summary formulas rebuilt!")
print("Saved as: 2025-10-26-Tracker-v28.xlsx")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nRegional_Summary formulas rebuilt for all 6 regions:")
print("  - AF (Africa)")
print("  - EAP (East Asia Pacific)")
print("  - EUR (Europe)")
print("  - NEA (Near East Asia)")
print("  - SCA (South/Central Asia)")
print("  - WHA (Western Hemisphere)")

print("\nColumns with formulas:")
print("  C: Project Count")
print("  D: Country Count")
print("  E: Active Budgets")
print("  F: Total Allocated")
print("  G: Total Obligated")
print("  H: Total Spent")
print("  I: Total ULO")
print("  J: ULO Percent")
print("  K: Execution Rate")

print("\nAll #REF! errors should be fixed!")
print("v27 is ready to use!")
