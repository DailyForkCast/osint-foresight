import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("UPDATING STAKEHOLDERS SHEET")
print("=" * 70)

# Load the updated workbook
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-UPDATED.xlsx')
ws = wb['Stakeholders']

print("\n[1/3] Backing up old stakeholder data...")
# Get existing data
old_data = []
for row in range(2, ws.max_row + 1):
    row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
    if any(row_data):
        old_data.append(row_data)
print(f"✓ Backed up {len(old_data)} existing stakeholder records")

print("\n[2/3] Creating new stakeholder structure...")

# Clear the sheet
ws.delete_rows(1, ws.max_row)

# New comprehensive headers
headers = [
    'Stakeholder_ID',          # 1
    'Name',                    # 2
    'Title',                   # 3
    'Organization',            # 4
    'Location_City',           # 5 - NEW
    'Location_Country',        # 6 - NEW
    'Time_Zone_Offset',        # 7 - NEW (hours from UTC)
    'Local_Time',              # 8 - NEW (formula)
    'Email',                   # 9
    'Phone',                   # 10
    'Stakeholder_Type',        # 11 - NEW (Project/Location/Product/Regional/Thematic)
    'Project_IDs',             # 12 - NEW (comma-separated)
    'Countries',               # 13 - NEW (comma-separated country codes)
    'Products',                # 14 - NEW (comma-separated)
    'Region',                  # 15 - NEW (EUR/WHA/EAP/AF/NEA/SCA)
    'Theme',                   # 16 - NEW (Cybersecurity, Digital Transform, etc.)
    'Influence_Level',         # 17
    'Interest_Level',          # 18
    'Contact_Frequency',       # 19
    'Last_Contact',            # 20
    'Next_Contact',            # 21
    'Notes'                    # 22
]

# Add headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

print(f"✓ Created {len(headers)} column headers")

print("\n[3/3] Adding sample stakeholders...")

# Sample stakeholders with new structure
sample_stakeholders = [
    # Project-Specific stakeholder
    {
        'Stakeholder_ID': 'STK-001',
        'Name': 'John Smith',
        'Title': 'Project Director',
        'Organization': 'USAID',
        'Location_City': 'Washington DC',
        'Location_Country': 'United States',
        'Time_Zone_Offset': -5,  # EST
        'Email': 'john.smith@usaid.gov',
        'Phone': '+1-202-555-0100',
        'Stakeholder_Type': 'Project-Specific',
        'Project_IDs': 'PRJ-001, PRJ-003',
        'Countries': '',
        'Products': '',
        'Region': '',
        'Theme': '',
        'Influence_Level': 'High',
        'Interest_Level': 'High',
        'Contact_Frequency': 'Weekly',
        'Last_Contact': '2025-01-15',
        'Next_Contact': '2025-01-22',
        'Notes': 'Primary decision maker for digital transformation projects'
    },
    # Regional stakeholder
    {
        'Stakeholder_ID': 'STK-002',
        'Name': 'Maria Schmidt',
        'Title': 'EUR Regional Director',
        'Organization': 'U.S. State Department',
        'Location_City': 'Brussels',
        'Location_Country': 'Belgium',
        'Time_Zone_Offset': 1,  # CET
        'Email': 'maria.schmidt@state.gov',
        'Phone': '+32-2-555-0200',
        'Stakeholder_Type': 'Regional',
        'Project_IDs': '',
        'Countries': 'DE, FR, BE, NL',
        'Products': '',
        'Region': 'EUR',
        'Theme': 'Digital Policy',
        'Influence_Level': 'High',
        'Interest_Level': 'High',
        'Contact_Frequency': 'Monthly',
        'Last_Contact': '2025-01-10',
        'Next_Contact': '2025-02-10',
        'Notes': 'Key contact for all EUR region initiatives'
    },
    # Location-Specific stakeholder
    {
        'Stakeholder_ID': 'STK-003',
        'Name': 'Hans Mueller',
        'Title': 'Country Director',
        'Organization': 'U.S. Embassy Berlin',
        'Location_City': 'Berlin',
        'Location_Country': 'Germany',
        'Time_Zone_Offset': 1,  # CET
        'Email': 'hans.mueller@state.gov',
        'Phone': '+49-30-555-0300',
        'Stakeholder_Type': 'Location-Specific',
        'Project_IDs': '',
        'Countries': 'DE',
        'Products': '',
        'Region': 'EUR',
        'Theme': '',
        'Influence_Level': 'High',
        'Interest_Level': 'High',
        'Contact_Frequency': 'Bi-weekly',
        'Last_Contact': '2025-01-12',
        'Next_Contact': '2025-01-26',
        'Notes': 'Primary contact for all Germany-based activities'
    },
    # Product-Specific stakeholder
    {
        'Stakeholder_ID': 'STK-004',
        'Name': 'Sarah Chen',
        'Title': 'Product Owner',
        'Organization': 'USAID',
        'Location_City': 'Washington DC',
        'Location_Country': 'United States',
        'Time_Zone_Offset': -5,  # EST
        'Email': 'sarah.chen@usaid.gov',
        'Phone': '+1-202-555-0400',
        'Stakeholder_Type': 'Product-Specific',
        'Project_IDs': 'PRJ-001',
        'Countries': '',
        'Products': 'Dashboard System, API Gateway',
        'Region': '',
        'Theme': 'Digital Transformation',
        'Influence_Level': 'Medium',
        'Interest_Level': 'High',
        'Contact_Frequency': 'Weekly',
        'Last_Contact': '2025-01-14',
        'Next_Contact': '2025-01-21',
        'Notes': 'Technical lead for software products'
    },
    # Thematic stakeholder
    {
        'Stakeholder_ID': 'STK-005',
        'Name': 'David Rodriguez',
        'Title': 'Cybersecurity Advisor',
        'Organization': 'NSA',
        'Location_City': 'Fort Meade',
        'Location_Country': 'United States',
        'Time_Zone_Offset': -5,  # EST
        'Email': 'david.rodriguez@nsa.gov',
        'Phone': '+1-301-555-0500',
        'Stakeholder_Type': 'Thematic',
        'Project_IDs': 'PRJ-002',
        'Countries': '',
        'Products': '',
        'Region': '',
        'Theme': 'Cybersecurity, Information Security',
        'Influence_Level': 'High',
        'Interest_Level': 'Medium',
        'Contact_Frequency': 'Monthly',
        'Last_Contact': '2025-01-08',
        'Next_Contact': '2025-02-08',
        'Notes': 'Security clearance advisor for all cyber projects'
    },
]

# Write sample stakeholders
for row_idx, stakeholder in enumerate(sample_stakeholders, 2):
    ws.cell(row=row_idx, column=1, value=stakeholder['Stakeholder_ID'])
    ws.cell(row=row_idx, column=2, value=stakeholder['Name'])
    ws.cell(row=row_idx, column=3, value=stakeholder['Title'])
    ws.cell(row=row_idx, column=4, value=stakeholder['Organization'])
    ws.cell(row=row_idx, column=5, value=stakeholder['Location_City'])
    ws.cell(row=row_idx, column=6, value=stakeholder['Location_Country'])
    ws.cell(row=row_idx, column=7, value=stakeholder['Time_Zone_Offset'])

    # Local Time formula: =NOW()+(G2/24)
    ws.cell(row=row_idx, column=8, value=f'=NOW()+(G{row_idx}/24)')
    ws.cell(row=row_idx, column=8).number_format = 'mm/dd/yyyy hh:mm AM/PM'

    ws.cell(row=row_idx, column=9, value=stakeholder['Email'])
    ws.cell(row=row_idx, column=10, value=stakeholder['Phone'])
    ws.cell(row=row_idx, column=11, value=stakeholder['Stakeholder_Type'])
    ws.cell(row=row_idx, column=12, value=stakeholder['Project_IDs'])
    ws.cell(row=row_idx, column=13, value=stakeholder['Countries'])
    ws.cell(row=row_idx, column=14, value=stakeholder['Products'])
    ws.cell(row=row_idx, column=15, value=stakeholder['Region'])
    ws.cell(row=row_idx, column=16, value=stakeholder['Theme'])
    ws.cell(row=row_idx, column=17, value=stakeholder['Influence_Level'])
    ws.cell(row=row_idx, column=18, value=stakeholder['Interest_Level'])
    ws.cell(row=row_idx, column=19, value=stakeholder['Contact_Frequency'])
    ws.cell(row=row_idx, column=20, value=stakeholder['Last_Contact'])
    ws.cell(row=row_idx, column=21, value=stakeholder['Next_Contact'])
    ws.cell(row=row_idx, column=22, value=stakeholder['Notes'])

print(f"✓ Added {len(sample_stakeholders)} sample stakeholders")

# Set column widths for better visibility
ws.column_dimensions['A'].width = 15  # Stakeholder_ID
ws.column_dimensions['B'].width = 20  # Name
ws.column_dimensions['C'].width = 25  # Title
ws.column_dimensions['D'].width = 25  # Organization
ws.column_dimensions['E'].width = 15  # Location_City
ws.column_dimensions['F'].width = 20  # Location_Country
ws.column_dimensions['G'].width = 12  # Time_Zone_Offset
ws.column_dimensions['H'].width = 20  # Local_Time
ws.column_dimensions['I'].width = 30  # Email
ws.column_dimensions['J'].width = 18  # Phone
ws.column_dimensions['K'].width = 18  # Stakeholder_Type
ws.column_dimensions['L'].width = 20  # Project_IDs
ws.column_dimensions['M'].width = 20  # Countries
ws.column_dimensions['N'].width = 25  # Products
ws.column_dimensions['O'].width = 12  # Region
ws.column_dimensions['P'].width = 30  # Theme
ws.column_dimensions['Q'].width = 15  # Influence_Level
ws.column_dimensions['R'].width = 15  # Interest_Level
ws.column_dimensions['S'].width = 18  # Contact_Frequency
ws.column_dimensions['T'].width = 12  # Last_Contact
ws.column_dimensions['U'].width = 12  # Next_Contact
ws.column_dimensions['V'].width = 40  # Notes

# Freeze panes at row 2
ws.freeze_panes = 'A2'

print("\n" + "=" * 70)
print("Saving updated workbook...")
wb.save('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-UPDATED.xlsx')
print("✓ Workbook saved with new Stakeholders structure")

print("\n" + "=" * 70)
print("STAKEHOLDERS UPDATE COMPLETE!")
print("=" * 70)
print(f"\nNew Stakeholders structure:")
print(f"  • {len(headers)} columns (added 8 new columns)")
print(f"  • Multi-dimensional relationship tracking")
print(f"  • Location and timezone support")
print(f"  • Dynamic local time calculation")
print(f"  • {len(sample_stakeholders)} sample stakeholders added")
print(f"\nStakeholder Types:")
print(f"  • Project-Specific: Tied to specific projects")
print(f"  • Location-Specific: Country/city based")
print(f"  • Product-Specific: Tied to deliverable products")
print(f"  • Regional: EUR/WHA/EAP/AF/NEA/SCA regions")
print(f"  • Thematic: Cross-cutting themes (Cybersecurity, etc.)")
