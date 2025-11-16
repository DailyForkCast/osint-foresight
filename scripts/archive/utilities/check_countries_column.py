"""
Check if Countries column is working in Master_Projects
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v18.xlsx', data_only=True)

print("="*80)
print("CHECKING COUNTRIES COLUMN (Column S)")
print("="*80)

ws_projects = wb['Master_Projects']

print("\nMaster_Projects - Column S values:")
print(f"{'Row':<5} {'Project ID':<15} {'Countries':<50}")
print("-"*70)

for row in range(2, ws_projects.max_row + 1):
    project_id = ws_projects.cell(row, 1).value or ""
    countries = ws_projects.cell(row, 19).value or ""  # Column S

    if project_id:
        print(f"{row:<5} {str(project_id):<15} {str(countries)[:48]:<50}")

print("\n" + "="*80)
print("CHECKING COUNTRY_BUDGETS DATA")
print("="*80)

ws_budgets = wb['Country_Budgets']

print("\nCountry_Budgets - Project to Country mapping:")
print(f"{'Row':<5} {'Project ID':<15} {'Country Name':<30}")
print("-"*50)

for row in range(2, min(ws_budgets.max_row + 1, 20)):  # First 20 rows
    project_id = ws_budgets.cell(row, 2).value or ""  # Column B
    country_name = ws_budgets.cell(row, 5).value or ""  # Column E

    if project_id and country_name:
        print(f"{row:<5} {str(project_id):<15} {str(country_name):<30}")

print("\n" + "="*80)
print("If Countries column shows blank, the TEXTJOIN formula may not be working.")
print("This could be because:")
print("  1. Formula wasn't entered as array formula (need Ctrl+Shift+Enter)")
print("  2. Your Excel version doesn't support TEXTJOIN")
print("  3. No country data exists in Country_Budgets yet")
print("="*80)
