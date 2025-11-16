"""
Build Tracker v16 - Phase 1: Data Foundation Only
Fast, clean, no tables, no merged cells
All dashboards/spotlights to be built manually in Excel
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

print("="*80)
print("BUILDING TRACKER V16 - DATA FOUNDATION")
print("="*80)
print(f"Started: {datetime.now().strftime('%H:%M:%S')}")

# Load v12 to get existing data
print("\nLoading v12 data...")
wb_old = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx')

# Create new workbook
wb = Workbook()
wb.remove(wb.active)

# Helper function
def create_header(ws, row, headers, bg_color='366092'):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

# State Department regions
STATE_DEPT_REGIONS = {
    'DZ': 'AF', 'AO': 'AF', 'BJ': 'AF', 'BW': 'AF', 'BF': 'AF', 'BI': 'AF', 'CM': 'AF', 'CV': 'AF',
    'CF': 'AF', 'TD': 'AF', 'KM': 'AF', 'CG': 'AF', 'CD': 'AF', 'CI': 'AF', 'DJ': 'AF', 'EG': 'AF',
    'GQ': 'AF', 'ER': 'AF', 'ET': 'AF', 'GA': 'AF', 'GM': 'AF', 'GH': 'AF', 'GN': 'AF', 'GW': 'AF',
    'KE': 'AF', 'LS': 'AF', 'LR': 'AF', 'LY': 'AF', 'MG': 'AF', 'MW': 'AF', 'ML': 'AF', 'MR': 'AF',
    'MU': 'AF', 'MA': 'AF', 'MZ': 'AF', 'NA': 'AF', 'NE': 'AF', 'NG': 'AF', 'RW': 'AF', 'ST': 'AF',
    'SN': 'AF', 'SC': 'AF', 'SL': 'AF', 'SO': 'AF', 'ZA': 'AF', 'SS': 'AF', 'SD': 'AF', 'SZ': 'AF',
    'TZ': 'AF', 'TG': 'AF', 'TN': 'AF', 'UG': 'AF', 'ZM': 'AF', 'ZW': 'AF',
    'AU': 'EAP', 'BN': 'EAP', 'KH': 'EAP', 'CN': 'EAP', 'FJ': 'EAP', 'ID': 'EAP', 'JP': 'EAP',
    'KI': 'EAP', 'LA': 'EAP', 'MY': 'EAP', 'MV': 'EAP', 'MH': 'EAP', 'FM': 'EAP', 'MN': 'EAP',
    'MM': 'EAP', 'NR': 'EAP', 'NZ': 'EAP', 'KP': 'EAP', 'PW': 'EAP', 'PG': 'EAP', 'PH': 'EAP',
    'WS': 'EAP', 'SG': 'EAP', 'SB': 'EAP', 'KR': 'EAP', 'TW': 'EAP', 'TH': 'EAP',
    'TL': 'EAP', 'TO': 'EAP', 'TV': 'EAP', 'VU': 'EAP', 'VN': 'EAP',
    'AL': 'EUR', 'AD': 'EUR', 'AM': 'EUR', 'AT': 'EUR', 'AZ': 'EUR', 'BY': 'EUR', 'BE': 'EUR',
    'BA': 'EUR', 'BG': 'EUR', 'HR': 'EUR', 'CY': 'EUR', 'CZ': 'EUR', 'DK': 'EUR', 'EE': 'EUR',
    'FI': 'EUR', 'FR': 'EUR', 'GE': 'EUR', 'DE': 'EUR', 'GR': 'EUR', 'HU': 'EUR', 'IS': 'EUR',
    'IE': 'EUR', 'IT': 'EUR', 'XK': 'EUR', 'LV': 'EUR', 'LI': 'EUR', 'LT': 'EUR', 'LU': 'EUR',
    'MK': 'EUR', 'MT': 'EUR', 'MD': 'EUR', 'MC': 'EUR', 'ME': 'EUR', 'NL': 'EUR', 'NO': 'EUR',
    'PL': 'EUR', 'PT': 'EUR', 'RO': 'EUR', 'RU': 'EUR', 'SM': 'EUR', 'RS': 'EUR', 'SK': 'EUR',
    'SI': 'EUR', 'ES': 'EUR', 'SE': 'EUR', 'CH': 'EUR', 'TR': 'EUR', 'UA': 'EUR', 'GB': 'EUR', 'VA': 'EUR',
    'BH': 'NEA', 'IQ': 'NEA', 'IL': 'NEA', 'JO': 'NEA', 'KW': 'NEA', 'LB': 'NEA', 'OM': 'NEA',
    'PS': 'NEA', 'QA': 'NEA', 'SA': 'NEA', 'SY': 'NEA', 'AE': 'NEA', 'YE': 'NEA',
    'AF': 'SCA', 'BD': 'SCA', 'BT': 'SCA', 'IN': 'SCA', 'KZ': 'SCA', 'KG': 'SCA', 'NP': 'SCA',
    'PK': 'SCA', 'LK': 'SCA', 'TJ': 'SCA', 'TM': 'SCA', 'UZ': 'SCA',
    'AG': 'WHA', 'AR': 'WHA', 'BS': 'WHA', 'BB': 'WHA', 'BZ': 'WHA', 'BO': 'WHA', 'BR': 'WHA',
    'CA': 'WHA', 'CL': 'WHA', 'CO': 'WHA', 'CR': 'WHA', 'CU': 'WHA', 'DM': 'WHA', 'DO': 'WHA',
    'EC': 'WHA', 'SV': 'WHA', 'GD': 'WHA', 'GT': 'WHA', 'GY': 'WHA', 'HT': 'WHA', 'HN': 'WHA',
    'JM': 'WHA', 'MX': 'WHA', 'NI': 'WHA', 'PA': 'WHA', 'PY': 'WHA', 'PE': 'WHA', 'KN': 'WHA',
    'LC': 'WHA', 'VC': 'WHA', 'SR': 'WHA', 'TT': 'WHA', 'US': 'WHA', 'UY': 'WHA', 'VE': 'WHA',
}

SUBREGIONS = {
    'DZ': 'Northern Africa', 'EG': 'Northern Africa', 'LY': 'Northern Africa', 'MA': 'Northern Africa',
    'TN': 'Northern Africa', 'SD': 'Northern Africa',
    'BJ': 'Western Africa', 'BF': 'Western Africa', 'CV': 'Western Africa', 'CI': 'Western Africa',
    'GM': 'Western Africa', 'GH': 'Western Africa', 'GN': 'Western Africa', 'GW': 'Western Africa',
    'LR': 'Western Africa', 'ML': 'Western Africa', 'MR': 'Western Africa', 'NE': 'Western Africa',
    'NG': 'Western Africa', 'SN': 'Western Africa', 'SL': 'Western Africa', 'TG': 'Western Africa',
    'BI': 'Eastern Africa', 'KM': 'Eastern Africa', 'DJ': 'Eastern Africa', 'ER': 'Eastern Africa',
    'ET': 'Eastern Africa', 'KE': 'Eastern Africa', 'MG': 'Eastern Africa', 'MW': 'Eastern Africa',
    'MU': 'Eastern Africa', 'MZ': 'Eastern Africa', 'RW': 'Eastern Africa', 'SC': 'Eastern Africa',
    'SO': 'Eastern Africa', 'SS': 'Eastern Africa', 'TZ': 'Eastern Africa', 'UG': 'Eastern Africa',
    'ZM': 'Eastern Africa', 'ZW': 'Eastern Africa',
    'AO': 'Middle Africa', 'CM': 'Middle Africa', 'CF': 'Middle Africa', 'TD': 'Middle Africa',
    'CG': 'Middle Africa', 'CD': 'Middle Africa', 'GQ': 'Middle Africa', 'GA': 'Middle Africa',
    'BW': 'Southern Africa', 'LS': 'Southern Africa', 'NA': 'Southern Africa', 'ZA': 'Southern Africa', 'SZ': 'Southern Africa',
    'DK': 'Northern Europe', 'EE': 'Northern Europe', 'FI': 'Northern Europe', 'IS': 'Northern Europe',
    'IE': 'Northern Europe', 'LV': 'Northern Europe', 'LT': 'Northern Europe', 'NO': 'Northern Europe',
    'SE': 'Northern Europe', 'GB': 'Northern Europe',
    'BY': 'Eastern Europe', 'BG': 'Eastern Europe', 'CZ': 'Eastern Europe', 'HU': 'Eastern Europe',
    'MD': 'Eastern Europe', 'PL': 'Eastern Europe', 'RO': 'Eastern Europe', 'RU': 'Eastern Europe',
    'SK': 'Eastern Europe', 'UA': 'Eastern Europe',
    'AL': 'Southern Europe', 'AD': 'Southern Europe', 'BA': 'Southern Europe', 'HR': 'Southern Europe',
    'CY': 'Southern Europe', 'GR': 'Southern Europe', 'IT': 'Southern Europe', 'XK': 'Southern Europe',
    'MK': 'Southern Europe', 'MT': 'Southern Europe', 'ME': 'Southern Europe', 'PT': 'Southern Europe',
    'RS': 'Southern Europe', 'SI': 'Southern Europe', 'ES': 'Southern Europe', 'VA': 'Southern Europe',
    'AT': 'Western Europe', 'BE': 'Western Europe', 'FR': 'Western Europe', 'DE': 'Western Europe',
    'LI': 'Western Europe', 'LU': 'Western Europe', 'MC': 'Western Europe', 'NL': 'Western Europe', 'CH': 'Western Europe',
    'AM': 'Caucasus', 'AZ': 'Caucasus', 'GE': 'Caucasus', 'TR': 'Western Asia',
    'CA': 'Northern America', 'US': 'Northern America',
    'BZ': 'Central America', 'CR': 'Central America', 'SV': 'Central America', 'GT': 'Central America',
    'HN': 'Central America', 'MX': 'Central America', 'NI': 'Central America', 'PA': 'Central America',
    'AG': 'Caribbean', 'BS': 'Caribbean', 'BB': 'Caribbean', 'CU': 'Caribbean', 'DM': 'Caribbean',
    'DO': 'Caribbean', 'GD': 'Caribbean', 'HT': 'Caribbean', 'JM': 'Caribbean', 'KN': 'Caribbean',
    'LC': 'Caribbean', 'VC': 'Caribbean', 'TT': 'Caribbean',
    'AR': 'South America', 'BO': 'South America', 'BR': 'South America', 'CL': 'South America',
    'CO': 'South America', 'EC': 'South America', 'GY': 'South America', 'PY': 'South America',
    'PE': 'South America', 'SR': 'South America', 'UY': 'South America', 'VE': 'South America',
    'CN': 'Eastern Asia', 'JP': 'Eastern Asia', 'KP': 'Eastern Asia', 'KR': 'Eastern Asia',
    'MN': 'Eastern Asia', 'TW': 'Eastern Asia',
    'BN': 'South-Eastern Asia', 'KH': 'South-Eastern Asia', 'ID': 'South-Eastern Asia',
    'LA': 'South-Eastern Asia', 'MY': 'South-Eastern Asia', 'MM': 'South-Eastern Asia',
    'PH': 'South-Eastern Asia', 'SG': 'South-Eastern Asia', 'TH': 'South-Eastern Asia',
    'TL': 'South-Eastern Asia', 'VN': 'South-Eastern Asia',
    'AF': 'Southern Asia', 'BD': 'Southern Asia', 'BT': 'Southern Asia', 'IN': 'Southern Asia',
    'MV': 'Southern Asia', 'NP': 'Southern Asia', 'PK': 'Southern Asia', 'LK': 'Southern Asia',
    'KZ': 'Central Asia', 'KG': 'Central Asia', 'TJ': 'Central Asia', 'TM': 'Central Asia', 'UZ': 'Central Asia',
    'AU': 'Australia and New Zealand', 'NZ': 'Australia and New Zealand',
    'FJ': 'Melanesia', 'PG': 'Melanesia', 'SB': 'Melanesia', 'VU': 'Melanesia',
    'KI': 'Micronesia', 'MH': 'Micronesia', 'FM': 'Micronesia', 'NR': 'Micronesia', 'PW': 'Micronesia',
    'WS': 'Polynesia', 'TO': 'Polynesia', 'TV': 'Polynesia',
    'BH': 'Arabian Peninsula', 'KW': 'Arabian Peninsula', 'OM': 'Arabian Peninsula',
    'QA': 'Arabian Peninsula', 'SA': 'Arabian Peninsula', 'AE': 'Arabian Peninsula', 'YE': 'Arabian Peninsula',
    'IL': 'Levant', 'JO': 'Levant', 'LB': 'Levant', 'PS': 'Levant', 'SY': 'Levant', 'IQ': 'Mesopotamia',
}

# ========== 1. Master_Projects ==========
print("1. Building Master_Projects...")
ws = wb.create_sheet("Master_Projects")
headers = [
    'Project_ID', 'Project_Unique_ID', 'Project_Name', 'Project_Status', 'Project_Priority',
    'Project_Progress', 'Project_Start_Date', 'Project_End_Date', 'Days_Remaining',
    'Implementer', 'Implementer_POC', 'Implementer_POC_Phone', 'Implementer_POC_Email',
    'Total_Proposed', 'Total_Allocation', 'Total_Obligated', 'Total_ULO', 'ULO_Percent',
    'Countries', 'Country_Count'
]
create_header(ws, 1, headers)

# Copy data from v12
ws_old = wb_old['Master_Projects']
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    # Copy A-M (basic fields)
    for col in range(1, 14):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value

    # Add budget formulas
    ws.cell(new_row, 14).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$F:$F)'
    ws.cell(new_row, 15).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$G:$G)'
    ws.cell(new_row, 16).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$H:$H)'
    ws.cell(new_row, 17).value = f'=SUMIF(Country_Budgets!$B:$B,A{new_row},Country_Budgets!$J:$J)'
    ws.cell(new_row, 18).value = f'=IF(P{new_row}>0,Q{new_row}/P{new_row},0)'
    ws.cell(new_row, 18).number_format = '0%'

    # Countries and count
    ws.cell(new_row, 19).value = ws_old.cell(old_row, 18).value
    ws.cell(new_row, 20).value = ws_old.cell(old_row, 19).value

print(f"   {ws_old.max_row - 1} projects")

# ========== 2. Country_Budgets ==========
print("2. Building Country_Budgets...")
ws = wb.create_sheet("Country_Budgets")
headers = [
    'Budget_ID', 'Unique_ID', 'My_Country', 'Country_Code', 'Country_Name',
    'Proposed_Amount', 'Allocated_Amount', 'Obligated_Amount', 'Spent_Amount',
    'ULO', 'ULO_Percent', 'Spend_Health'
]
create_header(ws, 1, headers)

ws_old = wb_old['Country_Budgets']
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    # Copy A-E
    for col in range(1, 6):
        ws.cell(new_row, col).value = ws_old.cell(old_row, col).value

    # F: Proposed (new, default 0)
    ws.cell(new_row, 6).value = 0

    # G-I: Allocated, Obligated, Spent
    ws.cell(new_row, 7).value = ws_old.cell(old_row, 6).value
    ws.cell(new_row, 8).value = ws_old.cell(old_row, 7).value
    ws.cell(new_row, 9).value = ws_old.cell(old_row, 8).value

    # J-K: ULO formulas
    ws.cell(new_row, 10).value = f'=H{new_row}-I{new_row}'
    ws.cell(new_row, 11).value = f'=IF(H{new_row}>0,J{new_row}/H{new_row},0)'
    ws.cell(new_row, 11).number_format = '0%'

    # L: Spend Health
    ws.cell(new_row, 12).value = ws_old.cell(old_row, 11).value

print(f"   {ws_old.max_row - 1} budget entries")

# ========== 3. Country_Regions ==========
print("3. Building Country_Regions...")
ws = wb.create_sheet("Country_Regions")
headers = ['Country_Code', 'Country_Name', 'Region', 'EU_Member', 'Subregion']
create_header(ws, 1, headers)

ws_old = wb_old['Country_Regions']
for old_row in range(2, ws_old.max_row + 1):
    new_row = old_row
    code = ws_old.cell(old_row, 1).value
    name = ws_old.cell(old_row, 2).value
    eu = ws_old.cell(old_row, 4).value

    ws.cell(new_row, 1).value = code
    ws.cell(new_row, 2).value = name
    ws.cell(new_row, 3).value = STATE_DEPT_REGIONS.get(code, 'EUR')
    ws.cell(new_row, 4).value = eu
    ws.cell(new_row, 5).value = SUBREGIONS.get(code, '')

print(f"   {ws_old.max_row - 1} countries")

# ========== 4-11. Supporting Sheets ==========
supporting_sheets = [
    'Milestones', 'Events', 'Risk_Register', 'Decision_Log', 'Stakeholders',
    'Project_Deliverables', 'Project_Audiences', 'Project_Products', 'Project_Technologies'
]

print("4. Building supporting sheets...")
for sheet_name in supporting_sheets:
    if sheet_name in wb_old.sheetnames:
        ws_old = wb_old[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Copy headers
        for col in range(1, ws_old.max_column + 1):
            val = ws_old.cell(1, col).value
            if val:
                cell = ws.cell(1, col, value=val)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Copy data
        for old_row in range(2, ws_old.max_row + 1):
            for col in range(1, ws_old.max_column + 1):
                ws.cell(old_row, col).value = ws_old.cell(old_row, col).value

        print(f"   {sheet_name}: {ws_old.max_row - 1} rows")

# ========== 12. Config_Lists ==========
print("5. Building Config_Lists...")
ws = wb.create_sheet("Config_Lists")
headers = ['List_Type', 'Value1', 'Value2', 'Value3', 'Value4', 'Value5']
create_header(ws, 1, headers)

# Copy existing config from v12
ws_old = wb_old['Config_Lists']
for old_row in range(2, ws_old.max_row + 1):
    for col in range(1, 7):
        ws.cell(old_row, col).value = ws_old.cell(old_row, col).value

# Add countries
ws_regions = wb['Country_Regions']
next_row = ws.max_row + 1
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws.cell(next_row, 1).value = "Country"
        ws.cell(next_row, 2).value = code
        ws.cell(next_row, 3).value = name
        ws.cell(next_row, 4).value = region
        next_row += 1

print(f"   {next_row - 2} entries (includes countries)")

# ========== 13. Country_PM_Assignments ==========
print("6. Building Country_PM_Assignments...")
ws = wb.create_sheet("Country_PM_Assignments")
headers = ['Country_Code', 'Country_Name', 'Region', 'Project_Manager', 'PM_Email', 'PM_Phone', 'Notes']
create_header(ws, 1, headers)

current_row = 2
for row in range(2, ws_regions.max_row + 1):
    code = ws_regions.cell(row, 1).value
    name = ws_regions.cell(row, 2).value
    region = ws_regions.cell(row, 3).value
    if code:
        ws.cell(current_row, 1).value = code
        ws.cell(current_row, 2).value = name
        ws.cell(current_row, 3).value = region
        ws.cell(current_row, 4).value = "TBD"
        current_row += 1

print(f"   {current_row - 2} countries")

# ========== 14. _SETUP Instructions ==========
print("7. Building _SETUP sheet...")
ws = wb.create_sheet("_SETUP")
ws['A1'] = "TRACKER V16 - DATA FOUNDATION"
ws['A1'].font = Font(size=16, bold=True)
ws['A3'] = "Phase 1 Complete: Data Foundation Built"
ws['A4'] = f"Build Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws['A5'] = "Version: 16.0"

ws['A7'] = "WHAT'S INCLUDED:"
ws['A7'].font = Font(bold=True)
ws['A8'] = "- Master_Projects: With Proposed -> Allocated -> Obligated -> Spent budget flow"
ws['A9'] = "- Country_Budgets: Budget breakdown by country with Proposed column"
ws['A10'] = "- Country_Regions: State Dept regions (AF, EAP, EUR, NEA, SCA, WHA) + subregions"
ws['A11'] = "- Supporting sheets: Milestones, Events, Risks, Decisions, Stakeholders"
ws['A12'] = "- Reference sheets: Deliverables, Audiences, Products, Technologies"
ws['A13'] = "- Config_Lists: All countries + dropdown values"
ws['A14'] = "- Country_PM_Assignments: PM by country (ready to fill in)"

ws['A16'] = "PHASE 2: BUILD MANUALLY IN EXCEL"
ws['A16'].font = Font(bold=True, size=12, color='FF0000')
ws['A17'] = "Create these sheets with your preferred design:"
ws['A18'] = "1. Portfolio_Dashboard - KPIs, charts, summary"
ws['A19'] = "2. Country_Dashboard - Country-specific views"
ws['A20'] = "3. Regional_Summary - Regional roll-ups"
ws['A21'] = "4. Spotlight_Executive - Executive summary"
ws['A22'] = "5. Spotlight_PMWorkspace - PM detailed view (Audiences/Tech -> Financial -> Deliverables -> Stakeholders)"
ws['A23'] = "6. Spotlight_Stakeholder - Stakeholder briefing"
ws['A24'] = "7. Calendar_Todo - Task calendar"

ws['A26'] = "FORMULA EXAMPLES:"
ws['A26'].font = Font(bold=True)
ws['A27'] = "Sum budget by project:"
ws['A28'] = "=SUMIF(Country_Budgets!$B:$B,A2,Country_Budgets!$F:$F)"
ws['A29'] = "Count countries:"
ws['A30'] = "=COUNTIF(Country_Budgets!$B:$B,A2)"
ws['A31'] = "Lookup project name:"
ws['A32'] = "=VLOOKUP(A2,Master_Projects!$A:$C,3,FALSE)"

ws['A34'] = "KEY FEATURES:"
ws['A34'].font = Font(bold=True)
ws['A35'] = "- NO Excel Tables (clean ranges only)"
ws['A36'] = "- NO merged cells in data sheets"
ws['A37'] = "- State Department region codes"
ws['A38'] = "- Budget progression: Proposed -> Allocated -> Obligated -> Spent"
ws['A39'] = "- ULO = Obligated - Spent (correct calculation)"
ws['A40'] = "- All countries synced across sheets"

# Save
print("\n8. Saving v16...")
wb.save('2025-10-26-Tracker-v16.xlsx')

print("\n" + "="*80)
print("BUILD COMPLETE!")
print("="*80)
print(f"Completed: {datetime.now().strftime('%H:%M:%S')}")
print("\nFile: 2025-10-26-Tracker-v16.xlsx")
print("\nData sheets built:")
print("  - Master_Projects (with budget formulas)")
print("  - Country_Budgets (with Proposed column)")
print("  - Country_Regions (State Dept regions)")
print("  - 9 supporting sheets")
print("  - Config_Lists (with countries)")
print("  - Country_PM_Assignments")
print("  - _SETUP (instructions)")
print("\nReady for Phase 2: Build dashboards/spotlights manually in Excel")
print("See _SETUP sheet for instructions and formula examples")
