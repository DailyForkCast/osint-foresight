import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.3.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v9.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING TRACKER V9 - EXCEL 365 COMPREHENSIVE OVERHAUL')
print('='*80)
print()
print('Implementing:')
print('  1. FILTER() functions for all Project_Spotlight sections')
print('  2. Target Audiences - dynamic filtering')
print('  3. Target Technologies - dynamic filtering')
print('  4. Key Deliverables - dynamic filtering')
print('  5. Stakeholders section - dynamic filtering')
print('  6. Visual enhancements')
print()
print('='*80)
print()

# Load workbook
print('Loading v8.3...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

# ============================================================================
# PROJECT SPOTLIGHT - COMPREHENSIVE OVERHAUL
# ============================================================================
print('='*80)
print('PROJECT SPOTLIGHT - Comprehensive Overhaul with FILTER()')
print('='*80)
print()

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']

    print('Section 1: TARGET AUDIENCES (Rows 6-14, Columns G-I)')
    print('-'*80)

    # Target Audiences - Use FILTER for the first audience, then spill down
    # G6 will be the anchor cell for FILTER
    ws['G6'].value = '=IFERROR(FILTER(T_Project_Audiences[Audience_Type],T_Project_Audiences[Project_ID]=$B$2),"")'
    ws['H6'].value = '=IFERROR(FILTER(T_Project_Audiences[Description],T_Project_Audiences[Project_ID]=$B$2),"")'
    ws['I6'].value = '=IFERROR(FILTER(T_Project_Audiences[Priority],T_Project_Audiences[Project_ID]=$B$2),"")'

    # Clear old formulas in other rows (G7-G14, H7-H14, I7-I14)
    # FILTER will automatically spill into these cells
    for row in range(7, 15):
        ws.cell(row, 7).value = None  # Clear G7-G14
        ws.cell(row, 8).value = None  # Clear H7-H14
        ws.cell(row, 9).value = None  # Clear I7-I14

    print('  [OK] G6: FILTER(Audiences by Project)')
    print('  [OK] H6: FILTER(Descriptions)')
    print('  [OK] I6: FILTER(Priorities)')
    print('  [NOTE] FILTER will auto-spill to rows 7-14')
    print()

    print('Section 2: TARGET TECHNOLOGIES (Rows 6-14, Columns J-K)')
    print('-'*80)

    # Target Technologies
    ws['J6'].value = '=IFERROR(FILTER(Project_Technologies[Technology],Project_Technologies[Project_ID]=$B$2),"")'
    ws['K6'].value = '=IFERROR(FILTER(Project_Technologies[Category],Project_Technologies[Project_ID]=$B$2),"")'

    # Clear old formulas
    for row in range(7, 15):
        ws.cell(row, 10).value = None  # Clear J7-J14
        ws.cell(row, 11).value = None  # Clear K7-K14

    print('  [OK] J6: FILTER(Technologies by Project)')
    print('  [OK] K6: FILTER(Categories)')
    print('  [NOTE] FILTER will auto-spill to rows 7-14')
    print()

    print('Section 3: KEY DELIVERABLES (Rows 18-27, Columns B-F)')
    print('-'*80)

    # Key Deliverables
    ws['B18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Deliverable_Name],T_Project_Deliverables[Project_ID]=$B$2),"")'
    ws['C18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Due_Date],T_Project_Deliverables[Project_ID]=$B$2),"")'
    ws['D18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Status],T_Project_Deliverables[Project_ID]=$B$2),"")'
    ws['E18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Owner],T_Project_Deliverables[Project_ID]=$B$2),"")'
    ws['F18'].value = '=IFERROR(FILTER(T_Project_Deliverables[Completion_Percent],T_Project_Deliverables[Project_ID]=$B$2),"")'

    # Clear old formulas
    for row in range(19, 28):
        for col in range(2, 7):  # B-F
            ws.cell(row, col).value = None

    print('  [OK] B18: FILTER(Deliverables by Project)')
    print('  [OK] C18: FILTER(Due Dates)')
    print('  [OK] D18: FILTER(Status)')
    print('  [OK] E18: FILTER(Owners)')
    print('  [OK] F18: FILTER(Completion %)')
    print('  [NOTE] FILTER will auto-spill to rows 19-27')
    print()

    print('Section 4: STAKEHOLDERS (Starting Row 24, Columns H-K)')
    print('-'*80)

    # Find stakeholders section - appears to be around row 24, column H
    # Let's add stakeholder filters

    # Check if there's a dedicated stakeholders area
    # For now, let's add it below deliverables

    stakeholder_start_row = 30  # Start below deliverables

    # Add header
    ws.cell(stakeholder_start_row - 1, 2).value = 'KEY STAKEHOLDERS'
    ws.cell(stakeholder_start_row - 1, 2).font = Font(bold=True, size=12)

    # Add column headers
    ws.cell(stakeholder_start_row, 2).value = 'Name'
    ws.cell(stakeholder_start_row, 3).value = 'Title'
    ws.cell(stakeholder_start_row, 4).value = 'Organization'
    ws.cell(stakeholder_start_row, 5).value = 'Email'
    ws.cell(stakeholder_start_row, 6).value = 'Type'

    for col in range(2, 7):
        ws.cell(stakeholder_start_row, col).font = Font(bold=True)

    # Stakeholder filters (if Project_IDs column exists)
    stakeholder_data_row = stakeholder_start_row + 1

    # Note: Stakeholders table has Project_IDs column (L) with potentially comma-separated values
    # For Excel 365, we can use FILTER with SEARCH or TEXTSPLIT
    ws.cell(stakeholder_data_row, 2).value = '=IFERROR(FILTER(Stakeholders[Name],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
    ws.cell(stakeholder_data_row, 3).value = '=IFERROR(FILTER(Stakeholders[Title],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
    ws.cell(stakeholder_data_row, 4).value = '=IFERROR(FILTER(Stakeholders[Organization],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
    ws.cell(stakeholder_data_row, 5).value = '=IFERROR(FILTER(Stakeholders[Email],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'
    ws.cell(stakeholder_data_row, 6).value = '=IFERROR(FILTER(Stakeholders[Stakeholder_Type],ISNUMBER(SEARCH($B$2,Stakeholders[Project_IDs]))),"")'

    # Clear rows below
    for row in range(stakeholder_data_row + 1, stakeholder_data_row + 10):
        for col in range(2, 7):
            ws.cell(row, col).value = None

    print(f'  [OK] Added Stakeholders section at row {stakeholder_start_row}')
    print('  [OK] FILTER with SEARCH (works with comma-separated Project_IDs)')
    print('  [NOTE] FILTER will auto-spill')
    print()

    # Visual enhancements
    print('Section 5: VISUAL ENHANCEMENTS')
    print('-'*80)

    # Add borders and styling to headers
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Style section headers
    section_headers = [
        (4, 7),   # TARGET AUDIENCES
        (4, 10),  # TARGET TECHNOLOGIES
        (16, 2),  # KEY DELIVERABLES (approximate)
        (stakeholder_start_row - 1, 2)  # KEY STAKEHOLDERS
    ]

    for row, col in section_headers:
        cell = ws.cell(row, col)
        if cell.value:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='left')

    print('  [OK] Applied styling to section headers')
    print('  [OK] Professional color scheme applied')

    print()
    print('[COMPLETE] Project_Spotlight overhauled with Excel 365 FILTER()')

else:
    print('  WARNING: Project_Spotlight sheet not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving v9...')
wb.save(output_file)
wb.close()

print('='*80)
print('EXCEL 365 COMPREHENSIVE OVERHAUL COMPLETE!')
print('='*80)
print()
print('Changes applied:')
print('  1. [OK] Target Audiences - FILTER() in G6, auto-spills to rows 7-14')
print('  2. [OK] Target Technologies - FILTER() in J6, auto-spills to rows 7-14')
print('  3. [OK] Key Deliverables - FILTER() in B18, auto-spills to rows 19-27')
print('  4. [OK] Key Stakeholders - FILTER() in row 31, with comma-separated ID support')
print('  5. [OK] Visual enhancements - Headers styled professionally')
print()
print('Excel 365 Features Used:')
print('  - FILTER() function (dynamic arrays)')
print('  - Auto-spilling (results spill into cells below automatically)')
print('  - SEARCH() with FILTER for comma-separated Project_IDs')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v9 in Excel 365')
print('  2. Go to Project_Spotlight')
print('  3. Select PRJ-001 in B2')
print('  4. Watch all sections auto-populate with filtered data!')
print()
print('NO Ctrl+Shift+Enter needed - FILTER() is dynamic!')
print()
print(f'Created: {timestamp}')
