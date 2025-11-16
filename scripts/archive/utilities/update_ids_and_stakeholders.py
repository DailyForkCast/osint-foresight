import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v2.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v3.xlsx'

print('='*80)
print('UPDATING: Decision_Log, Risk_Register IDs and Stakeholders')
print('='*80)
print()

# Load workbook
print('Loading workbook...')
wb = openpyxl.load_workbook(input_file, data_only=False)

# ============================================================================
# PART 1: Update Decision_Log IDs
# ============================================================================
print('='*80)
print('PART 1: Updating Decision_Log IDs')
print('='*80)

if 'Decision_Log' in wb.sheetnames:
    ws = wb['Decision_Log']
    updated_count = 0

    for row in range(2, ws.max_row + 1):
        decision_id = ws.cell(row, 1).value
        project_id = ws.cell(row, 3).value  # Column C: Unique_ID

        if decision_id and project_id:
            # Only update if not already in correct format
            if not str(decision_id).startswith('PRJ-'):
                # Extract DEC number (e.g., DEC-001 -> 001)
                if str(decision_id).startswith('DEC-'):
                    dec_num = str(decision_id).replace('DEC-', '')
                    new_id = f'{project_id}-DEC-{dec_num}'
                    ws.cell(row, 1).value = new_id
                    updated_count += 1
                    print(f'  Row {row}: {decision_id} -> {new_id}')

    print(f'COMPLETE: Updated {updated_count} decision IDs')
else:
    print('WARNING: Decision_Log sheet not found')

print()

# ============================================================================
# PART 2: Update Risk_Register IDs
# ============================================================================
print('='*80)
print('PART 2: Updating Risk_Register IDs')
print('='*80)

if 'Risk_Register' in wb.sheetnames:
    ws = wb['Risk_Register']
    updated_count = 0

    for row in range(2, ws.max_row + 1):
        risk_id = ws.cell(row, 1).value
        project_id = ws.cell(row, 3).value  # Column C: Unique_ID

        if risk_id and project_id:
            # Only update if not already in correct format
            if not str(risk_id).startswith('PRJ-'):
                # Extract RISK number (e.g., RISK-001 -> 001)
                if str(risk_id).startswith('RISK-'):
                    risk_num = str(risk_id).replace('RISK-', '')
                    new_id = f'{project_id}-RISK-{risk_num}'
                    ws.cell(row, 1).value = new_id
                    updated_count += 1
                    print(f'  Row {row}: {risk_id} -> {new_id}')

    print(f'COMPLETE: Updated {updated_count} risk IDs')
else:
    print('WARNING: Risk_Register sheet not found')

print()

# ============================================================================
# PART 3: Rebuild Stakeholders Sheet
# ============================================================================
print('='*80)
print('PART 3: Rebuilding Stakeholders Sheet (12 -> 22 columns)')
print('='*80)

if 'Stakeholders' in wb.sheetnames:
    ws = wb['Stakeholders']

    # Define the complete 22-column structure
    headers = [
        'Stakeholder_ID',      # A
        'Name',                # B
        'Title',               # C
        'Organization',        # D
        'Location_City',       # E
        'Location_Country',    # F
        'Time_Zone_Offset',    # G
        'Local_Time',          # H
        'Email',               # I
        'Phone',               # J
        'Stakeholder_Type',    # K
        'Project_IDs',         # L
        'Countries',           # M
        'Products',            # N
        'Region',              # O
        'Theme',               # P
        'Influence_Level',     # Q
        'Interest_Level',      # R
        'Contact_Frequency',   # S
        'Last_Contact',        # T
        'Next_Contact',        # U
        'Notes'                # V (22 columns)
    ]

    print(f'Expanding from {ws.max_column} to {len(headers)} columns')
    print()

    # Clear existing headers and data (we'll preserve STK-001 row if exists)
    old_data = []
    if ws.max_row > 1:
        # Save existing stakeholder IDs and names
        for row in range(2, ws.max_row + 1):
            stk_id = ws.cell(row, 1).value
            name = ws.cell(row, 2).value
            if stk_id or name:
                old_data.append({'id': stk_id, 'name': name})

    # Delete all rows except row 1
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Add new headers
    print('Adding 22 column headers...')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)

        # Format header: Bold, white text, dark blue background
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add formula for Local_Time in H2
    ws['H2'] = '=IF(G2<>"",NOW()+(G2/24),"")'

    # Set column widths
    column_widths = {
        'A': 15,  # Stakeholder_ID
        'B': 25,  # Name
        'C': 25,  # Title
        'D': 30,  # Organization
        'E': 20,  # Location_City
        'F': 20,  # Location_Country
        'G': 12,  # Time_Zone_Offset
        'H': 20,  # Local_Time
        'I': 30,  # Email
        'J': 18,  # Phone
        'K': 20,  # Stakeholder_Type
        'L': 20,  # Project_IDs
        'M': 20,  # Countries
        'N': 20,  # Products
        'O': 10,  # Region
        'P': 25,  # Theme
        'Q': 15,  # Influence_Level
        'R': 15,  # Interest_Level
        'S': 18,  # Contact_Frequency
        'T': 15,  # Last_Contact
        'U': 15,  # Next_Contact
        'V': 40   # Notes
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Add back old data to row 2 if it existed
    if old_data:
        ws['A2'] = old_data[0]['id']
        ws['B2'] = old_data[0]['name']

    print('New structure:')
    for idx, header in enumerate(headers, start=1):
        print(f'  {get_column_letter(idx)}: {header}')

    print()
    print('Special features:')
    print('  - Column H (Local_Time): Auto-calculates from Time_Zone_Offset')
    print('  - Formula in H2: =IF(G2<>"",NOW()+(G2/24),"")')
    print('  - Copy formula down for additional stakeholders')

    print(f'COMPLETE: Stakeholders expanded to 22 columns')
else:
    print('WARNING: Stakeholders sheet not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving workbook as v3...')
wb.save(output_file)
wb.close()

print('='*80)
print('ALL UPDATES COMPLETE!')
print('='*80)
print()
print('Summary of changes:')
print('  1. Decision_Log: IDs updated to PRJ-XXX-DEC-XXX format')
print('  2. Risk_Register: IDs updated to PRJ-XXX-RISK-XXX format')
print('  3. Stakeholders: Expanded from 12 to 22 columns')
print()
print(f'Output file: {output_file}')
print()
print('Stakeholder Column Mapping:')
print('  A: Stakeholder_ID        J: Phone                 S: Contact_Frequency')
print('  B: Name                  K: Stakeholder_Type      T: Last_Contact')
print('  C: Title                 L: Project_IDs           U: Next_Contact')
print('  D: Organization          M: Countries             V: Notes')
print('  E: Location_City         N: Products')
print('  F: Location_Country      O: Region')
print('  G: Time_Zone_Offset      P: Theme')
print('  H: Local_Time (formula)  Q: Influence_Level')
print('  I: Email                 R: Interest_Level')
