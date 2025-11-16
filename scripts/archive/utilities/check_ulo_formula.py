"""
Check ULO% formula in Country_Dashboard
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v43.xlsx')
ws_country = wb['Country_Dashboard']

print("="*80)
print("CHECKING ULO% LOCATION")
print("="*80)

# Check row 6 columns F, G, H
print("\nRow 6 around F6:")
for col in range(6, 10):
    cell = ws_country.cell(6, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter}6:")
    print(f"    Value: {cell.value}")
    print(f"    Type: {cell.data_type}")
    if cell.data_type == 'f':
        print(f"    Formula: {cell.value}")

# Check Project ID formulas
print("\n" + "="*80)
print("PROJECT ID FORMULAS")
print("="*80)

print("\nA12 (first Project ID):")
a12 = ws_country['A12']
print(f"  Value: {a12.value}")
print(f"  Type: {a12.data_type}")

print("\nA13 (second Project ID):")
a13 = ws_country['A13']
print(f"  Value: {a13.value}")
print(f"  Type: {a13.data_type}")

wb.close()

print("\n" + "="*80)
print("DONE")
print("="*80)
