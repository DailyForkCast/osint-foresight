"""
Check what's wrong with Portfolio_Dashboard in v35
"""
import openpyxl

print("="*80)
print("CHECKING PORTFOLIO_DASHBOARD IN V35")
print("="*80)

# Load v35
print("\nLoading v35...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v35.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("CHECKING STRUCTURE")
print("="*80)

print("\nRows 1-10:")
for row in range(1, 11):
    print(f"\nRow {row}:")
    for col in range(1, 17):  # A-P
        cell = ws_portfolio.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)

        if cell.value:
            if cell.data_type == 'f':
                formula = str(cell.value)[:60]
                print(f"  {col_letter}: FORMULA - {formula}...")
            else:
                clean_value = str(cell.value).encode('ascii', 'ignore').decode('ascii')[:40]
                print(f"  {col_letter}: {clean_value}")

print("\n" + "="*80)
print("ROW 10 HEADERS (PROJECT LIST)")
print("="*80)

print("\nHeaders:")
for col in range(1, 17):
    header = ws_portfolio.cell(10, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        clean_header = str(header).encode('ascii', 'ignore').decode('ascii')
        print(f"  {col_letter}10: {clean_header}")

print("\n" + "="*80)
print("ROW 11 FORMULAS (FIRST DATA ROW)")
print("="*80)

print("\nChecking row 11:")
for col in range(1, 17):
    cell = ws_portfolio.cell(11, col)
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.value:
        if cell.data_type == 'f':
            print(f"  {col_letter}11: HAS FORMULA")
        else:
            print(f"  {col_letter}11: {cell.value}")

print("\nDone!")
