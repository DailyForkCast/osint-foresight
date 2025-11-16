"""
Check v12 for Award Number and NCE columns
"""
import openpyxl

wb = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=True)

print("="*80)
print("V12 MASTER_PROJECTS - LOOKING FOR AWARD NUMBER & NCE")
print("="*80)

ws = wb['Master_Projects']

print("\nAll column headers:")
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    if header:
        # Check if Award or NCE related
        if 'AWARD' in str(header).upper() or 'NCE' in str(header).upper():
            print(f"  >>> {col_letter}: {header} <<<")
        else:
            print(f"  {col_letter}: {header}")

# Look for NCE specifically
print("\n" + "="*80)
print("NCE-RELATED COLUMNS")
print("="*80)

nce_columns = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header and 'NCE' in str(header).upper():
        col_letter = openpyxl.utils.get_column_letter(col)
        nce_columns.append((col_letter, header))

if nce_columns:
    print("\nFound NCE columns:")
    for col_letter, header in nce_columns:
        print(f"  {col_letter}: {header}")
        # Show sample data
        sample = ws.cell(2, openpyxl.utils.column_index_from_string(col_letter)).value
        print(f"      Sample value: {sample}")
else:
    print("\nNo NCE columns found in Master_Projects")

# Check for Award Number
print("\n" + "="*80)
print("AWARD NUMBER COLUMNS")
print("="*80)

award_columns = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header and 'AWARD' in str(header).upper():
        col_letter = openpyxl.utils.get_column_letter(col)
        award_columns.append((col_letter, header))

if award_columns:
    print("\nFound Award columns:")
    for col_letter, header in award_columns:
        print(f"  {col_letter}: {header}")
        sample = ws.cell(2, openpyxl.utils.column_index_from_string(col_letter)).value
        print(f"      Sample value: {sample}")
else:
    print("\nNo Award columns found in Master_Projects")

# Check Control sheet for NCE info
print("\n" + "="*80)
print("CHECKING OTHER SHEETS FOR NCE/AWARD")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in range(1, min(30, ws.max_column + 1)):
        header = ws.cell(1, col).value
        if header and ('NCE' in str(header).upper() or 'AWARD' in str(header).upper()):
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"\n{sheet_name} - {col_letter}: {header}")
