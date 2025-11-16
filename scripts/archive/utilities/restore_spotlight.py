import openpyxl
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 70)
print("RESTORING PROJECT SPOTLIGHT - UNDOING CHANGES")
print("=" * 70)

# Load the workbook
print("\n[1/2] Loading workbook...")
wb = load_workbook('C:/Projects/OSINT - Foresight/2025-10-05-Tracker-FINAL.xlsx')
ws = wb['Project_Spotlight']
print("✓ Workbook loaded")

print("\n[2/2] Cleaning up Project Spotlight...")

# Delete all the hardcoded content we added (rows 14 onwards)
max_row = ws.max_row
if max_row > 13:
    ws.delete_rows(14, max_row - 13)
    print(f"✓ Removed {max_row - 13} rows of hardcoded content")

# Keep it clean - just the header and basic project info
print("✓ Project Spotlight cleaned - ready for dynamic setup")

# Save
output_file = 'C:/Projects/OSINT - Foresight/2025-10-05-Tracker-CLEAN.xlsx'
print(f"\n{'='*70}")
print("Saving cleaned workbook...")
wb.save(output_file)
print(f"✓ Saved: {output_file}")

print(f"\n{'='*70}")
print("CLEANUP COMPLETE!")
print(f"{'='*70}")
print("\nProject Spotlight now has:")
print("  Rows 1-2:   Header & Project Selector")
print("  Rows 3-13:  Project Summary & Details")
print("  Row 14+:    EMPTY - ready for dynamic data")
print(f"\nFile: {output_file}")
print("\nNEXT STEPS:")
print("  We need to decide HOW to store and display:")
print("    • Target Technologies (per project)")
print("    • Target Audiences (per project)")
print("    • Key Deliverables (per project)")
print("\n  Options:")
print("    A) Add columns to Master_Projects sheet")
print("    B) Create separate data sheets that link to projects")
print("    C) Use existing Project_Deliverables, Project_Audiences")
