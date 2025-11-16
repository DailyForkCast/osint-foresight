"""
Check for missing formulas in v18
"""
import openpyxl

print("="*80)
print("MISSING FORMULAS CHECK")
print("="*80)

wb_v12 = openpyxl.load_workbook('2025-10-26-Tracker-v12.xlsx', data_only=False)

print("\n1. MASTER_PROJECTS - V12 FORMULAS")
print("="*80)

ws = wb_v12['Master_Projects']
print("\nFormulas found in row 2:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.data_type == 'f':
        formula = str(cell.value)
        # Clean up formula for display
        if len(formula) > 80:
            formula = formula[:77] + "..."
        print(f"  {col_letter:3} {str(header)[:25]:25} {formula}")

print("\n2. COUNTRY_BUDGETS - V12 FORMULAS")
print("="*80)

ws = wb_v12['Country_Budgets']
print("\nFormulas found in row 2:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)

    if cell.data_type == 'f':
        formula = str(cell.value)
        if len(formula) > 80:
            formula = formula[:77] + "..."
        print(f"  {col_letter:3} {str(header)[:25]:25} {formula}")

print("\n3. KEY MISSING FORMULAS TO ADD TO V18")
print("="*80)

print("\nMASTER_PROJECTS:")
print("  - Column I (Days_Remaining): Calculate days until project end")
print("  - Column S (Countries): List of countries in project")
print("  - Column T (Country_Count): Count of countries")

print("\nCOUNTRY_BUDGETS:")
print("  - Check if any formulas are missing")

print("\n4. V12 MASTER_PROJECTS COLUMN MAP")
print("="*80)

ws = wb_v12['Master_Projects']
print("\nAll columns in v12:")
for col in range(1, min(35, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter:3} {header}")

print("\n5. V12 COUNTRY_BUDGETS COLUMN MAP")
print("="*80)

ws = wb_v12['Country_Budgets']
print("\nAll columns in v12:")
for col in range(1, min(20, ws.max_column + 1)):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter:3} {header}")
