import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v11.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v12.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING THREE PROJECT_SPOTLIGHT VERSIONS')
print('='*80)
print()

wb = openpyxl.load_workbook(input_file, data_only=False)

# Styles
header_font = Font(bold=True, size=16, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
label_font = Font(bold=True, size=10)
metric_font = Font(bold=True, size=14)
column_header_font = Font(bold=True, size=9)
column_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Remove old Project_Spotlight
if 'Project_Spotlight' in wb.sheetnames:
    print('Removing old Project_Spotlight...')
    del wb['Project_Spotlight']

# ============================================================================
# VERSION 1: EXECUTIVE SUMMARY
# ============================================================================

print('='*80)
print('VERSION 1: EXECUTIVE SUMMARY')
print('='*80)
print()

ws1 = wb.create_sheet('Spotlight_Executive', 6)

print('Building Executive Summary layout...')

# Title and selector (Row 1-2)
ws1['A1'] = 'PROJECT EXECUTIVE SUMMARY'
ws1['A1'].font = header_font
ws1['A1'].fill = header_fill
ws1.merge_cells('A1:H1')

ws1['A2'] = 'Project ID:'
ws1['A2'].font = label_font
ws1['B2'] = 'PRJ-001'
ws1.merge_cells('B2:C2')

# Add dropdown
project_dv = DataValidation(
    type='list',
    formula1='Master_Projects!$B$2:$B$100',
    allow_blank=False,
    showDropDown=True
)
project_dv.add('B2')
ws1.add_data_validation(project_dv)

# Project basics (Row 4-9)
ws1['A4'] = 'PROJECT OVERVIEW'
ws1['A4'].font = section_font
ws1['A4'].fill = section_fill
ws1.merge_cells('A4:H4')

# Two-column layout
basics_labels = [
    ('A5', 'Project Name:', 'B5', '=IFERROR(INDEX(Master_Projects!$C:$C,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B5:D5'),
    ('E5', 'Status:', 'F5', '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'F5:H5'),
    ('A6', 'Priority:', 'B6', '=IFERROR(INDEX(Master_Projects!$E:$E,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B6:D6'),
    ('E6', 'Progress:', 'F6', '=IFERROR(INDEX(Master_Projects!$F:$F,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'F6:H6'),
    ('A7', 'Start Date:', 'B7', '=IFERROR(INDEX(Master_Projects!$G:$G,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B7:D7'),
    ('E7', 'End Date:', 'F7', '=IFERROR(INDEX(Master_Projects!$H:$H,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'F7:H7'),
    ('A8', 'Days Remaining:', 'B8', '=IFERROR(INDEX(Master_Projects!$I:$I,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B8:D8'),
    ('E8', 'Countries:', 'F8', '=IFERROR(INDEX(Master_Projects!$Q:$Q,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'F8:H8'),
]

for label_cell, label_text, value_cell, formula, merge_range in basics_labels:
    ws1[label_cell] = label_text
    ws1[label_cell].font = label_font
    ws1[value_cell] = formula
    if merge_range:
        ws1.merge_cells(merge_range)

# Summary (Row 9)
ws1['A9'] = 'Summary:'
ws1['A9'].font = label_font
ws1['B9'] = '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws1.merge_cells('B9:H9')
ws1['B9'].alignment = Alignment(wrap_text=True, vertical='top')
ws1.row_dimensions[9].height = 40

# Financials (Row 11-13) - Compact metrics
ws1['A11'] = 'FINANCIAL SUMMARY'
ws1['A11'].font = section_font
ws1['A11'].fill = section_fill
ws1.merge_cells('A11:H11')

ws1['A12'] = 'Allocated'
ws1['A12'].font = column_header_font
ws1['A12'].fill = column_header_fill
ws1['A12'].alignment = Alignment(horizontal='center')

ws1['C12'] = 'Obligated'
ws1['C12'].font = column_header_font
ws1['C12'].fill = column_header_fill
ws1['C12'].alignment = Alignment(horizontal='center')

ws1['E12'] = 'Spent'
ws1['E12'].font = column_header_font
ws1['E12'].fill = column_header_fill
ws1['E12'].alignment = Alignment(horizontal='center')

ws1['G12'] = 'ULO'
ws1['G12'].font = column_header_font
ws1['G12'].fill = column_header_fill
ws1['G12'].alignment = Alignment(horizontal='center')

ws1['A13'] = '=IFERROR(INDEX(Master_Projects!$M:$M,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws1['A13'].number_format = '$#,##0'
ws1['A13'].font = metric_font
ws1['A13'].alignment = Alignment(horizontal='center')

ws1['C13'] = '=IFERROR(INDEX(Master_Projects!$N:$N,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws1['C13'].number_format = '$#,##0'
ws1['C13'].font = metric_font
ws1['C13'].alignment = Alignment(horizontal='center')

ws1['E13'] = '=IFERROR(INDEX(Master_Projects!$T:$T,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws1['E13'].number_format = '$#,##0'
ws1['E13'].font = metric_font
ws1['E13'].alignment = Alignment(horizontal='center')

ws1['G13'] = '=IFERROR(INDEX(Master_Projects!$O:$O,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws1['G13'].number_format = '$#,##0'
ws1['G13'].font = metric_font
ws1['G13'].alignment = Alignment(horizontal='center')

# Key Deliverables (Row 15-21) - Compact list
ws1['A15'] = 'KEY DELIVERABLES'
ws1['A15'].font = section_font
ws1['A15'].fill = section_fill
ws1.merge_cells('A15:H15')

# Headers
deliverable_headers = [('A16', 'Deliverable'), ('D16', 'Due Date'), ('F16', 'Status'), ('G16', 'Owner')]
for cell, text in deliverable_headers:
    ws1[cell].value = text
    ws1[cell].font = column_header_font
    ws1[cell].fill = column_header_fill
    ws1[cell].border = thin_border

# 5 deliverable rows (compact for executive view)
for i in range(5):
    row = 17 + i
    # Using simple INDEX formulas (not arrays) - will need manual setup or simple lookups
    ws1[f'A{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Deliverable_Name],{i+1}),"")'
    ws1[f'D{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Due_Date],{i+1}),"")'
    ws1[f'F{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Status],{i+1}),"")'
    ws1[f'G{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Owner],{i+1}),"")'
    ws1.merge_cells(f'A{row}:C{row}')

# Audiences & Technologies (Row 23-26)
ws1['A23'] = 'TARGET AUDIENCES & TECHNOLOGIES'
ws1['A23'].font = section_font
ws1['A23'].fill = section_fill
ws1.merge_cells('A23:H23')

ws1['A24'] = 'Audiences:'
ws1['A24'].font = label_font
ws1['B24'] = '=IFERROR(INDEX(T_Project_Audiences[Audience_Type],1),"") & IF(INDEX(T_Project_Audiences[Audience_Type],2)<>"", ", " & INDEX(T_Project_Audiences[Audience_Type],2), "")'
ws1.merge_cells('B24:H24')

ws1['A25'] = 'Technologies:'
ws1['A25'].font = label_font
ws1['B25'] = '=IFERROR(INDEX(Project_Technologies[Technology],1),"") & IF(INDEX(Project_Technologies[Technology],2)<>"", ", " & INDEX(Project_Technologies[Technology],2), "")'
ws1.merge_cells('B25:H25')

# Key Stakeholders (Row 27-30)
ws1['A27'] = 'KEY STAKEHOLDERS'
ws1['A27'].font = section_font
ws1['A27'].fill = section_fill
ws1.merge_cells('A27:H27')

stakeholder_headers = [('A28', 'Name'), ('C28', 'Title'), ('E28', 'Organization'), ('G28', 'Email')]
for cell, text in stakeholder_headers:
    ws1[cell].value = text
    ws1[cell].font = column_header_font
    ws1[cell].fill = column_header_fill
    ws1[cell].border = thin_border

# 3 stakeholder rows (compact)
for i in range(3):
    row = 29 + i
    ws1[f'A{row}'] = f'=IFERROR(INDEX(Stakeholders[Name],{i+1}),"")'
    ws1[f'C{row}'] = f'=IFERROR(INDEX(Stakeholders[Title],{i+1}),"")'
    ws1[f'E{row}'] = f'=IFERROR(INDEX(Stakeholders[Organization],{i+1}),"")'
    ws1[f'G{row}'] = f'=IFERROR(INDEX(Stakeholders[Email],{i+1}),"")'
    ws1.merge_cells(f'A{row}:B{row}')
    ws1.merge_cells(f'C{row}:D{row}')
    ws1.merge_cells(f'E{row}:F{row}')
    ws1.merge_cells(f'G{row}:H{row}')

# Column widths
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 15
ws1.column_dimensions['G'].width = 15
ws1.column_dimensions['H'].width = 15

print('[OK] Executive Summary created')
print()

# ============================================================================
# VERSION 2: PM WORKSPACE
# ============================================================================

print('='*80)
print('VERSION 2: PM WORKSPACE')
print('='*80)
print()

ws2 = wb.create_sheet('Spotlight_PMWorkspace', 7)

print('Building PM Workspace layout...')

# Similar structure but with MORE rows for each section
# Title
ws2['A1'] = 'PROJECT MANAGER WORKSPACE'
ws2['A1'].font = header_font
ws2['A1'].fill = header_fill
ws2.merge_cells('A1:J1')

ws2['A2'] = 'Project ID:'
ws2['A2'].font = label_font
ws2['B2'] = 'PRJ-001'
ws2.merge_cells('B2:C2')

project_dv2 = DataValidation(
    type='list',
    formula1='Master_Projects!$B$2:$B$100',
    allow_blank=False,
    showDropDown=True
)
project_dv2.add('B2')
ws2.add_data_validation(project_dv2)

# Project Details - Expanded (Row 4-12)
ws2['A4'] = 'PROJECT DETAILS'
ws2['A4'].font = section_font
ws2['A4'].fill = section_fill
ws2.merge_cells('A4:J4')

details_labels = [
    ('A5', 'Project Name:', 'B5', '=IFERROR(INDEX(Master_Projects!$C:$C,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B5:E5'),
    ('F5', 'Status:', 'G5', '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G5:J5'),
    ('A6', 'Priority:', 'B6', '=IFERROR(INDEX(Master_Projects!$E:$E,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B6:E6'),
    ('F6', 'Progress:', 'G6', '=IFERROR(INDEX(Master_Projects!$F:$F,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G6:J6'),
    ('A7', 'Start Date:', 'B7', '=IFERROR(INDEX(Master_Projects!$G:$G,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B7:E7'),
    ('F7', 'End Date:', 'G7', '=IFERROR(INDEX(Master_Projects!$H:$H,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G7:J7'),
    ('A8', 'Days Remaining:', 'B8', '=IFERROR(INDEX(Master_Projects!$I:$I,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B8:E8'),
    ('F8', 'Implementer:', 'G8', '=IFERROR(INDEX(Master_Projects!$J:$J,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G8:J8'),
    ('A9', 'POC:', 'B9', '=IFERROR(INDEX(Master_Projects!$K:$K,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B9:E9'),
    ('F9', 'POC Email:', 'G9', '=IFERROR(INDEX(Master_Projects!$L:$L,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G9:J9'),
    ('A10', 'Countries:', 'B10', '=IFERROR(INDEX(Master_Projects!$Q:$Q,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'B10:E10'),
    ('F10', 'Country Count:', 'G10', '=IFERROR(INDEX(Master_Projects!$R:$R,MATCH($B$2,Master_Projects!$A:$A,0)),"")', 'G10:J10'),
]

for label_cell, label_text, value_cell, formula, merge_range in details_labels:
    ws2[label_cell] = label_text
    ws2[label_cell].font = label_font
    ws2[value_cell] = formula
    if merge_range:
        ws2.merge_cells(merge_range)

ws2['A11'] = 'Summary:'
ws2['A11'].font = label_font
ws2['B11'] = '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws2.merge_cells('B11:J11')
ws2['B11'].alignment = Alignment(wrap_text=True, vertical='top')
ws2.row_dimensions[11].height = 50

# Deliverables - Expanded (10 rows)
ws2['A13'] = 'KEY DELIVERABLES'
ws2['A13'].font = section_font
ws2['A13'].fill = section_fill
ws2.merge_cells('A13:J13')

deliverable_headers2 = [('A14', 'Deliverable'), ('E14', 'Due Date'), ('G14', 'Status'), ('H14', 'Owner'), ('I14', 'Complete %')]
for cell, text in deliverable_headers2:
    ws2[cell].value = text
    ws2[cell].font = column_header_font
    ws2[cell].fill = column_header_fill
    ws2[cell].border = thin_border

for i in range(10):
    row = 15 + i
    ws2[f'A{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Deliverable_Name],{i+1}),"")'
    ws2[f'E{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Due_Date],{i+1}),"")'
    ws2[f'G{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Status],{i+1}),"")'
    ws2[f'H{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Owner],{i+1}),"")'
    ws2[f'I{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Completion_Percent],{i+1}),"")'
    ws2.merge_cells(f'A{row}:D{row}')
    ws2.merge_cells(f'E{row}:F{row}')

# Audiences & Technologies - Expanded lists
ws2['A26'] = 'TARGET AUDIENCES'
ws2['A26'].font = section_font
ws2['A26'].fill = section_fill
ws2.merge_cells('A26:E26')

ws2['F26'] = 'TECHNOLOGIES'
ws2['F26'].font = section_font
ws2['F26'].fill = section_fill
ws2.merge_cells('F26:J26')

# 5 rows each
for i in range(5):
    row = 27 + i
    ws2[f'A{row}'] = f'=IFERROR(INDEX(T_Project_Audiences[Audience_Type],{i+1}),"")'
    ws2[f'F{row}'] = f'=IFERROR(INDEX(Project_Technologies[Technology],{i+1}),"")'
    ws2.merge_cells(f'A{row}:E{row}')
    ws2.merge_cells(f'F{row}:J{row}')

# Stakeholders - Expanded (8 rows)
ws2['A33'] = 'KEY STAKEHOLDERS'
ws2['A33'].font = section_font
ws2['A33'].fill = section_fill
ws2.merge_cells('A33:J33')

stakeholder_headers2 = [('A34', 'Name'), ('C34', 'Title'), ('E34', 'Organization'), ('G34', 'Email')]
for cell, text in stakeholder_headers2:
    ws2[cell].value = text
    ws2[cell].font = column_header_font
    ws2[cell].fill = column_header_fill
    ws2[cell].border = thin_border

for i in range(8):
    row = 35 + i
    ws2[f'A{row}'] = f'=IFERROR(INDEX(Stakeholders[Name],{i+1}),"")'
    ws2[f'C{row}'] = f'=IFERROR(INDEX(Stakeholders[Title],{i+1}),"")'
    ws2[f'E{row}'] = f'=IFERROR(INDEX(Stakeholders[Organization],{i+1}),"")'
    ws2[f'G{row}'] = f'=IFERROR(INDEX(Stakeholders[Email],{i+1}),"")'
    ws2.merge_cells(f'A{row}:B{row}')
    ws2.merge_cells(f'C{row}:D{row}')
    ws2.merge_cells(f'E{row}:F{row}')
    ws2.merge_cells(f'G{row}:J{row}')

# Financials at bottom
ws2['A44'] = 'FINANCIAL DETAILS'
ws2['A44'].font = section_font
ws2['A44'].fill = section_fill
ws2.merge_cells('A44:J44')

ws2['A45'] = 'Total Allocated:'
ws2['A45'].font = label_font
ws2['B45'] = '=IFERROR(INDEX(Master_Projects!$M:$M,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws2['B45'].number_format = '$#,##0'

ws2['D45'] = 'Total Obligated:'
ws2['D45'].font = label_font
ws2['E45'] = '=IFERROR(INDEX(Master_Projects!$N:$N,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws2['E45'].number_format = '$#,##0'

ws2['G45'] = 'Total Spent:'
ws2['G45'].font = label_font
ws2['H45'] = '=IFERROR(INDEX(Master_Projects!$T:$T,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws2['H45'].number_format = '$#,##0'

ws2['A46'] = 'ULO:'
ws2['A46'].font = label_font
ws2['B46'] = '=IFERROR(INDEX(Master_Projects!$O:$O,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws2['B46'].number_format = '$#,##0'

ws2['D46'] = 'ULO %:'
ws2['D46'].font = label_font
ws2['E46'] = '=IFERROR(INDEX(Master_Projects!$P:$P,MATCH($B$2,Master_Projects!$A:$A,0)),0)'
ws2['E46'].number_format = '0.0%'

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws2.column_dimensions[col].width = 15

print('[OK] PM Workspace created')
print()

# ============================================================================
# VERSION 3: STAKEHOLDER BRIEFING
# ============================================================================

print('='*80)
print('VERSION 3: STAKEHOLDER BRIEFING')
print('='*80)
print()

ws3 = wb.create_sheet('Spotlight_Stakeholder', 8)

print('Building Stakeholder Briefing layout...')

# Title
ws3['A1'] = 'PROJECT BRIEFING'
ws3['A1'].font = Font(bold=True, size=18, color='FFFFFF')
ws3['A1'].fill = header_fill
ws3.merge_cells('A1:H1')
ws3.row_dimensions[1].height = 35

# Selector
ws3['A2'] = 'Project:'
ws3['A2'].font = label_font
ws3['B2'] = 'PRJ-001'
ws3.merge_cells('B2:C2')

project_dv3 = DataValidation(
    type='list',
    formula1='Master_Projects!$B$2:$B$100',
    allow_blank=False,
    showDropDown=True
)
project_dv3.add('B2')
ws3.add_data_validation(project_dv3)

# Project name and status prominently
ws3['A4'] = '=IFERROR(INDEX(Master_Projects!$C:$C,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws3['A4'].font = Font(bold=True, size=14)
ws3.merge_cells('A4:H4')

ws3['A5'] = '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws3['A5'].alignment = Alignment(wrap_text=True)
ws3.merge_cells('A5:H5')
ws3.row_dimensions[5].height = 50

# Key metrics - Visual boxes
ws3['A7'] = 'STATUS'
ws3['A7'].font = section_font
ws3['A7'].fill = section_fill
ws3['A7'].alignment = Alignment(horizontal='center')
ws3.merge_cells('A7:B7')

ws3['C7'] = 'PROGRESS'
ws3['C7'].font = section_font
ws3['C7'].fill = section_fill
ws3['C7'].alignment = Alignment(horizontal='center')
ws3.merge_cells('C7:D7')

ws3['E7'] = 'PRIORITY'
ws3['E7'].font = section_font
ws3['E7'].fill = section_fill
ws3['E7'].alignment = Alignment(horizontal='center')
ws3.merge_cells('E7:F7')

ws3['G7'] = 'TIMELINE'
ws3['G7'].font = section_font
ws3['G7'].fill = section_fill
ws3['G7'].alignment = Alignment(horizontal='center')
ws3.merge_cells('G7:H7')

ws3['A8'] = '=IFERROR(INDEX(Master_Projects!$D:$D,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws3['A8'].font = metric_font
ws3['A8'].alignment = Alignment(horizontal='center')
ws3.merge_cells('A8:B8')

ws3['C8'] = '=IFERROR(INDEX(Master_Projects!$F:$F,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws3['C8'].font = metric_font
ws3['C8'].alignment = Alignment(horizontal='center')
ws3.merge_cells('C8:D8')

ws3['E8'] = '=IFERROR(INDEX(Master_Projects!$E:$E,MATCH($B$2,Master_Projects!$A:$A,0)),"")'
ws3['E8'].font = metric_font
ws3['E8'].alignment = Alignment(horizontal='center')
ws3.merge_cells('E8:F8')

ws3['G8'] = '=IFERROR(TEXT(INDEX(Master_Projects!$G:$G,MATCH($B$2,Master_Projects!$A:$A,0)),"mmm dd") & " - " & TEXT(INDEX(Master_Projects!$H:$H,MATCH($B$2,Master_Projects!$A:$A,0)),"mmm dd, yyyy"),"")'
ws3['G8'].font = Font(bold=True, size=9)
ws3['G8'].alignment = Alignment(horizontal='center')
ws3.merge_cells('G8:H8')

# Deliverables - Clean presentation
ws3['A10'] = 'KEY MILESTONES & DELIVERABLES'
ws3['A10'].font = section_font
ws3['A10'].fill = section_fill
ws3.merge_cells('A10:H10')

deliverable_headers3 = [('A11', 'Deliverable'), ('E11', 'Target Date'), ('G11', 'Status')]
for cell, text in deliverable_headers3:
    ws3[cell].value = text
    ws3[cell].font = column_header_font
    ws3[cell].fill = column_header_fill
    ws3[cell].border = thin_border

for i in range(6):
    row = 12 + i
    ws3[f'A{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Deliverable_Name],{i+1}),"")'
    ws3[f'E{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Due_Date],{i+1}),"")'
    ws3[f'G{row}'] = f'=IFERROR(INDEX(T_Project_Deliverables[Status],{i+1}),"")'
    ws3.merge_cells(f'A{row}:D{row}')
    ws3.merge_cells(f'E{row}:F{row}')
    ws3.merge_cells(f'G{row}:H{row}')

# Target Audiences - Presentation style
ws3['A19'] = 'TARGET AUDIENCES'
ws3['A19'].font = section_font
ws3['A19'].fill = section_fill
ws3.merge_cells('A19:H19')

ws3['A20'] = '=IFERROR(INDEX(T_Project_Audiences[Audience_Type],1),"") & IF(INDEX(T_Project_Audiences[Audience_Type],2)<>"", " • " & INDEX(T_Project_Audiences[Audience_Type],2), "") & IF(INDEX(T_Project_Audiences[Audience_Type],3)<>"", " • " & INDEX(T_Project_Audiences[Audience_Type],3), "")'
ws3.merge_cells('A20:H20')
ws3['A20'].alignment = Alignment(wrap_text=True)

# Technologies - Presentation style
ws3['A22'] = 'KEY TECHNOLOGIES'
ws3['A22'].font = section_font
ws3['A22'].fill = section_fill
ws3.merge_cells('A22:H22')

ws3['A23'] = '=IFERROR(INDEX(Project_Technologies[Technology],1),"") & IF(INDEX(Project_Technologies[Technology],2)<>"", " • " & INDEX(Project_Technologies[Technology],2), "") & IF(INDEX(Project_Technologies[Technology],3)<>"", " • " & INDEX(Project_Technologies[Technology],3), "")'
ws3.merge_cells('A23:H23')
ws3['A23'].alignment = Alignment(wrap_text=True)

# Key contacts
ws3['A25'] = 'KEY CONTACTS'
ws3['A25'].font = section_font
ws3['A25'].fill = section_fill
ws3.merge_cells('A25:H25')

stakeholder_headers3 = [('A26', 'Name'), ('C26', 'Role'), ('E26', 'Contact')]
for cell, text in stakeholder_headers3:
    ws3[cell].value = text
    ws3[cell].font = column_header_font
    ws3[cell].fill = column_header_fill
    ws3[cell].border = thin_border

for i in range(4):
    row = 27 + i
    ws3[f'A{row}'] = f'=IFERROR(INDEX(Stakeholders[Name],{i+1}),"")'
    ws3[f'C{row}'] = f'=IFERROR(INDEX(Stakeholders[Title],{i+1}),"")'
    ws3[f'E{row}'] = f'=IFERROR(INDEX(Stakeholders[Email],{i+1}),"")'
    ws3.merge_cells(f'A{row}:B{row}')
    ws3.merge_cells(f'C{row}:D{row}')
    ws3.merge_cells(f'E{row}:H{row}')

# Financial summary - De-emphasized at bottom
ws3['A32'] = 'Budget: '
ws3['A32'].font = Font(size=9)
ws3['B32'] = '=IFERROR(TEXT(INDEX(Master_Projects!$M:$M,MATCH($B$2,Master_Projects!$A:$A,0)),"$#,##0") & " allocated","")'
ws3['B32'].font = Font(size=9)
ws3.merge_cells('B32:D32')

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws3.column_dimensions[col].width = 15

print('[OK] Stakeholder Briefing created')
print()

# ============================================================================
# SAVE
# ============================================================================

print('='*80)
print('Saving v12...')
print()

try:
    wb.save(output_file)
    wb.close()
    print('[OK] File saved successfully!')
except Exception as e:
    print(f'[ERROR] Failed to save: {e}')
    exit(1)

print()
print('='*80)
print('THREE PROJECT_SPOTLIGHT VERSIONS COMPLETE!')
print('='*80)
print()
print('Created:')
print('  1. [OK] Spotlight_Executive - Clean, 1-page executive summary')
print('  2. [OK] Spotlight_PMWorkspace - Detailed working document (10 deliverables, 8 stakeholders)')
print('  3. [OK] Spotlight_Stakeholder - Professional presentation for external stakeholders')
print()
print('All three versions include:')
print('  - Project ID dropdown selector')
print('  - Project info + summary')
print('  - Target audiences & technologies')
print('  - Key deliverables with status')
print('  - Stakeholder list')
print('  - Financial summary')
print()
print('Key differences:')
print('  Executive: Compact (fits 1 page), metrics-focused')
print('  PM Workspace: Expanded (10 deliverables, 8 stakeholders, detailed info)')
print('  Stakeholder: Visual, presentation-ready, external-facing')
print()
print('NOTE: Formulas use simple INDEX lookups')
print('  - May need to filter by Project_ID manually')
print('  - Or we can add FILTER formulas knowing they might need manual entry')
print()
print(f'Output: {output_file}')
print(f'Created: {timestamp}')
print()
print('Compare all three and let me know which one you prefer!')
print('Or tell me what to combine from each!')
