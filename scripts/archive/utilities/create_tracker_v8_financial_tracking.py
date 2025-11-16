import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v7.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v8.xlsx'

timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print('='*80)
print('CREATING TRACKER V8 - FINANCIAL TRACKING SYSTEM')
print('='*80)
print()
print('Implementing:')
print('  1. Project lifecycle tracking (Proposed/Active/Archived)')
print('  2. Country ownership (My Countries vs All Countries)')
print('  3. Proper ULO calculation (Obligated - Spent)')
print('  4. Spent amount tracking')
print('  5. Conditional formatting (Proposed=italic, Archived=gray)')
print()
print('='*80)
print()

# Load workbook
print('Loading v7...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print('Loaded.')
print()

total_changes = 0

# ============================================================================
# PART 1: Update Master_Projects
# ============================================================================
print('='*80)
print('PART 1: Master_Projects - Add Financial Tracking Columns')
print('='*80)
print()

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    # Find existing columns
    headers = {}
    for col in range(1, 30):
        header = ws.cell(1, col).value
        if header:
            headers[header] = col

    print(f'Existing columns: {len(headers)}')
    print()

    # Determine next available column
    next_col = max(headers.values()) + 1 if headers else 1

    # Add new columns
    new_columns = []

    # Proposed_Amount (for proposed projects)
    if 'Proposed_Amount' not in headers:
        ws.cell(1, next_col).value = 'Proposed_Amount'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Proposed_Amount'] = next_col
        new_columns.append('Proposed_Amount')
        next_col += 1

    # Total_Spent (NEW - sum of spent from MY countries only)
    if 'Total_Spent' not in headers:
        ws.cell(1, next_col).value = 'Total_Spent'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Total_Spent'] = next_col
        new_columns.append('Total_Spent')
        next_col += 1

    # Project_Manager (if not exists)
    if 'Project_Manager' not in headers:
        ws.cell(1, next_col).value = 'Project_Manager'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Project_Manager'] = next_col
        new_columns.append('Project_Manager')
        next_col += 1

    # Include_In_Calcs (formula - excludes Proposed and Archived)
    if 'Include_In_Calcs' not in headers:
        ws.cell(1, next_col).value = 'Include_In_Calcs'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Include_In_Calcs'] = next_col
        new_columns.append('Include_In_Calcs')
        next_col += 1

    # My_Countries_Count
    if 'My_Countries_Count' not in headers:
        ws.cell(1, next_col).value = 'My_Countries_Count'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['My_Countries_Count'] = next_col
        new_columns.append('My_Countries_Count')
        next_col += 1

    print(f'Added {len(new_columns)} new columns:')
    for col in new_columns:
        print(f'  - {col}')
    print()

    # Update formulas for existing projects
    status_col = headers.get('Status', 5)
    unique_id_col = headers.get('Unique_ID', 2)
    total_obligated_col = headers.get('Total_Obligated', 15)
    total_ulo_col = headers.get('Total_ULO', 16)
    ulo_percent_col = headers.get('ULO_Percent', 17)

    print('Updating formulas for existing projects...')

    for row in range(2, 10):  # Process existing project rows
        unique_id = ws.cell(row, unique_id_col).value
        if not unique_id:
            break

        # Include_In_Calcs formula (exclude Proposed and Archived)
        status_cell = openpyxl.utils.get_column_letter(status_col) + str(row)
        include_col = headers['Include_In_Calcs']
        ws.cell(row, include_col).value = f'=NOT(OR({status_cell}="Proposed",{status_cell}="Archived"))'

        # Total_Spent formula (sum from Country_Budgets where My_Country = TRUE)
        unique_cell = openpyxl.utils.get_column_letter(unique_id_col) + str(row)
        spent_col = headers['Total_Spent']
        ws.cell(row, spent_col).value = f'=SUMIFS(Country_Budgets!$F:$F,Country_Budgets!$B:$B,{unique_cell},Country_Budgets!$C:$C,TRUE)'

        # Update Total_Obligated to sum only MY countries
        ws.cell(row, total_obligated_col).value = f'=SUMIFS(Country_Budgets!$F:$F,Country_Budgets!$B:$B,{unique_cell},Country_Budgets!$C:$C,TRUE)'

        # Update Total_ULO = Obligated - Spent
        obligated_cell = openpyxl.utils.get_column_letter(total_obligated_col) + str(row)
        spent_cell = openpyxl.utils.get_column_letter(spent_col) + str(row)
        ws.cell(row, total_ulo_col).value = f'={obligated_cell}-{spent_cell}'

        # Update ULO_Percent = ULO / Obligated
        ulo_cell = openpyxl.utils.get_column_letter(total_ulo_col) + str(row)
        ws.cell(row, ulo_percent_col).value = f'=IF({obligated_cell}=0,0,{ulo_cell}/{obligated_cell})'

        # My_Countries_Count
        my_countries_col = headers['My_Countries_Count']
        ws.cell(row, my_countries_col).value = f'=COUNTIFS(Country_Budgets!$B:$B,{unique_cell},Country_Budgets!$C:$C,TRUE)'

        # Set demo Project_Manager
        pm_col = headers.get('Project_Manager')
        if pm_col and not ws.cell(row, pm_col).value:
            ws.cell(row, pm_col).value = 'Smith' if row == 2 else 'Jones'

    print('  [OK] Updated formulas for existing projects')
    total_changes += 1

else:
    print('  WARNING: Master_Projects sheet not found')

print()

# ============================================================================
# PART 2: Update Country_Budgets
# ============================================================================
print('='*80)
print('PART 2: Country_Budgets - Add My_Country and Spent_Amount')
print('='*80)
print()

if 'Country_Budgets' in wb.sheetnames:
    ws = wb['Country_Budgets']

    # Find existing columns
    headers = {}
    for col in range(1, 15):
        header = ws.cell(1, col).value
        if header:
            headers[header] = col

    print(f'Existing columns: {len(headers)}')
    print()

    # Insert My_Country as Column C (before Country_Code)
    if 'My_Country' not in headers:
        ws.insert_cols(3)  # Insert before current column C (Country_Code)
        ws.cell(1, 3).value = 'My_Country'
        ws.cell(1, 3).font = Font(bold=True)
        print('  Added My_Country column (Column C)')

        # Update headers dict after insertion
        headers = {}
        for col in range(1, 15):
            header = ws.cell(1, col).value
            if header:
                headers[header] = col

    # Add Spent_Amount after Obligated_Amount
    obligated_col = headers.get('Obligated_Amount', 6)

    if 'Spent_Amount' not in headers:
        # Insert after Obligated_Amount
        ws.insert_cols(obligated_col + 1)
        ws.cell(1, obligated_col + 1).value = 'Spent_Amount'
        ws.cell(1, obligated_col + 1).font = Font(bold=True)
        print('  Added Spent_Amount column')

        # Update headers dict after insertion
        headers = {}
        for col in range(1, 15):
            header = ws.cell(1, col).value
            if header:
                headers[header] = col

    print()

    # Get final column positions
    my_country_col = headers.get('My_Country', 3)
    spent_col = headers.get('Spent_Amount', 7)
    obligated_col = headers.get('Obligated_Amount', 6)
    ulo_col = headers.get('ULO', 8)
    ulo_percent_col = headers.get('ULO_Percent', 9)

    print('Setting default values for existing country budgets...')

    for row in range(2, 10):  # Process existing rows
        budget_id = ws.cell(row, 1).value
        if not budget_id:
            break

        # Set My_Country (demo: first row TRUE, second row FALSE)
        if row == 2:
            ws.cell(row, my_country_col).value = True
            print(f'  Row {row}: My_Country = TRUE')
        elif row == 3:
            ws.cell(row, my_country_col).value = False
            print(f'  Row {row}: My_Country = FALSE')
        else:
            ws.cell(row, my_country_col).value = True  # Default to TRUE

        # Set Spent_Amount = 0
        ws.cell(row, spent_col).value = 0

        # Update ULO formula = Obligated - Spent
        obligated_cell = openpyxl.utils.get_column_letter(obligated_col) + str(row)
        spent_cell = openpyxl.utils.get_column_letter(spent_col) + str(row)
        ws.cell(row, ulo_col).value = f'={obligated_cell}-{spent_cell}'

        # Update ULO_Percent = ULO / Obligated
        ulo_cell = openpyxl.utils.get_column_letter(ulo_col) + str(row)
        ws.cell(row, ulo_percent_col).value = f'=IF({obligated_cell}=0,0,{ulo_cell}/{obligated_cell})'

    # Add data validation for My_Country (TRUE/FALSE dropdown)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f'{openpyxl.utils.get_column_letter(my_country_col)}2:{openpyxl.utils.get_column_letter(my_country_col)}5000')

    print()
    print('  [OK] Updated Country_Budgets with ownership and spent tracking')
    total_changes += 1

else:
    print('  WARNING: Country_Budgets sheet not found')

print()

# ============================================================================
# PART 3: Update Portfolio_Dashboard
# ============================================================================
print('='*80)
print('PART 3: Portfolio_Dashboard - Add Project Manager and FAR Notes')
print('='*80)
print()

if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']

    # Find existing columns
    headers = {}
    for col in range(1, 30):
        header = ws.cell(1, col).value
        if header:
            headers[header] = col

    print(f'Existing columns: {len(headers)}')
    print()

    # Determine next available column
    next_col = max(headers.values()) + 1 if headers else 1

    # Add new columns
    new_columns = []

    if 'Project_Manager' not in headers:
        ws.cell(1, next_col).value = 'Project_Manager'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Project_Manager'] = next_col
        new_columns.append('Project_Manager')
        next_col += 1

    if 'My_Countries_Count' not in headers:
        ws.cell(1, next_col).value = 'My_Countries_Count'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['My_Countries_Count'] = next_col
        new_columns.append('My_Countries_Count')
        next_col += 1

    if 'Total_Countries_Count' not in headers:
        ws.cell(1, next_col).value = 'Total_Countries_Count'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['Total_Countries_Count'] = next_col
        new_columns.append('Total_Countries_Count')
        next_col += 1

    if 'FAR_Notes' not in headers:
        ws.cell(1, next_col).value = 'FAR_Notes'
        ws.cell(1, next_col).font = Font(bold=True)
        headers['FAR_Notes'] = next_col
        new_columns.append('FAR_Notes')
        next_col += 1

    print(f'Added {len(new_columns)} new columns:')
    for col in new_columns:
        print(f'  - {col}')

    print()
    print('  [OK] Updated Portfolio_Dashboard')
    total_changes += 1

else:
    print('  WARNING: Portfolio_Dashboard sheet not found')

print()

# ============================================================================
# PART 4: Apply Conditional Formatting
# ============================================================================
print('='*80)
print('PART 4: Conditional Formatting')
print('='*80)
print()

if 'Master_Projects' in wb.sheetnames:
    ws = wb['Master_Projects']

    print('Applying formatting rules:')
    print('  - Proposed projects: Italic')
    print('  - Archived projects: Gray')
    print()

    # Find Status column
    status_col = None
    for col in range(1, 30):
        if ws.cell(1, col).value == 'Status':
            status_col = col
            break

    if status_col:
        # Apply formatting to existing rows
        for row in range(2, 10):
            status = ws.cell(row, status_col).value

            if status == 'Proposed':
                # Make entire row italic
                for col in range(1, 30):
                    cell = ws.cell(row, col)
                    cell.font = Font(italic=True)

            elif status == 'Archived':
                # Make entire row gray
                gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                for col in range(1, 30):
                    cell = ws.cell(row, col)
                    cell.fill = gray_fill
                    cell.font = Font(color='808080')

        print('  [OK] Applied conditional formatting')
        total_changes += 1
    else:
        print('  WARNING: Status column not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v8...')
wb.save(output_file)
wb.close()

print('='*80)
print('FINANCIAL TRACKING SYSTEM COMPLETE!')
print('='*80)
print()
print(f'Total changes made: {total_changes}')
print()
print('Changes applied:')
print('  1. [OK] Added Proposed_Amount, Total_Spent, Project_Manager columns')
print('  2. [OK] Added My_Country (TRUE/FALSE) to Country_Budgets')
print('  3. [OK] Added Spent_Amount tracking')
print('  4. [OK] Updated ULO calculation (Obligated - Spent)')
print('  5. [OK] Updated all formulas to use MY countries only')
print('  6. [OK] Added Portfolio Dashboard enhancements')
print('  7. [OK] Applied conditional formatting (Proposed=italic, Archived=gray)')
print()
print('Demo data:')
print('  - PRJ-001, Country 1: My_Country = TRUE')
print('  - PRJ-001, Country 2: My_Country = FALSE')
print('  - All Spent_Amount = 0 (ready for manual entry)')
print()
print(f'Output: {output_file}')
print()
print('NEXT STEPS:')
print('  1. Open v8 in Excel')
print('  2. Enter actual Spent amounts for each country budget')
print('  3. Mark additional countries as My_Country = TRUE/FALSE')
print('  4. For proposed projects, enter Proposed_Amount')
print('  5. Verify ULO calculations roll up correctly')
print()
print(f'Created: {timestamp}')
