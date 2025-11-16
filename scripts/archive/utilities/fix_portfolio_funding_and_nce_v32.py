"""
Fix Portfolio_Dashboard v32:
1. Fix Total Funding to use Allocated instead of Proposed
2. Add At Risk Funding display
3. Add Award Number column back
4. Add NCE columns back
"""
import openpyxl

print("="*80)
print("FIXING PORTFOLIO_DASHBOARD - CREATING V32")
print("="*80)

# Load v31
print("\nLoading v31...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v31.xlsx')

ws_portfolio = wb['Portfolio_Dashboard']

print("\n" + "="*80)
print("PART 1: FIX TOTAL FUNDING FORMULA")
print("="*80)

print("\nChanging D4 from Total Proposed to Total Allocated...")
print("  Old: =TEXT(Control!B15,\"$#,##0,K\") (Total Proposed)")
print("  New: =TEXT(Control!B16,\"$#,##0,K\") (Total Allocated)")

ws_portfolio['D4'] = '=TEXT(Control!B16,"$#,##0,K")'

print("  D4 updated")

print("\n" + "="*80)
print("PART 2: ADD AT RISK FUNDING")
print("="*80)

print("\nAdding At Risk display...")

# Check for merged cells in rows 3-4 and unmerge if needed
merged_to_unmerge = []
for merged in ws_portfolio.merged_cells:
    # Check if merged cell range overlaps with F3 or F4
    if (merged.min_row <= 4 and merged.max_row >= 3):
        if merged.min_col <= 6 and merged.max_col >= 6:  # Column F
            merged_to_unmerge.append(str(merged))

for merged_range in merged_to_unmerge:
    ws_portfolio.unmerge_cells(merged_range)
    print(f"  Unmerged: {merged_range}")

print("\n  Adding At Risk in columns F3-F4:")
ws_portfolio['F3'] = 'AT RISK'
ws_portfolio['F4'] = '=TEXT(Control!B20,"$#,##0,K")'

print("    F3: 'AT RISK'")
print("    F4: =TEXT(Control!B20,\"$#,##0,K\")")

print("\n" + "="*80)
print("PART 3: ADD AWARD NUMBER AND NCE COLUMNS")
print("="*80)

print("\nInserting 3 columns after F (Countries)...")

# Insert 3 columns at position G
ws_portfolio.insert_cols(7, 3)

print("  Inserted columns G, H, I")
print("  Old G-M shifted to J-P")

print("\nAdding headers in row 10...")
ws_portfolio['G10'] = 'Award Number'
ws_portfolio['H10'] = 'NCE Eligible'
ws_portfolio['I10'] = 'NCE Status'

print("  G10: Award Number")
print("  H10: NCE Eligible")
print("  I10: NCE Status")

print("\nAdding formulas for rows 11-20...")

for row in range(11, 21):
    # G: Award Number - using ROW()-10 for row offset
    ws_portfolio.cell(row, 7).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[Award_Number],ROW()-10)="","",INDEX(T_Master_Projects[Award_Number],ROW()-10)),""))'

    # H: NCE Eligible
    ws_portfolio.cell(row, 8).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[NCE_Eligible],ROW()-10)="","",INDEX(T_Master_Projects[NCE_Eligible],ROW()-10)),""))'

    # I: NCE Status
    ws_portfolio.cell(row, 9).value = f'=IF(A{row}="","",IFERROR(IF(INDEX(T_Master_Projects[NCE_Status],ROW()-10)="","",INDEX(T_Master_Projects[NCE_Status],ROW()-10)),""))'

print("  Formulas added for rows 11-20")

print("\n" + "="*80)
print("SAVING V32")
print("="*80)

wb.save('2025-10-26-Tracker-v32.xlsx')

print("\nOK - v32 created!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nFixed in v32:")

print("\n1. Total Funding (D4):")
print("   - Now uses Control!B16 (Total Allocated)")
print("   - Should show actual allocated amounts")

print("\n2. At Risk Funding (F3-F4):")
print("   - F3: 'AT RISK' label")
print("   - F4: =TEXT(Control!B20,\"$#,##0,K\")")
print("   - Shows projects with ULO > 50%")

print("\n3. Award Number and NCE Columns Added:")
print("   - Column G: Award Number")
print("   - Column H: NCE Eligible (Yes/No)")
print("   - Column I: NCE Status")
print("   - All show blank when no project or value is empty")

print("\nNew column structure:")
print("  A: ID")
print("  B: Project Name")
print("  C: Status")
print("  D: Priority")
print("  E: Progress")
print("  F: Countries")
print("  G: Award Number (NEW)")
print("  H: NCE Eligible (NEW)")
print("  I: NCE Status (NEW)")
print("  J: Total Proposed (shifted from G)")
print("  K: Total Allocation (shifted from H)")
print("  L: Total Obligated (shifted from I)")
print("  M: Total Spent (shifted from J)")
print("  N: Total ULO (shifted from K)")
print("  O: ULO % (shifted from L)")
print("  P: Days Remaining (shifted from M)")

print("\nv32 ready!")
