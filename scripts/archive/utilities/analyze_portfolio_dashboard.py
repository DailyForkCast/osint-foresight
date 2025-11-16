import openpyxl

file_path = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v10.xlsx'

print('='*80)
print('ANALYZING PORTFOLIO DASHBOARD STRUCTURE')
print('='*80)
print()

wb = openpyxl.load_workbook(file_path, data_only=False)

if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']

    print('PORTFOLIO DASHBOARD - Structure:')
    print('-'*80)
    print(f'Max row: {ws.max_row}')
    print(f'Max column: {ws.max_column}')
    print()

    print('First 20 rows (columns A-M):')
    print()
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(14, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value is None:
                value = ''
            else:
                # If it's a formula, show abbreviated version
                value_str = str(value)
                if value_str.startswith('='):
                    # Show just first 30 chars of formula
                    value = value_str[:30] + '...' if len(value_str) > 30 else value_str
                else:
                    value = value_str[:25]  # Truncate long values
            row_data.append(str(value))
        print(f'Row {row_idx:2d}: {" | ".join(row_data)}')
    print()

wb.close()

print('='*80)
print('Now checking what data we have for country-specific analysis...')
print('='*80)
print()

# What would a country dashboard show?
print('COUNTRY DASHBOARD CONCEPT:')
print('-'*80)
print('User selects a country from dropdown (e.g., "Germany")')
print()
print('Dashboard shows:')
print('  1. All projects that have budgets for that country')
print('  2. Total allocated/obligated/spent for that country across all projects')
print('  3. ULO for that country')
print('  4. Project list with status')
print('  5. Key metrics (number of projects, total budget, etc.)')
print()
print('Data sources:')
print('  - Country_Budgets (filtered by selected country)')
print('  - Master_Projects (linked via Unique_ID)')
print('  - Country_PM_Assignments (PM info for that country)')
print()
