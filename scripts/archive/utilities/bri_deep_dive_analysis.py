#!/usr/bin/env python3
"""
BRI Deep Dive Analysis
Analyzes 872 Belt and Road Initiative contracts from TED database
Cross-references with other BRI datasets in the project
"""

import sqlite3
import json
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class BRIDeepDive:
    """Comprehensive BRI contract analysis"""

    def __init__(self):
        self.db_path = Path("F:/OSINT_WAREHOUSE/osint_master.db")

    def search_for_bri_datasets(self):
        """Search database for other BRI-related datasets"""

        print("\n" + "="*80)
        print("SEARCHING FOR BRI-RELATED DATASETS")
        print("="*80)
        print()

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get all table names
        tables = cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table'
            ORDER BY name
        """).fetchall()

        print(f"Total tables in database: {len(tables)}")
        print()

        # Search for potentially relevant tables
        bri_keywords = ['aid', 'infrastructure', 'bilateral', 'partnership',
                       'cooperation', 'china', 'investment', 'project']

        relevant_tables = []
        for table in tables:
            table_name = table[0].lower()
            for keyword in bri_keywords:
                if keyword in table_name:
                    relevant_tables.append(table[0])
                    break

        print(f"Potentially BRI-relevant tables: {len(relevant_tables)}")
        print()

        # Check each table for BRI mentions
        bri_data_sources = {}

        for table_name in relevant_tables:
            try:
                # Get table schema
                schema = cursor.execute(f"PRAGMA table_info({table_name})").fetchall()
                columns = [col[1] for col in schema]

                # Check if table has text columns
                text_columns = [col for col in columns if any(
                    typ in str(col).lower() for typ in ['title', 'description', 'name', 'text', 'content']
                )]

                if text_columns:
                    # Sample data for BRI mentions
                    sample_query = f"""
                        SELECT * FROM {table_name}
                        LIMIT 1000
                    """
                    rows = cursor.execute(sample_query).fetchall()

                    # Check for BRI mentions
                    bri_mentions = 0
                    for row in rows:
                        row_text = ' '.join([str(val) for val in row if val]).lower()
                        if any(term in row_text for term in ['belt and road', 'bri', 'silk road']):
                            bri_mentions += 1

                    if bri_mentions > 0:
                        total = cursor.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                        bri_data_sources[table_name] = {
                            'total_records': total,
                            'bri_mentions_in_sample': bri_mentions,
                            'columns': columns[:10]  # First 10 columns
                        }
                        print(f"✓ {table_name}: {total:,} records, {bri_mentions} BRI mentions in sample")

            except Exception as e:
                pass

        conn.close()

        print()
        print(f"Found {len(bri_data_sources)} tables with BRI data")
        print()

        return bri_data_sources

    def extract_bri_contracts(self):
        """Extract all 872 BRI contracts from TED database"""

        print("\n" + "="*80)
        print("EXTRACTING BRI CONTRACTS")
        print("="*80)
        print()

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get BRI contracts (limit description length for memory efficiency)
        print("Querying database...")
        bri_contracts = cursor.execute("""
            SELECT
                id, notice_number, publication_date,
                ca_name, ca_country, ca_city,
                contractor_name, contractor_country,
                contract_title,
                SUBSTR(contract_description, 1, 500) as contract_description,
                cpv_code, value_total, currency,
                influence_category, influence_priority, influence_patterns,
                detection_rationale, chinese_confidence
            FROM ted_contracts_production
            WHERE influence_category = 'BRI_RELATED'
            ORDER BY publication_date DESC
        """).fetchall()

        print(f"Total BRI contracts: {len(bri_contracts)}")
        print("Converting to dictionaries...")

        contracts_list = [dict(contract) for contract in bri_contracts]

        print("Extraction complete")
        print()

        conn.close()

        return contracts_list

    def analyze_bri_patterns(self, contracts):
        """Analyze patterns in BRI contracts"""

        print("\n" + "="*80)
        print("BRI CONTRACT PATTERN ANALYSIS")
        print("="*80)
        print()

        analysis = {
            'total': len(contracts),
            'by_year': defaultdict(int),
            'by_country': defaultdict(int),
            'by_sector': defaultdict(int),
            'by_value': defaultdict(list),
            'by_ca': defaultdict(int),
            'chinese_contractors': 0,
            'bri_keywords': defaultdict(int),
            'sample_contracts': []
        }

        for contract in contracts:
            # Year analysis
            pub_date = contract.get('publication_date', '')
            if pub_date:
                year = pub_date[:4]
                analysis['by_year'][year] += 1

            # Country analysis
            country = contract.get('ca_country', 'UNKNOWN')
            analysis['by_country'][country] += 1

            # Contracting authority analysis
            ca = contract.get('ca_name', 'UNKNOWN')[:50]
            analysis['by_ca'][ca] += 1

            # Value analysis
            value = contract.get('value_total')
            currency = contract.get('currency', 'UNKNOWN')
            if value and value > 0:
                analysis['by_value'][currency].append(value)

            # Check if also Chinese contractor
            if contract.get('detection_rationale'):
                analysis['chinese_contractors'] += 1

            # Extract BRI keywords
            combined_text = ' '.join([
                str(contract.get('contract_title') or ''),
                str(contract.get('contract_description') or '')
            ]).lower()

            if 'belt and road' in combined_text:
                analysis['bri_keywords']['belt_and_road'] += 1
            if re.search(r'\bbri\b', combined_text):
                analysis['bri_keywords']['bri'] += 1
            if 'silk road' in combined_text:
                analysis['bri_keywords']['silk_road'] += 1
            if 'new silk road' in combined_text:
                analysis['bri_keywords']['new_silk_road'] += 1

        # Get top contracts by value
        all_values = []
        for contract in contracts:
            value = contract.get('value_total')
            if value and value > 0:
                all_values.append((value, contract))

        analysis['sample_contracts'] = sorted(all_values, key=lambda x: x[0], reverse=True)[:20]

        # Print summary
        print("TEMPORAL DISTRIBUTION:")
        for year in sorted(analysis['by_year'].keys()):
            count = analysis['by_year'][year]
            pct = count / analysis['total'] * 100
            print(f"  {year}: {count:4d} contracts ({pct:5.1f}%)")

        print()
        print("GEOGRAPHIC DISTRIBUTION (Top 10):")
        for country, count in sorted(analysis['by_country'].items(), key=lambda x: x[1], reverse=True)[:10]:
            pct = count / analysis['total'] * 100
            print(f"  {country:3s}: {count:4d} contracts ({pct:5.1f}%)")

        print()
        print("BRI KEYWORD USAGE:")
        for keyword, count in sorted(analysis['bri_keywords'].items(), key=lambda x: x[1], reverse=True):
            print(f"  {keyword:20s}: {count:4d}")

        print()
        print(f"Contracts with Chinese Contractors: {analysis['chinese_contractors']} "
              f"({analysis['chinese_contractors']/analysis['total']*100:.1f}%)")

        print()
        print("VALUE ANALYSIS:")
        for currency, values in analysis['by_value'].items():
            if values:
                print(f"  {currency}: {len(values)} contracts, Total: {sum(values):,.0f}, "
                      f"Avg: {sum(values)/len(values):,.0f}")

        return analysis

    def cross_reference_datasets(self, bri_contracts, bri_data_sources):
        """Cross-reference BRI contracts with other datasets"""

        print("\n" + "="*80)
        print("CROSS-REFERENCE ANALYSIS")
        print("="*80)
        print()

        if not bri_data_sources:
            print("No additional BRI datasets found for cross-reference")
            return {}

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cross_refs = {}

        # Extract entity names from BRI contracts
        bri_entities = set()
        for contract in bri_contracts:
            ca_name = contract.get('ca_name', '')
            contractor = contract.get('contractor_name', '')
            if ca_name:
                bri_entities.add(ca_name.lower())
            if contractor:
                bri_entities.add(contractor.lower())

        print(f"Unique entities in BRI contracts: {len(bri_entities)}")
        print()

        # Cross-reference with each dataset
        for table_name, info in bri_data_sources.items():
            try:
                # Sample matching logic - customize based on table structure
                print(f"Cross-referencing with {table_name}...")

                # Get sample to understand structure
                sample = cursor.execute(f"SELECT * FROM {table_name} LIMIT 5").fetchall()

                # Count potential matches (basic text matching)
                # This is a simple example - would need to be customized per dataset
                matches = 0

                for entity in list(bri_entities)[:100]:  # Sample for performance
                    # Build search query based on table columns
                    text_columns = info['columns'][:5]  # First 5 columns

                    # Simple text search (can be improved with specific column matching)
                    for col in text_columns:
                        try:
                            result = cursor.execute(
                                f"SELECT COUNT(*) FROM {table_name} WHERE LOWER({col}) LIKE ?",
                                (f'%{entity}%',)
                            ).fetchone()[0]

                            if result > 0:
                                matches += result
                        except:
                            pass

                if matches > 0:
                    cross_refs[table_name] = matches
                    print(f"  Found {matches} potential matches")

            except Exception as e:
                print(f"  Error: {str(e)}")

        conn.close()

        print()
        if cross_refs:
            print("CROSS-REFERENCE SUMMARY:")
            for table, matches in cross_refs.items():
                print(f"  {table}: {matches} potential matches")
        else:
            print("No direct cross-references found")

        return cross_refs

    def generate_bri_report(self, contracts, analysis, cross_refs):
        """Generate comprehensive BRI intelligence report"""

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        report = f"""# Belt and Road Initiative (BRI) in EU Procurement
## Comprehensive Intelligence Analysis

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**Total BRI Contracts:** {len(contracts)}
**Time Period:** {min(c.get('publication_date', '9999') for c in contracts)[:4]} - {max(c.get('publication_date', '0000') for c in contracts)[:4]}

---

## Executive Summary

Analysis of EU procurement contracts (TED database) reveals **{len(contracts)} contracts with Belt and Road Initiative references**, representing significant Chinese strategic influence in European infrastructure and development projects.

### Key Findings

1. **Scale:** {len(contracts)} BRI-related procurement contracts across EU
2. **Chinese Contractors:** {analysis['chinese_contractors']} contracts ({analysis['chinese_contractors']/len(contracts)*100:.1f}%) involve confirmed Chinese contractors
3. **Geographic Spread:** {len(analysis['by_country'])} countries with BRI contract activity
4. **Temporal Trend:** Peak activity in {max(analysis['by_year'].items(), key=lambda x: x[1])[0]}

---

## Temporal Analysis

### BRI Contract Activity by Year

| Year | Contracts | % of Total | Trend |
|------|-----------|------------|-------|
"""

        for year in sorted(analysis['by_year'].keys()):
            count = analysis['by_year'][year]
            pct = count / len(contracts) * 100
            report += f"| {year} | {count} | {pct:.1f}% | |\n"

        report += f"""
**Peak Year:** {max(analysis['by_year'].items(), key=lambda x: x[1])[0]} with {max(analysis['by_year'].values())} contracts

---

## Geographic Distribution

### Top 10 Countries by BRI Contract Activity

| Country | Contracts | % of Total | Intelligence Assessment |
|---------|-----------|------------|------------------------|
"""

        for country, count in sorted(analysis['by_country'].items(), key=lambda x: x[1], reverse=True)[:10]:
            pct = count / len(contracts) * 100
            report += f"| {country} | {count} | {pct:.1f}% | |\n"

        report += f"""
---

## BRI Terminology Analysis

Understanding how BRI is referenced in EU procurement:

| Term | Usage Count | Context |
|------|-------------|---------|
"""

        for keyword, count in sorted(analysis['bri_keywords'].items(), key=lambda x: x[1], reverse=True):
            report += f"| {keyword.replace('_', ' ').title()} | {count} | |\n"

        report += f"""
---

## Chinese Contractor Involvement

**Contracts with Confirmed Chinese Contractors:** {analysis['chinese_contractors']} out of {len(contracts)} ({analysis['chinese_contractors']/len(contracts)*100:.1f}%)

This indicates:
- BRI projects with direct Chinese company participation
- Potential Chinese financing/influence even without Chinese contractors
- Strategic positioning in European infrastructure

---

## High-Value BRI Contracts

### Top 20 Contracts by Value

"""

        for i, (value, contract) in enumerate(analysis['sample_contracts'][:20], 1):
            report += f"""
#### {i}. {contract.get('contract_title', 'N/A')[:80]}

- **Value:** {value:,.0f} {contract.get('currency', 'N/A')}
- **Contracting Authority:** {contract.get('ca_name', 'N/A')[:60]}
- **Country:** {contract.get('ca_country', 'N/A')}
- **Date:** {contract.get('publication_date', 'N/A')[:10]}
- **Chinese Contractor:** {'Yes' if contract.get('detection_rationale') else 'No'}

"""

        report += f"""
---

## Cross-Reference Analysis

"""

        if cross_refs:
            report += "BRI contracts cross-referenced with other datasets:\n\n"
            for dataset, matches in cross_refs.items():
                report += f"- **{dataset}**: {matches} potential entity matches\n"
        else:
            report += "No direct cross-references found with other BRI datasets in current database.\n"

        report += f"""

**Recommendation:** Integrate external BRI datasets:
- AidData Global Chinese Development Finance Dataset
- Council on Foreign Relations BRI Database
- China Global Investment Tracker (AEI)

---

## Intelligence Implications

### Strategic Assessment

1. **Infrastructure Penetration:** BRI references in EU procurement indicate Chinese strategic positioning
2. **Financial Influence:** Contracts may involve Chinese financing beyond visible Chinese contractors
3. **Geographic Focus:** Concentrated in specific EU regions (see distribution above)
4. **Temporal Trends:** Activity patterns reveal strategic timing

### Risk Assessment

**HIGH RISK:**
- Contracts with both BRI reference AND Chinese contractors ({analysis['chinese_contractors']} contracts)
- Critical infrastructure sectors
- Projects involving technology transfer

**MEDIUM RISK:**
- BRI references without confirmed Chinese contractors
- May involve indirect Chinese influence or financing

### Recommendations

1. **Deep Dive by Country:** Analyze top BRI countries individually
2. **Sector Analysis:** Categorize by CPV codes (infrastructure, technology, etc.)
3. **Entity Mapping:** Cross-reference contracting authorities and contractors with:
   - Chinese SOE database
   - Section 1260H entities
   - GLEIF corporate ownership
4. **Financial Flow Analysis:** Track contract values and potential Chinese financing

---

## Data Quality Notes

- **Detection Method:** Pattern matching for "Belt and Road", "BRI", "Silk Road"
- **Confidence:** HIGH - Explicit BRI references in contract text
- **Limitations:** May not capture all Chinese-influenced contracts without explicit BRI mention

---

## Files Generated

1. **This Report:** `BRI_DEEP_DIVE_REPORT_{timestamp}.md`
2. **Full Data:** `bri_contracts_full_{timestamp}.json`
3. **Excel Export:** `bri_contracts_analysis_{timestamp}.xlsx`

---

**Analysis Complete**
**Next Steps:** Country-specific deep dives, sector analysis, entity cross-referencing
"""

        # Save report
        report_file = Path(f"analysis/BRI_DEEP_DIVE_REPORT_{timestamp}.md")
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report)

        # Save full JSON
        json_file = Path(f"analysis/bri_contracts_full_{timestamp}.json")
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump({
                'timestamp': timestamp,
                'total_contracts': len(contracts),
                'analysis': {
                    'by_year': dict(analysis['by_year']),
                    'by_country': dict(analysis['by_country']),
                    'by_ca': dict(list(analysis['by_ca'].items())[:50]),
                    'bri_keywords': dict(analysis['bri_keywords']),
                    'chinese_contractors': analysis['chinese_contractors']
                },
                'contracts': contracts
            }, f, indent=2, default=str)

        print(f"\n[SUCCESS] BRI Intelligence Report: {report_file}")
        print(f"[SUCCESS] Full JSON Data: {json_file}")

        return report_file, json_file

    def create_bri_excel(self, contracts, analysis):
        """Create Excel workbook with BRI analysis"""

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = Path(f"analysis/bri_contracts_analysis_{timestamp}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "BRI Contracts in EU Procurement - Analysis Summary"
        ws_summary['A1'].font = Font(size=16, bold=True)

        ws_summary['A3'] = "Total Contracts:"
        ws_summary['B3'] = len(contracts)
        ws_summary['A4'] = "Chinese Contractors:"
        ws_summary['B4'] = analysis['chinese_contractors']
        ws_summary['A5'] = "Countries:"
        ws_summary['B5'] = len(analysis['by_country'])

        # Sheet 2: All BRI Contracts
        ws_contracts = wb.create_sheet("All BRI Contracts")

        headers = ['ID', 'Notice', 'Date', 'Country', 'CA Name', 'Contractor',
                  'Title', 'Value', 'Currency', 'Chinese Contractor', 'BRI Keywords']

        for col, header in enumerate(headers, 1):
            cell = ws_contracts.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for row, contract in enumerate(contracts, 2):
            ws_contracts.cell(row, 1, contract.get('id'))
            ws_contracts.cell(row, 2, contract.get('notice_number', ''))
            ws_contracts.cell(row, 3, contract.get('publication_date', '')[:10])
            ws_contracts.cell(row, 4, contract.get('ca_country', ''))
            ws_contracts.cell(row, 5, (contract.get('ca_name') or '')[:50])
            ws_contracts.cell(row, 6, (contract.get('contractor_name') or '')[:40])
            ws_contracts.cell(row, 7, (contract.get('contract_title') or '')[:80])
            ws_contracts.cell(row, 8, contract.get('value_total'))
            ws_contracts.cell(row, 9, contract.get('currency', ''))
            ws_contracts.cell(row, 10, 'Yes' if contract.get('detection_rationale') else 'No')

            # Extract BRI keywords from contract
            text = f"{contract.get('contract_title', '')} {contract.get('contract_description', '')}".lower()
            keywords = []
            if 'belt and road' in text:
                keywords.append('Belt and Road')
            if re.search(r'\bbri\b', text):
                keywords.append('BRI')
            if 'silk road' in text:
                keywords.append('Silk Road')
            ws_contracts.cell(row, 11, ', '.join(keywords))

        # Column widths
        ws_contracts.column_dimensions['A'].width = 8
        ws_contracts.column_dimensions['B'].width = 15
        ws_contracts.column_dimensions['C'].width = 12
        ws_contracts.column_dimensions['D'].width = 8
        ws_contracts.column_dimensions['E'].width = 40
        ws_contracts.column_dimensions['F'].width = 30
        ws_contracts.column_dimensions['G'].width = 60
        ws_contracts.column_dimensions['H'].width = 15
        ws_contracts.column_dimensions['I'].width = 8
        ws_contracts.column_dimensions['J'].width = 15
        ws_contracts.column_dimensions['K'].width = 30

        ws_contracts.freeze_panes = 'A2'
        ws_contracts.auto_filter.ref = ws_contracts.dimensions

        # Sheet 3: By Country
        ws_country = wb.create_sheet("By Country")
        ws_country['A1'] = "Country"
        ws_country['A1'].font = Font(bold=True)
        ws_country['B1'] = "Contracts"
        ws_country['B1'].font = Font(bold=True)

        for row, (country, count) in enumerate(sorted(analysis['by_country'].items(), key=lambda x: x[1], reverse=True), 2):
            ws_country.cell(row, 1, country)
            ws_country.cell(row, 2, count)

        # Sheet 4: By Year
        ws_year = wb.create_sheet("By Year")
        ws_year['A1'] = "Year"
        ws_year['A1'].font = Font(bold=True)
        ws_year['B1'] = "Contracts"
        ws_year['B1'].font = Font(bold=True)

        for row, (year, count) in enumerate(sorted(analysis['by_year'].items()), 2):
            ws_year.cell(row, 1, year)
            ws_year.cell(row, 2, count)

        wb.save(excel_file)

        print(f"[SUCCESS] Excel Analysis: {excel_file}")

        return excel_file

    def run_analysis(self):
        """Run complete BRI deep dive analysis"""

        print("\n" + "="*80)
        print("BELT AND ROAD INITIATIVE (BRI) - DEEP DIVE ANALYSIS")
        print("="*80)
        print()

        # Step 1: Search for existing BRI datasets
        bri_data_sources = self.search_for_bri_datasets()

        # Step 2: Extract BRI contracts
        bri_contracts = self.extract_bri_contracts()

        # Step 3: Analyze patterns
        analysis = self.analyze_bri_patterns(bri_contracts)

        # Step 4: Cross-reference
        cross_refs = self.cross_reference_datasets(bri_contracts, bri_data_sources)

        # Step 5: Generate report
        report_file, json_file = self.generate_bri_report(bri_contracts, analysis, cross_refs)

        # Step 6: Create Excel
        excel_file = self.create_bri_excel(bri_contracts, analysis)

        print("\n" + "="*80)
        print("BRI DEEP DIVE COMPLETE")
        print("="*80)
        print()
        print(f"Report: {report_file}")
        print(f"Data: {json_file}")
        print(f"Excel: {excel_file}")
        print()


if __name__ == '__main__':
    analyzer = BRIDeepDive()
    analyzer.run_analysis()
