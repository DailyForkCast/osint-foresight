"""
Move Egypt and Morocco from AF to NEA region
"""
import openpyxl

print("="*80)
print("MOVING EGYPT AND MOROCCO TO NEA REGION")
print("="*80)

# Load v30
print("\nLoading v30...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v30.xlsx')

print("\n" + "="*80)
print("PART 1: UPDATE COUNTRY_REGIONS TABLE")
print("="*80)

ws_regions = wb['Country_Regions']

print("\nSearching for Egypt and Morocco...")

countries_updated = []

# Search through the table
row = 2
while ws_regions.cell(row, 1).value:
    country_code = ws_regions.cell(row, 1).value
    country_name = ws_regions.cell(row, 2).value
    current_region = ws_regions.cell(row, 3).value

    if country_name in ['Egypt', 'Morocco']:
        print(f"\nFound: {country_name} ({country_code})")
        print(f"  Current region: {current_region}")
        print(f"  Changing to: NEA")

        # Update the region
        ws_regions.cell(row, 3).value = 'NEA'
        countries_updated.append((country_name, country_code, current_region, 'NEA'))

    row += 1

print(f"\nUpdated {len(countries_updated)} countries")

print("\n" + "="*80)
print("PART 2: UPDATE CONFIG_LISTS")
print("="*80)

ws_config = wb['Config_Lists']

print("\nSearching Config_Lists for Egypt and Morocco...")

# Config_Lists has countries in columns D-F
row = 2
while ws_config.cell(row, 4).value:  # Column D has country codes
    country_code = ws_config.cell(row, 4).value
    country_name = ws_config.cell(row, 5).value
    current_region = ws_config.cell(row, 6).value

    if country_name in ['Egypt', 'Morocco']:
        print(f"\nFound: {country_name} ({country_code})")
        print(f"  Current region: {current_region}")
        print(f"  Changing to: NEA")

        # Update the region
        ws_config.cell(row, 6).value = 'NEA'

    row += 1

print("\nConfig_Lists updated")

print("\n" + "="*80)
print("PART 3: RE-SORT CONFIG_LISTS BY REGION")
print("="*80)

print("\nRe-sorting countries by region, then country name...")

# Read all countries
countries = []
row = 2
while ws_config.cell(row, 4).value:
    country_code = ws_config.cell(row, 4).value
    country_name = ws_config.cell(row, 5).value
    region = ws_config.cell(row, 6).value
    countries.append((country_code, country_name, region))
    row += 1

print(f"  Found {len(countries)} countries")

# Sort by region, then country name
region_order = {'AF': 1, 'EAP': 2, 'EUR': 3, 'NEA': 4, 'SCA': 5, 'WHA': 6}
countries_sorted = sorted(countries, key=lambda x: (region_order.get(x[2], 99), x[1]))

# Clear existing data
for row in range(2, len(countries) + 2):
    ws_config.cell(row, 4).value = None
    ws_config.cell(row, 5).value = None
    ws_config.cell(row, 6).value = None

# Write sorted data
for idx, (country_code, country_name, region) in enumerate(countries_sorted, start=2):
    ws_config.cell(idx, 4).value = country_code
    ws_config.cell(idx, 5).value = country_name
    ws_config.cell(idx, 6).value = region

print("  Countries re-sorted")

print("\n" + "="*80)
print("SAVING V30")
print("="*80)

wb.save('2025-10-26-Tracker-v30.xlsx')

print("\nOK - Egypt and Morocco moved to NEA!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

print("\nCountries updated:")
for country_name, country_code, old_region, new_region in countries_updated:
    print(f"  {country_name} ({country_code}): {old_region} -> {new_region}")

print("\nUpdated in:")
print("  - Country_Regions table")
print("  - Config_Lists (and re-sorted)")

print("\nRegion classification:")
print("  AF: Africa (Sub-Saharan)")
print("  NEA: Near East Asia (includes North Africa/Middle East)")
print("  Egypt and Morocco now correctly in NEA")

print("\nv30 updated!")
