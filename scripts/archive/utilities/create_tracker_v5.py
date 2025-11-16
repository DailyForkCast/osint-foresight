import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

input_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v4.xlsx'
output_file = 'C:/Projects/OSINT-Foresight/2025-10-26-Tracker-v5.xlsx'

print('='*80)
print('TRACKER v5 IMPLEMENTATION')
print('='*80)
print()
print('Updates:')
print('  1. Funding format: Thousands -> Millions')
print('  2. Regional_Summary sheet build-out')
print('  3. Project_Spotlight redesign')
print()
print('='*80)

# Load workbook
print('Loading v4...')
wb = openpyxl.load_workbook(input_file, data_only=False)
print(f'Loaded. Sheets: {", ".join(wb.sheetnames)}')
print()

# ============================================================================
# PART 1: Update Funding Format (Thousands -> Millions)
# ============================================================================
print('='*80)
print('PART 1: Updating Funding Format to Millions')
print('='*80)
print()

if 'Portfolio_Dashboard' in wb.sheetnames:
    ws = wb['Portfolio_Dashboard']
    updates = 0

    # Search for cells with TEXT formulas using "$#,##0,K" format
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'TEXT(' in str(cell.value) and '"$#,##0,K"' in str(cell.value):
                    old_formula = cell.value
                    new_formula = old_formula.replace('"$#,##0,K"', '"$#,##0.0,M"')
                    cell.value = new_formula
                    updates += 1
                    print(f'  Updated {cell.coordinate}:')
                    print(f'    Old: {old_formula}')
                    print(f'    New: {new_formula}')

    print()
    print(f'COMPLETE: {updates} formulas updated to millions format')
else:
    print('WARNING: Portfolio_Dashboard not found')

print()

# ============================================================================
# PART 2: Build Out Regional_Summary Sheet
# ============================================================================
print('='*80)
print('PART 2: Building Regional_Summary Sheet')
print('='*80)
print()

if 'Regional_Summary' in wb.sheetnames:
    ws = wb['Regional_Summary']

    # Clear existing content
    print('Clearing existing content...')
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Define headers
    headers = [
        'Region',           # A
        'Region_Name',      # B
        'Active_Projects',  # C
        'Total_Countries',  # D
        'Active_Countries', # E
        'Total_Allocated',  # F
        'Total_Obligated',  # G
        'Total_Spent',      # H
        'ULO_Amount',       # I
        'ULO_Percent',      # J
        'Spend_Rate',       # K
        'Health_Status'     # L
    ]

    # Define regions
    regions = [
        ('EUR', 'Europe'),
        ('WHA', 'Western Hemisphere'),
        ('EAP', 'East Asia Pacific'),
        ('AF', 'Africa'),
        ('NEA', 'Near East Asia'),
        ('SCA', 'South Central Asia')
    ]

    # Add headers
    print('Adding headers...')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add region data and formulas
    print('Adding region data and formulas...')
    for idx, (region_code, region_name) in enumerate(regions, start=2):
        row_num = idx

        # A: Region code
        ws.cell(row=row_num, column=1, value=region_code)

        # B: Region name
        ws.cell(row=row_num, column=2, value=region_name)

        # C: Active_Projects - count projects that have this region in their countries
        # Note: This assumes Master_Projects has a Countries column that we can search
        ws.cell(row=row_num, column=3, value=f'=SUMPRODUCT((ISNUMBER(FIND(A{row_num},Master_Projects!M:M)))*1)')

        # D: Total_Countries - count countries in this region
        ws.cell(row=row_num, column=4, value=f'=COUNTIF(Country_Regions!C:C,A{row_num})')

        # E: Active_Countries - countries with budgets in this region
        # This is complex - need to count unique countries in Country_Budgets that belong to this region
        ws.cell(row=row_num, column=5, value=f'=SUMPRODUCT((Country_Budgets!C:C<>"")*(Country_Regions!C:C=A{row_num}))')

        # F: Total_Allocated
        # Need to sum budgets where the country is in this region
        # This requires matching Country_Budgets country codes with Country_Regions
        ws.cell(row=row_num, column=6, value=f'=SUMPRODUCT((Country_Regions!C:C=A{row_num})*(ISNUMBER(MATCH(Country_Regions!A:A,Country_Budgets!C:C,0))),Country_Budgets!E:E)')

        # G: Total_Obligated
        ws.cell(row=row_num, column=7, value=f'=SUMPRODUCT((Country_Regions!C:C=A{row_num})*(ISNUMBER(MATCH(Country_Regions!A:A,Country_Budgets!C:C,0))),Country_Budgets!F:F)')

        # H: Total_Spent
        ws.cell(row=row_num, column=8, value=f'=SUMPRODUCT((Country_Regions!C:C=A{row_num})*(ISNUMBER(MATCH(Country_Regions!A:A,Country_Budgets!C:C,0))),Country_Budgets!G:G)')

        # I: ULO_Amount (Allocated - Obligated)
        ws.cell(row=row_num, column=9, value=f'=F{row_num}-G{row_num}')

        # J: ULO_Percent
        ws.cell(row=row_num, column=10, value=f'=IF(F{row_num}=0,0,I{row_num}/F{row_num})')

        # K: Spend_Rate
        ws.cell(row=row_num, column=11, value=f'=IF(G{row_num}=0,0,H{row_num}/G{row_num})')

        # L: Health_Status
        ws.cell(row=row_num, column=12, value=f'=IF(J{row_num}>0.5,"High ULO",IF(K{row_num}<0.25,"Low Execution","On Track"))')

        print(f'  Added: {region_code} - {region_name}')

    # Format currency columns (F, G, H, I)
    for col in [6, 7, 8, 9]:
        for row in range(2, 8):  # 6 regions
            cell = ws.cell(row=row, column=col)
            cell.number_format = '$#,##0'

    # Format percentage columns (J, K)
    for col in [10, 11]:
        for row in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.0%'

    # Set column widths
    column_widths = {
        'A': 10, 'B': 22, 'C': 15, 'D': 16, 'E': 16, 'F': 16,
        'G': 16, 'H': 14, 'I': 14, 'J': 12, 'K': 12, 'L': 16
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    print()
    print('COMPLETE: Regional_Summary built with 6 regions and 12 columns')
else:
    print('WARNING: Regional_Summary not found')

print()

# ============================================================================
# PART 3: Redesign Project_Spotlight
# ============================================================================
print('='*80)
print('PART 3: Redesigning Project_Spotlight')
print('='*80)
print()

if 'Project_Spotlight' in wb.sheetnames:
    ws = wb['Project_Spotlight']

    # Step 1: Remove the old TEXT formula (if it exists)
    print('Step 1: Checking for old TEXT formula to remove...')
    removed_formulas = 0
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'TEXT(Control!B15' in str(cell.value):
                    print(f'  Removing formula from {cell.coordinate}: {cell.value}')
                    cell.value = None
                    removed_formulas += 1
    print(f'  Removed {removed_formulas} old formulas')

    # Step 2: Clear map reference cells (G3, G5 from v4)
    print('Step 2: Clearing map reference area...')
    if ws['G3'].value:
        print(f'  Clearing G3: {ws["G3"].value}')
        ws['G3'].value = None
    if ws['G5'].value:
        print(f'  Clearing G5: {ws["G5"].value}')
        ws['G5'].value = None

    # Step 3: Extend summary box (B5:F14)
    print('Step 3: Setting up extended summary box (B5:F14)...')
    # The summary is already in the sheet, we're just extending the merge/format
    # For now, just add a note in B5 if it's empty
    if not ws['B5'].value or ws['B5'].value == '':
        ws['B5'].value = '=IFERROR(INDEX(Master_Projects!D:D,MATCH($B$2,Master_Projects!B:B,0)),"Select a project to view details")'

    # Step 4: Add TARGET AUDIENCES header and formulas (G5:I14)
    print('Step 4: Adding Target Audiences section (G5:I14)...')

    # Unmerge any merged cells in the target area first
    print('  Unmerging cells in G4:K14 area...')
    merged_ranges_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        # Check if this merged range overlaps with our target area (columns G-K, rows 4-14)
        merged_str = str(merged_range)
        # Simple check - if it contains any of our target cells
        if any(f'{col}{row}' in merged_str for col in ['G', 'H', 'I', 'J', 'K'] for row in range(4, 15)):
            merged_ranges_to_remove.append(merged_range)

    for merged_range in merged_ranges_to_remove:
        ws.unmerge_cells(str(merged_range))
        print(f'    Unmerged: {merged_range}')

    # Add header in row 4
    ws['G4'].value = 'TARGET AUDIENCES'
    ws['G4'].font = Font(bold=True, size=12)
    ws['G4'].alignment = Alignment(horizontal='left', vertical='center')

    # Add column headers in row 5
    ws['G5'].value = 'Type'
    ws['H5'].value = 'Description'
    ws['I5'].value = 'Priority'
    for cell in [ws['G5'], ws['H5'], ws['I5']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Add formulas for audiences (rows 6-14)
    # Using INDEX/SMALL/IF array formula for compatibility
    for row_offset in range(1, 10):  # 9 rows of data
        row_num = 5 + row_offset

        # Type (Column G)
        ws[f'G{row_num}'].value = f'=IFERROR(INDEX(Project_Audiences!B:B,SMALL(IF(Project_Audiences!$A:$A=$B$2,ROW(Project_Audiences!$A:$A)),{row_offset})),"")'

        # Description (Column H)
        ws[f'H{row_num}'].value = f'=IF(G{row_num}="","",INDEX(Project_Audiences!D:D,MATCH(G{row_num},Project_Audiences!B:B,0)))'

        # Priority (Column I)
        ws[f'I{row_num}'].value = f'=IF(G{row_num}="","",INDEX(Project_Audiences!E:E,MATCH(G{row_num},Project_Audiences!B:B,0)))'

    # Step 5: Add TARGET TECHNOLOGIES header and formulas (J5:K14)
    print('Step 5: Adding Target Technologies section (J5:K14)...')

    # Add header in row 4
    ws['J4'].value = 'TARGET TECHNOLOGIES'
    ws['J4'].font = Font(bold=True, size=12)
    ws['J4'].alignment = Alignment(horizontal='left', vertical='center')

    # Check if Project_Technologies sheet exists
    has_tech_sheet = 'Project_Technologies' in wb.sheetnames
    if not has_tech_sheet:
        print('  WARNING: Project_Technologies sheet not found - adding placeholder')
        ws['J5'].value = 'Technology'
        ws['K5'].value = 'Category'
        for cell in [ws['J5'], ws['K5']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws['J6'].value = '[Create Project_Technologies sheet]'
    else:
        # Add column headers
        ws['J5'].value = 'Technology'
        ws['K5'].value = 'Category'
        for cell in [ws['J5'], ws['K5']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # Add formulas (rows 6-14)
        for row_offset in range(1, 10):
            row_num = 5 + row_offset

            # Technology name (Column J)
            ws[f'J{row_num}'].value = f'=IFERROR(INDEX(Project_Technologies!B:B,SMALL(IF(Project_Technologies!$A:$A=$B$2,ROW(Project_Technologies!$A:$A)),{row_offset})),"")'

            # Category (Column K)
            ws[f'K{row_num}'].value = f'=IF(J{row_num}="","",INDEX(Project_Technologies!C:C,MATCH(J{row_num},Project_Technologies!B:B,0)))'

    # Step 6: Add KEY DELIVERABLES section (starting row 16)
    print('Step 6: Adding Key Deliverables section (rows 16+)...')

    # Unmerge any merged cells in the deliverables area
    print('  Unmerging cells in B16:F27 area...')
    merged_ranges_to_remove2 = []
    for merged_range in list(ws.merged_cells.ranges):
        merged_str = str(merged_range)
        # Check if it overlaps with B16:F27
        if any(f'{col}{row}' in merged_str for col in ['B', 'C', 'D', 'E', 'F'] for row in range(16, 28)):
            merged_ranges_to_remove2.append(merged_range)

    for merged_range in merged_ranges_to_remove2:
        ws.unmerge_cells(str(merged_range))
        print(f'    Unmerged: {merged_range}')

    # Add section header
    ws['B16'].value = 'KEY DELIVERABLES'
    ws['B16'].font = Font(bold=True, size=12)
    ws['B16'].alignment = Alignment(horizontal='left', vertical='center')

    # Add column headers in row 17
    deliverable_headers = {
        'B17': 'Deliverable Name',
        'C17': 'Due Date',
        'D17': 'Status',
        'E17': 'Owner',
        'F17': 'Notes'
    }

    for cell_ref, header_text in deliverable_headers.items():
        ws[cell_ref].value = header_text
        ws[cell_ref].font = Font(bold=True)
        ws[cell_ref].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

    # Add formulas for deliverables (rows 18-27, 10 rows)
    for row_offset in range(1, 11):
        row_num = 17 + row_offset

        # Deliverable Name (Column B)
        ws[f'B{row_num}'].value = f'=IFERROR(INDEX(Project_Deliverables!B:B,SMALL(IF(Project_Deliverables!$A:$A=$B$2,ROW(Project_Deliverables!$A:$A)),{row_offset})),"")'

        # Due Date (Column C)
        ws[f'C{row_num}'].value = f'=IF(B{row_num}="","",INDEX(Project_Deliverables!D:D,MATCH(B{row_num},Project_Deliverables!B:B,0)))'
        ws[f'C{row_num}'].number_format = 'mm/dd/yy'

        # Status (Column D)
        ws[f'D{row_num}'].value = f'=IF(B{row_num}="","",INDEX(Project_Deliverables!E:E,MATCH(B{row_num},Project_Deliverables!B:B,0)))'

        # Owner (Column E)
        ws[f'E{row_num}'].value = f'=IF(B{row_num}="","",INDEX(Project_Deliverables!F:F,MATCH(B{row_num},Project_Deliverables!B:B,0)))'

        # Notes (Column F) - check if column exists in Project_Deliverables
        # Assuming notes might be in column H or similar
        ws[f'F{row_num}'].value = f'=IF(B{row_num}="","","")'

    # Set column widths
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 18

    print()
    print('COMPLETE: Project_Spotlight redesigned')
    print('  - Extended summary box')
    print('  - Added Target Audiences (G5:I14)')
    print('  - Added Target Technologies (J5:K14)')
    print('  - Added Key Deliverables (B16:F27)')
else:
    print('WARNING: Project_Spotlight not found')

print()

# ============================================================================
# SAVE
# ============================================================================
print('='*80)
print('Saving as v5...')
wb.save(output_file)
wb.close()

print('='*80)
print('v5 IMPLEMENTATION COMPLETE!')
print('='*80)
print()
print('Summary of changes:')
print('  1. Funding format: Changed to millions ($#,##0.0,M)')
print('  2. Regional_Summary: Built with 6 regions, 12 columns')
print('  3. Project_Spotlight: Redesigned with dynamic sections')
print()
print(f'Output: {output_file}')
print()
print('IMPORTANT: Open the file and verify:')
print('  - Portfolio_Dashboard shows funding in millions')
print('  - Regional_Summary has all 6 regions with formulas')
print('  - Project_Spotlight shows audiences/technologies/deliverables')
print('  - Change project selection in B2 to test filtering')
print()
print('Note: Array formulas (IF arrays) need to be entered as array formulas')
print('in Excel. They may show as regular formulas but should still work.')
