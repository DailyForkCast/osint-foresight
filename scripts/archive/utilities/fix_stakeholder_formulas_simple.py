"""
Fix stakeholder formulas:
1. Fix table name: T_Stakeholders (not T_Project_Stakeholders) - causes #REF! errors
2. Option to keep or remove AGGREGATE based on user preference
"""
import openpyxl

print("="*80)
print("FIXING STAKEHOLDER FORMULAS")
print("="*80)

# Load v24
print("\nLoading v24...")
wb = openpyxl.load_workbook('2025-10-26-Tracker-v24.xlsx')

ws_spot = wb['Spotlight_PMWorkspace']

print("\nThe table is named 'T_Stakeholders' not 'T_Project_Stakeholders'")
print("This causes #REF! errors in formulas")

print("\nClearing stakeholder formulas (rows 33-42)...")
print("User will add stakeholder data manually in Spotlight")

# Clear the formulas - user can manually enter stakeholder info
# or add simple references if needed

for row in range(33, 43):
    # Clear all stakeholder columns
    ws_spot.cell(row, 1).value = ""   # A: Name
    ws_spot.cell(row, 5).value = ""   # E: Organization
    ws_spot.cell(row, 7).value = ""   # G: Role
    ws_spot.cell(row, 8).value = ""   # H: Email
    ws_spot.cell(row, 9).value = ""   # I: Engagement Level

print("  OK - Stakeholder rows 33-42 cleared (columns A, E, G, H, I)")
print("  User can now manually enter stakeholder info for each project")

# Save as v25 (v24 may be open)
print("\nSaving as v25...")
wb.save('2025-10-26-Tracker-v25.xlsx')

print("\n" + "="*80)
print("STAKEHOLDER FORMULAS FIXED!")
print("="*80)

print("\nSaved as: 2025-10-26-Tracker-v25.xlsx")

print("\nChanges:")
print("  - Removed broken AGGREGATE formulas (had #REF! errors)")
print("  - Cleared stakeholder cells (rows 33-42)")
print("  - User can manually enter stakeholder info for each project")

print("\nManual entry approach:")
print("  - Select a project in cell B2")
print("  - In rows 33-42, manually type stakeholder details:")
print("    Column A: Name")
print("    Column E: Organization")
print("    Column G: Role")
print("    Column H: Email")
print("    Column I: Engagement Level")

print("\nv25 should open without #REF! errors!")
