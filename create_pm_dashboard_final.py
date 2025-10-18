import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import random

# Create the workbook
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# ================ HELPER FUNCTIONS ================
def create_styled_header(ws, row, columns, bg_color="366092"):
    """Create a styled header row"""
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def autosize_columns(ws):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                elif hasattr(cell, 'column'):
                    column_letter = get_column_letter(cell.column)
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# ================ DATA GENERATION ================

# Generate dates
today = datetime.now()
base_date = datetime(2025, 1, 1)

# 1. Master Projects Data (10 projects)
projects_data = []
project_statuses = ["In Progress", "Planning", "At Risk", "On Hold", "Complete"]
priorities = ["Critical", "High", "Medium", "Low"]
categories = ["Market Expansion", "Digital Transformation", "Infrastructure", "Research", "Innovation"]
funding_sources = ["Government", "EU Funds", "Private", "Mixed", "Industry Partners"]

for i in range(1, 11):
    project_id = f"PRJ-{i:03d}"
    names = [
        "UK Market Expansion", "German Digital Platform", "Nordic Innovation Hub",
        "Mediterranean Trade Route", "Eastern Europe Gateway", "Alpine Research Center",
        "Iberian Tech Corridor", "Baltic Sea Initiative", "Central Europe Network", "Atlantic Partnership"
    ]
    project_name = names[i-1]
    summary = f"Strategic initiative for {names[i-1].lower()} development and integration"
    status = random.choice(project_statuses[:3])  # Focus on active projects
    priority = random.choice(priorities[:2])  # Focus on high priority
    progress = random.uniform(0.1, 0.95)
    start_date = base_date + timedelta(days=random.randint(0, 60))
    end_date = start_date + timedelta(days=random.randint(180, 365))
    days_remaining = (end_date - today).days
    lead = random.choice(["John Smith", "Maria Garcia", "Hans Mueller", "Sophie Martin", "Erik Johansson"])
    sponsor = random.choice(["CEO", "CTO", "COO", "CFO", "Board"])
    partner = random.choice(["Acme Corp", "TechCo GmbH", "Nordic Solutions", "EuroTech", "GlobalNet"])
    category = categories[i % 5]
    funding = funding_sources[i % 5]

    projects_data.append([
        project_id, project_name, summary, status, priority, progress,
        start_date, end_date, days_remaining, lead, sponsor, partner,
        0, 0, 0,  # Budget placeholders (M, N, O)
        category, funding, None, None, f"Milestone {i}", start_date + timedelta(days=30),
        0, 0, "Active", "", ""  # Additional columns
    ])

df_projects = pd.DataFrame(projects_data, columns=[
    'Project_ID', 'Project_Name', 'Summary', 'Status', 'Priority', 'Progress',
    'Start_Date', 'End_Date', 'Days_Remaining', 'Lead', 'Sponsor', 'Partner',
    'Budget_Total', 'Budget_Spent', 'Budget_Percent', 'Category', 'Funding_Source',
    'Reserved1', 'Reserved2', 'Next_Milestone', 'Next_Milestone_Date',
    'Event_Count', 'Countries_Count', 'Phase', 'Health_Score', 'Health_Visual'
])

# 2. Project Country Budgets (~30 rows, 3-5 per project)
budget_data = []
budget_id = 1
european_countries = [
    "United Kingdom", "Germany", "France", "Italy", "Spain", "Netherlands",
    "Belgium", "Poland", "Sweden", "Norway", "Denmark", "Finland",
    "Austria", "Switzerland", "Portugal", "Greece", "Ireland", "Czech Republic"
]

for i in range(1, 11):
    project_id = f"PRJ-{i:03d}"
    project_name = df_projects[df_projects['Project_ID'] == project_id]['Project_Name'].iloc[0]
    num_countries = random.randint(2, 5)
    countries_used = random.sample(european_countries, num_countries)

    for j, country in enumerate(countries_used):
        budget_id_str = f"BUD-{budget_id:03d}"
        phase = f"Phase {(j % 3) + 1}"
        budget_total = random.randint(50000, 250000)
        budget_spent = random.randint(0, int(budget_total * 0.8))
        budget_remaining = budget_total - budget_spent
        percent_used = budget_spent / budget_total if budget_total > 0 else 0

        start_date = base_date + timedelta(days=j*30)
        end_date = start_date + timedelta(days=90)
        status = random.choice(["In Progress", "Planning", "Complete"])

        budget_data.append([
            budget_id_str, project_id, project_name, country, j+1, phase,
            budget_total, budget_spent, budget_remaining, percent_used,
            start_date, end_date, status, "", 0, 0, ""
        ])
        budget_id += 1

df_budgets = pd.DataFrame(budget_data, columns=[
    'Budget_ID', 'Project_ID', 'Project_Name', 'Country', 'Sequence', 'Phase',
    'Budget_Total', 'Budget_Spent', 'Budget_Remaining', 'Percent_Used',
    'Start_Date', 'End_Date', 'Status', 'Budget_Status', 'Days_in_Phase',
    'Days_Active', 'Burn_Rate_Alert'
])

# 3. Project Milestones (~30 rows, 3 per project)
milestone_data = []
milestone_id = 1
milestone_types = ["Planning", "Design", "Development", "Testing", "Deployment", "Review"]

for i in range(1, 11):
    project_id = f"PRJ-{i:03d}"
    for j in range(3):
        milestone_id_str = f"MIL-{milestone_id:03d}"
        name = ""  # Will be calculated by formula
        milestone_date = base_date + timedelta(days=(j+1)*60)
        status = random.choice(["Complete", "In Progress", "Pending"])
        phase = f"Phase {j+1}"
        criticality = random.choice(["Critical", "High", "Medium"])
        milestone_type = milestone_types[j % 6]
        days_until = (milestone_date - today).days

        milestone_data.append([
            milestone_id_str, name, project_id, milestone_date, status,
            phase, criticality, milestone_type, days_until, "", "", "", ""
        ])
        milestone_id += 1

df_milestones = pd.DataFrame(milestone_data, columns=[
    'Milestone_ID', 'Name', 'Project_ID', 'Due_Date', 'Status',
    'Phase', 'Criticality', 'Type', 'Days_Until', 'Time_Group',
    'Visual_Status', 'Project_Name', 'Alert'
])

# 4. Project Events (~20 rows, 2 per project)
event_data = []
event_id = 1
event_types = ["Meeting", "Review", "Workshop", "Conference", "Training", "Launch"]

for i in range(1, 11):
    project_id = f"PRJ-{i:03d}"
    for j in range(2):
        event_id_str = f"EVT-{event_id:03d}"
        event_name = f"{random.choice(event_types)} - {project_id}"
        country = random.choice(european_countries)
        city = random.choice(["London", "Berlin", "Paris", "Rome", "Madrid", "Amsterdam"])
        event_date = base_date + timedelta(days=random.randint(30, 180))
        duration = random.randint(1, 3)

        event_data.append([
            event_id_str, project_id, event_name, random.choice(event_types),
            country, city, event_date, duration, random.choice(["Confirmed", "Tentative"]),
            "", random.randint(5, 50)
        ])
        event_id += 1

df_events = pd.DataFrame(event_data, columns=[
    'Event_ID', 'Project_ID', 'Event_Name', 'Type', 'Country', 'City',
    'Date', 'Duration_Days', 'Status', 'Notes', 'Attendees'
])

# 5. Calendar Todo / Action Tracker (~30 rows)
todo_data = []
todo_id = 1
task_types = ["Review", "Approve", "Submit", "Complete", "Prepare", "Analyze"]

for i in range(30):
    todo_id_str = f"TSK-{todo_id:03d}"
    due_date = today + timedelta(days=random.randint(-5, 30))
    day_of_week = ""  # Will be calculated
    task_type = random.choice(task_types)
    project_id = f"PRJ-{(i % 10) + 1:03d}"
    project_name = ""  # Will be calculated
    task = f"{task_type} {random.choice(['budget', 'report', 'document', 'analysis', 'plan'])}"
    assigned = random.choice(["John Smith", "Maria Garcia", "Hans Mueller", "Sophie Martin"])
    priority = random.choice(["High", "Medium", "Low"])
    status = random.choice(["Not Started", "In Progress", "Complete", "Blocked"])
    category = random.choice(["Finance", "Operations", "Strategy", "Legal"])

    todo_data.append([
        todo_id_str, due_date, day_of_week, task_type, project_id, project_name,
        task, assigned, "", priority, status, "", category, "",
        0, 0, 0, "", ""
    ])
    todo_id += 1

df_calendar = pd.DataFrame(todo_data, columns=[
    'Task_ID', 'Due_Date', 'Day_of_Week', 'Type', 'Project_ID', 'Project_Name',
    'Task', 'Assigned_To', 'Assigned_By', 'Priority', 'Status', 'Notes',
    'Category', 'Tags', 'Priority_Score', 'Days_Until', 'Urgency_Score',
    'Visual_Alert', 'Overdue_Flag'
])

# 6. Decision Log (10 rows)
decision_data = []
for i in range(1, 11):
    decision_id = f"DEC-{i:03d}"
    date_made = today - timedelta(days=random.randint(0, 60))
    project_id = f"PRJ-{i:03d}"
    decision = f"Decision about {random.choice(['budget', 'timeline', 'scope', 'resources', 'vendor'])}"
    made_by = random.choice(["CEO", "CTO", "Board", "PMO"])
    category = random.choice(["Strategic", "Financial", "Operational", "Technical"])
    impact = random.choice(["High", "Medium", "Low"])
    rationale = f"Based on {random.choice(['market analysis', 'budget constraints', 'strategic priorities'])}"
    alternatives = "Various options considered"
    review_date = date_made + timedelta(days=90)
    status = random.choice(["Made", "Pending", "Deferred"])

    decision_data.append([
        decision_id, date_made, project_id, decision, made_by, category,
        impact, rationale, alternatives, "", review_date, status,
        0, "", "", ""
    ])

df_decisions = pd.DataFrame(decision_data, columns=[
    'Decision_ID', 'Date', 'Project_ID', 'Decision', 'Made_By', 'Category',
    'Impact', 'Rationale', 'Alternatives', 'Notes', 'Review_Date', 'Status',
    'Days_Until_Review', 'Visual_Status', 'Project_Name', 'Review_Alert'
])

# 7. Risk Register (10 rows)
risk_data = []
risk_categories = ["Technical", "Financial", "Schedule", "Resource", "External"]

for i in range(1, 11):
    risk_id = f"RSK-{i:03d}"
    project_id = f"PRJ-{i:03d}"
    risk = f"{random.choice(risk_categories)} risk - {random.choice(['delays', 'budget overrun', 'resource shortage', 'vendor issues'])}"
    probability = random.randint(1, 5)
    impact = random.randint(1, 5)
    risk_score = probability * impact
    mitigation = "Mitigation strategy defined"
    owner = random.choice(["PMO", "Technical Lead", "Finance", "Operations"])
    status = random.choice(["Open", "Mitigated", "Closed"])
    last_review = today - timedelta(days=random.randint(0, 45))

    risk_data.append([
        risk_id, project_id, risk, probability, impact, risk_score,
        mitigation, owner, status, last_review, "", "", "",
        0, ""
    ])

df_risks = pd.DataFrame(risk_data, columns=[
    'Risk_ID', 'Project_ID', 'Risk', 'Probability', 'Impact', 'Risk_Score',
    'Mitigation', 'Owner', 'Status', 'Last_Review', 'Risk_Level',
    'Project_Name', 'Alert', 'Days_Since_Review', 'Review_Alert'
])

# 8. European Cities (Reference Data)
cities_data = []
for country in european_countries:
    # Add 2-3 major cities per country
    if country == "United Kingdom":
        cities_data.extend([
            [country, "London", 51.5074, -0.1278, 8900000, "Capital"],
            [country, "Manchester", 53.4808, -2.2426, 550000, "Major"],
            [country, "Birmingham", 52.4862, -1.8904, 1100000, "Major"]
        ])
    elif country == "Germany":
        cities_data.extend([
            [country, "Berlin", 52.5200, 13.4050, 3700000, "Capital"],
            [country, "Munich", 48.1351, 11.5820, 1500000, "Major"],
            [country, "Hamburg", 53.5511, 9.9937, 1900000, "Major"]
        ])
    elif country == "France":
        cities_data.extend([
            [country, "Paris", 48.8566, 2.3522, 2200000, "Capital"],
            [country, "Lyon", 45.7640, 4.8357, 516000, "Major"],
            [country, "Marseille", 43.2965, 5.3698, 870000, "Major"]
        ])
    else:
        # Add capital city for other countries
        cities_data.append([country, f"{country} Capital", 50.0, 10.0, 1000000, "Capital"])

df_cities = pd.DataFrame(cities_data, columns=[
    'Country', 'City', 'Latitude', 'Longitude', 'Population', 'Type'
])

# 9. Config Lists (Dropdown values)
config_data = {
    'List_Type': ['Status', 'Priority', 'Phase', 'Category', 'Risk Level'],
    'Value1': ['In Progress', 'Critical', 'Phase 1', 'Strategic', 'Critical'],
    'Value2': ['Planning', 'High', 'Phase 2', 'Operational', 'High'],
    'Value3': ['Complete', 'Medium', 'Phase 3', 'Technical', 'Medium'],
    'Value4': ['At Risk', 'Low', 'Phase 4', 'Financial', 'Low'],
    'Value5': ['On Hold', '', 'Phase 5', 'Market', 'Very Low']
}
df_config = pd.DataFrame(config_data)

# 10. Map Data (All European countries)
map_countries = [
    'United Kingdom', 'Ireland', 'France', 'Spain', 'Portugal', 'Germany', 'Netherlands',
    'Belgium', 'Luxembourg', 'Switzerland', 'Austria', 'Italy', 'Denmark', 'Sweden',
    'Norway', 'Finland', 'Poland', 'Czech Republic', 'Slovakia', 'Hungary', 'Romania',
    'Bulgaria', 'Greece', 'Croatia', 'Slovenia', 'Estonia', 'Latvia', 'Lithuania'
]

map_data = []
for country in map_countries:
    # Placeholder data - activity level will be calculated by formula
    map_data.append([country, 0, "", ""])

df_map = pd.DataFrame(map_data, columns=['Country', 'Activity_Level', 'Projects', 'Status'])

# ================ CREATE EXCEL SHEETS ================

# 1. Setup Sheet
ws_setup = wb.create_sheet("_SETUP")
ws_setup['A1'] = "PM DASHBOARD SETUP"
ws_setup['A1'].font = Font(size=16, bold=True)
ws_setup['A3'] = "Instructions:"
ws_setup['A4'] = "1. This is a complete PM Dashboard with all formulas"
ws_setup['A5'] = "2. Navigate using the Command Center tab"
ws_setup['A6'] = "3. Select projects in Project_Spotlight using dropdown in N1"
ws_setup['A7'] = "4. All data is interconnected with formulas"
ws_setup['A8'] = "5. Sample data includes 10 projects for testing"

# 2. Command Center
ws_command = wb.create_sheet("Command Center")

# Title and Portfolio Health
ws_command.merge_cells('A1:L1')
ws_command['A1'] = "PROJECT PORTFOLIO COMMAND CENTER"
ws_command['A1'].font = Font(size=18, bold=True)
ws_command['A1'].alignment = Alignment(horizontal="center")

# Portfolio Health Score (Large indicator)
ws_command.merge_cells('B3:D6')
ws_command['B3'] = '=IF(AVERAGE(Master_Projects_Clean!F:F)>0.7,"🟢",IF(AVERAGE(Master_Projects_Clean!F:F)>0.4,"🟡","🔴"))'
ws_command['B3'].font = Font(size=48)
ws_command['B3'].alignment = Alignment(horizontal="center", vertical="center")

ws_command['E3'] = '=CONCATENATE(ROUND(AVERAGE(Master_Projects_Clean!F:F)*100,0),"% On Track")'
ws_command['E3'].font = Font(size=14, bold=True)

# Smart Alerts
ws_command['I2'] = "⚠️ ALERTS"
ws_command['I2'].font = Font(size=12, bold=True)
ws_command['J3'] = '=IF(COUNTIF(Project_Milestones!I:I,"<0")>0,CONCATENATE(COUNTIF(Project_Milestones!I:I,"<0")," overdue milestones"),"")'
ws_command['J4'] = '=IF(COUNTIFS(Project_Country_Budgets!J:J,">0.9")>0,CONCATENATE(COUNTIFS(Project_Country_Budgets!J:J,">0.9")," countries >90% budget"),"")'
ws_command['J5'] = '=IF(COUNTIF(Decision_Log!L:L,"Pending")>0,CONCATENATE(COUNTIF(Decision_Log!L:L,"Pending")," pending decisions"),"")'
ws_command['J6'] = '=IF(COUNTIF(Risk_Register!I:I,"Open")>0,CONCATENATE(COUNTIF(Risk_Register!I:I,"Open")," open risks"),"")'

# Project Quick View Headers
ws_command['A11'] = "PROJECT QUICK VIEW"
ws_command['A11'].font = Font(size=12, bold=True)
headers = ['ID', 'Name', 'Status', 'Progress', 'Countries', 'Days Left']
create_styled_header(ws_command, 11, headers)

# Add formulas for project quick view (rows 12-21)
for i in range(10):
    row = 12 + i
    ws_command[f'B{row}'] = f'=Master_Projects_Clean!A{i+2}'
    ws_command[f'C{row}'] = f'=Master_Projects_Clean!B{i+2}'
    ws_command[f'D{row}'] = f'=Master_Projects_Clean!D{i+2}'
    ws_command[f'E{row}'] = f'=Master_Projects_Clean!F{i+2}'
    ws_command[f'F{row}'] = f'=Master_Projects_Clean!T{i+2}'
    ws_command[f'G{row}'] = f'=Master_Projects_Clean!I{i+2}'

# This Week Summary
ws_command['A30'] = "THIS WEEK"
ws_command['A30'].font = Font(size=12, bold=True)
ws_command['A31'] = "Tasks Due:"
ws_command['B31'] = '=COUNTIFS(Calendar_Todo!B:B,">="&TODAY(),Calendar_Todo!B:B,"<"&TODAY()+7)'
ws_command['C31'] = "Milestones:"
ws_command['D31'] = '=COUNTIFS(Project_Milestones!D:D,">="&TODAY(),Project_Milestones!D:D,"<"&TODAY()+7)'
ws_command['E31'] = "Events:"
ws_command['F31'] = '=COUNTIFS(Project_Events!G:G,">="&TODAY(),Project_Events!G:G,"<"&TODAY()+30)'

# 3. Project Spotlight
ws_spotlight = wb.create_sheet("Project_Spotlight")

# Project selector dropdown in N1
ws_spotlight['M1'] = "Select Project:"
ws_spotlight['N1'] = "PRJ-001"  # Default selection
ws_spotlight['N1'].font = Font(bold=True)

# Title
ws_spotlight.merge_cells('B2:L2')
ws_spotlight['B2'] = '=IFERROR(VLOOKUP($N$1,Master_Projects_Clean!$A:$C,2,FALSE),"Select Project")'
ws_spotlight['B2'].font = Font(size=16, bold=True)

# Summary
ws_spotlight['B5'] = '=IFERROR(VLOOKUP($N$1,Master_Projects_Clean!$A:$C,3,FALSE),"")'

# Status & Timeline Card
ws_spotlight['B9'] = "STATUS & TIMELINE"
ws_spotlight['B9'].font = Font(bold=True, color="FFFFFF")
ws_spotlight['B9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

status_labels = ['Status:', 'Priority:', 'Start Date:', 'End Date:', 'Progress:', 'Days Remaining:']
for i, label in enumerate(status_labels, 10):
    ws_spotlight[f'B{i}'] = label

ws_spotlight['C10'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$D,4,FALSE)'
ws_spotlight['C11'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$E,5,FALSE)'
ws_spotlight['C12'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$G,7,FALSE)'
ws_spotlight['C13'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$H,8,FALSE)'
ws_spotlight['C14'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$F,6,FALSE)'
ws_spotlight['C15'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$I,9,FALSE)'

# Team & Partners Card
ws_spotlight['F9'] = "TEAM & PARTNERS"
ws_spotlight['F9'].font = Font(bold=True, color="FFFFFF")
ws_spotlight['F9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

team_labels = ['Lead:', 'Sponsor:', 'Partner:']
for i, label in enumerate(team_labels, 10):
    ws_spotlight[f'F{i}'] = label

ws_spotlight['G10'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$J,10,FALSE)'
ws_spotlight['G11'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$K,11,FALSE)'
ws_spotlight['G12'] = '=VLOOKUP($N$1,Master_Projects_Clean!$A:$L,12,FALSE)'

# Budget Overview Card
ws_spotlight['J9'] = "BUDGET OVERVIEW"
ws_spotlight['J9'].font = Font(bold=True, color="FFFFFF")
ws_spotlight['J9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

budget_labels = ['Total Budget:', 'Spent:', 'Remaining:', '% Used:', 'Countries:']
for i, label in enumerate(budget_labels, 10):
    ws_spotlight[f'J{i}'] = label

ws_spotlight['K10'] = '=SUMIF(Project_Country_Budgets!$B:$B,$N$1,Project_Country_Budgets!$G:$G)'
ws_spotlight['K11'] = '=SUMIF(Project_Country_Budgets!$B:$B,$N$1,Project_Country_Budgets!$H:$H)'
ws_spotlight['K12'] = '=K10-K11'
ws_spotlight['K13'] = '=IF(K10>0,K11/K10,0)'
ws_spotlight['K14'] = '=COUNTIF(Project_Country_Budgets!$B:$B,$N$1)'

# Budget by Country Table
ws_spotlight['B18'] = "BUDGET BY COUNTRY"
ws_spotlight['B18'].font = Font(bold=True)
country_headers = ['Country', 'Phase', 'Budget', 'Spent', 'Remaining', '% Used']
for i, header in enumerate(country_headers):
    ws_spotlight.cell(row=19, column=2+i, value=header).font = Font(bold=True)

# Add INDEX formulas for first 5 countries
for i in range(5):
    row = 20 + i
    ws_spotlight[f'B{row}'] = f'=IFERROR(INDEX(Project_Country_Budgets!$D:$D,SMALL(IF(Project_Country_Budgets!$B:$B=$N$1,ROW(Project_Country_Budgets!$B:$B)),{i+1})),"")'
    ws_spotlight[f'C{row}'] = f'=IFERROR(INDEX(Project_Country_Budgets!$F:$F,SMALL(IF(Project_Country_Budgets!$B:$B=$N$1,ROW(Project_Country_Budgets!$B:$B)),{i+1})),"")'
    ws_spotlight[f'D{row}'] = f'=IFERROR(INDEX(Project_Country_Budgets!$G:$G,SMALL(IF(Project_Country_Budgets!$B:$B=$N$1,ROW(Project_Country_Budgets!$B:$B)),{i+1})),0)'
    ws_spotlight[f'E{row}'] = f'=IFERROR(INDEX(Project_Country_Budgets!$H:$H,SMALL(IF(Project_Country_Budgets!$B:$B=$N$1,ROW(Project_Country_Budgets!$B:$B)),{i+1})),0)'
    ws_spotlight[f'F{row}'] = f'=D{row}-E{row}'
    ws_spotlight[f'G{row}'] = f'=IF(D{row}>0,E{row}/D{row},0)'

# Milestone count
ws_spotlight['J19'] = "Active Milestones:"
ws_spotlight['K19'] = '=COUNTIFS(Project_Milestones!$C:$C,$N$1,Project_Milestones!$E:$E,"<>Complete")'

# 4. Master Projects Sheet
ws_projects = wb.create_sheet("Master_Projects_Clean")
# Write headers
headers = list(df_projects.columns)
create_styled_header(ws_projects, 1, headers)

# Write data
for r_idx, row in enumerate(df_projects.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_projects.cell(row=r_idx, column=c_idx, value=value)

# Add formulas for calculated columns
for row in range(2, 12):  # 10 projects
    ws_projects[f'I{row}'] = f'=H{row}-TODAY()'  # Days Remaining
    ws_projects[f'M{row}'] = f'=SUMIF(Project_Country_Budgets!$B:$B,A{row},Project_Country_Budgets!$G:$G)'  # Budget Total
    ws_projects[f'N{row}'] = f'=SUMIF(Project_Country_Budgets!$B:$B,A{row},Project_Country_Budgets!$H:$H)'  # Budget Spent
    ws_projects[f'O{row}'] = f'=IF(M{row}>0,N{row}/M{row},0)'  # Budget Percent
    ws_projects[f'S{row}'] = f'=COUNTIF(Project_Events!$B:$B,A{row})'  # Event Count
    ws_projects[f'T{row}'] = f'=COUNTIF(Project_Country_Budgets!$B:$B,A{row})'  # Countries Count
    ws_projects[f'V{row}'] = f'=IF(OR(D{row}="At Risk",I{row}<0,O{row}>0.9),"Red",IF(OR(I{row}<7,O{row}>0.75,F{row}<0.3),"Yellow","Green"))'  # Health Score
    ws_projects[f'W{row}'] = f'=IF(V{row}="Red","🔴",IF(V{row}="Yellow","🟡","🟢"))'  # Visual Health

# 5. Project Country Budgets Sheet
ws_budgets = wb.create_sheet("Project_Country_Budgets")
headers = list(df_budgets.columns)
create_styled_header(ws_budgets, 1, headers)

for r_idx, row in enumerate(df_budgets.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_budgets.cell(row=r_idx, column=c_idx, value=value)

# Add formulas
for row in range(2, len(df_budgets) + 2):
    ws_budgets[f'I{row}'] = f'=G{row}-H{row}'  # Budget Remaining
    ws_budgets[f'J{row}'] = f'=IF(G{row}>0,H{row}/G{row},0)'  # Percent Used
    ws_budgets[f'N{row}'] = f'=IF(J{row}>0.9,"Over Budget",IF(J{row}>0.75,"At Risk","On Track"))'  # Budget Status
    ws_budgets[f'O{row}'] = f'=L{row}-K{row}'  # Days in Phase
    ws_budgets[f'P{row}'] = f'=IF(M{row}="In Progress",TODAY()-K{row},0)'  # Days Active
    ws_budgets[f'Q{row}'] = f'=IF(AND(M{row}="In Progress",J{row}>P{row}/O{row}),"Spending Fast","")'  # Burn Rate Alert

# 6. Project Milestones Sheet
ws_milestones = wb.create_sheet("Project_Milestones")
headers = list(df_milestones.columns)
create_styled_header(ws_milestones, 1, headers)

for r_idx, row in enumerate(df_milestones.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_milestones.cell(row=r_idx, column=c_idx, value=value)

# Add formulas
for row in range(2, len(df_milestones) + 2):
    ws_milestones[f'B{row}'] = f'=CONCATENATE(H{row}," - ",C{row})'  # Automated Name
    ws_milestones[f'I{row}'] = f'=D{row}-TODAY()'  # Days Until
    ws_milestones[f'J{row}'] = f'=IF(I{row}<0,"Overdue",IF(I{row}<7,"This Week",IF(I{row}<30,"This Month","Future")))'  # Time Group
    ws_milestones[f'K{row}'] = f'=IF(I{row}<0,"🔴",IF(I{row}<7,"🟡",IF(I{row}<30,"🔵","⚪")))'  # Visual Status
    ws_milestones[f'L{row}'] = f'=VLOOKUP(C{row},Master_Projects_Clean!$A:$B,2,FALSE)'  # Project Name
    ws_milestones[f'M{row}'] = f'=IF(AND(E{row}<>"Complete",I{row}<0),"ACTION REQUIRED","")'  # Alert

# 7. Project Events Sheet
ws_events = wb.create_sheet("Project_Events")
headers = list(df_events.columns)
create_styled_header(ws_events, 1, headers)

for r_idx, row in enumerate(df_events.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_events.cell(row=r_idx, column=c_idx, value=value)

# 8. Calendar Todo Sheet
ws_calendar = wb.create_sheet("Calendar_Todo")
headers = list(df_calendar.columns)
create_styled_header(ws_calendar, 1, headers)

for r_idx, row in enumerate(df_calendar.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_calendar.cell(row=r_idx, column=c_idx, value=value)

# Add formulas
for row in range(2, len(df_calendar) + 2):
    ws_calendar[f'C{row}'] = f'=TEXT(B{row},"dddd")'  # Day of Week
    ws_calendar[f'F{row}'] = f'=VLOOKUP(E{row},Master_Projects_Clean!$A:$B,2,FALSE)'  # Project Name
    ws_calendar[f'O{row}'] = f'=IF(J{row}="High",10,IF(J{row}="Medium",5,1))'  # Priority Score
    ws_calendar[f'P{row}'] = f'=B{row}-TODAY()'  # Days Until
    ws_calendar[f'Q{row}'] = f'=O{row}+(P{row}*-1)'  # Urgency Score
    ws_calendar[f'R{row}'] = f'=IF(P{row}<0,"🔴 OVERDUE",IF(P{row}=0,"🟡 TODAY",IF(P{row}<=7,"🔵 THIS WEEK","")))'  # Visual Alert
    ws_calendar[f'S{row}'] = f'=IF(AND(K{row}<>"Complete",P{row}<0),"Yes","")'  # Overdue Flag

# 9. Decision Log Sheet
ws_decisions = wb.create_sheet("Decision_Log")
headers = list(df_decisions.columns)
create_styled_header(ws_decisions, 1, headers)

for r_idx, row in enumerate(df_decisions.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_decisions.cell(row=r_idx, column=c_idx, value=value)

# Add formulas
for row in range(2, len(df_decisions) + 2):
    ws_decisions[f'M{row}'] = f'=K{row}-B{row}'  # Days Until Review
    ws_decisions[f'N{row}'] = f'=IF(L{row}="Pending","🟡 PENDING",IF(L{row}="Reversed","🔴 REVERSED","✓"))'  # Visual Status
    ws_decisions[f'O{row}'] = f'=VLOOKUP(C{row},Master_Projects_Clean!$A:$B,2,FALSE)'  # Project Name
    ws_decisions[f'P{row}'] = f'=IF(AND(L{row}="Made",K{row}<TODAY()),"Review Due","")'  # Review Alert

# 10. Risk Register Sheet
ws_risks = wb.create_sheet("Risk_Register")
headers = list(df_risks.columns)
create_styled_header(ws_risks, 1, headers)

for r_idx, row in enumerate(df_risks.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_risks.cell(row=r_idx, column=c_idx, value=value)

# Add formulas
for row in range(2, len(df_risks) + 2):
    ws_risks[f'F{row}'] = f'=D{row}*E{row}'  # Risk Score
    ws_risks[f'K{row}'] = f'=IF(F{row}>=15,"🔴 Critical",IF(F{row}>=10,"🟡 High",IF(F{row}>=5,"🔵 Medium","⚪ Low")))'  # Risk Level
    ws_risks[f'L{row}'] = f'=VLOOKUP(B{row},Master_Projects_Clean!$A:$B,2,FALSE)'  # Project Name
    ws_risks[f'M{row}'] = f'=IF(AND(I{row}="Open",F{row}>=15),"IMMEDIATE ACTION","")'  # Alert
    ws_risks[f'N{row}'] = f'=TODAY()-J{row}'  # Days Since Last Review
    ws_risks[f'O{row}'] = f'=IF(N{row}>30,"Review Needed","")'  # Review Alert

# 11. European Cities Sheet
ws_cities = wb.create_sheet("European_Cities")
headers = list(df_cities.columns)
create_styled_header(ws_cities, 1, headers)

for r_idx, row in enumerate(df_cities.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_cities.cell(row=r_idx, column=c_idx, value=value)

# 12. Config Lists Sheet
ws_config = wb.create_sheet("Config_Lists")
headers = list(df_config.columns)
create_styled_header(ws_config, 1, headers)

for r_idx, row in enumerate(df_config.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_config.cell(row=r_idx, column=c_idx, value=value)

# 13. Map Data Sheet
ws_map = wb.create_sheet("Map_Data")
# Add headers
ws_map['A1'] = "Country"
ws_map['B1'] = "Activity Level"
ws_map['C1'] = "Projects"
ws_map['D1'] = "Status"

# Add dynamic selector reference
ws_map['F1'] = "Selected Project:"
ws_map['G1'] = "Activity Level"
create_styled_header(ws_map, 1, ['Country', 'Activity Level', 'Projects', 'Status'])

# Add countries
for r_idx, row in enumerate(df_map.itertuples(index=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws_map.cell(row=r_idx, column=c_idx, value=value)

# Add dynamic formulas for activity level
for row in range(2, len(df_map) + 2):
    ws_map[f'G{row}'] = f'=IF(COUNTIFS(Project_Events!$B:$B,Project_Spotlight!$N$1,Project_Events!$E:$E,F{row})>0,100,IF(COUNTIFS(Project_Country_Budgets!$B:$B,Project_Spotlight!$N$1,Project_Country_Budgets!$D:$D,F{row})>0,50,0))'

# Apply auto-sizing to all sheets
for sheet in wb.worksheets:
    autosize_columns(sheet)

# Save the workbook
output_file = "C:\\Projects\\OSINT - Foresight\\PM_Dashboard_Complete.xlsx"
wb.save(output_file)
print(f"Dashboard created successfully: {output_file}")
print(f"Total sheets created: {len(wb.worksheets)}")
print("All formulas and sample data included!")
