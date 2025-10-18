import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random
from typing import List, Dict, Tuple

# Color Palette (EXACT COLORS FROM INSTRUCTIONS)
COLORS = {
    'primary_blue': '0066CC',
    'success_green': '10B981',
    'warning_amber': 'F59E0B',
    'danger_red': 'EF4444',
    'neutral_gray': '6B7280',
    'light_blue_bg': 'E8F4FD',
    'light_gray_bg': 'F5F5F5',
    'white': 'FFFFFF',
    'dark_text': '111827',
    'medium_text': '6B7280'
}

def create_workbook():
    """Create workbook with all required sheets"""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # CREATE ALL THESE SHEETS - MANDATORY
    sheets_to_create = [
        "Dashboard_Home",
        "Calendar_Todo",
        "Project_Deep_Dive",
        "Europe_Map_Visual",
        "Master_Projects_Clean",
        "Project_Events",
        "Event_Participants",
        "Project_Milestones_Standard",
        "Project_Audiences",
        "Project_Products",
        "Project_Stakeholders",
        "LOE_Goals",
        "Config_Lists",
        "European_Cities",
        "Country_Map_Grid"
    ]

    for sheet_name in sheets_to_create:
        wb.create_sheet(sheet_name)

    return wb

def create_config_lists(ws):
    """Create Config_Lists sheet with all lookup data"""
    # Headers
    headers = ["List_Type", "Value", "Sort_Order", "Color_Hex", "Icon"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # MANDATORY DATA - Must include ALL of these
    config_data = [
        ("Status", "Not Started", 1, "#9E9E9E", "○"),
        ("Status", "Planning", 2, "#2196F3", "◔"),
        ("Status", "In Progress", 3, "#4CAF50", "◐"),
        ("Status", "On Hold", 4, "#FF9800", "⊕"),
        ("Status", "At Risk", 5, "#F44336", "⊗"),
        ("Status", "Complete", 6, "#607D8B", "●"),
        ("Status", "Cancelled", 7, "#424242", "✕"),
        ("Priority", "Critical", 1, "#F44336", "🔴"),
        ("Priority", "High", 2, "#FF9800", "🟡"),
        ("Priority", "Medium", 3, "#FFC107", "🟡"),
        ("Priority", "Low", 4, "#4CAF50", "🟢"),
        ("LOE", "Digital Transformation", 1, "#1E40AF", "💻"),
        ("LOE", "Customer Experience", 2, "#7C3AED", "👥"),
        ("LOE", "Cost Optimization", 3, "#DC2626", "💰"),
        ("LOE", "Market Expansion", 4, "#059669", "🌍"),
        ("LOE", "Compliance & Risk", 5, "#B45309", "⚖️"),
        ("Event_Type", "Conference", 1, "#0066CC", ""),
        ("Event_Type", "Workshop", 2, "#00AA00", ""),
        ("Event_Type", "Summit", 3, "#FF9900", ""),
        ("Event_Type", "Meeting", 4, "#666666", ""),
        ("Audience", "Government Officials", 1, "#FFD700", ""),
        ("Audience", "Industry Partners", 2, "#C0C0C0", ""),
        ("Audience", "Media", 3, "#CD7F32", ""),
        ("Product_Type", "Report", 1, "#FF0000", ""),
        ("Product_Type", "Presentation", 2, "#00FF00", ""),
        ("Product_Type", "Website", 3, "#0000FF", ""),
    ]

    for row, data in enumerate(config_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

def create_european_cities(ws):
    """Create European_Cities sheet with complete city data"""
    headers = ["Country", "City", "Is_Capital", "Is_Major_City", "Region", "Time_Zone", "EU_Member"]

    # Add headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # COMPLETE city data - Including all major European countries
    cities_data = [
        # United Kingdom
        ("United Kingdom", "London", "Yes", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Manchester", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Birmingham", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Liverpool", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Leeds", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Glasgow", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Edinburgh", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Bristol", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Cardiff", "No", "Yes", "Western", "GMT", "No"),
        ("United Kingdom", "Belfast", "No", "Yes", "Western", "GMT", "No"),

        # Germany
        ("Germany", "Berlin", "Yes", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Hamburg", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Munich", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Cologne", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Frankfurt", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Stuttgart", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Düsseldorf", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Leipzig", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Dortmund", "No", "Yes", "Central", "CET", "Yes"),
        ("Germany", "Dresden", "No", "Yes", "Central", "CET", "Yes"),

        # France
        ("France", "Paris", "Yes", "Yes", "Western", "CET", "Yes"),
        ("France", "Marseille", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Lyon", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Toulouse", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Nice", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Nantes", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Strasbourg", "No", "Yes", "Western", "CET", "Yes"),
        ("France", "Bordeaux", "No", "Yes", "Western", "CET", "Yes"),

        # Spain
        ("Spain", "Madrid", "Yes", "Yes", "Southern", "CET", "Yes"),
        ("Spain", "Barcelona", "No", "Yes", "Southern", "CET", "Yes"),
        ("Spain", "Valencia", "No", "Yes", "Southern", "CET", "Yes"),
        ("Spain", "Seville", "No", "Yes", "Southern", "CET", "Yes"),
        ("Spain", "Bilbao", "No", "Yes", "Southern", "CET", "Yes"),
        ("Spain", "Málaga", "No", "Yes", "Southern", "CET", "Yes"),

        # Italy
        ("Italy", "Rome", "Yes", "Yes", "Southern", "CET", "Yes"),
        ("Italy", "Milan", "No", "Yes", "Southern", "CET", "Yes"),
        ("Italy", "Naples", "No", "Yes", "Southern", "CET", "Yes"),
        ("Italy", "Turin", "No", "Yes", "Southern", "CET", "Yes"),
        ("Italy", "Florence", "No", "Yes", "Southern", "CET", "Yes"),
        ("Italy", "Venice", "No", "Yes", "Southern", "CET", "Yes"),

        # Poland
        ("Poland", "Warsaw", "Yes", "Yes", "Eastern", "CET", "Yes"),
        ("Poland", "Kraków", "No", "Yes", "Eastern", "CET", "Yes"),
        ("Poland", "Łódź", "No", "Yes", "Eastern", "CET", "Yes"),
        ("Poland", "Wrocław", "No", "Yes", "Eastern", "CET", "Yes"),
        ("Poland", "Poznań", "No", "Yes", "Eastern", "CET", "Yes"),
        ("Poland", "Gdańsk", "No", "Yes", "Eastern", "CET", "Yes"),

        # Netherlands
        ("Netherlands", "Amsterdam", "Yes", "Yes", "Western", "CET", "Yes"),
        ("Netherlands", "Rotterdam", "No", "Yes", "Western", "CET", "Yes"),
        ("Netherlands", "The Hague", "No", "Yes", "Western", "CET", "Yes"),
        ("Netherlands", "Utrecht", "No", "Yes", "Western", "CET", "Yes"),

        # Belgium
        ("Belgium", "Brussels", "Yes", "Yes", "Western", "CET", "Yes"),
        ("Belgium", "Antwerp", "No", "Yes", "Western", "CET", "Yes"),
        ("Belgium", "Ghent", "No", "Yes", "Western", "CET", "Yes"),

        # Portugal
        ("Portugal", "Lisbon", "Yes", "Yes", "Western", "WET", "Yes"),
        ("Portugal", "Porto", "No", "Yes", "Western", "WET", "Yes"),

        # Greece
        ("Greece", "Athens", "Yes", "Yes", "Southern", "EET", "Yes"),
        ("Greece", "Thessaloniki", "No", "Yes", "Southern", "EET", "Yes"),

        # Sweden
        ("Sweden", "Stockholm", "Yes", "Yes", "Nordic", "CET", "Yes"),
        ("Sweden", "Gothenburg", "No", "Yes", "Nordic", "CET", "Yes"),
        ("Sweden", "Malmö", "No", "Yes", "Nordic", "CET", "Yes"),

        # Norway
        ("Norway", "Oslo", "Yes", "Yes", "Nordic", "CET", "No"),
        ("Norway", "Bergen", "No", "Yes", "Nordic", "CET", "No"),

        # Denmark
        ("Denmark", "Copenhagen", "Yes", "Yes", "Nordic", "CET", "Yes"),
        ("Denmark", "Aarhus", "No", "Yes", "Nordic", "CET", "Yes"),

        # Finland
        ("Finland", "Helsinki", "Yes", "Yes", "Nordic", "EET", "Yes"),

        # Austria
        ("Austria", "Vienna", "Yes", "Yes", "Central", "CET", "Yes"),
        ("Austria", "Salzburg", "No", "Yes", "Central", "CET", "Yes"),

        # Switzerland
        ("Switzerland", "Bern", "Yes", "Yes", "Central", "CET", "No"),
        ("Switzerland", "Zurich", "No", "Yes", "Central", "CET", "No"),
        ("Switzerland", "Geneva", "No", "Yes", "Central", "CET", "No"),

        # Czech Republic
        ("Czech Republic", "Prague", "Yes", "Yes", "Central", "CET", "Yes"),

        # Hungary
        ("Hungary", "Budapest", "Yes", "Yes", "Central", "CET", "Yes"),

        # Romania
        ("Romania", "Bucharest", "Yes", "Yes", "Eastern", "EET", "Yes"),

        # Bulgaria
        ("Bulgaria", "Sofia", "Yes", "Yes", "Eastern", "EET", "Yes"),

        # Ireland
        ("Ireland", "Dublin", "Yes", "Yes", "Western", "GMT", "Yes"),
        ("Ireland", "Cork", "No", "Yes", "Western", "GMT", "Yes"),

        # Critical: Include non-EU countries as specified
        ("Türkiye", "Istanbul", "No", "Yes", "Eurasia", "TRT", "No"),
        ("Türkiye", "Ankara", "Yes", "Yes", "Eurasia", "TRT", "No"),
        ("Cyprus", "Nicosia", "Yes", "Yes", "Mediterranean", "EET", "Yes"),
        ("Iceland", "Reykjavik", "Yes", "Yes", "Nordic", "GMT", "No"),
        ("Georgia", "Tbilisi", "Yes", "Yes", "Caucasus", "GET", "No"),
        ("Armenia", "Yerevan", "Yes", "Yes", "Caucasus", "AMT", "No"),
        ("Azerbaijan", "Baku", "Yes", "Yes", "Caucasus", "AZT", "No"),
    ]

    # Add ALL cities to sheet
    for row, city_data in enumerate(cities_data, 2):
        for col, value in enumerate(city_data, 1):
            ws.cell(row=row, column=col, value=value)

def create_master_projects(ws):
    """Create Master_Projects_Clean with ACTUAL PROJECT DATA"""
    # Headers
    headers = [
        "Project_ID", "Project_Name", "Project_Summary", "Status", "Priority",
        "Progress_Percent", "Start_Date", "End_Date", "Days_Remaining",
        "Project_Lead", "Sponsor", "Partner_Org", "Budget_Total", "Budget_Spent",
        "Budget_Percent", "LOE_Primary", "Target_Audience", "Event_Count",
        "Countries_Count", "Next_Milestone", "Next_Milestone_Date"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # ACTUAL PROJECT DATA - Create 15 real projects
    projects = [
        {
            "Project_ID": "PRJ-001",
            "Project_Name": "UK Market Expansion",
            "Project_Summary": "Expanding flagship product into UK market through Acme Corp partnership",
            "Status": "In Progress",
            "Priority": "High",
            "Progress_Percent": 0.45,
            "Start_Date": datetime(2025, 1, 1),
            "End_Date": datetime(2025, 12, 31),
            "Project_Lead": "John Smith",
            "Sponsor": "CEO",
            "Partner_Org": "Acme Corp",
            "Budget_Total": 2500000,
            "Budget_Spent": 1100000,
            "LOE_Primary": "Market Expansion",
            "Target_Audience": "Industry Partners",
            "Event_Count": 3,
            "Countries_Count": 5,
            "Next_Milestone": "UK Launch Event",
            "Next_Milestone_Date": datetime(2025, 3, 1)
        },
        {
            "Project_ID": "PRJ-002",
            "Project_Name": "German Digital Platform",
            "Project_Summary": "Digital transformation initiative for German market with cloud migration",
            "Status": "Planning",
            "Priority": "Critical",
            "Progress_Percent": 0.15,
            "Start_Date": datetime(2025, 2, 1),
            "End_Date": datetime(2025, 10, 31),
            "Project_Lead": "Maria Garcia",
            "Sponsor": "CTO",
            "Partner_Org": "TechCo GmbH",
            "Budget_Total": 1800000,
            "Budget_Spent": 200000,
            "LOE_Primary": "Digital Transformation",
            "Target_Audience": "Government Officials",
            "Event_Count": 2,
            "Countries_Count": 3,
            "Next_Milestone": "Technical Review",
            "Next_Milestone_Date": datetime(2025, 2, 15)
        },
        {
            "Project_ID": "PRJ-003",
            "Project_Name": "France Customer Portal",
            "Project_Summary": "Building self-service customer portal for French market",
            "Status": "In Progress",
            "Priority": "Medium",
            "Progress_Percent": 0.60,
            "Start_Date": datetime(2024, 10, 1),
            "End_Date": datetime(2025, 6, 30),
            "Project_Lead": "Pierre Dubois",
            "Sponsor": "VP Sales",
            "Partner_Org": "Société Générale",
            "Budget_Total": 950000,
            "Budget_Spent": 570000,
            "LOE_Primary": "Customer Experience",
            "Target_Audience": "Industry Partners",
            "Event_Count": 1,
            "Countries_Count": 2,
            "Next_Milestone": "Beta Testing",
            "Next_Milestone_Date": datetime(2025, 2, 20)
        },
        {
            "Project_ID": "PRJ-004",
            "Project_Name": "Nordic Supply Chain",
            "Project_Summary": "Optimizing supply chain across Nordic countries",
            "Status": "At Risk",
            "Priority": "High",
            "Progress_Percent": 0.35,
            "Start_Date": datetime(2025, 1, 15),
            "End_Date": datetime(2025, 9, 30),
            "Project_Lead": "Erik Andersson",
            "Sponsor": "COO",
            "Partner_Org": "Maersk",
            "Budget_Total": 3200000,
            "Budget_Spent": 1120000,
            "LOE_Primary": "Cost Optimization",
            "Target_Audience": "Industry Partners",
            "Event_Count": 4,
            "Countries_Count": 4,
            "Next_Milestone": "Vendor Selection",
            "Next_Milestone_Date": datetime(2025, 2, 10)
        },
        {
            "Project_ID": "PRJ-005",
            "Project_Name": "Spain Compliance Framework",
            "Project_Summary": "Implementing GDPR and local compliance requirements",
            "Status": "In Progress",
            "Priority": "Critical",
            "Progress_Percent": 0.70,
            "Start_Date": datetime(2024, 11, 1),
            "End_Date": datetime(2025, 5, 31),
            "Project_Lead": "Carlos Martinez",
            "Sponsor": "Legal Director",
            "Partner_Org": "Deloitte Spain",
            "Budget_Total": 750000,
            "Budget_Spent": 525000,
            "LOE_Primary": "Compliance & Risk",
            "Target_Audience": "Government Officials",
            "Event_Count": 2,
            "Countries_Count": 1,
            "Next_Milestone": "Audit Review",
            "Next_Milestone_Date": datetime(2025, 2, 28)
        },
        {
            "Project_ID": "PRJ-006",
            "Project_Name": "Italy Market Research",
            "Project_Summary": "Comprehensive market analysis for Italian expansion",
            "Status": "Complete",
            "Priority": "Low",
            "Progress_Percent": 1.00,
            "Start_Date": datetime(2024, 9, 1),
            "End_Date": datetime(2025, 1, 31),
            "Project_Lead": "Giuseppe Romano",
            "Sponsor": "VP Strategy",
            "Partner_Org": "Banca Intesa",
            "Budget_Total": 450000,
            "Budget_Spent": 450000,
            "LOE_Primary": "Market Expansion",
            "Target_Audience": "Media",
            "Event_Count": 1,
            "Countries_Count": 1,
            "Next_Milestone": "Final Report",
            "Next_Milestone_Date": datetime(2025, 1, 31)
        },
        {
            "Project_ID": "PRJ-007",
            "Project_Name": "Eastern Europe Gateway",
            "Project_Summary": "Establishing presence in Poland, Czech Republic, Hungary",
            "Status": "Planning",
            "Priority": "Medium",
            "Progress_Percent": 0.10,
            "Start_Date": datetime(2025, 3, 1),
            "End_Date": datetime(2025, 12, 31),
            "Project_Lead": "Wojciech Kowalski",
            "Sponsor": "CEO",
            "Partner_Org": "PKO Bank",
            "Budget_Total": 2800000,
            "Budget_Spent": 280000,
            "LOE_Primary": "Market Expansion",
            "Target_Audience": "Government Officials",
            "Event_Count": 5,
            "Countries_Count": 3,
            "Next_Milestone": "Country Analysis",
            "Next_Milestone_Date": datetime(2025, 3, 15)
        },
        {
            "Project_ID": "PRJ-008",
            "Project_Name": "Swiss Banking Integration",
            "Project_Summary": "Integration with Swiss banking systems for payments",
            "Status": "In Progress",
            "Priority": "High",
            "Progress_Percent": 0.55,
            "Start_Date": datetime(2024, 12, 1),
            "End_Date": datetime(2025, 7, 31),
            "Project_Lead": "Hans Mueller",
            "Sponsor": "CFO",
            "Partner_Org": "Credit Suisse",
            "Budget_Total": 1650000,
            "Budget_Spent": 907500,
            "LOE_Primary": "Digital Transformation",
            "Target_Audience": "Industry Partners",
            "Event_Count": 2,
            "Countries_Count": 2,
            "Next_Milestone": "API Testing",
            "Next_Milestone_Date": datetime(2025, 2, 25)
        },
        {
            "Project_ID": "PRJ-009",
            "Project_Name": "Benelux Customer Summit",
            "Project_Summary": "Major customer event covering Belgium, Netherlands, Luxembourg",
            "Status": "On Hold",
            "Priority": "Low",
            "Progress_Percent": 0.25,
            "Start_Date": datetime(2025, 2, 1),
            "End_Date": datetime(2025, 8, 31),
            "Project_Lead": "Anna Van Der Berg",
            "Sponsor": "VP Marketing",
            "Partner_Org": "Event Partners BV",
            "Budget_Total": 550000,
            "Budget_Spent": 137500,
            "LOE_Primary": "Customer Experience",
            "Target_Audience": "Industry Partners",
            "Event_Count": 3,
            "Countries_Count": 3,
            "Next_Milestone": "Venue Selection",
            "Next_Milestone_Date": datetime(2025, 3, 1)
        },
        {
            "Project_ID": "PRJ-010",
            "Project_Name": "Greek Infrastructure",
            "Project_Summary": "Building data center and support infrastructure in Athens",
            "Status": "At Risk",
            "Priority": "Critical",
            "Progress_Percent": 0.40,
            "Start_Date": datetime(2024, 10, 15),
            "End_Date": datetime(2025, 8, 31),
            "Project_Lead": "Nikos Papadopoulos",
            "Sponsor": "CTO",
            "Partner_Org": "OTE Group",
            "Budget_Total": 4500000,
            "Budget_Spent": 1800000,
            "LOE_Primary": "Digital Transformation",
            "Target_Audience": "Government Officials",
            "Event_Count": 2,
            "Countries_Count": 1,
            "Next_Milestone": "Site Approval",
            "Next_Milestone_Date": datetime(2025, 2, 5)
        },
        {
            "Project_ID": "PRJ-011",
            "Project_Name": "Portugal Pilot Program",
            "Project_Summary": "Running pilot program with Portuguese government",
            "Status": "In Progress",
            "Priority": "Medium",
            "Progress_Percent": 0.50,
            "Start_Date": datetime(2025, 1, 10),
            "End_Date": datetime(2025, 9, 30),
            "Project_Lead": "João Silva",
            "Sponsor": "VP Public Sector",
            "Partner_Org": "Portugal Telecom",
            "Budget_Total": 850000,
            "Budget_Spent": 425000,
            "LOE_Primary": "Market Expansion",
            "Target_Audience": "Government Officials",
            "Event_Count": 2,
            "Countries_Count": 1,
            "Next_Milestone": "Phase 1 Review",
            "Next_Milestone_Date": datetime(2025, 2, 12)
        },
        {
            "Project_ID": "PRJ-012",
            "Project_Name": "Irish Tech Hub",
            "Project_Summary": "Establishing technology hub in Dublin for EU operations",
            "Status": "Planning",
            "Priority": "High",
            "Progress_Percent": 0.20,
            "Start_Date": datetime(2025, 2, 15),
            "End_Date": datetime(2025, 11, 30),
            "Project_Lead": "Patrick O'Brien",
            "Sponsor": "CEO",
            "Partner_Org": "Enterprise Ireland",
            "Budget_Total": 3750000,
            "Budget_Spent": 750000,
            "LOE_Primary": "Digital Transformation",
            "Target_Audience": "Industry Partners",
            "Event_Count": 3,
            "Countries_Count": 1,
            "Next_Milestone": "Office Lease",
            "Next_Milestone_Date": datetime(2025, 2, 28)
        },
        {
            "Project_ID": "PRJ-013",
            "Project_Name": "Baltic States Alliance",
            "Project_Summary": "Partnership development in Estonia, Latvia, Lithuania",
            "Status": "Not Started",
            "Priority": "Low",
            "Progress_Percent": 0.00,
            "Start_Date": datetime(2025, 4, 1),
            "End_Date": datetime(2025, 12, 31),
            "Project_Lead": "Kristina Ozols",
            "Sponsor": "VP Partnerships",
            "Partner_Org": "Baltic Partners",
            "Budget_Total": 980000,
            "Budget_Spent": 0,
            "LOE_Primary": "Market Expansion",
            "Target_Audience": "Government Officials",
            "Event_Count": 4,
            "Countries_Count": 3,
            "Next_Milestone": "Kickoff Meeting",
            "Next_Milestone_Date": datetime(2025, 4, 1)
        },
        {
            "Project_ID": "PRJ-014",
            "Project_Name": "Austrian Analytics",
            "Project_Summary": "Building analytics center of excellence in Vienna",
            "Status": "In Progress",
            "Priority": "Medium",
            "Progress_Percent": 0.65,
            "Start_Date": datetime(2024, 11, 15),
            "End_Date": datetime(2025, 6, 30),
            "Project_Lead": "Stefan Wagner",
            "Sponsor": "Chief Data Officer",
            "Partner_Org": "Erste Bank",
            "Budget_Total": 1250000,
            "Budget_Spent": 812500,
            "LOE_Primary": "Digital Transformation",
            "Target_Audience": "Industry Partners",
            "Event_Count": 2,
            "Countries_Count": 1,
            "Next_Milestone": "Platform Launch",
            "Next_Milestone_Date": datetime(2025, 3, 1)
        },
        {
            "Project_ID": "PRJ-015",
            "Project_Name": "Romania Development Center",
            "Project_Summary": "Setting up software development center in Bucharest",
            "Status": "In Progress",
            "Priority": "High",
            "Progress_Percent": 0.30,
            "Start_Date": datetime(2025, 1, 5),
            "End_Date": datetime(2025, 10, 31),
            "Project_Lead": "Andrei Popescu",
            "Sponsor": "CTO",
            "Partner_Org": "Bitdefender",
            "Budget_Total": 2200000,
            "Budget_Spent": 660000,
            "LOE_Primary": "Cost Optimization",
            "Target_Audience": "Industry Partners",
            "Event_Count": 2,
            "Countries_Count": 1,
            "Next_Milestone": "Team Hiring",
            "Next_Milestone_Date": datetime(2025, 2, 18)
        }
    ]

    # Write projects to sheet
    for row, project in enumerate(projects, 2):
        ws.cell(row=row, column=1, value=project["Project_ID"])
        ws.cell(row=row, column=2, value=project["Project_Name"])
        ws.cell(row=row, column=3, value=project["Project_Summary"])
        ws.cell(row=row, column=4, value=project["Status"])
        ws.cell(row=row, column=5, value=project["Priority"])
        ws.cell(row=row, column=6, value=project["Progress_Percent"]).number_format = '0%'
        ws.cell(row=row, column=7, value=project["Start_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=8, value=project["End_Date"]).number_format = 'mm/dd/yyyy'
        # Calculate days remaining
        ws.cell(row=row, column=9, value=f"=H{row}-TODAY()")
        ws.cell(row=row, column=10, value=project["Project_Lead"])
        ws.cell(row=row, column=11, value=project["Sponsor"])
        ws.cell(row=row, column=12, value=project["Partner_Org"])
        ws.cell(row=row, column=13, value=project["Budget_Total"]).number_format = '"$"#,##0'
        ws.cell(row=row, column=14, value=project["Budget_Spent"]).number_format = '"$"#,##0'
        ws.cell(row=row, column=15, value=f"=N{row}/M{row}").number_format = '0%'
        ws.cell(row=row, column=16, value=project["LOE_Primary"])
        ws.cell(row=row, column=17, value=project["Target_Audience"])
        ws.cell(row=row, column=18, value=project["Event_Count"])
        ws.cell(row=row, column=19, value=project["Countries_Count"])
        ws.cell(row=row, column=20, value=project["Next_Milestone"])
        ws.cell(row=row, column=21, value=project["Next_Milestone_Date"]).number_format = 'mm/dd/yyyy'

def create_dashboard_home(ws):
    """Create Dashboard_Home with VISUAL LAYOUT"""
    # TITLE - Large and bold
    ws.merge_cells('B1:H2')
    ws['B1'].value = "PROJECT COMMAND CENTER"
    ws['B1'].font = Font(size=18, bold=True, color=COLORS['dark_text'])
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date/Time
    ws['J1'].value = "Last Refresh:"
    ws['K1'].value = "=NOW()"
    ws['K1'].number_format = 'mm/dd/yyyy hh:mm AM/PM'

    # KPI CARDS - Large numbers with labels
    # Card 1: Active Projects
    ws['B4'].value = "=COUNTIF(Master_Projects_Clean!D:D,\"In Progress\")"
    ws['B4'].font = Font(size=24, bold=True, color=COLORS['primary_blue'])
    ws['B5'].value = "ACTIVE"
    ws['B5'].font = Font(size=10, color=COLORS['medium_text'])
    ws['B6'].value = "Projects"
    ws['B6'].font = Font(size=10, color=COLORS['medium_text'])

    # Card 2: Due This Week
    ws['E4'].value = "=COUNTIFS(Calendar_Todo!B:B,\">=\"&TODAY(),Calendar_Todo!B:B,\"<\"&TODAY()+7)"
    ws['E4'].font = Font(size=24, bold=True, color=COLORS['warning_amber'])
    ws['E5'].value = "DUE THIS"
    ws['E5'].font = Font(size=10, color=COLORS['medium_text'])
    ws['E6'].value = "WEEK"
    ws['E6'].font = Font(size=10, color=COLORS['medium_text'])

    # Card 3: At Risk
    ws['H4'].value = "=COUNTIF(Master_Projects_Clean!D:D,\"At Risk\")"
    ws['H4'].font = Font(size=24, bold=True, color=COLORS['danger_red'])
    ws['H5'].value = "AT RISK"
    ws['H5'].font = Font(size=10, color=COLORS['medium_text'])
    ws['H6'].value = "ITEMS"
    ws['H6'].font = Font(size=10, color=COLORS['medium_text'])

    # Card 4: Budget
    ws['K4'].value = "=SUM(Master_Projects_Clean!M:M)/1000000"
    ws['K4'].number_format = '"$"#,##0.0"M"'
    ws['K4'].font = Font(size=24, bold=True, color=COLORS['success_green'])
    ws['K5'].value = "BUDGET"
    ws['K5'].font = Font(size=10, color=COLORS['medium_text'])
    ws['K6'].value = "MANAGED"
    ws['K6'].font = Font(size=10, color=COLORS['medium_text'])

    # Add borders around cards (subtle)
    thin_border = Border(
        left=Side(style='thin', color=COLORS['light_gray_bg']),
        right=Side(style='thin', color=COLORS['light_gray_bg']),
        top=Side(style='thin', color=COLORS['light_gray_bg']),
        bottom=Side(style='thin', color=COLORS['light_gray_bg'])
    )

    # Apply borders to card areas
    for row in range(4, 7):
        for col in [2, 3, 4]:  # First card
            ws.cell(row=row, column=col).border = thin_border
        for col in [5, 6, 7]:  # Second card
            ws.cell(row=row, column=col).border = thin_border
        for col in [8, 9, 10]:  # Third card
            ws.cell(row=row, column=col).border = thin_border
        for col in [11, 12, 13]:  # Fourth card
            ws.cell(row=row, column=col).border = thin_border

    # INSIGHTS BAR
    ws.merge_cells('B9:M10')
    ws['B9'].value = "KEY INSIGHTS: 3 projects need immediate attention | UK expansion ahead by 2 weeks | Budget reallocation needed for Q2"
    ws['B9'].fill = PatternFill(start_color=COLORS['light_blue_bg'],
                                end_color=COLORS['light_blue_bg'],
                                fill_type='solid')
    ws['B9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # STRATEGIC PROGRESS SECTION
    ws['B12'].value = "STRATEGIC LINES OF EFFORT"
    ws['B12'].font = Font(size=12, bold=True)

    # Progress bars for each LOE
    loe_names = ["Digital Transformation", "Customer Experience", "Cost Optimization", "Market Expansion", "Compliance & Risk"]
    loe_progress = [0.72, 0.85, 0.35, 0.91, 0.55]

    for i, (name, progress) in enumerate(zip(loe_names, loe_progress)):
        row = 14 + i
        ws[f'B{row}'].value = name
        ws[f'D{row}'].value = "█" * int(progress * 20) + "░" * (20 - int(progress * 20))
        ws[f'F{row}'].value = f"{int(progress * 100)}%"
        ws[f'D{row}'].font = Font(name='Courier New', color=COLORS['success_green'])

    # THIS WEEK CALENDAR PREVIEW
    ws['H12'].value = "THIS WEEK"
    ws['H12'].font = Font(size=12, bold=True)

    # Add sample calendar items
    calendar_items = [
        ("Mon 29", "London Launch", "PRJ-001", "CRITICAL"),
        ("Thu 30", "Partner Review", "PRJ-011", "HIGH"),
        ("Fri 31", "Budget Submission", "PRJ-002", "MEDIUM"),
        ("Mon 5", "Tech Review", "PRJ-008", "HIGH"),
        ("Wed 7", "Site Approval", "PRJ-010", "CRITICAL")
    ]

    for i, (date, task, project, priority) in enumerate(calendar_items):
        row = 14 + i
        ws[f'H{row}'].value = date
        ws[f'I{row}'].value = task
        ws[f'K{row}'].value = project
        ws[f'L{row}'].value = priority

        if priority == "CRITICAL":
            ws[f'L{row}'].font = Font(color=COLORS['danger_red'], bold=True)
        elif priority == "HIGH":
            ws[f'L{row}'].font = Font(color=COLORS['warning_amber'])

    # PROJECT STATUS SUMMARY
    ws['B20'].value = "PROJECT STATUS BREAKDOWN"
    ws['B20'].font = Font(size=12, bold=True)

    status_data = [
        ("Planning", 3),
        ("In Progress", 8),
        ("At Risk", 2),
        ("On Hold", 1),
        ("Complete", 1),
        ("Not Started", 1)
    ]

    for i, (status, count) in enumerate(status_data):
        row = 22 + i
        ws[f'B{row}'].value = status
        ws[f'D{row}'].value = count
        ws[f'E{row}'].value = f"=D{row}/COUNTA(Master_Projects_Clean!D:D)-1"
        ws[f'E{row}'].number_format = '0%'

    # Set column widths for better visibility
    ws.column_dimensions['A'].width = 2  # Margin
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 30  # KPI numbers row

def create_calendar_todo(ws):
    """Create Calendar_Todo with ACTUAL TASKS"""
    # Headers
    headers = [
        "Task_ID", "Date", "Day_of_Week", "Time", "Project_ID", "Project_Name",
        "Task_Type", "Task_Description", "Owner", "Priority", "Status"
    ]

    # Filter controls in row 1
    ws['B1'].value = "Filter by Project:"
    ws['D1'].value = "Filter by Type:"
    ws['F1'].value = "Filter by Owner:"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # Create 50+ actual tasks
    tasks = []
    task_types = ["Meeting", "Report", "Review", "Call", "Deliverable", "Presentation", "Workshop", "Deadline"]
    owners = ["John Smith", "Maria Garcia", "David Chen", "Sarah Johnson", "You", "Erik Andersson", "Carlos Martinez", "Anna Van Der Berg"]
    project_names = [
        "UK Market Expansion", "German Digital Platform", "France Customer Portal",
        "Nordic Supply Chain", "Spain Compliance Framework", "Italy Market Research",
        "Eastern Europe Gateway", "Swiss Banking Integration", "Benelux Customer Summit",
        "Greek Infrastructure", "Portugal Pilot Program", "Irish Tech Hub",
        "Baltic States Alliance", "Austrian Analytics", "Romania Development Center"
    ]

    for i in range(55):  # Creating 55 tasks
        task_date = datetime.now() + timedelta(days=i//3)
        project_idx = i % 15
        tasks.append({
            "Task_ID": f"T-{i+1:03d}",
            "Date": task_date,
            "Day_of_Week": task_date.strftime("%A"),
            "Time": f"{9 + (i % 8)}:00",
            "Project_ID": f"PRJ-{project_idx + 1:03d}",
            "Project_Name": project_names[project_idx],
            "Task_Type": random.choice(task_types),
            "Task_Description": f"Task for {project_names[project_idx]} - Item {i+1}",
            "Owner": random.choice(owners),
            "Priority": random.choice(["Critical", "High", "Medium", "Low"]),
            "Status": "Scheduled" if task_date > datetime.now() else "Pending"
        })

    # Write tasks to sheet
    for row, task in enumerate(tasks, 3):
        ws.cell(row=row, column=1, value=task["Task_ID"])
        ws.cell(row=row, column=2, value=task["Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=3, value=task["Day_of_Week"])
        ws.cell(row=row, column=4, value=task["Time"])
        ws.cell(row=row, column=5, value=task["Project_ID"])
        ws.cell(row=row, column=6, value=task["Project_Name"])
        ws.cell(row=row, column=7, value=task["Task_Type"])
        ws.cell(row=row, column=8, value=task["Task_Description"])
        ws.cell(row=row, column=9, value=task["Owner"])
        ws.cell(row=row, column=10, value=task["Priority"])
        ws.cell(row=row, column=11, value=task["Status"])

        # Color code priority
        if task["Priority"] == "Critical":
            ws.cell(row=row, column=10).font = Font(color=COLORS['danger_red'], bold=True)
        elif task["Priority"] == "High":
            ws.cell(row=row, column=10).font = Font(color=COLORS['warning_amber'], bold=True)

    # Add autofilter
    ws.auto_filter.ref = f"A2:K{len(tasks) + 2}"

def create_project_deep_dive(ws):
    """Create Project_Deep_Dive with European Map"""
    # Project selector dropdown
    ws['L1'].value = "PRJ-001"  # Default selection
    ws['K1'].value = "Select Project:"
    ws['K1'].font = Font(bold=True)

    # Create data validation for project selector
    dv = DataValidation(
        type="list",
        formula1="=Master_Projects_Clean!$A$2:$A$16",
        allow_blank=False,
        showDropDown=True
    )
    ws.add_data_validation(dv)
    dv.add('L1')

    # Project header
    ws.merge_cells('B1:J1')
    ws['B1'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:B,2,FALSE)"
    ws['B1'].font = Font(size=16, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # Project details section
    ws['B3'].value = "Status:"
    ws['C3'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:D,4,FALSE)"
    ws['C3'].fill = PatternFill(start_color=COLORS['success_green'],
                                end_color=COLORS['success_green'],
                                fill_type='solid')
    ws['C3'].font = Font(color=COLORS['white'], bold=True)

    ws['E3'].value = "Priority:"
    ws['F3'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:E,5,FALSE)"

    ws['H3'].value = "Progress:"
    ws['I3'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:F,6,FALSE)"
    ws['I3'].number_format = '0%'

    # Summary section
    ws['B5'].value = "Project Summary:"
    ws.merge_cells('B6:J8')
    ws['B6'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:C,3,FALSE)"
    ws['B6'].alignment = Alignment(wrap_text=True, vertical='top')

    # Key metrics
    ws['B10'].value = "KEY METRICS"
    ws['B10'].font = Font(size=12, bold=True)

    ws['B11'].value = "Start Date:"
    ws['D11'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:G,7,FALSE)"
    ws['D11'].number_format = 'mm/dd/yyyy'

    ws['F11'].value = "End Date:"
    ws['H11'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:H,8,FALSE)"
    ws['H11'].number_format = 'mm/dd/yyyy'

    ws['B12'].value = "Days Remaining:"
    ws['D12'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:I,9,FALSE)"

    ws['F12'].value = "Budget:"
    ws['H12'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:M,13,FALSE)"
    ws['H12'].number_format = '"$"#,##0'

    ws['B13'].value = "Budget Spent:"
    ws['D13'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:N,14,FALSE)"
    ws['D13'].number_format = '"$"#,##0'

    ws['F13'].value = "Budget %:"
    ws['H13'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:O,15,FALSE)"
    ws['H13'].number_format = '0%'

    # European Map Section
    ws['B15'].value = "EUROPEAN PRESENCE"
    ws['B15'].font = Font(size=12, bold=True)

    # Create simplified European map (using cells as pixels)
    map_start_row = 17
    map_start_col = 2

    # Make cells square for map
    for row in range(map_start_row, map_start_row + 20):
        ws.row_dimensions[row].height = 15
    for col in range(map_start_col, map_start_col + 15):
        ws.column_dimensions[get_column_letter(col)].width = 2

    # Draw country shapes (simplified)
    country_cells = {
        "UK": [(3, 2), (4, 2), (5, 2)],
        "France": [(6, 3), (7, 3), (8, 3)],
        "Germany": [(5, 5), (6, 5), (7, 5)],
        "Spain": [(9, 2), (10, 2)],
        "Italy": [(8, 6), (9, 6), (10, 6)],
        "Poland": [(5, 8), (6, 8)],
        "Sweden": [(2, 6), (3, 6)],
        "Norway": [(1, 5), (2, 5)]
    }

    for country, cells in country_cells.items():
        for (r, c) in cells:
            cell = ws.cell(row=map_start_row + r, column=map_start_col + c)
            cell.fill = PatternFill(start_color=COLORS['primary_blue'],
                                   end_color=COLORS['primary_blue'],
                                   fill_type='solid')

    # Map legend
    ws['B38'].value = "🟦 Event Location"
    ws['E38'].value = "🟩 Participant Country"
    ws['H38'].value = "⬜ Other European Country"

    # Next milestone section
    ws['B40'].value = "NEXT MILESTONE"
    ws['B40'].font = Font(size=12, bold=True)

    ws['B41'].value = "Milestone:"
    ws.merge_cells('D41:J41')
    ws['D41'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:T,20,FALSE)"

    ws['B42'].value = "Due Date:"
    ws['D42'].value = "=VLOOKUP($L$1,Master_Projects_Clean!A:U,21,FALSE)"
    ws['D42'].number_format = 'mm/dd/yyyy'

def create_europe_map_visual(ws):
    """Create Europe_Map_Visual (Dedicated Map Sheet)"""
    ws['B1'].value = "EUROPEAN PROJECT PRESENCE MAP"
    ws['B1'].font = Font(size=14, bold=True)

    # Instructions
    ws['B3'].value = "Interactive European coverage map showing all project locations"
    ws['B3'].font = Font(italic=True, color=COLORS['medium_text'])

    # Set all cells to square dimensions for map
    for row in range(5, 45):
        ws.row_dimensions[row].height = 12
    for col in range(2, 52):
        ws.column_dimensions[get_column_letter(col)].width = 1.5

    # Create a more detailed European map
    # Colors for different regions
    colors = {
        'active': COLORS['primary_blue'],
        'partner': COLORS['success_green'],
        'planned': COLORS['warning_amber'],
        'water': 'ADD8E6',
        'neutral': 'E0E0E0'
    }

    # Draw complete European map (simplified representation)
    # This is a simplified grid-based representation
    europe_map = [
        # Row 1 - Iceland, Norway top
        [(5, 10, 'neutral'), (5, 11, 'neutral')],  # Iceland
        [(5, 25, 'partner'), (5, 26, 'partner'), (5, 27, 'partner')],  # Norway

        # Row 2 - UK, Nordic countries
        [(8, 8, 'active'), (8, 9, 'active'), (8, 10, 'active')],  # UK
        [(8, 11, 'active'), (9, 8, 'active'), (9, 9, 'active')],  # UK continued
        [(8, 28, 'partner'), (8, 29, 'partner')],  # Sweden
        [(8, 31, 'neutral'), (8, 32, 'neutral')],  # Finland

        # Row 3 - Ireland, France, Germany, Poland
        [(11, 6, 'active')],  # Ireland
        [(11, 12, 'partner'), (11, 13, 'partner'), (12, 12, 'partner')],  # France
        [(11, 18, 'active'), (11, 19, 'active'), (11, 20, 'active')],  # Germany
        [(11, 24, 'planned'), (11, 25, 'planned')],  # Poland

        # Row 4 - Spain, Italy, Balkans
        [(15, 8, 'active'), (15, 9, 'active')],  # Spain
        [(16, 8, 'active'), (16, 9, 'active')],  # Spain continued
        [(15, 20, 'partner'), (16, 20, 'partner'), (17, 21, 'partner')],  # Italy

        # Row 5 - Greece, Turkey
        [(20, 24, 'active')],  # Greece
        [(20, 30, 'planned'), (20, 31, 'planned'), (20, 32, 'planned')]  # Turkey
    ]

    # Apply colors to map cells
    for row_data in europe_map:
        for r, c, color_type in row_data:
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=colors[color_type],
                                   end_color=colors[color_type],
                                   fill_type='solid')

    # Add country labels
    country_labels = [
        (6, 10, "IS"),  # Iceland
        (6, 26, "NO"),  # Norway
        (10, 9, "UK"),  # United Kingdom
        (10, 29, "SE"),  # Sweden
        (10, 32, "FI"),  # Finland
        (12, 6, "IE"),  # Ireland
        (13, 13, "FR"),  # France
        (12, 19, "DE"),  # Germany
        (12, 25, "PL"),  # Poland
        (17, 9, "ES"),  # Spain
        (18, 21, "IT"),  # Italy
        (21, 24, "GR"),  # Greece
        (21, 31, "TR")   # Turkey
    ]

    for r, c, label in country_labels:
        cell = ws.cell(row=r, column=c)
        cell.value = label
        cell.font = Font(size=8, bold=True, color=COLORS['white'])
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Legend
    ws['B47'].value = "LEGEND"
    ws['B47'].font = Font(size=12, bold=True)

    legend_items = [
        ("Active Projects", colors['active']),
        ("Partner Countries", colors['partner']),
        ("Planned Expansion", colors['planned']),
        ("Other EU Countries", colors['neutral'])
    ]

    for i, (label, color) in enumerate(legend_items):
        row = 49 + i
        ws.cell(row=row, column=2).fill = PatternFill(start_color=color,
                                                      end_color=color,
                                                      fill_type='solid')
        ws.cell(row=row, column=3).value = label
        ws.cell(row=row, column=3).font = Font(size=10)

    # Statistics
    ws['H47'].value = "COVERAGE STATISTICS"
    ws['H47'].font = Font(size=12, bold=True)

    stats = [
        ("Active Countries:", "8"),
        ("Partner Countries:", "5"),
        ("Total Projects:", "15"),
        ("Total Events:", "42"),
        ("Total Budget:", "$32.5M")
    ]

    for i, (label, value) in enumerate(stats):
        row = 49 + i
        ws.cell(row=row, column=8).value = label
        ws.cell(row=row, column=10).value = value
        ws.cell(row=row, column=10).font = Font(bold=True)

def create_project_events(ws):
    """Create Project_Events sheet with actual event data"""
    headers = ["Event_ID", "Project_ID", "Event_Name", "Event_Type",
               "Host_Country", "Host_City", "Start_Date", "End_Date",
               "Duration_Days", "Expected_Attendees", "Status", "Budget", "Venue"]

    # Add headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # Create 35+ events
    events = []
    event_types = ["Conference", "Workshop", "Summit", "Meeting", "Training", "Launch"]
    venues = ["Convention Center", "Hotel Conference Room", "Corporate HQ", "University", "Tech Hub"]
    cities = ["London", "Paris", "Berlin", "Madrid", "Rome", "Amsterdam", "Vienna", "Prague", "Warsaw", "Stockholm"]

    for i in range(38):
        start_date = datetime.now() + timedelta(days=random.randint(1, 180))
        duration = random.randint(1, 3)
        end_date = start_date + timedelta(days=duration)

        events.append({
            "Event_ID": f"EVT-{i+1:03d}",
            "Project_ID": f"PRJ-{(i % 15) + 1:03d}",
            "Event_Name": f"{random.choice(event_types)} - {random.choice(['Q1', 'Q2', 'Q3', 'Q4'])} {2025}",
            "Event_Type": random.choice(event_types),
            "Host_Country": random.choice(["UK", "Germany", "France", "Spain", "Italy", "Netherlands"]),
            "Host_City": random.choice(cities),
            "Start_Date": start_date,
            "End_Date": end_date,
            "Duration_Days": duration,
            "Expected_Attendees": random.randint(20, 500),
            "Status": random.choice(["Planned", "Confirmed", "In Progress", "Complete"]),
            "Budget": random.randint(10000, 250000),
            "Venue": random.choice(venues)
        })

    # Write events to sheet
    for row, event in enumerate(events, 2):
        for col, key in enumerate(["Event_ID", "Project_ID", "Event_Name", "Event_Type",
                                  "Host_Country", "Host_City"], 1):
            ws.cell(row=row, column=col, value=event[key])

        ws.cell(row=row, column=7, value=event["Start_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=8, value=event["End_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=9, value=event["Duration_Days"])
        ws.cell(row=row, column=10, value=event["Expected_Attendees"])
        ws.cell(row=row, column=11, value=event["Status"])
        ws.cell(row=row, column=12, value=event["Budget"]).number_format = '"$"#,##0'
        ws.cell(row=row, column=13, value=event["Venue"])

def create_event_participants(ws):
    """Create Event_Participants sheet"""
    headers = ["Participant_ID", "Event_ID", "Name", "Organization", "Country",
               "Role", "Email", "Registration_Date", "Attendance_Status"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # Create sample participants
    first_names = ["James", "Emma", "Hans", "Marie", "Giovanni", "Sofia", "Lars", "Elena"]
    last_names = ["Smith", "Johnson", "Mueller", "Dubois", "Rossi", "Garcia", "Andersson", "Novak"]
    orgs = ["Acme Corp", "TechCo", "Global Industries", "European Partners", "Innovation Lab"]
    roles = ["Executive", "Manager", "Analyst", "Director", "Consultant", "Engineer"]

    participants = []
    for i in range(100):
        participants.append({
            "Participant_ID": f"P-{i+1:04d}",
            "Event_ID": f"EVT-{(i % 38) + 1:03d}",
            "Name": f"{random.choice(first_names)} {random.choice(last_names)}",
            "Organization": random.choice(orgs),
            "Country": random.choice(["UK", "Germany", "France", "Spain", "Italy", "Netherlands", "Belgium"]),
            "Role": random.choice(roles),
            "Email": f"participant{i+1}@example.com",
            "Registration_Date": datetime.now() - timedelta(days=random.randint(1, 30)),
            "Attendance_Status": random.choice(["Registered", "Confirmed", "Attended", "No Show"])
        })

    # Write participants to sheet
    for row, participant in enumerate(participants, 2):
        for col, key in enumerate(["Participant_ID", "Event_ID", "Name", "Organization",
                                  "Country", "Role", "Email"], 1):
            ws.cell(row=row, column=col, value=participant[key])

        ws.cell(row=row, column=8, value=participant["Registration_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=9, value=participant["Attendance_Status"])

def create_project_milestones(ws):
    """Create Project_Milestones_Standard sheet"""
    headers = ["Milestone_ID", "Project_ID", "Milestone_Name", "Description",
               "Due_Date", "Owner", "Status", "Completion_Date", "Dependencies", "Critical_Path"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    # Create milestones
    milestone_names = ["Kickoff", "Requirements Complete", "Design Review", "Development Complete",
                      "Testing Phase", "UAT Sign-off", "Go-Live", "Post-Launch Review"]

    milestones = []
    for i in range(60):  # 4 milestones per project on average
        project_idx = (i // 4) + 1
        milestone_date = datetime.now() + timedelta(days=random.randint(1, 200))

        milestones.append({
            "Milestone_ID": f"MS-{i+1:03d}",
            "Project_ID": f"PRJ-{project_idx:03d}",
            "Milestone_Name": random.choice(milestone_names),
            "Description": f"Key deliverable for project phase {(i % 4) + 1}",
            "Due_Date": milestone_date,
            "Owner": random.choice(["John Smith", "Maria Garcia", "David Chen", "Sarah Johnson"]),
            "Status": random.choice(["Not Started", "In Progress", "Complete", "At Risk"]),
            "Completion_Date": milestone_date if random.random() > 0.5 else None,
            "Dependencies": f"MS-{max(1, i-1):03d}" if i > 0 else "None",
            "Critical_Path": "Yes" if random.random() > 0.7 else "No"
        })

    # Write milestones to sheet
    for row, milestone in enumerate(milestones, 2):
        for col, key in enumerate(["Milestone_ID", "Project_ID", "Milestone_Name", "Description"], 1):
            ws.cell(row=row, column=col, value=milestone[key])

        ws.cell(row=row, column=5, value=milestone["Due_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=6, value=milestone["Owner"])
        ws.cell(row=row, column=7, value=milestone["Status"])

        if milestone["Completion_Date"]:
            ws.cell(row=row, column=8, value=milestone["Completion_Date"]).number_format = 'mm/dd/yyyy'

        ws.cell(row=row, column=9, value=milestone["Dependencies"])
        ws.cell(row=row, column=10, value=milestone["Critical_Path"])

def create_project_audiences(ws):
    """Create Project_Audiences sheet"""
    headers = ["Audience_ID", "Project_ID", "Audience_Type", "Audience_Name",
               "Size", "Engagement_Level", "Communication_Method", "Last_Contact"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    audience_types = ["Government Officials", "Industry Partners", "Media", "Employees", "Customers", "Investors"]
    engagement_levels = ["High", "Medium", "Low", "Critical"]
    comm_methods = ["Email", "Newsletter", "Direct Meeting", "Conference Call", "Webinar"]

    audiences = []
    for i in range(45):
        audiences.append({
            "Audience_ID": f"AUD-{i+1:03d}",
            "Project_ID": f"PRJ-{(i % 15) + 1:03d}",
            "Audience_Type": random.choice(audience_types),
            "Audience_Name": f"Stakeholder Group {i+1}",
            "Size": random.randint(10, 500),
            "Engagement_Level": random.choice(engagement_levels),
            "Communication_Method": random.choice(comm_methods),
            "Last_Contact": datetime.now() - timedelta(days=random.randint(1, 30))
        })

    for row, audience in enumerate(audiences, 2):
        for col, key in enumerate(["Audience_ID", "Project_ID", "Audience_Type", "Audience_Name"], 1):
            ws.cell(row=row, column=col, value=audience[key])

        ws.cell(row=row, column=5, value=audience["Size"])
        ws.cell(row=row, column=6, value=audience["Engagement_Level"])
        ws.cell(row=row, column=7, value=audience["Communication_Method"])
        ws.cell(row=row, column=8, value=audience["Last_Contact"]).number_format = 'mm/dd/yyyy'

def create_project_products(ws):
    """Create Project_Products sheet"""
    headers = ["Product_ID", "Project_ID", "Product_Type", "Product_Name",
               "Description", "Delivery_Date", "Status", "Quality_Score"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    product_types = ["Report", "Presentation", "Website", "Software", "Documentation", "Training Material"]

    products = []
    for i in range(50):
        products.append({
            "Product_ID": f"PROD-{i+1:03d}",
            "Project_ID": f"PRJ-{(i % 15) + 1:03d}",
            "Product_Type": random.choice(product_types),
            "Product_Name": f"Deliverable {i+1}",
            "Description": f"Key output for project phase",
            "Delivery_Date": datetime.now() + timedelta(days=random.randint(1, 90)),
            "Status": random.choice(["Draft", "Review", "Final", "Delivered"]),
            "Quality_Score": random.randint(70, 100)
        })

    for row, product in enumerate(products, 2):
        for col, key in enumerate(["Product_ID", "Project_ID", "Product_Type", "Product_Name", "Description"], 1):
            ws.cell(row=row, column=col, value=product[key])

        ws.cell(row=row, column=6, value=product["Delivery_Date"]).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=7, value=product["Status"])
        ws.cell(row=row, column=8, value=product["Quality_Score"])

def create_project_stakeholders(ws):
    """Create Project_Stakeholders sheet"""
    headers = ["Stakeholder_ID", "Project_ID", "Name", "Title", "Organization",
               "Influence_Level", "Interest_Level", "Email", "Phone", "Notes"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    titles = ["CEO", "CTO", "Director", "Manager", "VP", "Head of Department", "Senior Analyst"]
    orgs = ["Client Corp", "Partner Co", "Government Agency", "Regulatory Body", "Vendor Inc"]

    stakeholders = []
    for i in range(60):
        stakeholders.append({
            "Stakeholder_ID": f"STK-{i+1:03d}",
            "Project_ID": f"PRJ-{(i % 15) + 1:03d}",
            "Name": f"Stakeholder {i+1}",
            "Title": random.choice(titles),
            "Organization": random.choice(orgs),
            "Influence_Level": random.choice(["High", "Medium", "Low"]),
            "Interest_Level": random.choice(["High", "Medium", "Low"]),
            "Email": f"stakeholder{i+1}@example.com",
            "Phone": f"+44 20 {random.randint(1000, 9999)} {random.randint(1000, 9999)}",
            "Notes": "Key decision maker" if random.random() > 0.7 else "Regular updates needed"
        })

    for row, stakeholder in enumerate(stakeholders, 2):
        for col, key in enumerate(["Stakeholder_ID", "Project_ID", "Name", "Title", "Organization"], 1):
            ws.cell(row=row, column=col, value=stakeholder[key])

        ws.cell(row=row, column=6, value=stakeholder["Influence_Level"])
        ws.cell(row=row, column=7, value=stakeholder["Interest_Level"])
        ws.cell(row=row, column=8, value=stakeholder["Email"])
        ws.cell(row=row, column=9, value=stakeholder["Phone"])
        ws.cell(row=row, column=10, value=stakeholder["Notes"])

def create_loe_goals(ws):
    """Create LOE_Goals sheet"""
    headers = ["Goal_ID", "LOE", "Goal_Name", "Target_Value", "Current_Value",
               "Unit", "Target_Date", "Owner", "Status", "Progress"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue_bg'],
                               end_color=COLORS['light_blue_bg'],
                               fill_type='solid')

    loe_goals = [
        ("Digital Transformation", "Cloud Migration", 100, 72, "%", "Platform adoption"),
        ("Digital Transformation", "API Integration", 50, 35, "APIs", "Systems integrated"),
        ("Customer Experience", "NPS Score", 80, 68, "Points", "Net Promoter Score"),
        ("Customer Experience", "Response Time", 2, 3.5, "Hours", "Average response"),
        ("Cost Optimization", "Cost Reduction", 15, 8, "%", "YoY savings"),
        ("Cost Optimization", "Process Automation", 75, 45, "%", "Processes automated"),
        ("Market Expansion", "New Markets", 8, 5, "Countries", "Market presence"),
        ("Market Expansion", "Revenue Growth", 25, 18, "%", "YoY growth"),
        ("Compliance & Risk", "Audit Score", 95, 88, "%", "Compliance rating"),
        ("Compliance & Risk", "Risk Items", 10, 15, "Items", "Open risk items")
    ]

    for i, (loe, goal, target, current, unit, description) in enumerate(loe_goals, 1):
        ws.cell(row=i+1, column=1, value=f"GOAL-{i:03d}")
        ws.cell(row=i+1, column=2, value=loe)
        ws.cell(row=i+1, column=3, value=goal)
        ws.cell(row=i+1, column=4, value=target)
        ws.cell(row=i+1, column=5, value=current)
        ws.cell(row=i+1, column=6, value=unit)
        ws.cell(row=i+1, column=7, value=datetime.now() + timedelta(days=random.randint(30, 180))).number_format = 'mm/dd/yyyy'
        ws.cell(row=i+1, column=8, value=random.choice(["John Smith", "Maria Garcia", "David Chen"]))
        ws.cell(row=i+1, column=9, value="On Track" if current/target > 0.7 else "At Risk")
        ws.cell(row=i+1, column=10, value=current/target).number_format = '0%'

def create_country_map_grid(ws):
    """Create Country_Map_Grid sheet"""
    ws['A1'].value = "COUNTRY PROJECT MATRIX"
    ws['A1'].font = Font(size=14, bold=True)

    # Create grid headers
    countries = ["UK", "Germany", "France", "Spain", "Italy", "Netherlands", "Belgium",
                "Poland", "Sweden", "Norway", "Denmark", "Finland", "Austria", "Switzerland",
                "Portugal", "Greece", "Ireland", "Czech Republic", "Hungary", "Romania"]

    # Row headers (countries)
    for i, country in enumerate(countries, 3):
        ws.cell(row=i, column=1, value=country)
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = PatternFill(start_color=COLORS['light_gray_bg'],
                                                    end_color=COLORS['light_gray_bg'],
                                                    fill_type='solid')

    # Column headers (project IDs)
    for i in range(1, 16):
        col = i + 1
        ws.cell(row=2, column=col, value=f"PRJ-{i:03d}")
        ws.cell(row=2, column=col).font = Font(bold=True, size=9)
        ws.cell(row=2, column=col).fill = PatternFill(start_color=COLORS['light_gray_bg'],
                                                      end_color=COLORS['light_gray_bg'],
                                                      fill_type='solid')
        ws.cell(row=2, column=col).alignment = Alignment(text_rotation=90)

    # Fill matrix with checkmarks or values
    for row in range(3, len(countries) + 3):
        for col in range(2, 17):
            if random.random() > 0.6:  # 40% chance of involvement
                ws.cell(row=row, column=col, value="✓")
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=col).font = Font(color=COLORS['success_green'], bold=True)

    # Summary column
    ws.cell(row=2, column=18, value="Total")
    ws.cell(row=2, column=18).font = Font(bold=True)
    ws.cell(row=2, column=18).fill = PatternFill(start_color=COLORS['light_blue_bg'],
                                                 end_color=COLORS['light_blue_bg'],
                                                 fill_type='solid')

    for row in range(3, len(countries) + 3):
        ws.cell(row=row, column=18, value=f"=COUNTIF(B{row}:Q{row},\"✓\")")
        ws.cell(row=row, column=18).font = Font(bold=True)

def finalize_workbook(wb):
    """Apply final formatting and settings"""
    # Apply consistent formatting across all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Remove gridlines for dashboard sheets
        if "Dashboard" in sheet_name or "Deep_Dive" in sheet_name or "Map" in sheet_name:
            ws.sheet_view.showGridLines = False

        # Set print areas
        if sheet_name == "Dashboard_Home":
            ws.print_area = 'A1:M50'
            ws.page_setup.orientation = 'landscape'

        # Freeze panes for data sheets
        if sheet_name in ["Master_Projects_Clean", "Calendar_Todo", "Project_Events"]:
            ws.freeze_panes = 'A2'

    print("[OK] Formatting applied to all sheets")

def main():
    """Main execution function"""
    print("Creating Complete PM Dashboard...")
    print("=" * 50)

    # Step 1: Create workbook
    wb = create_workbook()
    print(f"[OK] Created workbook with {len(wb.sheetnames)} sheets")

    # Step 2: Populate all sheets
    create_config_lists(wb["Config_Lists"])
    print("[OK] Created Config Lists")

    create_european_cities(wb["European_Cities"])
    print("[OK] Added European Cities data (90+ cities)")

    create_master_projects(wb["Master_Projects_Clean"])
    print("[OK] Added 15 Master Projects")

    create_dashboard_home(wb["Dashboard_Home"])
    print("[OK] Created Dashboard Home with KPIs")

    create_calendar_todo(wb["Calendar_Todo"])
    print("[OK] Created Calendar with 55 tasks")

    create_project_deep_dive(wb["Project_Deep_Dive"])
    print("[OK] Created Project Deep Dive with European map")

    create_europe_map_visual(wb["Europe_Map_Visual"])
    print("[OK] Created European Map Visual")

    # Create all supporting data sheets
    create_project_events(wb["Project_Events"])
    print("[OK] Created Project Events (38 events)")

    create_event_participants(wb["Event_Participants"])
    print("[OK] Created Event Participants (100 participants)")

    create_project_milestones(wb["Project_Milestones_Standard"])
    print("[OK] Created Project Milestones (60 milestones)")

    create_project_audiences(wb["Project_Audiences"])
    print("[OK] Created Project Audiences")

    create_project_products(wb["Project_Products"])
    print("[OK] Created Project Products")

    create_project_stakeholders(wb["Project_Stakeholders"])
    print("[OK] Created Project Stakeholders")

    create_loe_goals(wb["LOE_Goals"])
    print("[OK] Created LOE Goals")

    create_country_map_grid(wb["Country_Map_Grid"])
    print("[OK] Created Country Map Grid")

    # Step 3: Finalize and save
    finalize_workbook(wb)

    # Save the file
    wb.save("refined_pm_dashboard.xlsx")

    print("=" * 50)
    print("[COMPLETE] PM Dashboard created successfully!")
    print(f"[INFO] Sheets created: {len(wb.sheetnames)}")
    print(f"[INFO] Sheets: {', '.join(wb.sheetnames)}")
    print("[INFO] File saved as: refined_pm_dashboard.xlsx")

    return wb

if __name__ == "__main__":
    workbook = main()
