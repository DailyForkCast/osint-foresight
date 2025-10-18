#!/usr/bin/env python3
"""
Project Management Dashboard Excel Generator
Creates a complete PM dashboard workbook with all sheets, formulas, and formatting
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime, timedelta
import random
from typing import List, Dict, Any

class ProjectManagementDashboard:
    def __init__(self):
        """Initialize the workbook and styling"""
        self.wb = Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Define color scheme
        self.colors = {
            'header_dark': 'FF1E40AF',  # Dark blue
            'header_light': 'FF2196F3',  # Light blue
            'success': 'FF4CAF50',  # Green
            'warning': 'FFFFC107',  # Yellow
            'danger': 'FFF44336',   # Red
            'info': 'FF03A9F4',     # Light blue
            'light_gray': 'FFF5F5F5',
            'medium_gray': 'FFE0E0E0',
            'dark_gray': 'FF424242'
        }

        # Create styles
        self.create_styles()

        # Sample data storage
        self.config_data = []
        self.stakeholders = []
        self.projects = []
        self.milestones = []

    def create_styles(self):
        """Create reusable styles"""
        # Header style
        self.header_style = NamedStyle(name='header')
        self.header_style.font = Font(bold=True, color='FFFFFFFF', size=11)
        self.header_style.fill = PatternFill(start_color=self.colors['header_dark'],
                                             end_color=self.colors['header_dark'],
                                             fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')
        self.header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.wb.add_named_style(self.header_style)

        # Title style
        self.title_style = NamedStyle(name='title')
        self.title_style.font = Font(bold=True, size=14)
        self.title_style.alignment = Alignment(horizontal='left', vertical='center')
        self.wb.add_named_style(self.title_style)

        # Metric style
        self.metric_style = NamedStyle(name='metric')
        self.metric_style.font = Font(bold=True, size=24)
        self.metric_style.alignment = Alignment(horizontal='center', vertical='center')
        self.wb.add_named_style(self.metric_style)

    def create_config_lists_sheet(self):
        """Create the Config_Lists sheet with all dropdown values"""
        ws = self.wb.create_sheet('Config_Lists')

        # Headers
        headers = ['List_Type', 'Value', 'Sort_Order', 'Color_Hex', 'Icon']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Data
        config_data = [
            # Status values
            ('Status', 'Not Started', 1, '#9E9E9E', '○'),
            ('Status', 'Planning', 2, '#2196F3', '◔'),
            ('Status', 'In Progress', 3, '#4CAF50', '◐'),
            ('Status', 'On Hold', 4, '#FF9800', '⊕'),
            ('Status', 'At Risk', 5, '#F44336', '⊗'),
            ('Status', 'Complete', 6, '#607D8B', '●'),
            ('Status', 'Cancelled', 7, '#424242', '✕'),

            # Priority values
            ('Priority', 'Critical', 1, '#F44336', '🔴'),
            ('Priority', 'High', 2, '#FF9800', '🟠'),
            ('Priority', 'Medium', 3, '#FFC107', '🟡'),
            ('Priority', 'Low', 4, '#4CAF50', '🟢'),

            # LOE values
            ('LOE', 'Digital Transformation', 1, '#1E40AF', '💻'),
            ('LOE', 'Customer Experience', 2, '#7C3AED', '👥'),
            ('LOE', 'Cost Optimization', 3, '#DC2626', '💰'),
            ('LOE', 'Market Expansion', 4, '#059669', '🌍'),
            ('LOE', 'Compliance & Risk', 5, '#B45309', '⚖️'),

            # Delivery Status
            ('Delivery_Status', 'Not Started', 1, '#9E9E9E', ''),
            ('Delivery_Status', 'In Progress', 2, '#2196F3', ''),
            ('Delivery_Status', 'Delivered', 3, '#4CAF50', ''),
            ('Delivery_Status', 'Delayed', 4, '#F44336', ''),

            # Contact Status
            ('Contact_Status', 'Good', 1, '#4CAF50', '✓'),
            ('Contact_Status', 'Needs Attention', 2, '#FFC107', '!'),
            ('Contact_Status', 'Overdue', 3, '#F44336', '⚠'),

            # Contact Frequency
            ('Contact_Frequency', 'Weekly', 1, '', ''),
            ('Contact_Frequency', 'Bi-weekly', 2, '', ''),
            ('Contact_Frequency', 'Monthly', 3, '', ''),
            ('Contact_Frequency', 'Quarterly', 4, '', ''),
        ]

        for row_idx, row_data in enumerate(config_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        self.config_data = config_data

        # Create table
        table = Table(displayName="tbl_Config", ref=f"A1:E{len(config_data) + 1}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def create_loe_goals_sheet(self):
        """Create the LOE_Goals sheet with strategic initiatives"""
        ws = self.wb.create_sheet('LOE_Goals')

        # Headers
        headers = ['LOE_Name', 'FY25_Annual_Target', 'Q1_Target', 'Q1_Actual',
                  'Q2_Target', 'Q2_Actual', 'Progress_Percent', 'Status_Indicator',
                  'Owner', 'Success_Metric']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Sample data
        loe_data = [
            ['Digital Transformation', 20, 5, None, 5, None, None, None, 'John Smith', 'Projects delivered'],
            ['Customer Experience', 15, 4, None, 4, None, None, None, 'Jane Doe', 'NPS improvement'],
            ['Cost Optimization', 10, 3, None, 2, None, None, None, 'Mike Johnson', 'Cost savings $M'],
            ['Market Expansion', 12, 3, None, 3, None, None, None, 'Sarah Lee', 'New markets entered'],
            ['Compliance & Risk', 16, 4, None, 4, None, None, None, 'Tom Wilson', 'Audit findings']
        ]

        for row_idx, row_data in enumerate(loe_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Add formulas for calculated columns
                if col_idx == 4:  # Q1_Actual
                    formula = f'=COUNTIFS(Master_Projects[LOE_Primary],[@LOE_Name],Master_Projects[Delivery_Date],">=1/1/2025",Master_Projects[Delivery_Date],"<=3/31/2025",Master_Projects[Delivery_Status],"Delivered")'
                    cell.value = formula
                elif col_idx == 6:  # Q2_Actual
                    formula = f'=COUNTIFS(Master_Projects[LOE_Primary],[@LOE_Name],Master_Projects[Delivery_Date],">=4/1/2025",Master_Projects[Delivery_Date],"<=6/30/2025",Master_Projects[Delivery_Status],"Delivered")'
                    cell.value = formula
                elif col_idx == 7:  # Progress_Percent
                    formula = f'=IFERROR(D{row_idx}/C{row_idx},0)'
                    cell.value = formula
                    cell.number_format = '0%'
                elif col_idx == 8:  # Status_Indicator
                    formula = f'=IF(G{row_idx}>=0.9,"🟢 ON TRACK",IF(G{row_idx}>=0.7,"🟡 BEHIND","🔴 AT RISK"))'
                    cell.value = formula

        # Create table
        table = Table(displayName="LOE_Goals", ref=f"A1:J{len(loe_data) + 1}")
        style = TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit columns
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def create_stakeholder_master_sheet(self):
        """Create the Stakeholder_Master sheet with sample stakeholders"""
        ws = self.wb.create_sheet('Stakeholder_Master')

        # Headers
        headers = ['Stakeholder_ID', 'Full_Name', 'Organization', 'Email', 'Project_Count',
                  'Contact_Frequency', 'Last_Contact', 'Next_Required', 'Days_Overdue', 'Status']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Generate sample stakeholders
        first_names = ['John', 'Jane', 'Mike', 'Sarah', 'Tom', 'Lisa', 'David', 'Emily']
        last_names = ['Smith', 'Johnson', 'Williams', 'Brown', 'Davis', 'Miller', 'Wilson', 'Moore']
        organizations = ['Tech Corp', 'Finance Inc', 'Marketing Ltd', 'Sales Co', 'Operations LLC',
                        'HR Partners', 'Legal Firm', 'Consulting Group']
        frequencies = ['Weekly', 'Bi-weekly', 'Monthly', 'Quarterly']

        stakeholder_data = []
        for i in range(15):
            stakeholder_id = f'STK-{i+1:03d}'
            full_name = f'{random.choice(first_names)} {random.choice(last_names)}'
            organization = random.choice(organizations)
            email = f'{full_name.lower().replace(" ", ".")}@{organization.lower().replace(" ", "")}.com'
            contact_frequency = random.choice(frequencies)
            last_contact = datetime.now() - timedelta(days=random.randint(1, 60))

            stakeholder_data.append([
                stakeholder_id, full_name, organization, email, None,
                contact_frequency, last_contact, None, None, None
            ])

        self.stakeholders = stakeholder_data

        for row_idx, row_data in enumerate(stakeholder_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Add formulas
                if col_idx == 5:  # Project_Count
                    formula = f'=COUNTIF(Master_Projects[Primary_Stakeholder],B{row_idx})+COUNTIF(Master_Projects[Secondary_Stakeholder],B{row_idx})'
                    cell.value = formula
                elif col_idx == 7 and isinstance(value, datetime):  # Last_Contact
                    cell.number_format = 'mm/dd/yyyy'
                elif col_idx == 8:  # Next_Required
                    formula = f'=IFS(F{row_idx}="Weekly",G{row_idx}+7,F{row_idx}="Bi-weekly",G{row_idx}+14,F{row_idx}="Monthly",G{row_idx}+30,TRUE,G{row_idx}+90)'
                    cell.value = formula
                    cell.number_format = 'mm/dd/yyyy'
                elif col_idx == 9:  # Days_Overdue
                    formula = f'=IF(H{row_idx}<TODAY(),TODAY()-H{row_idx},0)'
                    cell.value = formula
                elif col_idx == 10:  # Status
                    formula = f'=IF(I{row_idx}>7,"🔴 Overdue",IF(I{row_idx}>0,"🟡 Due",IF(H{row_idx}-TODAY()<=3,"🟡 Upcoming","🟢 Good")))'
                    cell.value = formula

        # Create table
        table = Table(displayName="Stakeholder_Master", ref=f"A1:J{len(stakeholder_data) + 1}")
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit columns
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 16

    def create_master_projects_sheet(self):
        """Create the Master_Projects sheet with sample projects"""
        ws = self.wb.create_sheet('Master_Projects')

        # Headers
        headers = ['Project_ID', 'Project_Name', 'Status', 'Priority', 'LOE_Primary',
                  'Start_Date', 'End_Date', 'Days_Remaining', 'Progress_Percent',
                  'Primary_Stakeholder', 'Secondary_Stakeholder', 'Last_Updated',
                  'Delivery_Date', 'Delivery_Status', 'Budget', 'Budget_Used_Percent',
                  'Country_Primary', 'My_Project', 'Next_Milestone', 'Next_Milestone_Date',
                  'Health_Score', 'Description']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Generate sample projects
        project_names = [
            'Cloud Migration Phase 1',
            'Customer Portal Redesign',
            'Supply Chain Optimization',
            'Mobile App Development',
            'Data Analytics Platform',
            'Cybersecurity Enhancement',
            'ERP System Upgrade',
            'Marketing Automation',
            'AI/ML Implementation',
            'Process Automation Initiative'
        ]

        statuses = ['Not Started', 'Planning', 'In Progress', 'On Hold', 'At Risk']
        priorities = ['Critical', 'High', 'Medium', 'Low']
        loes = ['Digital Transformation', 'Customer Experience', 'Cost Optimization',
                'Market Expansion', 'Compliance & Risk']
        delivery_statuses = ['Not Started', 'In Progress', 'Delivered', 'Delayed']
        countries = ['USA', 'UK', 'Germany', 'France', 'Japan']

        project_data = []
        for i in range(10):
            project_id = f'PRJ-{i+1:03d}'
            project_name = project_names[i]
            status = random.choice(statuses)
            priority = random.choice(priorities)
            loe = random.choice(loes)
            start_date = datetime.now() - timedelta(days=random.randint(30, 180))
            end_date = start_date + timedelta(days=random.randint(60, 365))
            progress = random.randint(0, 100)
            primary_stakeholder = random.choice([s[1] for s in self.stakeholders[:10]])
            secondary_stakeholder = random.choice([s[1] for s in self.stakeholders[:10]])
            last_updated = datetime.now() - timedelta(days=random.randint(1, 14))
            delivery_date = end_date - timedelta(days=random.randint(0, 30))
            delivery_status = random.choice(delivery_statuses)
            budget = random.randint(50000, 500000)
            budget_used = random.randint(0, 100)
            country = random.choice(countries)
            my_project = random.choice(['Yes', 'No'])
            next_milestone = f'Complete {random.choice(["Design", "Development", "Testing", "Deployment"])}'
            next_milestone_date = datetime.now() + timedelta(days=random.randint(7, 60))
            description = f'Strategic initiative for {project_name.lower()}'

            project_data.append([
                project_id, project_name, status, priority, loe,
                start_date, end_date, None, progress, primary_stakeholder,
                secondary_stakeholder, last_updated, delivery_date, delivery_status,
                budget, budget_used, country, my_project, next_milestone,
                next_milestone_date, None, description
            ])

        self.projects = project_data

        for row_idx, row_data in enumerate(project_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format dates
                if col_idx in [6, 7, 12, 13, 20] and isinstance(value, datetime):
                    cell.number_format = 'mm/dd/yyyy'

                # Add formulas
                if col_idx == 8:  # Days_Remaining
                    formula = f'=MAX(0,G{row_idx}-TODAY())'
                    cell.value = formula
                elif col_idx == 9:  # Progress_Percent
                    cell.number_format = '0%'
                    cell.value = value / 100
                elif col_idx == 15:  # Budget
                    cell.number_format = '"$"#,##0'
                elif col_idx == 16:  # Budget_Used_Percent
                    cell.number_format = '0%'
                    cell.value = value / 100
                elif col_idx == 21:  # Health_Score
                    formula = f'=IF(OR(C{row_idx}="At Risk",H{row_idx}<7,P{row_idx}>0.9),"🔴 Critical",IF(OR(H{row_idx}<30,P{row_idx}>0.75),"🟡 Warning","🟢 Healthy"))'
                    cell.value = formula

        # Create table
        table = Table(displayName="Master_Projects", ref=f"A1:V{len(project_data) + 1}")
        style = TableStyleInfo(name="TableStyleMedium5", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit columns
        for col in range(1, 23):
            if col == 2:  # Project Name
                ws.column_dimensions[get_column_letter(col)].width = 25
            elif col == 22:  # Description
                ws.column_dimensions[get_column_letter(col)].width = 30
            else:
                ws.column_dimensions[get_column_letter(col)].width = 15

    def create_milestones_sheet(self):
        """Create the Milestones sheet with sample milestones"""
        ws = self.wb.create_sheet('Milestones')

        # Headers
        headers = ['Milestone_ID', 'Project_ID', 'Milestone_Name', 'Due_Date', 'Priority',
                  'Location', 'Owner', 'Status', 'Days_Until', 'Week_Group', 'Display_Line']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Generate sample milestones
        milestone_types = ['Design Review', 'Code Complete', 'Testing Complete', 'Go-Live',
                          'Training Complete', 'Documentation Review', 'Security Audit',
                          'Performance Testing', 'User Acceptance', 'Deployment']

        locations = ['Remote', 'HQ Office', 'NYC Office', 'London Office', 'Virtual']
        statuses = ['Not Started', 'In Progress', 'Complete', 'Delayed']

        milestone_data = []
        for i in range(20):
            milestone_id = f'M-{i+1:03d}'
            project_id = random.choice([p[0] for p in self.projects])
            milestone_name = random.choice(milestone_types)
            due_date = datetime.now() + timedelta(days=random.randint(-14, 90))
            priority = random.choice(['Critical', 'High', 'Medium', 'Low'])
            location = random.choice(locations)
            owner = random.choice([s[1] for s in self.stakeholders[:10]])
            status = random.choice(statuses)

            milestone_data.append([
                milestone_id, project_id, milestone_name, due_date, priority,
                location, owner, status, None, None, None
            ])

        self.milestones = milestone_data

        for row_idx, row_data in enumerate(milestone_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format dates
                if col_idx == 4 and isinstance(value, datetime):
                    cell.number_format = 'mm/dd/yyyy'

                # Add formulas
                if col_idx == 9:  # Days_Until
                    formula = f'=D{row_idx}-TODAY()'
                    cell.value = formula
                elif col_idx == 10:  # Week_Group
                    formula = f'=IF(I{row_idx}<0,"⚠️ OVERDUE",IF(I{row_idx}<=7,"THIS WEEK",IF(I{row_idx}<=14,"NEXT WEEK",IF(I{row_idx}<=30,"THIS MONTH","FUTURE"))))'
                    cell.value = formula
                elif col_idx == 11:  # Display_Line
                    formula = f'=TEXT(D{row_idx},"ddd dd")&" │ "&B{row_idx}&" │ "&C{row_idx}&" │ "&IF(E{row_idx}="Critical","🔴",IF(E{row_idx}="High","🟠","🟢"))&" "&E{row_idx}&" │ "&F{row_idx}'
                    cell.value = formula

        # Create table
        table = Table(displayName="Milestones", ref=f"A1:K{len(milestone_data) + 1}")
        style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit columns
        for col in range(1, 12):
            if col == 11:  # Display_Line
                ws.column_dimensions[get_column_letter(col)].width = 50
            else:
                ws.column_dimensions[get_column_letter(col)].width = 15

    def create_executive_dashboard_sheet(self):
        """Create the Executive Dashboard with complex formulas and formatting"""
        ws = self.wb.create_sheet('Executive_Dashboard')

        # Remove gridlines
        ws.sheet_view.showGridLines = False

        # Title Section
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = 'EXECUTIVE PORTFOLIO VIEW'
        title_cell.font = Font(bold=True, size=16, color='FF1E40AF')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws['O1'] = f'=TEXT(NOW(),"mmm dd, yyyy h:mm AM/PM")'
        ws['O1'].alignment = Alignment(horizontal='right')

        # Top Metrics Cards (Row 3-4)
        metric_headers = [
            ('C3', 'ACTIVE PROJECTS'),
            ('F3', 'DUE THIS WEEK'),
            ('I3', 'STAKEHOLDER TOUCHPOINTS'),
            ('L3', 'DELIVERED THIS MONTH'),
            ('O3', 'GOALS ON TRACK')
        ]

        for cell_ref, header in metric_headers:
            cell = ws[cell_ref]
            cell.value = header
            cell.font = Font(bold=True, size=10, color='FF666666')
            cell.alignment = Alignment(horizontal='center')

        # Metric Values with formulas
        ws['C4'] = '=COUNTIFS(Master_Projects[Status],"In Progress")+COUNTIFS(Master_Projects[Status],"At Risk")'
        ws['C4'].style = 'metric'

        ws['F4'] = '=COUNTIFS(Milestones[Due_Date],">="&TODAY(),Milestones[Due_Date],"<"&TODAY()+7,Milestones[Status],"<>"&"Complete")'
        ws['F4'].style = 'metric'

        ws['I4'] = '=COUNTIFS(Stakeholder_Master[Days_Overdue],">0")&" overdue"'
        ws['I4'].font = Font(bold=True, size=18)
        ws['I4'].alignment = Alignment(horizontal='center')

        ws['L4'] = '=COUNTIFS(Master_Projects[Delivery_Date],">="&EOMONTH(TODAY(),-1)+1,Master_Projects[Delivery_Date],"<="&EOMONTH(TODAY(),0),Master_Projects[Delivery_Status],"Delivered")&" of "&COUNTIFS(Master_Projects[Delivery_Date],"<="&EOMONTH(TODAY(),0))'
        ws['L4'].font = Font(bold=True, size=18)
        ws['L4'].alignment = Alignment(horizontal='center')

        ws['O4'] = '=COUNTIFS(LOE_Goals[Status_Indicator],"🟢 ON TRACK")&" of "&COUNTA(LOE_Goals[LOE_Name])'
        ws['O4'].font = Font(bold=True, size=18)
        ws['O4'].alignment = Alignment(horizontal='center')

        # Add borders to metric cards
        for col in ['C', 'F', 'I', 'L', 'O']:
            for row in [3, 4]:
                cell = ws[f'{col}{row}']
                cell.border = Border(
                    left=Side(style='thick', color='FFE0E0E0'),
                    right=Side(style='thick', color='FFE0E0E0'),
                    top=Side(style='thick', color='FFE0E0E0'),
                    bottom=Side(style='thick', color='FFE0E0E0')
                )

        # Lines of Effort Section
        ws.merge_cells('A7:O7')
        loe_title = ws['A7']
        loe_title.value = 'STRATEGIC LINES OF EFFORT'
        loe_title.font = Font(bold=True, size=12, color='FF1E40AF')
        loe_title.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')

        # LOE Headers
        ws['A9'] = 'Initiative'
        ws['G9'] = 'Progress'
        ws['L9'] = 'Status'
        ws['N9'] = 'Details'

        for cell in ['A9', 'G9', 'L9', 'N9']:
            ws[cell].font = Font(bold=True, size=10)
            ws[cell].fill = PatternFill(start_color='FFE8E8E8', end_color='FFE8E8E8', fill_type='solid')

        # LOE Data (rows 10-14)
        for i in range(5):
            row = 10 + i
            ws[f'A{row}'] = f'=INDEX(LOE_Goals[LOE_Name],{i+1})'
            ws[f'G{row}'] = f'=REPT("█",ROUND(INDEX(LOE_Goals[Progress_Percent],{i+1})*20,0))&REPT("░",20-ROUND(INDEX(LOE_Goals[Progress_Percent],{i+1})*20,0))&"  "&TEXT(INDEX(LOE_Goals[Progress_Percent],{i+1}),"0%")'
            ws[f'G{row}'].font = Font(name='Consolas', size=10)
            ws[f'L{row}'] = f'=INDEX(LOE_Goals[Status_Indicator],{i+1})'
            ws[f'N{row}'] = f'=INDEX(LOE_Goals[Q1_Actual],{i+1})&" of "&INDEX(LOE_Goals[Q1_Target],{i+1})&" projects"'

        # Milestone River Section
        ws.merge_cells('A22:O22')
        milestone_title = ws['A22']
        milestone_title.value = 'TIME MANAGEMENT: MILESTONE RIVER'
        milestone_title.font = Font(bold=True, size=12, color='FF1E40AF')
        milestone_title.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')

        # Overdue Milestones
        ws['A24'] = '=CONCATENATE("⚠️ OVERDUE (",COUNTIFS(Milestones[Week_Group],"⚠️ OVERDUE"),")")'
        ws['A24'].font = Font(bold=True, color='FFF44336')

        # This Week Milestones
        ws['A29'] = '=CONCATENATE("THIS WEEK (",COUNTIFS(Milestones[Week_Group],"THIS WEEK"),")")'
        ws['A29'].font = Font(bold=True, color='FF2196F3')

        # Next Week Milestones
        ws['A34'] = '=CONCATENATE("NEXT WEEK (",COUNTIFS(Milestones[Week_Group],"NEXT WEEK"),")")'
        ws['A34'].font = Font(bold=True, color='FF4CAF50')

        # Stakeholder Section
        ws.merge_cells('A42:O42')
        stakeholder_title = ws['A42']
        stakeholder_title.value = 'STAKEHOLDER ATTENTION NEEDED'
        stakeholder_title.font = Font(bold=True, size=12, color='FF1E40AF')
        stakeholder_title.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')

        ws['A44'] = '=CONCATENATE("🔴 OVERDUE CONTACT (",COUNTIFS(Stakeholder_Master[Status],"🔴 Overdue"),")")'
        ws['A44'].font = Font(bold=True, color='FFF44336')

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 3
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 3
        ws.column_dimensions['E'].width = 3
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 3
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 3
        ws.column_dimensions['K'].width = 3
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 3
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15

        # Freeze panes
        ws.freeze_panes = 'A6'

    def create_project_details_sheet(self):
        """Create the Project_Details sheet for deep dive views"""
        ws = self.wb.create_sheet('Project_Details')

        # Title
        ws['A1'] = 'PROJECT DEEP DIVE'
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = 'Select a project from Master_Projects sheet for detailed view'
        ws['A3'].font = Font(italic=True, color='FF666666')

        # Project selector area
        ws['A5'] = 'Project ID:'
        ws['B5'] = '[Select from dropdown]'
        ws['B5'].fill = PatternFill(start_color='FFFFFFE0', end_color='FFFFFFE0', fill_type='solid')

        # Detail sections
        sections = [
            ('A7', 'PROJECT OVERVIEW'),
            ('A15', 'TIMELINE & MILESTONES'),
            ('A25', 'BUDGET & RESOURCES'),
            ('A35', 'RISKS & ISSUES'),
            ('A45', 'STAKEHOLDERS')
        ]

        for cell_ref, section_title in sections:
            ws[cell_ref] = section_title
            ws[cell_ref].font = Font(bold=True, size=12, color='FF1E40AF')
            ws[cell_ref].fill = PatternFill(start_color='FFE8E8E8', end_color='FFE8E8E8', fill_type='solid')

        # Add data validation for project selector
        dv = DataValidation(type="list", formula1="Master_Projects[Project_ID]", allow_blank=False)
        dv.add('B5')
        ws.add_data_validation(dv)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30

    def create_archive_sheet(self):
        """Create the Archive sheet for completed projects"""
        ws = self.wb.create_sheet('Archive')

        # Headers (same as Master_Projects but with archive date)
        headers = ['Archive_Date', 'Project_ID', 'Project_Name', 'Status', 'Priority',
                  'LOE_Primary', 'Start_Date', 'End_Date', 'Final_Progress',
                  'Primary_Stakeholder', 'Budget', 'Budget_Used_Final', 'Lessons_Learned']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).style = 'header'

        # Add note
        ws['A3'] = 'Move completed projects here for historical reference'
        ws['A3'].font = Font(italic=True, color='FF666666')

        # Create table
        table = Table(displayName="Archive", ref=f"A1:M2")
        style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Column widths
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def add_data_validations(self):
        """Add data validation dropdowns across sheets"""
        # Master_Projects sheet validations
        mp_ws = self.wb['Master_Projects']

        # Status dropdown
        status_dv = DataValidation(
            type="list",
            formula1='"Not Started,Planning,In Progress,On Hold,At Risk,Complete,Cancelled"',
            allow_blank=False
        )
        status_dv.add('C2:C100')
        mp_ws.add_data_validation(status_dv)

        # Priority dropdown
        priority_dv = DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        )
        priority_dv.add('D2:D100')
        mp_ws.add_data_validation(priority_dv)

        # LOE dropdown
        loe_dv = DataValidation(
            type="list",
            formula1='"Digital Transformation,Customer Experience,Cost Optimization,Market Expansion,Compliance & Risk"',
            allow_blank=False
        )
        loe_dv.add('E2:E100')
        mp_ws.add_data_validation(loe_dv)

        # Delivery Status dropdown
        delivery_dv = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Delivered,Delayed"',
            allow_blank=False
        )
        delivery_dv.add('N2:N100')
        mp_ws.add_data_validation(delivery_dv)

        # My Project dropdown
        my_project_dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        my_project_dv.add('R2:R100')
        mp_ws.add_data_validation(my_project_dv)

        # Milestones sheet validations
        m_ws = self.wb['Milestones']

        # Priority dropdown for milestones
        m_priority_dv = DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        )
        m_priority_dv.add('E2:E100')
        m_ws.add_data_validation(m_priority_dv)

        # Status dropdown for milestones
        m_status_dv = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Delayed"',
            allow_blank=False
        )
        m_status_dv.add('H2:H100')
        m_ws.add_data_validation(m_status_dv)

    def add_conditional_formatting(self):
        """Add conditional formatting rules"""
        # Master_Projects sheet
        mp_ws = self.wb['Master_Projects']

        # Health Score formatting
        red_text = Font(color="FF9C0006")
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

        yellow_text = Font(color="FF9C6500")
        yellow_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")

        green_text = Font(color="FF006100")
        green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

        # Add rules for Health Score column
        mp_ws.conditional_formatting.add('U2:U100',
            CellIsRule(operator='containsText', formula=['Critical'], font=red_text, fill=red_fill))
        mp_ws.conditional_formatting.add('U2:U100',
            CellIsRule(operator='containsText', formula=['Warning'], font=yellow_text, fill=yellow_fill))
        mp_ws.conditional_formatting.add('U2:U100',
            CellIsRule(operator='containsText', formula=['Healthy'], font=green_text, fill=green_fill))

        # Progress bars color scale
        mp_ws.conditional_formatting.add('I2:I100',
            ColorScaleRule(start_type='percentile', start_value=0, start_color='FFFF0000',
                          mid_type='percentile', mid_value=50, mid_color='FFFFFF00',
                          end_type='percentile', end_value=100, end_color='FF00FF00'))

        # Executive Dashboard formatting
        ed_ws = self.wb['Executive_Dashboard']

        # Metric cards formatting
        ed_ws.conditional_formatting.add('C4',
            CellIsRule(operator='greaterThan', formula=['10'], font=red_text, fill=red_fill))
        ed_ws.conditional_formatting.add('F4',
            CellIsRule(operator='greaterThan', formula=['5'], font=yellow_text, fill=yellow_fill))

    def save_workbook(self, filename='Project_Management_Dashboard.xlsx'):
        """Save the workbook"""
        self.wb.save(filename)
        print(f"Workbook saved as: {filename}")

    def create_full_dashboard(self):
        """Main method to create the complete dashboard"""
        print("Creating Project Management Dashboard...")

        print("Creating Config_Lists sheet...")
        self.create_config_lists_sheet()

        print("Creating LOE_Goals sheet...")
        self.create_loe_goals_sheet()

        print("Creating Stakeholder_Master sheet...")
        self.create_stakeholder_master_sheet()

        print("Creating Master_Projects sheet...")
        self.create_master_projects_sheet()

        print("Creating Milestones sheet...")
        self.create_milestones_sheet()

        print("Creating Executive_Dashboard sheet...")
        self.create_executive_dashboard_sheet()

        print("Creating Project_Details sheet...")
        self.create_project_details_sheet()

        print("Creating Archive sheet...")
        self.create_archive_sheet()

        print("Adding data validations...")
        self.add_data_validations()

        print("Adding conditional formatting...")
        self.add_conditional_formatting()

        print("Saving workbook...")
        self.save_workbook('C:/Users/mrear/Downloads/Project_Management_Dashboard.xlsx')

        print("\nDashboard creation complete!")
        print("File saved to: C:/Users/mrear/Downloads/Project_Management_Dashboard.xlsx")
        print("\nNext steps:")
        print("1. Open the Excel file")
        print("2. Enable content if prompted")
        print("3. Review sample data in all sheets")
        print("4. Start customizing with your actual project data")
        print("5. Test the Executive Dashboard formulas")

if __name__ == "__main__":
    # Create the dashboard
    dashboard = ProjectManagementDashboard()
    dashboard.create_full_dashboard()
